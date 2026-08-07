"""
Script de demostración: genera Excel de auditoría de préstamos con datos ficticios.
Diseño contable profesional — formato para papel de trabajo de auditoría.
"""

import os
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Paleta de colores ─────────────────────────────────────────────────────────
_C_GRIS_HEADER = "FFB8B8B8"
_C_AZUL_OSCURO = "FF1F4E79"
_C_AZUL_MEDIO  = "FF2E75B6"
_C_GRIS_CLARO  = "FFF2F2F2"
_C_AMARILLO    = "FFFFFF00"
_C_ROJO        = "FFFF0000"
_C_BLANCO      = "FFFFFFFF"

fill_gris_header = PatternFill(start_color=_C_GRIS_HEADER, end_color=_C_GRIS_HEADER, fill_type="solid")
fill_azul_oscuro = PatternFill(start_color=_C_AZUL_OSCURO, end_color=_C_AZUL_OSCURO, fill_type="solid")
fill_azul_medio  = PatternFill(start_color=_C_AZUL_MEDIO,  end_color=_C_AZUL_MEDIO,  fill_type="solid")
fill_gris_claro  = PatternFill(start_color=_C_GRIS_CLARO,  end_color=_C_GRIS_CLARO,  fill_type="solid")
fill_amarillo    = PatternFill(start_color=_C_AMARILLO,     end_color=_C_AMARILLO,    fill_type="solid")
fill_rojo        = PatternFill(start_color=_C_ROJO,         end_color=_C_ROJO,        fill_type="solid")

font_bold_11   = Font(name="Calibri", bold=True, size=11)
font_bold_11_w = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
font_bold_10_w = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
font_bold_10   = Font(name="Calibri", bold=True, size=10)
font_data_10   = Font(name="Calibri", size=10)

FMT_MONEDA  = '$ #,##0.00'
COLS_GRILLA = ["CUOTA", "VENCIMIENTO", "CAPITAL", "INTERESES",
               "IVA/GASTOS", "MONTO A ABONAR", "SALDO RESTANTE"]
ANCHOS      = [8, 14, 16, 16, 16, 18, 18]
N_COLS      = len(COLS_GRILLA)  # 7

alin_izq    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
alin_der    = Alignment(horizontal="right",  vertical="center")
alin_centro = Alignment(horizontal="center", vertical="center")

_thin = Side(style="thin", color="FFB8B8B8")
_med  = Side(style="medium", color="FF2E75B6")
border_thin  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
border_grilla = Border(left=_med, right=_med, top=_med, bottom=_med)


# ── Amortización francesa (cuota fija) ───────────────────────────────────────
def _french(capital: float, n: int, tasa_anual: float, fecha_inicio: date,
            iva_tasa: float = 0.21) -> list[dict]:
    r = tasa_anual / 12
    if r == 0:
        cuota_fija = capital / n
    else:
        cuota_fija = capital * r / (1 - (1 + r) ** -n)

    cuotas = []
    saldo = capital
    for i in range(1, n + 1):
        interes = round(saldo * r, 2)
        cap_am  = round(cuota_fija - interes, 2)
        if i == n:                       # ajuste de redondeo en última cuota
            cap_am = round(saldo, 2)
        iva     = round(interes * iva_tasa, 2)
        monto   = round(cap_am + interes + iva, 2)
        saldo   = round(saldo - cap_am, 2)
        venc    = (fecha_inicio + relativedelta(months=i - 1)).strftime("%d/%m/%Y")
        cuotas.append({
            "cuota":          i,
            "vencimiento":    venc,
            "capital":        cap_am,
            "intereses":      interes,
            "iva_gastos":     iva,
            "monto_abonar":   monto,
            "saldo_restante": max(saldo, 0),
        })
    return cuotas


# ── Amortización alemana (capital constante) ─────────────────────────────────
def _german(capital: float, n: int, tasa_anual: float, fecha_inicio: date,
            iva_tasa: float = 0.21) -> list[dict]:
    r       = tasa_anual / 12
    cap_am  = round(capital / n, 2)
    cuotas  = []
    saldo   = capital
    for i in range(1, n + 1):
        if i == n:
            cap_am = round(saldo, 2)
        interes = round(saldo * r, 2)
        iva     = round(interes * iva_tasa, 2)
        monto   = round(cap_am + interes + iva, 2)
        saldo   = round(saldo - cap_am, 2)
        venc    = (fecha_inicio + relativedelta(months=i - 1)).strftime("%d/%m/%Y")
        cuotas.append({
            "cuota":          i,
            "vencimiento":    venc,
            "capital":        cap_am,
            "intereses":      interes,
            "iva_gastos":     iva,
            "monto_abonar":   monto,
            "saldo_restante": max(saldo, 0),
        })
    return cuotas


# ── Helpers de formato ────────────────────────────────────────────────────────
def _fila_merged(ws, fila: int, valor, fill, font, height: int = 22) -> None:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
    c = ws.cell(row=fila, column=1, value=valor)
    c.fill   = fill
    c.font   = font
    c.alignment = alin_izq
    ws.row_dimensions[fila].height = height


def _encabezado_grilla(ws, fila: int) -> None:
    for col, enc in enumerate(COLS_GRILLA, start=1):
        c = ws.cell(row=fila, column=col, value=enc)
        c.fill      = fill_azul_medio
        c.font      = font_bold_10_w
        c.alignment = alin_centro
        c.border    = border_grilla
    ws.row_dimensions[fila].height = 18


def _fila_cuota(ws, fila: int, datos: dict, es_par: bool) -> None:
    relleno = fill_gris_claro if es_par else None
    valores = [
        datos.get("cuota",          "—"),
        datos.get("vencimiento",    "—"),
        datos.get("capital",        "—"),
        datos.get("intereses",      "—"),
        datos.get("iva_gastos",     "—"),
        datos.get("monto_abonar",   "—"),
        datos.get("saldo_restante", "—"),
    ]
    for col, val in enumerate(valores, start=1):
        c = ws.cell(row=fila, column=col, value=val)
        c.font   = font_data_10
        c.border = border_thin
        if relleno:
            c.fill = relleno
        if col <= 2:
            c.alignment = alin_izq
        else:
            c.alignment = alin_der
            if isinstance(val, (int, float)):
                c.number_format = FMT_MONEDA
    ws.row_dimensions[fila].height = 15


def _ajustar_anchos(ws) -> None:
    for i, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _sum_cuotas(cuotas: list, campo: str) -> float:
    return sum(c.get(campo, 0) for c in cuotas if isinstance(c.get(campo), (int, float)))


# ── Escritura de un préstamo en la hoja ──────────────────────────────────────
def _escribir_prestamo(ws, fila: int, n: int, capital: float,
                       sistema: str, cuotas: list) -> int:
    # [A] Cabecera del préstamo
    titulo = (
        f"PRÉSTAMO N° {n}  |  Capital Original: $ {float(capital):,.0f}"
        f"  |  Sistema: {sistema}"
    )
    _fila_merged(ws, fila, titulo, fill_gris_header, font_bold_11, height=22)
    fila += 1

    # [B] Resumen anual — 3 filas: título (azul oscuro), labels, valores
    total_capital   = _sum_cuotas(cuotas, "capital")
    total_intereses = _sum_cuotas(cuotas, "intereses")
    total_iva       = _sum_cuotas(cuotas, "iva_gastos")

    anio_label = ""
    for q in cuotas:
        v = str(q.get("vencimiento", ""))
        if len(v) >= 10 and v[6:10].isdigit():
            anio_label = v[6:10]
            break

    titulo_resumen = f"Resumen Año {anio_label}" if anio_label else "Resumen Anual"
    _fila_merged(ws, fila, titulo_resumen, fill_azul_oscuro, font_bold_11_w, height=18)
    fila += 1

    enc_res = ["Total Capital Amortizado", "Total Intereses Devengados", "Total IVA/Gastos"]
    for col, enc in enumerate(enc_res, start=1):
        c = ws.cell(row=fila, column=col, value=enc)
        c.font      = font_bold_10
        c.alignment = alin_izq
    ws.row_dimensions[fila].height = 16
    fila += 1

    for col, val in enumerate([total_capital, total_intereses, total_iva], start=1):
        c = ws.cell(row=fila, column=col, value=val)
        c.font          = font_data_10
        c.number_format = FMT_MONEDA
        c.alignment     = alin_der
    ws.row_dimensions[fila].height = 15
    fila += 1
    fila += 1  # línea en blanco entre resumen y grilla

    # [C] Grilla cronológica
    _encabezado_grilla(ws, fila)
    fila += 1

    for idx_c, q in enumerate(cuotas):
        _fila_cuota(ws, fila, q, idx_c % 2 == 1)
        fila += 1

    # 4 filas en blanco entre préstamos
    fila += 4
    return fila


# ── Cierre contable de la hoja ────────────────────────────────────────────────
def _cierre_contable(ws, fila: int, saldo_ini: float, prestamos_lista: list) -> None:
    total_cap = sum(_sum_cuotas(p["cuotas"], "capital") for p in prestamos_lista)
    saldo_final = saldo_ini - total_cap

    fila += 2
    _fila_merged(ws, fila, "CONCILIACIÓN CONTABLE FINAL",
                 fill_azul_oscuro, font_bold_11_w, height=22)
    fila += 1

    filas_cierre = [
        ("Concepto",                                   "Importe",   None),
        ("Saldo Inicial del Banco",                    saldo_ini,   None),
        ("Total Capital Amortizado (todos préstamos)", total_cap,   None),
        ("Saldo Final Sugerido Mayor Contable",        saldo_final,
         fill_amarillo if saldo_final >= 0 else fill_rojo),
    ]

    for i, (concepto, importe, relleno_cierre) in enumerate(filas_cierre):
        c_concepto = ws.cell(row=fila, column=1, value=concepto)
        c_importe  = ws.cell(row=fila, column=2, value=importe)
        fnt = font_bold_10 if i == 0 else font_data_10
        c_concepto.font = fnt
        c_importe.font  = fnt
        c_concepto.alignment = alin_izq
        c_importe.alignment  = alin_der
        if i > 0 and isinstance(importe, (int, float)):
            c_importe.number_format = FMT_MONEDA
        if relleno_cierre:
            c_importe.fill  = relleno_cierre
            c_concepto.fill = relleno_cierre
        ws.row_dimensions[fila].height = 16
        fila += 1


# ── Construcción del workbook ─────────────────────────────────────────────────
def generar_demo() -> str:
    TASA_ANUAL = 0.40   # 40 % anual (demo Argentina)
    IVA_TASA   = 0.21

    bancos_data = {
        "Banco Galicia": [
            {
                "n":       1,
                "capital": 5_000_000,
                "sistema": "Francés",
                "cuotas":  _french(5_000_000, 12, TASA_ANUAL,
                                   date(2026, 1, 1), IVA_TASA),
            },
            {
                "n":       2,
                "capital": 2_000_000,
                "sistema": "Alemán",
                "cuotas":  _german(2_000_000, 6, TASA_ANUAL,
                                   date(2026, 7, 1), IVA_TASA),
            },
        ],
        "Banco Santander": [
            {
                "n":       1,
                "capital": 3_500_000,
                "sistema": "Francés",
                "cuotas":  _french(3_500_000, 18, TASA_ANUAL,
                                   date(2026, 1, 1), IVA_TASA),
            },
        ],
    }

    saldos_iniciales = {
        "Banco Galicia":   7_000_000,
        "Banco Santander": 3_500_000,
    }

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    filas_por_hoja = {}

    for nombre_banco, prestamos_lista in bancos_data.items():
        ws   = wb.create_sheet(nombre_banco)
        fila = 1

        for prestamo in prestamos_lista:
            fila = _escribir_prestamo(
                ws, fila,
                n=prestamo["n"],
                capital=prestamo["capital"],
                sistema=prestamo["sistema"],
                cuotas=prestamo["cuotas"],
            )

        _cierre_contable(ws, fila, saldos_iniciales[nombre_banco], prestamos_lista)
        _ajustar_anchos(ws)

        # Congelar encabezado y activar zoom
        ws.sheet_view.zoomScale = 90

        filas_por_hoja[nombre_banco] = fila + 10

    # ── Hoja de portada / resumen ejecutivo ──────────────────────────────────
    ws_cover = wb.create_sheet("Resumen Ejecutivo", 0)
    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 22
    ws_cover.column_dimensions["C"].width = 22

    cover_rows = [
        ("AUDITORÍA DE PRÉSTAMOS FINANCIEROS — DEMO",),
        ("Archivo de demostración con datos ficticios",),
        (),
        ("Fecha de generación", date.today().strftime("%d/%m/%Y"),),
        ("Versión", "Demo v1.0",),
        (),
        ("RESUMEN POR ENTIDAD", "Capital Total", "Saldo Inicial"),
    ]
    for banco, prestamos_lista in bancos_data.items():
        cap = sum(p["capital"] for p in prestamos_lista)
        sal = saldos_iniciales[banco]
        cover_rows.append((banco, cap, sal))

    cover_rows += [
        (),
        ("DETALLE DE PRÉSTAMOS",),
        ("Banco", "Préstamo", "Sistema | Cuotas | Capital"),
    ]
    for banco, prestamos_lista in bancos_data.items():
        for p in prestamos_lista:
            cover_rows.append((
                banco,
                f"Préstamo N° {p['n']}",
                f"{p['sistema']} | {len(p['cuotas'])} cuotas | $ {p['capital']:,.0f}",
            ))

    cover_rows += [
        (),
        ("NOTAS:",),
        ("• Tasa de interés utilizada", "40% anual (demo Argentina)",),
        ("• IVA sobre intereses", "21%",),
        ("• Sistema Francés", "Cuota fija — amortización creciente",),
        ("• Sistema Alemán", "Capital constante — cuota decreciente",),
        ("• Los datos son ficticios y no representan operaciones reales.", ),
        (),
        ("CONCILIACIÓN GLOBAL",),
        ("Total saldos iniciales",
         sum(saldos_iniciales.values()),),
        ("Total capital a amortizar",
         sum(sum(_sum_cuotas(p["cuotas"], "capital") for p in pl)
             for pl in bancos_data.values()),),
        ("Total intereses devengados",
         sum(sum(_sum_cuotas(p["cuotas"], "intereses") for p in pl)
             for pl in bancos_data.values()),),
        ("Total IVA/Gastos",
         sum(sum(_sum_cuotas(p["cuotas"], "iva_gastos") for p in pl)
             for pl in bancos_data.values()),),
    ]

    for i, row_data in enumerate(cover_rows, start=1):
        for j, val in enumerate(row_data, start=1):
            c = ws_cover.cell(row=i, column=j, value=val)
            c.font = font_data_10
            c.alignment = alin_izq
            if j > 1 and isinstance(val, (int, float)):
                c.number_format = FMT_MONEDA
                c.alignment = alin_der
        ws_cover.row_dimensions[i].height = 16

        # Estilos especiales para ciertas filas
        if i == 1:
            ws_cover.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws_cover.cell(1, 1).fill = fill_azul_oscuro
            ws_cover.cell(1, 1).font = font_bold_11_w
            ws_cover.row_dimensions[1].height = 26
        elif i == 2:
            ws_cover.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
            ws_cover.cell(2, 1).fill = fill_azul_medio
            ws_cover.cell(2, 1).font = Font(name="Calibri", size=10, color="FFFFFFFF", italic=True)
        elif row_data and row_data[0] in ("RESUMEN POR ENTIDAD", "DETALLE DE PRÉSTAMOS",
                                          "NOTAS:", "CONCILIACIÓN GLOBAL"):
            ws_cover.cell(i, 1).fill = fill_gris_header
            ws_cover.cell(i, 1).font = font_bold_10
        elif row_data and len(row_data) >= 2 and isinstance(row_data[1], float):
            for j in range(1, 4):
                c2 = ws_cover.cell(i, j)
                c2.border = border_thin

    ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "demo_prestamos_auditoria.xlsx")
    wb.save(ruta)
    return ruta, filas_por_hoja


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ruta, filas = generar_demo()
    tam_kb = os.path.getsize(ruta) / 1024

    print(f"\n{'='*60}")
    print(f"  DEMO EXCEL GENERADO")
    print(f"{'='*60}")
    print(f"  Ruta   : {ruta}")
    print(f"  Tamaño : {tam_kb:.1f} KB")
    print(f"\n  Hojas creadas:")
    for hoja, n_filas in filas.items():
        print(f"    • {hoja}: ~{n_filas} filas")
    print(f"{'='*60}\n")
