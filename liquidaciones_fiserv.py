#!/usr/bin/env python3
"""
Liquidaciones Fiserv / First Data → Excel de auditoría contable.

Lee resúmenes PDF de liquidaciones de tarjeta (cualquier comercio, cualquier
marca/modalidad) y genera un Excel con hoja Resumen + una hoja de detalle por PDF.

Uso:
  python liquidaciones_fiserv.py --carpeta-pdfs "T:\\...\\2026-07" --carpeta-salida "D:\\Salida"
  python liquidaciones_fiserv.py --carpeta-pdfs .\\pdfs --carpeta-salida .\\out --mes Julio --anio 2026
"""

from __future__ import annotations

import argparse
import re
import sys
import tempfile
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# ---------------------------------------------------------------------------
# Constantes de formato (réplica del Excel de referencia del estudio)
# ---------------------------------------------------------------------------

COLOR_TITULO = "1F4E78"
COLOR_HEADER_BG = "1F4E78"
COLOR_TOTALES = "DCE6F1"
COLOR_BORDE = "BFBFBF"
TOLERANCIA = 0.01

MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}
MESES_PDF = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

# Deducciones conocidas: (clave interna, regex del concepto, etiqueta consola)
DEDUCCIONES_CONOCIDAS: list[tuple[str, re.Pattern[str], str]] = [
    (
        "arancel",
        re.compile(r"-\s*ARANCEL\b", re.I),
        "ARANCEL",
    ),
    (
        "iva_arancel",
        re.compile(r"-\s*IVA\s*CRED\.?\s*FISC\.?\s*COMERCIO\s*S/?\s*ARANC", re.I),
        "IVA CRED.FISC. s/Arancel",
    ),
    (
        "per_iibb_ba",
        re.compile(r"-\s*PER\s*B\.?\s*A\.?\s*I\.?\s*BR\.?\s*DN\.?\s*01/?04", re.I),
        "PER B.A.I.BR.DN.01/04",
    ),
    (
        "ret_sirtac",
        re.compile(r"-\s*RETENCION\s*ING\.?\s*BRUTOS\s*SIRTAC", re.I),
        "RETENCION ING.BRUTOS SIRTAC",
    ),
    (
        "perc_iva_2408",
        re.compile(r"-\s*PERCEPCION\s*IVA\s*R\.?\s*G\.?\s*2408", re.I),
        "PERCEPCION IVA R.G. 2408",
    ),
    (
        "cargo_terminal",
        re.compile(r"-\s*CARGO\s*TERMINAL\s*FISERV", re.I),
        "CARGO TERMINAL FISERV",
    ),
]

# Líneas con "-" que NO son deducciones de liquidación (pies / legales)
RE_IGNORAR_GUION = re.compile(
    r"-\s*(Rep[uú]blica|República|Estimado|Centro de Atenci|"
    r"Acceda|Le comunicamos|a partir|MON\.|Usuario)",
    re.I,
)

RE_MONTO = re.compile(r"\$\s*([\d.]+,\d{2})")
RE_BLOQUE_VENTAS = re.compile(
    r"\+\s*VENTAS\s*C/?DESCUENTO\s*CONTADO\s*\$\s*([\d.]+,\d{2})",
    re.I,
)
RE_NETO = re.compile(r"IMPORTE\s*NETO\s*DE\s*PAGOS\s*\$\s*([\d.]+,\d{2})", re.I)
RE_F_PAGO = re.compile(
    r"F\.?\s*de\s*Pago:.*?el\s+d[ií]a\s+(\d{2}/\d{2}/\d{4}).*?Nro\.?\s*Liq:\s*(\d+)",
    re.I | re.S,
)
RE_TITULO = re.compile(
    r"TARJETA\s+DE\s+(\w+)\s+PESOS\s+(\w+)\s+(\d{4})",
    re.I,
)
RE_TOTAL_PRESENTADO = re.compile(r"Total\s+presentado:\s*([\d.]+,\d{2})", re.I)
RE_NETO_PAGOS = re.compile(r"Neto\s+de\s+pagos:\s*([\d.]+,\d{2})", re.I)
RE_TOTAL_LIQ = re.compile(
    r"TOTAL\s+LIQ\.?\s*TARJ\.?\s*\w+\s*:?\s*(\d+)",
    re.I,
)
RE_CUIT_COMERCIO = re.compile(
    r"Comercio.*?CUIT:\s*(\d{2}-\d{8}-\d)",
    re.I | re.S,
)
RE_NRO_COMERCIO = re.compile(r"N.?º?\s*Comercio:\s*([0-9/ ]+)", re.I)


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------


def parse_ar_money(s: str | None) -> float:
    if s is None:
        return 0.0
    t = str(s).strip().replace("$", "").replace(" ", "")
    if not t:
        return 0.0
    neg = t.startswith("-") or t.startswith("(")
    t = t.replace("-", "").replace("(", "").replace(")", "")
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        v = float(t)
    except ValueError:
        return 0.0
    return -v if neg else v


def money_eq(a: float, b: float, tol: float = TOLERANCIA) -> bool:
    return abs(a - b) <= tol


def sane_filename(name: str) -> str:
    bad = '<>:"/\\|?*'
    out = "".join("_" if c in bad else c for c in name)
    out = re.sub(r"\s+", "_", out.strip())
    return out[:180] or "Comercio"


def sane_sheet_name(name: str) -> str:
    bad = r"\/*?:[]"
    out = "".join("_" if c in bad else c for c in name).strip()
    return (out or "Hoja")[:31]


def title_modalidad(raw: str) -> str:
    u = (raw or "").strip().upper()
    mapa = {
        "DEBITO": "Debito",
        "DÉBITO": "Debito",
        "CREDITO": "Credito",
        "CRÉDITO": "Credito",
    }
    return mapa.get(u, raw.strip().title() if raw else "Modalidad")


def title_marca(raw: str) -> str:
    u = (raw or "").strip().upper()
    mapa = {
        "VISA": "Visa",
        "MASTERCARD": "Mastercard",
        "MASTER": "Mastercard",
        "MASTER CARD": "Mastercard",
        "CABAL": "Cabal",
        "AMEX": "Amex",
        "AMERICAN EXPRESS": "Amex",
        "NARANJA": "Naranja",
        "MAESTRO": "Maestro",
    }
    return mapa.get(u, raw.strip().title() if raw else "MarcaDesconocida")


# ---------------------------------------------------------------------------
# Detección de marca (logo del encabezado — no usar nombre de archivo)
# ---------------------------------------------------------------------------


def _score_logo_colores(im: Image.Image) -> dict[str, float]:
    """Heurística de color sobre el logo superior-izquierdo del PDF."""
    import numpy as np

    arr = np.array(im.convert("RGB"), dtype=float)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    # Excluir casi blanco / casi negro / grises planos
    mask = ~(
        ((r > 240) & (g > 240) & (b > 240))
        | ((r < 25) & (g < 25) & (b < 25))
    )
    n = int(mask.sum())
    if n < 30:
        return {}

    visa = float(((b > 100) & (b > r + 20) & (b > g + 20) & (r < 180)).sum()) / n
    mc_red = float(((r > 150) & (r > g + 40) & (r > b + 40)).sum()) / n
    mc_ora = float(
        ((r > 150) & (g > 80) & (g < 200) & (b < 100) & (r > b + 50)).sum()
    ) / n
    # Naranja (tarjeta): naranja intenso sin tanto rojo "Mastercard circle"
    naranja = float(
        ((r > 180) & (g > 70) & (g < 160) & (b < 60) & (r > g + 40)).sum()
    ) / n
    # Cabal / verdes
    cabal = float(((g > 100) & (g > r + 15) & (g > b + 10)).sum()) / n
    # Amex azul oscuro
    amex = float(((b > 80) & (r < 80) & (g < 100) & (b > r + 30)).sum()) / n

    return {
        "Visa": visa,
        "Mastercard": max(mc_red, mc_ora * 0.9 + mc_red * 0.1),
        "Naranja": naranja,
        "Cabal": cabal,
        "Amex": amex,
    }


def detectar_marca_logo(pdf_path: Path) -> tuple[str, str]:
    """
    Extrae el logo del encabezado (esquina superior izquierda) y clasifica la marca.
    Retorna (marca_titulo, metodo).
    """
    doc = fitz.open(str(pdf_path))
    try:
        page = doc[0]
        candidatos: list[tuple[float, fitz.Rect]] = []
        for img in page.get_images(full=True):
            xref = img[0]
            try:
                rects = page.get_image_rects(xref)
            except Exception:
                rects = []
            for r in rects:
                if r.y0 < 70 and r.x0 < 160 and r.width >= 25 and r.height >= 15:
                    # Preferir el más a la izquierda / arriba
                    score = (100 - r.x0) + (50 - r.y0) + min(r.width, 80) * 0.1
                    candidatos.append((score, r))
        if not candidatos:
            return "MarcaDesconocida", "sin_logo"

        candidatos.sort(key=lambda x: -x[0])
        rect = candidatos[0][1]
        pix = page.get_pixmap(matrix=fitz.Matrix(4, 4), clip=rect, alpha=False)
        im = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        scores = _score_logo_colores(im)
        if not scores:
            return "MarcaDesconocida", "logo_sin_color"

        mejor = max(scores, key=scores.get)
        if scores[mejor] < 0.35:
            # Intentar OCR opcional (easyocr) si está instalado
            ocr_marca = _ocr_marca_opcional(im)
            if ocr_marca:
                return title_marca(ocr_marca), "ocr_logo"
            return "MarcaDesconocida", f"baja_confianza({mejor}={scores[mejor]:.2f})"
        return title_marca(mejor), f"color_logo({mejor}={scores[mejor]:.2f})"
    finally:
        doc.close()


def _ocr_marca_opcional(im: Image.Image) -> str | None:
    try:
        import easyocr  # type: ignore
    except Exception:
        return None
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            reader = easyocr.Reader(["en", "es"], verbose=False)
            texts = reader.readtext(
                __import__("numpy").array(im),
                detail=0,
                paragraph=False,
            )
        blob = " ".join(str(t) for t in texts).upper()
        for key in (
            "MASTERCARD",
            "MASTER CARD",
            "VISA",
            "CABAL",
            "AMEX",
            "AMERICAN EXPRESS",
            "NARANJA",
            "MAESTRO",
        ):
            if key in blob:
                return key
    except Exception:
        return None
    return None


# ---------------------------------------------------------------------------
# Modelo de datos
# ---------------------------------------------------------------------------


@dataclass
class Liquidacion:
    fecha_pago: datetime
    nro_liq: str
    ventas: float
    arancel: float = 0.0
    iva_arancel: float = 0.0
    per_iibb_ba: float = 0.0
    ret_sirtac: float = 0.0
    perc_iva_2408: float = 0.0
    cargo_terminal: float = 0.0
    otras: float = 0.0
    neto: float = 0.0
    otras_detalle: list[str] = field(default_factory=list)


@dataclass
class ResumenPdf:
    path: Path
    razon_social: str
    nombre_fantasia: str
    cuit: str
    nro_comercio: str
    domicilio: str
    marca: str
    modalidad: str
    mes_nombre: str
    anio: int
    mes_num: int
    total_presentado_pdf: float
    neto_pagos_pdf: float
    cant_liq_pdf: int
    liquidaciones: list[Liquidacion]
    marca_metodo: str
    advertencias: list[str] = field(default_factory=list)

    @property
    def hoja(self) -> str:
        override = getattr(self, "_sheet_name", None)
        if override:
            return str(override)
        return sane_sheet_name(f"{self.marca} {self.modalidad}")
    @property
    def sum_ventas(self) -> float:
        return round(sum(x.ventas for x in self.liquidaciones), 2)

    @property
    def sum_neto(self) -> float:
        return round(sum(x.neto for x in self.liquidaciones), 2)


@dataclass
class ResultadoValidacion:
    ok: bool
    mensajes: list[str]


# ---------------------------------------------------------------------------
# Parser de PDF
# ---------------------------------------------------------------------------


def _extraer_texto(pdf_path: Path) -> str:
    with pdfplumber.open(str(pdf_path)) as pdf:
        partes = [(pg.extract_text() or "") for pg in pdf.pages]
    return "\n".join(partes)


def _extraer_bloque_comercio(text: str) -> tuple[str, str, str, str, str]:
    """
    Retorna razon_social, fantasia, cuit, nro_comercio, domicilio.
    """
    cuit = ""
    m = RE_CUIT_COMERCIO.search(text)
    if m:
        cuit = m.group(1).strip()
    else:
        # Fallback: primer CUIT del comercio (no el de First Data 30-52221156-3)
        cuits = re.findall(r"CUIT:\s*(\d{2}-\d{8}-\d)", text)
        for c in cuits:
            if not c.startswith("30-52221156"):
                cuit = c
                break
        if not cuit and cuits:
            cuit = cuits[-1]

    nro = ""
    m = RE_NRO_COMERCIO.search(text)
    if m:
        nro = re.sub(r"\s+", "", m.group(1)).strip(" /")

    # Bloque Comercio: líneas entre "Comercio" y "Total presentado" / "LA LEY"
    m_block = re.search(
        r"Comercio\s+N.?º?\s*Comercio:.*?\n(.*?)(?:Total presentado:|LA LEY)",
        text,
        re.I | re.S,
    )
    razon = ""
    fantasia = ""
    domicilio = ""
    if m_block:
        lineas = [
            ln.strip()
            for ln in m_block.group(1).splitlines()
            if ln.strip()
            and not re.match(r"CUIT:", ln, re.I)
            and not re.match(r"N.?º?\s*Ing", ln, re.I)
            and not re.match(r"Categor", ln, re.I)
            and not re.match(r"Hoja:", ln, re.I)
        ]
        # Típico: RAZON / FANTASIA / CALLE / CP-CIUDAD
        if lineas:
            razon = lineas[0]
        if len(lineas) >= 2 and not re.match(r"\d{5}", lineas[1]):
            # Si la 2ª no parece CP, es fantasía (salvo que sea calle con número)
            if not re.search(r"\d{3,}", lineas[1]) or re.match(
                r"^[A-ZÁÉÍÓÚÑ ]+$", lineas[1], re.I
            ):
                # Heurística: fantasía suele ser una sola palabra / marca corta
                if len(lineas[1].split()) <= 3 and not re.search(
                    r"\d{4,}", lineas[1]
                ):
                    fantasia = lineas[1]
                    resto = lineas[2:]
                else:
                    resto = lineas[1:]
            else:
                resto = lineas[1:]
        else:
            resto = lineas[1:]
        domicilio = ", ".join(resto[:3]) if resto else ""

    if not razon:
        # Fallback simple
        m2 = re.search(r"CUIT:\s*\d{2}-\d{8}-\d\s*\n([^\n]+)", text)
        if m2:
            razon = m2.group(1).strip()

    return razon.strip(), fantasia.strip(), cuit, nro, domicilio.strip()


def _clasificar_deduccion(linea: str) -> tuple[str | None, float]:
    """Retorna (clave_conocida | None, monto). None = no reconocida."""
    m_monto = RE_MONTO.search(linea)
    if not m_monto:
        return None, 0.0
    monto = parse_ar_money(m_monto.group(1))
    for clave, pat, _lab in DEDUCCIONES_CONOCIDAS:
        if pat.search(linea):
            return clave, monto
    return None, monto


def _parsear_liquidaciones(text: str, advertencias: list[str]) -> list[Liquidacion]:
    """
    Une bloques que cruzan salto de página: a veces IMPORTE NETO queda al pie
    de una hoja y F.de Pago al inicio de la siguiente.
    """
    flat = re.sub(r"[ \t]+", " ", text)
    flat = re.sub(r"\n+", "\n", flat)

    # Regex que no atraviesa otro "+ VENTAS..." entre ventas y F.de Pago
    bloque_re = re.compile(
        r"\+\s*VENTAS\s*C/?DESCUENTO\s*CONTADO\s*\$\s*([\d.]+,\d{2})"
        r"((?:(?!\+\s*VENTAS\s*C/?DESCUENTO\s*CONTADO).)*?)"
        r"IMPORTE\s*NETO\s*DE\s*PAGOS\s*\$\s*([\d.]+,\d{2})"
        r"((?:(?!\+\s*VENTAS\s*C/?DESCUENTO\s*CONTADO).)*?)"
        r"el\s+d[ií]a\s+(\d{2}/\d{2}/\d{4}).{0,160}?Nro\.?\s*Liq:\s*(\d+)",
        re.I | re.S,
    )

    out: list[Liquidacion] = []
    for m in bloque_re.finditer(flat):
        ventas = parse_ar_money(m.group(1))
        ded_tramo = m.group(2)
        neto = parse_ar_money(m.group(3))
        # grupo 4 = texto entre neto y fecha (ignorado)
        fecha = datetime.strptime(m.group(5), "%d/%m/%Y")
        nro = m.group(6)

        vals = {
            "arancel": 0.0,
            "iva_arancel": 0.0,
            "per_iibb_ba": 0.0,
            "ret_sirtac": 0.0,
            "perc_iva_2408": 0.0,
            "cargo_terminal": 0.0,
            "otras": 0.0,
        }
        otras_det: list[str] = []

        for m_ded in re.finditer(
            r"-\s*((?:(?!\$).){3,120}?)\$\s*([\d.]+,\d{2})",
            ded_tramo,
            re.I | re.S,
        ):
            concepto = m_ded.group(1).strip()
            line = f"- {concepto} $ {m_ded.group(2)}"
            if RE_IGNORAR_GUION.search(line):
                continue
            # Evitar basura de pies de página / teléfonos
            if re.search(r"\d{3,}-\d{3,}|0800|www\.|http", concepto, re.I):
                continue
            clave, monto = _clasificar_deduccion(line)
            if clave is None:
                # Solo advertir si parece una deducción contable (mayúsculas / %)
                if re.search(r"[A-ZÁÉÍÓÚÑ]{4,}", concepto):
                    vals["otras"] = round(vals["otras"] + monto, 2)
                    otras_det.append(line)
                    advertencias.append(
                        f"Deduccion no reconocida en liq {nro}: {line}"
                    )
            else:
                vals[clave] = round(vals[clave] + monto, 2)

        out.append(
            Liquidacion(
                fecha_pago=fecha,
                nro_liq=nro,
                ventas=ventas,
                arancel=vals["arancel"],
                iva_arancel=vals["iva_arancel"],
                per_iibb_ba=vals["per_iibb_ba"],
                ret_sirtac=vals["ret_sirtac"],
                perc_iva_2408=vals["perc_iva_2408"],
                cargo_terminal=vals["cargo_terminal"],
                otras=vals["otras"],
                neto=neto,
                otras_detalle=otras_det,
            )
        )

    # Detectar ventas sin match (para diagnóstico)
    n_ventas = len(RE_BLOQUE_VENTAS.findall(flat))
    if n_ventas != len(out):
        advertencias.append(
            f"Se detectaron {n_ventas} bloques de ventas y {len(out)} "
            f"liquidaciones con F.de Pago completo"
        )

    out.sort(key=lambda x: (x.fecha_pago, x.nro_liq))
    return out


def parsear_pdf(pdf_path: Path) -> ResumenPdf:
    text = _extraer_texto(pdf_path)
    adv: list[str] = []

    m_tit = RE_TITULO.search(text)
    if not m_tit:
        raise ValueError(
            f"{pdf_path.name}: no se encontró 'TARJETA DE [MODALIDAD] PESOS [MES] [AÑO]' "
            "en el encabezado."
        )
    modalidad = title_modalidad(m_tit.group(1))
    mes_raw = m_tit.group(2).upper()
    anio = int(m_tit.group(3))
    if mes_raw not in MESES_PDF:
        raise ValueError(f"{pdf_path.name}: mes no reconocido en encabezado: {mes_raw}")
    mes_num = MESES_PDF[mes_raw]
    mes_nombre = MESES_ES[mes_num]

    razon, fantasia, cuit, nro_comercio, domicilio = _extraer_bloque_comercio(text)
    if not cuit:
        raise ValueError(f"{pdf_path.name}: no se pudo extraer CUIT del comercio.")
    if not razon:
        raise ValueError(f"{pdf_path.name}: no se pudo extraer razón social.")

    marca, metodo = detectar_marca_logo(pdf_path)
    if marca == "MarcaDesconocida":
        adv.append(
            f"{pdf_path.name}: marca no detectada en logo del encabezado "
            f"(método={metodo}). Revisar manualmente el nombre de hoja."
        )

    m_tp = RE_TOTAL_PRESENTADO.search(text)
    m_np = RE_NETO_PAGOS.search(text)
    m_tl = RE_TOTAL_LIQ.search(text)
    if not m_tp or not m_np:
        raise ValueError(
            f"{pdf_path.name}: faltan 'Total presentado' / 'Neto de pagos' en el PDF."
        )
    total_presentado = parse_ar_money(m_tp.group(1))
    neto_pagos = parse_ar_money(m_np.group(1))
    cant_liq = int(m_tl.group(1)) if m_tl else -1

    liqs = _parsear_liquidaciones(text, adv)

    return ResumenPdf(
        path=pdf_path,
        razon_social=razon,
        nombre_fantasia=fantasia,
        cuit=cuit,
        nro_comercio=nro_comercio,
        domicilio=domicilio,
        marca=marca,
        modalidad=modalidad,
        mes_nombre=mes_nombre,
        anio=anio,
        mes_num=mes_num,
        total_presentado_pdf=total_presentado,
        neto_pagos_pdf=neto_pagos,
        cant_liq_pdf=cant_liq,
        liquidaciones=liqs,
        marca_metodo=metodo,
        advertencias=adv,
    )


def validar_resumen(r: ResumenPdf) -> ResultadoValidacion:
    msgs: list[str] = []
    ok = True

    if not money_eq(r.sum_ventas, r.total_presentado_pdf):
        ok = False
        msgs.append(
            f"[{r.path.name}] Total presentado PDF={r.total_presentado_pdf:,.2f} "
            f"vs suma Ventas={r.sum_ventas:,.2f} "
            f"(diff={r.sum_ventas - r.total_presentado_pdf:,.2f})"
        )
    else:
        msgs.append(
            f"[{r.path.name}] Total presentado OK ({r.total_presentado_pdf:,.2f})"
        )

    if not money_eq(r.sum_neto, r.neto_pagos_pdf):
        ok = False
        msgs.append(
            f"[{r.path.name}] Neto de pagos PDF={r.neto_pagos_pdf:,.2f} "
            f"vs suma Netos={r.sum_neto:,.2f} "
            f"(diff={r.sum_neto - r.neto_pagos_pdf:,.2f})"
        )
    else:
        msgs.append(f"[{r.path.name}] Neto de pagos OK ({r.neto_pagos_pdf:,.2f})")

    if r.cant_liq_pdf < 0:
        ok = False
        msgs.append(f"[{r.path.name}] No se leyó 'TOTAL LIQ. TARJ. ...'")
    elif len(r.liquidaciones) != r.cant_liq_pdf:
        ok = False
        msgs.append(
            f"[{r.path.name}] Cant. liquidaciones extraídas={len(r.liquidaciones)} "
            f"vs TOTAL LIQ PDF={r.cant_liq_pdf}"
        )
        for liq in r.liquidaciones:
            msgs.append(
                f"    - {liq.fecha_pago:%d/%m/%Y} liq {liq.nro_liq} "
                f"ventas={liq.ventas:,.2f} neto={liq.neto:,.2f}"
            )
    else:
        msgs.append(
            f"[{r.path.name}] Cant. liquidaciones OK ({r.cant_liq_pdf})"
        )

    return ResultadoValidacion(ok=ok, mensajes=msgs)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_FONT = "Arial"
_thin = Border(
    left=Side(style="thin", color=COLOR_BORDE),
    right=Side(style="thin", color=COLOR_BORDE),
    top=Side(style="thin", color=COLOR_BORDE),
    bottom=Side(style="thin", color=COLOR_BORDE),
)
_fill_header = PatternFill("solid", fgColor=COLOR_HEADER_BG)
_fill_tot = PatternFill("solid", fgColor=COLOR_TOTALES)
_font_title = Font(name=_FONT, bold=True, color=COLOR_TITULO, size=14)
_font_sub = Font(name=_FONT, size=11, color="666666")
_font_header = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
_font_normal = Font(name=_FONT, size=10)
_font_bold = Font(name=_FONT, bold=True, size=10)
_fmt_num = "#,##0.00"
_align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
_align_r = Alignment(horizontal="right", vertical="center")
_align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADERS_DETALLE = [
    "Fecha de pago",
    "N° Liquidación",
    "Ventas (Bruto)",
    "Arancel",
    "IVA Cred.Fisc. s/Arancel",
    "Percepción IIBB (Bs.As. Dto.01/04)",
    "Retención IIBB SIRTAC",
    "Percepción IVA R.G. 2408",
    "Cargo Terminal Fiserv",
    "Otras deducciones",
    "Importe Neto Pagado",
]

HEADERS_RESUMEN = [
    "Tarjeta",
    "Ventas (Bruto)",
    "Arancel",
    "IVA Cred.Fisc. s/Arancel",
    "Percepción IIBB",
    "Retención IIBB SIRTAC",
    "Percepción IVA RG2408",
    "Cargo Terminal Fiserv",
    "Otras deducciones",
    "Importe Neto",
]


def _escribir_detalle(ws, r: ResumenPdf) -> int:
    """Escribe hoja de detalle. Retorna fila de TOTALES (1-based)."""
    n = len(r.liquidaciones)
    fant = f" ({r.nombre_fantasia})" if r.nombre_fantasia else ""
    ws["A1"] = (
        f"Liquidación Mensual - {r.marca} {r.modalidad} - {r.mes_nombre} {r.anio}"
    )
    ws["A1"].font = _font_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    ws["A2"] = f"Comercio: {r.razon_social}{fant} - CUIT {r.cuit}"
    ws["A2"].font = _font_sub
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    header_row = 4
    for col, h in enumerate(HEADERS_DETALLE, 1):
        cell = ws.cell(header_row, col, h)
        cell.fill = _fill_header
        cell.font = _font_header
        cell.alignment = _align_c
        cell.border = _thin

    data_start = 5
    for i, liq in enumerate(r.liquidaciones):
        row = data_start + i
        valores = [
            liq.fecha_pago.strftime("%d/%m/%Y"),
            liq.nro_liq,
            liq.ventas,
            liq.arancel,
            liq.iva_arancel,
            liq.per_iibb_ba,
            liq.ret_sirtac,
            liq.perc_iva_2408,
            liq.cargo_terminal,
            liq.otras,
            liq.neto,
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row, col, val)
            cell.font = _font_normal
            cell.border = _thin
            if col <= 2:
                cell.alignment = _align_c
            else:
                cell.alignment = _align_r
                cell.number_format = _fmt_num

    data_end = data_start + n - 1 if n else data_start
    # Si no hay liquidaciones, igual dejar estructura
    if n == 0:
        data_end = data_start - 1

    control_row = (data_end + 1) if n else data_start
    tot_row = control_row + 1

    ws.cell(control_row, 1, "Control por línea (Neto calc. - Neto informado)").font = (
        _font_normal
    )
    # Control sobre fila TOTALES
    ctrl = (
        f"=C{tot_row}-D{tot_row}-E{tot_row}-F{tot_row}-G{tot_row}"
        f"-H{tot_row}-I{tot_row}-J{tot_row}-K{tot_row}"
    )
    c = ws.cell(control_row, 11, ctrl)
    c.font = _font_normal
    c.number_format = _fmt_num
    c.border = _thin

    ws.cell(tot_row, 1, "TOTALES").font = _font_bold
    ws.cell(tot_row, 1).fill = _fill_tot
    for col in range(1, 12):
        cell = ws.cell(tot_row, col)
        cell.fill = _fill_tot
        cell.border = _thin
        cell.font = _font_bold
    if n:
        for col in range(3, 12):
            letter = get_column_letter(col)
            cell = ws.cell(
                tot_row, col, f"=SUM({letter}{data_start}:{letter}{data_end})"
            )
            cell.font = _font_bold
            cell.fill = _fill_tot
            cell.border = _thin
            cell.number_format = _fmt_num
            cell.alignment = _align_r
    else:
        for col in range(3, 12):
            cell = ws.cell(tot_row, col, 0)
            cell.font = _font_bold
            cell.fill = _fill_tot
            cell.border = _thin
            cell.number_format = _fmt_num

    # Filas de control vs PDF
    row_tp = tot_row + 2
    row_np = tot_row + 3
    ws.cell(row_tp, 1, "Total presentado (según resumen PDF)").font = _font_normal
    c = ws.cell(row_tp, 3, r.total_presentado_pdf)
    c.number_format = _fmt_num
    c.font = _font_normal
    ws.cell(row_tp, 4, "Ventas calculadas (Col. C)").font = _font_normal
    ws.cell(row_tp, 6, f"=C{tot_row}").number_format = _fmt_num
    ws.cell(row_tp, 8, "Diferencia").font = _font_normal
    d = ws.cell(row_tp, 11, f"=C{row_tp}-F{row_tp}")
    d.number_format = _fmt_num

    ws.cell(row_np, 1, "Neto de pagos (según resumen PDF)").font = _font_normal
    c = ws.cell(row_np, 3, r.neto_pagos_pdf)
    c.number_format = _fmt_num
    c.font = _font_normal
    ws.cell(row_np, 4, "Neto calculado (Col. K)").font = _font_normal
    ws.cell(row_np, 6, f"=K{tot_row}").number_format = _fmt_num
    ws.cell(row_np, 8, "Diferencia").font = _font_normal
    d = ws.cell(row_np, 11, f"=C{row_np}-F{row_np}")
    d.number_format = _fmt_num

    ws.freeze_panes = "A5"
    anchos = [14, 14, 14, 12, 14, 16, 14, 14, 14, 14, 16]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return tot_row


def _escribir_resumen(
    ws,
    resumenees: list[ResumenPdf],
    filas_totales: dict[str, int],
) -> None:
    r0 = resumenees[0]
    fant = f" ({r0.nombre_fantasia})" if r0.nombre_fantasia else ""
    ws["A1"] = (
        f"Resumen Liquidaciones Tarjeta - {r0.razon_social}{fant} "
        f"- CUIT {r0.cuit} - {r0.mes_nombre} {r0.anio}"
    )
    ws["A1"].font = _font_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws["A2"] = (
        f"N° Comercio: {r0.nro_comercio}"
        + (f" | Domicilio: {r0.domicilio}" if r0.domicilio else "")
    )
    ws["A2"].font = _font_sub
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    header_row = 3
    for col, h in enumerate(HEADERS_RESUMEN, 1):
        cell = ws.cell(header_row, col, h)
        cell.fill = _fill_header
        cell.font = _font_header
        cell.alignment = _align_c
        cell.border = _thin

    # Columnas de detalle: C..K = ventas..neto → Resumen B..J
    # Mapeo: detalle col 3..11 → resumen col 2..10
    first_data = 4
    for i, r in enumerate(resumenees):
        row = first_data + i
        hoja = r.hoja
        tot = filas_totales[hoja]
        ws.cell(row, 1, hoja).font = _font_normal
        ws.cell(row, 1).border = _thin
        ws.cell(row, 1).alignment = _align_l
        for j, det_col in enumerate(range(3, 12), start=2):
            letter = get_column_letter(det_col)
            # openpyxl: quote sheet names with spaces
            formula = f"='{hoja}'!{letter}{tot}"
            cell = ws.cell(row, j, formula)
            cell.font = _font_normal
            cell.border = _thin
            cell.number_format = _fmt_num
            cell.alignment = _align_r

    last_data = first_data + len(resumenees) - 1
    tot_row = last_data + 1
    ws.cell(tot_row, 1, "TOTAL GENERAL").font = _font_bold
    for col in range(1, 11):
        cell = ws.cell(tot_row, col)
        cell.fill = _fill_tot
        cell.border = _thin
        cell.font = _font_bold
    for col in range(2, 11):
        letter = get_column_letter(col)
        cell = ws.cell(
            tot_row, col, f"=SUM({letter}{first_data}:{letter}{last_data})"
        )
        cell.fill = _fill_tot
        cell.font = _font_bold
        cell.border = _thin
        cell.number_format = _fmt_num
        cell.alignment = _align_r

    ctrl_row = tot_row + 2
    ws.cell(
        ctrl_row,
        1,
        "Control: Neto calculado (Bruto - deducciones) vs Neto informado (Col. Importe Neto)",
    ).font = _font_normal
    ws.merge_cells(
        start_row=ctrl_row, start_column=1, end_row=ctrl_row, end_column=6
    )
    # (B - C - D - E - F - G - H - I) - J
    formula = (
        f"=(B{tot_row}-C{tot_row}-D{tot_row}-E{tot_row}-F{tot_row}"
        f"-G{tot_row}-H{tot_row}-I{tot_row})-J{tot_row}"
    )
    c = ws.cell(ctrl_row, 7, formula)
    c.font = _font_bold
    c.number_format = _fmt_num

    ws.freeze_panes = "A4"
    anchos = [20, 14, 12, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _asignar_nombres_hoja(resumenes: list[ResumenPdf]) -> list[tuple[ResumenPdf, str]]:
    """Evita colisiones marca+modalidad; retorna (resumen, nombre_hoja)."""
    seen: dict[str, int] = {}
    out: list[tuple[ResumenPdf, str]] = []
    for r in resumenes:
        base = sane_sheet_name(f"{r.marca} {r.modalidad}")
        n = seen.get(base, 0) + 1
        seen[base] = n
        if n > 1:
            print(
                f"ADVERTENCIA: hoja duplicada '{base}' ({r.path.name}). "
                f"Se usará sufijo ({n})."
            )
            nombre = sane_sheet_name(f"{base} ({n})")
        else:
            nombre = base
        out.append((r, nombre))
    # Mantener el orden de los PDF recibidos (no reordenar alfabéticamente)
    return out


def generar_excel(resumenes: list[ResumenPdf], salida: Path) -> Path:
    pares = _asignar_nombres_hoja(resumenes)

    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"

    filas_totales: dict[str, int] = {}
    resumenes_orden: list[ResumenPdf] = []
    # Temporal: inyectar nombre de hoja vía atributo usado por _escribir_resumen
    for r, nombre in pares:
        r._sheet_name = nombre  # type: ignore[attr-defined]
        ws = wb.create_sheet(title=nombre)
        filas_totales[nombre] = _escribir_detalle(ws, r)
        resumenes_orden.append(r)

    _escribir_resumen(ws_res, resumenes_orden, filas_totales)

    r0 = resumenes[0]
    ws_m = wb.create_sheet(title="Metadatos")
    ws_m["A1"] = "Metadatos del comercio (extraídos del PDF)"
    ws_m["A1"].font = _font_title
    meta = [
        ("Razón social", r0.razon_social),
        ("Nombre de fantasía", r0.nombre_fantasia),
        ("CUIT", r0.cuit),
        ("N° de Comercio", r0.nro_comercio),
        ("Domicilio", r0.domicilio),
        ("Período", f"{r0.mes_nombre} {r0.anio}"),
        ("PDFs procesados", len(resumenes)),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws_m.cell(i, 1, k).font = _font_bold
        ws_m.cell(i, 2, v).font = _font_normal
    ws_m.column_dimensions["A"].width = 22
    ws_m.column_dimensions["B"].width = 50

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(salida))
    return salida

def recalcular_formulas_excel(path: Path) -> None:
    """Abre el libro con Excel (COM) para cachear valores de fórmulas."""
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        print(
            f"ADVERTENCIA: no se pudo recalcular fórmulas (pywin32 ausente: {exc}). "
            "Abrí el archivo en Excel para actualizar."
        )
        return

    excel = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(path.resolve()))
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
        print(f"Fórmulas recalculadas con Excel: {path}")
    except Exception as exc:
        print(f"ADVERTENCIA: recálculo Excel falló ({exc}). Abrí el archivo manualmente.")
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Orquestación / CLI
# ---------------------------------------------------------------------------


def listar_pdfs(carpeta: Path | None, pdfs: Iterable[str] | None) -> list[Path]:
    paths: list[Path] = []
    if pdfs:
        paths.extend(Path(p) for p in pdfs)
    if carpeta:
        paths.extend(sorted(carpeta.glob("*.pdf")))
        paths.extend(sorted(carpeta.glob("*.PDF")))
    # únicos preservando orden
    seen: set[str] = set()
    out: list[Path] = []
    for p in paths:
        key = str(p.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        if not p.is_file():
            raise FileNotFoundError(f"PDF no encontrado: {p}")
        out.append(p)
    return out


@dataclass
class ResultadoLiquidacionesFiserv:
    """Resultado usable desde CLI o desde la web Streamlit."""

    ok: bool
    excel_bytes: bytes | None = None
    nombre_archivo: str = ""
    mensajes: list[str] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)
    razon_social: str = ""
    nombre_fantasia: str = ""
    cuit: str = ""
    nro_comercio: str = ""
    periodo: str = ""
    hojas: list[str] = field(default_factory=list)
    error: str = ""


def procesar_pdfs_fiserv(
    pdfs: list[Path] | list[str],
    *,
    mes_cli: str | None = None,
    anio_cli: int | None = None,
    recalcular: bool = True,
    log: bool = False,
) -> ResultadoLiquidacionesFiserv:
    """
    Procesa PDFs Fiserv/First Data y devuelve el Excel en memoria.
    No lanza SystemExit: errores van en .error / .ok=False.
    """
    paths = [Path(p) for p in pdfs]
    if not paths:
        return ResultadoLiquidacionesFiserv(
            ok=False, error="No se encontraron PDF para procesar."
        )

    def _log(msg: str) -> None:
        if log:
            print(msg)

    mensajes: list[str] = []
    advertencias: list[str] = []
    resumenes: list[ResumenPdf] = []

    _log(f"Procesando {len(paths)} PDF(s)...")
    try:
        for p in paths:
            _log(f"  - {p.name}")
            r = parsear_pdf(p)
            _log(
                f"      -> {r.marca} {r.modalidad} | CUIT {r.cuit} | "
                f"{len(r.liquidaciones)} liq. | marca via {r.marca_metodo}"
            )
            for a in r.advertencias:
                advertencias.append(a)
                _log(f"      ADVERTENCIA: {a}")
            resumenes.append(r)
    except Exception as exc:
        return ResultadoLiquidacionesFiserv(ok=False, error=str(exc))

    cuits = {r.cuit for r in resumenes}
    if len(cuits) > 1:
        detalle = ", ".join(f"{r.path.name}={r.cuit}" for r in resumenes)
        return ResultadoLiquidacionesFiserv(
            ok=False,
            error=f"Los PDF no corresponden al mismo CUIT ({detalle}).",
        )

    periodos = {(r.mes_num, r.anio) for r in resumenes}
    if len(periodos) > 1:
        detalle = ", ".join(
            f"{r.path.name}={r.mes_nombre} {r.anio}" for r in resumenes
        )
        return ResultadoLiquidacionesFiserv(
            ok=False,
            error=f"Los PDF corresponden a períodos distintos ({detalle}).",
        )

    r0 = resumenes[0]
    if mes_cli:
        mes_ok = False
        for num, nom in MESES_ES.items():
            if mes_cli.strip().lower() == nom.lower():
                if num != r0.mes_num:
                    return ResultadoLiquidacionesFiserv(
                        ok=False,
                        error=(
                            f"--mes {mes_cli} no coincide con el PDF "
                            f"({r0.mes_nombre})."
                        ),
                    )
                mes_ok = True
                break
        if not mes_ok:
            return ResultadoLiquidacionesFiserv(
                ok=False, error=f"--mes no reconocido: {mes_cli}"
            )
    if anio_cli is not None and anio_cli != r0.anio:
        return ResultadoLiquidacionesFiserv(
            ok=False,
            error=f"--anio {anio_cli} no coincide con el PDF ({r0.anio}).",
        )

    todo_ok = True
    _log("\n=== Validaciones ===")
    for r in resumenes:
        vr = validar_resumen(r)
        mensajes.extend(vr.mensajes)
        for m in vr.mensajes:
            _log(m)
        if not vr.ok:
            todo_ok = False

    fant = f" ({r0.nombre_fantasia})" if r0.nombre_fantasia else ""
    periodo = f"{r0.mes_nombre} {r0.anio}"
    _log(
        f"\nComercio detectado: {r0.razon_social}{fant} | CUIT {r0.cuit} | "
        f"N° Comercio {r0.nro_comercio}"
    )
    _log(f"Período: {periodo}")

    if not todo_ok:
        return ResultadoLiquidacionesFiserv(
            ok=False,
            mensajes=mensajes,
            advertencias=advertencias,
            razon_social=r0.razon_social,
            nombre_fantasia=r0.nombre_fantasia,
            cuit=r0.cuit,
            nro_comercio=r0.nro_comercio,
            periodo=periodo,
            error=(
                "Las validaciones no cierran (tolerancia $0,01). "
                "No se genera el Excel."
            ),
        )

    nombre = (
        f"Liquidaciones_Tarjeta_"
        f"{sane_filename(r0.razon_social)}_"
        f"{r0.mes_nombre}{r0.anio}.xlsx"
    )
    with tempfile.TemporaryDirectory(prefix="fiserv_liq_") as tmp:
        salida = Path(tmp) / nombre
        generar_excel(resumenes, salida)
        hojas = ["Resumen"] + [r.hoja for r in resumenes] + ["Metadatos"]
        if recalcular:
            recalcular_formulas_excel(salida)
        data = salida.read_bytes()

    return ResultadoLiquidacionesFiserv(
        ok=True,
        excel_bytes=data,
        nombre_archivo=nombre,
        mensajes=mensajes,
        advertencias=advertencias,
        razon_social=r0.razon_social,
        nombre_fantasia=r0.nombre_fantasia,
        cuit=r0.cuit,
        nro_comercio=r0.nro_comercio,
        periodo=periodo,
        hojas=hojas,
    )


def procesar(
    pdfs: list[Path],
    carpeta_salida: Path,
    mes_cli: str | None = None,
    anio_cli: int | None = None,
) -> Path | None:
    res = procesar_pdfs_fiserv(
        pdfs,
        mes_cli=mes_cli,
        anio_cli=anio_cli,
        recalcular=True,
        log=True,
    )
    if res.error and not res.ok:
        print(f"\nERROR: {res.error}")
    if not res.ok or not res.excel_bytes:
        return None
    carpeta_salida.mkdir(parents=True, exist_ok=True)
    salida = carpeta_salida / res.nombre_archivo
    salida.write_bytes(res.excel_bytes)
    print(f"\nExcel generado: {salida}")
    return salida


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Convierte resúmenes PDF Fiserv/First Data en Excel de liquidaciones "
            "de tarjeta (genérico por comercio/marca/modalidad)."
        )
    )
    p.add_argument(
        "--carpeta-pdfs",
        type=Path,
        help="Carpeta con los PDF del mismo comercio y período",
    )
    p.add_argument(
        "--carpeta-salida",
        type=Path,
        required=True,
        help="Carpeta donde guardar el Excel",
    )
    p.add_argument(
        "--pdf",
        action="append",
        dest="pdfs",
        help="PDF individual (repetible). Se puede combinar con --carpeta-pdfs",
    )
    p.add_argument("--mes", help="Mes esperado (ej. Julio). Opcional, valida contra PDF")
    p.add_argument("--anio", type=int, help="Año esperado (ej. 2026). Opcional")
    return p


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    args = build_parser().parse_args(argv)
    if not args.carpeta_pdfs and not args.pdfs:
        print("ERROR: indica --carpeta-pdfs y/o --pdf", file=sys.stderr)
        return 2
    try:
        pdfs = listar_pdfs(args.carpeta_pdfs, args.pdfs)
        out = procesar(
            pdfs,
            Path(args.carpeta_salida),
            mes_cli=args.mes,
            anio_cli=args.anio,
        )
        return 0 if out else 1
    except SystemExit as e:
        if isinstance(e.code, int):
            return e.code
        print(str(e), file=sys.stderr)
        return 1
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
