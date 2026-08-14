# -*- coding: utf-8 -*-
"""Extractos FCI (cualquier banco) → Excel.

Salida fija:
Fecha | Descripcion | Cantidad de Cuotas | Valor CC | Total
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from excel_formato_estudio import (
    construir_informe_excel,
    guardar_informe_excel,
    informe_a_bytes,
)

_MONEY = Decimal("0.01")
_CUOTA = Decimal("0.01")
_VALOR_CC = Decimal("0.000001")

COLUMNAS = [
    "Fecha",
    "Descripcion",
    "Cantidad de Cuotas",
    "Valor CC",
    "Total",
    "Fondo",
    "Banco",
    "Periodo extracto",
    "Archivo",
]

TIPOS = (
    "RESCATE",
    "SUSCRIPCION",
    "SUSCRIPCIÓN",
    "SUSCRIPCION DE CUOTAS",
    "RESCATE DE CUOTAS",
    "COMPRA",
    "VENTA",
)

RE_FECHA = re.compile(r"\d{2}/\d{2}/\d{4}")
RE_NUM = re.compile(
    r"-?\d{1,3}(?:\.\d{3})*(?:,\d{2,8})|-?\d+,\d{2,8}|-?\d+\.\d{4,8}"
)
RE_MOV_GALICIA = re.compile(
    r"(?P<fecha>\d{2}/\d{2}/\d{4})\s+"
    r"(?P<desc>RESCATE|SUSCRIPCION|SUSCRIPCI[OÓ]N|COMPRA|VENTA)\s+"
    r"(?P<cuotas>-?\d{1,3}(?:\.\d{3})*,\d+)\s+"
    r"\$?\s*(?P<valor>-?\d{1,3}(?:\.\d{3})*,\d+)\s+"
    r"\$?\s*(?P<total>-?\d{1,3}(?:\.\d{3})*,\d+)",
    re.I,
)
RE_MOV_FLEX = re.compile(
    r"(?P<fecha>\d{2}/\d{2}/\d{4})\s+"
    r"(?P<desc>RESCATE(?:\s+DE\s+CUOTAS)?|SUSCRIPCI[OÓ]N(?:\s+DE\s+CUOTAS)?|COMPRA|VENTA)\b"
    r"(?P<rest>[^\n]{0,180})",
    re.I,
)

VALOR_CC_FMT = '_("$"* #,##0.000000_);_("$"* (#,##0.000000);_("$"* "-"??_);_(@_)'
CANTIDAD_FMT = "#,##0.00"

_BANCOS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Galicia", ("banco galicia", "galicia", "fima premium", "fima-fondos", "fima ")),
    ("Santander", ("santander", "superfondo", "super ahorro")),
    ("BBVA", ("bbva", "banco frances", "banco francés", " fba ")),
    ("Macro", ("banco macro", "macro fondos")),
    ("ICBC", ("icbc",)),
    ("HSBC", ("hsbc",)),
    ("Supervielle", ("supervielle",)),
    ("Nación", ("banco nacion", "banco nación", "nacion fondos")),
    ("Credicoop", ("credicoop",)),
    ("Patagonia", ("banco patagonia", "patagonia")),
    ("Ciudad", ("banco ciudad",)),
    ("Comafi", ("comafi",)),
    ("Bind", ("banco industrial", " bind ")),
)


def _num_ar(raw: Any, q: Decimal) -> Decimal:
    if isinstance(raw, Decimal):
        return raw.quantize(q, rounding=ROUND_HALF_UP)
    if isinstance(raw, (int, float)) and not pd.isna(raw):
        return Decimal(str(raw)).quantize(q, rounding=ROUND_HALF_UP)
    s = str(raw or "").strip().replace("$", "").replace(" ", "")
    if not s or s.lower() in {"nan", "none", "-"}:
        return Decimal("0")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    return Decimal(s).quantize(q, rounding=ROUND_HALF_UP)


def extraer_texto_pdf(data: bytes) -> str:
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(data)) as pdf:
            partes = [(page.extract_text() or "") for page in pdf.pages]
        texto = "\n".join(partes)
        if texto.strip():
            return texto
    except Exception:
        pass
    try:
        import fitz

        doc = fitz.open(stream=data, filetype="pdf")
        partes = [page.get_text("text") or "" for page in doc]
        doc.close()
        return "\n".join(partes)
    except Exception:
        return ""


def detectar_banco(texto: str, nombre: str = "") -> str:
    blob = f"{nombre}\n{texto[:4000]}".lower()
    for banco, keys in _BANCOS:
        if any(k in blob for k in keys):
            return banco
    return "Desconocido"


def _norm_desc(raw: str) -> str:
    t = re.sub(r"\s+", " ", (raw or "").upper()).strip()
    t = t.replace("SUSCRIPCIÓN", "SUSCRIPCION")
    if "SUSCRIP" in t:
        return "SUSCRIPCION"
    if "RESCATE" in t:
        return "RESCATE"
    if t.startswith("COMPRA"):
        return "COMPRA"
    if t.startswith("VENTA"):
        return "VENTA"
    return t[:40] if t else ""


def _periodos_en(texto: str) -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []
    for m in re.finditer(
        r"(?:Movimientos|Operaciones|Per[ií]odo)\s*(?:/\s*Operaciones)?\s*"
        r"(?:del\s+)?(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})",
        texto,
        re.I,
    ):
        out.append((m.start(), f"{m.group(1)} al {m.group(2)}"))
    return out


def _periodo_en(cortes: list[tuple[int, str]], pos: int) -> str:
    actual = ""
    for start, per in cortes:
        if start <= pos:
            actual = per
        else:
            break
    return actual


def _fondo_en(texto: str) -> str:
    m = re.search(r"FONDO\s*[-:]\s*([^\n]{3,80})", texto, re.I)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    m = re.search(
        r"\b((?:FIMA|SUPERFONDO|FBA|PIONEER|DELTA|PREMIER|ADCAP|SBS|COMPASS)\s+[A-ZÁÉÍÓÚÑ0-9 ]{3,50})",
        texto,
        re.I,
    )
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def _cuenta_en(texto: str) -> str:
    m = re.search(
        r"(?:Cuenta\s+Comitente|Nro\.?\s*Cuenta|Cuenta)\s*:?\s*(\d{4,})",
        texto,
        re.I,
    )
    return m.group(1) if m else ""


def extraer_posicion_fci(texto: str) -> dict[str, Any]:
    """
    Intenta leer la posición / saldo final del extracto de fondos:
    cuotas, valor cuota e importe en pesos.
    """
    out: dict[str, Any] = {
        "cuotas": None,
        "valor_cuota": None,
        "importe": None,
        "fuente": "",
    }
    if not (texto or "").strip():
        return out

    patrones = [
        # Cantidad … Valor (de la) cuota … Importe / Total
        re.compile(
            r"(?:Cantidad(?:\s+de)?\s+cuot(?:as|apartes)?|Saldo(?:\s+de)?\s+cuot(?:as|apartes)?|"
            r"Posici[oó]n(?:\s+al\s+cierre)?|Tenencia)"
            r"[^\d]{0,80}"
            r"(?P<cuotas>\d{1,3}(?:\.\d{3})*,\d+|\d+,\d+)"
            r".{0,120}?"
            r"(?:Valor\s+(?:de\s+la\s+)?cuota|Valor\s*CC|Cotizaci[oó]n)"
            r"[^\d]{0,40}"
            r"\$?\s*(?P<vc>\d{1,3}(?:\.\d{3})*,\d{4,8}|\d+,\d{4,8}|\d+\.\d{4,8})"
            r"(?:.{0,120}?"
            r"(?:Importe|Total|Valuaci[oó]n|Monto)"
            r"[^\d]{0,40}"
            r"\$?\s*(?P<imp>\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}))?",
            re.I | re.S,
        ),
        # Valor cuota primero, después cantidad
        re.compile(
            r"(?:Valor\s+(?:de\s+la\s+)?cuota|Valor\s*CC)"
            r"[^\d]{0,40}"
            r"\$?\s*(?P<vc>\d{1,3}(?:\.\d{3})*,\d{4,8}|\d+,\d{4,8}|\d+\.\d{4,8})"
            r".{0,120}?"
            r"(?:Cantidad(?:\s+de)?\s+cuot(?:as|apartes)?|Saldo(?:\s+de)?\s+cuot)"
            r"[^\d]{0,40}"
            r"(?P<cuotas>\d{1,3}(?:\.\d{3})*,\d+|\d+,\d+)",
            re.I | re.S,
        ),
    ]
    for pat in patrones:
        matches = list(pat.finditer(texto))
        if not matches:
            continue
        m = matches[-1]  # suele estar al cierre
        cuotas = _num_ar(m.group("cuotas"), _CUOTA)
        vc = _num_ar(m.group("vc"), _VALOR_CC)
        imp = None
        if "imp" in m.groupdict() and m.group("imp"):
            imp = _num_ar(m.group("imp"), _MONEY)
        if cuotas > 0 and vc > 0:
            if imp is None:
                imp = (cuotas * vc).quantize(_MONEY, rounding=ROUND_HALF_UP)
            out["cuotas"] = float(cuotas)
            out["valor_cuota"] = float(vc)
            out["importe"] = float(imp)
            out["fuente"] = "posicion_extracto"
            return out

    # Fallback: último valor cuota con muchos decimales + cantidad cercana
    m_vc = None
    for m in re.finditer(
        r"(?:Valor\s+(?:de\s+la\s+)?cuota|Valor\s*CC|Cotizaci[oó]n)\s*:?\s*\$?\s*"
        r"(\d{1,3}(?:\.\d{3})*,\d{4,8}|\d+,\d{4,8}|\d+\.\d{4,8})",
        texto,
        re.I,
    ):
        m_vc = m
    if m_vc:
        vc = _num_ar(m_vc.group(1), _VALOR_CC)
        ventana = texto[max(0, m_vc.start() - 200) : m_vc.end() + 200]
        m_c = re.search(
            r"(?:Cantidad|Cuotapartes|Saldo)[^\d]{0,40}(\d{1,3}(?:\.\d{3})*,\d+|\d+,\d+)",
            ventana,
            re.I,
        )
        if m_c and vc > 0:
            cuotas = _num_ar(m_c.group(1), _CUOTA)
            if cuotas > 0:
                out["cuotas"] = float(cuotas)
                out["valor_cuota"] = float(vc)
                out["importe"] = float((cuotas * vc).quantize(_MONEY, rounding=ROUND_HALF_UP))
                out["fuente"] = "posicion_parcial"
                return out
    return out


def _asignar_numeros(nums: list[Decimal]) -> tuple[Decimal, Decimal, Decimal] | None:
    """(cuotas, valor_cc, total) a partir de 2 o 3 números de la línea."""
    if len(nums) < 2:
        return None
    if len(nums) > 3:
        nums = nums[:3]
    if len(nums) == 2:
        a, b = nums
        # valor cuota suele tener más decimales y magnitud chica
        if a.copy_abs() < Decimal("10000") and b.copy_abs() >= a.copy_abs():
            valor, total = a, b
        elif b.copy_abs() < Decimal("10000"):
            valor, total = b, a
        else:
            valor, total = a, b
        if valor == 0:
            return None
        cuotas = (total / valor).quantize(_CUOTA, rounding=ROUND_HALF_UP)
        return cuotas, valor.quantize(_VALOR_CC), total.quantize(_MONEY)

    # 3 números: el de más decimales y magnitud chica = Valor CC
    scored = []
    for n in nums:
        decs = max(0, -n.as_tuple().exponent)
        scored.append((decs, 1 if n.copy_abs() < Decimal("100000") else 0, n))
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    valor = scored[0][2]
    resto = [n for n in nums if n is not valor]
    resto.sort(key=lambda x: x.copy_abs(), reverse=True)
    total, cuotas = resto[0], resto[1]
    # Si se invirtieron cuotas/total (total chico, cuotas enorme irreal)
    if cuotas.copy_abs() > total.copy_abs() and total.copy_abs() > Decimal("100"):
        cuotas, total = total, cuotas
    return (
        cuotas.quantize(_CUOTA),
        valor.quantize(_VALOR_CC),
        total.quantize(_MONEY),
    )


def _fila(
    *,
    fecha: date,
    desc: str,
    cuotas: Decimal,
    valor: Decimal,
    total: Decimal,
    fondo: str,
    banco: str,
    periodo: str,
    archivo: str,
) -> dict[str, Any]:
    return {
        "Fecha": fecha,
        "Descripcion": desc,
        "Cantidad de Cuotas": float(cuotas),
        "Valor CC": float(valor),
        "Total": float(total),
        "Fondo": fondo,
        "Banco": banco,
        "Periodo extracto": periodo,
        "Archivo": archivo,
    }


def _parse_lineas(texto: str, *, banco: str, fondo: str, archivo: str) -> list[dict]:
    cortes = _periodos_en(texto)
    filas: list[dict] = []
    for m in RE_MOV_GALICIA.finditer(texto):
        desc = _norm_desc(m.group("desc"))
        fecha = datetime.strptime(m.group("fecha"), "%d/%m/%Y").date()
        filas.append(
            _fila(
                fecha=fecha,
                desc=desc,
                cuotas=_num_ar(m.group("cuotas"), _CUOTA),
                valor=_num_ar(m.group("valor"), _VALOR_CC),
                total=_num_ar(m.group("total"), _MONEY),
                fondo=fondo,
                banco=banco,
                periodo=_periodo_en(cortes, m.start()),
                archivo=archivo,
            )
        )
    if filas:
        return filas

    for m in RE_MOV_FLEX.finditer(texto):
        desc = _norm_desc(m.group("desc"))
        fecha = datetime.strptime(m.group("fecha"), "%d/%m/%Y").date()
        nums = [_num_ar(x, Decimal("0.000001")) for x in RE_NUM.findall(m.group("rest") or "")]
        nums = [n for n in nums if n != 0]
        assigned = _asignar_numeros(nums)
        if not assigned:
            continue
        cuotas, valor, total = assigned
        filas.append(
            _fila(
                fecha=fecha,
                desc=desc,
                cuotas=cuotas,
                valor=valor,
                total=total,
                fondo=fondo,
                banco=banco,
                periodo=_periodo_en(cortes, m.start()),
                archivo=archivo,
            )
        )
    return filas


def _map_col(headers: list[str]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        n = re.sub(r"\s+", " ", str(h or "").lower())
        n = (
            n.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
        )
        if "fecha" in n and "liq" not in n and i not in idx.values():
            idx.setdefault("fecha", i)
        elif any(k in n for k in ("descrip", "concepto", "tipo", "operacion", "movim")):
            idx.setdefault("desc", i)
        elif any(k in n for k in ("cantidad", "cuota", "cuotapart")) and "valor" not in n:
            idx.setdefault("cuotas", i)
        elif any(k in n for k in ("valor", "vcp", "precio", "cotiz")):
            idx.setdefault("valor", i)
        elif any(k in n for k in ("monto", "total", "importe", "neto")):
            idx.setdefault("total", i)
    return idx


def _parse_tablas(data: bytes, *, banco: str, fondo: str, archivo: str) -> list[dict]:
    try:
        import pdfplumber
    except Exception:
        return []
    filas: list[dict] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            for tab in page.extract_tables() or []:
                if not tab or len(tab) < 2:
                    continue
                headers = [str(c or "") for c in tab[0]]
                idx = _map_col(headers)
                if "fecha" not in idx or len(idx) < 3:
                    # a veces el header está en la 2ª fila
                    idx = _map_col([str(c or "") for c in tab[1]])
                    body = tab[2:]
                else:
                    body = tab[1:]
                if "fecha" not in idx:
                    continue
                for row in body:
                    if not row or idx["fecha"] >= len(row):
                        continue
                    raw_f = str(row[idx["fecha"]] or "")
                    mf = RE_FECHA.search(raw_f)
                    if not mf:
                        continue
                    fecha = datetime.strptime(mf.group(0), "%d/%m/%Y").date()
                    desc = _norm_desc(str(row[idx["desc"]]) if "desc" in idx else "")
                    if not desc:
                        continue
                    cuotas = _num_ar(row[idx["cuotas"]], _CUOTA) if "cuotas" in idx else Decimal("0")
                    valor = _num_ar(row[idx["valor"]], _VALOR_CC) if "valor" in idx else Decimal("0")
                    total = _num_ar(row[idx["total"]], _MONEY) if "total" in idx else Decimal("0")
                    if total == 0 and cuotas and valor:
                        total = (cuotas * valor).quantize(_MONEY)
                    if cuotas == 0 and total and valor:
                        cuotas = (total / valor).quantize(_CUOTA)
                    if total == 0 and cuotas == 0:
                        continue
                    filas.append(
                        _fila(
                            fecha=fecha,
                            desc=desc,
                            cuotas=cuotas,
                            valor=valor,
                            total=total,
                            fondo=fondo,
                            banco=banco,
                            periodo="",
                            archivo=archivo,
                        )
                    )
    return filas


def parsear_pdf_fci(data: bytes, nombre: str = "extracto.pdf") -> dict[str, Any]:
    texto = extraer_texto_pdf(data)
    banco = detectar_banco(texto, nombre)
    fondo = _fondo_en(texto)
    cuenta = _cuenta_en(texto)
    cortes = _periodos_en(texto)
    periodo = f"{cortes[0][1].split(' al ')[0]} al {cortes[-1][1].split(' al ')[-1]}" if cortes else ""
    posicion = extraer_posicion_fci(texto)

    if not texto.strip():
        return {
            "ok": False,
            "requiere_ocr": True,
            "banco": banco,
            "fondo": fondo,
            "cuenta": cuenta,
            "periodo": periodo,
            "posicion": posicion,
            "filas": [],
            "error": "PDF sin texto (escaneado) — requiere OCR/manual",
            "archivo": nombre,
        }

    filas = _parse_lineas(texto, banco=banco, fondo=fondo, archivo=nombre)
    if not filas:
        filas = _parse_tablas(data, banco=banco, fondo=fondo, archivo=nombre)

    # Si no hay posición explícita, usar último valor cuota de movimientos como pista
    if (not posicion.get("valor_cuota")) and filas:
        ultimo = max(filas, key=lambda r: (r.get("Fecha") or date.min, float(r.get("Valor CC") or 0)))
        vc = float(ultimo.get("Valor CC") or 0)
        if vc > 0:
            posicion = {
                "cuotas": posicion.get("cuotas"),
                "valor_cuota": vc,
                "importe": posicion.get("importe"),
                "fuente": "ultimo_movimiento",
            }

    if not filas and not (posicion.get("cuotas") and posicion.get("valor_cuota")):
        return {
            "ok": False,
            "requiere_ocr": False,
            "banco": banco,
            "fondo": fondo,
            "cuenta": cuenta,
            "periodo": periodo,
            "posicion": posicion,
            "filas": [],
            "error": "No se leyeron movimientos de FCI",
            "archivo": nombre,
        }
    return {
        "ok": True,
        "requiere_ocr": False,
        "banco": banco,
        "fondo": fondo,
        "cuenta": cuenta,
        "periodo": periodo,
        "posicion": posicion,
        "filas": filas,
        "error": None,
        "archivo": nombre,
    }


def _dedupar(filas: list[dict]) -> list[dict]:
    vistos: set[tuple] = set()
    out: list[dict] = []
    for f in filas:
        clave = (
            f.get("Fecha"),
            f.get("Descripcion"),
            f.get("Cantidad de Cuotas"),
            f.get("Valor CC"),
            f.get("Total"),
            f.get("Fondo") or "",
        )
        if clave in vistos:
            continue
        vistos.add(clave)
        out.append(f)
    out.sort(key=lambda r: (r.get("Fecha") or date.min, str(r.get("Descripcion") or "")))
    return out


def _leer_entrada(item: Any) -> tuple[str, bytes]:
    if isinstance(item, (str, Path)):
        p = Path(item)
        return p.name, p.read_bytes()
    nombre = str(getattr(item, "name", "extracto.pdf"))
    data = item.read() if hasattr(item, "read") else item
    if hasattr(item, "seek"):
        try:
            item.seek(0)
        except Exception:
            pass
    if isinstance(data, str):
        data = data.encode("utf-8")
    return nombre, bytes(data)


def procesar_pdfs_fci(entradas: Iterable[Any]) -> dict[str, Any]:
    todas: list[dict] = []
    fallidos: list[dict] = []
    ocr: list[dict] = []
    bancos: set[str] = set()
    fondos: set[str] = set()
    cuentas: set[str] = set()
    periodos: list[str] = []
    ok_n = 0

    for item in entradas or []:
        nombre, data = _leer_entrada(item)
        res = parsear_pdf_fci(data, nombre)
        if res.get("banco") and res["banco"] != "Desconocido":
            bancos.add(res["banco"])
        if res.get("fondo"):
            fondos.add(res["fondo"])
        if res.get("cuenta"):
            cuentas.add(res["cuenta"])
        if res.get("periodo"):
            periodos.append(res["periodo"])
        if res.get("requiere_ocr"):
            ocr.append({"archivo": nombre, "motivo": res.get("error")})
            continue
        if not res.get("ok"):
            fallidos.append({"archivo": nombre, "motivo": res.get("error") or "Error"})
            continue
        ok_n += 1
        todas.extend(res["filas"])

    df = pd.DataFrame(_dedupar(todas))
    if not df.empty:
        cols = [c for c in COLUMNAS if c in df.columns]
        df = df[cols]

    periodo = ""
    if periodos:
        periodo = f"{periodos[0].split(' al ')[0]} al {periodos[-1].split(' al ')[-1]}"

    return {
        "detalle": df,
        "resumen": {
            "pdfs_ok": ok_n,
            "filas": len(df),
            "bancos": ", ".join(sorted(bancos)) or "—",
            "fondos": ", ".join(sorted(fondos)) or "—",
            "cuenta": next(iter(cuentas), "—"),
            "periodo": periodo,
            "pdfs_fallidos": len(fallidos),
            "pdfs_ocr": len(ocr),
        },
        "fallidos": fallidos,
        "requiere_ocr": ocr,
        "meta": {
            "banco": ", ".join(sorted(bancos)),
            "fondo": ", ".join(sorted(fondos)),
            "cuenta": next(iter(cuentas), ""),
            "periodo": periodo,
        },
    }


def _resumenes(df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    if df.empty:
        return []
    susc = df.loc[df["Descripcion"] == "SUSCRIPCION", "Total"].sum() if "Descripcion" in df.columns else 0
    resc = df.loc[df["Descripcion"] == "RESCATE", "Total"].sum() if "Descripcion" in df.columns else 0
    tipo = pd.DataFrame(
        [
            {"Descripcion": "SUSCRIPCION", "Movimientos": int((df["Descripcion"] == "SUSCRIPCION").sum()), "Total": float(susc)},
            {"Descripcion": "RESCATE", "Movimientos": int((df["Descripcion"] == "RESCATE").sum()), "Total": float(resc)},
        ]
    )
    hojas = [("Por tipo de movimiento", tipo)]
    if "Fecha" in df.columns:
        por_mes = (
            df.assign(Mes=df["Fecha"].map(lambda d: d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)[:7]))
            .groupby(["Mes", "Descripcion"], as_index=False)
            .agg(Movimientos=("Total", "size"), Total=("Total", "sum"))
            .sort_values(["Mes", "Descripcion"])
        )
        hojas.append(("Por mes", por_mes))
    return hojas


def _aplicar_formatos_ws(ws) -> None:
    headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    col_cant = headers.get("Cantidad de Cuotas")
    col_cc = headers.get("Valor CC")
    for row in range(2, ws.max_row + 1):
        if col_cant:
            ws.cell(row, col_cant).number_format = CANTIDAD_FMT
        if col_cc:
            ws.cell(row, col_cc).number_format = VALOR_CC_FMT


def _aplicar_formatos_wb(wb: Workbook) -> Workbook:
    if "Movimientos" in wb.sheetnames:
        _aplicar_formatos_ws(wb["Movimientos"])
    return wb


def _kpis(resultado: dict) -> list[tuple]:
    df = resultado.get("detalle")
    r = resultado.get("resumen") or {}
    if df is None or df.empty:
        return [
            ("Movimientos", 0, "int"),
            ("Bancos", r.get("bancos") or "—", "text"),
        ]
    susc = float(df.loc[df["Descripcion"] == "SUSCRIPCION", "Total"].sum())
    resc = float(df.loc[df["Descripcion"] == "RESCATE", "Total"].sum())
    return [
        ("Cuenta", r.get("cuenta") or "—", "text"),
        ("Banco", r.get("bancos") or "—", "text"),
        ("Movimientos", len(df), "int"),
        ("Suscripciones", susc, "money"),
        ("Rescates", resc, "money"),
    ]


def exportar_excel_bytes(resultado: dict, *, titulo: str = "", subtitulo: str = "") -> bytes:
    df = resultado.get("detalle")
    if df is None:
        df = pd.DataFrame(columns=COLUMNAS)
    r = resultado.get("resumen") or {}
    wb = construir_informe_excel(
        titulo=titulo or "Extracto FCI",
        subtitulo=subtitulo or r.get("fondos") or "",
        periodo=r.get("periodo") or "",
        kpis=_kpis(resultado),
        resumenes=_resumenes(df),
        detalle=df,
        hoja_detalle="Movimientos",
        col_moneda=["Valor CC", "Total"],
        col_fecha=["Fecha"],
        col_texto=["Periodo extracto", "Banco", "Archivo", "Fondo"],
        total_col="Total",
    )
    _aplicar_formatos_wb(wb)
    return informe_a_bytes(wb)


def extraer_y_guardar(pdf_path: str | Path, xlsx_path: str | Path | None = None) -> Path:
    pdf_path = Path(pdf_path)
    res = procesar_pdfs_fci([pdf_path])
    df = res.get("detalle")
    if df is None or df.empty:
        raise ValueError(f"No se leyeron movimientos en {pdf_path.name}")
    if xlsx_path is None:
        xlsx_path = pdf_path.with_name(f"Movimientos_FCI_{pdf_path.stem}.xlsx")
    xlsx_path = Path(xlsx_path)
    r = res.get("resumen") or {}
    guardar_informe_excel(
        xlsx_path,
        titulo="Extracto FCI",
        subtitulo=r.get("fondos") or pdf_path.name,
        periodo=r.get("periodo") or "",
        kpis=_kpis(res),
        resumenes=_resumenes(df),
        detalle=df,
        hoja_detalle="Movimientos",
        col_moneda=["Valor CC", "Total"],
        col_fecha=["Fecha"],
        col_texto=["Periodo extracto", "Banco", "Archivo", "Fondo"],
        total_col="Total",
    )
    wb = load_workbook(xlsx_path)
    _aplicar_formatos_wb(wb)
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


# Compatibilidad con el módulo Galicia anterior
def parsear_extracto_fci_galicia(texto: str) -> dict[str, Any]:
    banco = detectar_banco(texto)
    fondo = _fondo_en(texto)
    filas = _parse_lineas(texto, banco=banco, fondo=fondo, archivo="")
    cortes = _periodos_en(texto)
    periodo = cortes[0][1] if cortes else ""
    if cortes:
        periodo = f"{cortes[0][1].split(' al ')[0]} al {cortes[-1][1].split(' al ')[-1]}"
    return {
        "fondo": fondo,
        "cuenta": _cuenta_en(texto),
        "posicion": "",
        "periodo": periodo,
        "filas": filas,
    }
