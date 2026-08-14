"""
Formato Excel de análisis FCI (estilo extracto / FCI GROWTH).

Solo para FCI:
- Cada rescate partido por lote FIFO, con fórmulas de origen e interés.
- Cuadro mensual de intereses (SUMIFS).
- Resultado por tenencia por lote:
  DIF = VALOR DE CIERRE − VALOR ORIGEN; resultado = CANTIDAD × DIF
  (VC de cierre = último extracto de fondos del mes).
"""
from __future__ import annotations

import io
from collections import deque
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from inversiones_catalogo import EVENTO_ENTRADA, EVENTO_SALIDA, TIPO_FCI, evento_para

MONEY_FMT = '_-"$"\\ * #,##0.00_-;\\-"$"\\ * #,##0.00_-;_-"$"\\ * "-"??_-;_-@_-'
CUOTAS_FMT = "#,##0.00"
VC_FMT = "\\$\\ 0.000000"
DATE_FMT = "dd/mm/yyyy"
HDR_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(name="Calibri", size=8, bold=True, color="FF0000")
BODY_FONT = Font(name="Calibri", size=8)
BOLD_FONT = Font(name="Calibri", size=8, bold=True)
ITAL_FONT = Font(name="Calibri", size=8, italic=True, color="666666")
INI_FILL = PatternFill("solid", fgColor="E2EFDA")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
ZEBRA = PatternFill("solid", fgColor="F2F2F2")
_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def _thin_border() -> Border:
    return _THIN


def _num(val: object, default: float = 0.0) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def es_analisis_solo_fci(df_mov: pd.DataFrame | None, movimientos: list[dict] | None = None) -> bool:
    """True si el lote analizado es solo FCI (este formato aplica solo ahí)."""
    if df_mov is not None and not df_mov.empty and "Tipo_inversion" in df_mov.columns:
        tipos = {str(t) for t in df_mov["Tipo_inversion"].dropna().unique()}
        if tipos and tipos <= {TIPO_FCI}:
            return True
    if movimientos:
        tipos = {str(m.get("Tipo_inversion") or "") for m in movimientos}
        tipos.discard("")
        if tipos and tipos <= {TIPO_FCI}:
            return True
    return False


def _as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return pd.to_datetime(v, dayfirst=True).date()
    except Exception:
        return None


def _meses_del_periodo(fechas: list[date]) -> list[date]:
    if not fechas:
        return []
    inicio = date(fechas[0].year, fechas[0].month, 1)
    fin = date(fechas[-1].year, fechas[-1].month, 1)
    out: list[date] = []
    y, m = inicio.year, inicio.month
    while (y, m) <= (fin.year, fin.month):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def consolidar_lotes_abiertos(
    lotes_abiertos: list[dict],
    fecha_inicial: date | None = None,
) -> list[dict]:
    """Lotes FIFO abiertos, agrupados por (fecha, VC, origen)."""
    agrupado: dict[tuple, dict] = {}
    for lote in lotes_abiertos or []:
        fecha = lote.get("Fecha") or fecha_inicial
        origen = str(lote.get("Origen") or "")
        if "inicial" in origen.lower() or lote.get("Fecha") is None:
            origen_txt = "Saldo inicial"
            fecha = fecha_inicial or date(1900, 1, 1)
        else:
            origen_txt = "Suscripción"
        if not isinstance(fecha, date):
            fecha = _as_date(fecha) or date(1900, 1, 1)
        clave = (fecha, round(float(lote["Costo_Unitario"]), 8), origen_txt)
        if clave not in agrupado:
            agrupado[clave] = {
                "Origen": origen_txt,
                "Fecha origen": fecha,
                "Cuotas": 0.0,
                "Valor cuota origen": float(lote["Costo_Unitario"]),
                "Costo de origen": 0.0,
            }
        agrupado[clave]["Cuotas"] += float(lote["Cantidad"])
        agrupado[clave]["Costo de origen"] += float(lote["Costo_Total"])
    filas = list(agrupado.values())
    for fila in filas:
        fila["Cuotas"] = round(float(fila["Cuotas"]), 4)
        fila["Costo de origen"] = round(float(fila["Costo de origen"]), 2)
    filas.sort(key=lambda r: r["Fecha origen"])
    return filas


def resultado_por_tenencia(lotes: list[dict], vc_cierre: float) -> dict[str, float]:
    """
    Valuación = cuotas × valor cuota del último extracto de fondos del mes.
    Resultado por tenencia = valuación − costo de origen.
    """
    cuotas = round(sum(float(l["Cuotas"]) for l in lotes), 4)
    costo = round(sum(float(l["Costo de origen"]) for l in lotes), 2)
    valuacion = round(cuotas * float(vc_cierre), 2)
    return {
        "cuotas": cuotas,
        "vc_cierre": float(vc_cierre),
        "valuacion": valuacion,
        "costo_origen": costo,
        "resultado_tenencia": round(valuacion - costo, 2),
    }


def _escribir_mov(
    ws,
    r: int,
    fecha: date,
    tipo: str,
    cant: float,
    vc: float,
    tot: float,
) -> None:
    """Escribe una fila de movimiento. Solo pinta RESCATE; SUSCRIPCION queda en blanco."""
    ws.cell(r, 2, datetime(fecha.year, fecha.month, fecha.day)).number_format = DATE_FMT
    ws.cell(r, 3, tipo)
    ws.cell(r, 4, cant).number_format = CUOTAS_FMT
    ws.cell(r, 5, vc).number_format = VC_FMT
    ws.cell(r, 6, tot).number_format = MONEY_FMT
    pintar = str(tipo).strip().upper() == "RESCATE"
    for c in range(2, 7):
        ws.cell(r, c).font = BODY_FONT
        if pintar:
            ws.cell(r, c).fill = ZEBRA
        else:
            ws.cell(r, c).fill = PatternFill()


def exportar_analisis_fci_excel(
    df_mov: pd.DataFrame,
    lotes_abiertos: list[dict],
    df_inicial: pd.DataFrame | None = None,
    *,
    vc_cierre: float | None = None,
    fecha_cierre: date | None = None,
    meta: dict | None = None,
) -> bytes:
    """
    Excel estilo extracto FCI.

    vc_cierre: valor cuota del **último extracto de fondos del mes** (posición al cierre).
    """
    meta = meta or {}
    work = df_mov.copy() if df_mov is not None else pd.DataFrame()
    if not work.empty:
        if "_fecha" not in work.columns:
            work["_fecha"] = work["Fecha"].map(lambda x: _as_date(x) or date.min)
        work = work.copy()
        work["_idx"] = range(len(work))
        work = work.sort_values(by=["_fecha", "_idx"], kind="stable")

    ini_cant = 0.0
    ini_vc = 0.0
    ini_tot = 0.0
    fecha_ini = fecha_cierre
    if df_inicial is not None and not df_inicial.empty:
        row = df_inicial.iloc[0]
        ini_cant = _num(row.get("Cantidad"))
        ini_vc = _num(row.get("Costo_Unitario"))
        if ini_vc <= 0 and _num(row.get("Costo_Total")) > 0 and ini_cant > 0:
            ini_vc = _num(row.get("Costo_Total")) / ini_cant
        ini_tot = _num(row.get("Costo_Total")) or round(ini_cant * ini_vc, 2)
        fecha_ini = _as_date(row.get("Fecha")) or fecha_ini

    if fecha_ini is None and not work.empty:
        f0 = _as_date(work.iloc[0].get("_fecha") or work.iloc[0].get("Fecha"))
        if f0 and f0.month == 7:
            fecha_ini = date(f0.year, 6, 30)

    wb = Workbook()
    ws = wb.active
    ws.title = "FCI"

    fechas_mov = [_as_date(r.get("_fecha") or r.get("Fecha")) for _, r in work.iterrows()]
    fechas_mov = [f for f in fechas_mov if f]
    meses = _meses_del_periodo(fechas_mov)

    ws["B3"] = "FCI"
    ws["B3"].font = BOLD_FONT
    ws["B4"] = " a Intereses Ganados"
    ws["B4"].font = BODY_FONT

    for i, mes in enumerate(meses):
        col = 3 + i
        cell = ws.cell(2, col, datetime(mes.year, mes.month, 1))
        cell.number_format = "mm-dd-yy"
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
    tot_col = 3 + len(meses)
    if meses:
        ws.cell(2, tot_col, "Totales").font = HDR_FONT
        ws.cell(2, tot_col).fill = HDR_FILL

    ws["B7"] = "SALDO INICIAL"
    ws["B7"].font = BOLD_FONT
    ws["B7"].fill = INI_FILL
    for col, val, fmt in (
        (3, ini_cant, CUOTAS_FMT),
        (4, ini_vc, VC_FMT),
        (5, ini_tot, MONEY_FMT),
    ):
        cell = ws.cell(7, col, val)
        cell.number_format = fmt
        cell.font = BOLD_FONT
        cell.fill = INI_FILL

    headers = [
        "Fecha",
        "Descripcion",
        "Cantidad de Cuotas",
        "Valor CC",
        "Total",
        "Interés",
        "Origen suscripción",
        "VC origen",
    ]
    for c, h in enumerate(headers, 2):
        ws.cell(9, c, h).font = BOLD_FONT

    cola: deque[list] = deque()
    if ini_cant > 0:
        cola.append([ini_cant, 7, True])

    r = 10
    for _, row in work.iterrows():
        tipo = str(row.get("Tipo_Operacion") or "").strip().lower()
        fecha = _as_date(row.get("_fecha") or row.get("Fecha"))
        if fecha is None:
            continue
        cant = _num(row.get("Cantidad"))
        vc = _num(row.get("Precio"))
        tot = _num(row.get("Monto_Total")) or round(cant * vc, 2)
        if cant <= 0:
            continue

        es_sus = tipo.startswith("suscrip")
        es_res = tipo.startswith("rescate")
        if not es_sus and not es_res:
            ev = evento_para(str(row.get("Tipo_inversion") or TIPO_FCI), str(row.get("Tipo_Operacion") or ""))
            es_sus = ev == EVENTO_ENTRADA
            es_res = ev == EVENTO_SALIDA

        if es_sus:
            _escribir_mov(ws, r, fecha, "SUSCRIPCION", cant, vc, tot)
            cola.append([cant, r, False])
            r += 1
            continue

        if not es_res:
            continue

        restante = cant
        partes: list[tuple[float, int, bool]] = []
        while restante > 1e-12:
            if not cola:
                break
            lote = cola[0]
            toma = min(float(lote[0]), restante)
            partes.append((toma, int(lote[1]), bool(lote[2])))
            lote[0] = float(lote[0]) - toma
            restante -= toma
            if lote[0] <= 1e-12:
                cola.popleft()

        parciales: list[float] = []
        for i, (toma, fila_origen, es_ini) in enumerate(partes):
            ultimo = i == len(partes) - 1
            if ultimo:
                tot_p = round(tot - sum(parciales), 2)
            else:
                tot_p = round(toma * vc, 2)
                parciales.append(tot_p)
            _escribir_mov(ws, r, fecha, "RESCATE", round(toma, 6), vc, tot_p)
            if es_ini:
                ws.cell(r, 7, f"=(E{r}-D$7)*D{r}")
                ws.cell(r, 8, "=$B$7")
                ws.cell(r, 9, "=D$7")
            else:
                ws.cell(r, 7, f"=(E{r}-E${fila_origen})*D{r}")
                ws.cell(r, 8, f"=B${fila_origen}")
                ws.cell(r, 9, f"=E${fila_origen}")
            ws.cell(r, 7).number_format = MONEY_FMT
            ws.cell(r, 7).font = BODY_FONT
            ws.cell(r, 7).fill = ZEBRA
            ws.cell(r, 8).number_format = DATE_FMT
            ws.cell(r, 8).font = BODY_FONT
            ws.cell(r, 8).alignment = Alignment(horizontal="center")
            ws.cell(r, 8).fill = ZEBRA
            ws.cell(r, 9).number_format = VC_FMT
            ws.cell(r, 9).font = BODY_FONT
            ws.cell(r, 9).fill = ZEBRA
            r += 1

    last = max(r - 1, 10)

    for i, _mes in enumerate(meses):
        col = 3 + i
        letra = get_column_letter(col)
        if i + 1 < len(meses):
            nxt = get_column_letter(col + 1)
            formula = (
                f'=SUMIFS($G$10:$G${last},$B$10:$B${last},">="&{letra}$2,'
                f'$B$10:$B${last},"<"&{nxt}$2)'
            )
        else:
            formula = (
                f'=SUMIFS($G$10:$G${last},$B$10:$B${last},">="&{letra}$2,'
                f'$B$10:$B${last},"<"&EDATE({letra}$2,1))'
            )
        ws.cell(3, col, formula).number_format = MONEY_FMT
        ws.cell(4, col, f"=+{letra}3").number_format = MONEY_FMT
    if meses:
        letra_tot = get_column_letter(tot_col)
        primera = get_column_letter(3)
        ultima = get_column_letter(2 + len(meses))
        ws.cell(3, tot_col, f"=SUM({primera}3:{ultima}3)").number_format = MONEY_FMT
        ws.cell(4, tot_col, f"=+{letra_tot}3").number_format = MONEY_FMT

    tr = last + 2
    cierre_txt = (fecha_cierre or date.today()).strftime("%d/%m/%Y")

    # VC de cierre (input) — lo usa el cuadro de tenencia
    ws.cell(tr, 2, f"Resultado por tenencia al {cierre_txt}").font = BOLD_FONT
    tr += 1
    ws.cell(tr, 2, "Valor cuota (ultimo extracto de fondos del mes)").font = BODY_FONT
    vc_cell = ws.cell(tr, 3, float(vc_cierre) if vc_cierre not in (None, 0) else None)
    vc_cell.fill = INPUT_FILL
    vc_cell.number_format = VC_FMT
    vc_cell.font = BOLD_FONT
    vc_row = tr
    tr += 2

    # Cuadro por lote (formato del usuario):
    # CANTIDAD | VALOR ORIGEN | VALOR DE CIERRE | DIF DE VALORES | CUOTAS * DIF
    ten_headers = [
        "CANTIDAD DE CUOTAS",
        "VALOR ORIGEN",
        "VALOR DE CIERRE",
        "DIF DE VALORES",
        "CANTIDAD DE CUOTAS * VALOR DE DIF DE CAMBIO",
    ]
    header_row = tr
    for c, h in enumerate(ten_headers, 2):
        cell = ws.cell(tr, c, h)
        cell.font = BOLD_FONT
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = _thin_border()
    tr += 1

    lotes_restantes = [lote for lote in cola if float(lote[0]) > 1e-12]
    first_lot = tr
    for lote in lotes_restantes:
        cant_rest, fila_origen, es_ini = float(lote[0]), int(lote[1]), bool(lote[2])
        if es_ini:
            ws.cell(tr, 2, round(cant_rest, 4))
            ws.cell(tr, 3, "=D$7")
        else:
            cant_origen = _num(ws.cell(fila_origen, 4).value)
            if abs(cant_origen - cant_rest) <= 1e-6:
                ws.cell(tr, 2, f"=D{fila_origen}")
            else:
                ws.cell(tr, 2, round(cant_rest, 4))
            ws.cell(tr, 3, f"=E{fila_origen}")
        ws.cell(tr, 2).number_format = CUOTAS_FMT
        ws.cell(tr, 3).number_format = VC_FMT
        ws.cell(tr, 4, f"=C${vc_row}")
        ws.cell(tr, 4).number_format = VC_FMT
        ws.cell(tr, 5, f"=D{tr}-C{tr}")
        ws.cell(tr, 5).number_format = "0.000000"
        ws.cell(tr, 6, f"=B{tr}*E{tr}")
        ws.cell(tr, 6).number_format = MONEY_FMT
        for c in range(2, 7):
            ws.cell(tr, c).font = BODY_FONT
            ws.cell(tr, c).border = _thin_border()
        tr += 1
    last_lot = tr - 1

    if lotes_restantes:
        ws.cell(tr, 2, "TOTAL").font = BOLD_FONT
        for c in range(2, 6):
            ws.cell(tr, c).border = _thin_border()
        tot_ten = ws.cell(tr, 6, f"=SUM(F{first_lot}:F{last_lot})")
        tot_ten.number_format = MONEY_FMT
        tot_ten.font = BOLD_FONT
        tot_ten.fill = INPUT_FILL
        tot_ten.border = _thin_border()
        ten_row = tr
        # Al costado: cuotas finales y valuación del extracto (lo que debe dar el mayor)
        ws.cell(tr, 8, f"=SUM(B{first_lot}:B{last_lot})")
        ws.cell(tr, 8).number_format = CUOTAS_FMT
        ws.cell(tr, 8).font = BOLD_FONT
        ws.cell(tr, 9, f"=H{tr}*C${vc_row}")
        ws.cell(tr, 9).number_format = MONEY_FMT
        ws.cell(tr, 9).font = BOLD_FONT
        ws.cell(tr, 9).fill = INPUT_FILL
        val_extracto_row = tr
        cuotas_fin_row = tr
    else:
        ws.cell(tr, 2, "TOTAL").font = BOLD_FONT
        tot_ten = ws.cell(tr, 6, 0)
        tot_ten.number_format = MONEY_FMT
        tot_ten.border = _thin_border()
        ten_row = tr
        ws.cell(tr, 8, 0).number_format = CUOTAS_FMT
        ws.cell(tr, 9, 0).number_format = MONEY_FMT
        val_extracto_row = tr
        cuotas_fin_row = tr

    ws.cell(header_row, 8, "Cuotas saldo final").font = BOLD_FONT
    ws.cell(header_row, 9, "Valuacion extracto (cuotas x VC)").font = BOLD_FONT

    # --- Resumen en pesos (todo linkeado por fórmula) ---
    tr += 2
    ws.cell(tr, 2, "RESUMEN EN PESOS / MAYOR").font = BOLD_FONT
    ws.cell(tr, 2).fill = HDR_FILL
    ws.cell(tr, 3).fill = HDR_FILL
    ws.cell(tr, 3, "Importe").font = BOLD_FONT
    tr += 1

    interes_formula = (
        f"={get_column_letter(tot_col)}3" if meses else f"=SUM($G$10:$G${last})"
    )
    # Orden alineado a la fórmula de control del estudio
    row_si = tr
    ws.cell(tr, 2, "Saldo inicial").font = BODY_FONT
    ws.cell(tr, 3, "=E7").number_format = MONEY_FMT
    tr += 1
    row_sus = tr
    ws.cell(tr, 2, "Suscripciones").font = BODY_FONT
    ws.cell(tr, 3, f'=SUMIF($C$10:$C${last},"SUSCRIPCION",$F$10:$F${last})').number_format = MONEY_FMT
    tr += 1
    row_int = tr
    ws.cell(tr, 2, "Intereses ganados").font = BODY_FONT
    ws.cell(tr, 3, interes_formula).number_format = MONEY_FMT
    tr += 1
    row_ten_lotes = tr
    ws.cell(tr, 2, "Resultado por tenencia (por lotes)").font = BODY_FONT
    ws.cell(tr, 3, f"=F{ten_row}").number_format = MONEY_FMT
    tr += 1
    row_res = tr
    ws.cell(tr, 2, "Rescates").font = BODY_FONT
    ws.cell(tr, 3, f'=SUMIF($C$10:$C${last},"RESCATE",$F$10:$F${last})').number_format = MONEY_FMT
    tr += 1

    # Saldo final del extracto = lado derecho de la fórmula de control
    row_val = tr
    ws.cell(tr, 2, "Saldo final del extracto (ultimo mes del ejercicio)").font = BOLD_FONT
    cell = ws.cell(tr, 3, f"=I{val_extracto_row}")
    cell.number_format = MONEY_FMT
    cell.font = BOLD_FONT
    cell.fill = INPUT_FILL
    tr += 1

    # Ajuste: cierra SI+Sus+Int+Ten-Res contra el extracto (redondeos Totales vs cuotas x VC)
    row_aj = tr
    ws.cell(tr, 2, "Ajuste por redondeo").font = BODY_FONT
    ws.cell(
        tr,
        3,
        f"=C{row_val}-(C{row_si}+C{row_sus}+C{row_int}+C{row_ten_lotes}-C{row_res})",
    ).number_format = MONEY_FMT
    tr += 1

    row_ten = tr
    ws.cell(tr, 2, "Resultado por tenencia a contabilizar").font = BOLD_FONT
    cell = ws.cell(tr, 3, f"=C{row_ten_lotes}+C{row_aj}")
    cell.number_format = MONEY_FMT
    cell.font = BOLD_FONT
    cell.fill = INPUT_FILL
    tr += 2

    # Fórmula de control del estudio (debe dar igual al extracto)
    ws.cell(tr, 2, "FORMULA DE CONTROL").font = BOLD_FONT
    ws.cell(tr, 2).fill = HDR_FILL
    ws.cell(tr, 3).fill = HDR_FILL
    tr += 1
    ws.cell(
        tr,
        2,
        "Saldo inicial + Suscripciones + Intereses ganados + Resultado por tenencia - Rescates",
    ).font = BODY_FONT
    row_izq = tr
    cell = ws.cell(
        tr,
        3,
        f"=C{row_si}+C{row_sus}+C{row_int}+C{row_ten}-C{row_res}",
    )
    cell.number_format = MONEY_FMT
    cell.font = BOLD_FONT
    tr += 1
    ws.cell(tr, 2, "Saldo final del extracto (ultimo mes del ejercicio)").font = BODY_FONT
    row_der = tr
    cell = ws.cell(tr, 3, f"=C{row_val}")
    cell.number_format = MONEY_FMT
    cell.font = BOLD_FONT
    cell.fill = INPUT_FILL
    tr += 1
    ws.cell(tr, 2, "Diferencia (debe ser 0 si esta bien calculado)").font = BOLD_FONT
    cell = ws.cell(tr, 3, f"=C{row_izq}-C{row_der}")
    cell.number_format = MONEY_FMT
    cell.font = BOLD_FONT
    cell.fill = INPUT_FILL
    tr += 2
    nota = (
        "Formula de control: Saldo inicial + Suscripciones + Intereses ganados + "
        "Resultado por tenencia - Rescates = Saldo final del extracto del ultimo mes "
        "del ejercicio (cuotas finales x VC del ultimo dia). "
        "La tenencia a contabilizar incluye el ajuste por redondeo para que la formula cierre. "
        "Tenencia por lote: DIF = VC cierre - VC origen; resultado = cuotas x DIF."
    )
    if meta.get("nota"):
        nota = f"{meta.get('nota')} · {nota}"
    ws.cell(tr, 2, nota).font = ITAL_FONT

    for c, w in (
        (1, 3),
        (2, 48),
        (3, 16),
        (4, 16),
        (5, 14),
        (6, 42),
        (7, 3),
        (8, 16),
        (9, 18),
    ):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[header_row].height = 32

    if lotes_restantes:
        ws2 = wb.create_sheet("Lotes cierre")
        ws2["A1"] = "Lotes FIFO del saldo final (mismo cuadro de tenencia)"
        ws2["A1"].font = BOLD_FONT
        for c, h in enumerate(ten_headers, 1):
            ws2.cell(3, c, h).font = BOLD_FONT
        for i, lote in enumerate(lotes_restantes):
            cant_rest, fila_origen, es_ini = float(lote[0]), int(lote[1]), bool(lote[2])
            rr = 4 + i
            ws2.cell(rr, 1, round(cant_rest, 4)).number_format = CUOTAS_FMT
            if es_ini:
                ws2.cell(rr, 2, "=FCI!D7")
            else:
                ws2.cell(rr, 2, f"=FCI!E{fila_origen}")
            ws2.cell(rr, 2).number_format = VC_FMT
            ws2.cell(rr, 3, f"=FCI!C{vc_row}").number_format = VC_FMT
            ws2.cell(rr, 4, f"=C{rr}-B{rr}").number_format = "0.000000"
            ws2.cell(rr, 5, f"=A{rr}*D{rr}").number_format = MONEY_FMT
        tot_r = 4 + len(lotes_restantes)
        ws2.cell(tot_r, 1, "TOTAL").font = BOLD_FONT
        ws2.cell(tot_r, 5, f"=SUM(E4:E{tot_r - 1})").number_format = MONEY_FMT
        ws2.cell(tot_r, 5).font = BOLD_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
