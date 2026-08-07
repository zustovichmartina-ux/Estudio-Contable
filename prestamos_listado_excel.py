# -*- coding: utf-8 -*-
"""
Excel claro de préstamos — estilo Listado Histórico / GlobalRecife.

Layout aprobado para papel de trabajo (excepción demo préstamos del estudio):
  - Hoja Resumen: un préstamo por fila + totales
  - Una hoja por préstamo (P-XXXXXXXX) con cabecera, grilla de cuotas, totales y deuda al día

Usado por la ventana Préstamos de Streamlit y por generar_auditoria.
No usa guardar_informe_excel (layout fijo de auditoría de préstamos).
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── estilos (alineados al Excel GlobalRecife / demo préstamos) ───────────────
_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
_FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
_FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
_FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
_FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
_FONT_BOLD = Font(name="Calibri", bold=True, size=11)
_FONT_BOLD_10 = Font(name="Calibri", bold=True, size=10)
_FONT_10 = Font(name="Calibri", size=10)
_FONT_NOTE = Font(name="Calibri", size=9, color="808080")
_FONT_WARN = Font(name="Calibri", size=10, color="C00000")
_THIN = Border(
    left=Side(style="thin", color="B8B8B8"),
    right=Side(style="thin", color="B8B8B8"),
    top=Side(style="thin", color="B8B8B8"),
    bottom=Side(style="thin", color="B8B8B8"),
)
_MONEDA = '#,##0.00'
_FECHA = "DD/MM/YYYY"

COLS_CUOTA = [
    "Nro.Cuota",
    "Fecha Vto.",
    "Capital",
    "Intereses",
    "IVA/Gastos",
    "Total Cuota",
    "Saldo Restante",
    "Estado",
]


def _as_date(val: Any):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _sheet_id(prestamo_n: Any) -> str:
    raw = str(prestamo_n or "").strip()
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 6:
        return f"P-{digits[-8:]}"
    safe = re.sub(r"[^\w\-]", "", raw)[:20] or "Prestamo"
    return f"P-{safe}"[:31]


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    name = base[:31]
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        cand = f"{base[:28]}_{i}"[:31]
        if cand not in wb.sheetnames:
            return cand
        i += 1


def _flatten_prestamos(bancos_data: dict) -> list[dict]:
    """Lista plana de préstamos con banco y metadatos."""
    filas = []
    for banco, prestamos in (bancos_data or {}).items():
        for p in prestamos or []:
            item = dict(p)
            item["banco"] = banco
            filas.append(item)
    return filas


def _total_deuda(prestamo: dict) -> float | None:
    deuda = prestamo.get("total_deuda")
    if deuda is not None and deuda != "":
        try:
            return float(deuda)
        except (TypeError, ValueError):
            pass
    dd = prestamo.get("deuda_al_dia") or {}
    if dd:
        keys = (
            "capital_a_vencer",
            "ajuste_devengado_vigente",
            "interes_devengado_vigente",
            "capital_vencido_impago",
            "ajuste_vencido_impago",
            "interes_vencido_impago",
            "comision_vencida_impaga",
            "seguro_vida_vencido",
            "seguro_incendio_vencido",
            "ajuste_mora",
            "interes_mora",
        )
        vals = [float(dd.get(k) or 0) for k in keys]
        if any(v for v in vals):
            return round(sum(vals), 2)
    return None


def _escribir_resumen(ws, titulo: str, prestamos: list[dict], sheet_refs: list[str]) -> None:
    ws["A1"] = titulo
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws["A2"] = (
        "Resumen de préstamos (una hoja de detalle por operación). "
        "Formato listado histórico / auditoría."
    )
    ws["A2"].font = _FONT_NOTE

    headers = [
        "Archivo",
        "Nro. Prestamo",
        "Banco",
        "Fecha Inicio",
        "Fecha Vto.",
        "Convenio",
        "Capital Original",
        "Total Deuda",
        "N° Cuotas",
        "Hoja",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_BOLD_10
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (p, sheet_name) in enumerate(zip(prestamos, sheet_refs)):
        r = 5 + i
        nro = p.get("nro_prestamo") or p.get("prestamo_n") or ""
        vals = [
            p.get("archivo") or "",
            nro,
            p.get("banco") or "",
            _as_date(p.get("fecha_inicio")),
            _as_date(p.get("fecha_vto")),
            p.get("convenio") or "",
            float(p.get("capital_original") or 0),
            _total_deuda(p),
            len(p.get("cuotas") or []),
            sheet_name,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(r, col, val)
            cell.font = _FONT_10
            cell.border = _THIN
            if col in (4, 5) and isinstance(val, date):
                cell.number_format = _FECHA
            if col in (7, 8) and isinstance(val, (int, float)):
                cell.number_format = _MONEDA
            if col == 7 and sheet_name:
                # Capital desde celda B8 de la hoja detalle (si existe layout)
                cell.value = f"='{sheet_name}'!B8"

    last = 4 + len(prestamos)
    if prestamos:
        tot_r = last + 1
        ws.cell(tot_r, 6, "TOTAL:").font = _FONT_BOLD
        for col in (7, 8):
            cell = ws.cell(tot_r, col)
            letter = get_column_letter(col)
            cell.value = f"=SUM({letter}5:{letter}{last})"
            cell.fill = _FILL_TOTAL
            cell.font = _FONT_BOLD
            cell.number_format = _MONEDA

    widths = [42, 22, 16, 12, 12, 36, 14, 14, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"


def _escribir_hoja_prestamo(ws, prestamo: dict) -> None:
    banco = prestamo.get("banco") or ""
    nro = str(prestamo.get("nro_prestamo") or prestamo.get("prestamo_n") or "")
    titular = prestamo.get("titular") or ""
    cuit = prestamo.get("cuit") or ""
    domicilio = prestamo.get("domicilio") or ""

    ws["A1"] = banco.upper() if banco else "PRÉSTAMO BANCARIO"
    ws["A1"].font = _FONT_BOLD
    ws["J1"] = f"LISTADO HISTÓRICO DEL PRÉSTAMO Nro: {nro}" if nro else "DETALLE DEL PRÉSTAMO"
    ws["J1"].font = _FONT_BOLD

    meta_line = []
    if prestamo.get("sucursal"):
        meta_line.append(f"Sucursal: {prestamo['sucursal']}")
    if prestamo.get("fecha_emision"):
        fe = _as_date(prestamo["fecha_emision"])
        meta_line.append(f"Fecha emisión: {fe.strftime('%d/%m/%Y') if fe else prestamo['fecha_emision']}")
    ws["A2"] = "    ".join(meta_line)
    ws["A2"].font = _FONT_10

    ws["A4"] = f"Titular: {titular}" if titular else f"Banco: {banco}"
    ws["A4"].font = _FONT_BOLD
    if cuit:
        ws["A5"] = f"Doc. Identidad: CUIT {cuit}"
    if domicilio:
        ws["A6"] = f"Domicilio: {domicilio}"

    # Cabecera montos / fechas (fila 8 como en GlobalRecife → B8 = Cap. Orig.)
    ws["A8"] = "Cap. Orig.:"
    ws["A8"].font = _FONT_BOLD
    ws["B8"] = float(prestamo.get("capital_original") or 0)
    ws["B8"].number_format = _MONEDA

    ws["D8"] = "Moneda:"
    ws["D8"].font = _FONT_BOLD
    ws["E8"] = prestamo.get("moneda") or "ARS"

    ws["F8"] = "Inic.Contrato:"
    ws["F8"].font = _FONT_BOLD
    fi = _as_date(prestamo.get("fecha_inicio"))
    ws["G8"] = fi
    if fi:
        ws["G8"].number_format = _FECHA

    ws["H8"] = "Vto.Contrato:"
    ws["H8"].font = _FONT_BOLD
    fv = _as_date(prestamo.get("fecha_vto"))
    ws["I8"] = fv
    if fv:
        ws["I8"].number_format = _FECHA

    ws["A9"] = "Convenio:"
    ws["A9"].font = _FONT_BOLD
    ws["B9"] = prestamo.get("convenio") or ""

    ws["A10"] = "Sistema:"
    ws["A10"].font = _FONT_BOLD
    ws["B10"] = prestamo.get("sistema") or "—"

    sdo = prestamo.get("saldo_teorico")
    if sdo is None:
        sdo = prestamo.get("capital_original") or 0
    ws["A12"] = "Sdo.Teórico a la fecha:"
    ws["A12"].font = _FONT_BOLD
    ws["B12"] = float(sdo or 0)
    ws["B12"].number_format = _MONEDA

    # Grilla cuotas
    header_row = 14
    for col, h in enumerate(COLS_CUOTA, 1):
        cell = ws.cell(header_row, col, h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_BOLD_10
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    cuotas = sorted(
        prestamo.get("cuotas") or [],
        key=lambda c: (
            _as_date(c.get("vencimiento")) or date.min,
            int(c.get("cuota") or 0) if str(c.get("cuota") or "").isdigit() else 0,
        ),
    )
    first_data = header_row + 1
    for i, c in enumerate(cuotas):
        r = first_data + i
        fill = _FILL_ALT if i % 2 else PatternFill()
        nro_c = c.get("nro_cuota") or c.get("cuota") or (i + 1)
        fecha = _as_date(c.get("vencimiento"))
        cap = float(c.get("capital") or 0)
        inter = float(c.get("intereses") or 0)
        iva = float(c.get("iva_gastos") or 0)
        total = float(c.get("monto_abonar") or (cap + inter + iva))
        saldo = c.get("saldo_restante")
        estado = c.get("estado") or ""

        vals = [nro_c, fecha, cap, inter, iva, None, saldo if saldo is not None else "", estado]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(r, col, val)
            cell.font = _FONT_10
            cell.border = _THIN
            cell.fill = fill
            if col == 2 and isinstance(val, date):
                cell.number_format = _FECHA
            if col in (3, 4, 5, 7) and isinstance(val, (int, float)):
                cell.number_format = _MONEDA
        # Total Cuota = fórmula
        cell_t = ws.cell(r, 6)
        cell_t.value = f"=C{r}+D{r}+E{r}"
        cell_t.number_format = _MONEDA
        cell_t.font = _FONT_10
        cell_t.border = _THIN
        cell_t.fill = fill

    last_data = first_data + len(cuotas) - 1 if cuotas else header_row
    tot_row = last_data + 1 if cuotas else header_row + 1
    ws.cell(tot_row, 1, "TOTALES:").font = _FONT_BOLD
    ws.cell(tot_row, 1).fill = _FILL_TOTAL
    for col, letter in enumerate("CDEF", 3):
        cell = ws.cell(tot_row, col)
        if cuotas:
            cell.value = f"=SUM({letter}{first_data}:{letter}{last_data})"
        else:
            cell.value = 0
        cell.fill = _FILL_TOTAL
        cell.font = _FONT_BOLD
        cell.number_format = _MONEDA
        cell.border = _THIN
    # Control capital
    ctrl = tot_row + 1
    ws.cell(ctrl, 1, "Control Capital cuotas − Cap.Orig.:").font = _FONT_10
    ws.cell(ctrl, 3, f"=C{tot_row}-B8")
    ws.cell(ctrl, 3).number_format = _MONEDA

    # DEUDA AL DIA
    deuda_row = ctrl + 3
    ws.cell(deuda_row, 1, "DEUDA AL DÍA").font = _FONT_BOLD
    dd = prestamo.get("deuda_al_dia") or {}
    labels = [
        ("Fecha", "fecha"),
        ("Capital a Vencer:", "capital_a_vencer"),
        ("Ajuste Devengado Vigente:", "ajuste_devengado_vigente"),
        ("Interés Devengado Vigente:", "interes_devengado_vigente"),
        ("Capital Vencido Impago:", "capital_vencido_impago"),
        ("Ajuste Vencido Impago:", "ajuste_vencido_impago"),
        ("Interés Vencido Impago:", "interes_vencido_impago"),
        ("Comisión Vencida Impaga:", "comision_vencida_impaga"),
        ("Seguro de Vida Vencido Impago:", "seguro_vida_vencido"),
        ("Seguro de Incendio Vencido Impago:", "seguro_incendio_vencido"),
        ("Ajuste Devengado por Mora:", "ajuste_mora"),
        ("Interés Devengado por Mora:", "interes_mora"),
    ]
    if dd:
        start_vals = deuda_row + 1
        for j, (lbl, key) in enumerate(labels):
            r = start_vals + j
            ws.cell(r, 1, lbl).font = _FONT_10
            val = dd.get(key)
            if key == "fecha":
                d = _as_date(val)
                ws.cell(r, 2, d)
                if d:
                    ws.cell(r, 2).number_format = _FECHA
            else:
                ws.cell(r, 2, float(val or 0))
                ws.cell(r, 2).number_format = _MONEDA
        tot_d = start_vals + len(labels)
        ws.cell(tot_d, 1, "TOTAL DEUDA:").font = _FONT_BOLD
        ws.cell(tot_d, 2, f"=SUM(B{start_vals + 1}:B{tot_d - 1})")
        ws.cell(tot_d, 2).fill = _FILL_TOTAL
        ws.cell(tot_d, 2).font = _FONT_BOLD
        ws.cell(tot_d, 2).number_format = _MONEDA
        footer = tot_d + 2
    else:
        ws.cell(deuda_row + 1, 1, "(Sin bloque Deuda al día en el PDF / OCR)").font = _FONT_WARN
        footer = deuda_row + 3

    archivo = prestamo.get("archivo") or ""
    if archivo:
        ws.cell(footer, 1, f"Archivo origen: {archivo}").font = _FONT_NOTE

    widths = [22, 12, 14, 14, 12, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A15"
    ws.auto_filter.ref = f"A{header_row}:H{last_data}" if cuotas else f"A{header_row}:H{header_row}"


def generar_excel_listado_claro(
    bancos_data: dict,
    ruta_salida: str | Path,
    *,
    titulo: str = "Préstamos bancarios — Listado consolidado",
    saldos_iniciales: dict | None = None,
) -> Path:
    """
    Genera Excel claro (Resumen + una hoja por préstamo).
    bancos_data: {banco: [{prestamo_n, capital_original, cuotas, ...}, ...]}
    """
    ruta_salida = Path(ruta_salida)
    prestamos = _flatten_prestamos(bancos_data)
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"

    sheet_refs: list[str] = []
    for p in prestamos:
        base = _sheet_id(p.get("nro_prestamo") or p.get("prestamo_n"))
        name = _unique_sheet_name(wb, base)
        sheet_refs.append(name)
        ws = wb.create_sheet(name)
        _escribir_hoja_prestamo(ws, p)

    _escribir_resumen(ws_res, titulo, prestamos, sheet_refs)

    # Conciliación opcional por banco (papel de trabajo)
    if saldos_iniciales:
        ws_c = wb.create_sheet("Conciliacion")
        ws_c["A1"] = "Conciliación contable — saldos iniciales vs capital amortizado"
        ws_c["A1"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
        headers = ["Banco", "Saldo Inicial", "Capital Amortizado", "Saldo Final Sugerido"]
        for col, h in enumerate(headers, 1):
            cell = ws_c.cell(3, col, h)
            cell.fill = _FILL_HEADER
            cell.font = _FONT_BOLD_10
            cell.border = _THIN
        row = 4
        for banco, prest_list in (bancos_data or {}).items():
            saldo_ini = float((saldos_iniciales or {}).get(banco, 0) or 0)
            amort = sum(
                sum(float(c.get("capital") or 0) for c in (p.get("cuotas") or []))
                for p in prest_list
            )
            vals = [banco, saldo_ini, amort, saldo_ini - amort]
            for col, val in enumerate(vals, 1):
                cell = ws_c.cell(row, col, val)
                cell.border = _THIN
                cell.font = _FONT_10
                if col >= 2:
                    cell.number_format = _MONEDA
                    if col == 4:
                        cell.fill = _FILL_TOTAL
            row += 1
        for i, w in enumerate([22, 16, 18, 20], 1):
            ws_c.column_dimensions[get_column_letter(i)].width = w

    # Nota
    note_row = 5 + len(prestamos) + 3
    ws_res.cell(
        note_row,
        1,
        "Nota: cada hoja replica el detalle del préstamo (cuotas capital/interés). "
        "Los PDFs escaneados de Banco Provincia pueden requerir OCR; verifique totales vs el listado del banco.",
    ).font = _FONT_NOTE

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    return ruta_salida


def cargar_excel_listado(ruta: str | Path) -> dict:
    """
    Lee un Excel ya generado en este formato (o GlobalRecife) → bancos_data.
    Útil si el usuario sube el .xlsx en lugar de PDFs.
    """
    ruta = Path(ruta)
    wb = load_workbook(ruta, data_only=True)
    bancos: dict[str, list] = {}

    # Preferir hojas P-*
    detail_sheets = [s for s in wb.sheetnames if s.upper().startswith("P-") or s == "Detalle"]
    if not detail_sheets:
        detail_sheets = [s for s in wb.sheetnames if s not in ("Resumen", "Resumen Ejecutivo", "Conciliacion")]

    # Índice desde Resumen si existe
    meta_by_sheet: dict[str, dict] = {}
    if "Resumen" in wb.sheetnames:
        ws = wb["Resumen"]
        # detectar fila header
        header_row = None
        for r in range(1, 10):
            vals = [str(ws.cell(r, c).value or "").lower() for c in range(1, 12)]
            if any("prestamo" in v or "préstamo" in v for v in vals) and any(
                "capital" in v for v in vals
            ):
                header_row = r
                break
        if header_row:
            headers = {
                c: str(ws.cell(header_row, c).value or "").strip().lower()
                for c in range(1, 12)
            }
            for r in range(header_row + 1, ws.max_row + 1):
                hoja = None
                fila_meta = {}
                for c, h in headers.items():
                    val = ws.cell(r, c).value
                    if not h:
                        continue
                    if "hoja" in h:
                        hoja = str(val or "").strip()
                    elif "archivo" in h:
                        fila_meta["archivo"] = val
                    elif "nro" in h and "prest" in h:
                        fila_meta["nro_prestamo"] = val
                    elif h == "banco":
                        fila_meta["banco"] = val
                    elif "inicio" in h:
                        fila_meta["fecha_inicio"] = val
                    elif "vto" in h:
                        fila_meta["fecha_vto"] = val
                    elif "convenio" in h:
                        fila_meta["convenio"] = val
                    elif "capital" in h:
                        try:
                            fila_meta["capital_original"] = float(val or 0)
                        except (TypeError, ValueError):
                            pass
                    elif "deuda" in h:
                        fila_meta["total_deuda"] = val
                if hoja:
                    meta_by_sheet[hoja] = fila_meta

    for sheet_name in detail_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        meta = dict(meta_by_sheet.get(sheet_name) or {})
        banco = meta.get("banco") or "Banco Provincia"
        # Cap. Orig. en B8 (layout GlobalRecife / este módulo)
        if ws["B8"].value is not None and not meta.get("capital_original"):
            try:
                meta["capital_original"] = float(ws["B8"].value)
            except (TypeError, ValueError):
                pass
        if ws["B9"].value and not meta.get("convenio"):
            meta["convenio"] = ws["B9"].value
        if ws["G8"].value and not meta.get("fecha_inicio"):
            meta["fecha_inicio"] = ws["G8"].value
        if ws["I8"].value and not meta.get("fecha_vto"):
            meta["fecha_vto"] = ws["I8"].value
        nro = meta.get("nro_prestamo") or sheet_name.replace("P-", "")
        meta["prestamo_n"] = nro
        meta["nro_prestamo"] = meta.get("nro_prestamo") or nro

        # Detectar header de cuotas
        header_row = None
        for r in range(1, min(30, ws.max_row + 1)):
            row_txt = " ".join(str(ws.cell(r, c).value or "").lower() for c in range(1, 10))
            if "cuota" in row_txt and ("capital" in row_txt or "interes" in row_txt or "interés" in row_txt):
                header_row = r
                break
        cuotas = []
        if header_row:
            # map columns
            colmap = {}
            for c in range(1, 16):
                h = str(ws.cell(header_row, c).value or "").lower()
                if "nro" in h or h == "cuota":
                    colmap["cuota"] = c
                elif "vto" in h or "venc" in h:
                    colmap["vencimiento"] = c
                elif "capital" in h:
                    colmap["capital"] = c
                elif "interes" in h or "interés" in h:
                    colmap["intereses"] = c
                elif "iva" in h or "gasto" in h:
                    colmap["iva_gastos"] = c
                elif "total" in h:
                    colmap["monto_abonar"] = c
                elif "saldo" in h:
                    colmap["saldo_restante"] = c
                elif "estado" in h:
                    colmap["estado"] = c
            for r in range(header_row + 1, ws.max_row + 1):
                label = str(ws.cell(r, 1).value or "")
                if not label or label.upper().startswith("TOTAL"):
                    if label.upper().startswith("TOTAL"):
                        break
                    continue
                if "control" in label.lower() or "deuda" in label.lower():
                    break

                def _num(key, default=0.0):
                    c = colmap.get(key)
                    if not c:
                        return default
                    v = ws.cell(r, c).value
                    try:
                        return float(v or 0)
                    except (TypeError, ValueError):
                        return default

                cuota = {
                    "cuota": ws.cell(r, colmap["cuota"]).value if "cuota" in colmap else (len(cuotas) + 1),
                    "nro_cuota": ws.cell(r, colmap["cuota"]).value if "cuota" in colmap else None,
                    "vencimiento": ws.cell(r, colmap["vencimiento"]).value if "vencimiento" in colmap else None,
                    "capital": _num("capital"),
                    "intereses": _num("intereses"),
                    "iva_gastos": _num("iva_gastos"),
                    "monto_abonar": _num("monto_abonar"),
                    "saldo_restante": _num("saldo_restante"),
                    "estado": ws.cell(r, colmap["estado"]).value if "estado" in colmap else "",
                }
                if cuota["monto_abonar"] == 0:
                    cuota["monto_abonar"] = round(cuota["capital"] + cuota["intereses"] + cuota["iva_gastos"], 2)
                # saltar filas vacías de montos
                if cuota["capital"] == 0 and cuota["intereses"] == 0 and cuota["monto_abonar"] == 0:
                    # puede ser desembolso con capital en otra fila — igual incluir si hay fecha/nro
                    if not cuota.get("vencimiento") and not cuota.get("nro_cuota"):
                        continue
                cuotas.append(cuota)

        meta["cuotas"] = cuotas
        if not meta.get("capital_original"):
            meta["capital_original"] = sum(c.get("capital") or 0 for c in cuotas)
        bancos.setdefault(banco, []).append(meta)

    return bancos


def es_excel_listado_prestamos(ruta: str | Path) -> bool:
    """Heurística: ¿parece un Excel de listado de préstamos?"""
    try:
        wb = load_workbook(Path(ruta), read_only=True)
        names = [n.lower() for n in wb.sheetnames]
        wb.close()
        if any(n.startswith("p-") for n in names):
            return True
        if "resumen" in names and len(names) >= 2:
            return True
        return False
    except Exception:
        return False
