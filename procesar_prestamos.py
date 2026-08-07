# -*- coding: utf-8 -*-
"""
Script de procesamiento de préstamos para auditoría.
Extrae datos de 31 PDFs y genera Excel con formato aprobado.
"""

import re
import pdfplumber
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────
CARPETA_PDFS = Path(r"C:\Users\recep\Desktop\Estudio Contable\extractos bancarios\Prestamos Financieros")
SALIDA_EXCEL = Path(r"C:\Users\recep\Desktop\Estudio Contable\Auditoria_Prestamos_v2.xlsx")

# Colores
COLOR_HEADER_GRIS   = "B8B8B8"
COLOR_AZUL_OSCURO   = "1F4E79"
COLOR_AZUL_MEDIO    = "2E75B6"
COLOR_FILA_ALT      = "F2F2F2"
COLOR_AMARILLO      = "FFFF00"
COLOR_BLANCO        = "FFFFFF"

ANCHO_COLS = {
    "A": 8,   # CUOTA
    "B": 14,  # VENCIMIENTO
    "C": 16,  # CAPITAL
    "D": 16,  # INTERESES
    "E": 16,  # IVA/GASTOS
    "F": 18,  # MONTO A ABONAR
    "G": 18,  # SALDO RESTANTE
}

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE MONEDA
# ─────────────────────────────────────────────────────────────────────────────
def parse_moneda_arg(s):
    """Convierte string moneda argentina '$ 1.234.567,89' → float."""
    if s is None:
        return 0.0
    s = str(s).strip().replace("$", "").strip()
    if not s or s in ("-", "—"):
        return 0.0
    # Formato argentino: puntos como miles, coma como decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    # Eliminar espacios y paréntesis opcionales
    s = re.sub(r"[^\d\.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_moneda_int(s):
    """Convierte string moneda estilo int americano '1,234,567.89' → float."""
    if s is None:
        return 0.0
    s = str(s).strip().replace("$", "").strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0

# ─────────────────────────────────────────────────────────────────────────────
# DETECCIÓN DE BANCO Y NÚMERO DE PRÉSTAMO DESDE EL NOMBRE
# ─────────────────────────────────────────────────────────────────────────────
def detectar_banco(nombre):
    n = nombre.upper()
    if "SANTANDER" in n:
        return "Banco Santander"
    if "MERCADO" in n or "MERCADO_PAGO" in n or "MERCADOPAGO" in n:
        return "Mercado Pago"
    if "PROVINCIA" in n:
        return "Banco Provincia"
    if "NACION" in n:
        return "Banco Nación"
    if "GALICIA" in n:
        return "Banco Galicia"
    if "FRANCES" in n or "BBVA" in n or "FRANC" in n:
        return "Banco Francés (BBVA)"
    return "Banco Desconocido"

def detectar_numero_prestamo(nombre):
    stem = Path(nombre).stem
    # (NNNN) entre paréntesis – puede ser "8535-8407" → mantener como string
    m = re.search(r"\((\d[\d\-]+)\)", stem)
    if m:
        return m.group(1)
    # PréstamoN°_XXXXXXXX (ej: Santander __PréstamoN°_039100426399)
    m = re.search(r"Pr[eé]stamoN[°o]?[_\s]*(\w+)", stem, re.IGNORECASE)
    if m:
        return m.group(1)
    # Número largo precedido de underscore: _32024192, _808088208719
    m = re.search(r"_(\d{6,})", stem)
    if m:
        return m.group(1)
    # Número muy largo en cualquier posición (>=9 dígitos)
    m = re.search(r"(\d{9,})", stem)
    if m:
        return m.group(1)
    # Sufijo numérico al final del stem (ej: "Cuotas 0682" → 0682, "Prestamo 5848" → 5848)
    m = re.search(r"(\d{3,})\s*$", stem.strip())
    if m:
        return m.group(1)
    # Cualquier número >= 4 dígitos que NO sea una fecha YYYYMMDD
    nums = re.findall(r"\d{4,}", stem)
    # Filtrar fechas tipo 20250312
    non_dates = [n for n in nums if not re.match(r"^202[0-9][01]\d[0-3]\d$", n)]
    if non_dates:
        return non_dates[-1]
    if nums:
        return nums[-1]
    return stem[:20]

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTORES POR BANCO
# ─────────────────────────────────────────────────────────────────────────────

def extraer_mercado_pago(pdf_path):
    """
    Extrae cuotas de Mercado Pago.
    Tabla con headers: CUOTA | VENCIMIENTO | CAPITAL | INTERÉS | IVA | MONTO A ABONAR | SALDO
    Moneda: $ 388141,79  (coma decimal)
    """
    cuotas = []
    capital_prestamo = 0.0
    num_prestamo = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text() or ""
            # Extraer NRO. PRÉSTAMO (segunda columna de la línea tras los labels)
            if num_prestamo is None:
                # Formato: "CUST ID NRO. PRÉSTAMO\n407645112 990725421"
                m = re.search(r"NRO\.\s*PR\w+STAMO\s*\n\s*\d+\s+(\d+)", texto)
                if m:
                    num_prestamo = m.group(1)

            tablas = page.extract_tables() or []
            for tabla in tablas:
                for row in tabla:
                    if row is None or not row:
                        continue
                    # Cabecera de condiciones financieras
                    if row[0] is not None and "MONTO DEL CR" in str(row[0]).upper():
                        continue
                    # Extrae monto del crédito de la fila de valores de la primera tabla
                    if row[0] is None and row[3] is not None and "$" in str(row[3]):
                        capital_prestamo = parse_moneda_arg(row[3])
                        continue
                    # Fila de headers de cuotas
                    if str(row[0]).strip().upper() in ("CUOTA", "NRO", "#"):
                        continue
                    # Fila de datos
                    if len(row) >= 7 and row[0] is not None:
                        try:
                            nro = int(str(row[0]).strip())
                        except ValueError:
                            continue
                        venc  = str(row[1]).strip() if row[1] else ""
                        cap   = parse_moneda_arg(row[2])
                        inter = parse_moneda_arg(row[3])
                        iva   = parse_moneda_arg(row[4])
                        total = parse_moneda_arg(row[5])
                        saldo = parse_moneda_arg(row[6])
                        cuotas.append({
                            "cuota": nro,
                            "vencimiento": venc,
                            "capital": cap,
                            "intereses": inter,
                            "iva_gastos": iva,
                            "total": total,
                            "saldo": saldo,
                        })

    # Deduplicar por número de cuota (mantener primera ocurrencia)
    vistos = set()
    cuotas_unicas = []
    for c in cuotas:
        if c["cuota"] not in vistos:
            vistos.add(c["cuota"])
            cuotas_unicas.append(c)
    cuotas_unicas.sort(key=lambda x: x["cuota"])

    if capital_prestamo == 0.0 and cuotas_unicas:
        capital_prestamo = sum(c["capital"] for c in cuotas_unicas)

    return cuotas_unicas, capital_prestamo, num_prestamo


def extraer_galicia(pdf_path):
    """
    Banco Galicia: texto plano.
    Línea: N estado YYYY-MM-DD $ capital $ interes $ iva_interes $ iva_percep $ otros
    """
    cuotas = []
    capital_prestamo = 0.0
    num_prestamo = None

    with pdfplumber.open(pdf_path) as pdf:
        texto_total = ""
        for page in pdf.pages:
            texto_total += (page.extract_text() or "") + "\n"

    # Extraer importe del préstamo
    m = re.search(r"Importe:\s*\$\s*([\d,.]+)", texto_total)
    if m:
        capital_prestamo = parse_moneda_int(m.group(1))

    # Extraer número de préstamo
    m = re.search(r"Pr[eé]stamo\s+Nro:\s*(\d+)", texto_total)
    if m:
        num_prestamo = m.group(1)

    # Parsear cuotas
    # Formato: N A Vencer YYYY-MM-DD $ X $ X $ X $ X $ X $ X (*)
    patron = re.compile(
        r"^(\d+)\s+(?:A Vencer|Vencida|Cancelada)\s+"
        r"(\d{4}-\d{2}-\d{2})\s+"
        r"\$\s*([\d,.]+)\s+"        # monto total
        r"\$\s*([\d,.]+)\s+"        # capital
        r"\$\s*([\d,.]+)\s+"        # interés nominal
        r"\$\s*([\d,.]+)\s+"        # iva interés nominal
        r"\$\s*([\d,.]+)\s+"        # iva percepción
        r"\$\s*([\d,.]+)",          # otros gastos
        re.MULTILINE
    )
    for m in patron.finditer(texto_total):
        nro   = int(m.group(1))
        venc  = m.group(2)
        total = parse_moneda_int(m.group(3))
        cap   = parse_moneda_int(m.group(4))
        inter = parse_moneda_int(m.group(5))
        iva_i = parse_moneda_int(m.group(6))
        iva_p = parse_moneda_int(m.group(7))
        otros = parse_moneda_int(m.group(8))
        iva_gastos = iva_i + iva_p + otros
        saldo = 0.0
        cuotas.append({
            "cuota": nro,
            "vencimiento": venc,
            "capital": cap,
            "intereses": inter,
            "iva_gastos": iva_gastos,
            "total": total,
            "saldo": saldo,
        })

    if capital_prestamo == 0.0 and cuotas:
        capital_prestamo = sum(c["capital"] for c in cuotas)

    return cuotas, capital_prestamo, num_prestamo


def extraer_frances_simulacion(pdf_path):
    """
    Banco Francés simulación (P6 8467): texto fijo.
    Línea: N DD/MM/YY AMORTIZAC INTERESES SEGURO IMPUESTO CUOTA SALDO
    Moneda: 2,756,034.83 (coma miles, punto decimal)
    Capital desde: IMPORTE CONCEDIDO : 64,000,000.00
    """
    cuotas = []
    capital_prestamo = 0.0
    num_prestamo = None

    with pdfplumber.open(pdf_path) as pdf:
        texto_total = ""
        for page in pdf.pages:
            texto_total += (page.extract_text() or "") + "\n"

    m = re.search(r"IMPORTE CONCEDIDO\s*:\s*([\d,]+\.[\d]+)", texto_total)
    if m:
        capital_prestamo = parse_moneda_int(m.group(1))

    # Buscar número del producto en la primera línea: "C.C.C. : 00170214 82 9600148467"
    m = re.search(r"C\.C\.C\.\s*:\s*[\d\s]+(\d{4})\b", texto_total)
    if m:
        num_prestamo = m.group(1)

    # Líneas de cuotas: número, fecha, montos separados por espacios
    patron = re.compile(
        r"^\s*(\d+)\s+([\d/]+)\s+([\d,]+\.[\d]+)\s+([\d,]+\.[\d]+)\s+([\d,]+\.[\d]+)\s+([\d,]+\.[\d]+)\s+([\d,]+\.[\d]+)\s+([\d,]+\.[\d]+)",
        re.MULTILINE
    )
    prev_saldo = capital_prestamo
    for m in patron.finditer(texto_total):
        nro   = int(m.group(1))
        venc  = m.group(2)
        cap   = parse_moneda_int(m.group(3))
        inter = parse_moneda_int(m.group(4))
        seguro= parse_moneda_int(m.group(5))
        impto = parse_moneda_int(m.group(6))
        total = parse_moneda_int(m.group(7))
        saldo = parse_moneda_int(m.group(8))
        iva_gastos = seguro + impto
        cuotas.append({
            "cuota": nro,
            "vencimiento": venc,
            "capital": cap,
            "intereses": inter,
            "iva_gastos": iva_gastos,
            "total": total,
            "saldo": saldo,
        })

    return cuotas, capital_prestamo, num_prestamo


def extraer_frances_cuotas(pdf_path):
    """
    Banco Francés formato cuotas (P7/P8/P9 _Cuotas):
    CUOTA VTO. CAPITAL INTERÉS SEGUROS IMPUESTOS IMPORTE SALDO
    Moneda: $ 1.018.008,46 (punto miles, coma decimal)
    """
    cuotas = []
    capital_prestamo = 0.0
    num_prestamo = None

    with pdfplumber.open(pdf_path) as pdf:
        texto_total = ""
        for page in pdf.pages:
            texto_total += (page.extract_text() or "") + "\n"

    # Capital total = suma de capitales
    patron = re.compile(
        r"(\d+)\s+(\d{2}-\d{2}-\d{4})\s+"
        r"\$\s*([\d\.]+,\d{2})\s+"   # capital
        r"\$\s*([\d\.]+,\d{2})\s+"   # interés
        r"\$\s*([\d\.]+,\d{2})\s+"   # seguros
        r"\$\s*([\d\.]+,\d{2})\s+"   # impuestos
        r"\$\s*([\d\.]+,\d{2})\s+"   # importe
        r"\$\s*([\d\.]+,\d{2})",      # saldo
        re.MULTILINE
    )
    for m in patron.finditer(texto_total):
        nro   = int(m.group(1))
        venc  = m.group(2)
        cap   = parse_moneda_arg(m.group(3))
        inter = parse_moneda_arg(m.group(4))
        seg   = parse_moneda_arg(m.group(5))
        impto = parse_moneda_arg(m.group(6))
        total = parse_moneda_arg(m.group(7))
        saldo = parse_moneda_arg(m.group(8))
        iva_gastos = seg + impto
        cuotas.append({
            "cuota": nro,
            "vencimiento": venc,
            "capital": cap,
            "intereses": inter,
            "iva_gastos": iva_gastos,
            "total": total,
            "saldo": saldo,
        })

    if cuotas:
        capital_prestamo = sum(c["capital"] for c in cuotas)

    return cuotas, capital_prestamo, num_prestamo


def extraer_sin_texto(pdf_path):
    """Para PDFs escaneados sin texto nativo."""
    return [], 0.0, None


# ─────────────────────────────────────────────────────────────────────────────
# LÓGICA DE SELECCIÓN DE EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────
def procesar_pdf(pdf_path):
    nombre = pdf_path.name
    banco  = detectar_banco(nombre)
    num    = detectar_numero_prestamo(nombre)
    es_cuotas = "CUOTA" in nombre.upper() or "_CUOTA" in nombre.upper()
    es_contrato_frances = (
        banco == "Banco Francés (BBVA)" and
        not es_cuotas and
        "CUOTAS" not in nombre.upper()
    )

    # Verificar si tiene texto nativo
    tiene_texto = False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:2]:
                t = page.extract_text() or ""
                if len(t.strip()) > 50:
                    tiene_texto = True
                    break
    except Exception:
        pass

    if not tiene_texto:
        cuotas, capital, num_from_pdf = extraer_sin_texto(pdf_path)
        return banco, num, cuotas, capital, "ESCANEADO - sin texto nativo"

    # Seleccionar extractor según banco
    try:
        if banco == "Mercado Pago":
            cuotas, capital, num_from_pdf = extraer_mercado_pago(pdf_path)
        elif banco == "Banco Galicia":
            cuotas, capital, num_from_pdf = extraer_galicia(pdf_path)
        elif banco == "Banco Francés (BBVA)":
            if "_CUOTA" in nombre.upper() or "CUOTAS" in nombre.upper():
                cuotas, capital, num_from_pdf = extraer_frances_cuotas(pdf_path)
            else:
                cuotas, capital, num_from_pdf = extraer_frances_simulacion(pdf_path)
        else:
            cuotas, capital, num_from_pdf = [], 0.0, None
            return banco, num, cuotas, capital, f"Sin extractor para {banco} con texto"

        if num_from_pdf:
            num = num_from_pdf
        # Para Mercado Pago, si el filename tiene sufijo específico usarlo como complemento
        if banco == "Mercado Pago":
            fn_num = detectar_numero_prestamo(nombre)
            # Si el número del PDF tiene más dígitos que el del filename, preferir el PDF
            if len(str(num)) > len(str(fn_num)):
                pass  # keep PDF number
            else:
                num = fn_num  # use filename number
        estado = "OK" if cuotas else "SIN CUOTAS DETECTADAS"
        return banco, num, cuotas, capital, estado

    except Exception as e:
        return banco, num, [], 0.0, f"ERROR: {e}"


# ─────────────────────────────────────────────────────────────────────────────
# GENERAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def fill(ws, row, col, value, bg=None, bold=False, color="000000",
         fmt=None, align_h="left", align_v="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    font_color = color if color else "000000"
    cell.font = Font(name="Calibri", size=10, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=True)
    if fmt:
        cell.number_format = fmt
    return cell


def merge_fill(ws, row, col_start, col_end, value, bg, bold=False,
               color="000000", size=11):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def escribir_prestamo_en_hoja(ws, fila_inicio, prestamo_id, banco, cuotas,
                               capital_total):
    fila = fila_inicio

    # A) Cabecera del préstamo (merge A:G)
    label = f"PRÉSTAMO N° {prestamo_id}  |  Capital: $ {capital_total:,.0f}  |  Banco: {banco}"
    merge_fill(ws, fila, 1, 7, label, COLOR_HEADER_GRIS, bold=True,
               color="000000", size=11)
    fila += 1

    # B) Bloque resumen anual
    merge_fill(ws, fila, 1, 7, "RESUMEN AÑO 2025", COLOR_AZUL_OSCURO,
               bold=True, color=COLOR_BLANCO, size=10)
    fila += 1

    tot_cap  = sum(c["capital"]    for c in cuotas)
    tot_int  = sum(c["intereses"]  for c in cuotas)
    tot_iva  = sum(c["iva_gastos"] for c in cuotas)

    labels_resumen = ["Total Capital", "Total Intereses", "Total IVA/Gastos"]
    valores_resumen = [tot_cap, tot_int, tot_iva]
    for i, (lbl, val) in enumerate(zip(labels_resumen, valores_resumen)):
        col_l = i * 2 + 1  # 1, 3, 5
        col_v = col_l + 1  # 2, 4, 6
        ws.merge_cells(start_row=fila, start_column=col_l,
                       end_row=fila, end_column=col_v)
        c = ws.cell(row=fila, column=col_l, value=lbl)
        c.fill  = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
        c.font  = Font(name="Calibri", size=10, bold=True, color=COLOR_BLANCO)
        c.alignment = Alignment(horizontal="left", vertical="center")
    # Celdas de valores en fila siguiente
    fila += 1
    for i, val in enumerate(valores_resumen):
        col_l = i * 2 + 1
        col_v = col_l + 1
        ws.merge_cells(start_row=fila, start_column=col_l,
                       end_row=fila, end_column=col_v)
        c = ws.cell(row=fila, column=col_l, value=val)
        c.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
        c.font = Font(name="Calibri", size=10, color=COLOR_BLANCO)
        c.number_format = '"$ "#,##0.00'
        c.alignment = Alignment(horizontal="right", vertical="center")
    fila += 1

    # C) Encabezado de grilla
    headers = ["CUOTA", "VENCIMIENTO", "CAPITAL", "INTERESES",
               "IVA/GASTOS", "MONTO A ABONAR", "SALDO RESTANTE"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=fila, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=COLOR_AZUL_MEDIO)
        c.font = Font(name="Calibri", size=10, bold=True, color=COLOR_BLANCO)
        c.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    # D) Filas de cuotas
    FMT_MONEDA = '"$ "#,##0.00'
    for idx, cuota in enumerate(cuotas):
        bg = COLOR_BLANCO if idx % 2 == 0 else COLOR_FILA_ALT
        # Cuota (centro, texto)
        c = ws.cell(row=fila, column=1, value=cuota["cuota"])
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        # Vencimiento
        c = ws.cell(row=fila, column=2, value=cuota["vencimiento"])
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        # Montos
        campos_num = [
            ("capital",    3),
            ("intereses",  4),
            ("iva_gastos", 5),
            ("total",      6),
            ("saldo",      7),
        ]
        for campo, col in campos_num:
            v = cuota.get(campo, 0.0)
            c = ws.cell(row=fila, column=col, value=v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10)
            c.number_format = FMT_MONEDA
            c.alignment = Alignment(horizontal="right", vertical="center")
        fila += 1

    # Si no hay cuotas, fila placeholder
    if not cuotas:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
        c = ws.cell(row=fila, column=1,
                    value="⚠ PDF escaneado – datos pendientes de ingreso manual")
        c.fill = PatternFill("solid", fgColor="FFF2CC")
        c.font = Font(name="Calibri", size=10, italic=True, color="7F6000")
        c.alignment = Alignment(horizontal="left", vertical="center")
        fila += 1

    # E) 4 filas en blanco
    fila += 4
    return fila


def agregar_conciliacion(ws, fila, prestamos_hoja):
    """Bloque de conciliación al final de cada hoja de banco."""
    merge_fill(ws, fila, 1, 7, "CONCILIACIÓN CONTABLE FINAL",
               COLOR_AZUL_OSCURO, bold=True, color=COLOR_BLANCO, size=11)
    fila += 1

    tot_cap_global = sum(
        sum(c["capital"] for c in p["cuotas"])
        for p in prestamos_hoja
    )

    rows_conc = [
        ("Saldo Inicial",              0.0,              COLOR_BLANCO),
        ("Total Capital Amortizado",   tot_cap_global,   COLOR_BLANCO),
        ("Saldo Final Sugerido",        tot_cap_global,   COLOR_AMARILLO),
    ]
    for lbl, val, bg in rows_conc:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        c1 = ws.cell(row=fila, column=1, value=lbl)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.font = Font(name="Calibri", size=10, bold=True)
        c1.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=fila, start_column=4, end_row=fila, end_column=7)
        c2 = ws.cell(row=fila, column=4, value=val)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.font = Font(name="Calibri", size=10)
        c2.number_format = '"$ "#,##0.00'
        c2.alignment = Alignment(horizontal="right", vertical="center")
        fila += 1

    return fila


def crear_hoja_banco(wb, nombre_banco, prestamos):
    # Sanitizar nombre de hoja
    safe_name = (nombre_banco
                 .replace("(", "").replace(")", "")
                 .replace("/", "-").replace("\\", "-")
                 .replace(":", "")
                 .strip()[:31])
    ws = wb.create_sheet(title=safe_name)

    # Anchos de columna
    for letra, ancho in ANCHO_COLS.items():
        ws.column_dimensions[letra].width = ancho

    ws.row_dimensions[1].height = 20

    fila = 1
    for p in prestamos:
        fila = escribir_prestamo_en_hoja(
            ws, fila,
            prestamo_id=p["num_prestamo"],
            banco=nombre_banco,
            cuotas=p["cuotas"],
            capital_total=p["capital"],
        )

    agregar_conciliacion(ws, fila, prestamos)
    return ws


def crear_resumen_ejecutivo(wb, bancos_dict):
    ws = wb.create_sheet(title="Resumen Ejecutivo", index=0)

    for letra, ancho in ANCHO_COLS.items():
        ws.column_dimensions[letra].width = ancho
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    # Título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "AUDITORÍA DE PRÉSTAMOS FINANCIEROS – GLOBAL RECIFE SA – 2025"
    c.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
    c.font = Font(name="Calibri", size=13, bold=True, color=COLOR_BLANCO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    fila = 3
    # Headers
    hdrs = ["BANCO", "PRÉSTAMOS", "CUOTAS", "CAPITAL TOTAL", "INTERESES TOTAL", "TOTAL A PAGAR"]
    for i, h in enumerate(hdrs, start=1):
        c = ws.cell(row=fila, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=COLOR_AZUL_MEDIO)
        c.font = Font(name="Calibri", size=10, bold=True, color=COLOR_BLANCO)
        c.alignment = Alignment(horizontal="center", vertical="center")
    fila += 1

    totales = {"prestamos": 0, "cuotas": 0, "capital": 0.0,
               "intereses": 0.0, "total": 0.0}

    for banco, prestamos in sorted(bancos_dict.items()):
        n_prest = len(prestamos)
        n_cuotas = sum(len(p["cuotas"]) for p in prestamos)
        cap  = sum(sum(c["capital"]   for c in p["cuotas"]) for p in prestamos)
        inter = sum(sum(c["intereses"] for c in p["cuotas"]) for p in prestamos)
        tot  = sum(sum(c["total"]     for c in p["cuotas"]) for p in prestamos)

        bg = COLOR_FILA_ALT if fila % 2 == 0 else COLOR_BLANCO
        data = [banco, n_prest, n_cuotas, cap, inter, tot]
        fmts = [None, None, None, '"$ "#,##0.00', '"$ "#,##0.00', '"$ "#,##0.00']
        aligns = ["left", "center", "center", "right", "right", "right"]
        for i, (val, fmt, al) in enumerate(zip(data, fmts, aligns), start=1):
            c = ws.cell(row=fila, column=i, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Calibri", size=10)
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(horizontal=al, vertical="center")

        totales["prestamos"] += n_prest
        totales["cuotas"]    += n_cuotas
        totales["capital"]   += cap
        totales["intereses"] += inter
        totales["total"]     += tot
        fila += 1

    # Fila de totales
    data_tot = ["TOTAL GENERAL", totales["prestamos"], totales["cuotas"],
                totales["capital"], totales["intereses"], totales["total"]]
    fmts_tot = [None, None, None, '"$ "#,##0.00', '"$ "#,##0.00', '"$ "#,##0.00']
    for i, (val, fmt) in enumerate(zip(data_tot, fmts_tot), start=1):
        c = ws.cell(row=fila, column=i, value=val)
        c.fill = PatternFill("solid", fgColor=COLOR_AZUL_OSCURO)
        c.font = Font(name="Calibri", size=10, bold=True, color=COLOR_BLANCO)
        if fmt:
            c.number_format = fmt
        c.alignment = Alignment(
            horizontal="right" if i > 2 else "left", vertical="center"
        )

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    pdfs = sorted(CARPETA_PDFS.glob("*.pdf"))
    print(f"PDFs encontrados: {len(pdfs)}\n")

    # Para Banco Francés: si hay un archivo _Cuotas para el mismo Pn, saltar el contrato
    # Detectar qué "Pn" tienen archivo de cuotas
    frances_pn_con_cuotas = set()
    for pdf in pdfs:
        n_up = pdf.name.upper()
        if "FRANCES" in n_up and ("_CUOTAS" in n_up or "_CUOTA " in n_up):
            m = re.match(r"^(P\d+)\s", pdf.name, re.IGNORECASE)
            if m:
                frances_pn_con_cuotas.add(m.group(1).upper())  # "P7", "P8", "P9"

    # Mapeo Pn → número de préstamo desde contratos Francés
    # (para etiquetar correctamente los _Cuotas que no tienen número en el filename)
    frances_pn_numero = {}
    for pdf in pdfs:
        n_up = pdf.name.upper()
        if "FRANCES" in n_up and "_CUOTAS" not in n_up and "_CUOTA " not in n_up:
            m_pn = re.match(r"^(P\d+)\s", pdf.name, re.IGNORECASE)
            if m_pn:
                pn = m_pn.group(1).upper()
                # El número de 4 dígitos al final del stem del contrato
                m_num = re.search(r"\s(\d{4})\s*$", Path(pdf.name).stem)
                if m_num:
                    frances_pn_numero[pn] = m_num.group(1)

    # Para Santander: priorizar CUOTAS
    santander_cuotas = any("CUOTA" in p.name.upper() and "SANTANDER" in p.name.upper()
                           for p in pdfs)

    resultados = []
    saltados   = []

    for pdf_path in pdfs:
        nombre = pdf_path.name
        banco  = detectar_banco(nombre)

        # Saltar contratos de Banco Francés si ya hay _Cuotas para el mismo Pn
        if banco == "Banco Francés (BBVA)":
            n_up = nombre.upper()
            if "_CUOTAS" not in n_up and "_CUOTA " not in n_up:
                m_pn = re.match(r"^(P\d+)\s", nombre, re.IGNORECASE)
                if m_pn and m_pn.group(1).upper() in frances_pn_con_cuotas:
                    saltados.append((nombre, "Contrato Frances (se usa _Cuotas)"))
                    continue

        # Saltar contrato Santander sin CUOTAS si existe el de CUOTAS
        if banco == "Banco Santander" and santander_cuotas:
            if "CUOTA" not in nombre.upper():
                saltados.append((nombre, "Contrato Santander (se usa CUOTAS)"))
                continue

        print(f"Procesando: {nombre}")
        banco_det, num_prestamo, cuotas, capital, estado = procesar_pdf(pdf_path)

        # Para _Cuotas de Francés sin número explícito, usar el del contrato correspondiente
        if banco_det == "Banco Francés (BBVA)":
            n_up = nombre.upper()
            if "_CUOTAS" in n_up or "_CUOTA " in n_up:
                m_pn = re.match(r"^(P\d+)\s", nombre, re.IGNORECASE)
                if m_pn:
                    pn = m_pn.group(1).upper()
                    if pn in frances_pn_numero:
                        num_prestamo = frances_pn_numero[pn]

        print(f"  -> Banco: {banco_det} | Prestamo: {num_prestamo} | "
              f"Cuotas: {len(cuotas)} | Capital: ${capital:,.0f} | Estado: {estado}")

        resultados.append({
            "archivo":       nombre,
            "banco":         banco_det,
            "num_prestamo":  num_prestamo,
            "cuotas":        cuotas,
            "capital":       capital,
            "estado":        estado,
        })

    # Organizar por banco
    bancos = {}
    for r in resultados:
        b = r["banco"]
        if b not in bancos:
            bancos[b] = []
        bancos[b].append(r)

    # Crear Excel
    wb = Workbook()
    # Eliminar hoja default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    crear_resumen_ejecutivo(wb, bancos)

    for banco_nombre, prestamos in sorted(bancos.items()):
        crear_hoja_banco(wb, banco_nombre, prestamos)

    wb.save(SALIDA_EXCEL)
    print(f"\nExcel guardado: {SALIDA_EXCEL}")
    print(f"  Tamaño: {SALIDA_EXCEL.stat().st_size // 1024} KB")
    print(f"  Hojas: {wb.sheetnames}")

    # Resumen
    print("\n=== RESUMEN FINAL ===")
    total_cuotas = 0
    for banco_nombre, prestamos in sorted(bancos.items()):
        n_c = sum(len(p["cuotas"]) for p in prestamos)
        total_cuotas += n_c
        print(f"  {banco_nombre}: {len(prestamos)} préstamo(s), {n_c} cuota(s)")

    print(f"\n  TOTAL CUOTAS: {total_cuotas}")
    print(f"  PDFs saltados (contratos): {len(saltados)}")
    for n, r in saltados:
        print(f"    - {n}: {r}")

    pdfs_sin_datos = [r for r in resultados if not r["cuotas"]]
    if pdfs_sin_datos:
        print(f"\n  PDFs SIN DATOS EXTRAÍDOS ({len(pdfs_sin_datos)}):")
        for r in pdfs_sin_datos:
            print(f"    - {r['archivo']}: {r['estado']}")

    return wb, bancos, resultados


if __name__ == "__main__":
    main()
