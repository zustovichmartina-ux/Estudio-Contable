# -*- coding: utf-8 -*-
"""Netos de liquidación Meridiem 04/2025–05/2026 → Excel claro."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    COLOR_PRIMARIO,
    MONEY_FMT,
    TITLE_FONT,
    ZEBRA,
    guardar_informe_excel,
)

SRC = Path(r"T:\CLIENTES\GRUPO MERIDIEM\SUELDOS Y JORNALES\Liquidacion Sueldos Grupo Meridiem.xlsx")
OUT = Path(r"C:\Users\recep\Desktop\Netos_Sueldos_Meridiem_04-2025_05-2026.xlsx")

DESDE = (2025, 4)
HASTA = (2026, 5)

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def parse_sheet_periodo(nombre: str) -> tuple[int, int] | None:
    """'04-2025' → (2025, 4). Ignora SAC / F.931 / variantes."""
    m = re.fullmatch(r"(\d{2})-(\d{4})", str(nombre).strip())
    if not m:
        return None
    mes, anio = int(m.group(1)), int(m.group(2))
    if not (1 <= mes <= 12):
        return None
    return anio, mes


def en_rango(anio: int, mes: int) -> bool:
    return DESDE <= (anio, mes) <= HASTA


def limpia_nombre(n: str) -> str:
    t = re.sub(r"\s+", " ", str(n or "").strip())
    return t


def extraer_hoja(ws) -> list[dict]:
    """Empleados (fila 4) + SUELDO NETO por columna."""
    rows = list(ws.iter_rows(min_row=1, max_row=80, max_col=30, values_only=True))
    if len(rows) < 4:
        return []

    header = rows[3]  # fila 4
    # columnas de empleados: desde E hasta antes de TOTALES
    cols: list[tuple[int, str]] = []
    for idx, val in enumerate(header):
        if idx < 4:
            continue
        txt = limpia_nombre(val)
        if not txt:
            continue
        if txt.upper().startswith("TOTAL"):
            break
        cols.append((idx, txt))

    if not cols:
        return []

    neto_row = None
    for row in rows:
        lab = limpia_nombre(row[1] if len(row) > 1 else "")
        if re.search(r"sueldo\s*neto|^neto\b", lab, re.I):
            neto_row = row
            break
    if neto_row is None:
        return []

    out: list[dict] = []
    for idx, nombre in cols:
        raw = neto_row[idx] if idx < len(neto_row) else None
        if raw is None or raw == "":
            continue
        try:
            neto = float(raw)
        except (TypeError, ValueError):
            continue
        if abs(neto) < 0.005:
            continue
        out.append({"empleado": nombre, "neto": round(neto, 2)})
    return out


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"No está el archivo: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    registros: list[dict] = []
    for sheet_name in wb.sheetnames:
        per = parse_sheet_periodo(sheet_name)
        if not per or not en_rango(*per):
            continue
        anio, mes = per
        etiqueta = f"{MESES[mes]} {anio}"
        periodo = f"{mes:02d}/{anio}"
        empleados = extraer_hoja(wb[sheet_name])
        for e in empleados:
            registros.append(
                {
                    "Periodo": periodo,
                    "Mes": etiqueta,
                    "Empleado": e["empleado"],
                    "Neto": e["neto"],
                    "_ord": (anio, mes),
                }
            )
    wb.close()

    if not registros:
        raise SystemExit("No se extrajeron netos en el rango pedido.")

    registros.sort(key=lambda r: (r["_ord"], r["Empleado"]))
    for r in registros:
        del r["_ord"]

    df = pd.DataFrame(registros)
    resumen = (
        df.groupby("Mes", sort=False)
        .agg(Empleados=("Empleado", "count"), Total_netos=("Neto", "sum"))
        .reset_index()
    )
    # orden cronológico del resumen
    inv = {v: k for k, v in MESES.items()}

    def sk(mes_label: str):
        m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", mes_label)
        if not m:
            return (9999, 13)
        return (int(m.group(2)), inv.get(m.group(1), 13))

    resumen["_o"] = resumen["Mes"].map(sk)
    resumen = resumen.sort_values("_o").drop(columns="_o")

    guardar_informe_excel(
        OUT,
        titulo="Netos de sueldos — GRUPO MERIDIEM SRL",
        subtitulo="Empleados y sueldo neto por mes",
        periodo="04/2025 → 05/2026",
        kpis=[
            ("Liquidaciones (meses)", df["Periodo"].nunique()),
            ("Líneas", len(df)),
            ("Total netos", round(float(df["Neto"].sum()), 2)),
        ],
        resumenes=[("Totales por mes", resumen)],
        detalle=df[["Periodo", "Mes", "Empleado", "Neto"]],
        hoja_detalle="Lista",
        col_moneda=["Neto", "Total_netos"],
        total_col="Neto",
    )

    # Hoja clara: ABRIL 2025 / nombre + neto
    from openpyxl import load_workbook

    wb2 = load_workbook(OUT)
    if "Por mes" in wb2.sheetnames:
        del wb2["Por mes"]
    ws = wb2.create_sheet("Por mes", 0)
    ws["A1"] = "Netos de sueldos — vista clara"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "GRUPO MERIDIEM SRL · 04/2025 a 05/2026"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")

    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in registros:
        grupos[r["Mes"]].append(r)

    fila = 4
    for mes_label in sorted(grupos.keys(), key=sk):
        items = sorted(grupos[mes_label], key=lambda x: x["Empleado"])
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        c = ws.cell(fila, 1, mes_label)
        c.fill = fill_mes
        c.font = font_mes
        ws.cell(fila, 2).fill = fill_mes
        fila += 1
        for i, r in enumerate(items):
            ws.cell(fila, 1, r["Empleado"]).font = BODY_FONT
            ws.cell(fila, 2, r["Neto"]).number_format = MONEY_FMT
            ws.cell(fila, 2).font = BODY_FONT
            if i % 2 == 1:
                ws.cell(fila, 1).fill = ZEBRA
                ws.cell(fila, 2).fill = ZEBRA
            fila += 1
        tot = round(sum(x["Neto"] for x in items), 2)
        ws.cell(fila, 1, f"TOTAL {mes_label}").font = font_tot
        ws.cell(fila, 2, tot).number_format = MONEY_FMT
        ws.cell(fila, 2).font = font_tot
        fila += 2

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    wb2.save(OUT)

    print("OUT", OUT)
    print("MESES", df["Periodo"].nunique(), "LINEAS", len(df))
    print("EMPLEADOS", sorted(df["Empleado"].unique()))
    print(df.groupby("Periodo")["Neto"].sum().to_string())


if __name__ == "__main__":
    main()
