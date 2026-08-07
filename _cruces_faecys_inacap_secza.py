# -*- coding: utf-8 -*-
"""Cruces FAECYS / INACAP / SECZA vs débitos Galicia+Macro — Grupo Meridiem."""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_formato_estudio import (
    BODY_FONT,
    DATE_FMT,
    MONEY_FMT,
    ZEBRA,
    _escribir_encabezado_hoja,
    _pintar_header_fila,
    guardar_informe_excel,
)

ROOT = Path(__file__).resolve().parent
DESK = Path(r"C:\Users\recep\Desktop")
GAL = DESK / "Debitos_Galicia_Meridiem_claro.xlsx"
MAC = DESK / "Debitos_Macro_Meridiem_claro.xlsx"

TOL_IMP = 1.0
TOL_DIAS = 7


def parse_us_money(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if pd.isna(v):
            return None
        return float(v)
    s = str(v).strip().replace(" ", "")
    if not s or s.lower() == "nan":
        return None
    if re.match(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$", s) or re.match(r"^-?\d+(\.\d+)?$", s):
        return float(s.replace(",", ""))
    if "," in s and "." in s:
        # could be AR 1.234,56 or US 1,234.56
        if s.rfind(",") > s.rfind("."):
            return float(s.replace(".", "").replace(",", "."))
        return float(s.replace(",", ""))
    if "," in s:
        return float(s.replace(",", "."))
    return float(s)


def parse_ar_money(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(" ", "").replace("$", "")
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            return float(s.replace(".", "").replace(",", "."))
        return float(s.replace(",", ""))
    if "," in s:
        return float(s.replace(",", "."))
    return float(s)


def to_date(v):
    try:
        if v is None or pd.isna(v):
            return None
    except (TypeError, ValueError):
        if v is None:
            return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime().date()
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in {"nan", "nat", "none"}:
        return None
    ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.to_pydatetime().date()


def load_debitos() -> pd.DataFrame:
    frames = []
    for path, banco in [(GAL, "Galicia"), (MAC, "Macro")]:
        d = pd.read_excel(path, sheet_name="Lista")
        d["Fecha"] = pd.to_datetime(d["Fecha"], errors="coerce")
        d["Importe"] = pd.to_numeric(d["Importe"], errors="coerce").abs()
        d["Detalle"] = d["Detalle"].fillna("").astype(str)
        d["Banco"] = banco
        d = d.dropna(subset=["Fecha", "Importe"])
        d["fecha_d"] = d["Fecha"].dt.date
        frames.append(d[["Banco", "Fecha", "fecha_d", "Importe", "Detalle"]])
    return pd.concat(frames, ignore_index=True)


def fecha_ok(fecha_obl, fecha_b, tol=TOL_DIAS) -> bool:
    if fecha_obl is None or fecha_b is None:
        return False
    try:
        if pd.isna(fecha_obl) or pd.isna(fecha_b):
            return False
    except (TypeError, ValueError):
        pass
    try:
        dias = abs((fecha_b - fecha_obl).days)
    except TypeError:
        return False
    if dias <= tol:
        return True
    return fecha_b.year == fecha_obl.year and fecha_b.month == fecha_obl.month


def en_cobertura(fecha, cov_min, cov_max, margen=7) -> bool:
    if fecha is None:
        return False
    try:
        if pd.isna(fecha):
            return False
    except (TypeError, ValueError):
        pass
    from datetime import timedelta

    try:
        return (cov_min - timedelta(days=margen)) <= fecha <= (cov_max + timedelta(days=margen))
    except TypeError:
        return False


def match_items(
    items: list[dict],
    deb: pd.DataFrame,
    *,
    amount_key: str = "Total",
    date_key: str = "Fecha",
    prefer_kw: re.Pattern | None = None,
    bulk_key_fn=None,
) -> tuple[dict[int, int], dict[int, str], set[int]]:
    """Return matches idx->bank_idx, notes, used_bank."""
    used: set[int] = set()
    matches: dict[int, int] = {}
    notes: dict[int, str] = {}

    # bulk groups first
    if bulk_key_fn:
        groups: dict = defaultdict(list)
        for i, it in enumerate(items):
            k = bulk_key_fn(it)
            if k:
                groups[k].append(i)
        for key, idxs in groups.items():
            if len(idxs) < 2:
                continue
            suma = round(sum(items[i].get(amount_key) or 0 for i in idxs), 2)
            fecha = items[idxs[0]].get(date_key)
            if not fecha or suma <= 0:
                continue
            cands = []
            for j, b in deb.iterrows():
                if j in used:
                    continue
                if abs(float(b["Importe"]) - suma) <= TOL_IMP and fecha_ok(fecha, b["fecha_d"]):
                    pref = 0 if (prefer_kw and prefer_kw.search(b["Detalle"])) else 1
                    cands.append((pref, abs((b["fecha_d"] - fecha).days), abs(b["Importe"] - suma), j))
            cands.sort()
            if cands:
                j = cands[0][-1]
                used.add(j)
                for i in idxs:
                    matches[i] = j
                    notes[i] = f"Pago agrupado: suma ${suma:,.2f} = 1 débito"
                    if len(cands) > 1:
                        notes[i] += f" | {len(cands)} candidatos"

    for i, it in enumerate(items):
        if i in matches:
            continue
        amt = it.get(amount_key)
        fecha = it.get(date_key)
        if amt is None or (isinstance(amt, float) and pd.isna(amt)) or amt <= 0:
            continue
        if fecha is None:
            continue
        cands = []
        for j, b in deb.iterrows():
            if j in used:
                continue
            if abs(float(b["Importe"]) - float(amt)) > TOL_IMP:
                continue
            if not fecha_ok(fecha, b["fecha_d"]):
                continue
            pref = 0 if (prefer_kw and prefer_kw.search(b["Detalle"])) else 1
            cands.append((pref, abs((b["fecha_d"] - fecha).days), abs(b["Importe"] - amt), j))
        if not cands:
            continue
        cands.sort()
        j = cands[0][-1]
        used.add(j)
        matches[i] = j
        notes[i] = "Match importe±$1 y fecha"
        if prefer_kw and prefer_kw.search(deb.loc[j, "Detalle"]):
            notes[i] += " (keyword)"
        if len(cands) > 1:
            notes[i] += f" | {len(cands)} candidatos (elegido el más cercano)"

    return matches, notes, used


def add_candidatos_sheet(path: Path, deb: pd.DataFrame, used: set[int], kw: re.Pattern, titulo: str) -> None:
    wb = load_workbook(path)
    name = "Candidatos banco"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 2)
    fila = _escribir_encabezado_hoja(ws, titulo, "Débitos con keyword; indica si se usó en el cruce", "")
    cand = deb[deb["Detalle"].str.contains(kw, na=False)].copy()
    cand["Usado"] = cand.index.map(lambda j: "Sí" if j in used else "No")
    cand = cand.sort_values("Fecha", ascending=False)
    cols = ["Banco", "Fecha", "Importe", "Detalle", "Usado"]
    for c, h in enumerate(cols, 1):
        ws.cell(fila, c, h)
    _pintar_header_fila(ws, fila, len(cols))
    for ri, (_, r) in enumerate(cand.iterrows(), 1):
        rr = fila + ri
        ws.cell(rr, 1, r["Banco"]).font = BODY_FONT
        cf = ws.cell(rr, 2, r["Fecha"].to_pydatetime())
        cf.number_format = DATE_FMT
        cf.font = BODY_FONT
        ci = ws.cell(rr, 3, float(r["Importe"]))
        ci.number_format = MONEY_FMT
        ci.font = BODY_FONT
        ws.cell(rr, 4, r["Detalle"]).font = BODY_FONT
        ws.cell(rr, 5, r["Usado"]).font = BODY_FONT
        if ri % 2 == 0:
            for c in range(1, 6):
                ws.cell(rr, c).fill = ZEBRA
    if len(cand):
        ws.auto_filter.ref = f"A{fila}:{get_column_letter(5)}{fila + len(cand)}"
    ws.freeze_panes = f"A{fila + 1}"
    for c, w in enumerate([12, 14, 14, 40, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(path)


def build_detalle_and_save(
    *,
    out: Path,
    titulo: str,
    subtitulo: str,
    items: list[dict],
    detalle_rows: list[dict],
    deb: pd.DataFrame,
    used: set[int],
    kw_sheet: re.Pattern,
    cand_title: str,
    cov_min,
    cov_max,
) -> dict:
    detalle = pd.DataFrame(detalle_rows)
    n_pag = int((detalle["Estado"] == "PAGADO").sum())
    n_falta = int((detalle["Estado"] == "FALTA").sum())
    n_parcial = int((detalle["Estado"] == "PARCIAL").sum())
    n_sin = int((detalle["Estado"] == "SIN EXTRACTO").sum())
    n_sd = int((detalle["Estado"] == "SIN DATOS").sum()) if "SIN DATOS" in set(detalle["Estado"]) else 0

    def _sum_est(est):
        return float(detalle.loc[detalle["Estado"] == est, "Total"].fillna(0).sum())

    tot_pag = _sum_est("PAGADO")
    tot_falta = _sum_est("FALTA")
    tot_all = float(detalle["Total"].fillna(0).sum())

    res_est = (
        detalle.groupby("Estado", dropna=False)
        .agg(Cantidad=("Estado", "count"), Total=("Total", "sum"))
        .reset_index()
    )
    orden = {"PAGADO": 0, "PARCIAL": 1, "FALTA": 2, "SIN EXTRACTO": 3, "SIN DATOS": 4}
    res_est["_o"] = res_est["Estado"].map(lambda x: orden.get(x, 9))
    res_est = res_est.sort_values("_o").drop(columns="_o")

    kpis = [
        ("Ítems", len(detalle), "int"),
        ("Pagados", n_pag, "int"),
        ("Faltan (cobertura)", n_falta, "int"),
        ("Parciales", n_parcial, "int"),
        ("Sin extracto", n_sin, "int"),
        ("Total obligaciones", tot_all, "money"),
        ("Total pagado (cruzado)", tot_pag, "money"),
        ("Total falta (cobertura)", tot_falta, "money"),
        (
            "Cobertura extractos",
            f"{cov_min.strftime('%d/%m/%Y')} – {cov_max.strftime('%d/%m/%Y')}",
            "text",
        ),
    ]
    if n_sd:
        kpis.insert(5, ("Sin datos en fuente", n_sd, "int"))

    # sort detalle
    det_out = detalle.copy()
    det_out["_o"] = det_out["Estado"].map(lambda x: orden.get(x, 9))
    if "Periodo_ord" in det_out.columns:
        det_out = det_out.sort_values(["_o", "Periodo_ord"], ascending=[True, False])
        det_out = det_out.drop(columns=["_o", "Periodo_ord"])
    else:
        det_out = det_out.sort_values("_o").drop(columns=["_o"])

    guardar_informe_excel(
        out,
        titulo=titulo,
        subtitulo=subtitulo,
        periodo=(
            f"Extractos: {cov_min.strftime('%d/%m/%Y')} a {cov_max.strftime('%d/%m/%Y')} | "
            "Criterio: importe ±$1 y fecha ±7 días o mismo mes"
        ),
        kpis=kpis,
        resumenes=[("Por estado", res_est)],
        detalle=det_out,
        hoja_detalle="Detalle",
        col_moneda=[
            "Total",
            "Importe 0,5%",
            "Interés",
            "Importe banco",
            "Diferencia",
            "Monto Total",
            "2%",
            "Mora",
        ],
        col_fecha=["Fecha pago", "Fecha banco", "Fecha venc.", "Fecha pago FAECYS", "Fecha Pago Real", "Fecha Pago Declarada"],
        total_col="Total",
    )
    add_candidatos_sheet(out, deb, used, kw_sheet, cand_title)
    return {
        "path": str(out),
        "pagados": n_pag,
        "faltan": n_falta,
        "parciales": n_parcial,
        "sin_extracto": n_sin,
        "sin_datos": n_sd,
        "tot_pag": tot_pag,
        "tot_falta": tot_falta,
        "tot_all": tot_all,
    }


# ── FAECYS ───────────────────────────────────────────────────────────────────

def run_faecys(deb: pd.DataFrame, cov_min, cov_max) -> dict:
    path = Path(r"c:\Users\recep\Downloads\Pagos_Cuit_30714058386_05-08-2026.xls")
    raw = pd.read_excel(path, engine="xlrd", header=0)
    raw = raw[raw["Periodo"].notna()].copy()

    items = []
    for _, r in raw.iterrows():
        base = parse_us_money(r["Importe 0,5%"])
        total = parse_us_money(r["Total"])
        interes = None
        if total is not None and base is not None:
            interes = round(total - base, 2)
        elif base is not None:
            interes = 0.0
            total = base
        periodo = int(float(r["Periodo"]))
        y, m = divmod(periodo, 100)
        fecha = to_date(r["Fecha de pago"])
        acta = r["Nro. de Acta"]
        if pd.isna(acta):
            acta = None
        else:
            try:
                acta = int(float(acta))
            except Exception:
                acta = str(acta)
        items.append(
            {
                "Periodo": f"{m:02d}/{y}",
                "Periodo_ord": periodo,
                "Fecha": fecha,
                "Nro_Acta": acta,
                "Importe_05": base,
                "Interes": interes,
                "Total": total,
            }
        )

    kw = re.compile(r"faecys", re.I)

    def bulk_key(it):
        if it["Fecha"] and it["Nro_Acta"] and it["Total"]:
            return (it["Fecha"], it["Nro_Acta"])
        return None

    matches, notes, used = match_items(
        items, deb, amount_key="Total", date_key="Fecha", prefer_kw=kw, bulk_key_fn=bulk_key
    )

    rows = []
    for i, it in enumerate(items):
        total = it["Total"]
        fecha = it["Fecha"]
        if total is None:
            if it["Periodo_ord"] >= 202504:
                estado = "FALTA"
                nota = "Sin importe/fecha en FAECYS (período sin pago declarado)"
            else:
                estado = "SIN DATOS"
                nota = "Sin importe ni fecha en export FAECYS"
            rows.append(
                {
                    "Periodo": it["Periodo"],
                    "Periodo_ord": it["Periodo_ord"],
                    "Fecha pago FAECYS": fecha,
                    "Nro. Acta": it["Nro_Acta"],
                    "Importe 0,5%": it["Importe_05"],
                    "Interés": it["Interes"],
                    "Total": None,
                    "Estado": estado,
                    "Banco": None,
                    "Fecha banco": None,
                    "Importe banco": None,
                    "Diferencia": None,
                    "Detalle banco": None,
                    "Nota cruce": nota,
                }
            )
            continue

        if i in matches:
            j = matches[i]
            b = deb.loc[j]
            estado = "PAGADO"
            rows.append(
                {
                    "Periodo": it["Periodo"],
                    "Periodo_ord": it["Periodo_ord"],
                    "Fecha pago FAECYS": fecha,
                    "Nro. Acta": it["Nro_Acta"],
                    "Importe 0,5%": it["Importe_05"],
                    "Interés": it["Interes"],
                    "Total": total,
                    "Estado": estado,
                    "Banco": b["Banco"],
                    "Fecha banco": b["fecha_d"],
                    "Importe banco": float(b["Importe"]),
                    "Diferencia": round(float(b["Importe"]) - total, 2) if i not in notes or "agrupado" not in notes.get(i, "").lower() else 0.0,
                    "Detalle banco": b["Detalle"],
                    "Nota cruce": notes.get(i, ""),
                }
            )
            if "agrupado" in notes.get(i, "").lower():
                rows[-1]["Diferencia"] = 0.0
            continue

        if fecha and en_cobertura(fecha, cov_min, cov_max):
            estado = "FALTA"
            nota = "No hay débito por importe±$1 y fecha ±7d / mismo mes"
        elif fecha:
            estado = "SIN EXTRACTO"
            nota = f"Fuera de cobertura extractos ({cov_min} a {cov_max})"
        else:
            estado = "FALTA" if it["Periodo_ord"] >= 202504 else "SIN DATOS"
            nota = "Sin fecha de pago en FAECYS"

        rows.append(
            {
                "Periodo": it["Periodo"],
                "Periodo_ord": it["Periodo_ord"],
                "Fecha pago FAECYS": fecha,
                "Nro. Acta": it["Nro_Acta"],
                "Importe 0,5%": it["Importe_05"],
                "Interés": it["Interes"],
                "Total": total,
                "Estado": estado,
                "Banco": None,
                "Fecha banco": None,
                "Importe banco": None,
                "Diferencia": None,
                "Detalle banco": None,
                "Nota cruce": nota,
            }
        )

    return build_detalle_and_save(
        out=DESK / "Cruce_FAECYS_vs_Bancos_Meridiem.xlsx",
        titulo="Cruce FAECYS vs Bancos — Grupo Meridiem SRL",
        subtitulo="CUIT 30-71405838-6 | Aporte 0,5% FAECYS vs débitos Galicia/Macro",
        items=items,
        detalle_rows=rows,
        deb=deb,
        used=used,
        kw_sheet=kw,
        cand_title="Candidatos FAECYS en extractos",
        cov_min=cov_min,
        cov_max=cov_max,
    )


# ── INACAP ───────────────────────────────────────────────────────────────────

def run_inacap(deb: pd.DataFrame, cov_min, cov_max) -> dict:
    path = Path(r"c:\Users\recep\Downloads\reporteDePagos (3).xls")
    raw = pd.read_excel(path, engine="xlrd", header=1)
    items = []
    for _, r in raw.iterrows():
        if pd.isna(r.get("Periodo")):
            continue
        per = pd.to_datetime(r["Periodo"], errors="coerce")
        if pd.isna(per):
            continue
        monto = parse_us_money(r["Monto Total"])
        if monto is None:
            monto = float(r["Monto Total"]) if pd.notna(r["Monto Total"]) else None
        f_real = to_date(r["Fecha Pago Real"])
        f_decl = to_date(r["Fecha Pago Declarada"])
        fecha = f_real or f_decl
        estado_src = str(r.get("Estado") or "").strip()
        items.append(
            {
                "Periodo": per.strftime("%m/%Y"),
                "Periodo_ord": int(per.strftime("%Y%m")),
                "Empleados": int(r["Cantidad Empleados"]) if pd.notna(r["Cantidad Empleados"]) else None,
                "Fecha": fecha,
                "Fecha_real": f_real,
                "Fecha_decl": f_decl,
                "Total": round(monto, 2) if monto is not None else None,
                "Estado_src": estado_src,
            }
        )

    kw = re.compile(r"inacap|wnpower", re.I)
    matches, notes, used = match_items(
        items, deb, amount_key="Total", date_key="Fecha", prefer_kw=kw
    )

    rows = []
    for i, it in enumerate(items):
        total = it["Total"]
        fecha = it["Fecha"]
        if i in matches:
            j = matches[i]
            b = deb.loc[j]
            rows.append(
                {
                    "Periodo": it["Periodo"],
                    "Periodo_ord": it["Periodo_ord"],
                    "Empleados": it["Empleados"],
                    "Estado INACAP": it["Estado_src"],
                    "Fecha Pago Declarada": it["Fecha_decl"],
                    "Fecha Pago Real": it["Fecha_real"],
                    "Total": total,
                    "Estado": "PAGADO",
                    "Banco": b["Banco"],
                    "Fecha banco": b["fecha_d"],
                    "Importe banco": float(b["Importe"]),
                    "Diferencia": round(float(b["Importe"]) - (total or 0), 2),
                    "Detalle banco": b["Detalle"],
                    "Nota cruce": notes.get(i, ""),
                }
            )
            continue

        # Emitida without bank match
        if it["Estado_src"].lower() == "emitida":
            if fecha and en_cobertura(fecha, cov_min, cov_max):
                estado = "FALTA"
                nota = "INACAP Emitida — sin débito bancario en cobertura"
            elif fecha:
                estado = "SIN EXTRACTO"
                nota = "Emitida; fuera de cobertura extractos"
            else:
                estado = "FALTA"
                nota = "Emitida sin fecha de pago real"
        elif fecha and en_cobertura(fecha, cov_min, cov_max):
            estado = "FALTA"
            nota = "Pagada en INACAP pero sin débito ±$1/fecha en extractos"
        elif fecha:
            estado = "SIN EXTRACTO"
            nota = f"Fuera de cobertura extractos ({cov_min} a {cov_max})"
        else:
            estado = "FALTA"
            nota = "Sin fecha de pago"

        rows.append(
            {
                "Periodo": it["Periodo"],
                "Periodo_ord": it["Periodo_ord"],
                "Empleados": it["Empleados"],
                "Estado INACAP": it["Estado_src"],
                "Fecha Pago Declarada": it["Fecha_decl"],
                "Fecha Pago Real": it["Fecha_real"],
                "Total": total,
                "Estado": estado,
                "Banco": None,
                "Fecha banco": None,
                "Importe banco": None,
                "Diferencia": None,
                "Detalle banco": None,
                "Nota cruce": nota,
            }
        )

    return build_detalle_and_save(
        out=DESK / "Cruce_INACAP_vs_Bancos_Meridiem.xlsx",
        titulo="Cruce INACAP vs Bancos — Grupo Meridiem SRL",
        subtitulo="CUIT 30-71405838-6 | Cuotas INACAP vs débitos Galicia/Macro",
        items=items,
        detalle_rows=rows,
        deb=deb,
        used=used,
        kw_sheet=kw,
        cand_title="Candidatos INACAP/WNPOWER en extractos",
        cov_min=cov_min,
        cov_max=cov_max,
    )


# ── SECZA ────────────────────────────────────────────────────────────────────

def parse_secza_pdf(pdf_path: Path) -> list[dict]:
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    # Main data lines: 2026/06 1917325.29 ... total fvenc ...
    # Also payment lines above: 42,416.94 $ 21/07/2026
    line_re = re.compile(
        r"(?P<per>\d{4}/\d{2})\s+"
        r"(?P<sueldo>[\d.]+)\s+"
        r"(?P<norem>[\d.]+)\s+C/A\s+"
        r"(?P<emp>\d+)\s+"
        r"(?P<pct>[\d.]+)\s+"
        r"(?P<cuotas>[\d.]+)\s+"
        r"(?P<ctafam>[\d.]+)\s+"
        r"(?P<otros>[\d.]+)\s+"
        r"(?P<mora>[\d.,]+)\s+"
        r"(?P<total>[\d.]+)\s+"
        r"(?P<fvenc>\d{2}/\d{2}/\d{4})"
        r"(?P<rest>.*)"
    )
    pay_re = re.compile(
        r"(?P<amt>\d{1,3}(?:[.,]\d{3})*(?:[.,]\d{2})|\d+[.,]\d{2})\s*\$\s*(?P<f>\d{2}/\d{2}/\d{4})"
    )

    items = []
    lines = text.splitlines()
    pending_pay = None  # (amount, date) from line before period

    for line in lines:
        line = line.strip()
        mpay = pay_re.fullmatch(line) or pay_re.search(line)
        # payment-only lines like "42,416.94 $ 21/07/2026"
        if re.match(r"^[\d.,]+\s*\$\s*\d{2}/\d{2}/\d{4}", line):
            mm = pay_re.search(line)
            if mm:
                pending_pay = (parse_ar_money(mm.group("amt")), to_date(mm.group("f")))
            continue

        m = line_re.search(line)
        if not m:
            continue
        per = m.group("per")  # YYYY/MM
        y, mo = per.split("/")
        total = parse_us_money(m.group("total"))
        mora_raw = m.group("mora").replace(",", ".") if "," in m.group("mora") and "." not in m.group("mora") else m.group("mora")
        # mora sometimes 877,60
        mora = parse_ar_money(m.group("mora"))
        fvenc = to_date(m.group("fvenc"))
        rest = m.group("rest") or ""

        # payment embedded in same line: "39.475,24 $ 14/01/2026 Constancia"
        fecha_pago = None
        imp_pago = None
        emb = pay_re.search(rest)
        if emb:
            imp_pago = parse_ar_money(emb.group("amt"))
            fecha_pago = to_date(emb.group("f"))
        elif pending_pay:
            imp_pago, fecha_pago = pending_pay

        estado_src = []
        if "Constancia" in rest or (emb and "Constancia" in rest):
            estado_src.append("Constancia")
        if "Fuera de Termino" in rest or "P.Fuera" in line:
            estado_src.append("P.Fuera de Termino")
        # pending pay lines often paired with Fuera de Termino on next conceptual block
        if pending_pay and not emb:
            estado_src.append("P.Fuera de Termino")

        items.append(
            {
                "Periodo": f"{mo}/{y}",
                "Periodo_ord": int(y) * 100 + int(mo),
                "2%": parse_us_money(m.group("pct")),
                "Mora": mora,
                "Total": total,
                "Fecha_venc": fvenc,
                "Fecha": fecha_pago or fvenc,  # match date preference: real payment then venc
                "Fecha_pago": fecha_pago,
                "Importe_pago_pdf": imp_pago,
                "Estado_src": " / ".join(estado_src) if estado_src else "",
            }
        )
        pending_pay = None

    return items


def run_secza(deb: pd.DataFrame, cov_min, cov_max) -> dict:
    pdf = DESK / "Rehabilitaciones SECZA.pdf"
    items = parse_secza_pdf(pdf)
    kw = re.compile(r"secza|sindicat|rehabilit|cuota.?sind|emplead", re.I)
    prefer = re.compile(r"sindicat|secza|emplead", re.I)

    def bulk_key(it):
        # Agrupar por fecha de pago del PDF (varios períodos rehabilitados el mismo día)
        if it.get("Fecha_pago") and it.get("Total"):
            return it["Fecha_pago"]
        return None

    matches, notes, used = match_items(
        items,
        deb,
        amount_key="Total",
        date_key="Fecha",
        prefer_kw=prefer,
        bulk_key_fn=bulk_key,
    )

    # Bulk por misma fecha de pago: si la fecha PDF no cae cerca del banco,
    # probar suma vs débitos keyword (Sindicato) en cobertura.
    from collections import defaultdict as _dd

    by_pay = _dd(list)
    for i, it in enumerate(items):
        if i in matches:
            continue
        if it.get("Fecha_pago") and it.get("Total"):
            by_pay[it["Fecha_pago"]].append(i)
    for fecha_p, idxs in by_pay.items():
        if len(idxs) < 2:
            continue
        if any(i in matches for i in idxs):
            continue
        suma = round(sum(items[i]["Total"] or 0 for i in idxs), 2)
        fechas_try = {fecha_p}
        for i in idxs:
            if items[i].get("Fecha_venc"):
                fechas_try.add(items[i]["Fecha_venc"])
        cands = []
        for j, b in deb.iterrows():
            if j in used:
                continue
            if abs(float(b["Importe"]) - suma) > TOL_IMP:
                continue
            ok = any(fecha_ok(f, b["fecha_d"]) for f in fechas_try if f)
            # también aceptar si keyword sindicato y fecha banco dentro de cobertura
            if not ok and prefer.search(b["Detalle"]) and en_cobertura(b["fecha_d"], cov_min, cov_max):
                ok = True
            if not ok:
                continue
            pref = 0 if prefer.search(b["Detalle"]) else 1
            cands.append((pref, abs(float(b["Importe"]) - suma), j))
        if cands:
            cands.sort()
            j = cands[0][-1]
            used.add(j)
            for i in idxs:
                matches[i] = j
                notes[i] = f"Pago agrupado: suma ${suma:,.2f} = 1 débito"

    # Reintentar individuales con fecha de vencimiento / pago
    for i, it in enumerate(items):
        if i in matches or not it.get("Total"):
            continue
        for fecha_try in (it.get("Fecha_pago"), it.get("Fecha_venc")):
            if not fecha_try:
                continue
            cands = []
            for j, b in deb.iterrows():
                if j in used:
                    continue
                if abs(float(b["Importe"]) - float(it["Total"])) > TOL_IMP:
                    continue
                if not fecha_ok(fecha_try, b["fecha_d"]):
                    continue
                pref = 0 if prefer.search(b["Detalle"]) else 1
                cands.append((pref, abs((b["fecha_d"] - fecha_try).days), j))
            if cands:
                cands.sort()
                j = cands[0][-1]
                used.add(j)
                matches[i] = j
                notes[i] = f"Match por fecha {fecha_try}"
                break

    # Último recurso: mismo importe + keyword en cobertura (fecha PDF dudosa)
    for i, it in enumerate(items):
        if i in matches or not it.get("Total"):
            continue
        fecha_ref = it.get("Fecha_pago") or it.get("Fecha_venc")
        if fecha_ref and not en_cobertura(fecha_ref, cov_min, cov_max):
            continue
        cands = []
        for j, b in deb.iterrows():
            if j in used:
                continue
            if abs(float(b["Importe"]) - float(it["Total"])) > TOL_IMP:
                continue
            if not prefer.search(b["Detalle"]):
                continue
            if not en_cobertura(b["fecha_d"], cov_min, cov_max):
                continue
            cands.append((abs((b["fecha_d"] - (fecha_ref or b["fecha_d"])).days), j))
        if len(cands) == 1:
            j = cands[0][-1]
            used.add(j)
            matches[i] = j
            notes[i] = "Match importe+keyword (fecha PDF vs banco divergente)"

    rows = []
    for i, it in enumerate(items):
        total = it["Total"]
        fecha_ref = it.get("Fecha_pago") or it.get("Fecha_venc")
        if i in matches:
            j = matches[i]
            b = deb.loc[j]
            dif = round(float(b["Importe"]) - (total or 0), 2)
            if "agrupado" in notes.get(i, "").lower():
                dif = 0.0
            estado = "PAGADO" if abs(dif) <= TOL_IMP or "agrupado" in notes.get(i, "").lower() else "PARCIAL"
            rows.append(
                {
                    "Periodo": it["Periodo"],
                    "Periodo_ord": it["Periodo_ord"],
                    "Estado SECZA": it["Estado_src"],
                    "Fecha venc.": it["Fecha_venc"],
                    "Fecha pago": it["Fecha_pago"],
                    "2%": it["2%"],
                    "Mora": it["Mora"],
                    "Total": total,
                    "Estado": estado,
                    "Banco": b["Banco"],
                    "Fecha banco": b["fecha_d"],
                    "Importe banco": float(b["Importe"]),
                    "Diferencia": dif,
                    "Detalle banco": b["Detalle"],
                    "Nota cruce": notes.get(i, ""),
                }
            )
            continue

        if fecha_ref and en_cobertura(fecha_ref, cov_min, cov_max):
            estado = "FALTA"
            nota = "Sin débito bancario por importe±$1 y fecha cercana"
        elif fecha_ref:
            estado = "SIN EXTRACTO"
            nota = f"Fuera de cobertura extractos ({cov_min} a {cov_max})"
        else:
            estado = "FALTA"
            nota = "Sin fecha en PDF"

        rows.append(
            {
                "Periodo": it["Periodo"],
                "Periodo_ord": it["Periodo_ord"],
                "Estado SECZA": it["Estado_src"],
                "Fecha venc.": it["Fecha_venc"],
                "Fecha pago": it["Fecha_pago"],
                "2%": it["2%"],
                "Mora": it["Mora"],
                "Total": total,
                "Estado": estado,
                "Banco": None,
                "Fecha banco": None,
                "Importe banco": None,
                "Diferencia": None,
                "Detalle banco": None,
                "Nota cruce": nota,
            }
        )

    return build_detalle_and_save(
        out=DESK / "Cruce_SECZA_vs_Bancos_Meridiem.xlsx",
        titulo="Cruce SECZA (Rehabilitaciones) vs Bancos — Grupo Meridiem",
        subtitulo="Sindicato SECZA — cuotas/rehabilitaciones vs débitos Galicia/Macro",
        items=items,
        detalle_rows=rows,
        deb=deb,
        used=used,
        kw_sheet=kw,
        cand_title="Candidatos sindicato/SECZA en extractos",
        cov_min=cov_min,
        cov_max=cov_max,
    )


def main():
    deb = load_debitos()
    cov_min, cov_max = deb["fecha_d"].min(), deb["fecha_d"].max()
    print(f"Debitos: {len(deb)} | cobertura {cov_min} a {cov_max}")

    r1 = run_faecys(deb, cov_min, cov_max)
    print("FAECYS", {k: (round(v, 2) if isinstance(v, float) else v) for k, v in r1.items()})

    r2 = run_inacap(deb, cov_min, cov_max)
    print("INACAP", {k: (round(v, 2) if isinstance(v, float) else v) for k, v in r2.items()})

    r3 = run_secza(deb, cov_min, cov_max)
    print("SECZA", {k: (round(v, 2) if isinstance(v, float) else v) for k, v in r3.items()})

if __name__ == "__main__":
    main()
