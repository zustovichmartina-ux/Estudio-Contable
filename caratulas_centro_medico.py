# -*- coding: utf-8 -*-
"""
Extracción de carátulas / liquidaciones de honorarios
Centro Médico de Mar del Plata (Asociación Civil).

Uso:
    from caratulas_centro_medico import consolidar_carpeta
    consolidar_carpeta(r\"C:\\Users\\recep\\Desktop\\Centro Medico\")

PDF típico: Caratula_00796356_XXXXXXX.pdf
Montos en el PDF vienen estilo US (1,234.56).
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Conceptos del cuadro (orden fijo) ────────────────────────────────────────
# Incluye retenciones que restan (IIBB, derecho admin, etc.).
CONCEPTOS: list[tuple[str, str]] = [
    ("Valor Honorarios Exentos", r"Valor Honorarios Exento\s*\.*\s*:\s*([-\d,\.]+)"),
    (
        "Neto Gravado Honorarios",
        r"Valor Honorarios Exento\s*\.*\s*:\s*[-\d,\.]+\s*,\s*Neto Gravado:\s*([-\d,\.]+)",
    ),
    ("Valor Gastos Exentos", r"Valor Gastos Exento\s*\.*\s*:\s*([-\d,\.]+)"),
    (
        "Neto Gravado Gastos",
        r"Valor Gastos Exento\s*\.*\s*:\s*[-\d,\.]+\s*,\s*Neto Gravado:\s*([-\d,\.]+)",
    ),
    ("Valor IVA", r"Valor I\.?V\.?A\.?\s*:\s*([-\d,\.]+)"),
    # 5%, 4,5% o bonificado 5,75% — primer importe de la línea
    (
        "Derecho administrativo",
        r"DERECHO ADMINISTRATIVO(?:\s+BONIFICADO)?(?:\s+AL)?\s*[\d.,]+\s*%\s*([-\d,\.]+)",
    ),
    ("Aporte fondo compensador", r"APORTE FONDO COMPENSADOR\s+[\d.,]+\s*%\s*([-\d,\.]+)"),
    # Retención IIBB (NO la base imponible)
    (
        "Retencion IIBB",
        r"RET\.?\s*IIBB\s+DN[^\n]*?\([\d.\s]+%\)\s*([-\d,\.]+)",
    ),
    ("Jubilacion", r"JUBILACION\s*([-\d,\.]+)"),
    ("Impuesto a las ganancias", r"IMP\.\s*A LAS GCIAS\.[^\n]*?\s([-\d,\.]+)"),
    ("Recuperaciones bancarias", r"RECUP\.?\s*TRANSACCIONES BANCARIAS\s*([-\d,\.]+)"),
]

MONEY_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Calibri", bold=True, size=16, color="1F4E79")
FECHA_LBL = Font(name="Calibri", bold=True, size=12, color="1F4E79")
FECHA_VAL = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUB = Font(name="Calibri", size=11, color="666666")
BODY = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
ZEBRA = PatternFill("solid", fgColor="F2F2F2")
SECTION = Font(name="Calibri", bold=True, size=12, color="1F4E79")

DEFAULT_DIR = Path(r"C:\Users\recep\Desktop\Centro Medico")
DEFAULT_OUT = DEFAULT_DIR / "Liquidaciones_Centro_Medico_consolidado.xlsx"


def parse_us_money(s: str) -> float | None:
    """PDF Centro Médico: miles con coma, decimales con punto (1,234.56)."""
    s = re.sub(r"[^\d,.\-]", "", str(s).strip())
    if not s or s in ("-", "."):
        return None
    if re.search(r",\d{3}(\.|$)", s) or ("," in s and "." in s and s.rfind(".") > s.rfind(",")):
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def grab(text: str, pat: str) -> float | None:
    m = re.search(pat, text, re.I)
    return parse_us_money(m.group(1)) if m else None


def grab_sum(text: str, pat: str) -> float | None:
    """Suma todas las coincidencias (p.ej. varios derechos administrativos)."""
    vals = [parse_us_money(m) for m in re.findall(pat, text, re.I)]
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return round(sum(vals), 2)


# Conceptos que pueden aparecer más de una vez en la misma liquidación
_CONCEPTOS_SUMAR = {"Derecho administrativo"}


def parse_caratula(path: str | Path) -> dict[str, Any] | None:
    """Lee un PDF Caratula_*.pdf y devuelve dict con fecha + importes."""
    path = Path(path)
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception:
        return None
    if not text.strip():
        return None

    m = re.search(r"Fecha:\s*(\d{1,2}/\d{1,2}/\d{4})", text, re.I)
    fecha: date | None = datetime.strptime(m.group(1), "%d/%m/%Y").date() if m else None
    prof = re.search(r"del Dr\.:\s*(.+)", text)
    mat = re.search(r"Matr\.N.\s*:\s*([\d,]+)", text)
    liq = re.search(r"Liquidaci.n N./Cert\.:\s*([\d,]+)", text)
    cuit = re.search(r"CUIT:\s*([\d\-]+)", text)

    importes: dict[str, float | None] = {}
    for n, pat in CONCEPTOS:
        importes[n] = grab_sum(text, pat) if n in _CONCEPTOS_SUMAR else grab(text, pat)
    total = round(sum(v or 0.0 for v in importes.values()), 2)

    liq_nro = liq.group(1) if liq else ""
    if not liq_nro:
        m2 = re.search(r"_(\d+)\.pdf$", path.name, re.I)
        liq_nro = m2.group(1) if m2 else path.stem

    return {
        "archivo": path.name,
        "fecha": fecha,
        "profesional": (prof.group(1).strip() if prof else ""),
        "matricula": (mat.group(1) if mat else ""),
        "liquidacion": liq_nro,
        "cuit": (cuit.group(1) if cuit else ""),
        "importes": importes,
        "total": total,
        "tiene_iibb": bool(importes.get("Retencion IIBB")),
    }


def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        if c > 1:
            cell.alignment = Alignment(horizontal="right")


def escribir_excel_consolidado(registros: list[dict[str, Any]], destino: str | Path) -> Path:
    """
    Excel con 3 hojas:
      - Resumen: una fila por liquidación
      - Detalle: FECHA arriba + cuadro Concepto/Importe por liquidación
      - Lista: plana para filtros
    """
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    registros = sorted(
        registros,
        key=lambda r: (r["fecha"] or datetime.min.date(), str(r["liquidacion"])),
    )
    wb = Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Liquidaciones Centro Medico MDP - Consolidado"
    ws["A1"].font = TITLE
    n_iibb = sum(1 for r in registros if r.get("tiene_iibb"))
    ws["A2"] = f"{len(registros)} liquidaciones | {n_iibb} con retencion IIBB"
    ws["A2"].font = SUB

    headers = [
        "Fecha",
        "Liquidacion",
        "Profesional",
        "Matricula",
        "CUIT",
        *[n for n, _ in CONCEPTOS],
        "TOTAL",
        "Archivo",
    ]
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(header_row, i, h)
    _style_header(ws, header_row, len(headers))

    for i, rec in enumerate(registros):
        r = header_row + 1 + i
        ws.cell(r, 1, rec["fecha"]).number_format = "DD/MM/YYYY"
        ws.cell(r, 1).font = BODY
        ws.cell(r, 2, rec["liquidacion"]).font = BODY
        ws.cell(r, 3, rec["profesional"]).font = BODY
        ws.cell(r, 4, rec["matricula"]).font = BODY
        ws.cell(r, 5, rec["cuit"]).font = BODY
        for j, (nombre, _) in enumerate(CONCEPTOS):
            c = ws.cell(r, 6 + j, float((rec.get("importes") or {}).get(nombre) or 0.0))
            c.number_format = MONEY_FMT
            c.font = BODY
        tot = ws.cell(r, 6 + len(CONCEPTOS), float(rec.get("total") or 0.0))
        tot.number_format = MONEY_FMT
        tot.font = BOLD
        ws.cell(r, 7 + len(CONCEPTOS), rec.get("archivo") or "").font = BODY
        if i % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = ZEBRA

    last = header_row + len(registros)
    if registros:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = f"A{header_row + 1}"
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(42, max(12, len(h) + 2))

    # Detalle
    wd = wb.create_sheet("Detalle")
    wd["A1"] = "Detalle por liquidacion"
    wd["A1"].font = TITLE
    wd["A2"] = "FECHA arriba de cada cuadro Concepto / Importe"
    wd["A2"].font = SUB
    row = 4
    for rec in registros:
        wd.cell(row, 1, "FECHA").font = FECHA_LBL
        cfecha = wd.cell(row, 2, rec["fecha"])
        cfecha.number_format = "DD/MM/YYYY"
        cfecha.font = FECHA_VAL
        row += 1
        wd.cell(row, 1, f"Liquidacion {rec['liquidacion']}").font = SECTION
        wd.cell(row, 2, rec["profesional"]).font = SUB
        row += 1
        hdr = row
        wd.cell(hdr, 1, "Concepto").font = HDR_FONT
        wd.cell(hdr, 1).fill = HDR_FILL
        wd.cell(hdr, 2, "Importe").font = HDR_FONT
        wd.cell(hdr, 2).fill = HDR_FILL
        wd.cell(hdr, 2).alignment = Alignment(horizontal="right")
        row += 1
        for i, (nombre, _) in enumerate(CONCEPTOS):
            wd.cell(row, 1, nombre).font = BODY
            c = wd.cell(row, 2, float((rec.get("importes") or {}).get(nombre) or 0.0))
            c.number_format = MONEY_FMT
            c.font = BODY
            if i % 2 == 1:
                wd.cell(row, 1).fill = ZEBRA
                wd.cell(row, 2).fill = ZEBRA
            row += 1
        wd.cell(row, 1, "TOTAL").font = BOLD
        tot = wd.cell(row, 2, float(rec.get("total") or 0.0))
        tot.number_format = MONEY_FMT
        tot.font = BOLD
        row += 2
    wd.column_dimensions["A"].width = 50
    wd.column_dimensions["B"].width = 18
    wd.freeze_panes = "A4"

    # Lista
    wl = wb.create_sheet("Lista")
    wl["A1"] = "Lista plana (filtros)"
    wl["A1"].font = TITLE
    cols = ["Fecha", "Liquidacion", "Profesional", "Concepto", "Importe"]
    for i, h in enumerate(cols, 1):
        wl.cell(3, i, h)
    _style_header(wl, 3, len(cols))
    r = 4
    for rec in registros:
        for nombre, _ in CONCEPTOS:
            wl.cell(r, 1, rec["fecha"]).number_format = "DD/MM/YYYY"
            wl.cell(r, 2, rec["liquidacion"])
            wl.cell(r, 3, rec["profesional"])
            wl.cell(r, 4, nombre)
            c = wl.cell(r, 5, float((rec.get("importes") or {}).get(nombre) or 0.0))
            c.number_format = MONEY_FMT
            for col in range(1, 6):
                wl.cell(r, col).font = BODY
            r += 1
    if r > 4:
        wl.auto_filter.ref = f"A3:E{r - 1}"
    wl.freeze_panes = "A4"
    for i, w in enumerate([12, 14, 40, 48, 16], 1):
        wl.column_dimensions[get_column_letter(i)].width = w

    wb.save(destino)
    return destino


def consolidar_carpeta(
    carpeta: str | Path | None = None,
    destino: str | Path | None = None,
    glob_pat: str = "Caratula_*.pdf",
) -> Path:
    """Procesa todos los PDF Caratula_*.pdf de la carpeta y genera el Excel."""
    carpeta = Path(carpeta or DEFAULT_DIR)
    destino = Path(destino or (carpeta / DEFAULT_OUT.name))
    pdfs = sorted(carpeta.glob(glob_pat))
    if not pdfs:
        raise FileNotFoundError(f"No hay {glob_pat} en {carpeta}")
    registros = []
    for pdf in pdfs:
        rec = parse_caratula(pdf)
        if rec:
            registros.append(rec)
    if not registros:
        raise RuntimeError("No se pudo leer ninguna carátula")
    return escribir_excel_consolidado(registros, destino)


if __name__ == "__main__":
    out = consolidar_carpeta()
    print("OK", out)
