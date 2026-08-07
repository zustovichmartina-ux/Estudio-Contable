"""Prueba modo teórico (solo saldos iniciales, sin PDF ni OCR)."""
import time
from pathlib import Path

import openpyxl

from procesador import (
    detectar_modo_auditoria,
    ejecutar_auditoria_prestamos,
    generar_planilla_teorica_bancos,
    generar_reporte_auditoria_prestamos,
    _sanitizar_nombre_hoja_excel,
)

SALIDA = Path("exportaciones/Auditoria_Prestamos_teorico_test.xlsx")


def test_detectar_modo():
    assert detectar_modo_auditoria(["a.pdf"], [], None) == "teorico"
    assert detectar_modo_auditoria([], [], None) == "teorico"
    assert detectar_modo_auditoria(["a.pdf"], ["b.pdf"], None) == "parcial_extractos"
    assert detectar_modo_auditoria(["a.pdf"], ["b.pdf"], "mayor.xlsx") == "completo"
    assert detectar_modo_auditoria([], ["b.pdf"], None) == "solo_extractos"


def test_pipeline_teorico_sin_pdf():
    saldos = [
        {"banco": "Banco Galicia", "saldo_inicial": 500000.0},
        {"banco": "Banco Santander", "saldo_inicial": 1200000.0},
    ]

    t0 = time.perf_counter()
    resultado = ejecutar_auditoria_prestamos(
        pdf_prestamos=[],
        pdf_extractos=[],
        ruta_mayor=None,
        saldos_iniciales=saldos,
        nombre_cliente="TEST TEORICO",
        modo="teorico",
    )
    elapsed_pipeline = time.perf_counter() - t0
    assert resultado.modo == "teorico"
    assert resultado.cuotas == []
    assert len(resultado.saldos_por_banco) == 2

    t1 = time.perf_counter()
    excel = generar_planilla_teorica_bancos(saldos, "TEST TEORICO")
    elapsed_excel = time.perf_counter() - t1
    assert elapsed_excel < 3.0, f"Planilla teórica tardó {elapsed_excel:.2f}s (esperado <3s)"

    excel_reporte = generar_reporte_auditoria_prestamos(resultado, modo="teorico")
    assert len(excel_reporte) > 1000

    SALIDA.parent.mkdir(exist_ok=True)
    SALIDA.write_bytes(excel)

    wb = openpyxl.load_workbook(SALIDA)
    hoja_galicia = _sanitizar_nombre_hoja_excel("Banco Galicia")
    hoja_santander = _sanitizar_nombre_hoja_excel("Banco Santander")
    assert hoja_galicia in wb.sheetnames, wb.sheetnames
    assert hoja_santander in wb.sheetnames, wb.sheetnames
    assert "Consolidado" in wb.sheetnames

    ws = wb[hoja_galicia]
    headers = [ws.cell(5, c).value for c in range(1, 6)]
    assert headers == [
        "Fecha",
        "Importe Cuota (Capital)",
        "Impuestos",
        "Intereses",
        "Total a Debitar",
    ], headers
    textos = " ".join(str(ws.cell(r, c).value or "") for r in range(1, ws.max_row + 1) for c in range(1, 8))
    assert "Pendiente" in textos
    assert "500" in textos.replace(",", "").replace(".", "")
    wb.close()
    print(f"OK modo teórico sin PDF — pipeline {elapsed_pipeline:.3f}s, excel {elapsed_excel:.3f}s ->", SALIDA.resolve())


if __name__ == "__main__":
    test_detectar_modo()
    test_pipeline_teorico_sin_pdf()
    print("TODAS LAS PRUEBAS TEÓRICAS PASARON")
