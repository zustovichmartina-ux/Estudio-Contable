"""
Formato estándar de Excel del Estudio Contable (estilo Claude / informe prolijo).

Usar para todo Excel que se genere desde el chat o desde la web, salvo
plantillas oficiales (conciliación, asientos Tango, cuadros bancarios con
layout fijo, auditoría de préstamos en formato demo aprobado).
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Identidad visual ──────────────────────────────────────────────────────────
COLOR_PRIMARIO = "1F4E79"
COLOR_TEXTO_SUAVE = "666666"
COLOR_ZEBRA = "F2F2F2"
COLOR_HEADER_FG = "FFFFFF"

MONEY_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
MONEY_FMT_SIGNED = '#,##0.00;[Red]-#,##0.00'
DATE_FMT = "DD/MM/YYYY"
INT_FMT = "#,##0"

HDR_FILL = PatternFill("solid", fgColor=COLOR_PRIMARIO)
HDR_FONT = Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=COLOR_PRIMARIO)
SUB_FONT = Font(name="Calibri", size=11, color=COLOR_TEXTO_SUAVE)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=COLOR_PRIMARIO)
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
ZEBRA = PatternFill("solid", fgColor=COLOR_ZEBRA)


def _es_fecha_col(nombre: str) -> bool:
    n = str(nombre).lower()
    return n in {"fecha", "date", "f. pago", "f_pago"} or n.startswith("fecha")


def _es_monto_col(nombre: str) -> bool:
    n = str(nombre).lower()
    keys = (
        "débito", "debito", "crédito", "credito", "importe", "monto", "total",
        "saldo", "neto", "haber", "debe", "retencion", "retención", "percepcion",
        "percepción", "iva", "iibb", "capital", "interes", "interés",
    )
    return any(k in n for k in keys)


def _auto_ancho(ws: Worksheet, col_idx: int, series: Iterable[Any], header: str, minimo: int = 10, maximo: int = 42) -> None:
    largo = len(str(header))
    for v in series:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        largo = max(largo, min(len(str(v)), maximo))
    ws.column_dimensions[get_column_letter(col_idx)].width = max(minimo, min(largo + 2, maximo))


def _escribir_encabezado_hoja(
    ws: Worksheet,
    titulo: str,
    subtitulo: str = "",
    periodo: str = "",
) -> int:
    """Escribe título/subtítulo/período. Devuelve la próxima fila libre (1-based)."""
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    fila = 2
    if subtitulo:
        ws[f"A{fila}"] = subtitulo
        ws[f"A{fila}"].font = SUB_FONT
        fila += 1
    if periodo:
        ws[f"A{fila}"] = periodo
        ws[f"A{fila}"].font = SUB_FONT
        fila += 1
    return fila + 1  # deja una fila en blanco


def _pintar_header_fila(ws: Worksheet, fila: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(fila, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _escribir_tabla(
    ws: Worksheet,
    df: pd.DataFrame,
    start_row: int,
    *,
    col_moneda: Sequence[str] | None = None,
    col_fecha: Sequence[str] | None = None,
    zebra: bool = True,
    total_col: str | None = None,
    total_label_col: int | None = None,
) -> int:
    """Escribe df con header formateado. Devuelve la fila siguiente libre."""
    if df is None:
        df = pd.DataFrame()
    cols = list(df.columns)
    n = len(cols)
    if n == 0:
        return start_row

    monedas = {c.lower() for c in (col_moneda or [])}
    fechas = {c.lower() for c in (col_fecha or [])}

    for c, h in enumerate(cols, 1):
        ws.cell(start_row, c, h)
    _pintar_header_fila(ws, start_row, n)

    for i, (_, row) in enumerate(df.iterrows()):
        rr = start_row + 1 + i
        for c, h in enumerate(cols, 1):
            val = row[h]
            cell = ws.cell(rr, c)
            cell.font = BODY_FONT
            if pd.isna(val):
                cell.value = None
                continue
            hl = str(h).lower()
            if hl in fechas or _es_fecha_col(h):
                if hasattr(val, "strftime"):
                    cell.value = val
                else:
                    try:
                        cell.value = pd.to_datetime(val, dayfirst=True).date()
                    except Exception:
                        cell.value = val
                if hasattr(cell.value, "year"):
                    cell.number_format = DATE_FMT
            elif hl in monedas or _es_monto_col(h):
                try:
                    cell.value = float(val)
                    cell.number_format = MONEY_FMT
                    cell.alignment = Alignment(horizontal="right")
                except (TypeError, ValueError):
                    cell.value = val
            else:
                cell.value = val if not isinstance(val, float) or not pd.isna(val) else None
            if zebra and i % 2 == 1:
                cell.fill = ZEBRA

    last_data = start_row + len(df)
    next_row = last_data + 1

    if total_col and total_col in cols and len(df) > 0:
        c_tot = cols.index(total_col) + 1
        c_lbl = total_label_col or max(1, c_tot - 1)
        ws.cell(next_row, c_lbl, "TOTAL").font = BOLD_FONT
        tot_cell = ws.cell(next_row, c_tot, float(pd.to_numeric(df[total_col], errors="coerce").fillna(0).sum()))
        tot_cell.number_format = MONEY_FMT
        tot_cell.font = BOLD_FONT
        next_row += 1

    for c, h in enumerate(cols, 1):
        _auto_ancho(ws, c, df[h].tolist() if len(df) else [], str(h))

    # filtro + freeze sobre el header de la tabla
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(n)}{last_data}"
    if ws.freeze_panes is None:
        ws.freeze_panes = f"A{start_row + 1}"

    return next_row


def construir_informe_excel(
    *,
    titulo: str,
    subtitulo: str = "",
    periodo: str = "",
    kpis: Sequence[tuple[str, Any]] | None = None,
    resumenes: Sequence[tuple[str, pd.DataFrame]] | None = None,
    detalle: pd.DataFrame | None = None,
    hoja_detalle: str = "Movimientos",
    hojas_adicionales: Sequence[tuple[str, pd.DataFrame]] | None = None,
    col_moneda: Sequence[str] | None = None,
    col_fecha: Sequence[str] | None = None,
    total_col: str | None = None,
) -> Workbook:
    """
    Arma un workbook:
      - Hoja Resumen (título, KPIs, tablas de resumen)
      - Hoja de detalle (opcional)
      - Hojas adicionales (opcional): [(nombre_hoja, df), ...]
    """
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumen"

    fila = _escribir_encabezado_hoja(ws0, titulo, subtitulo, periodo)

    if kpis:
        for item in kpis:
            if len(item) == 3:
                label, value, kind = item[0], item[1], str(item[2]).lower()
            else:
                label, value = item[0], item[1]
                kind = "money" if isinstance(value, float) else "int" if isinstance(value, int) else "text"
            ws0.cell(fila, 1, label).font = BOLD_FONT
            cell = ws0.cell(fila, 2, value)
            cell.font = BODY_FONT
            if kind in {"money", "moneda", "$"}:
                try:
                    cell.value = float(value)
                except (TypeError, ValueError):
                    pass
                cell.number_format = MONEY_FMT
            elif kind in {"int", "cantidad", "n"}:
                cell.number_format = INT_FMT
            fila += 1
        fila += 1

    for titulo_tabla, df_res in resumenes or []:
        if df_res is None or df_res.empty:
            continue
        ws0.cell(fila, 1, titulo_tabla).font = SECTION_FONT
        fila += 1
        fila = _escribir_tabla(
            ws0,
            df_res,
            fila,
            col_moneda=col_moneda,
            col_fecha=col_fecha,
            zebra=False,
        )
        # quitar autofilter de tablas chicas de resumen (queda ruidoso)
        ws0.auto_filter.ref = None
        fila += 1

    ws0.column_dimensions["A"].width = max(ws0.column_dimensions["A"].width or 14, 34)
    ws0.column_dimensions["B"].width = max(ws0.column_dimensions["B"].width or 12, 14)
    ws0.column_dimensions["C"].width = max(ws0.column_dimensions["C"].width or 12, 18)

    if detalle is not None:
        ws = wb.create_sheet(hoja_detalle, 1)
        _escribir_tabla(
            ws,
            detalle,
            1,
            col_moneda=col_moneda,
            col_fecha=col_fecha,
            zebra=True,
            total_col=total_col,
        )

    for nombre_hoja, df_extra in hojas_adicionales or []:
        if df_extra is None:
            continue
        nombre = str(nombre_hoja or "Detalle").strip() or "Detalle"
        # Excel limita a 31 chars y prohíbe algunos caracteres
        for ch in (":", "\\", "/", "?", "*", "[", "]"):
            nombre = nombre.replace(ch, "-")
        nombre = nombre[:31]
        ws_extra = wb.create_sheet(nombre)
        _escribir_tabla(
            ws_extra,
            df_extra if not df_extra.empty else pd.DataFrame(),
            1,
            col_moneda=col_moneda,
            col_fecha=col_fecha,
            zebra=True,
            total_col=total_col if (not df_extra.empty and total_col in df_extra.columns) else None,
        )

    return wb


def informe_a_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_informe_excel(
    *,
    titulo: str,
    subtitulo: str = "",
    periodo: str = "",
    kpis: Sequence[tuple[str, Any]] | None = None,
    resumenes: Sequence[tuple[str, pd.DataFrame]] | None = None,
    detalle: pd.DataFrame | None = None,
    hoja_detalle: str = "Movimientos",
    hojas_adicionales: Sequence[tuple[str, pd.DataFrame]] | None = None,
    col_moneda: Sequence[str] | None = None,
    col_fecha: Sequence[str] | None = None,
    total_col: str | None = None,
) -> bytes:
    """Genera bytes .xlsx listos para st.download_button."""
    wb = construir_informe_excel(
        titulo=titulo,
        subtitulo=subtitulo,
        periodo=periodo,
        kpis=kpis,
        resumenes=resumenes,
        detalle=detalle,
        hoja_detalle=hoja_detalle,
        hojas_adicionales=hojas_adicionales,
        col_moneda=col_moneda,
        col_fecha=col_fecha,
        total_col=total_col,
    )
    return informe_a_bytes(wb)


def guardar_informe_excel(
    ruta: str | Path,
    *,
    titulo: str,
    subtitulo: str = "",
    periodo: str = "",
    kpis: Sequence[tuple[str, Any]] | None = None,
    resumenes: Sequence[tuple[str, pd.DataFrame]] | None = None,
    detalle: pd.DataFrame | None = None,
    hoja_detalle: str = "Movimientos",
    hojas_adicionales: Sequence[tuple[str, pd.DataFrame]] | None = None,
    col_moneda: Sequence[str] | None = None,
    col_fecha: Sequence[str] | None = None,
    total_col: str | None = None,
) -> Path:
    """Guarda el informe en disco (chat / scripts)."""
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = construir_informe_excel(
        titulo=titulo,
        subtitulo=subtitulo,
        periodo=periodo,
        kpis=kpis,
        resumenes=resumenes,
        detalle=detalle,
        hoja_detalle=hoja_detalle,
        hojas_adicionales=hojas_adicionales,
        col_moneda=col_moneda,
        col_fecha=col_fecha,
        total_col=total_col,
    )
    wb.save(ruta)
    return ruta


def aplicar_estilo_header_rango(ws: Worksheet, fila: int = 1, n_cols: int | None = None) -> None:
    """Aplica el header azul del Estudio a una fila ya escrita (para exports legacy)."""
    n = n_cols or ws.max_column
    _pintar_header_fila(ws, fila, n)
