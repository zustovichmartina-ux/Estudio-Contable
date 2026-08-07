"""Pruebas del completador de cuadros bancarios."""

from __future__ import annotations

import io
import unittest
from datetime import datetime
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

import completar_cuadro_bancario as modulo


def _excel_base() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "GALICIA"
    ws["B1"] = "GALICIA (plantilla)"
    ws["B2"] = "MOVIMIENTO"
    for idx, mes in enumerate(modulo.MESES, start=3):
        ws.cell(2, idx, mes)
    ws["B3"] = "TOTAL CREDITOS"
    ws["B4"] = "INTERESES GANADOS"
    ws["B6"] = "TRANSFERENCIAS RECIBIDAS"
    ws["B14"] = "OTROS CREDITOS"
    ws["B16"] = "TRANSFERENCIAS ENVIADAS"
    ws["B20"] = "PAGO DE SERVICIOS"
    ws["B30"] = "COMISIONES BANCARIAS"
    ws["B37"] = "OTROS DEBITOS"
    ws["B39"] = "TOTAL DEBITOS"
    ws["B41"] = "SALDO INICIAL"
    ws["B42"] = "CREDITO"
    ws["B43"] = "DEBITOS"
    ws["B44"] = "SALDO FINAL"
    ws["B45"] = "SALDO S/EXTRC"
    ws["B46"] = "DIF"
    # FEB preexistente: no debe borrarse al subir solo ENE.
    ws["D6"] = 777
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _grupo_ok() -> dict:
    df = pd.DataFrame(
        [
            {
                "Fecha": "02/01/2025",
                "Descripcion": "Transferencia recibida",
                "Detalle": "",
                "Importe": 100.0,
                "Saldo": 1100.0,
                "Clasificacion": "Transferencias recibidas",
                "Nueva_Clasificacion": None,
                "_fecha": pd.Timestamp("2025-01-02").date(),
                "_importe": 100.0,
                "_saldo": 1100.0,
                "_orden": 0,
            },
            {
                "Fecha": "03/01/2025",
                "Descripcion": "Transferencia emitida",
                "Detalle": "",
                "Importe": -40.0,
                "Saldo": 1060.0,
                "Clasificacion": "Transferencias emitidas",
                "Nueva_Clasificacion": None,
                "_fecha": pd.Timestamp("2025-01-03").date(),
                "_importe": -40.0,
                "_saldo": 1060.0,
                "_orden": 1,
            },
        ]
    )
    return {
        "banco_slug": "galicia",
        "banco": "Banco Galicia",
        "cuenta": "1234567/8",
        "cbu": "",
        "cliente": "Cliente",
        "archivos": ["enero.pdf"],
        "df": df,
    }


class CompletarCuadroBancarioTests(unittest.TestCase):
    def test_deriva_cuenta_provincia_desde_cbu(self):
        cuenta = modulo._cuenta_visible(
            {
                "cuenta": "",
                "cbu": "0140323503420057828560",
                "banco_slug": "provincia",
            },
            "extracto.pdf",
        )
        self.assertEqual(cuenta, "578285/6")

    def test_sin_numero_cuenta_va_a_caja(self):
        cuenta = modulo._cuenta_visible(
            {"cuenta": "", "cbu": "", "banco_slug": "macro", "banco": "Banco Macro"},
            "extracto_sin_cuenta.pdf",
        )
        self.assertEqual(cuenta, "Caja")

    def test_explorar_buzon_encuentra_excel_y_pdf(self):
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "Banco Galicia").mkdir()
            excel = root / "Banco Galicia" / "Ganancia 2025.xlsx"
            excel.write_bytes(b"PK\x03\x04fake")
            pdf = root / "Banco Galicia" / "extracto.pdf"
            pdf.write_bytes(b"%PDF-1.4")
            (root / "TENENCIAS 2025.pdf").write_bytes(b"%PDF")
            (root / "control.xlsx").write_bytes(b"PK")
            out = modulo.explorar_buzon_cuadros_bancarios(root, ejercicio=2025)
            self.assertTrue(out["excels"])
            self.assertTrue(out["excel_sugerido"].endswith("Ganancia 2025.xlsx"))
            self.assertEqual(len(out["pdfs"]), 1)
            self.assertTrue(out["pdfs"][0]["nombre"].endswith("extracto.pdf"))

    def test_completa_copia_controla_y_preserva_otro_mes(self):
        with patch.object(modulo, "procesar_pdfs_por_cuenta", return_value=([_grupo_ok()], [])):
            resultado = modulo.completar_cuadro_bancario_existente(
                _excel_base(),
                "Ganancias.xlsx",
                [("enero.pdf", b"pdf")],
            )

        self.assertTrue(resultado["excel"])
        self.assertEqual(resultado["errores"], [])
        self.assertEqual(len(resultado["resultados"]), 1)
        wb = load_workbook(io.BytesIO(resultado["excel"]), data_only=False)
        self.assertIn("_CONTROL_BANCOS", wb.sheetnames)
        hoja = resultado["resultados"][0]["hoja"]
        self.assertNotEqual(hoja, "GALICIA")
        ws = wb[hoja]
        self.assertEqual(ws["C6"].value, 100)
        self.assertEqual(ws["C16"].value, 40)
        self.assertEqual(ws["C41"].value, 1000)
        self.assertEqual(ws["C42"].value, "=C3")
        self.assertEqual(ws["C43"].value, "=C39")
        self.assertEqual(ws["C45"].value, 1060)
        self.assertEqual(ws["C46"].value, "=C44-C45")
        self.assertEqual(ws["D6"].value, 777)
        self.assertEqual(resultado["resultados"][0]["controles"][0]["Estado"], "OK")

    def test_bloquea_mes_sin_saldos(self):
        grupo = _grupo_ok()
        grupo["df"]["_saldo"] = None
        grupo["df"]["Saldo"] = None
        with patch.object(modulo, "procesar_pdfs_por_cuenta", return_value=([grupo], [])):
            resultado = modulo.completar_cuadro_bancario_existente(
                _excel_base(),
                "Ganancias.xlsx",
                [("enero.pdf", b"pdf")],
            )
        # Sin saldos no se puede armar el cuadro; queda el Excel de respaldo.
        self.assertEqual(resultado.get("modo"), "movimientos")
        self.assertTrue(resultado["excel"])
        self.assertTrue(resultado["errores"])
        wb = load_workbook(io.BytesIO(resultado["excel"]), data_only=False)
        self.assertIn("ERRORES", wb.sheetnames)
        self.assertIn("CONTROL", wb.sheetnames)

    def test_cadena_saldos_corrige_signo_ocr(self):
        """Si el OCR pone mal el signo, la cadena de saldos lo corrige."""
        df = pd.DataFrame(
            [
                {
                    "Fecha": "01/01/2025",
                    "Descripcion": "SALDO ANTERIOR",
                    "Detalle": "",
                    "Importe": 0.0,
                    "Saldo": 1000.0,
                    "Clasificacion": "",
                    "_fecha": pd.Timestamp("2025-01-01").date(),
                    "_importe": 0.0,
                    "_saldo": 1000.0,
                    "_orden": 0,
                },
                {
                    "Fecha": "02/01/2025",
                    "Descripcion": "Transferencia recibida",
                    "Detalle": "",
                    "Importe": -100.0,  # OCR al revés
                    "Saldo": 1100.0,
                    "Clasificacion": "Transferencias recibidas",
                    "_fecha": pd.Timestamp("2025-01-02").date(),
                    "_importe": -100.0,
                    "_saldo": 1100.0,
                    "_orden": 1,
                },
                {
                    "Fecha": "03/01/2025",
                    "Descripcion": "Pago",
                    "Detalle": "",
                    "Importe": 40.0,  # OCR al revés
                    "Saldo": 1060.0,
                    "Clasificacion": "Transferencias emitidas",
                    "_fecha": pd.Timestamp("2025-01-03").date(),
                    "_importe": 40.0,
                    "_saldo": 1060.0,
                    "_orden": 2,
                },
            ]
        )
        out = modulo._reparar_importes_cadena_saldos(df)
        movs = out[~out["Descripcion"].map(modulo._es_fila_saldo_ancla)].copy()
        self.assertEqual(float(movs.iloc[0]["_importe"]), 100.0)
        self.assertEqual(float(movs.iloc[1]["_importe"]), -40.0)
        movs = movs.assign(_orden=range(len(movs))).reset_index(drop=True)
        controles = modulo._controles_mensuales(movs)
        self.assertEqual(controles[0]["Estado"], "OK")
        self.assertEqual(controles[0]["DIF"], 0.0)

    def test_escribe_mes_aunque_dif_no_sea_cero(self):
        grupo = _grupo_ok()
        grupo["df"].loc[grupo["df"].index[-1], "_saldo"] = 9999.0
        grupo["df"].loc[grupo["df"].index[-1], "Saldo"] = 9999.0
        with patch.object(modulo, "procesar_pdfs_por_cuenta", return_value=([grupo], [])):
            resultado = modulo.completar_cuadro_bancario_existente(
                _excel_base(),
                "Ganancias.xlsx",
                [("enero.pdf", b"pdf")],
            )
        self.assertEqual(resultado.get("modo"), "cuadros")
        self.assertTrue(resultado["excel"])
        self.assertTrue(resultado["errores"])
        self.assertEqual(len(resultado["resultados"]), 1)
        ctrl = resultado["resultados"][0]["controles"][0]
        self.assertEqual(ctrl["Estado"], "ERROR")

    def test_inyecta_plantilla_estandar_si_cliente_no_la_tiene(self):
        """Si el Excel no tiene cuadro, se crea la plantilla estándar y se completa."""
        wb = Workbook()
        ws = wb.active
        ws.title = "M. Pago"
        ws["B2"] = "Mercado Pago"
        for idx, mes in enumerate(range(1, 13), start=3):
            ws.cell(2, idx, datetime(2025, mes, 1))
        ws["B5"] = "Transferencias recibidas"
        ws["B26"] = "Saldo al inicio"
        buf = io.BytesIO()
        wb.save(buf)

        grupo = _grupo_ok()
        grupo["banco_slug"] = "mercadopago"
        grupo["banco"] = "Mercado Pago"
        grupo["cuenta"] = "MP-001"
        with patch.object(modulo, "procesar_pdfs_por_cuenta", return_value=([grupo], [])):
            resultado = modulo.completar_cuadro_bancario_existente(
                buf.getvalue(),
                "GANANCIAS 2025.xlsx",
                [("enero.pdf", b"pdf")],
            )
        self.assertEqual(resultado.get("modo"), "cuadros")
        self.assertTrue(resultado["excel"])
        out = load_workbook(io.BytesIO(resultado["excel"]), data_only=False)
        # Libro limpio: no debe traer la planilla de Ganancias del cliente.
        self.assertNotIn("M. Pago", out.sheetnames)
        self.assertNotIn("2025", out.sheetnames)
        self.assertTrue(any("Mercado" in s or "MP" in s for s in out.sheetnames))
        self.assertIn("_CONTROL_BANCOS", out.sheetnames)
        hoja_nombre = resultado["resultados"][0]["hoja"]
        hoja = out[hoja_nombre]
        self.assertEqual(_normalizar_label(hoja["B3"].value), "total creditos")
        self.assertEqual(_normalizar_label(hoja["B46"].value), "dif")
        self.assertEqual(hoja["C6"].value, 100)
        self.assertTrue(str(hoja["C44"].value or "").startswith("="))
        self.assertTrue(str(hoja["C46"].value or "").startswith("="))
        self.assertTrue(str(resultado.get("nombre") or "").startswith("CUADROS_BANCARIOS_"))


def _normalizar_label(value) -> str:
    return modulo._normalizar(value)


if __name__ == "__main__":
    unittest.main()
