"""Completa cuadros bancarios mensuales dentro de un Excel existente.

El módulo no abre Excel ni ejecuta macros. Trabaja sobre bytes con openpyxl,
conserva VBA en .xlsm y devuelve un archivo nuevo para descargar.
"""

from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from procesador import (
    _procesar_un_pdf_extracto,
    enriquecer_df_extracto_formato_banco,
)


MESES = ("ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC")
COL_MES = {mes: idx + 3 for idx, mes in enumerate(MESES)}
TOLERANCIA = 0.05

_ALIASES_BANCO = {
    "provincia": ("provincia", "bapro"),
    "galicia": ("galicia",),
    "santander": ("santander", "rio"),
    "frances": ("frances", "bbva"),
    "nacion": ("nacion",),
    "macro": ("macro",),
    "credicoop": ("credicoop",),
    "icbc": ("icbc",),
    "hsbc": ("hsbc",),
    "supervielle": ("supervielle",),
    "ciudad": ("ciudad",),
    "comafi": ("comafi",),
    "mercadopago": ("mercado pago", "mercadopago"),
    "patagonia": ("patagonia",),
}

_EXCEL_IGNORAR = (
    "control", "prueba", "trabajo", "plantilla", "tenencia", "tenencias",
    "temp", "tmp", "~$", "cuadros_completados", "cuadro_completado",
)
_PDF_IGNORAR = (
    "tenencia", "tenencias", "thumbs",
)


def _normalizar(value: Any) -> str:
    text = str(value or "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def _digitos(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _puntaje_excel_destino(ruta: Path, ejercicio: int | None = None) -> int:
    """Cuánto parece un Excel de Ganancias / cuadros bancarios."""
    nombre = _normalizar(ruta.name)
    padres = _normalizar(" ".join(p.name for p in ruta.parents[:4]))
    if any(token in nombre for token in _EXCEL_IGNORAR):
        return -1000
    if ruta.name.startswith("~$"):
        return -1000
    score = 0
    if "ganancia" in nombre:
        score += 40
    if "ganancia" in padres:
        score += 15
    if any(b in padres or b in nombre for b in ("banco", "galicia", "provincia", "bbva", "santander")):
        score += 10
    if ejercicio and str(ejercicio) in nombre:
        score += 20
    if ejercicio and str(ejercicio) in padres:
        score += 10
    # Preferir libros más “centrales” (menos profundidad).
    score -= max(0, len(ruta.parts) - 8)
    return score


def _es_pdf_util(ruta: Path) -> bool:
    nombre = _normalizar(ruta.name)
    if any(token in nombre for token in _PDF_IGNORAR):
        return False
    return ruta.suffix.lower() == ".pdf"


_PDF_NO_EXTRACTO = (
    "tarjeta", "visa", "mastercard", "amex", "compras", "ventas",
    "factura", "constancia", "recibo", "ddjj", "f731", "f931",
)


def _es_pdf_extracto_probable(rel: str) -> bool:
    """True si la ruta relativa parece un extracto bancario (no tarjeta/compras)."""
    texto = _normalizar(rel)
    return not any(token in texto for token in _PDF_NO_EXTRACTO)


def explorar_buzon_cuadros_bancarios(
    carpeta: str | Path,
    *,
    ejercicio: int | None = None,
    max_depth: int = 6,
) -> dict:
    """
    Recorre una carpeta UNC/local y lista Excel destino + PDFs de extractos.

    Devuelve:
      carpeta, excels[{ruta,nombre,relativa,score,sugerido}], pdfs[{ruta,nombre,relativa}],
      excel_sugerido, advertencias
    """
    root = Path(str(carpeta or "").strip().strip('"').strip("'"))
    if not root.exists():
        raise FileNotFoundError(f"No existe la carpeta: {root}")
    if not root.is_dir():
        raise NotADirectoryError(f"La ruta no es una carpeta: {root}")

    excels: list[dict] = []
    pdfs: list[dict] = []
    advertencias: list[str] = []
    root_depth = len(root.parts)

    for path in root.rglob("*"):
        try:
            if not path.is_file():
                continue
            if len(path.parts) - root_depth > max_depth:
                continue
            rel = str(path.relative_to(root))
            suf = path.suffix.lower()
            if suf in {".xlsx", ".xlsm"}:
                score = _puntaje_excel_destino(path, ejercicio)
                if score <= -1000:
                    continue
                excels.append(
                    {
                        "ruta": str(path),
                        "nombre": path.name,
                        "relativa": rel,
                        "score": score,
                        "size": path.stat().st_size,
                    }
                )
            elif suf == ".pdf" and _es_pdf_util(path):
                pdfs.append(
                    {
                        "ruta": str(path),
                        "nombre": path.name,
                        "relativa": rel,
                        "size": path.stat().st_size,
                        "sugerido": _es_pdf_extracto_probable(rel),
                    }
                )
        except OSError as exc:
            advertencias.append(f"No pude leer {path}: {exc}")

    excels.sort(key=lambda x: (-int(x["score"]), x["relativa"].lower()))
    pdfs.sort(key=lambda x: x["relativa"].lower())
    for item in excels:
        item["sugerido"] = False
    if excels:
        excels[0]["sugerido"] = True

    if not excels:
        advertencias.append("No encontré Excel .xlsx/.xlsm candidatos (Ganancias / cuadros).")
    if not pdfs:
        advertencias.append("No encontré PDF de extractos en la carpeta.")

    return {
        "carpeta": str(root),
        "excels": excels,
        "pdfs": pdfs,
        "excel_sugerido": excels[0]["ruta"] if excels else None,
        "advertencias": advertencias,
    }


def _fecha(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parece_fecha_como_cuenta(cuenta: str) -> bool:
    """Evita tomar '02/01' o '18-09' (fechas del extracto) como número de cuenta."""
    c = str(cuenta or "").strip()
    return bool(re.fullmatch(r"\d{1,2}[/-]\d{1,2}(/\d{2,4})?", c))


def _cuenta_desde_ruta(nombre: str) -> str:
    partes = Path(str(nombre).replace("\\", "/")).parts
    if len(partes) >= 2:
        carpeta = partes[-2].strip()
        if carpeta and _normalizar(carpeta) not in {"pdf", "extractos", "download", "tmp"}:
            return carpeta
    return ""


def _cuenta_visible(meta: dict, nombre: str) -> str:
    carpeta = _cuenta_desde_ruta(nombre)
    carpeta_n = _normalizar(carpeta)
    carpeta_es_banco = bool(
        carpeta
        and any(
            a in carpeta_n
            for a in (
                "bbva", "mercado", "galicia", "frances", "provincia",
                "santander", "macro", "nacion", "credicoop",
            )
        )
    )
    # Si el PDF está en carpeta de banco, esa es la identidad visible.
    if carpeta_es_banco:
        return carpeta

    cuenta = str(meta.get("cuenta") or "").strip()
    if cuenta and not _parece_fecha_como_cuenta(cuenta):
        # Evitar basura tipo 0000000125
        dig = _digitos(cuenta)
        if not (len(dig) >= 6 and set(dig) <= {"0"}):
            return cuenta
    cbu = _digitos(meta.get("cbu"))
    if len(cbu) == 22 and meta.get("banco_slug") == "provincia":
        return f"{cbu[-8:-2]}/{cbu[-2]}"
    if len(cbu) >= 10:
        return cbu[-10:]
    for cand in re.findall(r"\d{7,22}", nombre):
        return cand
    if carpeta and _digitos(carpeta) and len(_digitos(carpeta)) >= 6:
        return carpeta
    # Sin número de cuenta: no agrupar todo bajo el nombre del banco (ej. Macro).
    # Va a Caja para que el cuadro no mezcle cuentas distintas.
    return "Caja"


def _clave_cuenta(meta: dict, nombre: str) -> str:
    """Clave estable por banco+cuenta (no por nombre de cada PDF)."""
    banco = _normalizar(meta.get("banco_slug") or meta.get("banco") or "")
    carpeta = _cuenta_desde_ruta(nombre)
    if carpeta:
        cn = _normalizar(carpeta)
        if any(
            a in cn
            for a in (
                "bbva", "mercado", "galicia", "frances", "provincia",
                "santander", "macro", "nacion", "credicoop",
            )
        ):
            return f"{banco}|{cn}"
        dig_carp = _digitos(carpeta)
        if len(dig_carp) >= 6:
            return dig_carp[-12:]
    visible = _cuenta_visible(meta, nombre)
    raw = _digitos(visible)
    if len(raw) >= 6 and not _parece_fecha_como_cuenta(visible) and set(raw) != {"0"}:
        return raw[-12:]
    return f"{banco}|{_normalizar(visible) or _normalizar(carpeta) or 'cuenta'}"


def _es_fila_saldo_ancla(descripcion: Any) -> bool:
    d = _normalizar(descripcion)
    return bool(
        re.search(
            r"\bsaldo\s+(?:anterior|inicial|al\s+inicio|final|al\s+cierre|de\s+cierre)\b",
            d,
        )
    )


def _reparar_importes_cadena_saldos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Motor estilo Trujillo: el saldo manda.

    Si hay saldos consecutivos, el importe real es:
        importe = saldo_actual − saldo_anterior
    Corrige signos y montos mal leídos por OCR sin romper fechas/descripciones.
    """
    if df is None or df.empty:
        return df
    if "_saldo" not in df.columns or "_importe" not in df.columns:
        return df

    out = df.sort_values(["_fecha", "_orden"], kind="stable").reset_index(drop=True).copy()
    prev_saldo: float | None = None
    nuevos: list[float | None] = []
    reparados = 0

    for _, row in out.iterrows():
        saldo = row.get("_saldo")
        importe = row.get("_importe")
        if pd.isna(saldo):
            nuevos.append(float(importe) if pd.notna(importe) else None)
            continue

        saldo_f = round(float(saldo), 2)
        if prev_saldo is None:
            # Primera ancla: si hay importe OCR, se respeta; si no, queda 0.
            if pd.notna(importe):
                nuevos.append(round(float(importe), 2))
            else:
                nuevos.append(0.0)
            prev_saldo = saldo_f
            continue

        chain = round(saldo_f - float(prev_saldo), 2)
        if abs(chain) < 0.005:
            # Sin movimiento neto (saldo repetido / ruido).
            nuevos.append(0.0)
            prev_saldo = saldo_f
            continue

        if pd.isna(importe):
            nuevos.append(chain)
            reparados += 1
        else:
            ocr = round(float(importe), 2)
            # Si el OCR discrepa en signo o monto, gana la cadena.
            if abs(ocr - chain) > 0.05:
                nuevos.append(chain)
                reparados += 1
            else:
                nuevos.append(ocr)
        prev_saldo = saldo_f

    out["_importe"] = nuevos
    out["Importe"] = out["_importe"]
    out.attrs["reparados_cadena"] = reparados
    return out


def _filas_movimiento(filas: list[dict]) -> pd.DataFrame:
    if not filas:
        return pd.DataFrame()
    df = enriquecer_df_extracto_formato_banco(pd.DataFrame(filas))
    if df.empty:
        return df
    df = df.copy()
    df["_fecha"] = df["Fecha"].map(_fecha)
    df["_importe"] = pd.to_numeric(df["Importe"], errors="coerce")
    df["_saldo"] = pd.to_numeric(df["Saldo"], errors="coerce")
    df = df[df["_fecha"].notna()].copy()
    df["_orden"] = range(len(df))
    df = df.sort_values(["_fecha", "_orden"], kind="stable").reset_index(drop=True)

    # 1) Reparar con cadena de saldos (incluye filas "Saldo anterior" como ancla).
    df = _reparar_importes_cadena_saldos(df)

    # 2) Sacar anclas de saldo y filas sin importe útil.
    if "Descripcion" in df.columns:
        df = df[~df["Descripcion"].map(_es_fila_saldo_ancla)].copy()
    df = df[df["_importe"].notna()].copy()
    df = df[df["_importe"].abs() >= 0.005].copy()

    # Extractos trimestrales/anuales pueden solaparse.
    subset = [c for c in ("Fecha", "Descripcion", "Detalle", "Importe", "Saldo") if c in df.columns]
    if subset:
        df = df.drop_duplicates(subset=subset, keep="last")
    return df.sort_values(["_fecha", "_orden"], kind="stable").reset_index(drop=True)


def _reforzar_continuidad_mensual(controles: list[dict]) -> list[dict]:
    """Si el mes N cierra en X, el N+1 debe abrir en X (cadena Trujillo)."""
    if len(controles) < 2:
        return controles
    out = [dict(c) for c in controles]
    for i in range(1, len(out)):
        ant, act = out[i - 1], out[i]
        if ant.get("Año") != act.get("Año"):
            continue
        if int(ant.get("Mes_num") or 0) + 1 != int(act.get("Mes_num") or 0):
            continue
        cierre = ant.get("Saldo extracto")
        apertura = act.get("Saldo inicial")
        if cierre is None or apertura is None:
            continue
        gap = round(float(apertura) - float(cierre), 2)
        act["Continuidad"] = gap
        if abs(gap) > TOLERANCIA:
            # Corregir apertura al cierre anterior y recalcular DIF.
            act["Saldo inicial"] = round(float(cierre), 2)
            cred = float(act.get("Créditos") or 0)
            deb = float(act.get("Débitos") or 0)
            calc = round(float(cierre) + cred - deb, 2)
            act["Saldo calculado"] = calc
            extracto = act.get("Saldo extracto")
            if extracto is not None:
                dif = round(calc - float(extracto), 2)
                act["DIF"] = dif
                act["Estado"] = "OK" if abs(dif) <= TOLERANCIA else "ERROR"
                act["Detalle"] = (
                    ""
                    if abs(dif) <= TOLERANCIA
                    else "El mes no reconcilia (tras encadenar saldo anterior)."
                )
            act["Continuidad"] = 0.0
            act["Detalle_cadena"] = f"Apertura corregida por cadena (gap era {gap:,.2f})."
    return out


def procesar_pdfs_por_cuenta(
    archivos: list[tuple[str, Any]],
) -> tuple[list[dict], list[dict]]:
    """Procesa los PDF de a uno (estable) y agrupa por banco + cuenta.

    Cada ítem es (nombre, bytes|Path). Si viene Path, se lee recién al procesarlo
    para no cargar docenas de PDF juntos en memoria.
    """
    grupos: dict[tuple[str, str], dict] = {}
    errores: list[dict] = []

    def _bytes_pdf(nombre: str, data: Any) -> bytes:
        if isinstance(data, (bytes, bytearray)):
            return bytes(data)
        return Path(data).read_bytes()

    # Secuencial a propósito: EasyOCR + muchos PDF en paralelo tumba el proceso.
    for nombre, data in archivos:
        try:
            raw = _bytes_pdf(nombre, data)
            filas, meta, error = _procesar_un_pdf_extracto(nombre, raw)
        except Exception as exc:
            errores.append(
                {
                    "archivo": nombre,
                    "banco": "",
                    "motivo": f"Error al leer el PDF: {exc}",
                }
            )
            continue
        if error:
            errores.append(error)
            continue
        banco_slug = str(meta.get("banco_slug") or "desconocido")
        cuenta = _cuenta_visible(meta, nombre)
        clave = (banco_slug, _clave_cuenta(meta, nombre))
        grupo = grupos.setdefault(
            clave,
            {
                "banco_slug": banco_slug,
                "banco": meta.get("banco") or banco_slug.title(),
                "cuenta": cuenta,
                "cbu": meta.get("cbu") or "",
                "cliente": meta.get("cliente") or "",
                "archivos": [],
                "filas": [],
            },
        )
        grupo["archivos"].append(nombre)
        grupo["filas"].extend(filas)
        for campo in ("cuenta", "cbu", "cliente"):
            if not grupo.get(campo) and meta.get(campo):
                grupo[campo] = meta[campo]

    salida = []
    for grupo in grupos.values():
        grupo["df"] = _filas_movimiento(grupo.pop("filas"))
        if grupo["df"].empty:
            errores.append(
                {
                    "archivo": " | ".join(grupo["archivos"]),
                    "banco": grupo["banco"],
                    "motivo": "No quedaron movimientos válidos para completar el cuadro.",
                }
            )
            continue
        salida.append(grupo)
    return salida, errores


def _tiene_plantilla_cuadro(wb) -> bool:
    for ws in wb.worksheets:
        try:
            _layout(ws)
            return True
        except Exception:
            continue
    return False


def _es_hoja_ganancias_categorias(ws: Worksheet) -> bool:
    """Planilla tipo Deprez: meses en fila 2 + categorías en col B + saldos abajo."""
    meses = 0
    for col in range(3, min(ws.max_column, 20) + 1):
        value = ws.cell(2, col).value
        if isinstance(value, datetime):
            meses += 1
        elif isinstance(value, date):
            meses += 1
    if meses < 3:
        return False
    labels = " ".join(
        _normalizar(ws.cell(row, 2).value)
        for row in range(1, min(ws.max_row, 40) + 1)
        if ws.cell(row, 2).value
    )
    return ("saldo al inicio" in labels or "saldo inicial" in labels) and (
        "transferencias" in labels or "depositos" in labels or "interes" in labels
    )


def _hojas_ganancias_categorias(wb) -> list[Worksheet]:
    return [ws for ws in wb.worksheets if _es_hoja_ganancias_categorias(ws)]


def _detectar_modo_libro(wb) -> str:
    if _tiene_plantilla_cuadro(wb):
        return "cuadro"
    if _hojas_ganancias_categorias(wb):
        return "ganancias"
    return "desconocido"


def _columnas_mes_ganancias(ws: Worksheet) -> dict[int, int]:
    """Mes (1-12) -> columna Excel."""
    out: dict[int, int] = {}
    for col in range(3, min(ws.max_column, 20) + 1):
        value = ws.cell(2, col).value
        if isinstance(value, datetime):
            out[int(value.month)] = col
        elif isinstance(value, date):
            out[int(value.month)] = col
    return out


def _filas_concepto_ganancias(ws: Worksheet) -> list[tuple[int, str, str]]:
    """Filas de categoría escribibles (no encabezados ni totales con SUM)."""
    filas: list[tuple[int, str, str]] = []
    skip_exact = {
        "depositos", "ingresos", "retiros y debitos", "retiros y debetos",
        "saldo al inicio", "saldo inicial", "saldo final", "totales",
    }
    for row in range(3, min(ws.max_row, 45) + 1):
        raw = ws.cell(row, 2).value
        if raw is None or not str(raw).strip():
            continue
        label = str(raw).strip()
        norm = _normalizar(label)
        if not norm or norm in skip_exact:
            continue
        if norm.startswith("a banco") or norm.startswith("a tarjeta"):
            continue
        sample = ws.cell(row, 3).value
        if isinstance(sample, str) and sample.startswith("=") and "SUM" in sample.upper():
            continue
        filas.append((row, norm, label))
    return filas


def _filas_saldo_ganancias(ws: Worksheet) -> dict[str, int]:
    found: dict[str, int] = {}
    for row in range(1, min(ws.max_row, 50) + 1):
        norm = _normalizar(ws.cell(row, 2).value)
        if not norm:
            continue
        if "saldo al inicio" in norm or norm == "saldo inicial":
            found["saldo_inicial"] = row
        elif norm.startswith("ingresos"):
            found["ingresos"] = row
        elif "retiro" in norm:
            found["egresos"] = row
        elif "saldo final" in norm:
            found["saldo_final"] = row
    return found


def _mapear_fila_ganancias(
    clasificacion: str,
    descripcion: str,
    importe: float,
    filas: list[tuple[int, str, str]],
) -> int | None:
    """Elige la fila de categoría del cliente según lógica contable."""
    texto = _normalizar(f"{clasificacion} {descripcion}")
    credito = importe >= 0

    reglas_credito = [
        (("interes",), ("interes",)),
        (("reintegro", "cashback", "devolucion"), ("reintegro",)),
        (("entre cuentas", "e cuentas", "propia"), ("movimiento e", "cuentas propia", "e cuentas")),
        (("galicia",), ("galicia",)),
        (("transfer", "recib", "acredit", "deposito", "pago recibido"), ("transferencias recibidas", "recibid")),
    ]
    reglas_debito = [
        (("visa",), ("visa",)),
        (("master", "mastercard"), ("master", "mstercard")),
        (("cheque",), ("cheque",)),
        (("servicio", "luz", "gas", "agua", "telecom", "internet"), ("pago de servicios", "servicio")),
        (("comision", "gasto banc", "mantenimiento"), ("comision", "gtos", "gasto")),
        (("retencion iibb", "iibb", "sircreb", "arba"), ("retencion iibb", "iibb")),
        (("impuesto pais", "pais"), ("impuesto pais", "pais")),
        (("percepcion iva", "perc iva"), ("percepcion iva", "percep")),
        (("rg 4815", "rg 5617", "rg4815"), ("rg 4815", "rg")),
        (("iva",), ("iva 21", "iva")),
        (("compra", "debito automatico", "pago comercio"), ("compra",)),
        (("transfer", "envi", "emitid", "realiz", "debito"), ("transferencias realizadas", "realiz")),
    ]

    reglas = reglas_credito if credito else reglas_debito
    for claves, destinos in reglas:
        if any(c in texto for c in claves):
            for row, norm, _label in filas:
                if any(d in norm for d in destinos):
                    return row

    # Fallback por signo: primera categoría de crédito/débito típica.
    preferidos = (
        ("transferencias recibidas", "interes", "reintegro", "movimiento")
        if credito
        else ("transferencias realizadas", "compra", "comision", "servicio")
    )
    for pref in preferidos:
        for row, norm, _label in filas:
            if pref in norm:
                return row
    return filas[0][0] if filas else None


def _buscar_hoja_ganancias(wb, banco_slug: str, cuenta: str) -> Worksheet | None:
    aliases = list(_ALIASES_BANCO.get(banco_slug, (banco_slug,)))
    # Alias extra por slug detectado.
    extra = {
        "frances": ("frances", "bbva", "frances"),
        "bbva": ("frances", "bbva"),
        "mercadopago": ("m pago", "mercado pago", "mercadopago"),
        "galicia": ("galicia",),
        "santander": ("santander",),
        "provincia": ("provincia", "bapro"),
    }
    aliases.extend(extra.get(banco_slug, ()))
    cuenta_d = _digitos(cuenta)
    sufijo = cuenta_d[-6:] if len(cuenta_d) >= 4 else cuenta_d
    candidatas = _hojas_ganancias_categorias(wb)
    # 1) coincidencia por cuenta en título/encabezado
    if sufijo:
        for ws in candidatas:
            text = _normalizar(f"{ws.title} {ws.cell(2, 2).value or ''}")
            digits = _digitos(f"{ws.title} {ws.cell(2, 2).value or ''}")
            if sufijo in digits and any(a in text for a in aliases):
                return ws
    # 2) solo banco
    for ws in candidatas:
        text = _normalizar(f"{ws.title} {ws.cell(2, 2).value or ''}")
        if any(a in text for a in aliases):
            return ws
    return None


def _escribir_grupo_ganancias(ws: Worksheet, grupo: dict, controles: list[dict]) -> None:
    """Completa una hoja de Ganancias por categorías (meses en fila 2)."""
    cols = _columnas_mes_ganancias(ws)
    if not cols:
        raise ValueError(f"La hoja '{ws.title}' no tiene meses en la fila 2.")
    filas = _filas_concepto_ganancias(ws)
    if not filas:
        raise ValueError(f"La hoja '{ws.title}' no tiene categorías escribibles.")
    saldos = _filas_saldo_ganancias(ws)
    df = grupo["df"]

    months = sorted({int(f.month) for f in df["_fecha"] if int(f.month) in cols})
    by_month = {c["Mes_num"]: c for c in controles if c.get("Mes_num") in cols}

    for month in months:
        col = cols[month]
        # Limpiar solo conceptos (no fórmulas de total).
        for row, _norm, _label in filas:
            ws.cell(row, col, None)

        bucket: dict[int, float] = defaultdict(float)
        mes_df = df[df["_fecha"].map(lambda f: f.month == month)]
        for _, row in mes_df.iterrows():
            importe = float(row.get("_importe") or 0)
            if abs(importe) < 0.005:
                continue
            clas = str(row.get("Nueva_Clasificacion") or row.get("Clasificacion") or "")
            desc = str(row.get("Descripcion") or row.get("Detalle") or "")
            dest = _mapear_fila_ganancias(clas, desc, importe, filas)
            if dest is None:
                continue
            # En la planilla los débitos se cargan en positivo (egresos).
            bucket[dest] += abs(importe) if importe < 0 else importe

        for row, total in bucket.items():
            ws.cell(row, col, round(total, 2))

        control = by_month.get(month) or {}
        if saldos.get("saldo_inicial") and control.get("Saldo inicial") is not None:
            ws.cell(saldos["saldo_inicial"], col, control["Saldo inicial"])
        if saldos.get("ingresos") and control.get("Créditos") is not None:
            ws.cell(saldos["ingresos"], col, control["Créditos"])
        if saldos.get("egresos") and control.get("Débitos") is not None:
            ws.cell(saldos["egresos"], col, control["Débitos"])
        if saldos.get("saldo_final") and control.get("Saldo extracto") is not None:
            ws.cell(saldos["saldo_final"], col, control["Saldo extracto"])


def _exportar_movimientos_detectados(grupos: list[dict], errores: list[dict]) -> bytes:
    """Excel de respaldo cuando el libro destino no admite el cuadro Trujillo."""
    from openpyxl import Workbook

    wb = Workbook()
    ws_err = wb.active
    ws_err.title = "ERRORES"
    ws_err.append(["Archivo", "Banco", "Cuenta", "Motivo"])
    for err in errores:
        ws_err.append(
            [
                err.get("archivo"),
                err.get("banco"),
                err.get("cuenta"),
                err.get("motivo"),
            ]
        )
    ws_mov = wb.create_sheet("MOVIMIENTOS")
    ws_mov.append(
        ["Banco", "Cuenta", "Archivo", "Fecha", "Descripcion", "Detalle", "Importe", "Saldo", "Clasificacion"]
    )
    for grupo in grupos:
        df = grupo.get("df")
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            ws_mov.append(
                [
                    grupo.get("banco"),
                    grupo.get("cuenta"),
                    " | ".join(grupo.get("archivos") or []),
                    row.get("Fecha"),
                    row.get("Descripcion"),
                    row.get("Detalle"),
                    row.get("Importe"),
                    row.get("Saldo"),
                    row.get("Clasificacion") or row.get("Nueva_Clasificacion"),
                ]
            )
    ws_ctrl = wb.create_sheet("CONTROL")
    ws_ctrl.append(
        ["Banco", "Cuenta", "Mes", "Año", "Saldo inicial", "Créditos", "Débitos",
         "Saldo calculado", "Saldo extracto", "DIF", "Estado", "Detalle"]
    )
    for grupo in grupos:
        for control in _controles_mensuales(grupo["df"]):
            ws_ctrl.append(
                [
                    grupo.get("banco"),
                    grupo.get("cuenta"),
                    control.get("Mes"),
                    control.get("Año"),
                    control.get("Saldo inicial"),
                    control.get("Créditos"),
                    control.get("Débitos"),
                    control.get("Saldo calculado"),
                    control.get("Saldo extracto"),
                    control.get("DIF"),
                    control.get("Estado"),
                    control.get("Detalle"),
                ]
            )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _buscar_fila(ws: Worksheet, patrones: tuple[str, ...], minimo: int = 1, maximo: int = 120) -> int | None:
    for row in range(minimo, min(maximo, ws.max_row) + 1):
        value = _normalizar(ws.cell(row, 2).value)
        if value and any(p in value for p in patrones):
            return row
    return None


def _layout(ws: Worksheet) -> dict:
    total_creditos = _buscar_fila(ws, ("total credit",))
    total_debitos = _buscar_fila(ws, ("total debit",))
    saldo_inicial = _buscar_fila(ws, ("saldo incial", "saldo inicial"))
    credito = _buscar_fila(ws, ("credito",), minimo=(saldo_inicial or 1))
    debitos = _buscar_fila(ws, ("debitos", "debito"), minimo=(credito or saldo_inicial or 1) + 1)
    saldo_final = _buscar_fila(ws, ("saldo final",), minimo=(debitos or 1))
    saldo_extracto = _buscar_fila(ws, ("saldo s extr", "saldo extract",), minimo=(saldo_final or 1))
    dif = _buscar_fila(ws, ("dif",), minimo=(saldo_extracto or 1))
    required = {
        "total_creditos": total_creditos,
        "total_debitos": total_debitos,
        "saldo_inicial": saldo_inicial,
        "credito": credito,
        "debitos": debitos,
        "saldo_final": saldo_final,
        "saldo_extracto": saldo_extracto,
        "dif": dif,
    }
    faltan = [k for k, v in required.items() if v is None]
    if faltan:
        raise ValueError(
            f"La hoja '{ws.title}' no tiene el cuadro bancario esperado. Faltan: {', '.join(faltan)}."
        )
    if not total_creditos < total_debitos < saldo_inicial:
        raise ValueError(f"La estructura de créditos/débitos de '{ws.title}' no es válida.")
    inicio_debitos = None
    # Orden importa: primero transferencias emitidas, después compras/pagos.
    marcas_debito = (
        "transferencias emitidas", "transferencia emitida", "transf emitida",
        "transferencias enviadas", "transf enviada", "debin",
        "compra", "pago mastercard", "pago visa", "pago de servicios",
        "comision", "impuesto", "retencion", "debito", "pago",
    )
    for row in range(total_creditos + 1, total_debitos):
        label = _normalizar(ws.cell(row, 2).value)
        if not label:
            continue
        if any(marca in label for marca in marcas_debito):
            inicio_debitos = row
            break
    if inicio_debitos is None:
        # Fallback para plantillas sin conceptos: usar el último separador real,
        # no un hueco intermedio entre dos conceptos de crédito.
        separadores = []
        for row in range(total_creditos + 1, total_debitos):
            actual = str(ws.cell(row, 2).value or "").strip()
            siguiente = str(ws.cell(row + 1, 2).value or "").strip()
            if not actual and siguiente:
                separadores.append(row + 1)
        if separadores:
            inicio_debitos = separadores[-1]
    if inicio_debitos is None or inicio_debitos <= total_creditos + 1:
        raise ValueError(f"No pude separar créditos y débitos en la hoja '{ws.title}'.")
    required["inicio_debitos"] = inicio_debitos
    return required


def _texto_hoja(ws: Worksheet) -> str:
    values = [ws.title]
    for row in range(1, min(ws.max_row, 12) + 1):
        values.extend(str(ws.cell(row, col).value or "") for col in range(1, 5))
    return _normalizar(" ".join(values))


def _buscar_plantilla(wb, banco_slug: str) -> Worksheet:
    aliases = _ALIASES_BANCO.get(banco_slug, (banco_slug,))
    candidatas = [ws for ws in wb.worksheets if any(a in _texto_hoja(ws) for a in aliases)]
    for ws in candidatas:
        try:
            _layout(ws)
            return ws
        except ValueError:
            continue
    # Una hoja llamada PROVINCIA/GALICIA/etc. sin texto de encabezado.
    for ws in wb.worksheets:
        if any(a in _normalizar(ws.title) for a in aliases):
            try:
                _layout(ws)
                return ws
            except ValueError:
                continue
    # Cualquier hoja ya en formato estándar del estudio.
    for ws in wb.worksheets:
        try:
            _layout(ws)
            return ws
        except ValueError:
            continue
    # Base única para todos los clientes.
    return _asegurar_plantilla_estandar(wb)


def _asegurar_plantilla_estandar(wb) -> Worksheet:
    """
    Plantilla única del estudio: meses | créditos | débitos | saldos | fórmulas.
    Si el libro del cliente no la tiene, se crea una vez y se copia por cuenta.
    """
    nombre = "_PLANTILLA_CUADRO"
    if nombre in wb.sheetnames:
        ws = wb[nombre]
        try:
            _layout(ws)
            return ws
        except ValueError:
            # Reconstruir si alguien la rompió.
            del wb[nombre]

    ws = wb.create_sheet(nombre)
    ws["B1"] = "PLANTILLA CUADRO BANCARIO (estándar estudio)"
    ws["B2"] = "MOVIMIENTO"
    for idx, mes in enumerate(MESES, start=3):
        ws.cell(2, idx, mes)
        col = ws.cell(2, idx).column_letter
        ws.cell(3, idx, f"=SUM({col}4:{col}14)")
        ws.cell(39, idx, f"=SUM({col}16:{col}37)")
        ws.cell(42, idx, f"={col}3")
        ws.cell(43, idx, f"={col}39")
        ws.cell(44, idx, f"={col}41+{col}42-{col}43")
        ws.cell(46, idx, f"={col}44-{col}45")

    ws["B3"] = "TOTAL CREDITOS"
    ws["B4"] = "INTERESES GANADOS"
    ws["B5"] = "REINTEGROS"
    ws["B6"] = "TRANSFERENCIAS RECIBIDAS"
    ws["B7"] = "MOV. ENTRE CUENTAS"
    ws["B8"] = "HABERES"
    ws["B14"] = "OTROS CREDITOS"

    ws["B16"] = "TRANSFERENCIAS EMITIDAS"
    ws["B17"] = "COMPRAS"
    ws["B18"] = "PAGO DE SERVICIOS"
    ws["B19"] = "PAGO MASTERCARD"
    ws["B20"] = "PAGO VISA"
    ws["B21"] = "DEBIN"
    ws["B22"] = "PLANES AFIP"
    ws["B30"] = "COMISIONES BANCARIAS"
    ws["B31"] = "IVA"
    ws["B32"] = "PERCEPCION IVA"
    ws["B33"] = "IMPUESTOS A LOS DEBITOS Y CREDITOS"
    ws["B34"] = "RETENCION ARBA / SIRCREB"
    ws["B35"] = "IMPUESTO PAIS"
    ws["B37"] = "OTROS DEBITOS"
    ws["B39"] = "TOTAL DEBITOS"

    ws["B41"] = "SALDO INCIAL"
    ws["B42"] = "CREDITO"
    ws["B43"] = "DEBITOS"
    ws["B44"] = "SALDO FINAL"
    ws["B45"] = "SALDO S/EXTRC"
    ws["B46"] = "DIF"

    try:
        ws.sheet_state = "hidden"
    except Exception:
        pass
    _layout(ws)
    return ws


def _obtener_hoja_destino_estandar(wb, grupo: dict) -> Worksheet:
    """Siempre trabaja sobre el formato estándar (copia plantilla por cuenta)."""
    ws = _buscar_hoja_cuenta(wb, grupo["banco_slug"], grupo["cuenta"])
    if ws is not None:
        try:
            _layout(ws)
            return ws
        except ValueError:
            pass
    # Reutiliza hoja estándar del cliente (GALICIA/PROVINCIA/etc.) o inyecta la del estudio.
    plantilla = _buscar_plantilla(wb, grupo["banco_slug"])
    nombre = _nombre_hoja(grupo["banco"], grupo["cuenta"], set(wb.sheetnames))
    return _copiar_hoja(wb, plantilla, nombre)


def _buscar_hoja_cuenta(wb, banco_slug: str, cuenta: str) -> Worksheet | None:
    cuenta_d = _digitos(cuenta)
    sufijo = cuenta_d[-6:]
    if not sufijo:
        return None
    aliases = _ALIASES_BANCO.get(banco_slug, (banco_slug,))
    for ws in wb.worksheets:
        text = _normalizar(f"{ws.title} {ws.cell(1, 2).value or ''}")
        digits = _digitos(f"{ws.title} {ws.cell(1, 2).value or ''}")
        if sufijo in digits and any(a in text for a in aliases):
            return ws
    return None


def _nombre_hoja(banco: str, cuenta: str, existentes: set[str]) -> str:
    banco_corto = re.sub(r"(?i)^banco\s+", "", banco).strip() or "BANCO"
    cuenta_corta = re.sub(r"\s+", "", cuenta).replace("/", "-")
    if _normalizar(cuenta) and (
        _normalizar(cuenta) in _normalizar(banco)
        or _normalizar(banco) in _normalizar(cuenta)
        or _normalizar(cuenta) in _normalizar(banco_corto)
    ):
        base = re.sub(r"[:\\/?*\[\]]", "-", banco_corto)[:31].strip()
    else:
        base = re.sub(r"[:\\/?*\[\]]", "-", f"{banco_corto} {cuenta_corta}")[:31].strip()
    nombre = base or "CUENTA BANCARIA"
    i = 2
    while nombre in existentes:
        suf = f" ({i})"
        nombre = f"{base[:31-len(suf)]}{suf}"
        i += 1
    return nombre


def _copiar_hoja(wb, plantilla: Worksheet, nombre: str) -> Worksheet:
    ws = wb.copy_worksheet(plantilla)
    ws.title = nombre
    try:
        ws.sheet_state = "visible"
    except Exception:
        pass
    return ws


def _rango_conceptos(layout: dict, credito: bool) -> range:
    if credito:
        return range(layout["total_creditos"] + 1, layout["inicio_debitos"])
    return range(layout["inicio_debitos"], layout["total_debitos"])


def _fila_concepto(ws: Worksheet, layout: dict, clasificacion: str, importe: float) -> int:
    credito = importe >= 0
    filas = list(_rango_conceptos(layout, credito))
    clas = _normalizar(clasificacion)
    reglas = [
        (("interes",), ("interes",)),
        (("transferencias recibidas", "pagos recibidos"), ("transf", "recibid")),
        (("transferencias emitidas",), ("transf", "envi")),
        (("pago arba",), ("arba",)),
        (("sircreb",), ("sircreb",)),
        (("gastos bancarios",), ("comision", "gasto banc")),
        (("impuestos a los debitos y creditos",), ("imp", "cred", "deb")),
        (("impuesto a los sellos",), ("sello",)),
        (("percepcion iva",), ("percep", "iva")),
        (("iva",), ("iva",)),
        (("iibb", "ingresos brutos"), ("iibb", "ing bruto")),
        (("pago de servicios",), ("servicio",)),
        (("pagos afip",), ("afip", "plan")),
        (("pago de haberes",), ("haber",)),
        (("compras",), ("compra",)),
        (("inversiones", "suscripcion fci"), ("inversion", "fci", "cocos")),
    ]
    keywords: tuple[str, ...] = ()
    for entradas, candidatos in reglas:
        if any(e in clas for e in entradas):
            keywords = candidatos
            break
    if keywords:
        for row in filas:
            label = _normalizar(ws.cell(row, 2).value)
            if any(k in label for k in keywords):
                return row

    # Fallback explícito para no perder importes.
    fallback = ("otros credit", "transf recibidas") if credito else ("otros debit",)
    for row in filas:
        label = _normalizar(ws.cell(row, 2).value)
        if any(k in label for k in fallback):
            return row

    # Usar la última fila libre del bloque y etiquetarla.
    for row in reversed(filas):
        if not str(ws.cell(row, 2).value or "").strip():
            ws.cell(row, 2, "OTROS CREDITOS" if credito else "OTROS DEBITOS")
            return row
    raise ValueError(
        f"No hay una fila disponible para '{clasificacion}' en la hoja '{ws.title}'."
    )


def _controles_mensuales(df: pd.DataFrame) -> list[dict]:
    controles = []
    for (year, month), group in df.groupby(
        [df["_fecha"].map(lambda f: f.year), df["_fecha"].map(lambda f: f.month)],
        sort=True,
    ):
        group = group.sort_values(["_fecha", "_orden"], kind="stable")
        saldos = group[group["_saldo"].notna()]
        if saldos.empty:
            controles.append(
                {
                    "Año": year,
                    "Mes": MESES[month - 1],
                    "Mes_num": month,
                    "Movimientos": len(group),
                    "Estado": "ERROR",
                    "Detalle": "El extracto no contiene saldos para controlar.",
                }
            )
            continue
        first = saldos.iloc[0]
        last = saldos.iloc[-1]
        apertura = round(float(first["_saldo"]) - float(first["_importe"]), 2)
        creditos = round(float(group.loc[group["_importe"] > 0, "_importe"].sum()), 2)
        debitos = round(float(-group.loc[group["_importe"] < 0, "_importe"].sum()), 2)
        cierre = round(float(last["_saldo"]), 2)
        calculado = round(apertura + creditos - debitos, 2)
        dif = round(calculado - cierre, 2)
        controles.append(
            {
                "Año": year,
                "Mes": MESES[month - 1],
                "Mes_num": month,
                "Movimientos": len(group),
                "Saldo inicial": apertura,
                "Créditos": creditos,
                "Débitos": debitos,
                "Saldo calculado": calculado,
                "Saldo extracto": cierre,
                "DIF": dif,
                "Estado": "OK" if abs(dif) <= TOLERANCIA else "ERROR",
                "Detalle": "" if abs(dif) <= TOLERANCIA else "El mes no reconcilia.",
            }
        )
    # Continuidad entre meses consecutivos presentes.
    for anterior, actual in zip(controles, controles[1:]):
        if (
            anterior.get("Año") == actual.get("Año")
            and anterior.get("Mes_num", 0) + 1 == actual.get("Mes_num")
            and anterior.get("Saldo extracto") is not None
            and actual.get("Saldo inicial") is not None
        ):
            gap = round(float(actual["Saldo inicial"]) - float(anterior["Saldo extracto"]), 2)
            actual["Continuidad"] = gap
            if abs(gap) > TOLERANCIA:
                actual["Estado"] = "ERROR"
                actual["Detalle"] = f"No continúa el mes anterior (diferencia {gap:,.2f})."
    return controles


def _escribir_grupo(ws: Worksheet, grupo: dict, controles: list[dict]) -> None:
    layout = _layout(ws)
    df = grupo["df"]
    years = sorted({f.year for f in df["_fecha"]})
    if len(years) != 1:
        raise ValueError(f"La cuenta {grupo['cuenta']} mezcla años: {years}.")
    ws.cell(1, 2, f"{str(grupo['banco']).upper()} - {grupo['cuenta']}")

    # Solo se reemplazan los meses subidos; el resto del Excel queda intacto.
    months = sorted({f.month for f in df["_fecha"]})
    for month in months:
        col = month + 2
        for row in _rango_conceptos(layout, True):
            ws.cell(row, col, None)
        for row in _rango_conceptos(layout, False):
            ws.cell(row, col, None)
        ws.cell(1, col, None)

    acumulado: dict[tuple[int, int], float] = defaultdict(float)
    for _, mov in df.iterrows():
        importe = float(mov["_importe"])
        month = mov["_fecha"].month
        clas = str(
            mov.get("Nueva_Clasificacion")
            or mov.get("Clasificacion")
            or ""
        )
        desc = str(mov.get("Descripcion") or mov.get("Detalle") or "")
        row = _fila_concepto(ws, layout, f"{clas} {desc}".strip() or "Sin clasificar", importe)
        acumulado[(row, month + 2)] += abs(importe)
    for (row, col), value in acumulado.items():
        ws.cell(row, col, round(value, 2))

    by_month = {c["Mes_num"]: c for c in controles}
    for month in months:
        control = by_month[month]
        col = month + 2
        letter = ws.cell(1, col).column_letter
        ws.cell(layout["saldo_inicial"], col, control["Saldo inicial"])
        ws.cell(layout["saldo_extracto"], col, control["Saldo extracto"])
        # Totales y puente de control SIEMPRE con fórmulas (formato estándar).
        ws.cell(
            layout["total_creditos"],
            col,
            f"=SUM({letter}{layout['total_creditos']+1}:{letter}{layout['inicio_debitos']-1})",
        )
        ws.cell(
            layout["total_debitos"],
            col,
            f"=SUM({letter}{layout['inicio_debitos']}:{letter}{layout['total_debitos']-1})",
        )
        ws.cell(layout["credito"], col, f"={letter}{layout['total_creditos']}")
        ws.cell(layout["debitos"], col, f"={letter}{layout['total_debitos']}")
        ws.cell(
            layout["saldo_final"],
            col,
            f"={letter}{layout['saldo_inicial']}+{letter}{layout['credito']}-{letter}{layout['debitos']}",
        )
        ws.cell(
            layout["dif"],
            col,
            f"={letter}{layout['saldo_final']}-{letter}{layout['saldo_extracto']}",
        )


def _agregar_control(wb, resultados: list[dict]) -> None:
    nombre = "_CONTROL_BANCOS"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)
    headers = [
        "Banco", "Cuenta", "Hoja", "Año", "Mes", "Movimientos", "Saldo inicial",
        "Créditos", "Débitos", "Saldo calculado", "Saldo extracto", "DIF",
        "Continuidad", "Estado", "Detalle", "Archivos",
    ]
    ws.append(headers)
    for result in resultados:
        for control in result["controles"]:
            ws.append(
                [
                    result["banco"],
                    result["cuenta"],
                    result["hoja"],
                    control.get("Año"),
                    control.get("Mes"),
                    control.get("Movimientos"),
                    control.get("Saldo inicial"),
                    control.get("Créditos"),
                    control.get("Débitos"),
                    control.get("Saldo calculado"),
                    control.get("Saldo extracto"),
                    control.get("DIF"),
                    control.get("Continuidad"),
                    control.get("Estado"),
                    control.get("Detalle"),
                    " | ".join(result["archivos"]),
                ]
            )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(name="Calibri", size=11, bold=True)
    for col in ("G", "H", "I", "J", "K", "L", "M"):
        for cell in ws[col][1:]:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
    widths = {
        "A": 20, "B": 20, "C": 28, "D": 10, "E": 10, "F": 12,
        "G": 15, "H": 15, "I": 15, "J": 16, "K": 16, "L": 12,
        "M": 13, "N": 12, "O": 45, "P": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def completar_cuadro_bancario_existente(
    excel_bytes: bytes,
    excel_nombre: str,
    pdfs: list[tuple[str, Any]],
    ejercicio: int | None = None,
) -> dict:
    """Devuelve bytes del Excel completado, controles y advertencias.

    `pdfs` acepta (nombre, bytes) o (nombre, Path) para no cargar todo en RAM.
    """
    if not excel_bytes:
        raise ValueError("Falta el Excel destino.")
    if not pdfs:
        raise ValueError("Falta al menos un extracto PDF.")
    extension = Path(excel_nombre).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise ValueError("El Excel destino debe ser .xlsx o .xlsm.")

    grupos, errores = procesar_pdfs_por_cuenta(pdfs)
    if not grupos:
        return {
            "excel": None,
            "resultados": [],
            "errores": errores,
            "nombre": excel_nombre,
            "modo": "sin_movimientos",
        }
    if ejercicio is None:
        match_year = re.search(r"\b(20\d{2})\b", Path(excel_nombre).stem)
        ejercicio = int(match_year.group(1)) if match_year else None
    if ejercicio is not None:
        grupos_filtrados = []
        for grupo in grupos:
            df_year = grupo["df"][grupo["df"]["_fecha"].map(lambda f: f.year == ejercicio)].copy()
            if df_year.empty:
                errores.append(
                    {
                        "archivo": " | ".join(grupo["archivos"]),
                        "banco": grupo["banco"],
                        "cuenta": grupo["cuenta"],
                        "motivo": f"No hay movimientos del ejercicio {ejercicio}.",
                    }
                )
                continue
            grupo = dict(grupo)
            grupo["df"] = df_year.reset_index(drop=True)
            grupos_filtrados.append(grupo)
        grupos = grupos_filtrados
    if not grupos:
        return {
            "excel": None,
            "resultados": [],
            "errores": errores,
            "nombre": excel_nombre,
            "modo": "sin_movimientos",
        }

    # Formato único del estudio.
    # - Si el libro del cliente YA tiene cuadro estándar → se completa ahí.
    # - Si no (p. ej. planilla de Ganancias) → se genera un Excel NUEVO limpio
    #   solo con hojas de cuadro (como Trujillo), sin mezclar hojas ajenas.
    wb_cliente = load_workbook(
        io.BytesIO(excel_bytes),
        keep_vba=(extension == ".xlsm"),
        keep_links=True,
        data_only=False,
    )
    usa_libro_cliente = _tiene_plantilla_cuadro(wb_cliente)
    if usa_libro_cliente:
        wb = wb_cliente
    else:
        wb = Workbook()
        default = wb.active
        default.title = "_tmp"
        _asegurar_plantilla_estandar(wb)
        if "_tmp" in wb.sheetnames:
            del wb["_tmp"]
        errores.insert(
            0,
            {
                "archivo": excel_nombre,
                "banco": "",
                "cuenta": "",
                "motivo": (
                    "El Excel del cliente no tenía el formato estándar de cuadro. "
                    "Se generó un libro NUEVO solo con cuadros bancarios "
                    "(meses / créditos / débitos / saldos / DIF con fórmulas), "
                    "sin mezclar la planilla de Ganancias."
                ),
            },
        )
        try:
            wb_cliente.close()
        except Exception:
            pass

    resultados = []
    for grupo in grupos:
        controles = _reforzar_continuidad_mensual(_controles_mensuales(grupo["df"]))
        ok_ctrl = [c for c in controles if c["Estado"] == "OK"]
        bad_ctrl = [c for c in controles if c["Estado"] != "OK"]
        for control in bad_ctrl:
            errores.append(
                {
                    "archivo": " | ".join(grupo["archivos"]),
                    "banco": grupo["banco"],
                    "cuenta": grupo["cuenta"],
                    "motivo": f"{control['Mes']} {control.get('Año')}: {control['Detalle']}",
                }
            )
        # Escribe todos los meses con saldos (aunque DIF ≠ 0).
        # El control queda en _CONTROL_BANCOS para revisar.
        escribibles = [
            c for c in controles
            if c.get("Saldo extracto") is not None and c.get("Mes_num") is not None
        ]
        if not escribibles:
            errores.append(
                {
                    "archivo": " | ".join(grupo["archivos"]),
                    "banco": grupo["banco"],
                    "cuenta": grupo["cuenta"],
                    "motivo": "Sin saldos de extracto: no se escribió la hoja.",
                }
            )
            continue

        try:
            meses_ok = {int(c["Mes_num"]) for c in escribibles}
            grupo_ok = dict(grupo)
            grupo_ok["df"] = grupo["df"][
                grupo["df"]["_fecha"].map(lambda f: int(f.month) in meses_ok)
            ].copy()
            ws = _obtener_hoja_destino_estandar(wb, grupo_ok)
            _escribir_grupo(ws, grupo_ok, escribibles)
            resultados.append(
                {
                    "banco": grupo["banco"],
                    "cuenta": grupo["cuenta"],
                    "hoja": ws.title,
                    "archivos": grupo["archivos"],
                    "controles": controles,
                    "formato": "cuadro_estandar",
                    "meses_ok": len(ok_ctrl),
                    "meses_con_dif": len(bad_ctrl),
                }
            )
        except Exception as exc:
            errores.append(
                {
                    "archivo": " | ".join(grupo["archivos"]),
                    "banco": grupo["banco"],
                    "cuenta": grupo["cuenta"],
                    "motivo": str(exc),
                }
            )

    if resultados:
        # No dejar la plantilla visible en el archivo entregado.
        if "_PLANTILLA_CUADRO" in wb.sheetnames and not usa_libro_cliente:
            try:
                del wb["_PLANTILLA_CUADRO"]
            except Exception:
                try:
                    wb["_PLANTILLA_CUADRO"].sheet_state = "hidden"
                except Exception:
                    pass
        _agregar_control(wb, resultados)
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass
        output = io.BytesIO()
        wb.save(output)
        stem = Path(excel_nombre).stem
        if usa_libro_cliente:
            nombre = f"{stem}_CUADROS_COMPLETADOS{extension}"
        else:
            # Nombre claro: no parece una copia de la DDJJ/Ganancias.
            anio = ejercicio or re.search(r"\b(20\d{2})\b", stem)
            anio = anio if isinstance(anio, int) else (int(anio.group(1)) if anio else "")
            nombre = f"CUADROS_BANCARIOS_{anio or 'ESTANDAR'}.xlsx"
        return {
            "excel": output.getvalue(),
            "resultados": resultados,
            "errores": errores,
            "nombre": nombre,
            "modo": "cuadros",
        }

    # Respaldo: siempre devolver algo usable en la carpeta del cliente.
    stem = Path(excel_nombre).stem
    nombre = f"{stem}_MOVIMIENTOS_DETECTADOS.xlsx"
    return {
        "excel": _exportar_movimientos_detectados(grupos, errores),
        "resultados": [
            {
                "banco": g.get("banco"),
                "cuenta": g.get("cuenta"),
                "hoja": "MOVIMIENTOS",
                "archivos": g.get("archivos") or [],
                "controles": _controles_mensuales(g["df"]),
            }
            for g in grupos
        ],
        "errores": errores,
        "nombre": nombre,
        "modo": "movimientos",
    }
