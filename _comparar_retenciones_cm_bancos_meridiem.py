"""Comparación mes a mes: Ret Bancos del papel CM vs retenciones Galicia+Macro."""
from __future__ import annotations

import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_formato_estudio import (  # noqa: E402
    _escribir_encabezado_hoja,
    _escribir_tabla,
    guardar_informe_excel,
)

CM_PATH = Path(r"T:\CLIENTES\GRUPO MERIDIEM\CONVENIO MULTILATERAL\LIQUIDACION CONVENIO MULTILATERAL.xlsx")
BANK_PATH = Path(r"C:\Users\recep\Desktop\Debitos_Retenciones_Galicia_Macro_Meridiem.xlsx")
OUT = Path(r"C:\Users\recep\Desktop\Comparacion_Retenciones_CM_vs_Bancos_Meridiem.xlsx")

PERIODOS = [
    "05-2025",
    "06-2025",
    "07-2025",
    "08-2025",
    "09-2025",
    "10-2025",
    "11-2025",
    "12-2025",
    "01-2026",
    "02-2026",
    "03-2026",
    "04-2026",
]
TOL = 1.0
INI, FIN = date(2025, 5, 1), date(2026, 4, 30)

TIPOS_OK = {
    "Ret. IIBB SIRCREB",
    "Ret. IIBB Tucumán",
    "Ret. IIBB",
    "Retención IIBB",
}


def periodo_from_date(d) -> str | None:
    if d is None or (isinstance(d, float) and pd.isna(d)):
        return None
    if isinstance(d, str):
        d = pd.to_datetime(d, dayfirst=True)
    if hasattr(d, "month"):
        return f"{d.month:02d}-{d.year}"
    return None


def as_date(f):
    if f is None:
        return None
    if isinstance(f, datetime):
        return f.date()
    return f


def fnum(x) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def es_ret_bancaria_comparable(tipo: str) -> bool:
    t = (tipo or "").strip()
    if t in TIPOS_OK:
        return True
    tl = t.lower()
    if "percep" in tl:
        return False
    if ("iibb" in tl or "sircreb" in tl) and "ret" in tl:
        return True
    return False


def extraer_cm() -> tuple[dict, list[dict], list[str]]:
    wb = openpyxl.load_workbook(CM_PATH, data_only=True, read_only=True)
    cm_mes: dict = {}
    cm_rows: list[dict] = []
    faltantes: list[str] = []

    for per in PERIODOS:
        if per not in wb.sheetnames:
            faltantes.append(per)
            cm_mes[per] = {
                "periodo": per,
                "retenciones": 0.0,
                "ret_bancos": 0.0,
                "perc_bancos": 0.0,
                "percepciones": 0.0,
                "fecha_hoja": None,
                "faltante": True,
            }
            continue

        ws = wb[per]
        rows = list(ws.iter_rows(min_row=1, max_row=25, max_col=14, values_only=True))
        provs = list(rows[7][3:11]) if len(rows) > 7 else []
        fecha_hoja = rows[3][10] if len(rows) > 3 else None
        ret = [None] * 8
        retb = [None] * 8
        percb = [None] * 8
        perc = [None] * 8
        for r in rows:
            label = r[2]
            if label == "Retenciones":
                ret = list(r[3:11])
            elif label == "Ret Bancos":
                retb = list(r[3:11])
            elif label == "Percepciones Bancos":
                percb = list(r[3:11])
            elif label == "Percepciones":
                perc = list(r[3:11])

        tot_ret = sum(fnum(x) for x in ret)
        tot_retb = sum(fnum(x) for x in retb)
        tot_percb = sum(fnum(x) for x in percb)
        tot_perc = sum(fnum(x) for x in perc)

        cm_mes[per] = {
            "periodo": per,
            "retenciones": round(tot_ret, 2),
            "ret_bancos": round(tot_retb, 2),
            "perc_bancos": round(tot_percb, 2),
            "percepciones": round(tot_perc, 2),
            "fecha_hoja": fecha_hoja,
            "faltante": False,
        }
        for i, prov in enumerate(provs):
            if not prov:
                continue
            v_ret = fnum(ret[i]) if i < len(ret) else 0.0
            v_retb = fnum(retb[i]) if i < len(retb) else 0.0
            if v_ret or v_retb:
                cm_rows.append(
                    {
                        "Período": per,
                        "Provincia": str(prov).strip(),
                        "Retenciones (clientes)": round(v_ret, 2) if v_ret else None,
                        "Ret Bancos (papel CM)": round(v_retb, 2) if v_retb else None,
                    }
                )

    wb.close()
    return cm_mes, cm_rows, faltantes


def extraer_bancos() -> list[dict]:
    wb = openpyxl.load_workbook(BANK_PATH, data_only=True, read_only=True)
    ws = wb["Lista"]
    bank_all: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        banco, mes_lbl, fecha, tipo, importe, desc = r[:6]
        if not fecha:
            continue
        bank_all.append(
            {
                "Banco": banco,
                "Mes extracto": mes_lbl,
                "Fecha": as_date(fecha),
                "Período": periodo_from_date(fecha),
                "Tipo": tipo,
                "Importe": float(importe or 0),
                "Descripción": desc,
            }
        )
    wb.close()
    return bank_all


def main() -> None:
    cm_mes, cm_rows, faltantes = extraer_cm()
    bank_all = extraer_bancos()

    bank_cmp = [b for b in bank_all if es_ret_bancaria_comparable(b["Tipo"])]
    bank_excl = [b for b in bank_all if not es_ret_bancaria_comparable(b["Tipo"])]

    bank_periodo = [
        b for b in bank_cmp if b["Fecha"] and INI <= b["Fecha"] <= FIN
    ]

    bank_by_mes: dict = defaultdict(
        lambda: {"Galicia": 0.0, "Macro": 0.0, "total": 0.0, "n": 0}
    )
    for b in bank_periodo:
        p = b["Período"]
        bank_by_mes[p][b["Banco"]] = round(bank_by_mes[p][b["Banco"]] + b["Importe"], 2)
        bank_by_mes[p]["total"] = round(bank_by_mes[p]["total"] + b["Importe"], 2)
        bank_by_mes[p]["n"] += 1

    fechas_g = [b["Fecha"] for b in bank_periodo if b["Banco"] == "Galicia"]
    fechas_m = [b["Fecha"] for b in bank_periodo if b["Banco"] == "Macro"]

    comp = []
    errores = []
    oks = []
    for per in PERIODOS:
        cm = cm_mes[per]
        cm_val = cm["ret_bancos"]
        b = bank_by_mes.get(per, {"Galicia": 0.0, "Macro": 0.0, "total": 0.0, "n": 0})
        bank_val = b["total"]
        dif = round(cm_val - bank_val, 2)
        if cm.get("faltante"):
            estado = "Falta hoja CM"
        elif abs(dif) <= TOL:
            estado = "OK"
        else:
            estado = "ERROR"
        row = {
            "Período": per,
            "Ret. CM (Ret Bancos)": cm_val,
            "Galicia": b["Galicia"],
            "Macro": b["Macro"],
            "Ret. bancos (total)": bank_val,
            "Diferencia (CM − Bancos)": dif,
            "Estado": estado,
            "Retenciones clientes (CM)": cm["retenciones"],
        }
        comp.append(row)
        if estado == "ERROR":
            errores.append(row)
        elif estado == "OK":
            oks.append(per)

    df_comp = pd.DataFrame(comp)
    df_bank = pd.DataFrame(bank_periodo)
    if not df_bank.empty:
        df_bank = df_bank.sort_values(["Período", "Banco", "Fecha"])
    df_cm_det = pd.DataFrame(cm_rows)

    por_mes = df_comp[
        [
            "Período",
            "Ret. CM (Ret Bancos)",
            "Galicia",
            "Macro",
            "Ret. bancos (total)",
            "Diferencia (CM − Bancos)",
            "Estado",
        ]
    ].copy()

    if not df_bank.empty:
        por_banco_mes = (
            df_bank.groupby(["Período", "Banco"], sort=False)
            .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
            .reset_index()
        )
    else:
        por_banco_mes = pd.DataFrame(columns=["Período", "Banco", "Cantidad", "Total"])

    tot_cm = round(df_comp["Ret. CM (Ret Bancos)"].sum(), 2)
    tot_bank = round(df_comp["Ret. bancos (total)"].sum(), 2)
    tot_dif = round(tot_cm - tot_bank, 2)
    n_err = sum(1 for r in comp if r["Estado"] == "ERROR")

    cob_g = f"{min(fechas_g)} → {max(fechas_g)}" if fechas_g else "sin datos en período"
    cob_m = f"{min(fechas_m)} → {max(fechas_m)}" if fechas_m else "sin datos en período"

    excl_periodo = [
        b for b in bank_excl if b.get("Fecha") and INI <= b["Fecha"] <= FIN
    ]
    excl_tot = round(sum(b["Importe"] for b in excl_periodo), 2)
    excl_n = len(excl_periodo)

    subtitulo = (
        "Comparación: papel CM fila «Ret Bancos» vs débitos bancarios IIBB/SIRCREB "
        f"(Galicia+Macro). Excluidas del lado bancos: Percepción IVA / Percepción IIBB "
        f"({excl_n} movs, ${excl_tot:,.2f} en el período)."
    )
    periodo_txt = "05/2025 → 04/2026"
    if faltantes:
        periodo_txt += f" | Hojas CM faltantes: {', '.join(faltantes)}"

    guardar_informe_excel(
        OUT,
        titulo="Comparación Retenciones CM vs Bancos — GRUPO MERIDIEM SRL",
        subtitulo=subtitulo,
        periodo=periodo_txt,
        kpis=[
            ("Meses comparados", len(PERIODOS), "int"),
            ("Meses con ERROR", n_err, "int"),
            ("Total Ret Bancos CM", tot_cm, "money"),
            ("Total ret. bancos", tot_bank, "money"),
            ("Diferencia neta", tot_dif, "money"),
            ("Cobertura Galicia (en período)", cob_g, "text"),
            ("Cobertura Macro (en período)", cob_m, "text"),
            ("Tolerancia", TOL, "money"),
        ],
        resumenes=[
            ("Comparación mes a mes (clave)", por_mes),
            ("Bancos por período y banco", por_banco_mes),
        ],
        detalle=(
            df_bank[
                [
                    "Banco",
                    "Período",
                    "Fecha",
                    "Tipo",
                    "Importe",
                    "Descripción",
                    "Mes extracto",
                ]
            ]
            if not df_bank.empty
            else pd.DataFrame()
        ),
        hoja_detalle="Detalle bancos",
        col_moneda=[
            "Importe",
            "Total",
            "Ret. CM (Ret Bancos)",
            "Galicia",
            "Macro",
            "Ret. bancos (total)",
            "Diferencia (CM − Bancos)",
            "Retenciones clientes (CM)",
        ],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    RED = PatternFill("solid", fgColor="FFC7CE")
    RED_F = Font(name="Calibri", size=11, color="9C0006", bold=True)
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    GREEN_F = Font(name="Calibri", size=11, color="006100")

    wb = load_workbook(OUT)
    ws = wb["Resumen"]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
        for c in row:
            if c.value == "ERROR":
                c.fill = RED
                c.font = RED_F
            elif c.value == "OK":
                c.fill = GREEN
                c.font = GREEN_F

    ws_c = wb.create_sheet("Comparación", 1)
    fila = _escribir_encabezado_hoja(
        ws_c,
        "Comparación mensual Ret Bancos CM vs bancos",
        "Diferencia = CM − (Galicia + Macro). ERROR si |dif| > $1.",
        "05/2025 → 04/2026",
    )
    _escribir_tabla(
        ws_c,
        por_mes,
        fila,
        col_moneda=[
            "Ret. CM (Ret Bancos)",
            "Galicia",
            "Macro",
            "Ret. bancos (total)",
            "Diferencia (CM − Bancos)",
        ],
    )
    for row in ws_c.iter_rows(min_row=fila + 1, max_row=fila + len(por_mes), max_col=7):
        est = row[6].value
        if est == "ERROR":
            row[6].fill = RED
            row[6].font = RED_F
        elif est == "OK":
            row[6].fill = GREEN
            row[6].font = GREEN_F

    ws_cm = wb.create_sheet("CM por provincia", 2)
    start = _escribir_encabezado_hoja(
        ws_cm,
        "Papel CM — Retenciones por provincia",
        "Ret Bancos = comparable con extractos. Retenciones (clientes) = informativo.",
        "05/2025 → 04/2026",
    )
    df_cm_out = (
        df_cm_det
        if not df_cm_det.empty
        else pd.DataFrame(
            columns=[
                "Período",
                "Provincia",
                "Retenciones (clientes)",
                "Ret Bancos (papel CM)",
            ]
        )
    )
    _escribir_tabla(
        ws_cm,
        df_cm_out,
        start,
        col_moneda=["Retenciones (clientes)", "Ret Bancos (papel CM)"],
    )

    if excl_periodo:
        df_excl = pd.DataFrame(excl_periodo)
        ws_e = wb.create_sheet("Excluidos (no IIBB ret)")
        f0 = _escribir_encabezado_hoja(
            ws_e,
            "Movimientos bancarios excluidos de la comparación",
            "Percepciones IVA / IIBB u otros que el papel CM no ubica en «Ret Bancos».",
            "05/2025 → 04/2026",
        )
        _escribir_tabla(
            ws_e,
            df_excl[["Banco", "Período", "Fecha", "Tipo", "Importe", "Descripción"]],
            f0,
            col_moneda=["Importe"],
            col_fecha=["Fecha"],
            total_col="Importe",
        )

    orden = ["Resumen", "Comparación", "CM por provincia", "Detalle bancos"]
    if "Excluidos (no IIBB ret)" in wb.sheetnames:
        orden.append("Excluidos (no IIBB ret)")
    for i, name in enumerate(orden):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    wb.save(OUT)

    print("OUT", OUT)
    print("FALTANTES_CM", faltantes)
    print("TOT_CM", tot_cm, "TOT_BANK", tot_bank, "DIF", tot_dif)
    print("COB_G", cob_g)
    print("COB_M", cob_m)
    print("EXCL", excl_n, excl_tot)
    print("--- ERRORES ---")
    for e in errores:
        print(
            f"{e['Período']}\tCM={e['Ret. CM (Ret Bancos)']}\t"
            f"Bancos={e['Ret. bancos (total)']}\tDif={e['Diferencia (CM − Bancos)']}\t"
            f"G={e['Galicia']}\tM={e['Macro']}"
        )
    print("--- OK ---", ", ".join(oks))
    print("--- ALL ---")
    for r in comp:
        print(
            f"{r['Período']}\t{r['Estado']}\tCM={r['Ret. CM (Ret Bancos)']}\t"
            f"B={r['Ret. bancos (total)']}\tD={r['Diferencia (CM − Bancos)']}"
        )


if __name__ == "__main__":
    main()
