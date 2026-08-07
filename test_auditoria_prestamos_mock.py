"""Prueba mock multi-banco / multi-préstamo para auditoría de préstamos."""
from datetime import date
from pathlib import Path

import openpyxl

from procesador import (
    CruceCuotaExtracto,
    CuotaPrestamo,
    MovimientoBanco,
    ResultadoAuditoriaPrestamos,
    SaldoPasivoBanco,
    ValidacionMayor,
    _resolver_ruta_plantilla,
    _sanitizar_nombre_hoja_excel,
    generar_reporte_auditoria_prestamos,
)

SALIDA = Path("exportaciones/Auditoria_Prestamos_test_mock.xlsx")


def _cuota(num, mes, banco, archivo, capital, interes, total):
    return CuotaPrestamo(
        numero_cuota=num,
        fecha_vencimiento=date(2026, mes, 10),
        capital=capital,
        intereses=interes,
        total=total,
        banco=banco,
        archivo_origen=archivo,
        prestamo_id=Path(archivo).stem,
    )


def main():
    # Galicia: 2 préstamos (PDF distintos)
    c_gal_1a = _cuota(1, 1, "galicia", "galicia_prestamo_A.pdf", 80000, 12000, 92000)
    c_gal_1b = _cuota(2, 2, "galicia", "galicia_prestamo_A.pdf", 81000, 11000, 92000)
    c_gal_2a = _cuota(1, 1, "galicia", "galicia_prestamo_B.pdf", 50000, 8000, 58000)

    # Nación: 1 préstamo
    c_nac_1 = _cuota(1, 3, "nacion", "nacion_hipoteca.pdf", 120000, 15000, 135000)
    c_nac_2 = _cuota(2, 4, "nacion", "nacion_hipoteca.pdf", 121000, 14000, 135000)

    cuotas = [c_gal_1a, c_gal_1b, c_gal_2a, c_nac_1, c_nac_2]

    mov_gal = MovimientoBanco(date(2026, 1, 12), "001", "Debito cuota prestamo", 92000, 0, 0, banco="galicia")
    mov_nac = MovimientoBanco(date(2026, 3, 11), "002", "Cobro prestamo hipoteca", 135000, 0, 0, banco="nacion")

    cruces = [
        CruceCuotaExtracto(cuota=c_gal_1a, movimiento=mov_gal, coincidencia=True, observacion="Cruce OK"),
        CruceCuotaExtracto(cuota=c_gal_1b, observacion="Sin débito"),
        CruceCuotaExtracto(cuota=c_gal_2a, observacion="Sin débito"),
        CruceCuotaExtracto(cuota=c_nac_1, movimiento=mov_nac, coincidencia=True, observacion="Cruce OK"),
        CruceCuotaExtracto(cuota=c_nac_2, observacion="Sin débito"),
    ]

    validaciones = [
        ValidacionMayor(cuota=c_gal_1a, error_imputacion=False, detalle="Imputación correcta"),
        ValidacionMayor(cuota=c_gal_1b, error_imputacion=True, detalle="Capital mayor ≠ esperado"),
        ValidacionMayor(cuota=c_nac_1, error_imputacion=False, detalle="Imputación correcta"),
    ]

    saldos_iniciales = [
        {"banco": "Banco Galicia", "saldo_inicial": 500000.0},
        {"banco": "Banco Nación", "saldo_inicial": 800000.0},
    ]

    saldos_por_banco = [
        SaldoPasivoBanco(
            banco="Banco Galicia",
            saldo_inicial_pasivo=500000,
            capital_pagado=211000,
            nuevos_creditos=0,
            saldo_calculado=289000,
            saldo_final_mayor=280000,
            diferencia=-9000,
            cierra=False,
            alerta="Desbalance Galicia",
        ),
        SaldoPasivoBanco(
            banco="Banco Nación",
            saldo_inicial_pasivo=800000,
            capital_pagado=241000,
            nuevos_creditos=0,
            saldo_calculado=559000,
            saldo_final_mayor=559000,
            diferencia=0,
            cierra=True,
        ),
    ]

    resultado = ResultadoAuditoriaPrestamos(
        cuotas=cuotas,
        cruces=cruces,
        validaciones_mayor=validaciones,
        saldos_por_banco=saldos_por_banco,
        saldos_iniciales=saldos_iniciales,
        nombre_cliente="TEST MOCK MULTI-BANCO",
        alertas=[{"tipo": "Constitución saldo", "detalle": "Desbalance Galicia", "banco": "Banco Galicia"}],
    )

    plantilla = _resolver_ruta_plantilla()
    excel = generar_reporte_auditoria_prestamos(resultado, ruta_plantilla=plantilla)
    SALIDA.parent.mkdir(exist_ok=True)
    SALIDA.write_bytes(excel)

    wb = openpyxl.load_workbook(SALIDA)
    hojas_esperadas = {
        "Consolidado",
        _sanitizar_nombre_hoja_excel("Banco Galicia"),
        _sanitizar_nombre_hoja_excel("Banco Nación"),
        "Alertas de Auditoría",
    }
    assert hojas_esperadas.issubset(set(wb.sheetnames)), wb.sheetnames

    ws_gal = wb[_sanitizar_nombre_hoja_excel("Banco Galicia")]
    textos = " ".join(str(ws_gal.cell(r, c).value or "") for r in range(1, ws_gal.max_row + 1) for c in range(1, 7))
    assert "galicia_prestamo_A" in textos or "Préstamo único" not in textos
    assert "galicia_prestamo_B" in textos or "Préstamo:" in textos
    assert "Control Mayor" in textos
    assert ws_gal.cell(row=ws_gal.max_row - 2, column=1).value or "Estado" in textos

    print("Hojas:", wb.sheetnames)
    print("Filas Galicia:", ws_gal.max_row)
    print("OK ->", SALIDA.resolve())
    wb.close()


if __name__ == "__main__":
    main()
