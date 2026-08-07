"""Prueba de clonación de plantilla y validación de saldos."""
from pathlib import Path

import openpyxl

from procesador import (
    BalanceMensual,
    ResultadoConciliacion,
    _extraer_lineas_pdf,
    _extraer_saldos_desde_lineas,
    _resolver_ruta_plantilla,
    aplicar_saldos_al_resultado,
    conciliar_movimientos,
    extraer_movimientos_anuales,
    generar_planilla_conciliacion,
)

EXTRACTO = Path("extractos bancarios/Extracto Banco Enero.pdf")
SALIDA = Path("exportaciones/Conciliacion_test.xlsx")
PLANTILLA = _resolver_ruta_plantilla()


def test_saldos_lineas():
    print("=== Test extracción saldos ===")
    lineas = _extraer_lineas_pdf(EXTRACTO)
    print(f"Líneas extraídas: {len(lineas)}")
    for l in lineas[:15]:
        if "saldo" in l.lower():
            print(" ", l[:80])
    si, sf = _extraer_saldos_desde_lineas(lineas)
    print(f"Saldo Inicial: {si}")
    print(f"Saldo Resumen/Final: {sf}")
    return si, sf


def test_planilla_completa():
    print("\n=== Test generación planilla ===")
    print(f"Plantilla: {PLANTILLA}")

    # Leer fuente original para comparar estilos
    wb_orig = openpyxl.load_workbook(PLANTILLA)
    ws_orig = wb_orig["Banco Santander"]
    font_orig_b57 = ws_orig["B57"].font.name
    font_orig_c60 = ws_orig["C60"].font.name
    formula_c60 = ws_orig["C60"].value
    formula_c62 = ws_orig["C62"].value
    wb_orig.close()

    movs, bancos, saldos = extraer_movimientos_anuales([EXTRACTO])
    print(f"Movimientos: {len(movs)}, Bancos: {bancos}")

    resultado = conciliar_movimientos(movs, [])
    resultado = aplicar_saldos_al_resultado(resultado, saldos)

    print(f"Balance cierra: {resultado.balance_cierra}")
    print(f"Saldo inicial: {resultado.saldo_inicial}")
    print(f"Ingresos: {resultado.ingresos}")
    print(f"Egresos: {resultado.egresos}")
    print(f"Saldo final: {resultado.saldo_final}")
    print(f"Diferencia: {resultado.diferencia_balance}")

    excel_bytes = generar_planilla_conciliacion(resultado, "TEST CLONACION PLANTILLA")
    SALIDA.parent.mkdir(exist_ok=True)
    SALIDA.write_bytes(excel_bytes)
    print(f"Guardado: {SALIDA.resolve()}")

    wb = openpyxl.load_workbook(SALIDA)
    ws = wb["Banco Santander"]

    # Enero 2026 = columna G
    celdas = {
        "G57": ws["G57"].value,
        "G58": ws["G58"].value,
        "G59": ws["G59"].value,
        "G60": ws["G60"].value,
        "G61": ws["G61"].value,
        "G62": ws["G62"].value,
    }
    print("\nCeldas clave (columna G - Enero 2026):")
    for addr, val in celdas.items():
        font = ws[addr].font.name
        print(f"  {addr} = {val!r}  [font={font}]")

    print(f"\nFórmula C60 original: {formula_c60!r}")
    print(f"Fórmula G60 generada: {ws['G60'].value!r}")
    print(f"Fórmula C62 original: {formula_c62!r}")
    print(f"Fórmula G62 generada: {ws['G62'].value!r}")

    # Detalle movimientos
    if "Detalle Extracto" in wb.sheetnames:
        det = wb["Detalle Extracto"]
        print(f"\nHoja Detalle Extracto: {det.max_row - 1} movimientos")
        print(f"  Encabezado B1 font: {det['B1'].font.name}")

    ok_font = ws["G57"].font.name == font_orig_b57
    ok_formula_60 = _es_formula(ws["G60"].value)
    ok_formula_62 = _es_formula(ws["G62"].value)
    ok_g57 = ws["G57"].value is not None
    ok_g61 = ws["G61"].value is not None

    print(f"\n✅ Font preservado (G57): {ok_font}")
    print(f"✅ G57 tiene saldo inicial: {ok_g57}")
    print(f"✅ G61 tiene saldo resumen: {ok_g61}")
    print(f"✅ G60 conserva fórmula: {ok_formula_60}")
    print(f"✅ G62 conserva fórmula: {ok_formula_62}")
    wb.close()


def _es_formula(valor) -> bool:
    return isinstance(valor, str) and valor.startswith("=")


if __name__ == "__main__":
    test_saldos_lineas()
    test_planilla_completa()
