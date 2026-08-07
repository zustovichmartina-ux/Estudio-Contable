"""Importación de legajos de sueldos desde Excel (export Tango / plantilla)."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, BinaryIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Columnas canónicas de la plantilla
COLUMNAS_PLANTILLA = (
    "CUIL",
    "Nombre",
    "Categoria",
    "SueldoBasico",
    "FechaIngreso",
    "AntiguedadAnios",
)

# Alias frecuentes (Tango / planillas del estudio) → canónico
_ALIAS: dict[str, str] = {
    "cuil": "CUIL",
    "c.u.i.l.": "CUIL",
    "cuit_cuil": "CUIL",
    "documento": "CUIL",
    "nombre": "Nombre",
    "apellido_y_nombre": "Nombre",
    "apeynom": "Nombre",
    "razon_social": "Nombre",
    "empleado": "Nombre",
    "categoria": "Categoria",
    "categoría": "Categoria",
    "cat": "Categoria",
    "convenio_categoria": "Categoria",
    "sueldobasico": "SueldoBasico",
    "sueldo_basico": "SueldoBasico",
    "sueldo basico": "SueldoBasico",
    "sueldo básico": "SueldoBasico",
    "basico": "SueldoBasico",
    "básico": "SueldoBasico",
    "remuneracion": "SueldoBasico",
    "remuneración": "SueldoBasico",
    "importe": "SueldoBasico",
    "fechaingreso": "FechaIngreso",
    "fecha_ingreso": "FechaIngreso",
    "fecha ingreso": "FechaIngreso",
    "fecha_de_ingreso": "FechaIngreso",
    "ingreso": "FechaIngreso",
    "alta": "FechaIngreso",
    "antiguedadanios": "AntiguedadAnios",
    "antiguedad_anios": "AntiguedadAnios",
    "antigüedad": "AntiguedadAnios",
    "antiguedad": "AntiguedadAnios",
    "antigüedad_años": "AntiguedadAnios",
    "anos_antiguedad": "AntiguedadAnios",
    "años": "AntiguedadAnios",
}


def _norm_header(h: Any) -> str:
    s = str(h or "").strip().lower()
    s = unicodedata_fold(s)
    s = re.sub(r"[\s./]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def unicodedata_fold(texto: str) -> str:
    import unicodedata

    t = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in t if not unicodedata.combining(c))


def generar_plantilla_legajos_bytes() -> bytes:
    """Excel vacío listo para completar o pegar export de Tango."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Legajos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(COLUMNAS_PLANTILLA, start=1):
        cell = ws.cell(1, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Fila de ejemplo (borrar / reemplazar)
    ws.append(
        [
            "20-12345678-9",
            "García, Ana",
            "Administrativo A",
            850000,
            date(2019, 3, 1),
            7,
        ]
    )
    ws["A2"].number_format = "@"
    ws["D2"].number_format = "#,##0.00"
    ws["E2"].number_format = "DD/MM/YYYY"

    ws2 = wb.create_sheet("Instrucciones")
    ws2["A1"] = "Plantilla de legajos — Estudio Contable"
    ws2["A1"].font = Font(bold=True, size=14)
    instrucciones = [
        "",
        "1. Completá la hoja «Legajos» o pegá columnas desde un export de Tango.",
        "2. Columnas obligatorias: CUIL, Nombre.",
        "3. SueldoBasico: número (sin $).",
        "4. FechaIngreso: DD/MM/AAAA o AAAA-MM-DD.",
        "5. AntiguedadAnios: años enteros. Si está vacío se calcula desde FechaIngreso.",
        "6. También se aceptan encabezados Tango: Apellido y Nombre, Básico, Fecha de ingreso, etc.",
        "7. Subí el archivo en Liquidación de Sueldos → Legajos y CCT → Importar desde Excel.",
        "",
        "Al importar, los CUIL existentes de la misma empresa se actualizan (no se duplican).",
    ]
    for i, linea in enumerate(instrucciones, start=2):
        ws2[f"A{i}"] = linea
    ws2.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_fecha(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return date.today().isoformat()
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        # Excel serial date
        try:
            return pd.to_datetime(val, unit="D", origin="1899-12-30").date().isoformat()
        except Exception:
            pass
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return date.today().isoformat()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).date().isoformat()
    except Exception:
        return date.today().isoformat()


def _parse_numero(val: Any, default: float = 0.0) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    s = str(val).strip().replace("$", "").replace(" ", "")
    if not s:
        return default
    # 1.234.567,89 → 1234567.89
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return default


def _calc_antiguedad(fecha_iso: str, explicit: Any) -> int:
    n = int(_parse_numero(explicit, -1))
    if n >= 0:
        return n
    try:
        ingreso = date.fromisoformat(fecha_iso)
        hoy = date.today()
        años = hoy.year - ingreso.year
        if (hoy.month, hoy.day) < (ingreso.month, ingreso.day):
            años -= 1
        return max(0, años)
    except ValueError:
        return 0


def _mapear_columnas(df: pd.DataFrame) -> pd.DataFrame:
    rename: dict[str, str] = {}
    for col in df.columns:
        key = _norm_header(col)
        canon = _ALIAS.get(key)
        if not canon:
            # probar sin guiones bajos
            canon = _ALIAS.get(key.replace("_", ""))
        if canon:
            rename[col] = canon
    out = df.rename(columns=rename)
    return out


def leer_legajos_desde_excel(fuente: str | BinaryIO | bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Lee Excel/CSV de legajos. Devuelve (filas_ok, errores).
    Cada fila: cuil, nombre, categoria, sueldo_basico, fecha_ingreso, antiguedad_anios.
    """
    if isinstance(fuente, bytes):
        fuente = BytesIO(fuente)

    name = ""
    if hasattr(fuente, "name"):
        name = str(getattr(fuente, "name") or "").lower()

    try:
        if name.endswith(".csv"):
            df = pd.read_csv(fuente, dtype=object)
        else:
            # Intentar Excel; si falla, CSV
            try:
                df = pd.read_excel(fuente, dtype=object)
            except Exception:
                if hasattr(fuente, "seek"):
                    fuente.seek(0)
                df = pd.read_csv(fuente, dtype=object)
    except Exception as exc:
        return [], [f"No se pudo leer el archivo: {exc}"]

    if df.empty:
        return [], ["El archivo no tiene filas de datos."]

    df = _mapear_columnas(df)
    faltantes = [c for c in ("CUIL", "Nombre") if c not in df.columns]
    if faltantes:
        cols = ", ".join(str(c) for c in df.columns)
        return [], [
            "Faltan columnas obligatorias: "
            + ", ".join(faltantes)
            + f". Columnas encontradas: {cols}. "
            "Descargá la plantilla o renombrá los encabezados del export de Tango."
        ]

    ok: list[dict[str, Any]] = []
    errores: list[str] = []

    for i, row in df.iterrows():
        excel_row = int(i) + 2  # header = 1
        cuil_raw = str(row.get("CUIL") or "").strip()
        nombre = str(row.get("Nombre") or "").strip()
        if (not cuil_raw or cuil_raw.lower() == "nan") and (
            not nombre or nombre.lower() == "nan"
        ):
            continue
        if not cuil_raw or cuil_raw.lower() == "nan":
            errores.append(f"Fila {excel_row}: CUIL vacío.")
            continue
        if not nombre or nombre.lower() == "nan":
            errores.append(f"Fila {excel_row}: Nombre vacío (CUIL {cuil_raw}).")
            continue

        cuil = re.sub(r"[^\d\-]", "", cuil_raw)
        if len(re.sub(r"\D", "", cuil)) < 11:
            errores.append(f"Fila {excel_row}: CUIL inválido «{cuil_raw}».")
            continue

        fecha = _parse_fecha(row.get("FechaIngreso"))
        categoria = str(row.get("Categoria") or "").strip()
        if categoria.lower() == "nan":
            categoria = ""
        sueldo = _parse_numero(row.get("SueldoBasico"), 0.0)
        antig = _calc_antiguedad(fecha, row.get("AntiguedadAnios"))

        ok.append(
            {
                "cuil": cuil,
                "nombre": nombre,
                "categoria": categoria,
                "sueldo_basico": sueldo,
                "fecha_ingreso": fecha,
                "antiguedad_anios": antig,
            }
        )

    if not ok and not errores:
        return [], ["No se encontraron filas válidas en el archivo."]

    return ok, errores
