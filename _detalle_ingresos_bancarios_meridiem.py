# -*- coding: utf-8 -*-
"""Detalle de ingresos bancarios (créditos) — Grupo Meridiem.

Extrae créditos Galicia + Macro con denominación legible del extracto
(origen, CUIT, TEF, etc.) y genera Excel formato estudio.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from excel_formato_estudio import (  # noqa: E402
    _escribir_tabla,
    guardar_informe_excel,
)

PDF_GALICIA = Path(r"c:\Users\recep\Downloads\ilovepdf_merged (24).pdf")
PDF_MACRO = Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged (1).pdf")
PDF_MACRO_FB = Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged.pdf")
OUT = Path(r"c:\Users\recep\Desktop\Ingresos_Bancarios_Meridiem_detalle.xlsx")

CUIT_CAJA = "30663205621"
CUIT_PROVINCIA = "30688254090"
CUIT_EXPERTA = "30687156168"
CUIT_BHN = "30693504186"
CUIT_MERIDIEM = "30714058386"
CUIT_MERIDIONAL = "30500051163"

CUIT_NOMBRES = {
    CUIT_CAJA: "CAJA DE SEGUROS S.A.",
    CUIT_PROVINCIA: "PROVINCIA ART / ASEGURADORA",
    CUIT_EXPERTA: "EXPERTA ART S.A.",
    CUIT_BHN: "BHN SEGUROS GENERALES S.A.",
    CUIT_MERIDIEM: "GRUPO MERIDIEM SRL",
    CUIT_MERIDIONAL: "LA MERIDIONAL CI",
}

MESES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

RE_MOV_GAL = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s+"
    r"(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s*$"
)
RE_MONEY = re.compile(r"-?\d{1,3}(?:\.\d{3})*(?:,\d{2})")
RE_MOV_MACRO = re.compile(r"^(\d{2}/\d{2}/\d{2})\s+(.+)$")
RE_SALDO_INI = re.compile(
    r"SALDO ULTIMO EXTRACTO AL .+?\s+(" + RE_MONEY.pattern + r")\s*$",
    re.I,
)
RE_CUIT = re.compile(r"\b(\d{11})\b")

# Continuaciones que no aportan origen (ruido de extracto Galicia)
RE_RUIDO_CONT = re.compile(
    r"^(VARIOS|BANCO |INDUSTRIAL AND|ACRED\.|Sucursal:|terminal:|"
    r"ENTRE BCOS|FIMA |P[aá]gina |Resumen de |Total \$|Consolidado|PERIODO |"
    r"TOTAL |Los dep|Dispon|El cr[eé]dito|Tasa |Datos de |Tipo de |N[uú]mero |"
    r"Cantidad |IVA:|CUIT |GRUPO MERIDIEM|Movimientos|Fecha Descri|CBU |"
    r"DT\.|REG\.|PROVEEDORES\s*$)",
    re.I,
)


def parse_ar_money(s: str) -> float:
    s = (s or "").strip().replace(" ", "")
    neg = s.startswith("-") or s.startswith("(")
    s = s.replace("-", "").replace("(", "").replace(")", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return -abs(float(s)) if neg else float(s)


def fmt_cuit(c: str) -> str:
    c = re.sub(r"\D", "", c or "")
    if len(c) == 11:
        return f"{c[:2]}-{c[2:10]}-{c[10]}"
    return c


def _es_no_cobranza(desc: str, conts: list[str] | None = None) -> bool:
    t = " ".join([desc] + (conts or [])).upper()
    if re.search(r"RESCATE FIMA|FIMA PREMIUM|Sol\.Resc|Liq\.Susc|SOL\.RESC", t, re.I):
        return True
    if re.search(r"MISMA\s*TITULARIDAD|TRANSFER\. CASH MISMA|CR TRANSF AUT SDO MISMO TIT", t, re.I):
        return True
    if CUIT_MERIDIEM in t.replace("-", ""):
        return True
    if re.search(r"BENEFICIO PYME|CASHBACK|INTERES|INTERÉS|AJUSTE", t, re.I):
        return True
    return False


def clasificar(desc: str, conts: list[str] | None = None) -> tuple[str, str | None]:
    t = " ".join([desc] + (conts or []))
    if _es_no_cobranza(desc, conts):
        if re.search(r"RESCATE|FIMA|Sol\.Resc|Liq\.Susc", t, re.I):
            return "Fondos / rescate", None
        if re.search(r"MISMA|TITULARIDAD|MERIDIEM", t, re.I):
            return "Traspaso mismo titular", CUIT_MERIDIEM
        if re.search(r"BENEFICIO PYME|CASHBACK", t, re.I):
            return "Beneficio / cashback", None
        if re.search(r"INTERES|INTERÉS", t, re.I):
            return "Intereses", None
        return "Otro crédito", None
    cuits = RE_CUIT.findall(t.replace("-", ""))
    for c in RE_CUIT.findall(t):
        if c not in cuits:
            cuits.append(c)
    cuit = None
    for c in cuits:
        if c != CUIT_MERIDIEM:
            cuit = c
            break
    tu = t.upper()
    if "CAJA DE SEGUROS" in tu or cuit == CUIT_CAJA:
        return "Cobranza cliente", CUIT_CAJA
    if "EXPERTA" in tu or cuit == CUIT_EXPERTA:
        return "Cobranza cliente", CUIT_EXPERTA
    if "PROVINCIA" in tu or cuit == CUIT_PROVINCIA:
        return "Cobranza cliente", CUIT_PROVINCIA
    if "BHN" in tu or cuit == CUIT_BHN:
        return "Cobranza cliente", CUIT_BHN
    if "MERIDIONAL" in tu or cuit == CUIT_MERIDIONAL:
        return "Cobranza cliente", CUIT_MERIDIONAL
    if re.search(r"TEF DATANET|TRANSFERENCIAS CASH|SNP PAGO|SERVICIO PAGO|N/C|ACRED", tu):
        return "Cobranza cliente", cuit
    return "Otro ingreso", cuit


def denominacion_galicia(desc: str, conts: list[str]) -> tuple[str, str]:
    """Denominación legible + referencia (si hay)."""
    origen = None
    cuit = None
    refs: list[str] = []
    fondo = None

    for c in conts:
        s = c.strip()
        if not s or RE_RUIDO_CONT.match(s):
            continue
        if re.match(r"^\d{11}$", s):
            if s != CUIT_MERIDIEM:
                cuit = s
            continue
        if re.match(r"^\d{6,10}$", s):
            refs.append(s)
            continue
        if re.search(r"FIMA PREMIUM|PREMIUM CLASE", s, re.I):
            fondo = s
            continue
        if re.search(r"CASHBACK|BENEFICIO", s, re.I):
            origen = origen or s
            continue
        if re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]", s) and len(s) >= 3:
            # Preferir razón social sobre rótulos genéricos
            if s.upper() in {"PROVEEDORES"}:
                continue
            if origen is None or ("CAJA" in s.upper() or "SEGUR" in s.upper() or "ART" in s.upper()
                                  or "PROVINCIA" in s.upper() or "EXPERTA" in s.upper()
                                  or "BHN" in s.upper() or "MERIDIONAL" in s.upper()):
                origen = s

    tipo = re.sub(r"\s+", " ", desc).strip()
    partes = [tipo]
    if origen:
        partes.append(origen)
    elif cuit and cuit in CUIT_NOMBRES:
        partes.append(CUIT_NOMBRES[cuit])
    if fondo:
        partes.append(fondo)
    if cuit:
        partes.append(f"CUIT {fmt_cuit(cuit)}")
    denom = " — ".join(partes)
    ref = refs[0] if refs else ""
    return denom[:200], ref


def denominacion_macro(desc: str) -> tuple[str, str, str | None]:
    """Denominación + referencia + CUIT opcional."""
    d = re.sub(r"\s+", " ", (desc or "").strip())
    cuit = None
    ref = ""

    m = re.match(r"TEF DATANET PR\s+(.+?)\s+(\d{11})\s+(\d+)\s*$", d, re.I)
    if m:
        nombre, cuit, ref = m.group(1).strip(), m.group(2), m.group(3)
        nombre = CUIT_NOMBRES.get(cuit, nombre)
        return f"TEF DATANET — {nombre} — CUIT {fmt_cuit(cuit)}", ref, cuit

    m = re.match(r"TEF DATANET PR\s+(.+?)\s+(\d{11})\b", d, re.I)
    if m:
        nombre, cuit = m.group(1).strip(), m.group(2)
        nombre = CUIT_NOMBRES.get(cuit, nombre)
        return f"TEF DATANET — {nombre} — CUIT {fmt_cuit(cuit)}", "", cuit

    if re.search(r"Sol\.Resc", d, re.I):
        mref = re.search(r"Sol\.Resc\s+(\d+)", d, re.I)
        if mref:
            ref = mref.group(1)
        return "Rescate fondos (Sol.Resc)", ref, None

    if re.search(r"Liq\.Susc", d, re.I):
        return "Suscripción / liquidación fondos (Liq.Susc)", "", None

    if re.search(r"CR TRANSF AUT SDO MISMO TIT|TRF MO CCDO|MISMO TIT", d, re.I):
        return "Transferencia mismo titular", "", CUIT_MERIDIEM

    if re.search(r"^N/C", d, re.I):
        return f"Nota de crédito — {d}", "", None

    m = re.search(r"(\d{11})", d)
    if m:
        cuit = m.group(1)
        nombre = CUIT_NOMBRES.get(cuit)
        if nombre:
            return f"{d[:80]} — {nombre}", "", cuit

    # Limpiar SUC / refs largas al final para la denom, guardar ref
    mref = re.search(r"(\d{7,})\s*$", d)
    if mref:
        ref = mref.group(1)
        d_clean = d[: mref.start()].strip()
    else:
        d_clean = d
    d_clean = re.sub(r"\s+SUC\.:\s*\d+", "", d_clean, flags=re.I).strip()
    return (d_clean or d)[:200], ref, cuit


def _dedupe(creds: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    out = []
    for c in creds:
        key = (c["banco"], c["fecha"], c["importe"], c["descripcion_raw"][:80])
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return out


def extraer_creditos_galicia(pdf: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf) as pdf_obj:
        for page in pdf_obj.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    creds: list[dict] = []
    i = 0
    while i < len(lineas):
        m = RE_MOV_GAL.match(lineas[i])
        if not m:
            i += 1
            continue
        fecha_s, desc, mon1, _mon2 = m.groups()
        monto = parse_ar_money(mon1)
        if monto <= 0.009 or desc.lower().startswith("total"):
            i += 1
            continue
        conts: list[str] = []
        j = i + 1
        while j < len(lineas):
            nxt = lineas[j]
            if RE_MOV_GAL.match(nxt) or nxt.startswith("Total $") or nxt.startswith("Resumen de"):
                break
            if nxt.startswith("Página") or ("Página" in nxt and "/" in nxt):
                break
            conts.append(nxt)
            j += 1
        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            i = j
            continue

        cat, cuit = clasificar(desc, conts)
        denom, ref = denominacion_galicia(desc, conts)
        if not cuit:
            m_c = RE_CUIT.search(" ".join(conts))
            if m_c and m_c.group(1) != CUIT_MERIDIEM:
                cuit = m_c.group(1)

        creds.append(
            {
                "banco": "Galicia",
                "fecha": fecha,
                "importe": round(monto, 2),
                "denominacion": denom,
                "referencia": ref,
                "descripcion_raw": desc,
                "categoria": cat,
                "cuit": cuit,
            }
        )
        i = j
    return _dedupe(creds)


def extraer_creditos_macro(pdf: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf) as pdf_obj:
        for page in pdf_obj.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    creds: list[dict] = []
    saldo_prev: float | None = None
    en_cc = False

    for ln in lineas:
        if "CUENTA CORRIENTE BANCARIA NRO" in ln.upper():
            en_cc = True
            saldo_prev = None
            continue
        if ln.startswith("CUENTA CORRIENTE ESPECIAL"):
            en_cc = False
            saldo_prev = None
            continue
        if not en_cc:
            continue

        m_ini = RE_SALDO_INI.search(ln)
        if m_ini:
            saldo_prev = parse_ar_money(m_ini.group(1))
            continue

        m = RE_MOV_MACRO.match(ln)
        if not m:
            continue
        fecha_s, resto = m.groups()
        montos = RE_MONEY.findall(resto)
        if len(montos) < 2:
            continue
        saldo = parse_ar_money(montos[-1])
        importe = abs(parse_ar_money(montos[-2]))
        if importe < 0.005:
            saldo_prev = saldo
            continue

        desc = resto
        for mon in montos[-2:]:
            idx = desc.rfind(mon)
            if idx >= 0:
                desc = (desc[:idx] + desc[idx + len(mon) :]).strip()
        desc = re.sub(r"\s+", " ", desc).strip()

        es_debito: bool | None = None
        if saldo_prev is not None:
            delta = round(saldo - saldo_prev, 2)
            if abs(abs(delta) - importe) <= 0.05:
                es_debito = delta < 0
            elif delta < -0.01:
                es_debito = True
            elif delta > 0.01:
                es_debito = False
        if es_debito is None:
            if re.search(r"^N/D|^IMP\.|^TRANSF:|^WNPOWER|^INACAP|^DEBITO|^RETIRO|^DB ", desc, re.I):
                es_debito = True
            elif re.search(r"^N/C|^TEF DATANET|Sol\.Resc|Liq\.Susc", desc, re.I):
                es_debito = False
            else:
                es_debito = bool(re.search(r"N/D|TRANSF:|Comision|COMISION|IMP\. AFIP", desc, re.I))

        saldo_prev = saldo
        if es_debito:
            continue

        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            continue

        denom, ref, cuit_desc = denominacion_macro(desc)
        cat, cuit = clasificar(desc)
        if not cuit:
            cuit = cuit_desc

        creds.append(
            {
                "banco": "Macro",
                "fecha": fecha,
                "importe": round(importe, 2),
                "denominacion": denom,
                "referencia": ref,
                "descripcion_raw": desc,
                "categoria": cat,
                "cuit": cuit,
            }
        )
    return _dedupe(creds)


def main() -> None:
    print("Extrayendo créditos Galicia...")
    creds_g = extraer_creditos_galicia(PDF_GALICIA)
    print(f"  {len(creds_g)} créditos — ${sum(c['importe'] for c in creds_g):,.2f}")

    pdf_m = PDF_MACRO if PDF_MACRO.exists() else PDF_MACRO_FB
    print(f"Extrayendo créditos Macro ({pdf_m.name})...")
    creds_m = extraer_creditos_macro(pdf_m)
    print(f"  {len(creds_m)} créditos — ${sum(c['importe'] for c in creds_m):,.2f}")

    todos = sorted(creds_g + creds_m, key=lambda c: (c["fecha"], c["banco"], -c["importe"]))

    filas = []
    for c in todos:
        mes_lbl = f"{MESES[c['fecha'].month]} {c['fecha'].year}"
        filas.append(
            {
                "Fecha": c["fecha"],
                "Banco": c["banco"],
                "Importe": c["importe"],
                "Denominación": c["denominacion"],
                "Referencia": c["referencia"] or "",
                "Clasificación": c["categoria"],
                "CUIT origen": fmt_cuit(c["cuit"]) if c["cuit"] else "",
                "Mes": mes_lbl,
                "Descripción extracto": c["descripcion_raw"],
            }
        )
    df = pd.DataFrame(filas)

    fmin, fmax = df["Fecha"].min(), df["Fecha"].max()
    periodo = f"{fmin.strftime('%d/%m/%Y')} — {fmax.strftime('%d/%m/%Y')}"
    tot = float(df["Importe"].sum())
    tot_g = float(df.loc[df["Banco"] == "Galicia", "Importe"].sum())
    tot_m = float(df.loc[df["Banco"] == "Macro", "Importe"].sum())
    n_g = int((df["Banco"] == "Galicia").sum())
    n_m = int((df["Banco"] == "Macro").sum())

    por_banco = (
        df.groupby("Banco", as_index=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .sort_values("Banco")
    )
    por_mes = (
        df.assign(_ym=df["Fecha"].apply(lambda d: d.strftime("%Y-%m")))
        .groupby(["_ym", "Mes"], as_index=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .sort_values("_ym")
        .drop(columns=["_ym"])
    )
    por_mes_banco = (
        df.assign(_ym=df["Fecha"].apply(lambda d: d.strftime("%Y-%m")))
        .groupby(["_ym", "Mes", "Banco"], as_index=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .sort_values(["_ym", "Banco"])
        .drop(columns=["_ym"])
    )
    por_clase = (
        df.groupby("Clasificación", as_index=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .sort_values("Total", ascending=False)
    )

    detalle_cols = [
        "Fecha",
        "Banco",
        "Importe",
        "Denominación",
        "Referencia",
        "Clasificación",
        "CUIT origen",
        "Mes",
        "Descripción extracto",
    ]

    guardar_informe_excel(
        OUT,
        titulo="Ingresos bancarios — Grupo Meridiem",
        subtitulo="Créditos / cobranzas y otros ingresos (Galicia + Macro)",
        periodo=periodo,
        kpis=[
            ("Cantidad de créditos", len(df), "int"),
            ("Total ingresos", tot, "money"),
            ("Galicia — cantidad", n_g, "int"),
            ("Galicia — total", tot_g, "money"),
            ("Macro — cantidad", n_m, "int"),
            ("Macro — total", tot_m, "money"),
            ("Período", periodo, "text"),
        ],
        resumenes=[
            ("Totales por banco", por_banco),
            ("Totales por mes", por_mes),
            ("Por mes y banco", por_mes_banco),
            ("Por clasificación", por_clase),
        ],
        detalle=df[detalle_cols],
        hoja_detalle="Detalle",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    # Hoja extra Por mes (totales mes × banco)
    wb = load_workbook(OUT)
    ws = wb.create_sheet("Por mes", 2)
    _escribir_tabla(
        ws,
        por_mes_banco,
        1,
        col_moneda=["Total"],
        zebra=True,
        total_col="Total",
    )
    wb.save(OUT)

    print(f"\nOK -> {OUT}")
    print(f"Creditos: {len(df)} | Total: ${tot:,.2f}")
    print(f"Fechas: {periodo}")
    print(f"Galicia: {n_g} / ${tot_g:,.2f} | Macro: {n_m} / ${tot_m:,.2f}")
    print("\nMuestra denominaciones Galicia:")
    for _, r in df[df["Banco"] == "Galicia"].head(6).iterrows():
        print(f"  {r['Fecha']} ${r['Importe']:,.2f} | {r['Denominacion']}")
    print("Muestra denominaciones Macro:")
    for _, r in df[df["Banco"] == "Macro"].head(6).iterrows():
        print(f"  {r['Fecha']} ${r['Importe']:,.2f} | {r['Denominacion']}")


if __name__ == "__main__":
    main()
