# -*- coding: utf-8 -*-
"""Cruce pagos FAECYS vs débitos Galicia/Macro — Grupo Meridiem SRL."""
from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from excel_formato_estudio import (
    BODY_FONT,
    DATE_FMT,
    MONEY_FMT,
    ZEBRA,
    _escribir_encabezado_hoja,
    _pintar_header_fila,
    guardar_informe_excel,
)

FAECYS = Path(r"c:\Users\recep\Downloads\Pagos_Cuit_30714058386_05-08-2026.xls")
GAL = Path(r"C:\Users\recep\Desktop\Debitos_Galicia_Meridiem_claro.xlsx")
MAC = Path(r"C:\Users\recep\Desktop\Debitos_Macro_Meridiem_claro.xlsx")
OUT = Path(r"C:\Users\recep\Desktop\Cruce_FAECYS_vs_Bancos_Meridiem.xlsx")

TOL_IMPORTE = 1.0
TOL_DIAS = 7
KW = re.compile(
    r"faecys|sindicat|cuota.?sind|aporte.?sind|emplead.*comerc|comerc.*emplead",
    re.I,
)


def parse_us_money(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(" ", "")
    if not s or s.lower() == "nan":
        return None
    if re.match(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$", s) or re.match(r"^-?\d+(\.\d+)?$", s):
        return float(s.replace(",", ""))
    if "," in s and "." in s:
        return float(s.replace(".", "").replace(",", "."))
    if "," in s:
        return float(s.replace(",", "."))
    return float(s)


def parse_interes(v, base, total):
    """FAECYS a veces guarda interés sin coma (16525 ≈ 165.25). Preferir Total - Base."""
    if total is not None and base is not None:
        return round(total - base, 2)
    return parse_us_money(v) or 0.0


def periodo_label(p):
    if p is None or (isinstance(p, float) and pd.isna(p)):
        return None
    s = str(int(float(p)))
    if len(s) == 6:
        return f"{s[4:6]}/{s[:4]}"
    return s


def load_lista(path: Path, banco: str) -> pd.DataFrame:
    d = pd.read_excel(path, sheet_name="Lista")
    d["Fecha"] = pd.to_datetime(d["Fecha"], errors="coerce")
    d["Importe"] = pd.to_numeric(d["Importe"], errors="coerce").abs()
    d["Detalle"] = d["Detalle"].astype(str)
    d["Banco"] = banco
    out = d[["Banco", "Fecha", "Importe", "Detalle"]].dropna(subset=["Fecha", "Importe"])
    out["fecha_d"] = out["Fecha"].dt.date
    out["es_candidato_kw"] = out["Detalle"].str.contains(KW, na=False)
    return out


def main() -> None:
    raw = pd.read_excel(FAECYS, engine="xlrd", header=0)
    raw = raw.dropna(how="all", subset=["Periodo", "Fecha de pago", "Total", "Importe 0,5%"])
    raw = raw[raw["Periodo"].notna()].copy()

    faecys = []
    for _, r in raw.iterrows():
        base = parse_us_money(r["Importe 0,5%"])
        total = parse_us_money(r["Total"])
        interes = parse_interes(r["Interes"], base, total)
        if total is None and base is not None:
            total = round(base + (interes or 0), 2)
        fp = r["Fecha de pago"]
        if pd.isna(fp) or str(fp).strip() == "":
            fecha = None
        else:
            fecha = pd.to_datetime(str(fp).strip(), dayfirst=True, errors="coerce")
            fecha = None if pd.isna(fecha) else fecha.to_pydatetime().date()
        acta = r["Nro. de Acta"]
        if pd.isna(acta):
            acta = None
        else:
            try:
                acta = int(float(acta))
            except Exception:
                acta = str(acta)
        faecys.append(
            {
                "Periodo": int(float(r["Periodo"])),
                "Periodo_lbl": periodo_label(r["Periodo"]),
                "Fecha_pago_FAECYS": fecha,
                "Nro_Acta": acta,
                "Importe_05": base,
                "Interes": interes,
                "Total": total,
            }
        )

    df_f = pd.DataFrame(faecys)
    deb = pd.concat(
        [load_lista(GAL, "Galicia"), load_lista(MAC, "Macro")],
        ignore_index=True,
    )
    cov_min = deb["fecha_d"].min()
    cov_max = deb["fecha_d"].max()

    used_bank: set[int] = set()
    matches: dict[int, list[int]] = {}
    notes: dict[int, str] = {}
    bulk_matched_items: set[int] = set()

    bulk_groups: dict[tuple, list[int]] = defaultdict(list)
    for i, row in df_f.iterrows():
        if row["Fecha_pago_FAECYS"] and row["Nro_Acta"]:
            bulk_groups[(row["Fecha_pago_FAECYS"], row["Nro_Acta"])].append(i)

    for key, idxs in bulk_groups.items():
        if len(idxs) < 2:
            continue
        suma = round(sum(df_f.loc[i, "Total"] or 0 for i in idxs), 2)
        fecha = key[0]
        cands = []
        for j, b in deb.iterrows():
            if j in used_bank:
                continue
            if abs(b["Importe"] - suma) <= TOL_IMPORTE:
                dias = abs((b["fecha_d"] - fecha).days)
                if dias <= TOL_DIAS:
                    cands.append((dias, abs(b["Importe"] - suma), j))
        cands.sort()
        if cands:
            j = cands[0][2]
            used_bank.add(j)
            for i in idxs:
                matches[i] = [j]
                notes[i] = f"Pago agrupado Acta {key[1]}: suma ${suma:,.2f} = 1 débito banco"
                if len(cands) > 1:
                    notes[i] += f" | otros candidatos: {len(cands) - 1}"
                bulk_matched_items.add(i)

    for i, row in df_f.iterrows():
        if i in matches:
            continue
        fecha = row["Fecha_pago_FAECYS"]
        total = row["Total"]
        base = row["Importe_05"]
        if total is None and base is None:
            continue

        targets = []
        if total is not None:
            targets.append(("Total", total))
        if base is not None and (total is None or abs(base - total) > 0.05):
            targets.append(("Importe_05", base))

        multi = []
        for label, amt in targets:
            for j, b in deb.iterrows():
                if j in used_bank:
                    continue
                diff = abs(b["Importe"] - amt)
                if diff > TOL_IMPORTE:
                    continue
                if fecha is None:
                    multi.append((0 if b["es_candidato_kw"] else 1, diff, 999, j, label))
                else:
                    dias = abs((b["fecha_d"] - fecha).days)
                    if dias > TOL_DIAS:
                        if not (
                            b["fecha_d"].year == fecha.year
                            and b["fecha_d"].month == fecha.month
                        ):
                            continue
                    multi.append((0 if b["es_candidato_kw"] else 1, dias, diff, j, label))
        if not multi:
            continue
        multi.sort()
        j = multi[0][-2]
        label = multi[0][-1]
        used_bank.add(j)
        matches[i] = [j]
        note = f"Match por {label}"
        if len(multi) > 1:
            note += f" | {len(multi)} candidatos (tomado el más cercano)"
        notes[i] = note

    rows_det = []
    for i, row in df_f.iterrows():
        fecha_f = row["Fecha_pago_FAECYS"]
        total = row["Total"] or 0.0
        if fecha_f:
            en_cobertura = cov_min <= fecha_f <= cov_max
        else:
            en_cobertura = int(row["Periodo"]) >= 202504

        if i in matches:
            j = matches[i][0]
            b = deb.loc[j]
            imp_b = float(b["Importe"])
            dif = round(imp_b - (total or 0), 2)
            if i in bulk_matched_items:
                estado = "PAGADO"
                dif = 0.0
            elif abs(imp_b - (total or 0)) <= TOL_IMPORTE:
                estado = "PAGADO"
            elif abs(imp_b - (row["Importe_05"] or 0)) <= TOL_IMPORTE:
                estado = "PAGADO"
                dif = round(imp_b - (row["Importe_05"] or 0), 2)
            else:
                estado = "PARCIAL"
            banco = b["Banco"]
            fecha_b = b["fecha_d"]
            det_b = b["Detalle"]
            nota = notes.get(i, "")
        else:
            banco = fecha_b = imp_b = dif = det_b = None
            if fecha_f is None and (not total):
                estado = "FALTA"
                nota = "Sin fecha ni importe en FAECYS (período declarado sin pago)"
            elif not en_cobertura:
                estado = "SIN EXTRACTO"
                nota = f"Fuera de cobertura extractos ({cov_min} a {cov_max})"
            else:
                estado = "FALTA"
                nota = "No se encontró débito bancario por importe±$1 y fecha ±7d / mismo mes"
            amt = total or row["Importe_05"]
            if amt and estado == "FALTA":
                soft = []
                for j, b in deb.iterrows():
                    if abs(b["Importe"] - amt) <= TOL_IMPORTE:
                        soft.append(
                            f"{b['Banco']} {b['fecha_d']} ${b['Importe']:,.2f} ({b['Detalle'][:40]})"
                        )
                if soft:
                    nota += " | candidatos solo por importe: " + "; ".join(soft[:3])

        rows_det.append(
            {
                "Periodo": row["Periodo_lbl"],
                "Periodo_YYYYMM": row["Periodo"],
                "Fecha pago FAECYS": fecha_f,
                "Nro. Acta": row["Nro_Acta"],
                "Importe 0,5%": row["Importe_05"],
                "Interés": row["Interes"],
                "Total FAECYS": total if total else None,
                "Estado": estado,
                "Banco": banco,
                "Fecha banco": fecha_b,
                "Importe banco": imp_b,
                "Diferencia": dif,
                "Detalle banco": det_b,
                "Nota cruce": nota,
            }
        )

    detalle = pd.DataFrame(rows_det)
    n_pag = int((detalle["Estado"] == "PAGADO").sum())
    n_falta = int((detalle["Estado"] == "FALTA").sum())
    n_parcial = int((detalle["Estado"] == "PARCIAL").sum())
    n_sin = int((detalle["Estado"] == "SIN EXTRACTO").sum())
    tot_faecys = float(detalle["Total FAECYS"].fillna(0).sum())
    tot_pag = float(detalle.loc[detalle["Estado"] == "PAGADO", "Total FAECYS"].fillna(0).sum())
    tot_falta = float(detalle.loc[detalle["Estado"] == "FALTA", "Total FAECYS"].fillna(0).sum())

    res_est = (
        detalle.groupby("Estado", dropna=False)
        .agg(Cantidad=("Periodo", "count"), **{"Total FAECYS": ("Total FAECYS", "sum")})
        .reset_index()
    )
    orden = {"PAGADO": 0, "PARCIAL": 1, "FALTA": 2, "SIN EXTRACTO": 3}
    res_est["_o"] = res_est["Estado"].map(lambda x: orden.get(x, 9))
    res_est = res_est.sort_values("_o").drop(columns="_o")

    det_cov = detalle[detalle["Estado"] != "SIN EXTRACTO"].copy()
    resumenes = [("Por estado", res_est)]
    if not det_cov.empty:
        det_cov["Año período"] = det_cov["Periodo_YYYYMM"].astype(str).str[:4]
        res_anio = (
            det_cov.groupby(["Año período", "Estado"])
            .agg(Cantidad=("Periodo", "count"), Total=("Total FAECYS", "sum"))
            .reset_index()
        )
        resumenes.append(("En cobertura extractos — por año/estado", res_anio))

    kpis = [
        ("Ítems FAECYS (períodos)", len(detalle), "int"),
        ("Pagados (cruzados con banco)", n_pag, "int"),
        ("Faltan (en cobertura extractos)", n_falta, "int"),
        ("Parciales", n_parcial, "int"),
        ("Sin extracto bancario", n_sin, "int"),
        ("Total FAECYS declarado", tot_faecys, "money"),
        ("Total pagado (cruzado)", tot_pag, "money"),
        ("Total falta (cobertura)", tot_falta, "money"),
        (
            "Cobertura extractos",
            f"{cov_min.strftime('%d/%m/%Y')} – {cov_max.strftime('%d/%m/%Y')}",
            "text",
        ),
    ]

    detalle_out = detalle.drop(columns=["Periodo_YYYYMM"]).copy()
    detalle_out["_ord"] = detalle_out["Estado"].map(
        {"FALTA": 0, "PARCIAL": 1, "PAGADO": 2, "SIN EXTRACTO": 3}
    )
    detalle_out["_per"] = detalle["Periodo_YYYYMM"]
    detalle_out = detalle_out.sort_values(["_ord", "_per"], ascending=[True, False]).drop(
        columns=["_ord", "_per"]
    )

    guardar_informe_excel(
        OUT,
        titulo="Cruce FAECYS vs Bancos — Grupo Meridiem SRL",
        subtitulo="CUIT 30-71405838-6 | Obligaciones FAECYS (0,5%) cruzadas con débitos Galicia y Macro",
        periodo=(
            f"Extractos: {cov_min.strftime('%d/%m/%Y')} a {cov_max.strftime('%d/%m/%Y')} | "
            "Criterio: importe ±$1 y fecha ±7 días o mismo mes"
        ),
        kpis=kpis,
        resumenes=resumenes,
        detalle=detalle_out,
        hoja_detalle="Detalle",
        col_moneda=[
            "Importe 0,5%",
            "Interés",
            "Total FAECYS",
            "Importe banco",
            "Diferencia",
            "Total",
            "Importe",
        ],
        col_fecha=["Fecha pago FAECYS", "Fecha banco", "Fecha"],
        total_col="Total FAECYS",
    )

    wb = load_workbook(OUT)
    ws = wb.create_sheet("Candidatos banco", 2)
    fila = _escribir_encabezado_hoja(
        ws,
        "Débitos bancarios candidatos FAECYS/sindicato",
        "Keyword FAECYS / Sindicato Empleados; indica si se usó en el cruce",
        "",
    )
    all_cand = deb[deb["es_candidato_kw"]].copy()
    all_cand["Usado en cruce"] = all_cand.index.map(lambda j: "Sí" if j in used_bank else "No")
    all_cand = all_cand.sort_values("Fecha", ascending=False)
    cols = ["Banco", "Fecha", "Importe", "Detalle", "Usado en cruce"]
    for c, h in enumerate(cols, 1):
        ws.cell(fila, c, h)
    _pintar_header_fila(ws, fila, len(cols))
    for ri, (_, r) in enumerate(all_cand.iterrows(), 1):
        rr = fila + ri
        ws.cell(rr, 1, r["Banco"]).font = BODY_FONT
        cell_f = ws.cell(rr, 2, r["Fecha"].to_pydatetime())
        cell_f.number_format = DATE_FMT
        cell_f.font = BODY_FONT
        cell_i = ws.cell(rr, 3, float(r["Importe"]))
        cell_i.number_format = MONEY_FMT
        cell_i.font = BODY_FONT
        ws.cell(rr, 4, r["Detalle"]).font = BODY_FONT
        ws.cell(rr, 5, r["Usado en cruce"]).font = BODY_FONT
        if ri % 2 == 0:
            for c in range(1, 6):
                ws.cell(rr, c).fill = ZEBRA
    if len(all_cand):
        ws.auto_filter.ref = f"A{fila}:{get_column_letter(5)}{fila + len(all_cand)}"
    ws.freeze_panes = f"A{fila + 1}"
    for c, w in enumerate([12, 14, 14, 36, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wb.save(OUT)

    print("OUT", OUT)
    print("PAGADO", n_pag, f"${tot_pag:,.2f}")
    print("FALTA", n_falta, f"${tot_falta:,.2f}")
    print("PARCIAL", n_parcial)
    print("SIN_EXTRACTO", n_sin)
    print(
        detalle[detalle["Estado"].isin(["PAGADO", "FALTA", "PARCIAL"])][
            [
                "Periodo",
                "Fecha pago FAECYS",
                "Total FAECYS",
                "Estado",
                "Banco",
                "Fecha banco",
                "Importe banco",
                "Nota cruce",
            ]
        ].to_string()
    )


if __name__ == "__main__":
    main()
