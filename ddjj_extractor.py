# -*- coding: utf-8 -*-
"""Extractor de datos de DDJJs argentinas desde PDF/texto."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional


def _limpiar_monto(s) -> float:
    """Parsea montos en formato argentino (1.234,56) o anglosajón (1,234.56)."""
    if not s:
        return 0.0
    txt = str(s).replace("$", "").replace("\xa0", "").replace(" ", "").strip()
    if not txt or txt in ("-", "—", "–"):
        return 0.0
    # Formato argentino: 1.234.567,89 → eliminar puntos de miles, coma→punto decimal
    if re.search(r"\d\.\d{3}", txt) and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt and re.search(r",\d{1,2}$", txt):
        txt = txt.replace(",", ".")
    else:
        txt = txt.replace(",", "")
    try:
        return round(float(txt), 2)
    except ValueError:
        return 0.0


def extraer_texto_pdf(pdf_path) -> list[str]:
    """
    Extrae texto de cada página del PDF.
    Usa pdfplumber para texto nativo; si hay páginas con < 50 chars usa
    fitz + EasyOCR como fallback.
    """
    import pdfplumber

    pdf_path = Path(pdf_path)
    paginas: list[str] = []

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                texto = page.extract_text() or ""
                paginas.append(texto)
    except Exception:
        paginas = []

    # Fallback OCR para páginas escaneadas
    paginas_finales: list[str] = []
    for i, texto in enumerate(paginas):
        if len(texto.strip()) >= 50:
            paginas_finales.append(texto)
        else:
            try:
                import fitz  # PyMuPDF
                import easyocr
                import numpy as np

                doc = fitz.open(str(pdf_path))
                page = doc[i]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                    pix.height, pix.width, pix.n
                )
                if pix.n == 4:
                    img_array = img_array[:, :, :3]
                reader = easyocr.Reader(["es"], verbose=False)
                resultados = reader.readtext(img_array)
                texto_ocr = " ".join(r[1] for r in resultados)
                paginas_finales.append(texto_ocr)
                doc.close()
            except Exception:
                paginas_finales.append(texto)

    return paginas_finales


def _parsear_periodo(texto: str, nombre_archivo: str = "") -> Optional[str]:
    """Extrae período MM/AAAA del texto o del nombre de archivo como fallback."""
    texto_lower = texto.lower()

    # Formato AFIP F.2002/F.2051: "Período: 202504" o "Período Secuencia\n202605"
    # Primero intento directo (F.2002: período en misma línea)
    m = re.search(r"per[ií]odo[:\s]+(\d{4})(\d{2})\b", texto_lower)
    if m:
        anio, mes = m.group(1), m.group(2)
        if 1 <= int(mes) <= 12:
            return f"{mes}/{anio}"
    # F.2051: período en línea siguiente (header "Período Secuencia" + siguiente línea "202605 Original")
    m = re.search(r"per[ií]odo\b[^\n]*\n\s*(\d{4})(0[1-9]|1[0-2])\b", texto_lower)
    if m:
        anio, mes = m.group(1), m.group(2)
        return f"{mes}/{anio}"
    # Búsqueda de YYYYMM standalone (cualquier 6 dígitos válidos como período)
    m = re.search(r"\b(20\d{2})(0[1-9]|1[0-2])\b", texto_lower)
    if m:
        anio, mes = m.group(1), m.group(2)
        return f"{mes}/{anio}"

    # Formato ARBA/genérico: "período MM/AAAA" o "Período: MM/AAAA" o "Período 4/2025"
    m = re.search(r"per[ií]odo[:\s]+(\d{1,2})/(\d{4})", texto_lower)
    if m:
        mes_str = m.group(1).zfill(2)
        return f"{mes_str}/{m.group(2)}"

    # Buscar "vencimiento DD/MM/AAAA" → extraer MM/AAAA
    m = re.search(r"vencimiento\s+\d{2}/(\d{2})/(\d{4})", texto_lower)
    if m:
        return f"{m.group(1)}/{m.group(2)}"

    # Patrón genérico en texto: MM/AAAA
    matches = re.findall(r"\b(\d{2})/(\d{4})\b", texto)
    for mes_str, anio_str in matches:
        mes = int(mes_str)
        anio = int(anio_str)
        if 1 <= mes <= 12 and 2000 <= anio <= 2099:
            return f"{mes_str}/{anio_str}"

    # Fallback desde nombre de archivo: "IVA Rele 2025-07.pdf" → "07/2025"
    if nombre_archivo:
        m = re.search(r"(\d{4})[_\-](\d{2})", nombre_archivo)
        if m:
            anio, mes = m.group(1), m.group(2)
            if 1 <= int(mes) <= 12 and 2000 <= int(anio) <= 2099:
                return f"{mes}/{anio}"
        m = re.search(r"(\d{2})[_\-](\d{4})", nombre_archivo)
        if m:
            mes, anio = m.group(1), m.group(2)
            if 1 <= int(mes) <= 12 and 2000 <= int(anio) <= 2099:
                return f"{mes}/{anio}"

    return None


def _parsear_fecha_vto(texto: str) -> Optional[str]:
    """Extrae fecha de vencimiento DD/MM/AAAA."""
    m = re.search(r"vencimiento[:\s]+(\d{2}/\d{2}/\d{4})", texto.lower())
    if m:
        return m.group(1)
    m = re.search(r"vto\.?\s*:?\s*(\d{2}/\d{2}/\d{4})", texto.lower())
    if m:
        return m.group(1)
    return None


# Keywords exclusivos del F.2002 de AFIP para boost de confianza
_KEYWORDS_IVA_AFIP = [
    "impuesto al valor agregado",
    "f. 2002", "f.2002", "formulario 2002",
    "f.2051", "f. 2051",           # formato nuevo ARCA 2026
    "nro. de transacci",  # "Nro. de Transacción"
    "mis aplicaciones web",
    "saldo técnico a favor",
    "saldo del impuesto a favor de arca",
    "saldo de impuesto a favor de arca",   # formato nuevo (sin "del")
    "saldo de libre disponibilidad",
    "débito fiscal",
    "crédito fiscal",
]


def _parsear_iva_afip(texto: str, nombre_archivo: str = "") -> Optional[dict]:
    """
    Detecta y parsea formulario F.2002 (IVA AFIP).
    Extrae: debito_fiscal, credito_fiscal, retenciones_percepciones, saldo_a_ingresar.
    """
    texto_lower = texto.lower()
    keywords_basicos = ["debito fiscal", "débito fiscal", "credito fiscal", "crédito fiscal",
                "saldo a ingresar", "formulario 731", "f. 731", "f.731", " iva ",
                "impuesto al valor agregado", "f. 2002", "f.2051", "nro. de transacci"]
    if not any(kw in texto_lower for kw in keywords_basicos):
        return None

    datos: dict = {}

    # Débito fiscal (busca "Total del Débito Fiscal" o simplemente "Débito Fiscal")
    m = re.search(r"total\s+del\s+d[eé]bito\s+fiscal(?:\s+del\s+per[ií]odo)?[\s\$]*([\d.,]+)", texto_lower)
    if not m:
        m = re.search(r"d[eé]bito\s+fiscal[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["debito_fiscal"] = _limpiar_monto(m.group(1))

    # Crédito fiscal
    m = re.search(r"total\s+del\s+cr[eé]dito\s+fiscal(?:\s+del\s+per[ií]odo)?[\s\$]*([\d.,]+)", texto_lower)
    if not m:
        m = re.search(r"cr[eé]dito\s+fiscal[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["credito_fiscal"] = _limpiar_monto(m.group(1))

    # Retenciones, percepciones y pagos a cuenta
    m = re.search(
        r"retenciones.*?pagos\s+a\s+cuenta.*?neto\s+de\s+restituciones[\s\$]*([\d.,]+)",
        texto_lower,
        re.DOTALL,
    )
    if not m:
        m = re.search(
            r"retenciones.*?pagos\s+a\s+cuenta.*?\$\s*([\d.,]+)",
            texto_lower,
            re.DOTALL,
        )
    if m:
        datos["retenciones_percepciones"] = _limpiar_monto(m.group(1))

    # Saldo del Impuesto a Favor de ARCA/AFIP — positivo = contribuyente DEBE pagar
    # El signo se preserva: si el PDF muestra "-1.234,56" el regex lo captura
    m = re.search(r"saldo\s+(?:del\s+)?impuesto\s+a\s+favor\s+de\s+(?:arca|afip)[\s\$]*(-?[\d.,]+)", texto_lower)
    if m:
        datos["saldo_a_ingresar"] = _limpiar_monto(m.group(1))

    # Fallback: "Saldo a Ingresar" o cálculo manual
    if "saldo_a_ingresar" not in datos:
        m = re.search(r"saldo\s+a\s+ingresar[\s:$]*(-?[0-9.,]+)", texto_lower)
        if m:
            datos["saldo_a_ingresar"] = _limpiar_monto(m.group(1))
        elif "debito_fiscal" in datos and "credito_fiscal" in datos:
            retenc = datos.get("retenciones_percepciones", 0)
            datos["saldo_a_ingresar"] = round(
                datos["debito_fiscal"] - datos["credito_fiscal"] - retenc, 2
            )

    # Campos de crédito del contribuyente → saldo_a_ingresar negativo
    # "Saldo Técnico a Favor": crédito IVA técnico pendiente de compensar
    if not datos.get("saldo_a_ingresar"):
        m = re.search(r"saldo\s+t[eé]cnico\s+a\s+favor[\s\$:]*(-?[\d.,]+)", texto_lower)
        if m:
            val = _limpiar_monto(m.group(1))
            if val > 0:
                datos["saldo_a_ingresar"] = -val  # negativo = crédito del contribuyente

    # "Saldo de Libre Disponibilidad": crédito disponible para devolución
    if not datos.get("saldo_a_ingresar"):
        m = re.search(r"saldo\s+(?:de\s+)?libre\s+disponibilidad[\s\$:]*(-?[\d.,]+)", texto_lower)
        if m:
            val = _limpiar_monto(m.group(1))
            if val > 0:
                datos["saldo_a_ingresar"] = -val

    datos["periodo"] = _parsear_periodo(texto, nombre_archivo)
    datos["fecha_vto"] = _parsear_fecha_vto(texto)
    return datos if datos else None


def _parsear_iibb_arba(texto: str) -> Optional[dict]:
    """
    Detecta y parsea declaración IIBB ARBA (Provincia de Buenos Aires).
    """
    texto_lower = texto.lower()
    keywords = ["arba", "ingresos brutos", "impuesto determinado",
                "provincia de buenos aires", "agip"]
    if not any(kw in texto_lower for kw in keywords):
        return None

    datos: dict = {}

    # Impuesto determinado
    m = re.search(r"impuesto\s+determinado[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["impuesto_determinado"] = _limpiar_monto(m.group(1))
    else:
        # Buscar "total a pagar" como alternativa
        m = re.search(r"total\s+a\s+pagar[\s:$]*([0-9.,]+)", texto_lower)
        if m:
            datos["impuesto_determinado"] = _limpiar_monto(m.group(1))

    # Saldo a favor del organismo (alternativa a impuesto determinado)
    if "impuesto_determinado" not in datos:
        m = re.search(r"saldo\s+a\s+favor\s+del\s+organismo[\s:$]*([\d.,]+)", texto_lower)
        if m:
            datos["impuesto_determinado"] = _limpiar_monto(m.group(1))

    # Total impuesto determinado / mínimo (formato ARBA)
    if "impuesto_determinado" not in datos:
        m = re.search(r"total\s+impuesto\s+determinado\s*/\s*m[ií]nimo[\s:$]*([\d.,]+)", texto_lower)
        if m:
            datos["impuesto_determinado"] = _limpiar_monto(m.group(1))

    # Retenciones de agentes (primer total, no banco)
    m = re.search(r"retenciones:\s+\$\s*([\d.,]+)", texto_lower)
    if m:
        datos["retenciones_agentes"] = _limpiar_monto(m.group(1))

    # Retenciones banco
    m = re.search(r"retenciones\s+banco:\s+\$\s*([\d.,]+)", texto_lower)
    if m:
        datos["retenciones_banco"] = _limpiar_monto(m.group(1))

    # Percepciones (total neto al final de la línea)
    m = re.search(r"percepciones:.*?\$\s*([\d.,]+)\s*$", texto_lower, re.MULTILINE)
    if m:
        datos["percepciones"] = _limpiar_monto(m.group(1))

    # Saldo a pagar / saldo DJ — signo preservado: negativo = saldo a favor del contribuyente
    m = re.search(r"saldo\s+de\s+la\s+dj.*?\$\s*(-?[\d.,]+)", texto_lower, re.DOTALL)
    if not m:
        m = re.search(
            r"saldo\s+acumulado\s+al\s+cierre.*?\$\s*(-?[\d.,]+)", texto_lower, re.DOTALL
        )
    if m:
        datos["saldo_a_pagar"] = _limpiar_monto(m.group(1))

    datos["periodo"] = _parsear_periodo(texto)
    datos["fecha_vto"] = _parsear_fecha_vto(texto)
    return datos if datos else None


def _parsear_iibb_cm03(texto: str) -> Optional[dict]:
    """
    Detecta y parsea CM03 (Convenio Multilateral).
    """
    texto_lower = texto.lower()
    keywords = ["convenio multilateral", "cm03", "cm 03", "comision arbitral",
                "comisión arbitral", "multilateral"]
    if not any(kw in texto_lower for kw in keywords):
        return None

    datos: dict = {}

    m = re.search(r"monto\s+total[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["monto_total"] = _limpiar_monto(m.group(1))
    else:
        m = re.search(r"impuesto\s+determinado[\s:$]*([0-9.,]+)", texto_lower)
        if m:
            datos["monto_total"] = _limpiar_monto(m.group(1))
            datos["impuesto_determinado"] = datos["monto_total"]

    datos["periodo"] = _parsear_periodo(texto)
    datos["fecha_vto"] = _parsear_fecha_vto(texto)
    return datos if datos else None


def _parsear_tsh(texto: str) -> Optional[dict]:
    """
    Detecta y parsea Tasa de Seguridad e Higiene (TSH) municipal.
    """
    texto_lower = texto.lower()
    keywords = ["tasa de seguridad e higiene", "tasa habilitacion", "tasa de habilitación",
                "tsh", "municipalidad", "mar del plata", "mgp"]
    if not any(kw in texto_lower for kw in keywords):
        return None

    datos: dict = {}

    m = re.search(r"tasa\s+determinada[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["tasa_determinada"] = _limpiar_monto(m.group(1))
    else:
        m = re.search(r"total[\s:$]*([0-9.,]+)", texto_lower)
        if m:
            datos["tasa_determinada"] = _limpiar_monto(m.group(1))

    datos["periodo"] = _parsear_periodo(texto)
    datos["fecha_vto"] = _parsear_fecha_vto(texto)
    return datos if datos else None


def _parsear_sueldos_f931(texto: str) -> Optional[dict]:
    """
    Detecta y parsea F.931 (declaración jurada de seguridad social / aportes patronales).
    """
    texto_lower = texto.lower()
    keywords = ["f. 931", "f.931", "f931", "declaracion jurada",
                "contribuciones patronales", "seguridad social", "sicoss"]
    if not any(kw in texto_lower for kw in keywords):
        return None

    datos: dict = {}

    m = re.search(r"sueldo\s+bruto[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["sueldo_bruto"] = _limpiar_monto(m.group(1))

    m = re.search(r"contribuciones\s+patronales[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["contribuciones_patronales"] = _limpiar_monto(m.group(1))

    m = re.search(r"retenciones?\s+(?:y\s+)?aportes?[\s:$]*([0-9.,]+)", texto_lower)
    if m:
        datos["retenciones_aportes"] = _limpiar_monto(m.group(1))
    else:
        m = re.search(r"aportes?\s+del?\s+trabajador[\s:$]*([0-9.,]+)", texto_lower)
        if m:
            datos["retenciones_aportes"] = _limpiar_monto(m.group(1))

    datos["periodo"] = _parsear_periodo(texto)
    datos["fecha_vto"] = _parsear_fecha_vto(texto)
    return datos if datos else None


def _detectar_tipo_forzado(nombre_archivo: str, texto: str) -> Optional[str]:
    """
    Detecta el tipo de DDJJ forzado por nombre de archivo o keywords provinciales exclusivos.
    Retorna el tipo forzado o None si no se puede determinar con certeza.
    """
    nombre_upper = nombre_archivo.upper()
    texto_lower = texto.lower()

    # IIBB: keywords exclusivos que no aparecen en IVA
    keywords_iibb_exclusivos = [
        "ingresos brutos", "arba", "agip", "r-606m", "r-001",
        "sifere", "cm03", "convenio multilateral", "impuesto sobre los ingresos brutos",
        "agencia de recaudacion provincia de buenos aires",
        "anticipo impuesto iibb",
    ]
    nombre_dice_iibb = any(kw in nombre_upper for kw in ["IIBB", "INGRESOS BRUTOS", "ARBA"])
    texto_dice_iibb = any(kw in texto_lower for kw in keywords_iibb_exclusivos)

    if nombre_dice_iibb or texto_dice_iibb:
        return "IIBB"

    # IVA: keywords exclusivos
    keywords_iva_exclusivos = [
        "f. 2002", "f.2002", "f.2051", "f. 2051",
        "impuesto al valor agregado",
        "debito fiscal", "débito fiscal",
        "credito fiscal", "crédito fiscal",
    ]
    nombre_dice_iva = "IVA" in nombre_upper
    texto_dice_iva = any(kw in texto_lower for kw in keywords_iva_exclusivos)

    if nombre_dice_iva and texto_dice_iva:
        return "IVA"
    if nombre_dice_iva and not texto_dice_iibb:
        return "IVA"

    # TSH
    if any(kw in nombre_upper for kw in ["TSH", "TISH", "SEGURIDAD E HIGIENE"]):
        return "TSH"

    # SUELDOS
    if any(kw in nombre_upper for kw in ["SUELDO", "F931", "F.931", "SICOSS"]):
        return "SUELDOS"

    return None  # No forzado, dejar que los parsers decidan


def extraer_datos_ddjj(pdf_path) -> dict:
    """
    Función pública principal.
    Extrae texto del PDF y detecta el tipo de DDJJ y sus montos.
    Incluye pre-clasificación por nombre de archivo y review loop automático.
    """
    pdf_path = Path(pdf_path)
    nombre_archivo = pdf_path.name

    paginas = extraer_texto_pdf(pdf_path)
    texto_completo = "\n".join(paginas)
    texto_crudo = texto_completo[:500]
    texto_lower = texto_completo.lower()

    # PRE-CLASIFICACIÓN: detectar tipo forzado por nombre/keywords antes de parsear
    tipo_forzado = _detectar_tipo_forzado(nombre_archivo, texto_completo)

    # Mapeo tipo forzado → subtipos compatibles
    _SUBTIPOS = {
        "IIBB": ["IIBB_ARBA", "IIBB_CM03", "IIBB"],
        "IVA": ["IVA"],
        "TSH": ["TSH"],
        "SUELDOS": ["SUELDOS"],
    }

    # Ordenar parsers: si hay tipo forzado, poner sus parsers primero
    parsers_conf = [
        ("IVA", lambda t: _parsear_iva_afip(t, nombre_archivo)),
        ("IIBB_CM03", _parsear_iibb_cm03),
        ("IIBB_ARBA", _parsear_iibb_arba),
        ("TSH", _parsear_tsh),
        ("SUELDOS", _parsear_sueldos_f931),
    ]

    if tipo_forzado == "IIBB":
        parsers_conf = [
            ("IIBB_CM03", _parsear_iibb_cm03),
            ("IIBB_ARBA", _parsear_iibb_arba),
        ]
    elif tipo_forzado == "IVA":
        parsers_conf = [("IVA", lambda t: _parsear_iva_afip(t, nombre_archivo))]
    elif tipo_forzado == "TSH":
        parsers_conf = [("TSH", _parsear_tsh)]
    elif tipo_forzado == "SUELDOS":
        parsers_conf = [("SUELDOS", _parsear_sueldos_f931)]

    # Campos de montos principales por tipo
    campos_principales = {
        "IVA": ["debito_fiscal", "credito_fiscal"],
        "IIBB_ARBA": ["impuesto_determinado"],
        "IIBB_CM03": ["monto_total"],
        "TSH": ["tasa_determinada"],
        "SUELDOS": ["sueldo_bruto", "contribuciones_patronales"],
    }

    # Keywords de alta confianza por tipo
    keywords_alta_confianza = {
        "IVA": [
            "impuesto al valor agregado", "nro. de transacci", "mis aplicaciones web",
            "saldo del impuesto a favor de arca", "saldo de impuesto a favor de arca",
            "saldo técnico a favor", "saldo de libre disponibilidad",
            "f. 2002", "f.2051",
        ],
        "IIBB_ARBA": [
            "agencia de recaudación provincia de buenos aires",
            "agencia de recaudacion provincia de buenos aires",
            "r-606m", "saldo de la dj", "percepciones aduaneras",
            "anticipo impuesto iibb", "ingresos brutos",
        ],
        "IIBB_CM03": ["comisión arbitral", "comision arbitral", "convenio multilateral", "cm03", "sifere"],
        "TSH": ["tasa de seguridad e higiene", "municipalidad de", "mgp"],
        "SUELDOS": ["f. 931", "sicoss", "seguridad social"],
    }

    resultado_final = None

    for tipo, parser in parsers_conf:
        datos = parser(texto_completo)
        if datos is None:
            continue

        periodo = datos.pop("periodo", None) or ""
        if not periodo:
            periodo = _parsear_periodo("", nombre_archivo) or ""

        fecha_vto = datos.pop("fecha_vto", None)
        montos = {k: v for k, v in datos.items() if isinstance(v, (int, float))}

        principales = campos_principales.get(tipo, [])
        encontrados = sum(1 for c in principales if c in montos and montos[c] > 0)

        kw_alta = keywords_alta_confianza.get(tipo, [])
        kw_encontrados = sum(1 for kw in kw_alta if kw in texto_lower)

        # Determinar confianza
        if kw_encontrados >= 2 and encontrados >= 1 and periodo:
            confianza = 0.97
        elif kw_encontrados >= 1 and encontrados >= 1 and periodo:
            confianza = 0.92
        elif periodo and encontrados >= len(principales):
            confianza = 1.0
        elif periodo and encontrados >= 1:
            confianza = 0.75
        elif encontrados >= 1:
            confianza = 0.5
            if not periodo:
                periodo = _parsear_periodo("", nombre_archivo) or ""
        else:
            confianza = 0.3
            if not periodo:
                periodo = _parsear_periodo("", nombre_archivo) or ""

        resultado_final = {
            "tipo": tipo,
            "periodo": periodo,
            "fecha_vto": fecha_vto,
            "montos": montos,
            "confianza": confianza,
            "texto_crudo": texto_crudo,
        }
        break

    # REVIEW LOOP: si el resultado no coincide con el tipo forzado, corregir
    if resultado_final and tipo_forzado:
        tipo_obtenido = resultado_final["tipo"]
        subtipos_validos = _SUBTIPOS.get(tipo_forzado, [tipo_forzado])
        if tipo_obtenido not in subtipos_validos:
            if tipo_forzado == "IIBB":
                datos_iibb = _parsear_iibb_arba(texto_completo) or _parsear_iibb_cm03(texto_completo)
                if datos_iibb:
                    periodo = datos_iibb.pop("periodo", None) or _parsear_periodo("", nombre_archivo) or ""
                    fecha_vto = datos_iibb.pop("fecha_vto", None)
                    montos = {k: v for k, v in datos_iibb.items() if isinstance(v, (int, float))}
                    resultado_final = {
                        "tipo": "IIBB_ARBA",
                        "periodo": periodo,
                        "fecha_vto": fecha_vto,
                        "montos": montos,
                        "confianza": 0.80,
                        "texto_crudo": texto_crudo,
                    }
                else:
                    resultado_final["tipo"] = "IIBB_ARBA"
                    resultado_final["confianza"] = 0.60

    if resultado_final:
        return resultado_final

    # No se detectó ningún tipo
    periodo_fallback = _parsear_periodo(texto_completo, nombre_archivo) or ""
    return {
        "tipo": "DESCONOCIDO",
        "periodo": periodo_fallback,
        "fecha_vto": _parsear_fecha_vto(texto_completo),
        "montos": {},
        "confianza": 0.2,
        "texto_crudo": texto_crudo,
    }


def parsear_mis_retenciones(archivo, periodo: str) -> dict:
    """
    Lee el archivo 'Mis Retenciones Impositivas' de AFIP (.xls o .xlsx).
    Filtra por período (MM/AAAA) usando la columna 'Fecha Ret./Perc.' (col 6).
    Discrimina RETENCION vs PERCEPCION por la columna 'Descripción Operación' (col 8).

    Retorna dict con:
        retenciones: float    - suma de filas tipo RETENCION del período
        percepciones: float   - suma de filas tipo PERCEPCION del período
        total: float          - retenciones + percepciones
        cantidad: int         - cantidad de registros del período
        detalle: list[dict]   - filas individuales para tabla en UI
        periodo_filtrado: str - período usado para filtrar
    """
    import tempfile
    from pathlib import Path as _Path

    # Guardar el archivo subido en un temporal si es necesario
    if hasattr(archivo, "read"):
        nombre = getattr(archivo, "name", "retenciones.xls")
        sufijo = _Path(nombre).suffix.lower() or ".xls"
        with tempfile.NamedTemporaryFile(delete=False, suffix=sufijo) as tmp:
            tmp.write(archivo.read())
            ruta = _Path(tmp.name)
        archivo.seek(0)
    else:
        ruta = _Path(archivo)
        nombre = ruta.name

    # Parsear período objetivo: "MM/AAAA" → mes y año
    try:
        partes = periodo.split("/")
        mes_obj = int(partes[0])
        anio_obj = int(partes[1])
    except Exception:
        return {
            "retenciones": 0.0, "percepciones": 0.0, "total": 0.0,
            "cantidad": 0, "detalle": [], "periodo_filtrado": periodo,
            "error": f"Período inválido: {periodo}",
        }

    filas: list[dict] = []

    try:
        if ruta.suffix.lower() == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(ruta))
            ws = wb.sheet_by_index(0)
            for i in range(1, ws.nrows):
                fecha_str = str(ws.cell_value(i, 6)).strip()
                try:
                    partes_f = fecha_str.split("/")
                    mes_f = int(partes_f[1])
                    anio_f = int(partes_f[2])
                except Exception:
                    continue
                if mes_f != mes_obj or anio_f != anio_obj:
                    continue
                try:
                    monto = float(ws.cell_value(i, 9))
                except Exception:
                    monto = 0.0
                tipo_op = str(ws.cell_value(i, 8)).strip().upper()
                filas.append({
                    "Agente": str(ws.cell_value(i, 1))[:40],
                    "Fecha": fecha_str,
                    "Tipo": tipo_op,
                    "Régimen": str(ws.cell_value(i, 5))[:40],
                    "Importe": monto,
                })
        else:
            import openpyxl
            wb = openpyxl.load_workbook(str(ruta), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
            col_fecha = next((j for j, h in enumerate(headers) if "fecha ret" in h.lower()), 6)
            col_tipo = next((j for j, h in enumerate(headers) if "operaci" in h.lower()), 8)
            col_monto = next((j for j, h in enumerate(headers) if "importe" in h.lower()), 9)
            col_agente = next((j for j, h in enumerate(headers) if "denominaci" in h.lower()), 1)
            col_regimen = next((j for j, h in enumerate(headers) if "descripci" in h.lower() and j > 4), 5)
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha_val = row[col_fecha]
                if fecha_val is None:
                    continue
                fecha_str = str(fecha_val).strip()
                try:
                    import datetime as _dt
                    if isinstance(fecha_val, (_dt.date, _dt.datetime)):
                        mes_f, anio_f = fecha_val.month, fecha_val.year
                    else:
                        partes_f = fecha_str.split("/")
                        mes_f = int(partes_f[1])
                        anio_f = int(partes_f[2])
                except Exception:
                    continue
                if mes_f != mes_obj or anio_f != anio_obj:
                    continue
                try:
                    monto = float(row[col_monto] or 0)
                except Exception:
                    monto = 0.0
                tipo_op = str(row[col_tipo] or "").strip().upper()
                filas.append({
                    "Agente": str(row[col_agente] or "")[:40],
                    "Fecha": fecha_str,
                    "Tipo": tipo_op,
                    "Régimen": str(row[col_regimen] or "")[:40],
                    "Importe": monto,
                })
    except Exception as exc:
        return {
            "retenciones": 0.0, "percepciones": 0.0, "total": 0.0,
            "cantidad": 0, "detalle": [], "periodo_filtrado": periodo,
            "error": str(exc),
        }

    retenciones = round(sum(f["Importe"] for f in filas if f["Tipo"] == "RETENCION"), 2)
    percepciones = round(sum(f["Importe"] for f in filas if f["Tipo"] == "PERCEPCION"), 2)
    total = round(retenciones + percepciones, 2)

    return {
        "retenciones": retenciones,
        "percepciones": percepciones,
        "total": total,
        "cantidad": len(filas),
        "detalle": filas,
        "periodo_filtrado": periodo,
    }


# ---------------------------------------------------------------------------
# Helpers internos para match_retenciones_iva
# ---------------------------------------------------------------------------

def _periodo_a_tuple(periodo: str):
    """'MM/AAAA' → (mes: int, año: int)"""
    partes = periodo.strip().split("/")
    return int(partes[0]), int(partes[1])


def _tuple_a_periodo(mes: int, anio: int) -> str:
    return f"{mes:02d}/{anio}"


def _periodo_adyacente(periodo: str, delta: int) -> str:
    """Retorna el período desplazado +delta meses (puede ser negativo)."""
    mes, anio = _periodo_a_tuple(periodo)
    total_meses = mes - 1 + delta          # 0-indexed
    anio_nuevo = anio + total_meses // 12
    mes_nuevo = total_meses % 12 + 1
    return _tuple_a_periodo(mes_nuevo, anio_nuevo)


def _leer_filas_xls(ruta, periodos: list) -> list:
    """
    Lee el archivo (xlrd o openpyxl) y devuelve las filas que pertenecen
    a cualquiera de los períodos indicados (lista de 'MM/AAAA').
    """
    import datetime as _dt
    from pathlib import Path as _Path

    targets = set()
    for p in periodos:
        try:
            m, a = _periodo_a_tuple(p)
            targets.add((m, a))
        except Exception:
            pass

    filas: list[dict] = []
    ruta = _Path(ruta)

    if ruta.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(ruta))
        ws = wb.sheet_by_index(0)
        for i in range(1, ws.nrows):
            fecha_str = str(ws.cell_value(i, 6)).strip()
            try:
                partes_f = fecha_str.split("/")
                mes_f = int(partes_f[1])
                anio_f = int(partes_f[2])
            except Exception:
                continue
            if (mes_f, anio_f) not in targets:
                continue
            try:
                monto = float(ws.cell_value(i, 9))
            except Exception:
                monto = 0.0
            tipo_op = str(ws.cell_value(i, 8)).strip().upper()
            filas.append({
                "Agente": str(ws.cell_value(i, 1))[:40],
                "Fecha": fecha_str,
                "Tipo": tipo_op,
                "Régimen": str(ws.cell_value(i, 5))[:40],
                "Importe": monto,
                "Período": _tuple_a_periodo(mes_f, anio_f),
            })
    else:
        import openpyxl
        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows())]
        col_fecha  = next((j for j, h in enumerate(headers) if "fecha ret" in h.lower()), 6)
        col_tipo   = next((j for j, h in enumerate(headers) if "operaci" in h.lower()), 8)
        col_monto  = next((j for j, h in enumerate(headers) if "importe" in h.lower()), 9)
        col_agente = next((j for j, h in enumerate(headers) if "denominaci" in h.lower()), 1)
        col_regimen = next((j for j, h in enumerate(headers) if "descripci" in h.lower() and j > 4), 5)
        for row in ws.iter_rows(min_row=2, values_only=True):
            fecha_val = row[col_fecha]
            if fecha_val is None:
                continue
            fecha_str = str(fecha_val).strip()
            try:
                if isinstance(fecha_val, (_dt.date, _dt.datetime)):
                    mes_f, anio_f = fecha_val.month, fecha_val.year
                else:
                    partes_f = fecha_str.split("/")
                    mes_f = int(partes_f[1])
                    anio_f = int(partes_f[2])
            except Exception:
                continue
            if (mes_f, anio_f) not in targets:
                continue
            try:
                monto = float(row[col_monto] or 0)
            except Exception:
                monto = 0.0
            tipo_op = str(row[col_tipo] or "").strip().upper()
            filas.append({
                "Agente": str(row[col_agente] or "")[:40],
                "Fecha": fecha_str,
                "Tipo": tipo_op,
                "Régimen": str(row[col_regimen] or "")[:40],
                "Importe": monto,
                "Período": _tuple_a_periodo(mes_f, anio_f),
            })

    return filas


def _intentar_match(filas: list, monto_target: float, tolerancia: float):
    """
    Aplica los 3 escenarios de match sobre las filas dadas.
    Retorna (escenario, suma_ret, suma_perc) o None si ninguno coincide.
    """
    suma_ret  = round(sum(f["Importe"] for f in filas if f["Tipo"] == "RETENCION"), 2)
    suma_perc = round(sum(f["Importe"] for f in filas if f["Tipo"] == "PERCEPCION"), 2)

    if abs(suma_ret - monto_target) <= tolerancia:
        return "solo_retenciones", suma_ret, 0.0
    if abs(suma_perc - monto_target) <= tolerancia:
        return "solo_percepciones", 0.0, suma_perc
    if abs(suma_ret + suma_perc - monto_target) <= tolerancia:
        return "mixto", suma_ret, suma_perc
    return None


# ---------------------------------------------------------------------------
# Función principal: match_retenciones_iva
# ---------------------------------------------------------------------------

def match_retenciones_iva(
    archivo_xls,
    periodo: str,
    monto_target: float,
    tolerancia: float = 0.01,
) -> dict:
    """
    Algoritmo de Match por Desglose con Self-Correction Loop.

    Parámetros:
        archivo_xls  – archivo XLS/XLSX de AFIP (Mis Retenciones Impositivas),
                       puede ser un objeto file-like (UploadedFile de Streamlit)
                       o una ruta en disco.
        periodo      – período del PDF en formato 'MM/AAAA'.
        monto_target – valor de 'retenciones_percepciones' extraído del PDF.
        tolerancia   – diferencia máxima aceptable para considerar match (default 0.01).

    Retorna dict con:
        escenario             : str   – 'solo_retenciones' | 'solo_percepciones' |
                                        'mixto' | 'expandido' | 'fallback'
        retenciones_asignadas : float
        percepciones_asignadas: float
        diferencia            : float – monto_target - ret_asignadas - perc_asignadas
        detalle               : list[dict]  – filas del Excel usadas
        periodos_usados       : list[str]   – períodos incluidos en la búsqueda
    """
    import tempfile
    from pathlib import Path as _Path

    # ── 1. Materializar archivo en disco si es file-like ────────────────────
    if hasattr(archivo_xls, "read"):
        nombre = getattr(archivo_xls, "name", "retenciones.xls")
        sufijo = _Path(nombre).suffix.lower() or ".xls"
        with tempfile.NamedTemporaryFile(delete=False, suffix=sufijo) as tmp:
            tmp.write(archivo_xls.read())
            ruta = _Path(tmp.name)
        archivo_xls.seek(0)
    else:
        ruta = _Path(archivo_xls)

    # ── 2. Validar período ───────────────────────────────────────────────────
    try:
        _periodo_a_tuple(periodo)
    except Exception:
        return {
            "escenario": "fallback",
            "retenciones_asignadas": 0.0,
            "percepciones_asignadas": 0.0,
            "diferencia": monto_target,
            "detalle": [],
            "periodos_usados": [],
            "error": f"Período inválido: {periodo}",
        }

    # ── 3. Paso 1 – período exacto ───────────────────────────────────────────
    periodos_exactos = [periodo]
    try:
        filas_exactas = _leer_filas_xls(ruta, periodos_exactos)
    except Exception as exc:
        return {
            "escenario": "fallback",
            "retenciones_asignadas": 0.0,
            "percepciones_asignadas": 0.0,
            "diferencia": monto_target,
            "detalle": [],
            "periodos_usados": periodos_exactos,
            "error": str(exc),
        }

    resultado = _intentar_match(filas_exactas, monto_target, tolerancia)
    if resultado:
        esc, ret, perc = resultado
        filas_usadas = filas_exactas
    else:
        # ── 4. Paso 2 – expandir ±1 mes ─────────────────────────────────────
        periodo_ant = _periodo_adyacente(periodo, -1)
        periodo_post = _periodo_adyacente(periodo, +1)
        periodos_ampliados = [periodo_ant, periodo, periodo_post]
        try:
            filas_ampliadas = _leer_filas_xls(ruta, periodos_ampliados)
        except Exception:
            filas_ampliadas = filas_exactas

        resultado_amp = _intentar_match(filas_ampliadas, monto_target, tolerancia)
        if resultado_amp:
            esc, ret, perc = resultado_amp
            esc = "expandido"
            filas_usadas = filas_ampliadas
            periodos_exactos = periodos_ampliados
        else:
            # ── 5. Fallback ──────────────────────────────────────────────────
            suma_ret_fb  = round(sum(f["Importe"] for f in filas_exactas if f["Tipo"] == "RETENCION"), 2)
            suma_perc_fb = round(sum(f["Importe"] for f in filas_exactas if f["Tipo"] == "PERCEPCION"), 2)
            esc = "fallback"
            ret = suma_ret_fb
            perc = suma_perc_fb
            filas_usadas = filas_exactas

    diferencia = round(monto_target - ret - perc, 2)

    return {
        "escenario": esc,
        "retenciones_asignadas": ret,
        "percepciones_asignadas": perc,
        "diferencia": diferencia,
        "detalle": filas_usadas,
        "periodos_usados": periodos_exactos,
    }
