# -*- coding: utf-8 -*-
"""Cruza liquidaciones Centro Médico 2025 vs créditos bancarios (Pago a proveedores recibido)."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from caratulas_centro_medico import CONCEPTOS, consolidar_carpeta, parse_caratula

DIR = Path(r"C:\Users\recep\Desktop\Centro Medico")
EXTRACTO = Path(r"C:\Users\recep\Downloads\Extracto_Santander_00720067005000850315_2025.xlsx")
OUT = DIR / "Cruce_Liquidaciones_vs_Banco_Centro_Medico_2025.xlsx"

MONEY = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE = Font(name="Calibri", bold=True, size=16, color="1F4E79")
SUB = Font(name="Calibri", size=11, color="666666")
BODY = Font(name="Calibri", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
ZEBRA = PatternFill("solid", fgColor="F2F2F2")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")

TOL_DIAS = 5
TOL_MONTO = 1.0  # $1


def _style_header(ws, row: int, n: int) -> None:
    for c in range(1, n + 1):
        cell = ws.cell(row, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL


def main() -> None:
    # Liquidaciones (reparse para montos frescos)
    pdfs = sorted(DIR.glob("Caratula_*.pdf"))
    liqs = []
    for pdf in pdfs:
        rec = parse_caratula(pdf)
        if not rec or not rec.get("fecha"):
            continue
        if rec["fecha"].year != 2025:
            continue
        imp = rec["importes"] or {}
        iva = float(imp.get("Valor IVA") or 0)
        total = float(rec["total"] or 0)
        # El banco abona el neto sin IVA (IVA va aparte al facturar)
        neto_banco = round(total - iva, 2)
        bruto = round(
            float(imp.get("Valor Honorarios Exentos") or 0)
            + float(imp.get("Neto Gravado Honorarios") or 0)
            + float(imp.get("Valor Gastos Exentos") or 0)
            + float(imp.get("Neto Gravado Gastos") or 0),
            2,
        )
        descuentos = round(
            float(imp.get("Derecho administrativo") or 0)
            + float(imp.get("Aporte fondo compensador") or 0)
            + float(imp.get("Retencion IIBB") or 0)
            + float(imp.get("Jubilacion") or 0)
            + float(imp.get("Impuesto a las ganancias") or 0)
            + float(imp.get("Recuperaciones bancarias") or 0),
            2,
        )
        liqs.append({
            **rec,
            "bruto": bruto,
            "iva": iva,
            "descuentos": descuentos,
            "total_con_iva": total,
            "neto_a_cobrar": neto_banco,
        })
    liqs.sort(key=lambda r: (r["fecha"], str(r["liquidacion"])))

    # Extracto: créditos Centro Médico (recibido / proveedores / honorarios)
    ext = pd.read_excel(EXTRACTO)
    ext["Fecha"] = pd.to_datetime(ext["Fecha"], dayfirst=True, errors="coerce")
    desc = ext["Descripcion"].astype(str)
    det = ext["Detalle"].astype(str) if "Detalle" in ext.columns else pd.Series([""] * len(ext))
    mask_cm = (
        det.str.contains("centro medic", case=False, na=False)
        | desc.str.contains("centro medic", case=False, na=False)
    ) & (pd.to_numeric(ext["Importe"], errors="coerce") > 0)
    pagos = ext.loc[mask_cm].copy().reset_index(drop=True)
    pagos["importe"] = pd.to_numeric(pagos["Importe"], errors="coerce").abs()
    pagos["usado"] = False

    calces = []
    usados_p = set()

    # 1) monto exacto + fecha cercana
    for liq in liqs:
        cand = []
        for i, p in pagos.iterrows():
            if i in usados_p:
                continue
            if pd.isna(p["Fecha"]):
                continue
            dif = abs(float(p["importe"]) - float(liq["neto_a_cobrar"]))
            if dif > TOL_MONTO:
                continue
            dias = abs((p["Fecha"].date() - liq["fecha"]).days)
            if dias > TOL_DIAS:
                continue
            cand.append((dias, dif, i, p))
        if not cand:
            continue
        cand.sort(key=lambda x: (x[0], x[1]))
        dias, dif, i, p = cand[0]
        usados_p.add(i)
        pagos.at[i, "usado"] = True
        # comprobante del detalle (último token numérico)
        det_txt = str(p.get("Detalle") or "")
        mcomp = re.search(r"(\d{5,})\s*$", det_txt.replace("\n", " "))
        calces.append({
            "estado": "CALZADO",
            "fecha_liq": liq["fecha"],
            "liquidacion": liq["liquidacion"],
            "bruto": liq["bruto"],
            "iva": liq["iva"],
            "descuentos": liq["descuentos"],
            "neto_a_cobrar": liq["neto_a_cobrar"],
            "fecha_banco": p["Fecha"].date(),
            "importe_banco": float(p["importe"]),
            "dif_monto": round(dif, 2),
            "dias": dias,
            "comprobante_banco": mcomp.group(1) if mcomp else "",
            "detalle_banco": det_txt.replace("\n", " | "),
            "archivo": liq["archivo"],
        })

    calzados_liq = {c["liquidacion"] for c in calces}
    for liq in liqs:
        if liq["liquidacion"] in calzados_liq:
            continue
        # ¿hay mismo monto en otro día?
        near = None
        for i, p in pagos.iterrows():
            if i in usados_p or pd.isna(p["Fecha"]):
                continue
            dif = abs(float(p["importe"]) - float(liq["neto_a_cobrar"]))
            if dif > TOL_MONTO:
                continue
            dias = abs((p["Fecha"].date() - liq["fecha"]).days)
            if near is None or dias < near[0]:
                near = (dias, dif, i, p)
        if near and near[0] <= 30:
            dias, dif, i, p = near
            usados_p.add(i)
            pagos.at[i, "usado"] = True
            det_txt = str(p.get("Detalle") or "")
            mcomp = re.search(r"(\d{5,})\s*$", det_txt.replace("\n", " "))
            calces.append({
                "estado": "CALZADO_FECHA_LEJOS",
                "fecha_liq": liq["fecha"],
                "liquidacion": liq["liquidacion"],
                "bruto": liq["bruto"],
                "iva": liq["iva"],
                "descuentos": liq["descuentos"],
                "neto_a_cobrar": liq["neto_a_cobrar"],
                "fecha_banco": p["Fecha"].date(),
                "importe_banco": float(p["importe"]),
                "dif_monto": round(dif, 2),
                "dias": dias,
                "comprobante_banco": mcomp.group(1) if mcomp else "",
                "detalle_banco": det_txt.replace("\n", " | "),
                "archivo": liq["archivo"],
            })
        else:
            calces.append({
                "estado": "SIN_PAGO_BANCO",
                "fecha_liq": liq["fecha"],
                "liquidacion": liq["liquidacion"],
                "bruto": liq["bruto"],
                "iva": liq["iva"],
                "descuentos": liq["descuentos"],
                "neto_a_cobrar": liq["neto_a_cobrar"],
                "fecha_banco": None,
                "importe_banco": None,
                "dif_monto": None,
                "dias": None,
                "comprobante_banco": "",
                "detalle_banco": "",
                "archivo": liq["archivo"],
            })

    pagos_sin = []
    for i, p in pagos.iterrows():
        if i in usados_p:
            continue
        det_txt = str(p.get("Detalle") or "")
        pagos_sin.append({
            "fecha_banco": p["Fecha"].date() if not pd.isna(p["Fecha"]) else None,
            "importe_banco": float(p["importe"]),
            "detalle_banco": det_txt.replace("\n", " | "),
            "descripcion": str(p.get("Descripcion") or ""),
        })

    calces.sort(key=lambda c: (c["fecha_liq"] or datetime.min.date(), str(c["liquidacion"])))

    n_ok = sum(1 for c in calces if c["estado"] == "CALZADO")
    n_lejos = sum(1 for c in calces if c["estado"] == "CALZADO_FECHA_LEJOS")
    n_sin = sum(1 for c in calces if c["estado"] == "SIN_PAGO_BANCO")
    sum_neto = round(sum(c["neto_a_cobrar"] for c in calces), 2)
    sum_banco = round(sum(c["importe_banco"] or 0 for c in calces if c["importe_banco"]), 2)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Cruce"
    ws["A1"] = "Cruce liquidaciones Centro Medico vs pagos bancarios 2025"
    ws["A1"].font = TITLE
    ws["A2"] = (
        f"Calzados ±{TOL_DIAS}d: {n_ok} | Fecha lejos: {n_lejos} | Sin pago: {n_sin} | "
        f"Pagos CM en banco sin liq: {len(pagos_sin)} | "
        f"Neto a cobrar: {sum_neto:,.2f} | Banco calzado: {sum_banco:,.2f}"
    )
    ws["A2"].font = SUB
    ws["A3"] = "Regla: neto a cobrar = TOTAL liquidacion − IVA (el IVA se abona al facturar)."
    ws["A3"].font = SUB

    headers = [
        "Estado", "Fecha liq", "Liquidacion", "Bruto", "IVA", "Descuentos/retenciones",
        "Neto a cobrar", "Fecha banco", "Importe banco", "Dif monto", "Dias",
        "Comprobante banco", "Detalle banco", "Archivo",
    ]
    hr = 5
    for i, h in enumerate(headers, 1):
        ws.cell(hr, i, h)
    _style_header(ws, hr, len(headers))

    money_cols = {4, 5, 6, 7, 9, 10}
    for i, c in enumerate(calces):
        r = hr + 1 + i
        vals = [
            c["estado"], c["fecha_liq"], c["liquidacion"], c["bruto"], c["iva"],
            c["descuentos"], c["neto_a_cobrar"], c["fecha_banco"], c["importe_banco"],
            c["dif_monto"], c["dias"], c["comprobante_banco"], c["detalle_banco"], c["archivo"],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(r, col, v)
            cell.font = BODY
            if col in (2, 8) and v is not None:
                cell.number_format = "DD/MM/YYYY"
            if col in money_cols and isinstance(v, (int, float)):
                cell.number_format = MONEY
        fill = OK_FILL if c["estado"] == "CALZADO" else (
            WARN_FILL if c["estado"] == "CALZADO_FECHA_LEJOS" else BAD_FILL
        )
        ws.cell(r, 1).fill = fill
        if i % 2 == 1 and c["estado"] == "CALZADO":
            for col in range(2, len(headers) + 1):
                ws.cell(r, col).fill = ZEBRA

    last = hr + len(calces)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = f"A{hr + 1}"
    widths = [18, 12, 12, 14, 12, 18, 14, 12, 14, 10, 8, 16, 45, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Pagos sin liquidación
    wp = wb.create_sheet("Pagos_CM_sin_liq")
    wp["A1"] = "Pagos Centro Medico en banco sin liquidacion calzada"
    wp["A1"].font = TITLE
    ph = ["Fecha banco", "Importe", "Descripcion", "Detalle"]
    for i, h in enumerate(ph, 1):
        wp.cell(3, i, h)
    _style_header(wp, 3, len(ph))
    for i, p in enumerate(pagos_sin):
        r = 4 + i
        wp.cell(r, 1, p["fecha_banco"]).number_format = "DD/MM/YYYY"
        wp.cell(r, 2, p["importe_banco"]).number_format = MONEY
        wp.cell(r, 3, p["descripcion"])
        wp.cell(r, 4, p["detalle_banco"])
        for c in range(1, 5):
            wp.cell(r, c).font = BODY
    for i, w in enumerate([12, 14, 28, 50], 1):
        wp.column_dimensions[get_column_letter(i)].width = w

    # Resumen
    wr = wb.create_sheet("Resumen", 0)
    wr["A1"] = "Resumen cruce 2025"
    wr["A1"].font = TITLE
    rows = [
        ("Liquidaciones 2025", len(liqs), sum_neto),
        ("Calzados (±5 días)", n_ok, round(sum(c["importe_banco"] or 0 for c in calces if c["estado"] == "CALZADO"), 2)),
        ("Calzados fecha lejos", n_lejos, round(sum(c["importe_banco"] or 0 for c in calces if c["estado"] == "CALZADO_FECHA_LEJOS"), 2)),
        ("Liquidaciones sin pago", n_sin, round(sum(c["neto_a_cobrar"] for c in calces if c["estado"] == "SIN_PAGO_BANCO"), 2)),
        ("Pagos CM banco sin liq", len(pagos_sin), round(sum(p["importe_banco"] for p in pagos_sin), 2)),
        ("Pagos CM en extracto (total)", len(pagos), round(float(pagos["importe"].sum()), 2)),
    ]
    for i, h in enumerate(["Concepto", "Cantidad", "Importe"], 1):
        wr.cell(3, i, h)
    _style_header(wr, 3, 3)
    for i, (c, q, imp) in enumerate(rows):
        r = 4 + i
        wr.cell(r, 1, c).font = BODY
        wr.cell(r, 2, q).font = BODY
        wr.cell(r, 3, imp).number_format = MONEY
        wr.cell(r, 3).font = BODY
    wr.column_dimensions["A"].width = 36
    wr.column_dimensions["B"].width = 12
    wr.column_dimensions["C"].width = 16
    wr["A11"] = "Neto a cobrar = Bruto − descuentos/retenciones − (el IVA no entra al pago bancario típico)."
    wr["A11"].font = SUB

    wb.save(OUT)
    print("OK", OUT)
    print("liq", len(liqs), "calzados", n_ok, "lejos", n_lejos, "sin", n_sin, "banco_sin", len(pagos_sin))
    print("neto", sum_neto, "banco_calzado", sum_banco)
    for c in calces:
        if c["estado"] != "CALZADO":
            print(c["estado"], c["fecha_liq"], c["liquidacion"], c["neto_a_cobrar"], c.get("fecha_banco"), c.get("importe_banco"))


if __name__ == "__main__":
    main()
