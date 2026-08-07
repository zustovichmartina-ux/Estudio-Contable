"""
Caja en USD — cuenta corriente en dólares + diferencia de cotización.

Separado del analizador de inversiones (FIFO de especies).

Método default (cta_cte):
- Bolsillos separados: dólares físicos vs bancarios.
- Ingreso → actualiza TC de posición (promedio ponderado).
- Egreso → Dif = monto × (TC_movimiento − TC_posición).

Método alternativo (fifo): lotes por ingreso al TC de la operación.
"""

from __future__ import annotations

import io
import re
from collections import deque
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inversiones import (
    _movimientos_galicia_caja_usd,
    _parse_fecha,
    _texto_pdf,
    _to_float,
    _norm,
)


# ---------------------------------------------------------------------------
# Tipos
# ---------------------------------------------------------------------------

@dataclass
class LoteUsd:
    fecha: date
    cantidad: float
    tc_costo: float
    origen: str = ""


@dataclass
class ResultadoCajaUsd:
    caja: list[dict] = field(default_factory=list)
    aplicaciones: list[dict] = field(default_factory=list)
    lotes_cierre: list[dict] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)
    resumen: dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Cotizaciones
# ---------------------------------------------------------------------------

def leer_cotizaciones_excel(data: bytes, archivo: str = "cotizaciones.xlsx") -> tuple[pd.DataFrame, list[str]]:
    """
    Excel: Fecha | TC  (alias: cotizacion, tipo_cambio, dolar, bna, oficial).
    Devuelve DataFrame ordenado con columnas Fecha (date) y TC (float).
    """
    avisos: list[str] = []
    try:
        df = pd.read_excel(io.BytesIO(data), dtype=object)
    except Exception as exc:
        return pd.DataFrame(columns=["Fecha", "TC"]), [f"{archivo}: no se pudo leer ({exc})"]

    if df.empty:
        return pd.DataFrame(columns=["Fecha", "TC"]), [f"{archivo}: vacío"]

    cols = {_norm(c): c for c in df.columns}
    col_f = None
    for a in ("fecha", "date", "dia", "día"):
        if a in cols:
            col_f = cols[a]
            break
    col_tc = None
    for a in (
        "tc", "tipo de cambio", "tipodecambio", "cotizacion", "cotización",
        "dolar", "dólar", "bna", "oficial", "px", "precio",
    ):
        if a in cols:
            col_tc = cols[a]
            break
    if col_f is None or col_tc is None:
        # Si hay exactamente 2 columnas, asumir orden Fecha / TC
        if len(df.columns) >= 2:
            col_f = df.columns[0]
            col_tc = df.columns[1]
            avisos.append(f"{archivo}: se asumió col1=Fecha, col2=TC.")
        else:
            return pd.DataFrame(columns=["Fecha", "TC"]), [
                f"{archivo}: faltan columnas Fecha y TC."
            ]

    filas: list[dict] = []
    for _, row in df.iterrows():
        f = row.get(col_f)
        if isinstance(f, datetime):
            f = f.date()
        elif isinstance(f, date):
            pass
        else:
            f = _parse_fecha(f)
        tc = _to_float(row.get(col_tc))
        if not f or not tc or tc <= 0:
            continue
        filas.append({"Fecha": f, "TC": float(tc)})

    out = pd.DataFrame(filas)
    if out.empty:
        avisos.append(f"{archivo}: sin filas de cotización válidas.")
        return out, avisos
    out = out.drop_duplicates("Fecha", keep="last").sort_values("Fecha").reset_index(drop=True)
    avisos.append(f"{archivo}: {len(out)} cotización(es) cargadas.")
    return out, avisos


def plantilla_cotizaciones_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"
    ws.append(["Fecha", "TC"])
    ws.append([date(2025, 1, 2), 1032.50])
    ws.append([date(2025, 1, 3), 1035.00])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws2 = wb.create_sheet("Instrucciones")
    ws2["A1"] = (
        "Cargá una fila por día hábil. TC = tipo de cambio a usar para valuación "
        "(habitual: BNA comprador, o el que use el estudio). "
        "Si falta un día, se usa el TC del día hábil anterior más cercano."
    )
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _tc_para_fecha(cotiz: dict[date, float], fechas_ord: list[date], f: date) -> float | None:
    """TC del día o último anterior disponible."""
    if not cotiz:
        return None
    if f in cotiz:
        return cotiz[f]
    # búsqueda hacia atrás
    prev = [x for x in fechas_ord if x <= f]
    if prev:
        return cotiz[prev[-1]]
    # si la fecha es anterior a toda la serie, usar la primera
    return cotiz[fechas_ord[0]] if fechas_ord else None


# ---------------------------------------------------------------------------
# Movimientos desde extracto CA USD
# ---------------------------------------------------------------------------

def movimientos_desde_pdfs_caja_usd(
    archivos: list[Any],
) -> tuple[pd.DataFrame, list[dict]]:
    """Lee PDFs de caja de ahorro en dólares (Galicia / similar)."""
    rows: list[dict] = []
    errs: list[dict] = []
    for up in archivos or []:
        nombre = getattr(up, "name", "archivo.pdf")
        data = up.getvalue() if hasattr(up, "getvalue") else bytes(up)
        if not str(nombre).lower().endswith(".pdf"):
            errs.append({"archivo": nombre, "motivo": "Solo PDF de extracto CA USD en este paso."})
            continue
        texto = _texto_pdf(data)
        if not (texto or "").strip():
            errs.append({"archivo": nombre, "motivo": "PDF sin texto extractable (¿escaneado?)"})
            continue
        movs = _movimientos_galicia_caja_usd(texto, nombre)
        if not movs:
            # Intentar parse genérico de líneas con signo
            movs = _movimientos_usd_generico(texto, nombre)
        if not movs:
            errs.append({
                "archivo": nombre,
                "motivo": "No se detectaron movimientos en el extracto.",
            })
            continue
        rows.extend(movs)

    if not rows:
        return pd.DataFrame(), errs

    df = pd.DataFrame(rows)
    df["_fecha"] = pd.to_datetime(df["_fecha"], errors="coerce")
    df = df.sort_values("_fecha", kind="mergesort").reset_index(drop=True)
    return df, errs


def _movimientos_usd_generico(texto: str, archivo: str) -> list[dict]:
    """Fallback: líneas Fecha … ±importe saldo."""
    pat = re.compile(
        r"(?P<fecha>\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<importe>[+\-]?\d{1,3}(?:\.\d{3})*,\d{2})\s+"
        r"(?P<saldo>\d{1,3}(?:\.\d{3})*,\d{2})\s*$"
    )
    movs: list[dict] = []
    for raw in (texto or "").splitlines():
        ln = re.sub(r"[ \t]+", " ", raw).strip()
        m = pat.match(ln)
        if not m:
            continue
        fecha = _parse_fecha(m.group("fecha"))
        if not fecha:
            continue
        imp = _to_float(m.group("importe"))
        if imp is None or abs(imp) < 0.0001:
            continue
        desc = m.group("desc").strip()
        tipo = "Ingreso" if imp > 0 else "Egreso"
        movs.append({
            "Fecha": fecha.strftime("%d/%m/%Y"),
            "_fecha": fecha,
            "Especie": "USD",
            "Grupo": "Dólar / MEP",
            "Tipo_Operacion": tipo,
            "Cantidad": abs(imp),
            "Precio": None,
            "Monto_Total": abs(imp),
            "Moneda": "USD",
            "Descripcion": desc,
            "Archivo origen": archivo,
            "Nueva_Clasificacion": None,
        })
    return movs


def movimientos_desde_excel_extracto(data: bytes, archivo: str = "extracto.xlsx") -> pd.DataFrame:
    """
    Acepta Excel del convertidor de extractos (Sheet1: Fecha, Descripcion, Importe)
    o columnas Debe/Haber / Credito/Debito en USD.
    """
    try:
        df = pd.read_excel(io.BytesIO(data), dtype=object)
    except Exception:
        return pd.DataFrame()

    cols = {_norm(c): c for c in df.columns}
    col_f = cols.get("fecha")
    col_desc = cols.get("descripcion") or cols.get("concepto") or cols.get("detalle")
    col_imp = cols.get("importe") or cols.get("monto") or cols.get("cantidad")

    rows: list[dict] = []
    if col_f and col_imp:
        for _, row in df.iterrows():
            f = row.get(col_f)
            if isinstance(f, datetime):
                f = f.date()
            elif not isinstance(f, date):
                f = _parse_fecha(f)
            if not f:
                continue
            imp = _to_float(row.get(col_imp))
            if imp is None or abs(imp) < 0.0001:
                continue
            desc = str(row.get(col_desc) or "").strip() if col_desc else ""
            tipo = "Ingreso" if imp > 0 else "Egreso"
            rows.append({
                "Fecha": f.strftime("%d/%m/%Y"),
                "_fecha": f,
                "Especie": "USD",
                "Grupo": "Dólar / MEP",
                "Tipo_Operacion": tipo,
                "Cantidad": abs(imp),
                "Precio": None,
                "Monto_Total": abs(imp),
                "Moneda": "USD",
                "Descripcion": desc,
                "Archivo origen": archivo,
                "Nueva_Clasificacion": None,
            })
        if rows:
            out = pd.DataFrame(rows)
            out["_fecha"] = pd.to_datetime(out["_fecha"], errors="coerce")
            return out.sort_values("_fecha", kind="mergesort").reset_index(drop=True)

    # Debe / Haber
    col_debe = None
    col_haber = None
    for a in ("debe", "debito", "débito", "egreso"):
        if a in cols:
            col_debe = cols[a]
            break
    for a in ("haber", "credito", "crédito", "ingreso"):
        if a in cols:
            col_haber = cols[a]
            break
    if col_f and (col_debe or col_haber):
        for _, row in df.iterrows():
            f = row.get(col_f)
            if isinstance(f, datetime):
                f = f.date()
            elif not isinstance(f, date):
                f = _parse_fecha(f)
            if not f:
                continue
            debe = _to_float(row.get(col_debe)) if col_debe else 0.0
            haber = _to_float(row.get(col_haber)) if col_haber else 0.0
            debe = debe or 0.0
            haber = haber or 0.0
            if abs(debe) < 0.0001 and abs(haber) < 0.0001:
                continue
            desc = str(row.get(col_desc) or "").strip() if col_desc else ""
            if haber > 0:
                rows.append({
                    "Fecha": f.strftime("%d/%m/%Y"),
                    "_fecha": f,
                    "Especie": "USD",
                    "Grupo": "Dólar / MEP",
                    "Tipo_Operacion": "Ingreso",
                    "Cantidad": abs(haber),
                    "Precio": None,
                    "Monto_Total": abs(haber),
                    "Moneda": "USD",
                    "Descripcion": desc,
                    "Archivo origen": archivo,
                    "Nueva_Clasificacion": None,
                })
            if debe > 0:
                rows.append({
                    "Fecha": f.strftime("%d/%m/%Y"),
                    "_fecha": f,
                    "Especie": "USD",
                    "Grupo": "Dólar / MEP",
                    "Tipo_Operacion": "Egreso",
                    "Cantidad": abs(debe),
                    "Precio": None,
                    "Monto_Total": abs(debe),
                    "Moneda": "USD",
                    "Descripcion": desc,
                    "Archivo origen": archivo,
                    "Nueva_Clasificacion": None,
                })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["_fecha"] = pd.to_datetime(out["_fecha"], errors="coerce")
    return out.sort_values("_fecha", kind="mergesort").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Motor cta. cte. dólares (físico / bancario) — diferencia de cambio
# ---------------------------------------------------------------------------

def procesar_cuenta_corriente_dolares(
    movimientos: list[dict],
    saldo_inicial_fisico: float = 0.0,
    saldo_inicial_bancario: float = 0.0,
    tc_inicial_fisico: float | None = None,
    tc_inicial_bancario: float | None = None,
    *,
    tc_inicial: float | None = None,
) -> list[dict]:
    """
    Loop iterativo de caja USD separando dólares físicos y bancarios.

    Cada bolsillo lleva su propio saldo y un TC de posición (costo promedio
    ponderado). Así no queda “pegado” un único tc_inicial cuando hubo ingresos
    posteriores a otra cotización.

    Movimiento esperado:
      tipo: 'fisico' | 'bancario'
      operacion: 'ingreso' | 'egreso'
      monto: float (USD)
      tc_movimiento: float
      fecha / descripcion / archivo (opcionales)

    Egreso → diferencia_cambio = monto × (tc_movimiento − tc_posición)
    Ingreso → diferencia_cambio = 0 y recalcula TC de posición.
    """
    if tc_inicial is not None:
        if tc_inicial_fisico is None:
            tc_inicial_fisico = tc_inicial
        if tc_inicial_bancario is None:
            tc_inicial_bancario = tc_inicial

    saldo_fisico = float(saldo_inicial_fisico or 0.0)
    saldo_bancario = float(saldo_inicial_bancario or 0.0)
    tc_fisico = float(tc_inicial_fisico or 0.0)
    tc_bancario = float(tc_inicial_bancario or 0.0)

    cta_cte: list[dict] = []

    # Traza de apertura (si hay saldo)
    for tipo, saldo, tc in (
        ("fisico", saldo_fisico, tc_fisico),
        ("bancario", saldo_bancario, tc_bancario),
    ):
        if saldo <= 0:
            continue
        cta_cte.append({
            "fecha": "",
            "descripcion": f"Saldo inicial {tipo}",
            "tipo_dolar": tipo,
            "operacion": "ingreso",
            "monto": round(saldo, 2),
            "tc_posicion_anterior": round(tc, 6),
            "tc_inicial": round(tc, 6),
            "tc_movimiento": round(tc, 6),
            "tc_posicion": round(tc, 6),
            "saldo_resultante": round(saldo, 2),
            "saldo_fisico": round(saldo_fisico, 2),
            "saldo_bancario": round(saldo_bancario, 2),
            "diferencia_cambio": 0.0,
            "ars_al_tc_op": round(saldo * tc, 2),
            "ars_costo": round(saldo * tc, 2),
            "archivo": "",
        })

    for mov in movimientos or []:
        tipo = str(mov.get("tipo") or "").strip().lower()
        if tipo in {"físico", "fisicos", "efectivo", "billete"}:
            tipo = "fisico"
        elif tipo in {"banco", "bancarios", "cuenta", "ca", "cc"}:
            tipo = "bancario"
        if tipo not in {"fisico", "bancario"}:
            raise ValueError(f"tipo_dolar desconocido: {mov.get('tipo')!r}")

        operacion = str(mov.get("operacion") or "").strip().lower()
        if operacion in {"credito", "crédito", "compra", "haber"}:
            operacion = "ingreso"
        elif operacion in {"debito", "débito", "venta", "debe"}:
            operacion = "egreso"
        if operacion not in {"ingreso", "egreso"}:
            raise ValueError(f"Operación desconocida: {operacion}")

        monto = float(mov.get("monto") or 0)
        if monto <= 0:
            continue
        tc_movimiento = float(mov.get("tc_movimiento") or 0)
        fecha = mov.get("fecha") or ""
        desc = mov.get("descripcion") or ""
        archivo = mov.get("archivo") or ""

        if tipo == "fisico":
            saldo_anterior = saldo_fisico
            tc_pos = tc_fisico
        else:
            saldo_anterior = saldo_bancario
            tc_pos = tc_bancario

        if operacion == "ingreso":
            nuevo_saldo = round(saldo_anterior + monto, 2)
            # Costo promedio ponderado → sin desvío acumulado vs un TC fijo viejo
            if saldo_anterior > 1e-9 and tc_pos > 0:
                tc_nuevo = (saldo_anterior * tc_pos + monto * tc_movimiento) / (saldo_anterior + monto)
            else:
                tc_nuevo = tc_movimiento
            diferencia_cambio = 0.0
            ars_costo = round(monto * tc_movimiento, 2)
        else:
            nuevo_saldo = round(saldo_anterior - monto, 2)
            # Salida contra el TC de posición (histórico / promedio) del bolsillo
            diferencia_cambio = round(monto * (tc_movimiento - tc_pos), 2)
            tc_nuevo = tc_pos  # el costo unitario del remanente no cambia
            ars_costo = round(monto * tc_pos, 2)

        if tipo == "fisico":
            saldo_fisico = nuevo_saldo
            tc_fisico = float(tc_nuevo)
        else:
            saldo_bancario = nuevo_saldo
            tc_bancario = float(tc_nuevo)

        cta_cte.append({
            "fecha": fecha,
            "descripcion": desc,
            "tipo_dolar": tipo,
            "operacion": operacion,
            "monto": round(monto, 2),
            "tc_posicion_anterior": round(tc_pos, 6),
            "tc_inicial": round(tc_pos, 6),  # alias: TC histórico del bolsillo al momento
            "tc_movimiento": round(tc_movimiento, 6),
            "tc_posicion": round(float(tc_nuevo), 6),
            "saldo_resultante": nuevo_saldo,
            "saldo_fisico": round(saldo_fisico, 2),
            "saldo_bancario": round(saldo_bancario, 2),
            "diferencia_cambio": diferencia_cambio,
            "ars_al_tc_op": round(monto * tc_movimiento, 2),
            "ars_costo": ars_costo,
            "archivo": archivo,
        })

    return cta_cte


def clasificar_tipo_dolar(descripcion: str, tipo_explicit: str | None = None) -> str:
    """Heurística: extracto bancario → bancario; efectivo/billetes → fisico."""
    if tipo_explicit:
        t = _norm(tipo_explicit)
        if "fisic" in t or "efect" in t or "billete" in t:
            return "fisico"
        if "banc" in t or t in {"ca", "cc", "cuenta"}:
            return "bancario"
    d = _norm(descripcion or "")
    if any(k in d for k in ("deposito en efectivo", "depósito en efectivo", "extraccion", "extracción", "billetera", "caja fuerte")):
        # El movimiento en el extracto bancario sigue siendo bancario;
        # el físico se carga aparte si hay Excel de caja chica.
        pass
    return "bancario"


# ---------------------------------------------------------------------------
# Motor caja + diferencia de cotización (wrapper DataFrame / Excel)
# ---------------------------------------------------------------------------

def _es_ingreso(tipo: str, desc: str = "") -> bool:
    t = _norm(tipo)
    d = _norm(desc)
    if t in {"ingreso", "compra", "credito", "crédito", "haber", "rescate"}:
        return True
    if t in {"egreso", "venta", "debito", "débito", "debe", "suscripcion", "suscripción"}:
        return False
    if "suscrip" in d:
        return False
    if "rescate" in d:
        return True
    return t in {"ingreso", "compra"}


def armar_caja_usd(
    df_mov: pd.DataFrame,
    df_cotiz: pd.DataFrame,
    *,
    saldo_inicial_usd: float = 0.0,
    saldo_inicial_fisico: float = 0.0,
    saldo_inicial_bancario: float | None = None,
    tc_inicial: float | None = None,
    tc_inicial_fisico: float | None = None,
    tc_inicial_bancario: float | None = None,
    fecha_inicial: date | None = None,
    metodo: str = "cta_cte",
) -> ResultadoCajaUsd:
    """
    Arma la caja USD.

    metodo='cta_cte' (default): físico + bancario con TC de posición ponderado
    (procesar_cuenta_corriente_dolares).

    metodo='fifo': lotes FIFO por ingreso (comportamiento anterior).
    """
    if metodo == "fifo":
        banc = saldo_inicial_bancario if saldo_inicial_bancario is not None else saldo_inicial_usd
        return _armar_caja_usd_fifo(
            df_mov,
            df_cotiz,
            saldo_inicial_usd=float(banc or 0) + float(saldo_inicial_fisico or 0),
            tc_inicial=tc_inicial or tc_inicial_bancario or tc_inicial_fisico,
            fecha_inicial=fecha_inicial,
        )

    res = ResultadoCajaUsd()
    cotiz_map: dict[date, float] = {}
    fechas_ord: list[date] = []
    if df_cotiz is not None and not df_cotiz.empty:
        for _, r in df_cotiz.iterrows():
            f = r["Fecha"]
            if isinstance(f, datetime):
                f = f.date()
            cotiz_map[f] = float(r["TC"])
        fechas_ord = sorted(cotiz_map.keys())

    if not cotiz_map:
        res.avisos.append(
            "Sin cotizaciones: se arma el saldo USD, pero la dif. de cambio queda en 0 / incompleta."
        )

    banc_ini = (
        float(saldo_inicial_bancario)
        if saldo_inicial_bancario is not None
        else float(saldo_inicial_usd or 0)
    )
    fis_ini = float(saldo_inicial_fisico or 0)

    movs: list[dict] = []
    work = df_mov.copy() if df_mov is not None and not df_mov.empty else pd.DataFrame()
    if not work.empty:
        if "_fecha" not in work.columns:
            work["_fecha"] = work["Fecha"].map(lambda x: _parse_fecha(x) or date.min)
        work["_fecha"] = pd.to_datetime(work["_fecha"], errors="coerce")
        work = work.sort_values("_fecha", kind="mergesort")

        for _, row in work.iterrows():
            fecha = row.get("_fecha")
            if isinstance(fecha, pd.Timestamp):
                fecha = fecha.date()
            elif isinstance(fecha, datetime):
                fecha = fecha.date()
            elif not isinstance(fecha, date):
                fecha = _parse_fecha(row.get("Fecha")) or date.min
            fecha_txt = fecha.strftime("%d/%m/%Y") if isinstance(fecha, date) and fecha.year > 1900 else str(row.get("Fecha") or "")
            desc = str(row.get("Descripcion") or "").strip()
            tipo_op = str(row.get("Tipo_Operacion") or "")
            cant = float(row.get("Cantidad") or 0)
            if cant <= 0:
                continue

            tc_op = row.get("Precio")
            if tc_op is None or (isinstance(tc_op, float) and pd.isna(tc_op)) or float(tc_op or 0) <= 0:
                tc_op = _tc_para_fecha(cotiz_map, fechas_ord, fecha) if cotiz_map else None
            else:
                tc_op = float(tc_op)
            if tc_op is None or tc_op <= 0:
                res.avisos.append(f"{fecha_txt}: sin TC — movimiento omitido del cálculo ARS.")
                tc_op = 0.0

            ingreso = _es_ingreso(tipo_op, desc)
            if _norm(tipo_op) == "compra":
                ingreso = True
            elif _norm(tipo_op) == "venta":
                ingreso = False

            tipo_dolar = clasificar_tipo_dolar(
                desc,
                str(row.get("Tipo_Dolar") or row.get("tipo_dolar") or ""),
            )
            movs.append({
                "tipo": tipo_dolar,
                "operacion": "ingreso" if ingreso else "egreso",
                "monto": cant,
                "tc_movimiento": float(tc_op),
                "fecha": fecha_txt,
                "descripcion": desc or tipo_op,
                "archivo": row.get("Archivo origen") or "",
            })

    tc_f = tc_inicial_fisico if tc_inicial_fisico and tc_inicial_fisico > 0 else tc_inicial
    tc_b = tc_inicial_bancario if tc_inicial_bancario and tc_inicial_bancario > 0 else tc_inicial
    if (fis_ini > 0 and (not tc_f or tc_f <= 0)) or (banc_ini > 0 and (not tc_b or tc_b <= 0)):
        # Buscar TC del día de apertura
        f0 = fecha_inicial or date(1900, 1, 1)
        tc_dia = _tc_para_fecha(cotiz_map, fechas_ord, f0) if cotiz_map else None
        if not tc_f or tc_f <= 0:
            tc_f = tc_dia
        if not tc_b or tc_b <= 0:
            tc_b = tc_dia

    try:
        cta = procesar_cuenta_corriente_dolares(
            movs,
            saldo_inicial_fisico=fis_ini,
            saldo_inicial_bancario=banc_ini,
            tc_inicial_fisico=tc_f,
            tc_inicial_bancario=tc_b,
        )
    except ValueError as exc:
        res.avisos.append(str(exc))
        return res

    total_dif = 0.0
    total_ing = 0.0
    total_egr = 0.0
    for row in cta:
        if row["descripcion"].startswith("Saldo inicial"):
            # fila de apertura
            pass
        elif row["operacion"] == "ingreso":
            total_ing += row["monto"]
        else:
            total_egr += row["monto"]
            total_dif += row["diferencia_cambio"]

        credito = row["monto"] if row["operacion"] == "ingreso" else 0.0
        debito = row["monto"] if row["operacion"] == "egreso" else 0.0
        res.caja.append({
            "Fecha": row["fecha"] or ("Saldo inicial" if "Saldo inicial" in row["descripcion"] else ""),
            "Descripcion": row["descripcion"],
            "Tipo_Dolar": row["tipo_dolar"],
            "Credito_USD": credito,
            "Debito_USD": debito,
            "Saldo_USD": row["saldo_resultante"],
            "Saldo_Fisico": row["saldo_fisico"],
            "Saldo_Bancario": row["saldo_bancario"],
            "TC_Operacion": row["tc_movimiento"],
            "TC_Posicion": row["tc_posicion"],
            "TC_Costo_FIFO": row["tc_posicion_anterior"] if row["operacion"] == "egreso" else row["tc_movimiento"],
            "ARS_al_TC_op": row["ars_al_tc_op"],
            "ARS_Costo": row["ars_costo"],
            "Dif_Cotizacion": row["diferencia_cambio"],
            "Saldo_ARS_Costo": round(
                row["saldo_fisico"] * (tc_f or 0) + row["saldo_bancario"] * (tc_b or 0),
                2,
            ),  # se recalcula abajo con posición real
            "Detalle": (
                "Ingreso → actualiza TC posición"
                if row["operacion"] == "ingreso"
                else f"Egreso → dif = monto×(TC_mov − TC_pos) = {row['diferencia_cambio']}"
            ),
            "Archivo origen": row.get("archivo") or "",
        })
        if row["operacion"] == "egreso":
            res.aplicaciones.append({
                "Fecha_egreso": row["fecha"],
                "Descripcion": row["descripcion"],
                "Tipo_Dolar": row["tipo_dolar"],
                "USD": row["monto"],
                "TC_Costo": row["tc_posicion_anterior"],
                "TC_Egreso": row["tc_movimiento"],
                "ARS_Costo": row["ars_costo"],
                "ARS_Egreso": row["ars_al_tc_op"],
                "Dif_Cotizacion": row["diferencia_cambio"],
            })

    # Recalcular Saldo_ARS_Costo corrido con TC de posición de cada bolsillo
    # (última fila ya tiene saldos; recalculamos recorriendo cta)
    # Mejor: al final usar última posición por tipo desde cta
    if cta:
        last = cta[-1]
        # TC posición final: buscar último tc_posicion por tipo
        tc_f_fin = tc_f or 0.0
        tc_b_fin = tc_b or 0.0
        for row in cta:
            if row["tipo_dolar"] == "fisico":
                tc_f_fin = row["tc_posicion"]
            else:
                tc_b_fin = row["tc_posicion"]
        saldo_ars = round(last["saldo_fisico"] * tc_f_fin + last["saldo_bancario"] * tc_b_fin, 2)
        # actualizar última columna en filas de forma incremental
        sf, sb = fis_ini, banc_ini
        tcf, tcb = float(tc_f or 0), float(tc_b or 0)
        for i, row in enumerate(cta):
            if row["tipo_dolar"] == "fisico":
                sf = row["saldo_resultante"]
                tcf = row["tc_posicion"]
            else:
                sb = row["saldo_resultante"]
                tcb = row["tc_posicion"]
            if i < len(res.caja):
                res.caja[i]["Saldo_ARS_Costo"] = round(sf * tcf + sb * tcb, 2)
        res.lotes_cierre = []
        if last["saldo_fisico"] > 0:
            res.lotes_cierre.append({
                "Fecha_lote": "posición",
                "Cantidad_USD": last["saldo_fisico"],
                "TC_Costo": tc_f_fin,
                "ARS_Costo": round(last["saldo_fisico"] * tc_f_fin, 2),
                "Origen": "fisico",
            })
        if last["saldo_bancario"] > 0:
            res.lotes_cierre.append({
                "Fecha_lote": "posición",
                "Cantidad_USD": last["saldo_bancario"],
                "TC_Costo": tc_b_fin,
                "ARS_Costo": round(last["saldo_bancario"] * tc_b_fin, 2),
                "Origen": "bancario",
            })
        res.resumen = {
            "Saldo_USD": round(last["saldo_fisico"] + last["saldo_bancario"], 2),
            "Saldo_Fisico": last["saldo_fisico"],
            "Saldo_Bancario": last["saldo_bancario"],
            "Saldo_ARS_Costo": saldo_ars,
            "Total_Ingresos_USD": round(total_ing, 2),
            "Total_Egresos_USD": round(total_egr, 2),
            "Total_Dif_Cotizacion": round(total_dif, 2),
            "Lotes_abiertos": len(res.lotes_cierre),
            "Metodo": "cta_cte (físico/bancario + TC posición)",
        }
    else:
        res.resumen = {
            "Saldo_USD": round(fis_ini + banc_ini, 2),
            "Saldo_Fisico": round(fis_ini, 2),
            "Saldo_Bancario": round(banc_ini, 2),
            "Saldo_ARS_Costo": 0.0,
            "Total_Ingresos_USD": 0.0,
            "Total_Egresos_USD": 0.0,
            "Total_Dif_Cotizacion": 0.0,
            "Lotes_abiertos": 0,
            "Metodo": "cta_cte (físico/bancario + TC posición)",
        }

    # Avisos de saldo negativo
    for row in cta:
        if row["saldo_resultante"] < -0.01:
            res.avisos.append(
                f"{row['fecha']} [{row['tipo_dolar']}]: saldo negativo ({row['saldo_resultante']:.2f})."
            )
    return res


def _armar_caja_usd_fifo(
    df_mov: pd.DataFrame,
    df_cotiz: pd.DataFrame,
    *,
    saldo_inicial_usd: float = 0.0,
    tc_inicial: float | None = None,
    fecha_inicial: date | None = None,
) -> ResultadoCajaUsd:
    """FIFO por lote de ingreso (alternativa)."""
    res = ResultadoCajaUsd()
    cotiz_map: dict[date, float] = {}
    fechas_ord: list[date] = []
    if df_cotiz is not None and not df_cotiz.empty:
        for _, r in df_cotiz.iterrows():
            f = r["Fecha"]
            if isinstance(f, datetime):
                f = f.date()
            cotiz_map[f] = float(r["TC"])
        fechas_ord = sorted(cotiz_map.keys())

    if not cotiz_map:
        res.avisos.append(
            "Sin cotizaciones: se puede armar el saldo USD, "
            "pero no la diferencia de cotización en ARS."
        )

    lotes: deque[LoteUsd] = deque()
    saldo_usd = 0.0
    saldo_ars_costo = 0.0
    total_dif = 0.0
    total_ing = 0.0
    total_egr = 0.0

    if saldo_inicial_usd and saldo_inicial_usd > 0:
        f0 = fecha_inicial or date(1900, 1, 1)
        tc0 = tc_inicial
        if tc0 is None or tc0 <= 0:
            tc0 = _tc_para_fecha(cotiz_map, fechas_ord, f0) if cotiz_map else 1.0
        if tc0 is None or tc0 <= 0:
            tc0 = 1.0
            res.avisos.append("Saldo inicial sin TC: se usó TC=1 (revisar).")
        lotes.append(LoteUsd(f0, float(saldo_inicial_usd), float(tc0), "Saldo inicial"))
        saldo_usd = float(saldo_inicial_usd)
        saldo_ars_costo = round(saldo_usd * float(tc0), 2)
        res.caja.append({
            "Fecha": f0.strftime("%d/%m/%Y") if f0.year > 1900 else "Saldo inicial",
            "Descripcion": "Saldo inicial caja USD",
            "Tipo_Dolar": "bancario",
            "Credito_USD": saldo_usd,
            "Debito_USD": 0.0,
            "Saldo_USD": round(saldo_usd, 2),
            "Saldo_Fisico": 0.0,
            "Saldo_Bancario": round(saldo_usd, 2),
            "TC_Operacion": float(tc0),
            "TC_Posicion": float(tc0),
            "ARS_al_TC_op": round(saldo_usd * float(tc0), 2),
            "TC_Costo_FIFO": float(tc0),
            "ARS_Costo": round(saldo_usd * float(tc0), 2),
            "Dif_Cotizacion": 0.0,
            "Saldo_ARS_Costo": round(saldo_ars_costo, 2),
            "Detalle": "Apertura",
            "Archivo origen": "",
        })

    work = df_mov.copy() if df_mov is not None and not df_mov.empty else pd.DataFrame()
    if not work.empty:
        if "_fecha" not in work.columns:
            work["_fecha"] = work["Fecha"].map(lambda x: _parse_fecha(x) or date.min)
        work["_fecha"] = pd.to_datetime(work["_fecha"], errors="coerce")
        work = work.sort_values("_fecha", kind="mergesort")

    for _, row in work.iterrows():
        fecha = row.get("_fecha")
        if isinstance(fecha, pd.Timestamp):
            fecha = fecha.date()
        elif isinstance(fecha, datetime):
            fecha = fecha.date()
        elif not isinstance(fecha, date):
            fecha = _parse_fecha(row.get("Fecha")) or date.min
        fecha_txt = fecha.strftime("%d/%m/%Y") if fecha.year > 1900 else str(row.get("Fecha") or "")
        desc = str(row.get("Descripcion") or "").strip()
        tipo = str(row.get("Tipo_Operacion") or "")
        cant = float(row.get("Cantidad") or 0)
        if cant <= 0:
            continue

        tc_op = row.get("Precio")
        if tc_op is None or (isinstance(tc_op, float) and pd.isna(tc_op)) or float(tc_op or 0) <= 0:
            tc_op = _tc_para_fecha(cotiz_map, fechas_ord, fecha)
        else:
            tc_op = float(tc_op)
        if tc_op is None or tc_op <= 0:
            tc_op = None

        ingreso = _es_ingreso(tipo, desc)
        if _norm(tipo) == "compra":
            ingreso = True
        elif _norm(tipo) == "venta":
            ingreso = False

        if ingreso:
            if tc_op is None:
                res.avisos.append(f"{fecha_txt}: ingreso USD {cant:.2f} sin TC — lote a TC=0 (completar cotizaciones).")
                tc_op = 0.0
            lotes.append(LoteUsd(fecha, cant, float(tc_op), desc or tipo))
            saldo_usd = round(saldo_usd + cant, 2)
            ars_op = round(cant * float(tc_op), 2)
            saldo_ars_costo = round(saldo_ars_costo + ars_op, 2)
            total_ing += cant
            res.caja.append({
                "Fecha": fecha_txt,
                "Descripcion": desc or tipo,
                "Tipo_Dolar": "bancario",
                "Credito_USD": cant,
                "Debito_USD": 0.0,
                "Saldo_USD": saldo_usd,
                "Saldo_Fisico": 0.0,
                "Saldo_Bancario": saldo_usd,
                "TC_Operacion": float(tc_op),
                "TC_Posicion": float(tc_op),
                "ARS_al_TC_op": ars_op,
                "TC_Costo_FIFO": float(tc_op),
                "ARS_Costo": ars_op,
                "Dif_Cotizacion": 0.0,
                "Saldo_ARS_Costo": saldo_ars_costo,
                "Detalle": "Ingreso → lote al TC de la operación",
                "Archivo origen": row.get("Archivo origen") or "",
            })
            continue

        if tc_op is None:
            res.avisos.append(f"{fecha_txt}: egreso USD {cant:.2f} sin TC del día — Dif_cotiz incompleta.")
            tc_op = 0.0

        restante = cant
        costo_ars = 0.0
        detalle_parts: list[str] = []
        tc_costo_pond = 0.0
        aplicado = 0.0

        while restante > 1e-9 and lotes:
            lote = lotes[0]
            toma = min(lote.cantidad, restante)
            costo_ars += toma * lote.tc_costo
            tc_costo_pond += toma * lote.tc_costo
            aplicado += toma
            detalle_parts.append(
                f"{toma:.4f} @ TC costo {lote.tc_costo:.4f} ({lote.fecha.strftime('%d/%m/%Y')})"
            )
            res.aplicaciones.append({
                "Fecha_egreso": fecha_txt,
                "Descripcion": desc or tipo,
                "Tipo_Dolar": "bancario",
                "USD": round(toma, 4),
                "TC_Costo": lote.tc_costo,
                "Fecha_lote": lote.fecha.strftime("%d/%m/%Y") if lote.fecha.year > 1900 else "Saldo inicial",
                "ARS_Costo": round(toma * lote.tc_costo, 2),
                "TC_Egreso": float(tc_op),
                "ARS_Egreso": round(toma * float(tc_op), 2),
                "Dif_Cotizacion": round(toma * (float(tc_op) - lote.tc_costo), 2),
            })
            lote.cantidad = round(lote.cantidad - toma, 6)
            restante = round(restante - toma, 6)
            if lote.cantidad <= 1e-9:
                lotes.popleft()

        if restante > 1e-6:
            res.avisos.append(
                f"{fecha_txt}: egreso sin stock suficiente (faltan {restante:.4f} USD)."
            )
            detalle_parts.append(f"FALTANTE {restante:.4f} USD")

        ars_egreso = round(cant * float(tc_op), 2)
        ars_costo = round(costo_ars, 2)
        dif = round(ars_egreso - ars_costo, 2)
        tc_fifo = round(tc_costo_pond / aplicado, 6) if aplicado > 0 else 0.0
        saldo_usd = round(saldo_usd - cant, 2)
        saldo_ars_costo = round(saldo_ars_costo - ars_costo, 2)
        total_egr += cant
        total_dif += dif

        res.caja.append({
            "Fecha": fecha_txt,
            "Descripcion": desc or tipo,
            "Tipo_Dolar": "bancario",
            "Credito_USD": 0.0,
            "Debito_USD": cant,
            "Saldo_USD": saldo_usd,
            "Saldo_Fisico": 0.0,
            "Saldo_Bancario": saldo_usd,
            "TC_Operacion": float(tc_op),
            "TC_Posicion": tc_fifo,
            "ARS_al_TC_op": ars_egreso,
            "TC_Costo_FIFO": tc_fifo,
            "ARS_Costo": ars_costo,
            "Dif_Cotizacion": dif,
            "Saldo_ARS_Costo": saldo_ars_costo,
            "Detalle": " | ".join(detalle_parts) if detalle_parts else "Egreso",
            "Archivo origen": row.get("Archivo origen") or "",
        })

    for lote in lotes:
        if lote.cantidad <= 1e-9:
            continue
        res.lotes_cierre.append({
            "Fecha_lote": lote.fecha.strftime("%d/%m/%Y") if lote.fecha.year > 1900 else "Saldo inicial",
            "Cantidad_USD": round(lote.cantidad, 4),
            "TC_Costo": lote.tc_costo,
            "ARS_Costo": round(lote.cantidad * lote.tc_costo, 2),
            "Origen": lote.origen,
        })

    res.resumen = {
        "Saldo_USD": round(saldo_usd, 2),
        "Saldo_Fisico": 0.0,
        "Saldo_Bancario": round(saldo_usd, 2),
        "Saldo_ARS_Costo": round(saldo_ars_costo, 2),
        "Total_Ingresos_USD": round(total_ing, 2),
        "Total_Egresos_USD": round(total_egr, 2),
        "Total_Dif_Cotizacion": round(total_dif, 2),
        "Lotes_abiertos": len(res.lotes_cierre),
        "Metodo": "fifo",
    }
    return res


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def exportar_caja_usd_excel(
    resultado: ResultadoCajaUsd,
    *,
    df_cotiz: pd.DataFrame | None = None,
    meta: dict | None = None,
) -> bytes:
    wb = Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    ws["A1"] = "Caja USD — Diferencia de cotización"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = (meta or {}).get("nota") or ""
    ws["A3"] = (
        "Criterio cta. cte.: físico/bancario separados; "
        "ingreso actualiza TC de posición (promedio); "
        "egreso → Dif = monto × (TC_mov − TC_posición). "
        "Alternativa FIFO disponible en el motor."
    )

    r = 5
    for k, v in (resultado.resumen or {}).items():
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    r += 1
    ws.cell(r, 1, "Avisos")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    for a in resultado.avisos or ["(sin avisos)"]:
        ws.cell(r, 1, a)
        r += 1

    # Caja
    ws_c = wb.create_sheet("Caja")
    cols_c = [
        "Fecha", "Descripcion", "Tipo_Dolar", "Credito_USD", "Debito_USD",
        "Saldo_USD", "Saldo_Fisico", "Saldo_Bancario",
        "TC_Operacion", "TC_Posicion", "ARS_al_TC_op", "TC_Costo_FIFO", "ARS_Costo",
        "Dif_Cotizacion", "Saldo_ARS_Costo", "Detalle", "Archivo origen",
    ]
    ws_c.append(cols_c)
    for cell in ws_c[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in resultado.caja:
        ws_c.append([row.get(c) for c in cols_c])

    # Aplicaciones (detalle egresos)
    ws_a = wb.create_sheet("Aplicaciones")
    cols_a = [
        "Fecha_egreso", "Descripcion", "Tipo_Dolar", "USD", "TC_Costo", "Fecha_lote",
        "ARS_Costo", "TC_Egreso", "ARS_Egreso", "Dif_Cotizacion",
    ]
    ws_a.append(cols_a)
    for cell in ws_a[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in resultado.aplicaciones:
        ws_a.append([row.get(c) for c in cols_a])

    # Lotes / posición cierre
    ws_l = wb.create_sheet("Posicion_cierre")
    cols_l = ["Fecha_lote", "Cantidad_USD", "TC_Costo", "ARS_Costo", "Origen"]
    ws_l.append(cols_l)
    for cell in ws_l[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in resultado.lotes_cierre:
        ws_l.append([row.get(c) for c in cols_l])

    # Cotizaciones usadas
    ws_t = wb.create_sheet("Cotizaciones")
    ws_t.append(["Fecha", "TC"])
    for cell in ws_t[1]:
        cell.fill = header_fill
        cell.font = header_font
    if df_cotiz is not None and not df_cotiz.empty:
        for _, row in df_cotiz.iterrows():
            f = row["Fecha"]
            if isinstance(f, datetime):
                f = f.date()
            ws_t.append([f, float(row["TC"])])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter = get_column_letter(col[0].column)
            maxlen = 0
            for cell in col[:80]:
                maxlen = max(maxlen, len(str(cell.value or "")))
            sheet.column_dimensions[letter].width = min(max(maxlen + 2, 10), 45)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
