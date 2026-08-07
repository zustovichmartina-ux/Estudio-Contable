# -*- coding: utf-8 -*-
"""
Script para procesar todos los PDFs de prestamos y generar el Excel de auditoria.
Ejecutar: python generar_auditoria.py
"""
import io
import os
import re
import sys
import shutil
import threading

# Forzar UTF-8 en la salida de consola (Windows cp1252 falla con caracteres especiales).
# Con pythonw / servicio oculto sys.stdout puede ser None.
if sys.stdout is not None and getattr(sys.stdout, "encoding", None):
    if str(sys.stdout.encoding).lower() != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
if sys.stderr is not None and getattr(sys.stderr, "encoding", None):
    if str(sys.stderr.encoding).lower() != "utf-8":
        try:
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
from collections import defaultdict
from datetime import date
from pathlib import Path

import fitz
import numpy as np
import pdfplumber
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)

BASE = Path(__file__).parent
CARPETA = BASE / "extractos bancarios" / "Prestamos Financieros"

# ── helpers de limpieza ──────────────────────────────────────────────────────

def _limpiar_monto(s):
    """Convierte string de monto a float.
    Maneja: formato argentino (1.234,56), US (1,234.56), OCR-mangled (1.234.56),
    montos con espacios (24 .800 .000 , 00), errores OCR o/l por 0/1.
    Estrategia: el ULTIMO separador (. o ,) es el decimal; el resto son miles.
    """
    if not s:
        return 0.0
    s = str(s).strip().replace("$", "").replace(" ", "")
    # Corregir errores OCR comunes: letra o/O por 0, l/| por 1
    s = s.replace("o", "0").replace("O", "0").replace("l", "1").replace("|", "1")
    # Solo debe contener digitos y separadores
    if not re.match(r"^[\d.,]+$", s):
        s = re.sub(r"[^0-9.,]", "", s)
    if not s:
        return 0.0
    # Buscar el ultimo separador
    last_dot = s.rfind(".")
    last_comma = s.rfind(",")
    last_sep = max(last_dot, last_comma)
    if last_sep == -1:
        # Sin separadores: entero puro
        try:
            return float(s)
        except ValueError:
            return 0.0
    decimal_part = s[last_sep + 1:]
    integer_part = s[:last_sep]
    if re.match(r"^\d{1,2}$", decimal_part):
        # Decimal de 1-2 digitos: correcto
        int_clean = re.sub(r"[.,]", "", integer_part)
        s = (int_clean or "0") + "." + decimal_part
    else:
        # Decimal con 3+ digitos: no es decimal, remover todos los separadores
        s = re.sub(r"[.,]", "", s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parsear_fecha(txt):
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d"):
        try:
            from datetime import datetime
            return datetime.strptime(txt.strip(), fmt).date()
        except (ValueError, AttributeError):
            pass
    return None


# ── detección de banco desde nombre de archivo ───────────────────────────────

def _detectar_banco(nombre):
    n = nombre.lower()
    if "santander" in n:
        return "Banco Santander"
    if "mercado pago" in n or "mercado_pago" in n:
        return "Mercado Pago"
    if "provincia" in n:
        return "Banco Provincia"
    if "nacion" in n or "nación" in n:
        return "Banco Nación"
    if "galicia" in n:
        return "Banco Galicia"
    if "frances" in n or "francés" in n or "bbva" in n:
        return "Banco Francés"
    return "Banco Desconocido"


def _detectar_id_prestamo(nombre, texto=""):
    # Buscar en texto primero
    for patron in [
        r"[Pp]r[eé]stamo\s*[Nn][°o]?\s*:?\s*([\w][\w\-/]+)",
        r"[Nn]ro\.?\s*[Pp]r[eé]stamo\s*:?\s*([\w\-]+)",
        r"[Oo]peraci[oó]n\s*[Nn][°o]?\s*:?\s*([\w\-/]+)",
        r"N[°º]\s*(\d{5,})",
        r"Contrato\s*:?\s*([\w\-]+)",
    ]:
        m = re.search(patron, texto)
        if m:
            c = m.group(1).strip().rstrip(".,;:")
            if len(c) >= 3 and re.search(r"\d", c):
                return c

    # Desde nombre de archivo: números entre paréntesis
    m = re.search(r"\((\d{4,})\)", nombre)
    if m:
        return m.group(1)
    # Último grupo de dígitos largo
    m = re.search(r"(\d{6,})", nombre)
    if m:
        return m.group(1)
    stem = Path(nombre).stem
    return stem


# ── extracción de texto por método ──────────────────────────────────────────

def _texto_nativo(pdf_path):
    textos = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            textos.append(t)
    return textos


_OCR_READER = None  # cache global compartido entre threads
_OCR_READER_LOCK = threading.Lock()  # protege la inicialización
_OCR_INFER_LOCK = threading.Lock()   # serializa llamadas a readtext (PyTorch CPU)


def _get_ocr_reader():
    global _OCR_READER
    if _OCR_READER is None:
        with _OCR_READER_LOCK:
            if _OCR_READER is None:
                import easyocr
                print("[INFO] Inicializando EasyOCR reader...", flush=True)
                _OCR_READER = easyocr.Reader(["es", "en"], gpu=False, verbose=False)
    return _OCR_READER


def _ocr_fitz(pdf_path, rotar_grados=0, dpi=150):
    """OCR con EasyOCR a traves de imagenes fitz. rotar_grados se aplica a la imagen."""
    lector = _get_ocr_reader()
    doc = fitz.open(str(pdf_path))
    textos = []
    for page in doc:
        pix = page.get_pixmap(dpi=dpi)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        if rotar_grados != 0:
            img = img.rotate(rotar_grados, expand=True)
        with _OCR_INFER_LOCK:
            resultados = lector.readtext(np.array(img))
        # Reconstruir lineas por coordenada Y (bucket de 15px)
        filas = defaultdict(list)
        for bbox, txt, _ in resultados:
            y = int((bbox[0][1] + bbox[2][1]) / 2 / 15) * 15
            filas[y].append((bbox[0][0], txt))
        lineas = [
            " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0]))
            for y in sorted(filas)
        ]
        textos.append("\n".join(lineas))
    return textos


# ── parsers por banco ─────────────────────────────────────────────────────────

def _parsear_frances_cuotas(textos):
    """Banco Frances/BBVA.
    Formato nuevo (_Cuotas): 'N VTO.  $ CAPITAL  $ INTERES  $ SEGUROS  $ IMPUESTOS  $ IMPORTE  $ SALDO'
    Fecha usa guiones: 30-08-2025
    Formato viejo (simulacion): 'N dd/mm/yy  AMORTIZAC  INTERESES  SEGURO  IMPUESTO  CUOTA  SALDO'
    Numeros con coma como miles: 2,756,034.83
    """
    cuotas = []

    # Formato nuevo: $ antes de cada monto, fecha con guiones o barras, 6 montos
    pat_nuevo = re.compile(
        r"(\d{1,3})\s+"
        r"(\d{2}[-/]\d{2}[-/]\d{4})\s+"
        r"\$\s*([\d\.,]+)\s+"    # CAPITAL
        r"\$\s*([\d\.,]+)\s+"    # INTERES
        r"\$\s*([\d\.,]+)\s+"    # SEGUROS (puede ser 0,00)
        r"\$\s*([\d\.,]+)\s+"    # IMPUESTOS
        r"\$\s*([\d\.,]+)"       # IMPORTE total
    )
    for texto in textos:
        for m in pat_nuevo.finditer(texto):
            n = int(m.group(1))
            fecha = _parsear_fecha(m.group(2))
            if not fecha or n > 600:
                continue
            capital = _limpiar_monto(m.group(3))
            intereses = _limpiar_monto(m.group(4))
            impuestos = _limpiar_monto(m.group(6))
            importe = _limpiar_monto(m.group(7))
            if importe <= 0:
                importe = round(capital + intereses + impuestos, 2)
            cuotas.append({
                "cuota": n, "vencimiento": str(fecha),
                "capital": capital, "intereses": intereses,
                "iva_gastos": impuestos, "monto_abonar": importe, "saldo_restante": 0.0
            })

    # Formato viejo simulacion BBVA: N dd/mm/yy  AMORTIZAC  INTERESES  SEGURO  IMPUESTO  CUOTA  SALDO
    if not cuotas:
        pat_viejo = re.compile(
            r"^\s*(\d{1,3})\s+"
            r"(\d{2}/\d{2}/\d{2,4})\s+"
            r"([\d,]+\.\d{2})\s+"    # AMORTIZACION (capital)
            r"([\d,]+\.\d{2})\s+"    # INTERESES
            r"([\d,]+\.\d{2})\s+"    # SEGURO
            r"([\d,]+\.\d{2})\s+"    # IMPUESTO
            r"([\d,]+\.\d{2})",      # CUOTA (total)
            re.MULTILINE
        )
        for texto in textos:
            for m in pat_viejo.finditer(texto):
                n = int(m.group(1))
                fecha = _parsear_fecha(m.group(2))
                if not fecha or n > 600:
                    continue
                capital = _limpiar_monto(m.group(3))
                intereses = _limpiar_monto(m.group(4))
                impuesto = _limpiar_monto(m.group(6))
                total = _limpiar_monto(m.group(7))
                if capital <= 0:
                    continue
                cuotas.append({
                    "cuota": n, "vencimiento": str(fecha),
                    "capital": capital, "intereses": intereses,
                    "iva_gastos": impuesto, "monto_abonar": total, "saldo_restante": 0.0
                })
    return cuotas


def _parsear_galicia(textos):
    """Banco Galicia: 'N [estado] YYYY-MM-DD $ monto_total $ capital $ interes ...'
    Montos sin separador de miles, punto como decimal: 1558534.38
    """
    cuotas = []
    # Estado puede ser 'A Vencer', 'Vencida', 'Cobrada', etc.
    pat = re.compile(
        r"(\d{1,2})\s+"
        r"[A-Za-z][A-Za-z ]*?\s+"          # estado (greedy via backtrack)
        r"(\d{4}-\d{2}-\d{2}|\d{2}[-/]\d{2}[-/]\d{4})\s+"
        r"\$\s*([\d\.,]+)\s+"              # Monto total
        r"\$\s*([\d\.,]+)\s+"              # Capital
        r"\$\s*([\d\.,]+)"                 # Interes nominal
    )
    for texto in textos:
        for m in pat.finditer(texto):
            n = int(m.group(1))
            fecha = _parsear_fecha(m.group(2))
            if not fecha or n > 600:
                continue
            total = _limpiar_monto(m.group(3))
            capital = _limpiar_monto(m.group(4))
            intereses = _limpiar_monto(m.group(5))
            if capital <= 0 and total <= 0:
                continue
            cuotas.append({
                "cuota": n, "vencimiento": str(fecha),
                "capital": capital, "intereses": intereses,
                "iva_gastos": 0.0, "monto_abonar": total, "saldo_restante": 0.0
            })
    return cuotas


def _parsear_mercadopago(textos):
    """Mercado Pago - tabla: CUOTA VENCIMIENTO CAPITAL INTERES IVA MONTO_A_ABONAR SALDO
    Formato: '1 08/05/2025 $ 388141,79 $ 920547,94 $ 193315,07 $ 1502004,80 $ 13611858,21'
    """
    cuotas = []
    texto_completo = "\n".join(textos)

    # Patron completo con los 5 montos (capital, interes, iva, monto_abonar, saldo)
    pat = re.compile(
        r"^(\d{1,3})\s+"
        r"(\d{2}/\d{2}/\d{4})\s+"
        r"\$?\s*([\d\.,]+)\s+"    # CAPITAL
        r"\$?\s*([\d\.,]+)\s+"    # INTERES
        r"\$?\s*([\d\.,]+)\s+"    # IVA
        r"\$?\s*([\d\.,]+)\s+"    # MONTO A ABONAR
        r"\$?\s*([\d\.,]+)",      # SALDO
        re.MULTILINE
    )
    for m in pat.finditer(texto_completo):
        n = int(m.group(1))
        fecha = _parsear_fecha(m.group(2))
        if not fecha or n > 600:
            continue
        capital = _limpiar_monto(m.group(3))
        intereses = _limpiar_monto(m.group(4))
        iva = _limpiar_monto(m.group(5))
        total = _limpiar_monto(m.group(6))
        saldo = _limpiar_monto(m.group(7))
        if capital <= 0 and total <= 0:
            continue
        cuotas.append({
            "cuota": n, "vencimiento": str(fecha),
            "capital": capital, "intereses": intereses,
            "iva_gastos": iva, "monto_abonar": total, "saldo_restante": saldo
        })

    # Fallback: patron con 3 montos (algunos PDFs de cuota unica)
    if not cuotas:
        pat3 = re.compile(
            r"^(\d{1,3})\s+"
            r"(\d{2}/\d{2}/\d{4})\s+"
            r"\$?\s*([\d\.,]+)\s+"
            r"\$?\s*([\d\.,]+)\s+"
            r"\$?\s*([\d\.,]+)",
            re.MULTILINE
        )
        for m in pat3.finditer(texto_completo):
            n = int(m.group(1))
            fecha = _parsear_fecha(m.group(2))
            if not fecha or n > 600:
                continue
            capital = _limpiar_monto(m.group(3))
            intereses = _limpiar_monto(m.group(4))
            total = _limpiar_monto(m.group(5))
            if capital <= 0 and total <= 0:
                continue
            cuotas.append({
                "cuota": n, "vencimiento": str(fecha),
                "capital": capital, "intereses": intereses,
                "iva_gastos": 0.0, "monto_abonar": total, "saldo_restante": 0.0
            })
    return cuotas


def _parsear_generico_ocr(textos):
    """Parser fallback: busca lineas con fecha + montos numericos.
    Sirve para cualquier banco cuyo OCR no produce el formato esperado.
    """
    cuotas = []
    n = 0
    pat_fecha = re.compile(r"\b(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})\b")
    # Monto: 1+ digitos, opcionalmente grupos de miles, luego separador + 2 digitos decimales
    pat_monto = re.compile(r"(?<!\d)[\d]+(?:[.,]\d{3})*[.,]\d{2}(?!\d)")

    for texto in textos:
        for linea in texto.splitlines():
            fechas = pat_fecha.findall(linea)
            if not fechas:
                continue
            montos_raw = pat_monto.findall(linea)
            importes = []
            for mr in montos_raw:
                v = _limpiar_monto(mr)
                if v >= 100:
                    importes.append(v)
            if not importes:
                continue
            fecha = _parsear_fecha(fechas[0])
            if not fecha:
                continue
            n += 1
            capital = importes[0]
            intereses = importes[1] if len(importes) > 1 else 0.0
            total = importes[-1] if len(importes) >= 3 else round(capital + intereses, 2)
            cuotas.append({
                "cuota": n, "vencimiento": str(fecha),
                "capital": capital, "intereses": intereses,
                "iva_gastos": 0.0, "monto_abonar": total, "saldo_restante": 0.0
            })
    return cuotas


def _parsear_fecha_flexible(txt):
    """Parsea fecha tolerando errores OCR como ' o * , en lugar de /."""
    # Reemplazar separadores OCR incorrectos
    txt_clean = re.sub(r"(?<=\d)['\*,](?=\d)", "/", txt)
    txt_clean = re.sub(r"(?<=\d)\.(?=\d{4})", "/", txt_clean)
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", txt_clean)
    if m:
        return _parsear_fecha(m.group(1))
    return None


def _parsear_santander_ocr(textos):
    """Santander OCR: tabla 'Fecha Vto | Capital | Cuota | Total | Saldo'.
    Maneja: filas divididas en 2 lineas, fechas con errores OCR, montos con
    todos puntos (ej. 3.006.617.31), montos con letras OCR (0.0o).
    """
    cuotas = []
    n = 0
    en_tabla = False
    pat_monto = re.compile(r"(?<!\d)[\d]+(?:[.,]\d{3})*[.,]\d{2}(?!\d)")

    for texto in textos:
        lineas = texto.splitlines()
        i = 0
        while i < len(lineas):
            linea = lineas[i]
            low = linea.lower()
            # Activar tabla
            if not en_tabla:
                if any(k in low for k in ("fecha", "vencim", "capital", "desarrollo", "cuota")):
                    en_tabla = True
                i += 1
                continue
            # Buscar fecha en linea (flexible ante errores OCR)
            fecha = _parsear_fecha_flexible(linea)
            if not fecha:
                i += 1
                continue
            # Combinar con linea siguiente si no tiene suficientes montos
            linea_ext = linea
            if i + 1 < len(lineas):
                siguiente = lineas[i + 1]
                # Agregar linea siguiente solo si no tiene fecha propia
                if not _parsear_fecha_flexible(siguiente):
                    linea_ext = linea + " " + siguiente
            nums = pat_monto.findall(linea_ext)
            importes = [_limpiar_monto(x) for x in nums if _limpiar_monto(x) >= 100]
            if not importes:
                i += 1
                continue
            n += 1
            capital = importes[0]
            intereses = importes[1] if len(importes) > 1 else 0.0
            total = importes[2] if len(importes) > 2 else round(capital + intereses, 2)
            saldo = importes[3] if len(importes) > 3 else 0.0
            cuotas.append({
                "cuota": n, "vencimiento": str(fecha),
                "capital": capital, "intereses": intereses,
                "iva_gastos": 0.0, "monto_abonar": total, "saldo_restante": saldo
            })
            i += 1

    if not cuotas:
        cuotas = _parsear_generico_ocr(textos)
    return cuotas


def _parsear_provincia_ocr(textos):
    """Provincia: Listado Historico del Prestamo (rotado 90 grados).
    Formato OCR: cada cuota tiene DOS filas (capital + interes), separadas pero con la misma fecha.
    Soporta fechas en formato ISO (YYYY-MM-DD), DD/MM/YYYY, DD-MM-YYYY y YYYYMMDD.
    Amounts: espacios internos ej. '717666 , 66' -> normalizar primero.
    """
    from datetime import date as date_type

    # Debug: imprimir primeras 20 lineas si DEBUG_PROVINCIA=1
    if os.environ.get("DEBUG_PROVINCIA") == "1":
        print("[DEBUG PROVINCIA] Primeras 20 lineas del texto OCR:")
        todas = "\n".join(textos).splitlines()
        for i, l in enumerate(todas[:20], 1):
            print(f"  {i:02d}: {l}")

    # Paso 1: normalizar espacios dentro de numeros y fechas en todo el texto
    textos_n = []
    for texto in textos:
        t = re.sub(r"(\d)\s+([,.])\s*(\d)", r"\1\2\3", texto)
        t = re.sub(r"([,.])\s+(\d)", r"\1\2", t)
        # Normalizar fechas ISO con espacio extra: "2025-03 -28" → "2025-03-28"
        t = re.sub(r"(\d{4}-\d{2})\s+-(\d{2})", r"\1-\2", t)
        # Normalizar "2025 03-28" (espacio en primer sep) → "2025-03-28"
        t = re.sub(r"\b(20\d{2})\s+(0[1-9]|1[0-2])-(\d{2})\b", r"\1-\2-\3", t)
        textos_n.append(t)

    pat_monto = re.compile(r"(?<!\d)[\d]+(?:[.,]\d{3})*[.,]\d{2}(?!\d)")

    def _buscar_fecha(linea):
        """Busca fecha en la linea en múltiples formatos. Retorna date o None."""
        # ISO: YYYY-MM-DD o YYYY MM-DD (con espacio como primer sep)
        m = re.search(r"\b(\d{4})[\s\-](\d{2})[\s\-](\d{2})\b", linea)
        if m:
            try:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if 2000 <= y <= 2040 and 1 <= mo <= 12 and 1 <= d <= 31:
                    return date_type(y, mo, d)
            except ValueError:
                pass
        # DD/MM/YYYY o DD-MM-YYYY
        m = re.search(r"\b(\d{2})[/\-](\d{2})[/\-](\d{4})\b", linea)
        if m:
            try:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
                if 2000 <= y <= 2040 and 1 <= mo <= 12 and 1 <= d <= 31:
                    return date_type(y, mo, d)
            except ValueError:
                pass
        # YYYYMMDD sin separador (solo si 8 digitos exactos delimitados)
        m = re.search(r"(?<!\d)(20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)", linea)
        if m:
            try:
                return date_type(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
        return None

    # Paso 2: agrupar pagos por fecha
    pagos_por_fecha = defaultdict(list)

    for texto in textos_n:
        for linea in texto.splitlines():
            fecha = _buscar_fecha(linea)
            if not fecha:
                continue
            # Extraer importes de esta fila
            nums = pat_monto.findall(linea)
            importes = [_limpiar_monto(x) for x in nums if _limpiar_monto(x) >= 100]
            if not importes:
                continue
            pagos_por_fecha[fecha].append(importes)

    # Paso 3: por cada fecha, combinar las filas para armar la cuota
    cuotas = []
    for i, fecha in enumerate(sorted(pagos_por_fecha.keys()), 1):
        filas = pagos_por_fecha[fecha]
        # Buscar el mayor importe de cada fila como "total de esa fila"
        totales_filas = [max(f) for f in filas]
        totales_filas_filtrados = [v for v in totales_filas if v >= 100]

        if not totales_filas_filtrados:
            continue

        # Si hay 2 filas: una es capital, otra es interes
        # La fila con mayor monto suele ser el capital
        if len(totales_filas_filtrados) >= 2:
            totales_ord = sorted(totales_filas_filtrados, reverse=True)
            capital = totales_ord[0]
            intereses = totales_ord[1]
        elif len(totales_filas_filtrados) == 1:
            # Solo interes (cuota temprana sin amortizacion de capital)
            capital = 0.0
            intereses = totales_filas_filtrados[0]
        else:
            continue

        total = round(capital + intereses, 2)
        cuotas.append({
            "cuota": i, "vencimiento": str(fecha),
            "capital": capital, "intereses": intereses,
            "iva_gastos": 0.0, "monto_abonar": total, "saldo_restante": 0.0
        })

    if not cuotas:
        cuotas = _parsear_generico_ocr(textos_n)
    return cuotas


def _parsear_nacion_ocr(textos):
    """Banco Nacion recibo: pago unico con capital e importe.
    OCR produce montos con espacios: '24 .800 .000 , 00' -> necesita normalizacion.
    """
    texto_completo = "\n".join(textos)
    capital = 0.0
    fecha_vto = None

    for linea in texto_completo.splitlines():
        low = linea.lower()
        if any(k in low for k in ("capital", "importe neto", "total db")):
            # Eliminar espacios dentro de numeros para manejar OCR: '24 .800 .000 , 00'
            linea_n = re.sub(r"(\d)\s+([.,])", r"\1\2", linea)
            linea_n = re.sub(r"([.,])\s+(\d)", r"\1\2", linea_n)
            nums = re.findall(r"(?<!\d)[\d]+(?:[.,]\d{3})*[.,]\d{2}(?!\d)", linea_n)
            candidatos = [_limpiar_monto(x) for x in nums if _limpiar_monto(x) >= 1000]
            if candidatos:
                capital = max(candidatos)
        if any(k in low for k in ("ult.vto", "ult vto", "1r.vto", "primer vto", "vencim", "fecha ir")):
            m = re.search(r"(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})", linea)
            if m:
                d = _parsear_fecha(re.sub(r"[.]", "/", m.group(1)))
                if d:
                    fecha_vto = d

    if capital > 0 and fecha_vto:
        return [{"cuota": 1, "vencimiento": str(fecha_vto),
                 "capital": capital, "intereses": 0.0,
                 "iva_gastos": 0.0, "monto_abonar": capital, "saldo_restante": 0.0}]

    cuotas = _parsear_generico_ocr(textos)
    return cuotas if cuotas else []




# ── validación y estrategia de reintentos ────────────────────────────────────

def _validar_cuotas(cuotas, banco, pid):
    """Verifica que las cuotas extraídas sean razonables.
    Retorna (True, "") si OK, o (False, motivo) si hay problema.
    """
    if len(cuotas) == 0:
        return False, "sin_cuotas"

    total = len(cuotas)

    vacios = sum(
        1 for c in cuotas
        if c.get("capital", 0) == 0 and c.get("monto_abonar", 0) == 0
    )
    if vacios / total > 0.5:
        return False, "montos_vacios"

    sin_fecha = sum(
        1 for c in cuotas
        if not c.get("vencimiento") or c.get("vencimiento") in (None, "None", "")
    )
    if sin_fecha / total > 0.5:
        return False, "fechas_vacias"

    if any(c.get("monto_abonar", 0) < 0 or c.get("capital", 0) < 0 for c in cuotas):
        return False, "montos_negativos"

    return True, ""


def _ajustar_estrategia(motivo_error, intento):
    """Retorna un dict de parámetros de extracción según el motivo del intento previo.
    intento es 0-based; intento >= 2 activa la estrategia máxima (último).
    """
    if intento >= 2:
        return {"forzar_ocr": True, "rotacion": 270}

    tabla = {
        ("sin_cuotas",      1): {"forzar_ocr": True, "rotacion": 0},
        ("sin_cuotas",      2): {"forzar_ocr": True, "rotacion": 90},
        ("montos_vacios",   1): {"umbral_confianza_ocr": 0.2},
        ("montos_vacios",   2): {"forzar_ocr": True, "rotacion": 90},
        ("fechas_vacias",   1): {"forzar_ocr": True, "rotacion": 0},
        ("fechas_vacias",   2): {"forzar_ocr": True, "rotacion": 180},
        ("montos_negativos",1): {"forzar_ocr": True, "rotacion": 0},
    }
    return tabla.get((motivo_error, intento), {})


def _extraer_cuotas(ruta_pdf, banco, forzar_ocr=False, rotacion_override=None):
    """Extrae cuotas de un PDF aplicando la estrategia indicada.

    Si forzar_ocr=False se intenta primero extracción nativa; si el PDF no tiene
    texto nativo suficiente se cae automáticamente a OCR.
    rotacion_override sobreescribe la rotación por defecto por banco.
    """
    if not forzar_ocr:
        textos_nativos = _texto_nativo(ruta_pdf)
        tiene_texto = any(len(t.strip()) > 50 for t in textos_nativos)
        if tiene_texto:
            if banco == "Banco Galicia":
                return _parsear_galicia(textos_nativos)
            elif banco == "Banco Francés":
                return _parsear_frances_cuotas(textos_nativos)
            elif banco == "Mercado Pago":
                return _parsear_mercadopago(textos_nativos)
            elif banco == "Banco Santander":
                return _parsear_santander_ocr(textos_nativos)
            elif banco == "Banco Nación":
                return _parsear_nacion_ocr(textos_nativos)
            else:
                return _parsear_mercadopago(textos_nativos)
        # Sin texto nativo suficiente → caer a OCR

    rotacion = rotacion_override if rotacion_override is not None else (
        90 if banco == "Banco Provincia" else 0
    )
    dpi_ocr = 200 if banco == "Banco Provincia" else 150
    print(f"    [OCR] rotacion={rotacion}° dpi={dpi_ocr}...", flush=True)
    textos_ocr = _ocr_fitz(ruta_pdf, rotar_grados=rotacion, dpi=dpi_ocr)

    if banco == "Banco Provincia":
        return _parsear_provincia_ocr(textos_ocr)
    elif banco == "Banco Santander":
        return _parsear_santander_ocr(textos_ocr)
    elif banco == "Banco Nación":
        return _parsear_nacion_ocr(textos_ocr)
    elif banco == "Banco Galicia":
        return _parsear_galicia(textos_ocr)
    elif banco == "Banco Francés":
        return _parsear_frances_cuotas(textos_ocr)
    else:
        return _parsear_mercadopago(textos_ocr)


def _procesar_un_pdf(ruta_pdf):
    """Procesa un único PDF con loop de auto-validación y reintentos.

    Retorna: {"banco": str, "pid": str, "cuotas": list,
              "capital_original_extra": float, "error": str}
    Compatible con ThreadPoolExecutor (función de módulo, sin closures).
    """
    ruta_pdf = Path(ruta_pdf)
    banco = _detectar_banco(ruta_pdf.name)
    pid = _detectar_id_prestamo(ruta_pdf.name)
    INTENTOS = 3
    motivo_error = ""
    capital_original_extra = 0.0

    # Pre-check Banco Francés: puede ser comprobante de alta (texto nativo con capital, sin cuotas)
    if banco == "Banco Francés":
        try:
            textos_nativos = _texto_nativo(ruta_pdf)
            texto_concat = "\n".join(textos_nativos)
            pid_texto = _detectar_id_prestamo(ruta_pdf.name, texto_concat)
            pid = pid_texto or pid
            m_cap = re.search(r"CAPITAL\s+([\d\.,]+)", texto_concat)
            if not m_cap:
                m_cap = re.search(r"IMPORTE CONCEDIDO\s*:\s*([\d\.,]+)", texto_concat)
            if m_cap:
                cap_val = _limpiar_monto(m_cap.group(1))
                if cap_val > 0:
                    capital_original_extra = cap_val
        except Exception:
            pass

    for intento in range(INTENTOS):
        params = _ajustar_estrategia(motivo_error if intento > 0 else "", intento)
        forzar_ocr = params.get("forzar_ocr", False)
        rotacion = params.get("rotacion", None)

        cuotas = _extraer_cuotas(
            ruta_pdf, banco,
            forzar_ocr=forzar_ocr,
            rotacion_override=rotacion,
        )
        es_valido, motivo_error = _validar_cuotas(cuotas, banco, pid)

        if es_valido:
            print(f"  [OK] {ruta_pdf.name} | Banco: {banco} | Cuotas: {len(cuotas)}")
            return {
                "banco": banco, "pid": pid, "cuotas": cuotas,
                "capital_original_extra": capital_original_extra, "error": "",
            }

        # Francés alta: sin cuotas pero capital conocido → no tiene sentido reintentar OCR
        if banco == "Banco Francés" and capital_original_extra > 0 and intento == 0:
            print(f"  [Francés alta] {ruta_pdf.name} | capital={capital_original_extra:,.0f}")
            return {
                "banco": banco, "pid": pid, "cuotas": [],
                "capital_original_extra": capital_original_extra, "error": "",
            }

        if intento < INTENTOS - 1:
            print(
                f"  [REINTENTO {intento + 1}/{INTENTOS}] "
                f"{ruta_pdf.name} | Motivo: {motivo_error}"
            )

    print(f"  [FALLO] {ruta_pdf.name} | Sin datos tras {INTENTOS} intentos")
    return {
        "banco": banco, "pid": pid, "cuotas": [],
        "capital_original_extra": capital_original_extra, "error": motivo_error,
    }


# ── procesamiento principal ──────────────────────────────────────────────────

def procesar_todos(carpeta: Path):
    """Procesa todos los PDFs en paralelo y retorna dict banco → lista de préstamos."""
    from concurrent.futures import ThreadPoolExecutor
    import multiprocessing

    pdfs = sorted(carpeta.glob("*.pdf"))
    print(f"PDFs encontrados: {len(pdfs)}")

    workers = max(1, min(multiprocessing.cpu_count() - 1, 4))
    print(f"[INFO] Procesando {len(pdfs)} PDFs en paralelo con {workers} workers...", flush=True)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        resultados_raw = list(executor.map(_procesar_un_pdf, pdfs))

    # Consolidar resultados en el dict prestamos_por_banco
    prestamos_por_banco = defaultdict(dict)  # banco → {id_prestamo: {"cuotas": [], ...}}

    for res in resultados_raw:
        banco = res["banco"]
        pid = res["pid"]
        cuotas = res["cuotas"]
        capital_original_extra = res["capital_original_extra"]

        if pid not in prestamos_por_banco[banco]:
            prestamos_por_banco[banco][pid] = {
                "prestamo_n": pid, "capital_original": 0.0,
                "sistema": "Francés", "cuotas": []
            }

        # Comprobante de alta francés: sólo actualiza el capital
        if capital_original_extra > 0:
            prestamos_por_banco[banco][pid]["capital_original"] = capital_original_extra

        # Deduplicar cuotas por número
        existentes = {c["cuota"] for c in prestamos_por_banco[banco][pid]["cuotas"]}
        for c in cuotas:
            if c["cuota"] not in existentes:
                prestamos_por_banco[banco][pid]["cuotas"].append(c)
                existentes.add(c["cuota"])

        # Estimar capital original desde cuotas si todavía es 0
        if cuotas and prestamos_por_banco[banco][pid]["capital_original"] == 0.0:
            saldo_max = max((c.get("saldo_restante", 0) for c in cuotas), default=0)
            cap_sum = sum(c.get("capital", 0) for c in cuotas)
            prestamos_por_banco[banco][pid]["capital_original"] = saldo_max or cap_sum

    # Convertir a formato esperado por generar_excel
    resultado = {}
    for banco, prest_dict in prestamos_por_banco.items():
        resultado[banco] = list(prest_dict.values())
    return resultado


# ── generación Excel ──────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _alin(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)

COLS = ["CUOTA", "VENCIMIENTO", "CAPITAL", "INTERESES", "IVA/GASTOS", "MONTO A ABONAR", "SALDO RESTANTE"]
WIDTHS = [8, 14, 16, 16, 14, 18, 18]
MONEDA = '"$ "#,##0.00'
FILL_HEADER_PRESTAMO = _fill("B8B8B8")
FILL_RESUMEN_TITLE = _fill("1F4E79")
FILL_GRID_HEADER = _fill("2E75B6")
FILL_ALT = _fill("F2F2F2")
FILL_CIERRE_TITLE = _fill("1F4E79")
FILL_AMARILLO = _fill("FFFF00")
FILL_ROJO = _fill("FF0000")


def _escribir_prestamo(ws, fila_inicio, prestamo, banco):
    f = fila_inicio
    n_cols = len(COLS)
    last_col = chr(ord("A") + n_cols - 1)

    # A) Cabecera del préstamo
    cell = ws.cell(f, 1)
    ws.merge_cells(f"A{f}:{last_col}{f}")
    capital_fmt = f"$ {prestamo['capital_original']:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    cell.value = f"PRÉSTAMO N° {prestamo['prestamo_n']} | Capital: {capital_fmt} | Banco: {banco}"
    cell.fill = FILL_HEADER_PRESTAMO
    cell.font = _font(bold=True, size=11)
    cell.alignment = _alin("center")
    ws.row_dimensions[f].height = 22
    f += 1

    cuotas = sorted(prestamo["cuotas"], key=lambda x: x.get("cuota", 0))
    # Calcular totales anuales
    total_cap = sum(c.get("capital", 0) for c in cuotas)
    total_int = sum(c.get("intereses", 0) for c in cuotas)
    total_iva = sum(c.get("iva_gastos", 0) for c in cuotas)

    # B) Bloque resumen anual
    ws.merge_cells(f"A{f}:{last_col}{f}")
    c = ws.cell(f, 1)
    c.value = "RESUMEN ANUAL"
    c.fill = FILL_RESUMEN_TITLE
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _alin("center")
    f += 1

    labels = ["Total Capital Amortizado", "Total Intereses Devengados", "Total IVA/Gastos"]
    values = [total_cap, total_int, total_iva]
    for col, (lbl, val) in enumerate(zip(labels, values), 1):
        cl = ws.cell(f, col)
        cl.value = lbl
        cl.font = _font(bold=True)
        cl.fill = _fill("D9E1F2")
        cl.alignment = _alin("center")
    f += 1
    for col, val in enumerate(values, 1):
        cv = ws.cell(f, col)
        cv.value = val
        cv.number_format = MONEDA
        cv.font = _font(bold=True)
        cv.fill = _fill("EBF3FB")
        cv.alignment = _alin("right")
    f += 1

    # C) Encabezado grilla
    for col, lbl in enumerate(COLS, 1):
        ch = ws.cell(f, col)
        ch.value = lbl
        ch.fill = FILL_GRID_HEADER
        ch.font = _font(bold=True, color="FFFFFF", size=11)
        ch.alignment = _alin("center")
        ch.border = _border()
    f += 1

    # D) Filas de cuotas
    for i, cuota in enumerate(cuotas):
        fill = FILL_ALT if i % 2 else _fill("FFFFFF")
        vals = [
            cuota.get("cuota", i + 1),
            cuota.get("vencimiento", ""),
            cuota.get("capital", 0.0),
            cuota.get("intereses", 0.0),
            cuota.get("iva_gastos", 0.0),
            cuota.get("monto_abonar", 0.0),
            cuota.get("saldo_restante", 0.0),
        ]
        for col, val in enumerate(vals, 1):
            cd = ws.cell(f, col)
            cd.value = val
            cd.fill = fill
            cd.border = _border()
            cd.font = _font(size=10)
            if col >= 3:
                cd.number_format = MONEDA
                cd.alignment = _alin("right")
            else:
                cd.alignment = _alin("left" if col == 2 else "center")
        ws.row_dimensions[f].height = 15
        f += 1

    return f + 4  # 4 filas en blanco entre préstamos


def _escribir_cierre(ws, fila, prestamos, saldo_inicial):
    last_col = chr(ord("A") + len(COLS) - 1)
    f = fila + 1

    ws.merge_cells(f"A{f}:{last_col}{f}")
    c = ws.cell(f, 1)
    c.value = "CONCILIACIÓN CONTABLE FINAL"
    c.fill = FILL_CIERRE_TITLE
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = _alin("center")
    f += 1

    total_amort = sum(
        sum(q.get("capital", 0) for q in p["cuotas"])
        for p in prestamos
    )
    saldo_final = saldo_inicial - total_amort
    filas_cierre = [
        ("Saldo Inicial del Banco", saldo_inicial),
        ("Total Capital Amortizado (todos los préstamos)", total_amort),
        ("Saldo Final Sugerido Mayor Contable", saldo_final),
    ]
    for concepto, valor in filas_cierre:
        ws.cell(f, 1).value = concepto
        ws.cell(f, 1).font = _font(bold=True)
        ws.cell(f, 2).value = valor
        ws.cell(f, 2).number_format = MONEDA
        ws.cell(f, 2).alignment = _alin("right")
        if concepto.startswith("Saldo Final"):
            fill_c = FILL_AMARILLO if saldo_final >= 0 else FILL_ROJO
            ws.cell(f, 1).fill = fill_c
            ws.cell(f, 2).fill = fill_c
        f += 1


def generar_excel(bancos_data, saldos_iniciales, ruta_salida):
    wb = Workbook()

    # Hoja resumen ejecutivo
    ws_res = wb.active
    ws_res.title = "Resumen Ejecutivo"
    ws_res.append(["Banco", "Préstamos", "Total Cuotas", "Capital Total"])
    for banco, prestamos in bancos_data.items():
        n_prest = len(prestamos)
        n_cuotas = sum(len(p["cuotas"]) for p in prestamos)
        cap_total = sum(p.get("capital_original", 0) for p in prestamos)
        ws_res.append([banco, n_prest, n_cuotas, cap_total])

    # Una hoja por banco
    for banco, prestamos in bancos_data.items():
        nombre_hoja = banco[:31]
        ws = wb.create_sheet(title=nombre_hoja)
        for col, (lbl, w) in enumerate(zip(COLS, WIDTHS), 1):
            ws.column_dimensions[chr(ord("A") + col - 1)].width = w
        ws.sheet_view.zoomScale = 90

        fila = 1
        for prestamo in prestamos:
            fila = _escribir_prestamo(ws, fila, prestamo, banco)

        saldo = saldos_iniciales.get(banco, 0.0)
        _escribir_cierre(ws, fila, prestamos, saldo)

    wb.save(str(ruta_salida))
    print(f"\nExcel guardado: {ruta_salida} ({ruta_salida.stat().st_size // 1024} KB)")


# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("GENERADOR DE AUDITORÍA DE PRÉSTAMOS")
    print("=" * 60)

    bancos_data = procesar_todos(CARPETA)

    print("\nResumen:")
    for banco, prestamos in bancos_data.items():
        total_cuotas = sum(len(p["cuotas"]) for p in prestamos)
        print(f"  {banco}: {len(prestamos)} préstamos, {total_cuotas} cuotas")

    saldos = {b: 0.0 for b in bancos_data}
    ruta = BASE / "Auditoria_Prestamos_Completa.xlsx"
    generar_excel(bancos_data, saldos, ruta)
    print("LISTO.")
