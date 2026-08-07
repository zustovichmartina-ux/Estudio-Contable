"""
Demo v1 — Proyección de Ganancias (formato Estudio).
Lee el resumen de Gastro y genera Excel + PDF de ejemplo.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from excel_formato_estudio import COLOR_PRIMARIO, guardar_informe_excel

SRC = Path(
    r"\\TANGOSRV\Compartido\CLIENTES\GASTROENTEROLOGIA Y ENDOSCOPIA DIGESTIVA "
    r"MAR DEL PLATA S.A\Impuestos\Proyecciones de Ganancias\Proyeccion de Ganancias 2025.xlsx"
)
OUT_DIR = Path(r"C:\Users\recep\Desktop")
CLIENTE = "GASTROENTEROLOGIA Y ENDOSCOPIA DIGESTIVA MAR DEL PLATA S.A."
PERIODO = "Ejercicio 01/09/2025 al 31/08/2026"


def _money(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def leer_resumen_gastro(path: Path) -> tuple[list[str], pd.DataFrame, dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["2025-2026"]

    # Columnas mes: C..L = sep..jun (10 meses cargados); M/N jul-ago vacíos
    meses: list[str] = []
    for c in range(3, 15):
        v = ws.cell(10, c).value
        if isinstance(v, datetime):
            meses.append(v.strftime("%m/%Y"))
        elif v:
            meses.append(str(v))
        else:
            meses.append(f"mes{c}")

    def fila(r: int) -> list[float]:
        return [_money(ws.cell(r, c).value) for c in range(3, 15)]

    ingresos = [a + b + c + d for a, b, c, d in zip(fila(11), fila(12), fila(13), fila(14))]
    sueldos = [a + b for a, b in zip(fila(20), fila(21))]  # sueldos + cargas
    iibb = fila(22)
    # resto de gastos (filas 23-36) sin sueldos/cargas/iibb
    otros = [0.0] * 12
    for r in range(23, 37):
        vals = fila(r)
        for i, v in enumerate(vals):
            otros[i] += v
    gastos_tot = [s + i + o for s, i, o in zip(sueldos, iibb, otros)]
    resultado = [ing - gas for ing, gas in zip(ingresos, gastos_tot)]
    deudores = fila(51)  # ventas cobradas / mov deudores bruto
    proveedores = fila(52)

    rows = []
    for i, mes in enumerate(meses):
        if ingresos[i] == 0 and gastos_tot[i] == 0 and deudores[i] == 0:
            continue
        rows.append(
            {
                "Mes": mes,
                "Ingresos netos IVA": round(ingresos[i], 2),
                "Sueldos y cargas": round(sueldos[i], 2),
                "Impuestos (IIBB)": round(iibb[i], 2),
                "Otros gastos netos": round(otros[i], 2),
                "Total gastos netos": round(gastos_tot[i], 2),
                "Resultado proyectado": round(resultado[i], 2),
                "Deudores (bruto)": round(deudores[i], 2),
                "Proveedores (bruto)": round(proveedores[i], 2),
            }
        )

    df = pd.DataFrame(rows)
    tot_ing = float(df["Ingresos netos IVA"].sum())
    tot_sue = float(df["Sueldos y cargas"].sum())
    tot_imp = float(df["Impuestos (IIBB)"].sum())
    tot_otr = float(df["Otros gastos netos"].sum())
    tot_gas = float(df["Total gastos netos"].sum())
    kpis = {
        "ingresos": tot_ing,
        "sueldos": tot_sue,
        "impuestos": tot_imp,
        "otros": tot_otr,
        "gastos": tot_gas,
        "resultado": tot_ing - tot_gas,
        "pct_sueldos": (tot_sue / tot_ing * 100) if tot_ing else 0,
        "pct_impuestos": (tot_imp / tot_ing * 100) if tot_ing else 0,
        "pct_otros": (tot_otr / tot_ing * 100) if tot_ing else 0,
        "pct_resultado": ((tot_ing - tot_gas) / tot_ing * 100) if tot_ing else 0,
    }
    return meses, df, kpis


def armar_grafico_png(kpis: dict, path: Path) -> Path:
    labels = ["Sueldos", "Impuestos", "Otros gastos", "Resultado"]
    vals = [kpis["pct_sueldos"], kpis["pct_impuestos"], kpis["pct_otros"], kpis["pct_resultado"]]
    colores = ["#5B9BD5", "#ED7D31", "#A5A5A5", "#70AD47"]

    fig, ax = plt.subplots(figsize=(9, 4.2))
    bars = ax.bar(labels, vals, color=colores, width=0.62)
    ax.set_ylabel("% de los ingresos netos")
    ax.set_title("Distribución de gastos y resultado sobre ingresos netos de IVA")
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_ylim(min(0, min(vals) - 5), max(vals) + 12)
    for b, v in zip(bars, vals):
        ax.text(
            b.get_x() + b.get_width() / 2,
            v + (1.5 if v >= 0 else -4),
            f"{v:.1f}%",
            ha="center",
            va="bottom" if v >= 0 else "top",
            fontsize=10,
            fontweight="bold",
        )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=140, bbox_inches="tight")
    plt.close(fig)
    return path


def agregar_chart_excel(xlsx: Path, df: pd.DataFrame) -> None:
    """Agrega hoja 'Composición %' con % mensuales + gráfico de barras."""
    wb = load_workbook(xlsx)
    if "Composición %" in wb.sheetnames:
        del wb["Composición %"]
    ws = wb.create_sheet("Composición %", 1)

    headers = ["Mes", "% Sueldos / Ing", "% Impuestos / Ing", "% Otros / Ing", "% Resultado / Ing"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    for i, (_, row) in enumerate(df.iterrows()):
        ing = float(row["Ingresos netos IVA"]) or 1.0
        r = i + 2
        ws.cell(r, 1, row["Mes"])
        ws.cell(r, 2, float(row["Sueldos y cargas"]) / ing).number_format = "0.0%"
        ws.cell(r, 3, float(row["Impuestos (IIBB)"]) / ing).number_format = "0.0%"
        ws.cell(r, 4, float(row["Otros gastos netos"]) / ing).number_format = "0.0%"
        ws.cell(r, 5, float(row["Resultado proyectado"]) / ing).number_format = "0.0%"

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "% de ingresos netos — Sueldos / Impuestos / Otros"
    chart.y_axis.title = "% ingresos"
    data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=len(df) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "G2")

    for col, w in enumerate([12, 16, 18, 16, 18], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(xlsx)


def armar_pdf(path: Path, df: pd.DataFrame, kpis: dict, chart_png: Path) -> Path:
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "Titulo",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1F4E79"),
        spaceAfter=4,
    )
    sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#666666"))
    story = [
        Paragraph("Proyección de Ganancias — DEMO v1", title),
        Paragraph(CLIENTE, sub),
        Paragraph(PERIODO, sub),
        Spacer(1, 0.4 * cm),
    ]

    kpi_data = [
        ["Ingresos netos IVA", f"$ {kpis['ingresos']:,.0f}".replace(",", ".")],
        ["Total gastos netos", f"$ {kpis['gastos']:,.0f}".replace(",", ".")],
        ["  · Sueldos / cargas", f"$ {kpis['sueldos']:,.0f}  ({kpis['pct_sueldos']:.1f}% ing)".replace(",", ".")],
        ["  · Impuestos (IIBB)", f"$ {kpis['impuestos']:,.0f}  ({kpis['pct_impuestos']:.1f}% ing)".replace(",", ".")],
        ["  · Otros gastos", f"$ {kpis['otros']:,.0f}  ({kpis['pct_otros']:.1f}% ing)".replace(",", ".")],
        ["Resultado proyectado", f"$ {kpis['resultado']:,.0f}  ({kpis['pct_resultado']:.1f}% ing)".replace(",", ".")],
    ]
    t_kpi = Table(kpi_data, colWidths=[7 * cm, 8 * cm])
    t_kpi.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FE")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t_kpi)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Image(str(chart_png), width=16 * cm, height=7.2 * cm))
    story.append(Spacer(1, 0.5 * cm))

    # tabla mensual compacta
    headers = ["Mes", "Ingresos", "Sueldos", "Impuestos", "Otros", "Resultado", "Deudores", "Proveedores"]
    body = [headers]
    for _, row in df.iterrows():
        body.append(
            [
                row["Mes"],
                f"{row['Ingresos netos IVA']:,.0f}".replace(",", "."),
                f"{row['Sueldos y cargas']:,.0f}".replace(",", "."),
                f"{row['Impuestos (IIBB)']:,.0f}".replace(",", "."),
                f"{row['Otros gastos netos']:,.0f}".replace(",", "."),
                f"{row['Resultado proyectado']:,.0f}".replace(",", "."),
                f"{row['Deudores (bruto)']:,.0f}".replace(",", "."),
                f"{row['Proveedores (bruto)']:,.0f}".replace(",", "."),
            ]
        )
    t = Table(body, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(Paragraph("Cuadro mensual (ingresos/gastos netos de IVA; deudores/proveedores brutos)", sub))
    story.append(Spacer(1, 0.2 * cm))
    story.append(t)
    doc.build(story)
    return path


def main() -> None:
    _, df, kpis = leer_resumen_gastro(SRC)

    # resumen por rubro para hoja Resumen
    df_rubros = pd.DataFrame(
        [
            {"Concepto": "Ingresos netos de IVA", "Importe": kpis["ingresos"], "% ingresos": 100.0},
            {"Concepto": "Sueldos y cargas sociales", "Importe": kpis["sueldos"], "% ingresos": round(kpis["pct_sueldos"], 1)},
            {"Concepto": "Impuestos (IIBB)", "Importe": kpis["impuestos"], "% ingresos": round(kpis["pct_impuestos"], 1)},
            {"Concepto": "Otros gastos netos", "Importe": kpis["otros"], "% ingresos": round(kpis["pct_otros"], 1)},
            {"Concepto": "Total gastos netos", "Importe": kpis["gastos"], "% ingresos": round(kpis["gastos"] / kpis["ingresos"] * 100, 1)},
            {"Concepto": "Resultado proyectado", "Importe": kpis["resultado"], "% ingresos": round(kpis["pct_resultado"], 1)},
            {"Concepto": "Deudores (bruto) — total período", "Importe": float(df["Deudores (bruto)"].sum()), "% ingresos": None},
            {"Concepto": "Proveedores (bruto) — total período", "Importe": float(df["Proveedores (bruto)"].sum()), "% ingresos": None},
        ]
    )

    xlsx = OUT_DIR / "DEMO_Proyeccion_Ganancias_v1_Gastro.xlsx"
    png = OUT_DIR / "DEMO_Proyeccion_Ganancias_v1_grafico.png"
    pdf = OUT_DIR / "DEMO_Proyeccion_Ganancias_v1_Gastro.pdf"

    guardar_informe_excel(
        xlsx,
        titulo="Proyección de Ganancias — DEMO v1",
        subtitulo=CLIENTE,
        periodo=PERIODO,
        kpis=[
            ("Ingresos netos de IVA", kpis["ingresos"], "money"),
            ("Total gastos netos", kpis["gastos"], "money"),
            ("Resultado proyectado", kpis["resultado"], "money"),
            ("% Sueldos / ingresos", round(kpis["pct_sueldos"], 1)),
            ("% Impuestos / ingresos", round(kpis["pct_impuestos"], 1)),
            ("% Otros / ingresos", round(kpis["pct_otros"], 1)),
        ],
        resumenes=[("Composición del ejercicio", df_rubros)],
        detalle=df,
        hoja_detalle="Cuadro mensual",
        col_moneda=[
            "Ingresos netos IVA",
            "Sueldos y cargas",
            "Impuestos (IIBB)",
            "Otros gastos netos",
            "Total gastos netos",
            "Resultado proyectado",
            "Deudores (bruto)",
            "Proveedores (bruto)",
            "Importe",
        ],
    )
    agregar_chart_excel(xlsx, df)
    armar_grafico_png(kpis, png)
    armar_pdf(pdf, df, kpis, png)

    print("Excel:", xlsx)
    print("PDF  :", pdf)
    print("PNG  :", png)
    print(
        f"Ing={kpis['ingresos']:,.0f} Gas={kpis['gastos']:,.0f} "
        f"Res={kpis['resultado']:,.0f} | "
        f"Sue={kpis['pct_sueldos']:.1f}% Imp={kpis['pct_impuestos']:.1f}% "
        f"Otr={kpis['pct_otros']:.1f}%"
    )


if __name__ == "__main__":
    main()
