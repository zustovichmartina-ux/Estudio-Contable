"""Prueba rápida de clonación sin OCR (datos simulados enero 2026)."""
from datetime import date
from pathlib import Path

import openpyxl

from procesador import (
    BalanceMensual,
    MovimientoBanco,
    ResultadoConciliacion,
    _resolver_ruta_plantilla,
    generar_planilla_conciliacion,
)

PLANTILLA = _resolver_ruta_plantilla()
SALIDA = Path("exportaciones/Conciliacion_test_mock.xlsx")


def main():
    movs = [
        MovimientoBanco(date(2026, 1, 2), "001", "Pago haberes", 2000000, 0, 13033278.09),
        MovimientoBanco(date(2026, 1, 5), "002", "Debito automatico Telecom", 200721.92, 0, 12797783.76),
        MovimientoBanco(date(2026, 1, 8), "003", "Transferencia recibida", 0, 500000, 13297783.76),
    ]
    balance = BalanceMensual(
        anio=2026, mes=1,
        saldo_inicial=15037742.99,
        saldo_final=20907272.55,
        saldo_resumen=20907272.55,
        ingresos=sum(m.credito for m in movs),
        egresos=sum(m.debito for m in movs),
        balance_cierra=True,
        diferencia_balance=0.0,
        archivo_origen="mock",
    )

    resultado = ResultadoConciliacion(
        mes_referencia=date(2026, 1, 1),
        saldos_por_mes={(2026, 1): balance},
        saldo_inicial=balance.saldo_inicial,
        saldo_final=balance.saldo_final,
        saldo_extracto=balance.saldo_resumen,
        ingresos=balance.ingresos,
        egresos=balance.egresos,
        balance_cierra=True,
        movimientos_todos=movs,
        bancos_detectados=["santander"],
    )

    wb_orig = openpyxl.load_workbook(PLANTILLA)
    font_orig = wb_orig["Banco Santander"]["B57"].font.name
    wb_orig.close()

    excel = generar_planilla_conciliacion(resultado, "TEST MOCK ENERO")
    SALIDA.parent.mkdir(exist_ok=True)
    SALIDA.write_bytes(excel)

    wb = openpyxl.load_workbook(SALIDA)
    ws = wb["Banco Santander"]
    print("G57 (Saldo inicio):", ws["G57"].value, "font:", ws["G57"].font.name)
    print("G60 (Saldo final fórmula):", ws["G60"].value)
    print("G61 (Saldo resumen):", ws["G61"].value)
    print("G62 (Diferencia fórmula):", ws["G62"].value)
    print("Detalle filas:", wb["Detalle Extracto"].max_row if "Detalle Extracto" in wb.sheetnames else 0)
    print("Font match:", ws["G57"].font.name == font_orig)
    print("OK ->", SALIDA.resolve())
    wb.close()


if __name__ == "__main__":
    main()
