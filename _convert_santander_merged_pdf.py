"""Convierte extracto Santander digital (PDF merged) a Excel ordenado."""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import fitz
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

PDF = Path(r"C:\Users\recep\Downloads\2025-04_merged.pdf")
OUT = Path(r"C:\Users\recep\Downloads\Extracto_Santander_Oftalmologia_RELE_merged.xlsx")
OUT2 = Path(r"C:\Users\recep\Desktop\Extracto_Santander_Oftalmologia_RELE_merged.xlsx")

RE_FECHA = re.compile(r"\b(\d{2}/\d{2}/\d{2,4})\b")
RE_PESOS = re.compile(r"pesos\s*([\d.]+,\d{2})", re.IGNORECASE)
RE_PERIODO = re.compile(
    r"Desde:\s*(\d{2}/\d{2}/\d{2,4}).*?Hasta:\s*(\d{2}/\d{2}/\d{2,4})",
    re.IGNORECASE | re.DOTALL,
)
RE_CUIT = re.compile(r"CUIT:\s*([\d\-]+)", re.IGNORECASE)
RE_CUENTA = re.compile(r"Cuenta Corriente\s*N[º°o.\s]*([\d\-/]+)", re.IGNORECASE)
RE_CBU = re.compile(r"CBU:\s*(\d{22})", re.IGNORECASE)

PALABRAS_DEBITO = (
    "debito",
    "débito",
    "compra",
    "transferencia realizada",
    "comision",
    "comisión",
    "impuesto",
    "iva ",
    "pago de ",
    "retencion",
    "retención",
    "cargo",
    "mantenimiento",
    "interes",
    "interés",
)


def parse_fecha(txt: str) -> date | None:
    txt = txt.strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def parse_monto(txt: str) -> float:
    return float(txt.replace(".", "").replace(",", "."))


def es_ruido(linea: str) -> bool:
    n = linea.lower().strip()
    if not n:
        return True
    # Encabezados / legales exactos o casi exactos (no filtrar "Impuesto ley..." de movimientos)
    if n in {
        "fecha",
        "comprobante",
        "movimiento",
        "débito",
        "debito",
        "crédito",
        "credito",
        "saldo en cuenta",
        "movimientos en pesos",
        "resumen de cuenta",
        "* salvo error u omisión",
        "salvo error u omisión",
    }:
        return True
    ruido_prefijo = (
        "banco santander argentina",
        "movimientos en pesos",
        "saldo en cuenta",
        "cuenta corriente n",
        "período",
        "periodo",
        "emisión mensual",
        "emision mensual",
        "total en ",
        "resumen de cuenta",
        "responsable inscri",
        "unidad de información",
        "fondos comunes",
        "ningún accionista",
        "correlativo",
        "saldo total detalle impositivo",
        "detalle impositivo",
        "tipo de impuesto",
        "totales de retencion",
        "totales de retención",
    )
    if any(n.startswith(r) for r in ruido_prefijo):
        return True
    if re.fullmatch(r"\d+\s*-\s*\d+", n):
        return True
    return False


def clasificar_debito_credito(desc: str, monto: float, saldo_prev: float | None, saldo: float | None) -> tuple[float, float]:
    if saldo_prev is not None and saldo is not None:
        delta = round(saldo - saldo_prev, 2)
        if abs(abs(delta) - abs(monto)) <= 0.05:
            if delta < 0:
                return abs(monto), 0.0
            if delta > 0:
                return 0.0, abs(monto)
    dlow = desc.lower()
    if any(p in dlow for p in PALABRAS_DEBITO):
        # "impuesto ... credito" still reduces balance → debit
        if "pago comercios" in dlow or "pago a proveedores recibido" in dlow or "pago con transferencia" in dlow:
            if "realizada" not in dlow:
                return 0.0, abs(monto)
        if "transferencia recibida" in dlow or "acredita" in dlow or "deposito" in dlow or "depósito" in dlow:
            return 0.0, abs(monto)
        return abs(monto), 0.0
    if any(p in dlow for p in ("credito", "crédito", "recibid", "deposito", "depósito", "transferencia recibida", "pago comercios", "pago a proveedores")):
        return 0.0, abs(monto)
    return abs(monto), 0.0


def extraer_texto_paginas(pdf: Path) -> list[tuple[int, str]]:
    doc = fitz.open(pdf)
    out = []
    for i in range(doc.page_count):
        t = doc[i].get_text("text") or ""
        out.append((i + 1, t))
    doc.close()
    return out


def parsear_movimientos(paginas: list[tuple[int, str]]) -> tuple[list[dict], dict]:
    meta = {
        "cliente": "",
        "cuit": "",
        "cuenta": "",
        "cbu": "",
        "periodos": [],
    }
    movs: list[dict] = []
    periodo_actual = ""
    saldo_prev: float | None = None

    # Flatten lines with page + period tracking
    lineas: list[tuple[int, str, str]] = []
    for pag, texto in paginas:
        if "OFTALMOLOGIA" in texto.upper() and not meta["cliente"]:
            for ln in texto.splitlines():
                if "OFTALMOLOGIA" in ln.upper():
                    meta["cliente"] = ln.strip()
                    break
        m_cuit = RE_CUIT.search(texto)
        if m_cuit and not meta["cuit"]:
            meta["cuit"] = m_cuit.group(1)
        m_cta = RE_CUENTA.search(texto)
        if m_cta and not meta["cuenta"]:
            meta["cuenta"] = m_cta.group(1).strip()
        m_cbu = RE_CBU.search(texto)
        if m_cbu and not meta["cbu"]:
            meta["cbu"] = m_cbu.group(1)
        m_per = RE_PERIODO.search(texto)
        if m_per:
            dsd, hst = m_per.group(1), m_per.group(2)
            periodo_actual = f"{dsd} a {hst}"
            if periodo_actual not in meta["periodos"]:
                meta["periodos"].append(periodo_actual)

        for raw in texto.splitlines():
            ln = re.sub(r"[ \t]+", " ", raw).strip()
            if not ln:
                continue
            lineas.append((pag, ln, periodo_actual))

    i = 0
    periodo_saldo: str | None = None
    while i < len(lineas):
        pag, ln, periodo = lineas[i]
        if periodo != periodo_saldo:
            saldo_prev = None
            periodo_saldo = periodo

        if es_ruido(ln) and not RE_FECHA.match(ln):
            i += 1
            continue

        m_fecha = RE_FECHA.match(ln)
        if not m_fecha:
            i += 1
            continue

        fecha = parse_fecha(m_fecha.group(1))
        if not fecha:
            i += 1
            continue

        resto = ln[m_fecha.end() :].strip()
        comprobante = ""
        m_comp = re.match(r"^(\d{4,12})\b", resto)
        if m_comp:
            comprobante = m_comp.group(1)
            resto = resto[m_comp.end() :].strip()

        bloque = [resto] if resto else []
        j = i + 1
        montos: list[float] = []
        while j < len(lineas) and j < i + 12:
            p2, l2, per2 = lineas[j]
            if RE_FECHA.match(l2) and j > i:
                break
            if es_ruido(l2) and not RE_PESOS.search(l2):
                j += 1
                continue
            for m in RE_PESOS.finditer(l2):
                montos.append(parse_monto(m.group(1)))
            desc_part = RE_PESOS.sub("", l2).strip(" -")
            if desc_part and not re.fullmatch(r"\d{4,12}", desc_part):
                bloque.append(desc_part)
            elif re.fullmatch(r"\d{4,12}", l2.strip()) and not comprobante:
                comprobante = l2.strip()
            j += 1

        descripcion = re.sub(r"\s+", " ", " ".join(b for b in bloque if b)).strip(" -|")
        if not descripcion and not montos:
            i = j
            continue

        # Descartar basura de pie de extracto
        dlow = descripcion.lower()
        if "detalle impositivo" in dlow or "saldo total detalle" in dlow:
            i = j
            continue

        if "saldo inicial" in dlow:
            saldo = montos[0] if montos else None
            movs.append(
                {
                    "Fecha": fecha.strftime("%d/%m/%Y"),
                    "Periodo extracto": periodo,
                    "Comprobante": comprobante,
                    "Descripcion": descripcion or "Saldo Inicial",
                    "Debito": None,
                    "Credito": None,
                    "Saldo": saldo,
                    "Pagina PDF": pag,
                    "Tipo fila": "Saldo inicial",
                }
            )
            saldo_prev = saldo
            i = j
            continue

        if not montos:
            i = j
            continue

        if len(montos) >= 2:
            monto_mov, saldo = montos[-2], montos[-1]
        else:
            monto_mov, saldo = montos[0], None

        debito, credito = clasificar_debito_credito(descripcion, monto_mov, saldo_prev, saldo)
        if saldo_prev is not None and saldo is not None:
            delta = round(saldo - saldo_prev, 2)
            if abs(abs(delta) - abs(monto_mov)) <= 0.05:
                if delta < 0:
                    debito, credito = abs(monto_mov), 0.0
                elif delta > 0:
                    debito, credito = 0.0, abs(monto_mov)

        movs.append(
            {
                "Fecha": fecha.strftime("%d/%m/%Y"),
                "Periodo extracto": periodo,
                "Comprobante": comprobante,
                "Descripcion": descripcion or "Sin descripción",
                "Debito": debito if debito else None,
                "Credito": credito if credito else None,
                "Saldo": saldo,
                "Pagina PDF": pag,
                "Tipo fila": "Movimiento",
            }
        )
        if saldo is not None:
            saldo_prev = saldo
        i = j

    return movs, meta


def main() -> None:
    print("Leyendo PDF...", flush=True)
    paginas = extraer_texto_paginas(PDF)
    print(f"Paginas: {len(paginas)}", flush=True)
    movs, meta = parsear_movimientos(paginas)
    print(f"Filas parseadas: {len(movs)}", flush=True)

    df = pd.DataFrame(movs)
    # Dedup exact duplicates (merged PDF may repeat)
    before = len(df)
    if not df.empty:
        df = df.drop_duplicates(
            subset=["Fecha", "Comprobante", "Descripcion", "Debito", "Credito", "Saldo"],
            keep="first",
        ).reset_index(drop=True)
    print(f"Tras deduplicar: {len(df)} (antes {before})", flush=True)

    total_deb = float(df["Debito"].fillna(0).sum()) if len(df) else 0.0
    total_cred = float(df["Credito"].fillna(0).sum()) if len(df) else 0.0
    n_mov = int((df["Tipo fila"] == "Movimiento").sum()) if len(df) else 0

    resumen = pd.DataFrame(
        [
            {"Campo": "Cliente", "Valor": meta.get("cliente") or "OFTALMOLOGIA RELE MAR DEL SRL"},
            {"Campo": "CUIT", "Valor": meta.get("cuit") or ""},
            {"Campo": "Cuenta", "Valor": meta.get("cuenta") or ""},
            {"Campo": "CBU", "Valor": meta.get("cbu") or ""},
            {"Campo": "Banco", "Valor": "Banco Santander"},
            {"Campo": "PDF origen", "Valor": str(PDF)},
            {"Campo": "Paginas PDF", "Valor": len(paginas)},
            {"Campo": "Periodos detectados", "Valor": " | ".join(meta.get("periodos") or [])},
            {"Campo": "Filas totales", "Valor": len(df)},
            {"Campo": "Movimientos", "Valor": n_mov},
            {"Campo": "Total Debitos", "Valor": round(total_deb, 2)},
            {"Campo": "Total Creditos", "Valor": round(total_cred, 2)},
            {"Campo": "Generado", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M")},
        ]
    )

    # Por periodo sheet
    por_periodo = (
        df[df["Tipo fila"] == "Movimiento"]
        .groupby("Periodo extracto", dropna=False)
        .agg(
            Movimientos=("Descripcion", "count"),
            Total_Debito=("Debito", "sum"),
            Total_Credito=("Credito", "sum"),
        )
        .reset_index()
        if len(df)
        else pd.DataFrame()
    )

    header_font = Font(name="Calibri", bold=True, size=12, color="000000")
    body_font = Font(name="Calibri", size=11, color="000000")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fill_h = PatternFill("solid", fgColor="D9E2F3")

    def style_sheet(ws, money_cols=()):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill_h
            cell.border = thin
        for row in ws.iter_rows(min_row=2, max_row=max(2, ws.max_row), max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.border = thin
                if cell.column in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    for r in dataframe_to_rows(resumen, index=False, header=True):
        ws.append(r)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100

    ws2 = wb.create_sheet("Movimientos")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)
    style_sheet(ws2, money_cols=(5, 6, 7))
    for col, w in {
        "A": 12,
        "B": 22,
        "C": 14,
        "D": 55,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 12,
        "I": 14,
    }.items():
        ws2.column_dimensions[col].width = w
    if ws2.max_row > 1:
        ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Por_periodo")
    for r in dataframe_to_rows(por_periodo, index=False, header=True):
        ws3.append(r)
    style_sheet(ws3, money_cols=(3, 4))
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 16

    wb.save(OUT)
    print("Guardado:", OUT, flush=True)
    try:
        wb.save(OUT2)
        print("Copia Desktop:", OUT2, flush=True)
    except Exception as e:
        print("No se pudo copiar a Desktop:", e, flush=True)
    print("OK", flush=True)


if __name__ == "__main__":
    main()
