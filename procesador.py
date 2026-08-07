"""Procesamiento de extractos bancarios, conciliación y exportación Tango/Excel."""

from __future__ import annotations

import io
import calendar
import hashlib
import json
import re
import zipfile
import shutil
import subprocess
import sys
import threading
import unicodedata
from copy import copy, deepcopy
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
import os
from typing import Optional

import fitz
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from rapidfuzz import fuzz, process

BASE_DIR = Path(__file__).resolve().parent
RUTA_RAIZ_CLIENTES = BASE_DIR / "clientes"
BALANCE_LOCAL_POR_SOCIEDAD: dict[str, str] = {
    "oftalmologia rele mar del plata s.r.l.": r"./Copia de OFTALMOLOGIA RELE Balance 2026.xlsx",
    "30718022742": r"./Copia de OFTALMOLOGIA RELE Balance 2026.xlsx",
    "177": r"./Copia de OFTALMOLOGIA RELE Balance 2026.xlsx",
}
BALANCE_EXCEL_PROYECTO = r"./Copia de OFTALMOLOGIA RELE Balance 2026.xlsx"
PLANTILLA_CONCILIACION = BASE_DIR / "planilla de conciliacion" / "Planilla de Conciliacion.xlsx"
PLAN_CUENTAS_DEFAULT = BASE_DIR / "planes de cuentas" / "Cuentas contables (4).xlsx"
COMPRAS_TANGO_PATH = BASE_DIR / "Compras_Tango.xlsx"
PLANTILLA_ASIENTOS_CONTABLES = BASE_DIR / "Asientos contables (11).xlsx"
PLANTILLA_ASIENTOS_TANGO_VACIO = (
    BASE_DIR / "Asientos contables VACIO PARA LLENAR E IMPORTAR.xlsx"
)
HOJA_ASIENTOS = "Asientos contables"
HOJA_RENGLONES = "Renglones"
COLUMNAS_GRILLA_PRESTAMOS = [
    "Fecha",
    "Importe Cuota (Capital)",
    "Impuestos",
    "Intereses",
    "Total a Debitar",
]
HOJA_BANCO_DEFAULT = "Banco Santander"
TOLERANCIA_DIAS_CLEARING = 3
TOLERANCIA_IMPORTE = 0.02
MAX_PDFS_ANUALES = 12

# Filas fijas de la hoja "Banco Santander" en la plantilla original
FILAS_PLANTILLA = {
    "saldo_inicio": 57,
    "ingresos": 58,
    "retiros": 59,
    "saldo_final": 60,
    "saldo_resumen": 61,
    "diferencia": 62,
}
HOJA_DETALLE_MOVIMIENTOS = "Detalle Extracto"
FILA_INICIO_DETALLE = 2  # Fila 1 = encabezados
ANIO_MIN_EXTRACTO = 1990
ANIO_MAX_EXTRACTO = date.today().year + 1

TAX_REGISTRY: dict[str, dict] = {
    "IVA": {
        "slug": "iva",
        "codigo_tango": "IVA",
        "solapas": ["IVA", "Iva"],
        "inputs_contingencia": [
            {"clave": "saldos_tecnicos_list", "prefix": "iva_stec", "titulo": "Saldo Técnico IVA Período Anterior ($)"},
            {"clave": "saldos_libre_list", "prefix": "iva_slib", "titulo": "Saldo Libre Disponibilidad IVA Período Anterior ($)"},
        ],
        "inputs_manuales": [
            {"clave": "retenciones", "titulo": "Total Retenciones IVA del Mes ($)"},
            {"clave": "percepciones", "titulo": "Total Percepciones IVA del Mes ($)"},
        ],
        "cuenta_ajuste_centavos_rol": "ventas_21",
        "cuenta_ajuste_label": "IVA Débito Fiscal 21%",
        "motor": "iva",
        "palabras_clave": [],
    },
    "Ingresos Brutos": {
        "slug": "iibb",
        "codigo_tango": "IIBB",
        "solapas": ["IIBB", "Ingresos Brutos", "ingresos_brutos"],
        "inputs_contingencia": [
            {"clave": "saldos_favor_iibb_list", "prefix": "iibb_sfav", "titulo": "Saldo a Favor IIBB Período Anterior ($)"},
        ],
        "inputs_manuales": [
            {"clave": "retenciones", "titulo": "Total Retenciones IIBB del Mes ($)"},
            {"clave": "percepciones", "titulo": "Total Percepciones IIBB del Mes ($)"},
            {"clave": "retenciones_bancarias", "titulo": "Total Retenciones Bancarias (Sircreb) del Mes ($)"},
        ],
        "cuenta_ajuste_centavos_rol": "impuesto_determinado",
        "cuenta_ajuste_label": "IIBB Impuesto Determinado",
        "motor": "iibb",
        "palabras_clave": [],
    },
    "Convenio Multilateral": {
        "slug": "cm",
        "codigo_tango": "CM",
        "solapas": [
            "CM",
            "Convenio Multilateral",
            "Convenio",
            "Ingresos Brutos CM",
            "Convenio Mult",
        ],
        "inputs_contingencia": [
            {"clave": "saldos_favor_cm_list", "prefix": "cm_sfav", "titulo": "Saldo a Favor CM Período Anterior ($)"},
        ],
        "inputs_manuales": [
            {"clave": "retenciones", "titulo": "Total Retenciones CM del Mes ($)"},
            {"clave": "percepciones", "titulo": "Total Percepciones CM del Mes ($)"},
            {"clave": "retenciones_bancarias", "titulo": "Total Retenciones Bancarias (Sircreb) del Mes ($)"},
        ],
        "cuenta_ajuste_centavos_rol": "impuesto_determinado",
        "cuenta_ajuste_label": "CM Impuesto Determinado",
        "motor": "iibb",
        "palabras_clave": [],
    },
    "Sueldos": {
        "slug": "sueldos",
        "codigo_tango": "SUELDOS",
        "solapas": ["Sueldos", "SUELDOS", "Cargas Sociales", "F931"],
        "inputs_contingencia": [
            {"clave": "saldos_favor_sueldos_list", "prefix": "suel_sfav", "titulo": "Saldos a Favor Asignaciones ($)"},
        ],
        "inputs_manuales": [],
        "cuenta_ajuste_centavos_rol": "sueldos_jornales",
        "cuenta_ajuste_label": "Sueldos y Jornales",
        "motor": "sueldos",
        "palabras_clave": [
            {"clave": "sueldos_brutos", "keywords": ["sueldos brutos", "sueldo bruto", "total bruto"]},
            {"clave": "sueldos_netos", "keywords": ["sueldos netos", "sueldo neto", "total neto"]},
            {"clave": "aportes", "keywords": ["aportes", "f931", "aportes empleados"]},
            {"clave": "contribuciones", "keywords": ["contribuciones", "contribuciones patronales", "cargas sociales devengadas"]},
            {"clave": "sueldos_pagar", "keywords": ["sueldos a pagar", "sueldos y jornales a pagar", "jornales a pagar"]},
            {"clave": "cargas_sociales_pagar", "keywords": ["cargas sociales a pagar", "cargas a pagar", "cargas sociales pagar"]},
            {"clave": "sindicato", "keywords": ["sindicato", "cuota sindical"]},
        ],
    },
    "TISH": {
        "slug": "tish",
        "codigo_tango": "TISH",
        "solapas": ["TISH", "Seguridad e Higiene", "Tasas Municipales"],
        "inputs_contingencia": [
            {"clave": "saldos_favor_tish_list", "prefix": "tish_sfav", "titulo": "Saldo a Favor TISH ($)"},
        ],
        "inputs_manuales": [],
        "cuenta_ajuste_centavos_rol": "gasto_tasa",
        "cuenta_ajuste_label": "Gasto Tasa / Tasa Determinada",
        "motor": "tish",
        "palabras_clave": [
            {"clave": "tasa_determinada", "keywords": ["tasa determinada", "impuesto determinado tish", "tasa tish"]},
            {"clave": "retenciones_tish", "keywords": ["retenciones tish", "retencion tish", "retenciones seguridad e higiene"]},
            {"clave": "derecho_oficina", "keywords": ["derecho de oficina", "derecho oficina", "derecho municipal"]},
        ],
    },
}


BANK_REGISTRY: dict[str, dict] = {
    "Banco Galicia": {
        "slug": "galicia",
        "codigo_tango": "GALICIA",
        "solapas": ["Galicia", "Banco Galicia", "GALICIA", "Banco Galicia S.A."],
        "cuenta_ajuste_codigo": "11102",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Banco Nación": {
        "slug": "nacion",
        "codigo_tango": "NACION",
        "solapas": ["Nacion", "Banco Nacion", "BANCO NACION", "Banco Nación", "NACION"],
        "cuenta_ajuste_codigo": "11103",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Banco Provincia": {
        "slug": "provincia",
        "codigo_tango": "PROVINCIA",
        "solapas": ["Provincia", "Banco Provincia", "BANCO PROVINCIA", "PROVINCIA"],
        "cuenta_ajuste_codigo": "11105",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Santander": {
        "slug": "santander",
        "codigo_tango": "SANTANDER",
        "solapas": ["SANTANDER", "Santander", "Banco Santander", "Santander Rio", "Banco Santander Rio"],
        "cuenta_ajuste_codigo": "11104",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "BBVA": {
        "slug": "bbva",
        "codigo_tango": "BBVA",
        "solapas": ["BBVA", "Banco Frances", "Francés", "FRANCES"],
        "cuenta_ajuste_codigo": "11106",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Banco Macro": {
        "slug": "macro",
        "codigo_tango": "MACRO",
        "solapas": ["Macro", "Banco Macro", "MACRO"],
        "cuenta_ajuste_codigo": "11107",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Banco Credicoop": {
        "slug": "credicoop",
        "codigo_tango": "CREDICOOP",
        "solapas": ["Credicoop", "Banco Credicoop", "CREDICOOP"],
        "cuenta_ajuste_codigo": "11108",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "HSBC": {
        "slug": "hsbc",
        "codigo_tango": "HSBC",
        "solapas": ["HSBC", "Banco HSBC"],
        "cuenta_ajuste_codigo": "11109",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "ICBC": {
        "slug": "icbc",
        "codigo_tango": "ICBC",
        "solapas": ["ICBC", "Banco ICBC"],
        "cuenta_ajuste_codigo": "11110",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
    "Mercado Pago": {
        "slug": "mercadopago",
        "codigo_tango": "MERCADOPAGO",
        "solapas": [
            "Mercado Pago",
            "MercadoPago",
            "MERCADO PAGO",
            "Mercado Pago S.A.",
            "M Pago",
            "MP",
        ],
        # Código usado por el perfil bancario general. La cuenta de cada
        # sociedad se resuelve igualmente contra su plan de cuentas activo.
        "cuenta_ajuste_codigo": "11114",
        "cuenta_ajuste_centavos_rol": "cuenta_banco_principal",
        "motor": "banco",
    },
}


def obtener_ficha_banco(nombre: str) -> dict:
    """Retorna la ficha del banco desde BANK_REGISTRY (match tolerante)."""
    clave = _normalizar_nombre_solapa_balance(nombre)
    for reg_key, ficha in BANK_REGISTRY.items():
        if _normalizar_nombre_solapa_balance(reg_key) == clave:
            return ficha
        slug = ficha.get("slug", "")
        if slug and _normalizar_nombre_solapa_balance(slug) == clave:
            return ficha
    raise ValueError(f"Banco no registrado en BANK_REGISTRY: {nombre!r}")


def solapas_banco(nombre: str) -> tuple[str, ...]:
    """Nombres de solapa aceptados según el banco seleccionado."""
    ficha = obtener_ficha_banco(nombre)
    solapas = ficha.get("solapas") or ()
    return tuple(str(s) for s in solapas)


def obtener_ficha_impuesto(nombre: str) -> dict:
    """Retorna la ficha del impuesto desde TAX_REGISTRY (match tolerante)."""
    clave = _normalizar_nombre_solapa_balance(nombre)
    for reg_key, ficha in TAX_REGISTRY.items():
        if _normalizar_nombre_solapa_balance(reg_key) == clave:
            return ficha
        slug = ficha.get("slug", "")
        if slug and _normalizar_nombre_solapa_balance(slug) == clave:
            return ficha
    raise ValueError(f"Impuesto no registrado en TAX_REGISTRY: {nombre!r}")


def solapas_impuesto(nombre: str) -> tuple[str, ...]:
    """Nombres de solapa aceptados según el impuesto (desde TAX_REGISTRY)."""
    ficha = obtener_ficha_impuesto(nombre)
    solapas = ficha.get("solapas") or ()
    return tuple(str(s) for s in solapas)


def sanitizar_ruta_unc(ruta: str) -> str:
    """
    Limpia rutas UNC pegadas desde Windows («Copiar como ruta»).
    Elimina espacios y comillas simples/dobles al inicio y al final.
    """
    ruta_limpia = str(ruta or "").strip().strip('"').strip("'").strip()
    return ruta_limpia


def es_ruta_http_legacy(ruta: str) -> bool:
    """True si la ruta es una URL web legacy (Excel Cloud) en lugar de archivo local/UNC."""
    limpia = sanitizar_ruta_unc(ruta).lower()
    return limpia.startswith("http://") or limpia.startswith("https://")


def ruta_balance_local_por_sociedad(
    *,
    nombre: str = "",
    cuit: str = "",
    sociedad_id: int | None = None,
) -> str:
    """Ruta relativa al balance local preconfigurado para una sociedad, o cadena vacía."""
    claves = (
        _normalizar_nombre_solapa_balance(nombre) if nombre else "",
        re.sub(r"\D", "", str(cuit or "")),
        str(sociedad_id) if sociedad_id is not None else "",
    )
    for clave in claves:
        if clave and clave in BALANCE_LOCAL_POR_SOCIEDAD:
            return BALANCE_LOCAL_POR_SOCIEDAD[clave]
    return ""


def resolver_ruta_balance_archivo(ruta: str) -> Path:
    """Resuelve ruta UNC, absoluta o relativa (./) contra BASE_DIR del proyecto."""
    ruta_limpia = sanitizar_ruta_unc(ruta)
    if not ruta_limpia:
        raise ValueError("La ruta del balance está vacía.")
    if es_ruta_http_legacy(ruta_limpia):
        raise ValueError(
            "La ruta apunta a una URL web (Excel Cloud). "
            "Usá una ruta local relativa (./archivo.xlsx) o UNC (\\\\servidor\\carpeta\\balance.xlsx)."
        )
    path = Path(ruta_limpia)
    if path.is_file():
        return path.resolve()
    candidatas: list[Path] = []
    if ruta_limpia.startswith(("./", ".\\")):
        candidatas.append(BASE_DIR / ruta_limpia[2:])
    if not path.is_absolute():
        candidatas.append(BASE_DIR / ruta_limpia)
    candidatas.append(path)
    for candidata in candidatas:
        if candidata.is_file():
            return candidata.resolve()
    raise FileNotFoundError(
        f"No se encontró el archivo de balance en: {ruta_limpia!r}. "
        "Verificá la ruta relativa al proyecto, la UNC o los permisos de red."
    )


def cargar_balance_desde_ruta_unc(ruta: str) -> io.BytesIO:
    """Lee balance Excel/CSV desde ruta UNC, absoluta o relativa al proyecto."""
    path = resolver_ruta_balance_archivo(ruta)
    data = path.read_bytes()
    if not data:
        raise ValueError(f"El archivo de balance está vacío: {path.name}")
    buf = io.BytesIO(data)
    buf.name = path.name
    if path.suffix.lower() not in (".csv",):
        try:
            listar_solapas_excel(buf)
        except Exception as exc:
            raise ValueError(f"El archivo no es un Excel válido: {path.name}") from exc
    buf.seek(0)
    return buf


cargar_balance_desde_ruta = cargar_balance_desde_ruta_unc


def _matcher_etiqueta_por_palabras_clave(concepto: str, reglas: list[dict]) -> str | None:
    """Resuelve clave de datos según palabras clave (columna A)."""
    if not concepto:
        return None
    for regla in reglas:
        clave = regla.get("clave", "")
        keywords = regla.get("keywords") or ()
        for kw in keywords:
            if _normalizar_texto(str(kw)) in concepto:
                return str(clave)
    return None


def leer_datos_balance_por_ficha(
    source,
    impuesto: str,
    *,
    es_csv: bool = False,
    periodo_mensual: str | None = None,
) -> dict:
    """Escaneo genérico columna A usando palabras_clave de la ficha del impuesto."""
    ficha = obtener_ficha_impuesto(impuesto)
    motor = ficha.get("motor", "")
    if motor in ("iva", "iibb"):
        raise ValueError(
            f"leer_datos_balance_por_ficha no aplica al motor '{motor}'; "
            "use leer_planilla_iva_* / leer_planilla_iibb_* en app.py."
        )

    reglas = ficha.get("palabras_clave") or []
    datos: dict = {str(r.get("clave")): 0.0 for r in reglas if r.get("clave")}
    datos.update({"periodo_texto": "", "periodo_mes": None, "periodo_anio": None})

    if es_csv:
        if hasattr(source, "seek"):
            source.seek(0)
        df_mes = pd.read_csv(source, header=None)
    else:
        df_mes = leer_dataframe_balance_solapa(source, impuesto, es_csv=False, header=None)

    asignados: set[str] = set()
    col_montos_idx = 1 if df_mes.shape[1] > 1 else 0
    if periodo_mensual:
        coords = congelar_coordenadas_balance(df_mes, periodo_mensual)
        if coords is not None:
            col_montos_idx = coords.columna_indice
    if df_mes.shape[1] > 0:
        col_a = df_mes[0].fillna("").astype(str).map(_normalizar_texto)
        col_montos = df_mes[col_montos_idx]
        for concepto_raw, monto_raw in zip(col_a, col_montos):
            concepto = re.sub(r"\s+", " ", str(concepto_raw).strip())
            if not concepto or "periodo" in concepto:
                continue
            clave = _matcher_etiqueta_por_palabras_clave(concepto, reglas)
            if clave is None or clave in asignados:
                continue
            monto = _celda_a_float_balance(monto_raw)
            if monto != 0.0:
                datos[clave] = round(float(datos.get(clave, 0.0)) + monto, 2)
                asignados.add(clave)

    if periodo_mensual:
        parsed = _parsear_periodo_balance_celda(periodo_mensual.replace("/", "-"))
        if parsed:
            datos["periodo_mes"], datos["periodo_anio"] = parsed
            datos["periodo_texto"] = periodo_mensual.replace("/", "-")
    elif df_mes.shape[1] > 0:
        col_a_norm = df_mes[0].fillna("").astype(str).map(_normalizar_texto)
        col_b = df_mes[1] if df_mes.shape[1] > 1 else df_mes[0]
        for concepto_raw, val in zip(col_a_norm, col_b):
            concepto = re.sub(r"\s+", " ", str(concepto_raw).strip())
            if "periodo" in concepto or concepto in ("mes", "mes/anio", "mes/año"):
                periodo = _parsear_periodo_balance_celda(val)
                if periodo:
                    datos["periodo_mes"], datos["periodo_anio"] = periodo
                    if isinstance(val, (pd.Timestamp, datetime, date)):
                        datos["periodo_texto"] = f"{val.month:02d}-{val.year}"
                    else:
                        datos["periodo_texto"] = str(val).strip()
                    break

    return datos


def _celda_a_float_balance(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return round(float(val), 2)
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return 0.0
    negativo = False
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()
        negativo = True
    elif s.startswith("-"):
        negativo = True
        s = s[1:].strip()
    try:
        monto = round(float(s.replace(".", "").replace(",", ".")), 2)
        return -monto if negativo else monto
    except ValueError:
        pass
    matches = re.findall(r"[\d\.]+,[\d]{2}", s)
    if matches:
        monto = round(float(matches[-1].replace(".", "").replace(",", ".")), 2)
        return -monto if negativo else monto
    return 0.0


MESES_MAP: dict[str, int] = {
    "enero": 1, "ene": 1, "febrero": 2, "feb": 2, "marzo": 3, "mar": 3,
    "abril": 4, "abr": 4, "mayo": 5, "may": 5, "junio": 6, "jun": 6,
    "julio": 7, "jul": 7, "agosto": 8, "ago": 8, "septiembre": 9, "sep": 9,
    "octubre": 10, "oct": 10, "noviembre": 11, "nov": 11, "diciembre": 12, "dic": 12,
    # Alias tolerantes (Excel / exportaciones mixtas)
    "jan": 1, "apr": 4, "aug": 8, "sept": 9, "set": 9, "dec": 12,
}

_MESES_NOMBRE_A_NUM = MESES_MAP
_MESES_TOKENS_ORDENADOS = sorted(MESES_MAP.keys(), key=len, reverse=True)


def _texto_cabecera_normalizado(val) -> str:
    """Minúsculas, sin acentos y sin espacios (Traductor de Períodos Textuales)."""
    if pd.isna(val):
        return ""
    return re.sub(r"\s+", "", _normalizar_texto(str(val).strip()))


def _anio_desde_texto_cabecera(texto: str) -> int | None:
    m = re.search(r"\b(19|20)\d{2}\b", str(texto or ""))
    if m:
        anio = int(m.group(0))
        if 1990 <= anio <= 2050:
            return anio
    return None


def _mes_desde_texto_cabecera(val) -> int | None:
    """
    Detecta mes numérico en cabeceras textuales: 'Abril', 'ABRIL', 'Sueldos Abril', 'Abr'.
    """
    s = _texto_cabecera_normalizado(val)
    if not s:
        return None
    for token in _MESES_TOKENS_ORDENADOS:
        if token in s:
            return MESES_MAP[token]
    return None


def _parsear_ejercicio_cierre_balance(df: pd.DataFrame, max_filas: int = 8) -> tuple[int, int]:
    """Mes y año de cierre fiscal desde filas tipo 'Ejercicio cerrado al 31/03/2026'."""
    if df is None or df.empty:
        return 12, date.today().year
    patron = re.compile(
        r"(?:ejercicio|cierre).*?(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})",
        re.IGNORECASE,
    )
    for i in range(min(max_filas, len(df))):
        for j in range(min(6, df.shape[1])):
            s = str(df.iat[i, j] or "")
            m = patron.search(s)
            if m:
                return int(m.group(2)), int(m.group(3))
    return 12, date.today().year


def _anio_fiscal_para_mes(mes: int, mes_cierre: int, anio_cierre: int) -> int:
    """Meses posteriores al cierre pertenecen al ejercicio anterior (ej. Abr–Dic → 2025 si cierra 03/2026)."""
    if mes > mes_cierre:
        return anio_cierre - 1
    return anio_cierre


def celda_coincide_periodo_seleccionado(val, mes_objetivo: int, anio_objetivo: int) -> bool:
    """
    Loop Review — matching robusto fecha nativa o texto español para columna del período.
    """
    if pd.isna(val) or _celda_parece_importe(val):
        return False
    parsed = _parsear_periodo_balance_celda(val)
    if parsed:
        mes, anio = parsed
        return mes == mes_objetivo and anio == anio_objetivo
    s_orig = str(val).strip()
    if not s_orig:
        return False
    mes_txt = _mes_desde_texto_cabecera(val)
    if mes_txt != mes_objetivo:
        return False
    anio_txt = _anio_desde_texto_cabecera(s_orig)
    if anio_txt is not None:
        return anio_txt == anio_objetivo
    return True


def celda_coincide_periodo_flexible(val, mes_objetivo: int, anio_objetivo: int) -> bool:
    """
    Matching ampliado para mes de cierre (Marzo/03) y cabeceras tolerantes.
    Acepta 'marzo', 'mar', '03', '3' con año, y fechas 31/03/AAAA.
    """
    if celda_coincide_periodo_seleccionado(val, mes_objetivo, anio_objetivo):
        return True
    if pd.isna(val) or _celda_parece_importe(val):
        return False
    try:
        s_orig = str(val).strip()
        if not s_orig or len(s_orig) > 80:
            return False
        norm = _normalizar_texto(s_orig)
        if "total" in norm or norm in ("check", "subtotal"):
            return False
        m_cierre = re.search(
            r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})",
            s_orig,
        )
        if m_cierre:
            mes_f = int(m_cierre.group(2))
            anio_f = int(m_cierre.group(3))
            if mes_f == mes_objetivo and anio_f == anio_objetivo:
                return True
        if str(anio_objetivo) in s_orig:
            if re.search(rf"\b0?{mes_objetivo}\b", s_orig):
                return True
            mes_txt = _mes_desde_texto_cabecera(val)
            if mes_txt == mes_objetivo:
                return True
        if mes_objetivo == 3:
            if any(tok in norm for tok in ("marzo", "mar")) and (
                str(anio_objetivo) in s_orig
                or _anio_desde_texto_cabecera(s_orig) in (None, anio_objetivo)
            ):
                return True
    except Exception:
        pass
    return False


def _celda_es_columna_total_o_invalida_haber(val) -> bool:
    """True si la columna Haber contigua no es válida (TOTAL, CHECK, etc.)."""
    if pd.isna(val):
        return False
    norm = _normalizar_texto(str(val))
    return any(kw in norm for kw in ("total general", "total", "check", "subtotal"))


def _enriquecer_mapa_meses_texto(
    df: pd.DataFrame,
    fila_encabezado: int,
    mapa: dict[str, int],
) -> dict[str, int]:
    """Completa el mapa MM/YYYY con cabeceras solo-texto ('Abril', 'Sueldos Abril', 'Abr')."""
    if df is None or df.empty or fila_encabezado >= len(df):
        return mapa
    mes_cierre, anio_cierre = _parsear_ejercicio_cierre_balance(df)
    mapa_out = dict(mapa)
    for j in range(df.shape[1]):
        val = df.iat[fila_encabezado, j]
        if _celda_parece_importe(val):
            continue
        if any(col == j for col in mapa_out.values()):
            continue
        parsed = _parsear_periodo_balance_celda(val)
        if parsed:
            mes, anio = parsed
            mapa_out.setdefault(f"{mes:02d}/{anio}", j)
            continue
        mes = _mes_desde_texto_cabecera(val)
        if not mes:
            continue
        anio = _anio_desde_texto_cabecera(str(val)) or _anio_fiscal_para_mes(
            mes, mes_cierre, anio_cierre,
        )
        mapa_out.setdefault(f"{mes:02d}/{anio}", j)
    return mapa_out


def _parsear_periodo_balance_celda(val) -> tuple[int, int] | None:
    if pd.isna(val):
        return None
    if isinstance(val, str):
        s = str(val).strip()
        if not s or len(s) > 80:
            return None
    if isinstance(val, pd.Timestamp):
        return int(val.month), int(val.year)
    if isinstance(val, datetime):
        return int(val.month), int(val.year)
    if isinstance(val, date):
        return int(val.month), int(val.year)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        serial = float(val)
        # Solo seriales Excel plausibles (~1990–2050); evita confundir importes con fechas.
        if 32874 <= serial <= 55153:
            try:
                ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=serial)
                mes, anio = int(ts.month), int(ts.year)
                if 1990 <= anio <= 2050:
                    return mes, anio
            except (ValueError, OverflowError):
                pass
        return None
    s = str(val).strip()
    m = re.match(r"^(0?[1-9]|[12][0-9]|3[01])[/\-](0?[1-9]|1[0-2])[/\-](\d{4})$", s)
    if m:
        return int(m.group(2)), int(m.group(3))
    m = re.match(r"^(0?[1-9]|1[0-2])[/\-](\d{4})$", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^(0?[1-9]|1[0-2])[/\-](\d{2})$", s)
    if m:
        yy = int(m.group(2))
        anio = 2000 + yy if yy < 50 else 1900 + yy
        return int(m.group(1)), anio
    m = re.match(r"^(\d{4})[/\-](0?[1-9]|1[0-2])(?:[/\-]\d{1,2})?", s)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.match(r"^([A-Za-zÁÉÍÓÚáéíóú]{3,12})[/\-\s]+(\d{4})$", s, re.IGNORECASE)
    if m:
        mes_nombre = _normalizar_texto(m.group(1))[:3]
        mes = _MESES_NOMBRE_A_NUM.get(mes_nombre) or _MESES_NOMBRE_A_NUM.get(
            _normalizar_texto(m.group(1))
        )
        if mes:
            return mes, int(m.group(2))
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(ts):
            mes, anio = int(ts.month), int(ts.year)
            if 1990 <= anio <= 2050:
                return mes, anio
    except Exception:
        pass
    mes_txt = _mes_desde_texto_cabecera(val)
    if mes_txt:
        anio_txt = _anio_desde_texto_cabecera(s)
        if anio_txt:
            return mes_txt, anio_txt
    return None


def etiqueta_periodo_desde_celda(val) -> str | None:
    """Convierte celda de encabezado/fecha a etiqueta MM/YYYY."""
    periodo = _parsear_periodo_balance_celda(val)
    if not periodo:
        return None
    mes, anio = periodo
    return f"{mes:02d}/{anio}"


def _celda_parece_importe(val) -> bool:
    """True si el valor parece un importe contable y no una fecha de encabezado."""
    if pd.isna(val):
        return False
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return abs(float(val)) >= 1000
    return False


def localizar_encabezado_meses_balance(
    df: pd.DataFrame,
    max_filas: int = 30,
) -> tuple[int, dict[str, int]]:
    """
    Escanea filas superiores de sábanas masivas y detecta la fila de encabezados
    de meses con el mapa {MM/YYYY: índice_columna}. Retorna (fila_encabezado, mapa).
    """
    if df is None or df.empty:
        return 0, {}
    mejor_mapa: dict[str, int] = {}
    mejor_fila = 0
    mejor_score = -1
    n_filas = min(max_filas, len(df))
    for i in range(n_filas):
        mapa_fila: dict[str, int] = {}
        texto_meses = 0
        for j in range(df.shape[1]):
            val = df.iat[i, j]
            if _celda_parece_importe(val):
                continue
            etiqueta = etiqueta_periodo_desde_celda(val)
            if etiqueta:
                mapa_fila[etiqueta] = j
            elif _mes_desde_texto_cabecera(val):
                texto_meses += 1
        score = len(mapa_fila) * 100 + texto_meses
        if score > mejor_score:
            mejor_mapa = mapa_fila
            mejor_fila = i
            mejor_score = score
    if mejor_score >= 0:
        mejor_mapa = _enriquecer_mapa_meses_texto(df, mejor_fila, mejor_mapa)
    return mejor_fila, mejor_mapa


def mapear_columnas_periodo_balance(df: pd.DataFrame, max_filas: int = 30) -> dict[str, int]:
    """Mapa MM/YYYY → índice de columna desde encabezados tolerantes (fechas completas o Excel)."""
    _, mapa = localizar_encabezado_meses_balance(df, max_filas=max_filas)
    return mapa


def _normalizar_etiqueta_periodo(periodo: str) -> str | None:
    """Normaliza MM/YYYY, MM-YYYY u otras variantes a MM/YYYY."""
    if not periodo:
        return None
    s = str(periodo).strip().replace("-", "/")
    parsed = _parsear_periodo_balance_celda(s)
    if parsed:
        mes, anio = parsed
        return f"{mes:02d}/{anio}"
    return None


def formatear_periodo_mm_yyyy(val) -> str:
    """String estricto MM/YYYY para columna Período de la grilla (sin hora ni día)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    etiqueta = _normalizar_etiqueta_periodo(str(val).strip().replace("-", "/"))
    if etiqueta:
        return etiqueta
    parsed = _parsear_periodo_balance_celda(val)
    if parsed:
        mes, anio = parsed
        return f"{mes:02d}/{anio}"
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return ""
    if "T" in s:
        s = s.split("T", 1)[0]
    if " " in s:
        s = s.split(" ", 1)[0]
    return s


def formatear_fecha_dd_mm_yyyy(val) -> str:
    """String estricto DD/MM/YYYY para columna Fecha de la grilla (sin 00:00:00 ni ISO)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, pd.Timestamp):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return ""
    if "T" in s:
        s = s.split("T", 1)[0]
    if " " in s and len(s) > 10:
        s = s.split(" ", 1)[0]
    m = re.match(r"^(0?[1-9]|[12][0-9]|3[01])/(0?[1-9]|1[0-2])/(\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(ts):
            return ts.strftime("%d/%m/%Y")
    except Exception:
        pass
    return s


def monto_neto_fila_grilla(row: dict) -> float:
    """Importe neto (Debe o Haber) de una fila editable de grilla Streamlit."""
    stored = row.get("_monto")
    if stored is not None:
        try:
            val = round(float(stored), 2)
            if val > 0:
                return val
        except (TypeError, ValueError):
            pass
    debe = round(float(row.get("Debe") or row.get("debe") or 0), 2)
    haber = round(float(row.get("Haber") or row.get("haber") or 0), 2)
    return max(debe, haber)


def aplicar_monto_editable_fila(row: dict, monto: float, tipo: str | None = None) -> dict:
    """Persiste importe editado manualmente en la fila (partida simple Debe/Haber)."""
    monto = round(max(float(monto or 0), 0.0), 2)
    lado = str(tipo or row.get("_tipo") or "Debe")
    row["_tipo"] = lado
    row["_monto"] = monto
    if lado == "Debe":
        row["Debe"] = monto
        row["Haber"] = 0.0
    else:
        row["Debe"] = 0.0
        row["Haber"] = monto
    return row


def detectar_periodos_en_balance_df(df: pd.DataFrame, max_filas: int = 30) -> list[str]:
    """Escanea encabezados de meses en sábanas multi-columna (fechas completas o MM/YYYY)."""
    mapa = mapear_columnas_periodo_balance(df, max_filas=max_filas)
    if mapa:
        periodos = list(mapa.keys())
        periodos.sort(key=lambda p: (
            int(_parsear_periodo_balance_celda(p.replace("/", "-"))[1]),
            int(_parsear_periodo_balance_celda(p.replace("/", "-"))[0]),
        ))
        return periodos
    if df is None or df.empty:
        return []
    periodos: list[str] = []
    vistos: set[str] = set()
    n_filas = min(max_filas, len(df))
    for i in range(n_filas):
        for j in range(df.shape[1]):
            val = df.iat[i, j]
            if _celda_parece_importe(val):
                continue
            etiqueta = etiqueta_periodo_desde_celda(val)
            if etiqueta and etiqueta not in vistos:
                vistos.add(etiqueta)
                periodos.append(etiqueta)
    periodos.sort(key=lambda p: (
        int(_parsear_periodo_balance_celda(p.replace("/", "-"))[1]),
        int(_parsear_periodo_balance_celda(p.replace("/", "-"))[0]),
    ))
    return periodos


def resolver_indice_columna_periodo(df: pd.DataFrame, periodo: str, max_filas: int = 30) -> int | None:
    """Índice de columna cuyo encabezado coincide con el período MM/YYYY del selector."""
    coords = congelar_coordenadas_balance(df, periodo, max_filas=max_filas)
    return coords.columna_indice if coords else None


def _periodos_default_anio(anio: int | None = None) -> list[str]:
    anio = anio or date.today().year
    return [f"{m:02d}/{anio}" for m in range(1, 13)]


def listar_periodos_disponibles_balance(
    source,
    impuesto: str,
    *,
    es_csv: bool = False,
) -> list[str]:
    """Períodos detectados en la solapa del balance o 01..12 del año actual."""
    try:
        if es_csv:
            if hasattr(source, "seek"):
                source.seek(0)
            df = pd.read_csv(source, header=None)
        else:
            if hasattr(source, "seek"):
                source.seek(0)
            df = leer_dataframe_balance_solapa(source, impuesto, es_csv=False, header=None)
        periodos = detectar_periodos_en_balance_df(df)
        if periodos:
            return periodos
    except Exception:
        pass
    return _periodos_default_anio()


_RE_CODIGO_CUENTA_CONCEPTO = re.compile(r"^(\d{5})\b")


def _codigo_cinco_digitos_desde_celda(val) -> str | None:
    """Extrae código Tango de 5 dígitos desde celda numérica o textual."""
    if pd.isna(val):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = int(float(val))
        if 10000 <= n <= 99999:
            return f"{n:05d}"
        return None
    s = str(val).strip()
    m = re.match(r"^(\d{5})(?:\.0+)?$", s)
    if m:
        return m.group(1)
    m = _RE_CODIGO_CUENTA_CONCEPTO.match(s)
    if m:
        return m.group(1)
    return None


def _limite_columnas_concepto_balance(
    df: pd.DataFrame,
    col_montos_idx: int,
    *,
    mapa_meses: dict[str, int] | None = None,
) -> int:
    """
    Columnas de concepto = antes de la primera columna de mes.
    Evita que notas en celdas de otros meses (PAYWAY, etc.) entren al asiento.
    """
    if mapa_meses:
        try:
            primera = min(int(v) for v in mapa_meses.values() if v is not None)
            return max(1, min(primera, df.shape[1]))
        except (TypeError, ValueError):
            pass
    if col_montos_idx > 0:
        return max(1, min(col_montos_idx, 5, df.shape[1]))
    return min(5, df.shape[1])


def _construir_concepto_fila_balance(
    df: pd.DataFrame,
    row_idx: int,
    col_montos_idx: int,
    *,
    mapa_meses: dict[str, int] | None = None,
) -> str | None:
    """
    Arma el concepto de una fila escaneando solo columnas de cuenta/descripcion.
    Tolera balances con códigos en col B/C y descripción en col D (OFTALMOLOGIA).
    """
    limite = _limite_columnas_concepto_balance(
        df, col_montos_idx, mapa_meses=mapa_meses,
    )
    codigos: list[str] = []
    textos: list[str] = []
    for j in range(limite):
        val = df.iat[row_idx, j]
        if pd.isna(val):
            continue
        cod = _codigo_cinco_digitos_desde_celda(val)
        if cod:
            codigos.append(cod)
            s = str(val).strip()
            # Float Excel "42405.0" → solo código, sin residual ".0" como texto.
            if re.fullmatch(r"\d{5}(?:\.0+)?", s.replace(" ", "")):
                continue
            m = _RE_CODIGO_CUENTA_CONCEPTO.match(s)
            if m:
                resto = s[m.end():].strip()
                if resto and not re.fullmatch(r"\.0+", resto):
                    textos.append(resto)
            continue
        s = str(val).strip()
        if not s or _celda_parece_importe(val):
            continue
        if etiqueta_periodo_desde_celda(val):
            continue
        textos.append(s)
    if not codigos and not textos:
        return None
    codigo = codigos[-1] if codigos else None
    texto = " ".join(textos).strip()
    if codigo and texto:
        return f"{codigo} {texto}"
    return codigo or texto


@dataclass(frozen=True)
class CoordenadasBalanceCongeladas:
    """Índices de columna fijados una sola vez antes del barrido vertical (Loop Review)."""
    idx_debe: int
    idx_haber: int
    fila_encabezado: int
    columna_cabecera_texto: str | None
    columna_indice: int
    mellizas_debe_haber: bool = True


def _es_celda_institucional_no_periodo(val) -> bool:
    """Excluye razón social, ejercicio cerrado y títulos que no son columna de mes."""
    if pd.isna(val):
        return False
    norm = _normalizar_texto(str(val))
    return any(
        kw in norm
        for kw in (
            "oftalmolog", "s.r.l", "ejercicio cerrado", "asiento de",
            "devengamiento", "mar del plata", "representante afip",
        )
    )


def _extraer_mes_anio_celda_seguro(val) -> tuple[int, int] | None:
    """Extrae mes/año sin lanzar excepciones (texto institucional, vacíos, formatos raros)."""
    try:
        if pd.isna(val):
            return None
        if isinstance(val, str):
            s = val.strip()
            if not s or len(s) > 80:
                return None
            if _es_celda_institucional_no_periodo(val):
                return None
            norm = _normalizar_texto(s)
            if any(
                kw in norm
                for kw in (
                    "oftalmolog", "s.r.l", "ejercicio cerrado", "asiento de",
                    "devengamiento", "mar del plata", "representante afip",
                )
            ):
                return None
        if isinstance(val, (pd.Timestamp, datetime, date)):
            return int(val.month), int(val.year)
        return _parsear_periodo_balance_celda(val)
    except (TypeError, ValueError, OverflowError):
        return None


def _formato_celda_cabecera_periodo_seguro(val) -> str:
    try:
        return _formato_celda_cabecera_periodo(val)
    except (TypeError, ValueError, AttributeError):
        if pd.isna(val):
            return "(vacío)"
        return str(val).strip() or "(vacío)"


def _columna_contigua_es_par_mellizo(
    df: pd.DataFrame,
    fila_encabezado: int,
    col_mes: int,
    mes_obj: int,
    anio_obj: int,
) -> bool:
    """
    True si idx_haber=col+1 es columna Haber del mismo período (celda combinada).
    False si la contigua encabeza otro mes (sábana mensual simple → solo leer idx_debe).
    """
    if col_mes + 1 >= df.shape[1]:
        return False
    try:
        val_sig = df.iat[fila_encabezado, col_mes + 1]
        if pd.isna(val_sig) or str(val_sig).strip() == "":
            return True
        if _celda_es_etiqueta_debe_haber(val_sig, "haber"):
            return True
        parsed = _extraer_mes_anio_celda_seguro(val_sig)
        if parsed is not None:
            return parsed == (mes_obj, anio_obj)
        mes_sig = _mes_desde_texto_cabecera(val_sig)
        if mes_sig is not None:
            return mes_sig == mes_obj
        if celda_coincide_periodo_seleccionado(val_sig, mes_obj, anio_obj):
            return True
        return False
    except Exception:
        return False


def _celda_es_etiqueta_debe_haber(val, lado: str) -> bool:
    if pd.isna(val):
        return False
    return _normalizar_texto(str(val)) == lado


def _congelar_indices_mellizos(
    df: pd.DataFrame,
    fila_encabezado: int,
    col_mes: int,
    mes_obj: int,
    anio_obj: int,
    cabecera: str | None,
) -> CoordenadasBalanceCongeladas:
    """Bloqueo absoluto: idx_debe=col, idx_haber=col+1 (inmutable para toda la solapa)."""
    idx_debe = col_mes
    idx_haber = col_mes + 1 if col_mes + 1 < df.shape[1] else col_mes
    if idx_haber >= df.shape[1]:
        idx_haber = idx_debe
    else:
        try:
            if _celda_es_columna_total_o_invalida_haber(df.iat[fila_encabezado, idx_haber]):
                idx_haber = idx_debe
        except Exception:
            idx_haber = idx_debe
    mellizas = _columna_contigua_es_par_mellizo(df, fila_encabezado, col_mes, mes_obj, anio_obj)
    return CoordenadasBalanceCongeladas(
        idx_debe=idx_debe,
        idx_haber=idx_haber,
        fila_encabezado=fila_encabezado,
        columna_cabecera_texto=cabecera,
        columna_indice=idx_debe,
        mellizas_debe_haber=mellizas,
    )


def _escanear_cabecera_periodo_seguro(
    df: pd.DataFrame,
    periodo_mensual: str,
    max_filas: int = 10,
) -> CoordenadasBalanceCongeladas | None:
    """
    Escaneo blindado de cabecera (primeras filas): try/except por celda, congela idx_debe/idx_haber.
    Columnas mellizas: idx_debe = col del mes, idx_haber = col + 1 (inmutable).
    """
    if df is None or df.empty or df.shape[1] == 0:
        return None
    objetivo = _normalizar_etiqueta_periodo(periodo_mensual)
    if not objetivo:
        return None
    parsed_obj = _parsear_periodo_balance_celda(objetivo.replace("/", "-"))
    if not parsed_obj:
        m = re.match(r"^(0?[1-9]|1[0-2])[/\-](\d{4})$", objetivo)
        if not m:
            return None
        mes_obj, anio_obj = int(m.group(1)), int(m.group(2))
    else:
        mes_obj, anio_obj = parsed_obj

    n_filas = min(max_filas, len(df))
    for i in range(n_filas):
        for j in range(df.shape[1]):
            try:
                val = df.iat[i, j]
                if _es_celda_institucional_no_periodo(val):
                    continue
                parsed = _extraer_mes_anio_celda_seguro(val)
                coincide = parsed == (mes_obj, anio_obj) if parsed else False
                if not coincide:
                    coincide = celda_coincide_periodo_flexible(val, mes_obj, anio_obj)
                if not coincide:
                    continue
                cabecera = _formato_celda_cabecera_periodo_seguro(val)
                return _congelar_indices_mellizos(
                    df, i, j, mes_obj, anio_obj, cabecera,
                )
            except Exception:
                pass
    return None


def congelar_coordenadas_balance(
    df: pd.DataFrame,
    periodo_mensual: str,
    max_filas: int = 10,
) -> CoordenadasBalanceCongeladas | None:
    """
    Localiza el período solo en filas de cabecera y congela idx_debe/idx_haber
    antes del loop vertical. Nunca recalcular dentro del for i in range(...).
    """
    coords = _escanear_cabecera_periodo_seguro(df, periodo_mensual, max_filas=max_filas)
    if coords is not None:
        return coords
    try:
        fila_encabezado, col_idx, cabecera_col, _mapa = localizar_columna_periodo_estricto(
            df, periodo_mensual, max_filas=max_filas,
        )
        if col_idx is None:
            return None
        parsed = _parsear_periodo_balance_celda(
            (_normalizar_etiqueta_periodo(periodo_mensual) or "").replace("/", "-"),
        )
        if not parsed:
            return None
        return _congelar_indices_mellizos(
            df, fila_encabezado, col_idx, parsed[0], parsed[1], cabecera_col,
        )
    except Exception:
        return None


@dataclass
class ResultadoExtraccionBalance:
    """Resultado estructurado del escaneo matricial fila por fila."""
    filas: list[dict] = field(default_factory=list)
    error: str | None = None
    error_tipo: str | None = None
    solapa_resuelta: str | None = None
    periodos_disponibles: list[str] = field(default_factory=list)
    columna_indice: int | None = None
    fila_encabezado: int | None = None
    columna_cabecera_texto: str | None = None
    coordenadas: CoordenadasBalanceCongeladas | None = None
    idx_debe: int | None = None
    idx_haber: int | None = None


def _leer_monto_celda_balance(val) -> float:
    """Convierte celda a float con pd.to_numeric; tolera formatos contables argentinos."""
    n = pd.to_numeric(val, errors="coerce")
    if pd.notna(n):
        return round(float(n), 2)
    return _celda_a_float_balance(val)


def _montos_debe_haber_fila_balance(
    df: pd.DataFrame,
    fila: int,
    coords: CoordenadasBalanceCongeladas,
) -> tuple[float, float]:
    """Lee montos con índices congelados (succión horizontal pura)."""
    try:
        val_debe = pd.to_numeric(df.iloc[fila, coords.idx_debe], errors="coerce")
        col_haber = coords.idx_haber
        if col_haber >= df.shape[1]:
            col_haber = coords.idx_debe
        val_haber = pd.to_numeric(df.iloc[fila, col_haber], errors="coerce")
    except (IndexError, KeyError, TypeError, ValueError):
        return 0.0, 0.0
    monto_debe = 0.0 if pd.isna(val_debe) else round(float(val_debe), 2)
    monto_haber = 0.0 if pd.isna(val_haber) else round(float(val_haber), 2)
    return monto_debe, monto_haber


def _fila_completamente_en_blanco(df: pd.DataFrame, fila: int) -> bool:
    if fila < 0 or fila >= len(df):
        return True
    for j in range(df.shape[1]):
        val = df.iat[fila, j]
        if pd.notna(val) and str(val).strip():
            return False
    return True


def _texto_columna_a_fila(df: pd.DataFrame, fila: int) -> str:
    """Texto de concepto desde columna A (0) con fallback a B/C para balances OFTALMOLOGIA."""
    partes: list[str] = []
    for j in range(min(4, df.shape[1])):
        val = df.iat[fila, j]
        if pd.isna(val):
            continue
        s = str(val).strip()
        if s and s.lower() != "nan":
            partes.append(s)
    return " ".join(partes).strip()


def _codigo_tango_desde_columna_a(df: pd.DataFrame, fila: int) -> str | None:
    """Último código Tango de 5 dígitos en columnas A–E (prioridad a la derecha)."""
    encontrado: str | None = None
    for j in range(min(5, df.shape[1])):
        cod = _codigo_cinco_digitos_desde_celda(df.iat[fila, j])
        if cod:
            encontrado = cod
    return encontrado


def _codigo_tango_columna_cero(df: pd.DataFrame, fila: int) -> str | None:
    """Código Tango estricto desde la columna 0 de la fila (partida doble bancos)."""
    if df.shape[1] < 1:
        return None
    return _codigo_cinco_digitos_desde_celda(df.iat[fila, 0])


def _montos_periodo_fila_banco(
    df: pd.DataFrame,
    fila: int,
    coords: CoordenadasBalanceCongeladas,
) -> tuple[float, float]:
    """
    idx_debe = columna Debe del mes; idx_haber = columna Haber contigua (+1).
    Si no hay par mellizo, idx_haber apunta al mes siguiente y se ignora para Haber.
    """
    monto_debe, monto_haber = _montos_debe_haber_fila_balance(df, fila, coords)
    if not coords.mellizas_debe_haber:
        return monto_debe, 0.0
    return monto_debe, monto_haber


def _inyeccion_partida_doble_banco(
    monto_debe: float,
    monto_haber: float,
) -> list[tuple[float, float, str]]:
    """Inyección directa: columna idx_debe → Debe; columna idx_haber → Haber."""
    salida: list[tuple[float, float, str]] = []
    if abs(monto_debe) > 0.01:
        salida.append((round(abs(monto_debe), 2), 0.0, "Debe"))
    if abs(monto_haber) > 0.01:
        salida.append((0.0, round(abs(monto_haber), 2), "Haber"))
    return salida


def _concepto_fila_balance_seguro(
    df: pd.DataFrame,
    fila: int,
    col_montos_idx: int,
    *,
    mapa_meses: dict[str, int] | None = None,
) -> str | None:
    """
    Arma concepto con códigos + descripción (p.ej. col E en Bahía Chica).
    No cortar en A–C: ahí solo suelen estar los códigos Tango.
    """
    construido = _construir_concepto_fila_balance(
        df, fila, col_montos_idx, mapa_meses=mapa_meses,
    )
    if construido:
        return construido
    texto_a = _texto_columna_a_fila(df, fila)
    return texto_a or None


def _montos_grilla_desde_columnas_mellizas(
    monto_debe: float,
    monto_haber: float,
    *,
    coords: CoordenadasBalanceCongeladas,
    concepto: str,
    codigo: str,
    impuesto: str,
    fila_idx: int = 0,
) -> tuple[float, float, str]:
    """
    Inyección Loop Review desde columnas congeladas idx_debe / idx_haber.
    Retorna (debe_grilla, haber_grilla, tipo_principal).
    """
    es_banco = es_entidad_banco_balance(impuesto)

    if es_banco:
        inyecciones = _inyeccion_partida_doble_banco(monto_debe, monto_haber)
        if not inyecciones:
            return 0.0, 0.0, "Debe"
        debe_grilla, haber_grilla, tipo = inyecciones[0]
        return debe_grilla, haber_grilla, tipo

    debe_grilla = 0.0
    haber_grilla = 0.0
    if es_banco or coords.mellizas_debe_haber:
        if abs(monto_debe) > 0.01:
            debe_grilla = round(abs(monto_debe), 2)
        if abs(monto_haber) > 0.01:
            haber_grilla = round(abs(monto_haber), 2)
        if debe_grilla > 0 and haber_grilla <= 0:
            return debe_grilla, haber_grilla, "Debe"
        if haber_grilla > 0 and debe_grilla <= 0:
            return debe_grilla, haber_grilla, "Haber"
        if debe_grilla > 0 and haber_grilla > 0:
            return debe_grilla, haber_grilla, "Debe"
        tipo = inferir_tipo_movimiento_desde_concepto(
            concepto, monto_original=0.0, codigo=codigo, impuesto=impuesto,
        )
        return 0.0, 0.0, tipo

    if abs(monto_debe) > 0.01:
        monto = round(abs(monto_debe), 2)
        tipo = inferir_tipo_movimiento_desde_concepto(
            concepto, monto_original=monto_debe, codigo=codigo, impuesto=impuesto,
        )
        if tipo == "Haber":
            return 0.0, monto, "Haber"
        return monto, 0.0, "Debe"
    # Mes en cero: igual inferir lado para proyectar el concepto del Excel.
    tipo = inferir_tipo_movimiento_desde_concepto(
        concepto, monto_original=0.0, codigo=codigo, impuesto=impuesto,
    )
    return 0.0, 0.0, tipo


def _formato_celda_cabecera_periodo(val) -> str:
    """Texto legible de la celda de encabezado de columna (auditoría Loop Review)."""
    if pd.isna(val):
        return "(vacío)"
    if isinstance(val, pd.Timestamp):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    return str(val).strip()


def localizar_columna_periodo_estricto(
    df: pd.DataFrame,
    periodo_mensual: str,
    max_filas: int = 30,
) -> tuple[int | None, int | None, str | None, int, dict[str, int]]:
    """
    Puntería estricta MM/YYYY → (fila_encabezado, col_idx, texto_cabecera, mapa_meses).
    Solo considera la fila de encabezados detectada; sin desfases ±1 ni filas vecinas.
    """
    fila_encabezado, mapa_meses = localizar_encabezado_meses_balance(df, max_filas=max_filas)
    objetivo = _normalizar_etiqueta_periodo(periodo_mensual)
    if not objetivo or df is None or df.empty:
        return fila_encabezado, None, None, mapa_meses

    parsed_obj = _parsear_periodo_balance_celda(objetivo.replace("/", "-"))
    mes_obj = parsed_obj[0] if parsed_obj else None
    anio_obj = parsed_obj[1] if parsed_obj else None
    if mes_obj is None or anio_obj is None:
        m = re.match(r"^(0?[1-9]|1[0-2])[/\-](\d{4})$", objetivo)
        if m:
            mes_obj, anio_obj = int(m.group(1)), int(m.group(2))

    col_idx = mapa_meses.get(objetivo)
    if col_idx is not None and fila_encabezado < len(df):
        celda_mapa = df.iat[fila_encabezado, col_idx]
        if etiqueta_periodo_desde_celda(celda_mapa) != objetivo:
            if mes_obj is None or not celda_coincide_periodo_seleccionado(
                celda_mapa, mes_obj, anio_obj,
            ):
                col_idx = None

    if col_idx is None and fila_encabezado < len(df):
        candidato_texto: int | None = None
        for j in range(df.shape[1]):
            val = df.iat[fila_encabezado, j]
            if _celda_parece_importe(val):
                continue
            if etiqueta_periodo_desde_celda(val) == objetivo:
                col_idx = j
                mapa_meses[objetivo] = j
                break
            if (
                mes_obj is not None
                and anio_obj is not None
                and celda_coincide_periodo_seleccionado(val, mes_obj, anio_obj)
            ):
                candidato_texto = j
        if col_idx is None and candidato_texto is not None:
            col_idx = candidato_texto
            mapa_meses.setdefault(objetivo, candidato_texto)

    cabecera = (
        _formato_celda_cabecera_periodo(df.iat[fila_encabezado, col_idx])
        if col_idx is not None and fila_encabezado < len(df)
        else None
    )
    return fila_encabezado, col_idx, cabecera, mapa_meses


def listar_bancos_conciliacion() -> list[str]:
    """Entidades financieras registradas para conciliación desde balance Excel."""
    return list(BANK_REGISTRY.keys())


def es_entidad_banco_balance(nombre: str) -> bool:
    """True si el nombre corresponde a una ficha de BANK_REGISTRY."""
    try:
        obtener_ficha_banco(nombre)
        return True
    except ValueError:
        return False


def _es_celda_cuenta_banco_vacia(texto: str) -> bool:
    """True si la celda no aporta cuenta (vacía, guiones, puntos)."""
    s = str(texto or "").strip()
    if not s or s.lower() == "nan":
        return True
    if re.fullmatch(r"[-–—_\s\.]+", s):
        return True
    return False


def _codigo_cinco_digitos_desde_texto_banco(concepto: str) -> str:
    """Extrae código Tango de 5 dígitos desde texto de columna 0."""
    s = str(concepto or "").strip()
    if not s:
        return ""
    m = _RE_CODIGO_CUENTA_CONCEPTO.match(s)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{5})\b", s)
    if m:
        return m.group(1)
    codigo_pre, _ = extraer_codigo_cuenta_tango_desde_concepto(s)
    if codigo_pre and codigo_pre != "99999":
        return codigo_pre
    return ""


def _buscar_codigo_en_plan_banco(
    codigo: str,
    plan_cuentas: list[tuple[str, str]],
) -> tuple[str, str] | None:
    cod = str(codigo or "").strip()
    if not cod:
        return None
    for c, d in plan_cuentas:
        if str(c).strip() == cod:
            return c, d
    return None


def _buscar_cuenta_plan_por_etiquetas(
    plan_cuentas: list[tuple[str, str]],
    *grupos: tuple[str, ...],
) -> tuple[str, str] | None:
    for codigo, desc in plan_cuentas:
        desc_norm = _normalizar_texto(desc)
        for grupo in grupos:
            if all(palabra in desc_norm for palabra in grupo):
                return codigo, desc
    return None


_ALIASES_BANCO_PLAN: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("galicia", ("galicia",)),
    ("macro", ("macro",)),
    ("santander", ("santander",)),
    ("nacion", ("nacion", "bna")),
    ("provincia", ("provincia", "bapro")),
    ("bbva", ("bbva", "frances")),
    ("credicoop", ("credicoop",)),
    ("hsbc", ("hsbc",)),
    ("icbc", ("icbc",)),
    ("patagonia", ("patagonia",)),
    ("supervielle", ("supervielle",)),
    ("ciudad", ("ciudad",)),
    ("comafi", ("comafi",)),
    ("mercadopago", ("mercado pago", "mercadopago")),
)


def _extraer_numeros_cuenta_bancaria(texto: str) -> list[str]:
    """
    Números de cuenta visibles en el concepto (ej. 3-684-0942508301-8, CBU, nro largo).
    Ignora códigos Tango de 5 dígitos al inicio y fechas cortas.
    """
    raw = str(texto or "")
    candidatos: list[str] = []
    # Formatos con guiones / barras típicos de cuentas
    for m in re.finditer(r"\b\d{1,4}[-/]\d{2,6}[-/]\d{4,14}(?:[-/]\d{1,4})?\b", raw):
        dig = re.sub(r"\D", "", m.group(0))
        if len(dig) >= 6:
            candidatos.append(dig)
    # Bloques de dígitos largos (cuenta / CBU parcial)
    for m in re.finditer(r"\d{7,22}", raw):
        dig = m.group(0)
        if len(dig) == 8 and dig.startswith("20"):  # posible año+algo
            continue
        candidatos.append(dig)
    # Únicos, priorizando los más largos
    vistos: set[str] = set()
    ordenados: list[str] = []
    for dig in sorted(candidatos, key=len, reverse=True):
        if dig not in vistos:
            vistos.add(dig)
            ordenados.append(dig)
    return ordenados


def _detectar_banco_en_concepto(concepto_norm: str) -> str | None:
    for slug, aliases in _ALIASES_BANCO_PLAN:
        if any(a in concepto_norm for a in aliases):
            return slug
    if "banco" in concepto_norm or "bco" in concepto_norm:
        return "banco"
    return None


def _buscar_caja_en_plan(plan_cuentas: list[tuple[str, str]]) -> tuple[str, str] | None:
    matches = [(c, d) for c, d in plan_cuentas if "caja" in _normalizar_texto(d)]
    if not matches:
        return None
    for c, d in matches:
        dn = _normalizar_texto(d)
        if dn == "caja" or dn.startswith("caja ") or "caja en" in dn or "caja $" in dn:
            return c, d
    return matches[0]


def _buscar_cuenta_banco_por_numero_en_plan(
    numeros: list[str],
    plan_cuentas: list[tuple[str, str]],
    banco_slug: str | None = None,
) -> tuple[str, str] | None:
    """Busca en el plan una cuenta bancaria que contenga el número (y opcionalmente el banco)."""
    if not numeros or not plan_cuentas:
        return None
    candidatos_banco: list[tuple[str, str]] = []
    if banco_slug and banco_slug != "banco":
        aliases = dict(_ALIASES_BANCO_PLAN).get(banco_slug, (banco_slug,))
        for c, d in plan_cuentas:
            dn = _normalizar_texto(d)
            if any(a in dn for a in aliases) or "banco" in dn or "bco" in dn:
                candidatos_banco.append((c, d))
    pool = candidatos_banco or list(plan_cuentas)

    for num in numeros:
        sufijos = [num]
        if len(num) > 8:
            sufijos.append(num[-10:])
            sufijos.append(num[-8:])
        for c, d in pool:
            dig_desc = re.sub(r"\D", "", d)
            if any(s and s in dig_desc for s in sufijos):
                return c, d
        # También contra descripción con guiones tal cual
        for c, d in pool:
            if any(s in d.replace(" ", "") for s in sufijos if len(s) >= 6):
                return c, d
    return None


def _buscar_cuenta_banco_por_nombre_en_plan(
    banco_slug: str,
    plan_cuentas: list[tuple[str, str]],
) -> tuple[str, str] | None:
    """Si hay una sola cuenta del banco en el plan, la usa; si hay varias, no adivina."""
    if not banco_slug or banco_slug == "banco" or not plan_cuentas:
        return None
    aliases = dict(_ALIASES_BANCO_PLAN).get(banco_slug, (banco_slug,))
    matches = [
        (c, d) for c, d in plan_cuentas
        if any(a in _normalizar_texto(d) for a in aliases)
    ]
    if len(matches) == 1:
        return matches[0]
    return None


def _buscar_cuenta_banco_por_palabras_clave(
    concepto_norm: str,
    plan_cuentas: list[tuple[str, str]],
    *,
    concepto_raw: str = "",
) -> tuple[str, str] | None:
    """PASO 2: caja / banco+nro / impuestos fijos / fuzzy gasto-comisión-iva."""
    if not concepto_norm or not plan_cuentas:
        return None

    if "caja" in concepto_norm:
        hit_caja = _buscar_caja_en_plan(plan_cuentas)
        if hit_caja:
            return hit_caja

    numeros = _extraer_numeros_cuenta_bancaria(concepto_raw or concepto_norm)
    banco_slug = _detectar_banco_en_concepto(concepto_norm)

    # Con número: buscar esa cuenta en el plan (prioridad absoluta).
    if numeros:
        hit_num = _buscar_cuenta_banco_por_numero_en_plan(numeros, plan_cuentas, banco_slug)
        if hit_num:
            return hit_num

    # Menciona banco pero sin número: si hay una sola cuenta de ese banco en el plan, usarla;
    # si hay varias (o ninguna), mandar a Caja — no adivinar Macro/Galicia al azar.
    if banco_slug is not None and not numeros:
        if banco_slug != "banco":
            hit_unica = _buscar_cuenta_banco_por_nombre_en_plan(banco_slug, plan_cuentas)
            if hit_unica:
                return hit_unica
        hit_caja = _buscar_caja_en_plan(plan_cuentas)
        if hit_caja:
            return hit_caja

    reglas_texto_fijas: list[tuple[tuple[str, ...], tuple[tuple[str, ...], ...]]] = [
        (("sircreb",), (("sircreb",),)),
        (("ing", "brutos"), (("iibb", "retenc"), ("ing", "bruto"))),
        (("ley", "25413"), (("25413",),)),
        (("percep", "iva"), (("percep", "iva"),)),
    ]
    for palabras, grupos in reglas_texto_fijas:
        if all(p in concepto_norm for p in palabras):
            hit = _buscar_cuenta_plan_por_etiquetas(plan_cuentas, *grupos)
            if hit:
                return hit

    keywords_fuzzy = ("gasto", "comision", "iva")
    if any(kw in concepto_norm for kw in keywords_fuzzy):
        try:
            from rapidfuzz import fuzz, process
        except ImportError:
            process = None  # type: ignore[assignment,misc]
            fuzz = None  # type: ignore[assignment,misc]

        candidatos = list(plan_cuentas)
        for kw in keywords_fuzzy:
            if kw in concepto_norm:
                filtrados = [(c, d) for c, d in plan_cuentas if kw in _normalizar_texto(d)]
                if filtrados:
                    candidatos = filtrados
                break

        if process is not None and fuzz is not None:
            desc_por_norm = {_normalizar_texto(d): (c, d) for c, d in candidatos}
            if desc_por_norm:
                resultado = process.extractOne(
                    concepto_norm,
                    list(desc_por_norm.keys()),
                    scorer=fuzz.token_set_ratio,
                )
                if resultado and resultado[1] >= 55:
                    return desc_por_norm[resultado[0]]

        for kw in keywords_fuzzy:
            if kw in concepto_norm:
                for c, d in candidatos:
                    if kw in _normalizar_texto(d):
                        return c, d
    return None


def resolver_cuenta_banco_hibrida(
    concepto: str,
    plan_cuentas: list[tuple[str, str]] | None = None,
    *,
    monto: float | None = None,
    es_acreditacion: bool | None = None,
    df_subdiario_tango: pd.DataFrame | None = None,
) -> tuple[str, str]:
    """
    Intérprete híbrido bancario (código → subdiario Tango → texto/keywords → fallback vacío).
    Retorna (codigo, descripcion). codigo vacío → selector manual en grilla.

    Reglas de cuenta bancaria:
    - Si el concepto trae número de cuenta → busca esa cuenta en el plan.
    - Si menciona un banco pero no tiene número → Caja.
    """
    plan = plan_cuentas or []
    texto = str(concepto or "").strip()
    if _es_celda_cuenta_banco_vacia(texto):
        return "", ""

    codigo = _codigo_cinco_digitos_desde_texto_banco(texto)
    if codigo and plan:
        hit = _buscar_codigo_en_plan_banco(codigo, plan)
        if hit:
            return hit
    elif codigo and not plan:
        _, resto = extraer_codigo_cuenta_tango_desde_concepto(texto)
        return codigo, resto or texto

    if monto is not None and monto > 0 and es_acreditacion is not None:
        hit_tango = sugerir_cuenta_conciliacion_tango(
            texto, monto, es_acreditacion, df_subdiario_tango, plan,
        )
        if hit_tango[0]:
            return hit_tango

    concepto_norm = _normalizar_texto(texto)
    kw_hit = _buscar_cuenta_banco_por_palabras_clave(
        concepto_norm, plan, concepto_raw=texto,
    )
    if kw_hit:
        return kw_hit

    return "", texto

def _extraer_codigo_opcional_fila_banco(
    df: pd.DataFrame,
    fila: int,
    concepto: str,
) -> str:
    """Hint de código en col. 0 (resolución final en intérprete híbrido con plan)."""
    codigo = _codigo_tango_columna_cero(df, fila)
    if codigo:
        return codigo
    return _codigo_cinco_digitos_desde_texto_banco(concepto)


def _concepto_columna_cero_fila_banco(df: pd.DataFrame, fila: int) -> str:
    """Texto crudo de la columna 0 (sin validar ni descartar)."""
    if df.shape[1] < 1:
        return ""
    val = df.iat[fila, 0]
    if pd.isna(val):
        return ""
    texto = str(val).strip()
    return "" if texto.lower() == "nan" else texto


def _extraer_filas_banco_lectura_sincronica(
    df: pd.DataFrame,
    coords: CoordenadasBalanceCongeladas,
    impuesto: str,
) -> list[dict]:
    """
    Prioridad absoluta: cantidad y orden secuencial del Excel anclados al monto.
    Único disparador: importe en idx_debe o idx_haber (|valor| > 0.01).
    Cuenta flexible: código opcional; fila nunca se descarta por nombre/código.
    """
    filas: list[dict] = []
    orden_secuencial = 0

    for idx, _row in df.iterrows():
        i = int(idx)

        monto_debe, monto_haber = _montos_periodo_fila_banco(df, i, coords)
        if abs(monto_debe) <= 0.01 and abs(monto_haber) <= 0.01:
            continue

        concepto = _concepto_columna_cero_fila_banco(df, i)
        if not concepto:
            concepto = f"Movimiento Excel fila {i + 1}"

        codigo = _extraer_codigo_opcional_fila_banco(df, i, concepto)
        _, descripcion = extraer_codigo_cuenta_tango_desde_concepto(concepto)
        desc = descripcion if descripcion and codigo else concepto

        for debe_grilla, haber_grilla, tipo in _inyeccion_partida_doble_banco(
            monto_debe, monto_haber,
        ):
            monto = max(debe_grilla, haber_grilla)
            filas.append({
                "concepto_raw": concepto,
                "codigo": codigo,
                "descripcion": desc,
                "monto": monto,
                "debe": debe_grilla,
                "haber": haber_grilla,
                "monto_original": monto_debe if tipo == "Debe" else -monto_haber,
                "tipo": tipo,
                "fila_idx": i,
                "orden_secuencial": orden_secuencial,
            })
            orden_secuencial += 1
    return filas


def extraer_filas_universales_balance_por_banco_con_errores(
    source,
    nombre_banco: str,
    periodo_mensual: str,
    *,
    es_csv: bool = False,
) -> ResultadoExtraccionBalance:
    """Wrapper bancario → mismo motor de coordenadas mellizas con solapa dinámica."""
    return extraer_filas_universales_balance_por_periodo_con_errores(
        source, nombre_banco, periodo_mensual, es_csv=es_csv,
    )


def extraer_filas_universales_balance_por_periodo_con_errores(
    source,
    impuesto: str,
    periodo_mensual: str,
    *,
    es_csv: bool = False,
) -> ResultadoExtraccionBalance:
    """
    Extractor universal con errores estructurados (solapa/período no encontrado).
    """
    solapa_resuelta: str | None = None
    try:
        if es_csv:
            if hasattr(source, "seek"):
                source.seek(0)
            df = pd.read_csv(source, header=None)
        else:
            if hasattr(source, "seek"):
                source.seek(0)
            solapa_resuelta = resolver_solapa_balance(source, impuesto)
            if hasattr(source, "seek"):
                source.seek(0)
            df = pd.read_excel(source, sheet_name=solapa_resuelta, header=None, engine="openpyxl")
    except ValueError as exc:
        msg = str(exc)
        if "No se encontró la solapa" in msg:
            return ResultadoExtraccionBalance(
                error=msg,
                error_tipo="sheet_not_found",
            )
        return ResultadoExtraccionBalance(error=msg, error_tipo="read_error")
    except Exception as exc:
        return ResultadoExtraccionBalance(error=str(exc), error_tipo="read_error")

    if df is None or df.empty or df.shape[1] == 0:
        return ResultadoExtraccionBalance(
            error=f"La solapa «{solapa_resuelta or impuesto}» está vacía.",
            error_tipo="empty_sheet",
            solapa_resuelta=solapa_resuelta,
        )

    coords = congelar_coordenadas_balance(df, periodo_mensual)
    periodos_disponibles = detectar_periodos_en_balance_df(df)
    if coords is None:
        return ResultadoExtraccionBalance(
            error=(
                "No se localizó la columna para el período seleccionado. "
                "Verifique que la solapa contenga la fecha correspondiente."
            ),
            error_tipo="month_column_not_found",
            solapa_resuelta=solapa_resuelta,
            periodos_disponibles=periodos_disponibles,
        )

    col_idx = coords.columna_indice
    fila_encabezado = coords.fila_encabezado
    cabecera_col = coords.columna_cabecera_texto
    es_banco = es_entidad_banco_balance(impuesto)
    _fila_enc_mapa, mapa_meses = localizar_encabezado_meses_balance(df)

    if es_banco:
        filas = _extraer_filas_banco_lectura_sincronica(df, coords, impuesto)
    else:
        filas = []
        for i in range(len(df)):
            if _fila_completamente_en_blanco(df, i):
                continue
            concepto = _concepto_fila_balance_seguro(
                df, i, col_idx, mapa_meses=mapa_meses,
            )
            if not concepto:
                continue
            codigo_col = _codigo_tango_desde_columna_a(df, i)
            codigo_pre, _ = extraer_codigo_cuenta_tango_desde_concepto(concepto)
            if codigo_pre == "99999":
                codigo_pre = ""
            if _es_fila_ruido_balance(concepto):
                continue

            monto_debe, monto_haber = _montos_debe_haber_fila_balance(df, i, coords)
            codigo = codigo_col or codigo_pre or ""
            debe_grilla, haber_grilla, tipo = _montos_grilla_desde_columnas_mellizas(
                monto_debe,
                monto_haber,
                coords=coords,
                concepto=concepto,
                codigo=codigo,
                impuesto=impuesto,
                fila_idx=i,
            )
            # Solo líneas de asiento (cuenta Tango o etiqueta «a …»); no razón social.
            if not _es_linea_proyectable_balance(concepto, codigo):
                continue
            if debe_grilla <= 0.01 and haber_grilla <= 0.01:
                if not codigo or not re.fullmatch(r"\d{5}", str(codigo)):
                    codigo = "99999"

            # Heredar cuenta de «Ajuste por redondeo» del Debe si el Haber no trae código.
            if (not codigo or codigo == "99999") and "ajuste por redondeo" in _normalizar_texto(concepto):
                for prev in filas:
                    if (
                        str(prev.get("codigo") or "").isdigit()
                        and len(str(prev.get("codigo"))) == 5
                        and "ajuste por redondeo" in _normalizar_texto(str(prev.get("concepto_raw") or ""))
                    ):
                        codigo = str(prev["codigo"])
                        break

            descripcion = _descripcion_concepto_balance_limpia(concepto, codigo)
            monto = max(debe_grilla, haber_grilla)
            filas.append({
                "concepto_raw": concepto,
                "codigo": codigo,
                "descripcion": descripcion if descripcion else concepto,
                "monto": monto,
                "debe": debe_grilla,
                "haber": haber_grilla,
                "monto_original": monto_debe if debe_grilla > 0 else (-monto_haber if haber_grilla > 0 else 0.0),
                "tipo": tipo,
                "fila_idx": i,
            })
    return ResultadoExtraccionBalance(
        filas=filas,
        solapa_resuelta=solapa_resuelta,
        periodos_disponibles=periodos_disponibles,
        columna_indice=col_idx,
        fila_encabezado=fila_encabezado,
        columna_cabecera_texto=cabecera_col,
        coordenadas=coords,
        idx_debe=coords.idx_debe,
        idx_haber=coords.idx_haber,
    )


def extraer_filas_universales_balance_por_periodo(
    source,
    impuesto: str,
    periodo_mensual: str,
    *,
    es_csv: bool = False,
) -> list[dict]:
    """
    Extractor universal fila por fila: recorre columnas de concepto y extrae todo monto
    distinto de cero (|valor| > 0.01) en la columna del período seleccionado.
    """
    resultado = extraer_filas_universales_balance_por_periodo_con_errores(
        source, impuesto, periodo_mensual, es_csv=es_csv,
    )
    return resultado.filas


def extraer_codigo_cuenta_tango_desde_concepto(texto: str) -> tuple[str, str]:
    """
    Extrae código Tango de 5 dígitos al inicio del concepto (columna A).
    Ej: '42405 Impuesto...' → ('42405', 'Impuesto...'); '11418 a Retenciones...' → ('11418', 'a Retenciones...').
    """
    s = str(texto or "").strip()
    if not s:
        return "99999", ""
    m = _RE_CODIGO_CUENTA_CONCEPTO.match(s)
    if m:
        codigo = m.group(1)
        resto = s[m.end():].strip()
        return codigo, resto if resto else s
    return "99999", s


def _es_concepto_nota_credito(concepto: str) -> bool:
    """Detecta NC / nota de crédito (movimientos reales del asiento IVA)."""
    t = _normalizar_texto(str(concepto or ""))
    if not t:
        return False
    if "nota" in t and "credito" in t:
        return True
    return bool(re.search(r"\bnc\b", t))


def inferir_tipo_movimiento_desde_concepto(
    texto: str,
    *,
    monto_original: float | None = None,
    codigo: str | None = None,
    impuesto: str | None = None,
) -> str:
    """
    Heurística Debe/Haber desde concepto, signo del Excel y cuenta Tango.
    Pasivos/negativos en sábanas de Sueldos → Haber; gastos positivos → Debe.
    """
    raw = str(texto or "").strip()
    if not raw:
        return "Debe"

    raw_lower = raw.lower()
    t = _normalizar_texto(raw)
    imp_norm = _normalizar_texto(str(impuesto or ""))

    # NC IVA: proyectar en asiento (antes del signo del Excel).
    # NC compras → Debe; NC ventas → Haber (aunque el monto reste en la posición).
    if _es_concepto_nota_credito(raw):
        if "compr" in t:
            return "Debe"
        if "vent" in t:
            return "Haber"

    # Saldo a favor: en Debe del asiento, salvo que la fila diga «a Saldo a favor…»
    if "saldo a favor" in t or "saldo favor" in t:
        if re.search(r"(^|\s)a\s+saldo\s+a?\s*favor", t) or t.startswith("a saldo"):
            return "Haber"
        return "Debe"

    if monto_original is not None and monto_original < -0.01:
        return "Haber"

    cod = str(codigo or "").strip()
    if cod.startswith("213"):
        return "Haber"
    if cod.startswith(("422", "421", "430")):
        return "Debe"

    _PALABRAS_HABER = (
        "a pagar", "f931", "f 931", "retenciones de ley", "sindicato", "sueldos netos",
        "cuota sindical", "seg social a pagar", "obra social a pagar", "lrt a pagar",
        "seguro de vida", "retencion ganancias", "retencion de ganancias",
    )
    for palabra in _PALABRAS_HABER:
        if palabra in t:
            return "Haber"
    if "cargas sociales" in t and ("pagar" in t or cod.startswith("213")):
        return "Haber"

    # Partida «a Cuenta…» del asiento (tras códigos 11418.0 / 11418).
    texto_sin_cod = re.sub(
        r"^(?:\d{5}(?:\.0+)?\s+|\.0+\s*)+", "", raw_lower,
    ).strip()
    if texto_sin_cod.startswith("a ") or re.search(
        r"\ba\s+(retencion|percepcion|saldo|ingresos|ajuste|impuesto)\b", t,
    ):
        return "Haber"
    if "pagar" in t:
        return "Haber"

    if imp_norm in ("sueldos", "sueldo"):
        _GASTO_SUELDOS = (
            "sueldos y jornales", "jornales", "sueldo bruto", "remuneracion",
            "contribuciones patronales", "aseguradora", "art ", "cargas sociales",
            "ajuste por redondeo",
        )
        for palabra in _GASTO_SUELDOS:
            if palabra in t and "pagar" not in t:
                return "Debe"

    return "Debe"


def _es_fila_encabezado_balance(concepto: str) -> bool:
    """Filas de encabezado que no son movimientos contables."""
    if not concepto or not str(concepto).strip():
        return True
    norm = _normalizar_texto(str(concepto))
    if norm in ("concepto", "detalle", "cuenta", "descripcion", "descripción", "mes", "periodo", "rubro"):
        return True
    if etiqueta_periodo_desde_celda(concepto):
        return True
    return False


_TITULOS_RUIDO_BANCO_EXTRACTO = frozenset({
    "depositos", "depósitos",
    "ingresos", "retiros y debitos", "saldo al inicio", "saldo final",
    "saldo s resumen", "saldo s/ resumen", "diferencia c resumen",
    "diferencia c/ resumen", "comisiones y gtos", "representante afip",
    "banco santander rio", "afip", "sicore", "caja", "alquileres",
    "compras tarjeta debito", "transferencia electronica",
})


def _inferir_tipo_movimiento_banco(concepto: str, codigo: str, fila_idx: int) -> str:
    """Heurística Debe/Haber para extractos bancarios con una columna por mes."""
    cod = str(codigo or "").strip()
    if cod.startswith(("213", "214", "11203")):
        return "Debe"
    if cod.startswith(("425", "424", "114")):
        return "Debe"
    if cod.startswith("11301"):
        return "Haber"
    if cod.startswith("11101"):
        return "Debe"
    if cod == "99999":
        return "Haber" if fila_idx < 17 else "Debe"
    if cod.startswith("111"):
        return "Debe"
    return "Debe"


def _es_fila_ruido_banco_extracto(concepto: str, codigo: str, fila_idx: int) -> bool:
    """Excluye totales, cabeceras y líneas agregadas de solapas tipo extracto bancario."""
    if _es_fila_ruido_balance(concepto):
        return True
    norm = _normalizar_texto(str(concepto or ""))
    if not norm:
        return True
    if any(
        kw in norm
        for kw in (
            "oftalmolog", "ejercicio cerrado", "mar del plata", "representante afip",
            "cuit 33", "cbu 15", "cbu 0",
        )
    ):
        return True
    if norm in _TITULOS_RUIDO_BANCO_EXTRACTO:
        return True
    if "saldo" in norm and any(k in norm for k in ("inicio", "final", "resumen")):
        return True
    cod = str(codigo or "").strip()
    if cod.startswith("11104") and any(
        kw in norm for kw in ("cta. cte", "santander rio cta", "a banco santander")
    ):
        return True
    return False


_PATRONES_RUIDO_BALANCE = (
    r"^total\b",
    r"^sub\s*total\b",
    r"^subtotal\b",
    r"^suma\b",
    r"^gran total\b",
    r"^total general\b",
    r"^total debe\b",
    r"^total haber\b",
    r"^control\b",
    r"verificacion",
    r"validacion",
    r"cuadro de control",
    r"^diferencia\b",
    r"^resultado\b",
    r"^neto\b",
    r"^saldo inicial\b",
    r"^saldo final\b",
    r"^check\b",
    r"^posicion\b",
    r"\bposicion\b",
    r"^ejercicio\b",
    r"\bejercicio\b",
    r"^observacion",
    r"^nota\b",
    r"^---+$",
    r"^===+$",
    r"\bpayway\b",
    r"asiento de apertura",
)


def _es_fila_ruido_balance(concepto: str) -> bool:
    """
    Purificador anti-ruido: excluye totales, controles y filas intermedias
    de sábanas masivas antes de inyectar en la grilla.
    """
    if _es_fila_encabezado_balance(concepto):
        return True
    norm = _normalizar_texto(str(concepto))
    if not norm or len(norm) < 2:
        return True
    # Notas de crédito: no son "nota" de pie; van al asiento aunque resten en neto.
    es_nc = _es_concepto_nota_credito(concepto)
    for patron in _PATRONES_RUIDO_BALANCE:
        if es_nc and patron == r"^nota\b":
            continue
        if re.search(patron, norm):
            return True
    if re.fullmatch(r"[\d./\-]+", norm):
        # Código Tango solo (11419 / 11419.0) no es ruido.
        if re.fullmatch(r"\d{5}", norm) or re.fullmatch(r"\d{5}\.0+", norm):
            return False
        return True
    if not _RE_CODIGO_CUENTA_CONCEPTO.match(str(concepto).strip()):
        titulos_seccion = (
            "ingresos brutos", "iva", "debito fiscal", "credito fiscal",
            "retenciones y percepciones", "liquidacion", "liquidación",
        )
        if norm in titulos_seccion or norm.endswith(":"):
            return True
    return False


def _descripcion_concepto_balance_limpia(concepto: str, codigo: str | None = None) -> str:
    """Texto de concepto sin códigos/flotantes repetidos (legible en grilla)."""
    raw = str(concepto or "").strip()
    if not raw:
        return ""
    _, resto = extraer_codigo_cuenta_tango_desde_concepto(raw)
    texto = (resto or raw).strip()
    texto = re.sub(r"^(?:\d{5}(?:\.0+)?\s+)+", "", texto).strip()
    texto = re.sub(r"\.0+\b", "", texto).strip()
    if codigo and texto.startswith(str(codigo)):
        texto = texto[len(str(codigo)):].strip(" -–—")
    return texto or raw


def _es_linea_proyectable_balance(concepto: str, codigo: str | None) -> bool:
    """Líneas de asiento con cuenta Tango (aunque el mes esté en $0)."""
    if _es_fila_ruido_balance(concepto):
        return False
    t = _normalizar_texto(concepto)
    if not t:
        return False
    # Títulos de bloque / razón social: no son movimientos.
    if any(
        k in t
        for k in (
            "asiento de", "devengamiento de", "oftalmolog", "s.r.l", "srl",
            "ejercicio cerrado", "representante afip", "papeles de trabajo",
            "estados contables",
        )
    ):
        return False
    # Razón social suelta (sin cuenta ni «a …»).
    cod = str(codigo or "").strip()
    if cod in ("99999",) or not re.fullmatch(r"\d{5}", cod):
        if not any(
            k in t
            for k in (
                "ajuste por redondeo", "a saldo a favor", "saldo a favor imp",
                "impuesto determinado", "impuesto sobre los ingresos brutos",
                "ingresos brutos a pagar", "iva a pagar",
                "a retencion", "a percepcion", "retenciones bancarias", "sircreb",
                "a impuesto",
            )
        ) and not t.startswith("a "):
            return False
        return True
    return True


def escanear_carpeta_cliente(cuit: str, ruta_raiz: str | Path | None = None) -> tuple[list[Path], Optional[Path], Optional[Path]]:
    """
    Escanea la carpeta del cliente buscando PDFs de extractos, archivo de compras y plan de cuentas.
    Retorna (rutas_pdfs, ruta_compras, ruta_cuentas).
    """
    raiz = Path(ruta_raiz) if ruta_raiz else RUTA_RAIZ_CLIENTES
    carpeta_cliente = raiz / cuit
    
    rutas_pdfs = []
    ruta_compras = None
    ruta_cuentas = None
    
    if not carpeta_cliente.exists():
        return rutas_pdfs, ruta_compras, ruta_cuentas
        
    for archivo in carpeta_cliente.iterdir():
        if not archivo.is_file() or archivo.name.startswith("~$"):
            continue
            
        nombre_lower = archivo.name.lower()
        if nombre_lower.endswith(".pdf"):
            rutas_pdfs.append(archivo)
        elif nombre_lower.endswith((".xlsx", ".xls", ".csv")):
            if "compras" in nombre_lower:
                ruta_compras = archivo
            elif "cuentas" in nombre_lower or "plan" in nombre_lower:
                ruta_cuentas = archivo
                
    # Ordenar PDFs alfabéticamente para intentar mantener cronología
    rutas_pdfs.sort()
    
    return rutas_pdfs[:MAX_PDFS_ANUALES], ruta_compras, ruta_cuentas


def _resolver_ruta_plantilla() -> Path:
    """Resuelve la ruta de la plantilla original (plantillas/ o planilla de conciliacion/)."""
    candidatas = [
        BASE_DIR / "plantillas" / "Planilla de Conciliacion.xlsx",
        BASE_DIR / "planilla de conciliacion" / "Planilla de Conciliacion.xlsx",
    ]
    for ruta in candidatas:
        if ruta.exists():
            return ruta
    return PLANTILLA_CONCILIACION


def _resolver_ruta_plantilla_asientos() -> Path:
    """Resuelve plantilla Excel de asientos contables para importación Tango."""
    candidatas = [
        BASE_DIR / "Asientos contables (11).xlsx",
        BASE_DIR / "asientos contables.xlsx",
        BASE_DIR / "informacion sobre tango" / "asientos contables.xlsx",
    ]
    for ruta in candidatas:
        if ruta.exists():
            return ruta
    return PLANTILLA_ASIENTOS_CONTABLES


def _resolver_ruta_plantilla_asientos_tango_vacio() -> Path:
    """Template vacío oficial Tango (raíz o plantillas/; Cloud-friendly)."""
    nombre = "Asientos contables VACIO PARA LLENAR E IMPORTAR.xlsx"
    candidatas = [
        BASE_DIR / nombre,
        BASE_DIR / "plantillas" / nombre,
        Path(__file__).resolve().parent / nombre,
        Path(__file__).resolve().parent / "plantillas" / nombre,
    ]
    for ruta in candidatas:
        if ruta.exists():
            return ruta
    return PLANTILLA_ASIENTOS_TANGO_VACIO


def leer_estructura_plantilla_asientos(ruta: str | Path | None = None) -> dict[str, list[str]]:
    """Lee encabezados exactos de las hojas Asientos contables y Renglones."""
    origen = Path(ruta) if ruta else _resolver_ruta_plantilla_asientos()
    if not origen.exists():
        raise FileNotFoundError(f"No se encontró plantilla de asientos: {origen}")
    wb = openpyxl.load_workbook(origen, data_only=True)
    estructura: dict[str, list[str]] = {}
    for nombre_hoja in (HOJA_ASIENTOS, HOJA_RENGLONES):
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            raise ValueError(f"La plantilla no contiene la hoja '{nombre_hoja}'")
        ws = wb[nombre_hoja]
        encabezados = []
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor is None and col > 1:
                break
            encabezados.append(str(valor or "").strip())
        estructura[nombre_hoja] = encabezados
    wb.close()
    return estructura


# Perfiles de bancos detectables automáticamente desde encabezados PDF
PERFILES_BANCO: dict[str, dict] = {
    "santander": {
        "keywords": ["santander rio", "banco santander", "santander", "067-0", "0720067020"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11103",
        "nombre_display": "Banco Santander",
    },
    "galicia": {
        "keywords": ["banco galicia", "bco galicia", "galicia", "0070127", "office banking"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11109",
        "nombre_display": "Banco Galicia",
    },
    "frances": {
        "keywords": [
            "bbva frances", "banco frances", "bbva banco", "bbva argentina",
            "banco bbva", "frances", "bbva", "332532", "332533",
        ],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11106",
        "nombre_display": "Banco Francés",
    },
    "credicoop": {
        "keywords": ["banco credicoop", "bco credicoop", "credicoop", "191011"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11104",
        "nombre_display": "Banco Credicoop",
    },
    "provincia": {
        "keywords": [
            "banco de la provincia de buenos aires",
            "banco de la provincia",
            "banco provincia",
            "bco provincia",
            "bapro",
            "014012",
            "provincia",
        ],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11105",
        "nombre_display": "Banco Provincia",
    },
    "macro": {
        "keywords": ["banco macro", "bco macro", "285012"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11103",
        "nombre_display": "Banco Macro",
    },
    "nacion": {
        "keywords": [
            "banco de la nacion", "banco nacion argentina", "banco nacion",
            "b. nacion", "bna",
        ],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11107",
        "nombre_display": "Banco Nación",
    },
    "icbc": {
        "keywords": ["icbc", "industrial and commercial", "industrial bank", "bco industrial"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11108",
        "nombre_display": "ICBC",
    },
    "hsbc": {
        "keywords": ["hsbc", "hongkong", "midland"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11110",
        "nombre_display": "HSBC",
    },
    "supervielle": {
        "keywords": ["supervielle", "banco supervielle", "bco supervielle"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11111",
        "nombre_display": "Banco Supervielle",
    },
    "ciudad": {
        "keywords": ["banco de la ciudad", "banco ciudad", "bco ciudad"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11112",
        "nombre_display": "Banco Ciudad",
    },
    "comafi": {
        "keywords": ["comafi", "banco comafi", "bco comafi"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11113",
        "nombre_display": "Banco Comafi",
    },
    "mercadopago": {
        "keywords": ["mercado pago", "mercadopago", "mp financiero", "mp credito", "mp credit", "mercadopago s.a"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11114",
        "nombre_display": "Mercado Pago",
    },
    "patagonia": {
        "keywords": ["patagonia", "banco patagonia", "bco patagonia"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11115",
        "nombre_display": "Banco Patagonia",
    },
    "brubank": {
        "keywords": ["brubank", "bru bank"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11116",
        "nombre_display": "Brubank",
    },
    "naranja": {
        "keywords": ["naranja x", "naranjax", "tarjeta naranja", "cuenta naranja"],
        "hoja_planilla": "Banco Santander",
        "cuenta_contable": "11117",
        "nombre_display": "Naranja X",
    },
    "desconocido": {
        "keywords": [],
        "hoja_planilla": "Banco Desconocido",
        "cuenta_contable": "11199",
        "nombre_display": "Banco Desconocido",
    },
}

# Etiquetas contables segregadas
ETIQUETAS = {
    "proveedor": "Proveedor habitual",
    "prestamo": "Préstamo bancario",
    "inversion": "Inversión / Plazo Fijo",
    "plan_fiscal": "Plan de pago fiscal",
    "impuesto": "Impuesto / Retención",
    "haberes": "Sueldos y haberes",
    "transferencia": "Transferencia",
    "comision": "Comisión bancaria",
    "anomalia": "Anomalía / Revisar",
}

MAPEO_CATEGORIAS_PLANILLA: dict[str, str] = {
    "proveedor": "Proveedores",
    "prestamo": "Capital del Prestamo",
    "inversion": "Plazo Fijo en pesos",
    "plan_fiscal": "IIBB a pagar",
    "impuesto": "IVA a Pagar",
    "haberes": "Sueldos y jornales a pagar ",
    "haber": "Sueldos y jornales a pagar ",
    "jornal": "Sueldos y jornales a pagar ",
    "sueldo": "Sueldos y jornales a pagar ",
    "transferencia": "Transferencia Recibida misma cuenta",
    "comision": "Comis. y Gtos  Bcarios.",
    "comis": "Comis. y Gtos  Bcarios.",
    "anomalia": "Trasnferencia de 3eros",
    "deudor": "Deduores por ventas",
    "cheque recibido": "Cheques Recibidos",
    "iva": "IVA a Pagar",
    "iibb": "IIBB a pagar",
    "alquiler": "Alquiler Pagado",
    "impuesto 25.413": "Imp. Déb. Ley.25413",
    "telecom": "Pago Telecom",
    "edea": "Edea",
    "interes prestamo": "Intereses Prestamos",
    "ganancias": "Ganancias Sociedades",
    "contribucion ss": "Contribucion SS (351)",
}

MAPEO_CUENTAS_KEYWORDS: dict[str, str] = {
    "proveedor": "21101",
    "prestamo": "22101",
    "inversion": "11202",
    "plan_fiscal": "21403",
    "haberes": "21401",
    "comision": "42501",
    "iva": "21404",
    "iibb": "21403",
    "alquiler": "51201",
    "telecom": "51202",
    "edea": "51203",
}

_lector_ocr = None
_lector_ocr_init_lock = threading.Lock()
_lector_ocr_run_lock = threading.Lock()


@dataclass
class MovimientoBanco:
    """Movimiento extraído del extracto bancario."""

    fecha: date
    comprobante: str
    descripcion: str
    debito: float
    credito: float
    saldo: Optional[float] = None
    pagina: int = 0
    banco: str = "santander"
    cuit_contraparte: str = ""
    categoria_contable: str = "transferencia"
    etiqueta: str = ETIQUETAS["transferencia"]
    es_anomalia: bool = False
    archivo_origen: str = ""

    @property
    def importe_neto(self) -> float:
        if self.credito > 0:
            return self.credito
        if self.debito > 0:
            return -self.debito
        return 0.0

    @property
    def importe_absoluto(self) -> float:
        return abs(self.importe_neto)


@dataclass
class MovimientoContable:
    fecha: date
    descripcion: str
    importe: float
    cuenta: str
    codigo_cuenta: str = ""
    referencia: str = ""


@dataclass
class BalanceMensual:
    """Saldos de control mensuales: SI + Ingresos − Egresos = SF."""

    anio: int
    mes: int
    saldo_inicial: Optional[float] = None
    saldo_final: Optional[float] = None
    saldo_resumen: Optional[float] = None
    total_ingresos: float = 0.0  # Créditos del mes
    total_egresos: float = 0.0  # Débitos del mes
    balance_cierra: bool = True
    diferencia_balance: float = 0.0
    archivo_origen: str = ""


@dataclass
class ResultadoConciliacion:
    conciliados: list[dict] = field(default_factory=list)
    solo_banco: list[MovimientoBanco] = field(default_factory=list)
    solo_contabilidad: list[MovimientoContable] = field(default_factory=list)
    resumen_por_categoria: dict[str, float] = field(default_factory=dict)
    resumen_anual_por_mes: dict[tuple[int, int], dict[str, float]] = field(default_factory=dict)
    saldos_por_mes: dict[tuple[int, int], BalanceMensual] = field(default_factory=dict)
    saldo_extracto: Optional[float] = None
    saldo_inicial: Optional[float] = None
    saldo_final: Optional[float] = None
    total_depositos: float = 0.0
    total_retiros: float = 0.0
    balance_cierra: bool = True
    diferencia_balance: float = 0.0
    mes_referencia: Optional[date] = None
    bancos_detectados: list[str] = field(default_factory=list)
    anomalias: list[dict] = field(default_factory=list)
    total_movimientos: int = 0
    movimientos_todos: list[MovimientoBanco] = field(default_factory=list)


_lector_ocr_unavailable = False


class _NoOpOcrReader:
    """Fallback cuando EasyOCR no carga: no crashea la app."""

    def readtext(self, *args, **kwargs):
        return []


def _es_entorno_cloud_ocr() -> bool:
    """Cloud/Linux: DPI bajo y sin paralelizar EasyOCR. Windows local = estudio."""
    if str(os.environ.get("ESTUDIO_FORCE_LOCAL", "")).strip().lower() in {"1", "true", "yes"}:
        return False
    flags = (
        os.environ.get("STREAMLIT_SHARING_MODE"),
        os.environ.get("STREAMLIT_CLOUD"),
        os.environ.get("IS_STREAMLIT_CLOUD"),
        os.environ.get("STREAMLIT_RUNTIME_ENVIRONMENT"),
    )
    if any(str(f).strip().lower() in {"1", "true", "yes", "cloud"} for f in flags if f):
        return True
    if Path("/mount/src").is_dir() or Path("/home/appuser").is_dir():
        return True
    if os.name != "nt":
        return True
    return False


def _obtener_lector_ocr():
    """Lazy singleton EasyOCR (o NoOp si falla). Compatible con @st.cache_resource."""
    global _lector_ocr, _lector_ocr_unavailable
    if _lector_ocr is not None:
        return _lector_ocr
    with _lector_ocr_init_lock:
        if _lector_ocr is not None:
            return _lector_ocr
        if _lector_ocr_unavailable:
            _lector_ocr = _NoOpOcrReader()
            return _lector_ocr
        try:
            import easyocr
            _lector_ocr = easyocr.Reader(["es"], gpu=False, verbose=False)
        except Exception as exc:
            _lector_ocr_unavailable = True
            print(f"[WARN] EasyOCR no disponible, OCR desactivado: {exc}", flush=True)
            _lector_ocr = _NoOpOcrReader()
    return _lector_ocr


def _lector_ocr_streamlit_cached():
    """Una sola carga del reader por proceso Streamlit."""
    try:
        import streamlit as st
    except Exception:
        return _obtener_lector_ocr()
    try:
        return _st_cached_easyocr_reader()
    except Exception:
        return _obtener_lector_ocr()


def _st_cached_easyocr_reader_impl():
    return _obtener_lector_ocr()


try:
    import streamlit as _st_for_ocr

    @_st_for_ocr.cache_resource(show_spinner="Cargando OCR (una sola vez)...")
    def _st_cached_easyocr_reader():
        return _st_cached_easyocr_reader_impl()
except Exception:
    def _st_cached_easyocr_reader():
        return _obtener_lector_ocr()


def _extraer_texto_nativo_pdf(ruta: Path, max_paginas: int | None = None) -> str:
    """Texto nativo vía pdfplumber y pymupdf (sin OCR)."""
    partes: list[str] = []

    with pdfplumber.open(ruta) as pdf:
        paginas = pdf.pages[:max_paginas] if max_paginas is not None else pdf.pages
        for pagina in paginas:
            texto = pagina.extract_text() or ""
            if texto.strip():
                partes.append(texto)

    if partes:
        return "\n".join(partes).strip()

    documento = fitz.open(ruta)
    try:
        limite = len(documento)
        if max_paginas is not None:
            limite = min(max_paginas, limite)
        for idx in range(limite):
            texto = documento[idx].get_text() or ""
            if texto.strip():
                partes.append(texto)
    finally:
        documento.close()

    return "\n".join(partes).strip()


def _extraer_texto_pdf_hibrido(ruta: Path, max_paginas: int | None = None) -> str:
    """
    Extracción híbrida: texto nativo (ms en PDFs digitales) y OCR solo si vacío.
    """
    texto = _extraer_texto_nativo_pdf(ruta, max_paginas=max_paginas)
    if texto.strip():
        return texto

    documento = fitz.open(ruta)
    try:
        lector = _obtener_lector_ocr()
        limite = len(documento)
        if max_paginas is not None:
            limite = min(max_paginas, limite)
        partes: list[str] = []
        dpi = 150 if _es_entorno_cloud_ocr() else 180
        for idx in range(limite):
            pix = documento[idx].get_pixmap(dpi=dpi)
            imagen = np.array(Image.open(io.BytesIO(pix.tobytes("png"))))
            with _lector_ocr_run_lock:
                for _bbox, t, _ in lector.readtext(imagen):
                    partes.append(t)
        return " ".join(partes)
    finally:
        documento.close()


def pdf_requiere_ocr(ruta: str | Path, max_paginas: int | None = None) -> bool:
    """Indica si el PDF carece de texto nativo y necesitará OCR."""
    return len(_extraer_texto_nativo_pdf(Path(ruta), max_paginas=max_paginas).strip()) < 30


def _normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c)).lower().strip()


def _parsear_importe(texto: str) -> Optional[float]:
    if not texto:
        return None
    limpio = texto.replace("$", "").replace(" ", "").strip()
    if re.search(r"\d,\d{2}$", limpio):
        limpio = limpio.replace(".", "").replace(",", ".")
    else:
        limpio = limpio.replace(",", "")
    try:
        return round(float(limpio), 2)
    except ValueError:
        return None


def _limpiar_monto(s: object) -> float:
    """
    Parsea montos argentinos y variantes OCR frecuentes.
    Acepta: 1.234,56 | 1234,56 | 1.234 | 12.50 | 1.234.56 (OCR de 1.234,56) | (1.234,56)
    """
    if not s:
        return 0.0
    txt = str(s).replace("$", "").replace("\xa0", " ").replace(" ", "").strip()
    if not txt or txt in ("-", "—", "–"):
        return 0.0
    negativo = False
    if txt.startswith("(") and txt.endswith(")"):
        negativo = True
        txt = txt[1:-1]
    elif txt.startswith("-"):
        negativo = True
        txt = txt[1:]
    # Solo dígitos / separadores
    if not re.fullmatch(r"[\d.,]+", txt):
        txt = re.sub(r"[^\d.,]", "", txt)
        if not txt:
            return 0.0

    # OCR argentino: 1.234.56 → miles con punto y decimales con punto (debería ser 1.234,56)
    if re.fullmatch(r"\d{1,3}(?:\.\d{3})+\.\d{2}", txt):
        partes = txt.split(".")
        txt = "".join(partes[:-1]) + "." + partes[-1]
    # Clásico AR: 1.234.567,89
    elif re.search(r"\d\.\d{3}", txt) and "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    # Miles AR sin decimales: 1.234 / 50.000
    elif re.fullmatch(r"\d{1,3}(?:\.\d{3})+", txt):
        txt = txt.replace(".", "")
    # Decimales con coma: 1234,56 o 1,5
    elif "," in txt and re.search(r",\d{1,2}$", txt):
        txt = txt.replace(".", "").replace(",", ".")
    # US / Excel: 1,234.56
    elif "," in txt and "." in txt and txt.rfind(".") > txt.rfind(","):
        txt = txt.replace(",", "")
    else:
        # 12.50 o 1000.00 (decimal punto) — no tocar el punto
        txt = txt.replace(",", "")

    try:
        val = round(float(txt), 2)
        return -val if negativo else val
    except ValueError:
        return 0.0


def _fecha_plausible_extracto(fecha: date) -> bool:
    """Filtra fechas OCR erróneas (ej. año 9265 por dígitos de importe)."""
    return ANIO_MIN_EXTRACTO <= fecha.year <= ANIO_MAX_EXTRACTO and 1 <= fecha.month <= 12


def _parsear_fecha(texto: str) -> Optional[date]:
    """Parsea fechas admitidas; devuelve None ante cadenas inválidas (ej. 'OFTALMOLOG')."""
    if not texto:
        return None
    try:
        texto = str(texto).strip()
        if not texto:
            return None
        # YYYYMMDD — solo si son exactamente 8 dígitos y forman fecha válida
        if re.fullmatch(r"\d{8}", texto):
            parsed = datetime.strptime(texto, "%Y%m%d").date()
            return parsed if _fecha_plausible_extracto(parsed) else None
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
            try:
                parsed = datetime.strptime(texto, fmt).date()
                return parsed if _fecha_plausible_extracto(parsed) else None
            except ValueError:
                continue
    except (ValueError, TypeError, OverflowError):
        return None
    return None


def _extraer_lineas_pdf(ruta: Path, permitir_ocr: bool = True) -> list[str]:
    """Obtiene líneas del PDF: nativo primero, OCR solo si no hay texto."""
    texto_nativo = _extraer_texto_nativo_pdf(ruta)
    if texto_nativo.strip():
        return [l.strip() for l in texto_nativo.splitlines() if l.strip()]

    if not permitir_ocr:
        return []

    documento = fitz.open(ruta)
    try:
        lineas: list[str] = []
        for idx in range(len(documento)):
            lineas.extend(t for _, t in _ocr_pagina(documento[idx]))
        return lineas
    finally:
        documento.close()


def _extraer_saldos_desde_lineas(lineas: list[str]) -> tuple[Optional[float], Optional[float]]:
    """
    Busca Saldo Inicial y Saldo Final / Saldo total en cuentas en el extracto.
    Retorna (saldo_inicial, saldo_resumen_final).
    """
    saldo_inicial: Optional[float] = None
    saldo_resumen: Optional[float] = None
    patron_importe = re.compile(r"\$\s*[\d\.\,]+")

    for linea in lineas:
        lower = _normalizar_texto(linea)
        importes = [_parsear_importe(m.group()) for m in patron_importe.finditer(linea)]
        importes = [v for v in importes if v is not None]

        if not importes:
            continue

        if "saldo inicial" in lower or "saldo anterior" in lower:
            saldo_inicial = importes[-1]
        elif any(
            k in lower
            for k in (
                "saldo total en cuentas",
                "saldo final",
                "saldo de cierre",
                "saldo al cierre",
            )
        ):
            saldo_resumen = importes[-1]

    return saldo_inicial, saldo_resumen


def _calcular_totales_movimientos(movimientos: list[MovimientoBanco]) -> tuple[float, float]:
    """Suma Ingresos (créditos) y Egresos (débitos) del período."""
    ingresos = sum(m.credito for m in movimientos if m.credito > 0)
    egresos = sum(m.debito for m in movimientos if m.debito > 0)
    return round(ingresos, 2), round(egresos, 2)


def _validar_balance(
    saldo_inicial: Optional[float],
    ingresos: float,
    egresos: float,
    saldo_final: Optional[float],
) -> tuple[bool, float]:
    """
    Valida la ecuación contable mensual:
    Saldo Inicial + Ingresos − Egresos = Saldo Final
    """
    if saldo_inicial is None or saldo_final is None:
        return True, 0.0
    calculado = round(saldo_inicial + ingresos - egresos, 2)
    diferencia = round(saldo_final - calculado, 2)
    return abs(diferencia) <= TOLERANCIA_IMPORTE, diferencia


def _calcular_balances_mensuales_encadenados(
    movimientos: list[MovimientoBanco],
    saldos_pdf: dict[tuple[int, int], BalanceMensual],
) -> dict[tuple[int, int], BalanceMensual]:
    """
    Calcula el balance de cada mes a partir de los movimientos consolidados.
    Fórmula: Saldo Inicial + Ingresos − Egresos = Saldo Final.
    El Saldo Final del mes N pasa automáticamente como Saldo Inicial del mes N+1.
    """
    por_mes: dict[tuple[int, int], list[MovimientoBanco]] = {}
    for mov in movimientos:
        if not _fecha_plausible_extracto(mov.fecha):
            continue
        clave = (mov.fecha.year, mov.fecha.month)
        por_mes.setdefault(clave, []).append(mov)

    if not por_mes:
        return dict(saldos_pdf)

    meses_ordenados = sorted(por_mes.keys())
    balances: dict[tuple[int, int], BalanceMensual] = {}
    saldo_cadena: Optional[float] = None

    for clave in meses_ordenados:
        anio, mes = clave
        movs_mes = por_mes[clave]
        ingresos, egresos = _calcular_totales_movimientos(movs_mes)
        pdf_bal = saldos_pdf.get(clave)

        # Saldo inicial: encadenado desde mes anterior, o del PDF en el primer mes
        if saldo_cadena is not None:
            saldo_ini = saldo_cadena
        elif pdf_bal and pdf_bal.saldo_inicial is not None:
            saldo_ini = pdf_bal.saldo_inicial
        else:
            saldo_ini = None

        # Saldo final = SI + Ingresos − Egresos (regla contable obligatoria)
        if saldo_ini is not None:
            saldo_fin = round(saldo_ini + ingresos - egresos, 2)
        elif pdf_bal and pdf_bal.saldo_final is not None:
            saldo_fin = pdf_bal.saldo_final
        else:
            saldo_fin = None

        saldo_res = (pdf_bal.saldo_resumen if pdf_bal else None) or saldo_fin
        cierra, dif = _validar_balance(saldo_ini, ingresos, egresos, saldo_fin)

        # Validar también contra saldo del resumen bancario si difiere
        if saldo_res is not None and saldo_fin is not None:
            dif_resumen = round(saldo_res - saldo_fin, 2)
            if abs(dif_resumen) > TOLERANCIA_IMPORTE:
                cierra = False
                dif = max(abs(dif), abs(dif_resumen))

        balances[clave] = BalanceMensual(
            anio=anio,
            mes=mes,
            saldo_inicial=saldo_ini,
            saldo_final=saldo_fin,
            saldo_resumen=saldo_res,
            total_ingresos=ingresos,
            total_egresos=egresos,
            balance_cierra=cierra,
            diferencia_balance=dif,
            archivo_origen=pdf_bal.archivo_origen if pdf_bal else "",
        )

        if saldo_fin is not None:
            saldo_cadena = saldo_fin

    return balances


def _mes_predominante(movimientos: list[MovimientoBanco]) -> tuple[int, int]:
    """Determina el mes con más movimientos (solo fechas plausibles)."""
    conteo: dict[tuple[int, int], int] = {}
    for m in movimientos:
        if not _fecha_plausible_extracto(m.fecha):
            continue
        clave = (m.fecha.year, m.fecha.month)
        conteo[clave] = conteo.get(clave, 0) + 1
    if not conteo:
        hoy = date.today()
        return hoy.year, hoy.month
    return max(conteo, key=conteo.get)


def _construir_balance_mensual(
    movimientos: list[MovimientoBanco],
    lineas_pdf: list[str],
    archivo: str,
) -> BalanceMensual:
    """Construye el balance de control para un extracto mensual."""
    anio, mes = _mes_predominante(movimientos)
    saldo_ini, saldo_res = _extraer_saldos_desde_lineas(lineas_pdf)
    ingresos, egresos = _calcular_totales_movimientos(movimientos)

    saldos_mov = [m.saldo for m in movimientos if m.saldo is not None]
    saldo_resumen_pdf = saldo_res

    # Derivar saldo inicial si no viene en el encabezado del PDF
    if saldo_ini is None and saldos_mov and movimientos:
        primer = movimientos[0]
        if primer.saldo is not None:
            if primer.credito > 0:
                saldo_ini = round(primer.saldo - primer.credito, 2)
            elif primer.debito > 0:
                saldo_ini = round(primer.saldo + primer.debito, 2)

    # Saldo Final = Saldo Inicial + Ingresos − Egresos
    if saldo_ini is not None:
        saldo_final = round(saldo_ini + ingresos - egresos, 2)
    elif saldo_resumen_pdf is not None:
        saldo_final = saldo_resumen_pdf
    elif saldos_mov:
        saldo_final = saldos_mov[-1]
    else:
        saldo_final = None

    cierra, dif = _validar_balance(saldo_ini, ingresos, egresos, saldo_final)

    return BalanceMensual(
        anio=anio,
        mes=mes,
        saldo_inicial=saldo_ini,
        saldo_final=saldo_final,
        saldo_resumen=saldo_resumen_pdf or saldo_final,
        total_ingresos=ingresos,
        total_egresos=egresos,
        balance_cierra=cierra,
        diferencia_balance=dif,
        archivo_origen=archivo,
    )


def _copiar_estilo_celda(origen, destino) -> None:
    """Copia formato visual de una celda openpyxl a otra."""
    if origen.has_style:
        destino.font = copy(origen.font)
        destino.border = copy(origen.border)
        destino.fill = copy(origen.fill)
        destino.number_format = copy(origen.number_format)
        destino.protection = copy(origen.protection)
        destino.alignment = copy(origen.alignment)


def _es_formula(valor) -> bool:
    return isinstance(valor, str) and valor.startswith("=")


def _escribir_celda_valor(hoja, celda_addr: str, valor, celda_estilo_ref: str, forzar: bool = False) -> None:
    """Escribe un valor numérico preservando el estilo. Si forzar=True, reemplaza fórmulas."""
    celda = hoja[celda_addr]
    if _es_formula(celda.value) and not forzar:
        return
    ref = hoja[celda_estilo_ref]
    _copiar_estilo_celda(ref, celda)
    celda.value = valor


def _escribir_celda_formula(hoja, celda_addr: str, formula: str, celda_estilo_ref: str) -> None:
    """Escribe una fórmula preservando estilo (solo si la celda no es fórmula de plantilla distinta)."""
    celda = hoja[celda_addr]
    ref = hoja[celda_estilo_ref]
    _copiar_estilo_celda(ref, celda)
    celda.value = formula


def _columna_anterior(col_letra: str) -> Optional[str]:
    """Devuelve la columna anterior (C->B no válida para meses; D->C)."""
    idx = openpyxl.utils.column_index_from_string(col_letra)
    if idx <= 3:
        return None
    return get_column_letter(idx - 1)


def _extraer_texto_encabezado(ruta: Path, max_paginas: int = 2) -> str:
    """Extrae texto de las primeras páginas para detectar el banco."""
    texto = _extraer_texto_pdf_hibrido(ruta, max_paginas=max_paginas)
    return _normalizar_texto(texto)


def detectar_banco_pdf(ruta: str | Path) -> str:
    """
    Inspecciona las primeras 3 páginas del PDF y el nombre del archivo para deducir el banco.
    Devuelve clave del perfil (santander, galicia, etc.) o 'desconocido' si no hay match.
    """
    ruta = Path(ruta)
    # Texto de las primeras 3 páginas + nombre del archivo (el prefijo ec_ preserva el nombre original)
    texto_pdf = _extraer_texto_encabezado(ruta, max_paginas=3)
    # Extraer la parte del nombre original del prefijo ec_{nombre}_...
    nombre_arch = _normalizar_texto(ruta.name)
    # Eliminar prefijo ec_ y sufijo de hash para recuperar el nombre original
    nombre_sin_prefijo = nombre_arch.replace("ec_", "").split("_")[:-1]
    nombre_original = " ".join(nombre_sin_prefijo)
    texto_combinado = f"{texto_pdf} {nombre_original}"

    mejor_banco = ""
    mejor_puntaje = 0

    for clave, perfil in PERFILES_BANCO.items():
        puntaje = sum(1 for kw in perfil["keywords"] if kw in texto_combinado)
        if puntaje > mejor_puntaje:
            mejor_puntaje = puntaje
            mejor_banco = clave

    return mejor_banco if mejor_puntaje > 0 else "desconocido"


_PATTERNS_N_PRESTAMO = [
    re.compile(r"[Pp]r[eé]stamo\s*[Nn][°ºo]?\s*:?\s*([\w][\w\-/]+)"),
    re.compile(r"[Cc]r[eé]dito\s*[Nn][°ºo]?\s*:?\s*([\w][\w\-/]+)"),
    re.compile(r"[Oo]peraci[oó]n\s*[Nn][°ºo]?\s*:?\s*([\w][\w\-/]+)"),
    re.compile(r"[Cc]ontrato\s*[Nn][°ºo]?\s*:?\s*([\w][\w\-/]+)"),
    re.compile(r"[Nn][°ºo]\s*[Oo]peraci[oó]n\s*:?\s*([\w][\w\-/]+)"),
    re.compile(r"N[°º]\s*(\d{4,})"),
    re.compile(r"[Nn]ro\.?\s*(\d{4,})"),
    re.compile(r"[Cc]uenta\s*[Nn][°ºo]?\s*:?\s*([\w][\w\-/]+)"),
]


def _extraer_numero_prestamo_pdf(ruta: Path, texto_pdf: str = "") -> str:
    """
    Busca el número de contrato/préstamo dentro del texto de las primeras 2 páginas del PDF.
    Si no lo encuentra, devuelve el stem del nombre del archivo (ya preservado con el nombre original).
    """
    texto = texto_pdf or _extraer_texto_encabezado(ruta, max_paginas=2)
    for patron in _PATTERNS_N_PRESTAMO:
        m = patron.search(texto)
        if m:
            candidato = m.group(1).strip().rstrip(".,;:")
            # Filtrar candidatos muy cortos o que sean solo puntuación
            if len(candidato) >= 3 and re.search(r"\d", candidato):
                return candidato

    # Fallback: usar el nombre original del archivo (sin el prefijo ec_ ni el hash final)
    stem = ruta.stem
    # Si viene de _guardar_upload, el formato es "ec_{nombre_original}_{hash}"
    if stem.startswith("ec_"):
        partes = stem[3:].rsplit("_", 1)
        nombre_limpio = partes[0] if partes else stem[3:]
        if nombre_limpio:
            return nombre_limpio
    return stem or "Prestamo_1"


def _ocr_pagina(pagina_fitz, dpi: int = 250) -> list[tuple[float, str]]:
    lector = _obtener_lector_ocr()
    pix = pagina_fitz.get_pixmap(dpi=dpi)
    img_pil = Image.open(io.BytesIO(pix.tobytes("png")))

    # Corrección de rotación por metadata de fitz (CW → PIL usa CCW → invertir signo)
    rot_metadata = getattr(pagina_fitz, "rotation", 0) or 0
    if rot_metadata != 0:
        img_pil = img_pil.rotate(-rot_metadata, expand=True)

    # Heurística: imagen más ancha que alta indica contenido landscape dentro de portrait
    w, h = img_pil.size
    if rot_metadata == 0 and w > h:
        img_pil = img_pil.rotate(90, expand=True)

    def _correr_ocr(img):
        return lector.readtext(np.array(img))

    resultados = _correr_ocr(img_pil)

    # Si confianza promedio muy baja (<0.35), rotar 90° adicionales y quedarse con el mejor
    if resultados:
        conf_prom = sum(c for _, _, c in resultados) / len(resultados)
        if conf_prom < 0.35:
            for angulo in (90, 180, 270):
                img_rot = img_pil.rotate(angulo, expand=True)
                r2 = _correr_ocr(img_rot)
                if r2:
                    conf2 = sum(c for _, _, c in r2) / len(r2)
                    if conf2 > conf_prom + 0.1:
                        resultados = r2
                        conf_prom = conf2
                        break

    filas: dict[int, list[tuple[float, str]]] = {}
    for bbox, texto, _ in resultados:
        y_centro = (bbox[0][1] + bbox[2][1]) / 2
        clave = int(y_centro / 18) * 18
        filas.setdefault(clave, []).append((bbox[0][0], texto))
    return [(y, " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0]))) for y in sorted(filas)]


def _extraer_movimientos_desde_texto(
    lineas: list[str], pagina: int, banco: str = "santander", archivo: str = ""
) -> list[MovimientoBanco]:
    """
    Interpreta líneas OCR/texto plano del extracto bancario.
    Soporta:
      - Montos con $ (Santander y similares)
      - Montos ARS sin $ (1.234,56)
      - Marcadores D/H (Nación, BBVA/Francés y variantes)
    """
    movimientos: list[MovimientoBanco] = []
    patron_fecha = re.compile(
        r"\b(\d{2}/\d{2}/\d{2,4}|\d{2}-\d{2}-\d{2,4}|\d{8})\b"
    )
    # $1.234,56  |  1.234,56 D/H  |  D 1.234,56
    patron_importe_dh = re.compile(
        r"(?:\$\s*)?(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|\d+\.\d{2})"
        r"(?:\s*([DHdh]))?\b"
        r"|"
        r"\b([DHdh])\s*\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})"
    )
    patron_cuit = re.compile(r"\b(20|23|24|27|30|33|34)\-?\d{8}\-?\d\b")
    fecha_movimiento_anterior: Optional[date] = None
    usa_dh = banco in {"nacion", "frances", "bbva", "macro", "credicoop", "provincia", "icbc", "hsbc"}

    i = 0
    while i < len(lineas):
        linea = lineas[i]
        match_fecha = patron_fecha.search(linea)
        # Fecha sola: anclar día y continuar (no exigir fecha en cada egreso/ingreso)
        if match_fecha and not _montos_en_linea_extracto(linea):
            solo = linea[match_fecha.end():].strip()
            if len(re.sub(r"\d", "", solo)) < 4 and not re.search(
                r"(?i)impuesto|transf|pago|compra|debito|credito|debin|retiro", solo
            ):
                f_ancla = _parsear_fecha(match_fecha.group(1))
                if f_ancla and _fecha_plausible_extracto(f_ancla):
                    fecha_movimiento_anterior = f_ancla
                i += 1
                continue
        if not match_fecha:
            # Sin fecha en la línea: si hay día anclado y la línea parece movimiento, usarlo
            if fecha_movimiento_anterior is None:
                i += 1
                continue
            if not (
                _es_concepto_transferencia_terceros(linea)
                or _RE_CONCEPTO_MOV_EXT.match(linea)
                or re.search(r"(?i)\b(pesos|\$)\b", linea)
                or re.search(r"\d{1,3}(?:\.\d{3})*,\d{2}", linea)
            ):
                i += 1
                continue
            fecha = fecha_movimiento_anterior
            texto_extra_descripcion = ""
            # Sintético: no había match_fecha; seguir flujo con fecha anclada
            bloque = [linea]
            j = i + 1
            es_terceros = _es_concepto_transferencia_terceros(linea)
            max_lineas_bloque = 8 if es_terceros else 5
            while j < len(lineas) and j < i + max_lineas_bloque:
                if patron_fecha.search(lineas[j]):
                    break
                bloque.append(lineas[j])
                if not es_terceros and _es_concepto_transferencia_terceros(lineas[j]):
                    es_terceros = True
                    max_lineas_bloque = 8
                j += 1
            texto_bloque = " ".join(bloque)
            # Reutilizar el armado de importes más abajo vía goto-like: set vars and jump
            importes_tipados = []
            for m in patron_importe_dh.finditer(texto_bloque):
                if m.group(1):
                    val = _limpiar_monto(m.group(1))
                    marca = (m.group(2) or "").upper() or None
                else:
                    val = _limpiar_monto(m.group(4))
                    marca = (m.group(3) or "").upper() or None
                if val and val > 0:
                    importes_tipados.append((val, marca))
            debito = credito = saldo = 0.0
            comprobante = ""
            cuit_match = patron_cuit.search(texto_bloque)
            cuit_contraparte = cuit_match.group(0).replace("-", "") if cuit_match else ""
            desc_lower = _normalizar_texto(texto_bloque)
            if len(importes_tipados) >= 2:
                saldo = importes_tipados[-1][0]
                monto_mov, marca = importes_tipados[-2]
                if marca == "D":
                    debito = monto_mov
                elif marca == "H":
                    credito = monto_mov
                else:
                    debito, credito = _clasificar_debito_credito_extracto(
                        desc_lower, monto_mov, None, None
                    )
            elif len(importes_tipados) == 1:
                monto_mov, marca = importes_tipados[0]
                if marca == "D":
                    debito = monto_mov
                elif marca == "H":
                    credito = monto_mov
                else:
                    debito, credito = _clasificar_debito_credito_extracto(
                        desc_lower, monto_mov, None, None
                    )
            nums = re.findall(r"\b\d{5,12}\b", texto_bloque)
            if nums:
                comprobante = nums[0]
            descripcion = patron_fecha.sub("", texto_bloque)
            descripcion = patron_importe_dh.sub("", descripcion)
            descripcion = re.sub(r"\b\d{5,12}\b", "", descripcion)
            descripcion = re.sub(r"\s+", " ", descripcion).strip(" -|")
            if debito > 0 or credito > 0:
                movimientos.append(
                    MovimientoBanco(
                        fecha=fecha,
                        comprobante=comprobante,
                        descripcion=descripcion or "Sin descripción",
                        debito=debito,
                        credito=credito,
                        saldo=saldo if saldo else None,
                        banco=banco,
                        archivo_origen=archivo,
                        pagina=pagina,
                        cuit_contraparte=cuit_contraparte,
                    )
                )
            i = j if j > i else i + 1
            continue

        fecha = _parsear_fecha(match_fecha.group(1))
        texto_extra_descripcion = ""
        if not fecha:
            if fecha_movimiento_anterior is None:
                i += 1
                continue
            fecha = fecha_movimiento_anterior
            texto_extra_descripcion = linea.strip() + " "
        else:
            if not _fecha_plausible_extracto(fecha):
                if fecha_movimiento_anterior is None:
                    i += 1
                    continue
                fecha = fecha_movimiento_anterior
                texto_extra_descripcion = linea.strip() + " "
            else:
                fecha_movimiento_anterior = fecha

        bloque = [linea]
        j = i + 1
        # Transferencias a/de terceros suelen partir descripción / montos en varias líneas
        es_terceros = _es_concepto_transferencia_terceros(texto_extra_descripcion or linea)
        max_lineas_bloque = 8 if es_terceros else 5
        while j < len(lineas) and j < i + max_lineas_bloque:
            if patron_fecha.search(lineas[j]) and not es_terceros:
                break
            if patron_fecha.search(lineas[j]) and es_terceros and j > i + 1:
                break
            bloque.append(lineas[j])
            # Si ya aparece el concepto terceros en el bloque, ampliar ventana
            if not es_terceros and _es_concepto_transferencia_terceros(lineas[j]):
                es_terceros = True
                max_lineas_bloque = 8
            j += 1

        texto_bloque = " ".join(bloque)
        importes_tipados: list[tuple[float, str | None]] = []
        for m in patron_importe_dh.finditer(texto_bloque):
            if m.group(1):
                val = _limpiar_monto(m.group(1))
                marca = (m.group(2) or "").upper() or None
            else:
                val = _limpiar_monto(m.group(4))
                marca = (m.group(3) or "").upper() or None
            if val and val > 0:
                importes_tipados.append((val, marca))

        # Si el saldo no trae marca D/H y el movimiento sí, no confundir saldo con movimiento
        # (ya se toma penúltimo = movimiento, último = saldo cuando hay 2+)
        debito = credito = saldo = 0.0
        comprobante = ""
        cuit_match = patron_cuit.search(texto_bloque)
        cuit_contraparte = cuit_match.group(0).replace("-", "") if cuit_match else ""
        desc_lower = _normalizar_texto(texto_bloque)

        if len(importes_tipados) >= 2:
            saldo = importes_tipados[-1][0]
            monto_mov, marca = importes_tipados[-2]
            if marca == "D":
                debito = monto_mov
            elif marca == "H":
                credito = monto_mov
            else:
                debito, credito = _clasificar_debito_credito_extracto(
                    desc_lower, monto_mov, None, None
                )
                # Sin saldo previo en este parser de bloque: usar palabras + no forzar crédito
        elif len(importes_tipados) == 1:
            monto_mov, marca = importes_tipados[0]
            if marca == "D":
                debito = monto_mov
            elif marca == "H":
                credito = monto_mov
            else:
                debito, credito = _clasificar_debito_credito_extracto(
                    desc_lower, monto_mov, None, None
                )

        nums = re.findall(r"\b\d{5,12}\b", texto_bloque)
        if nums:
            comprobante = nums[0]

        descripcion = patron_fecha.sub("", texto_bloque)
        descripcion = patron_importe_dh.sub("", descripcion)
        descripcion = re.sub(r"\b\d{5,12}\b", "", descripcion)
        descripcion = re.sub(r"\s+", " ", descripcion).strip(" -|")
        if texto_extra_descripcion:
            descripcion = (texto_extra_descripcion + descripcion).strip()

        if debito > 0 or credito > 0:
            movimientos.append(
                MovimientoBanco(
                    fecha=fecha,
                    comprobante=comprobante,
                    descripcion=descripcion or "Sin descripción",
                    debito=debito,
                    credito=credito,
                    saldo=saldo if saldo else None,
                    pagina=pagina,
                    banco=banco,
                    cuit_contraparte=cuit_contraparte,
                    archivo_origen=archivo,
                )
            )
        i = j

    return movimientos


def extraer_datos_pdf_galicia(pdf_file) -> pd.DataFrame:
    """Lee tablas del PDF digital de Galicia y devuelve un DataFrame estructurado."""
    filas_totales: list[list] = []
    with pdfplumber.open(pdf_file) as pdf:
        for pagina in pdf.pages:
            tabla = pagina.extract_table()
            if not tabla:
                continue
            # Detectar orden de columnas por encabezado (Débito/Crédito a veces vienen invertidos)
            idx_map = {"fecha": 0, "desc": 1, "origen": 2, "credito": 3, "debito": 4, "saldo": 5}
            header_found = False
            for fila in tabla:
                if not fila:
                    continue
                cells = [("" if c is None else str(c)).replace("\r", " ").replace("\n", " ").strip() for c in fila]
                heads = [_normalizar_texto(c) for c in cells]
                if any("fecha" in h for h in heads) and any(
                    ("debit" in h or "credit" in h or "saldo" in h) for h in heads
                ):
                    header_found = True
                    for i, h in enumerate(heads):
                        if "fecha" in h:
                            idx_map["fecha"] = i
                        elif "desc" in h or "concepto" in h or "movimiento" in h:
                            idx_map["desc"] = i
                        elif "origen" in h or "comprob" in h or "referencia" in h:
                            idx_map["origen"] = i
                        elif "debit" in h or h in {"debe", "egreso"}:
                            idx_map["debito"] = i
                        elif "credit" in h or h in {"haber", "ingreso"}:
                            idx_map["credito"] = i
                        elif "saldo" in h:
                            idx_map["saldo"] = i
                    continue
                if cells and _normalizar_texto(cells[0]) == "fecha":
                    continue

                def _cell(key: str) -> str:
                    i = idx_map.get(key, 0)
                    return cells[i] if i < len(cells) else ""

                fecha_txt = _cell("fecha")
                # Filas de continuación (sin fecha): fusionar descripción / montos a la anterior
                if (not fecha_txt or not re.search(r"\d{2}[/-]\d{2}", fecha_txt)) and filas_totales:
                    prev = filas_totales[-1]
                    desc_extra = _cell("desc") or " ".join(c for c in cells if c and not re.fullmatch(r"[\d.,\-\s]+", c))
                    if desc_extra:
                        prev[1] = f"{prev[1]} {desc_extra}".strip()
                    for col_i, key in ((3, "credito"), (4, "debito"), (5, "saldo")):
                        val = _cell(key)
                        if val and (not prev[col_i] or str(prev[col_i]).strip() in ("", "-", "None")):
                            prev[col_i] = val
                    # Si la "continuación" trae montos y la fila anterior no, ya quedó arriba
                    continue

                if not fecha_txt:
                    continue
                if len(cells) < 4 and not header_found:
                    continue
                filas_totales.append([
                    fecha_txt,
                    _cell("desc"),
                    _cell("origen"),
                    _cell("credito"),
                    _cell("debito"),
                    _cell("saldo"),
                ])
    return pd.DataFrame(
        filas_totales,
        columns=["Fecha", "Descripción", "Origen", "Crédito", "Débito", "Saldo"],
    )


def extraer_movimientos_galicia_tabla(
    pdf_path: str | Path | bytes,
    archivo: str = "",
) -> list[MovimientoBanco]:
    """Convierte tablas nativas del extracto digital Galicia en MovimientoBanco."""
    movimientos: list[MovimientoBanco] = []
    try:
        df = extraer_datos_pdf_galicia(pdf_path)
    except Exception:
        return []
    if df.empty:
        return []

    nombre_archivo = archivo or (
        Path(pdf_path).name if isinstance(pdf_path, (str, Path)) else "galicia.pdf"
    )
    for _, row in df.iterrows():
        fecha = _parsear_fecha(str(row.get("Fecha", "")).strip())
        if not fecha or not _fecha_plausible_extracto(fecha):
            continue
        descripcion = str(row.get("Descripción", "") or "").strip()
        credito_raw = _limpiar_monto(row.get("Crédito"))
        debito_raw = _limpiar_monto(row.get("Débito"))
        # Galicia: egresos suelen venir como crédito NEGATIVO (resta = importe −)
        if credito_raw < 0 and abs(debito_raw) < 0.01:
            importe = round(credito_raw, 2)  # ya negativo
        elif debito_raw < 0 and abs(credito_raw) < 0.01:
            importe = round(debito_raw, 2)
        elif debito_raw > 0 and credito_raw <= 0:
            importe = -round(debito_raw, 2)
        elif credito_raw > 0:
            importe = round(credito_raw, 2)
        else:
            importe = None
        debito = abs(importe) if importe is not None and importe < 0 else 0.0
        credito = abs(importe) if importe is not None and importe > 0 else 0.0
        saldo_val = abs(_limpiar_monto(row.get("Saldo")))
        if credito <= 0 and debito <= 0:
            continue
        movimientos.append(
            MovimientoBanco(
                fecha=fecha,
                comprobante=str(row.get("Origen", "") or "").strip(),
                descripcion=descripcion or "Sin descripción",
                debito=debito if debito > 0 else 0.0,
                credito=credito if credito > 0 else 0.0,
                saldo=saldo_val if saldo_val else None,
                banco="galicia",
                archivo_origen=nombre_archivo,
            )
        )
    return movimientos


def normalizar_subdiario_tango(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza export de subdiarios Tango (Cuenta, Razón social, Debe, Haber)."""
    if df is None or df.empty:
        return pd.DataFrame(columns=["cuenta", "razon_social", "debe", "haber"])
    work = df.copy()
    work.columns = [_normalizar_texto(str(c)).replace(" ", "_") for c in work.columns]
    col_cuenta = next((c for c in work.columns if c == "cuenta" or c.startswith("codigo")), None)
    col_razon = next(
        (c for c in work.columns if "razon" in c or ("social" in c and "cuenta" not in c)),
        None,
    )
    col_debe = next((c for c in work.columns if "debe" in c), None)
    col_haber = next((c for c in work.columns if "haber" in c), None)
    if not col_cuenta:
        return pd.DataFrame(columns=["cuenta", "razon_social", "debe", "haber"])
    out = pd.DataFrame()
    out["cuenta"] = work[col_cuenta].astype(str).str.strip()
    out["razon_social"] = work[col_razon].astype(str).str.strip() if col_razon else ""
    out["debe"] = pd.to_numeric(work[col_debe], errors="coerce").fillna(0.0) if col_debe else 0.0
    out["haber"] = pd.to_numeric(work[col_haber], errors="coerce").fillna(0.0) if col_haber else 0.0
    return out


def cargar_subdiario_tango(archivo) -> pd.DataFrame:
    """Carga Excel/CSV de subdiarios Tango para conciliación bancaria."""
    nombre = getattr(archivo, "name", str(archivo)).lower()
    if nombre.endswith(".csv"):
        df = pd.read_csv(archivo)
    else:
        df = pd.read_excel(archivo)
    return normalizar_subdiario_tango(df)


def _nombre_coincide_descripcion_banco(razon: str, descripcion: str) -> bool:
    razon_norm = _normalizar_texto(razon)
    desc_norm = _normalizar_texto(descripcion)
    if not razon_norm or not desc_norm:
        return False
    if razon_norm in desc_norm:
        return True
    return any(palabra in desc_norm for palabra in razon_norm.split() if len(palabra) > 3)


def sugerir_cuenta_conciliacion_tango(
    descripcion: str,
    monto: float,
    es_acreditacion: bool,
    df_tango: pd.DataFrame | None,
    plan_cuentas: list[tuple[str, str]] | None = None,
) -> tuple[str, str]:
    """
    Cruza un movimiento bancario con subdiarios Tango (21101 proveedores / 11301 deudores)
    y reglas fijas de gastos. Retorna (codigo, descripcion); codigo vacío si no hay match.
    """
    plan = plan_cuentas or []
    desc_norm = _normalizar_texto(descripcion)
    monto = round(abs(float(monto or 0)), 2)
    if monto <= 0:
        return "", ""

    prefijo = "11301" if es_acreditacion else "21101"
    etiqueta_default = "Deudores por Ventas" if es_acreditacion else "Proveedores"

    if df_tango is not None and not df_tango.empty:
        sub = normalizar_subdiario_tango(df_tango)
        sub_filtrado = sub[sub["cuenta"].str.startswith(prefijo, na=False)]
        col_monto = "debe" if es_acreditacion else "haber"
        match_monto = sub_filtrado[sub_filtrado[col_monto].round(2) == monto]
        if not match_monto.empty:
            hit = _buscar_codigo_en_plan_banco(prefijo, plan)
            return hit if hit else (prefijo, etiqueta_default)
        for _, fila in sub_filtrado.iterrows():
            if _nombre_coincide_descripcion_banco(str(fila.get("razon_social", "")), descripcion):
                hit = _buscar_codigo_en_plan_banco(prefijo, plan)
                return hit if hit else (prefijo, etiqueta_default)

    reglas_fijas: list[tuple[tuple[str, ...], tuple[tuple[str, ...], ...]]] = [
        (("comision",), (("gasto", "bancari"), ("comis",), ("bancari",))),
        (("percep", "iva"), (("percep", "iva"),)),
        (("iva",), (("iva", "credito"), ("credito", "fiscal"))),
        (("sircreb",), (("sircreb",),)),
        (("ing", "brutos"), (("iibb", "retenc"), ("ing", "bruto"))),
        (("ley", "25413"), (("25413",), ("ley", "25413"))),
    ]
    for palabras, grupos_plan in reglas_fijas:
        if all(p in desc_norm for p in palabras):
            if plan:
                hit = _buscar_cuenta_plan_por_etiquetas(plan, *grupos_plan)
                if hit:
                    return hit
            fallbacks = {
                ("comision",): ("42400", "Gastos Bancarios"),
                ("percep", "iva"): ("11405", "Percepciones IVA"),
                ("iva",): ("11402", "IVA Crédito Fiscal"),
                ("sircreb",): ("11410", "Retención SIRCREB"),
                ("ing", "brutos"): ("11409", "Retención IIBB"),
                ("ley", "25413"): ("42405", "Impuesto Ley 25413"),
            }
            clave = tuple(palabras)
            if clave in fallbacks:
                cod, etiqueta = fallbacks[clave]
                hit = _buscar_codigo_en_plan_banco(cod, plan)
                return hit if hit else (cod, etiqueta)

    return "", ""


def conciliar_banco_con_tango(
    df_galicia: pd.DataFrame,
    df_tango: pd.DataFrame | None,
    plan_cuentas: list[tuple[str, str]] | None = None,
) -> list[dict]:
    """
    Cruza movimientos de Galicia con subdiarios Tango y devuelve líneas de asiento sugeridas.
    """
    lineas_asiento: list[dict] = []
    for _, row in df_galicia.iterrows():
        fecha_banco = row.get("Fecha")
        descripcion = str(row.get("Descripción", ""))
        credito_val = row.get("Crédito")
        debito_val = row.get("Débito")

        monto = 0.0
        tipo = ""
        es_acreditacion = False

        if credito_val is not None and str(credito_val).strip() not in ("", "None", "nan"):
            monto = _limpiar_monto(credito_val)
            if monto <= 0:
                continue
            tipo = "Haber"
            es_acreditacion = True
        elif debito_val is not None and str(debito_val).strip() not in ("", "None", "nan"):
            monto = _limpiar_monto(str(debito_val).replace("-", ""))
            if monto <= 0:
                continue
            tipo = "Debe"
            es_acreditacion = False
        else:
            continue

        codigo, desc_cuenta = sugerir_cuenta_conciliacion_tango(
            descripcion, monto, es_acreditacion, df_tango, plan_cuentas,
        )
        if codigo:
            cuenta_sugerida = f"{codigo} - {desc_cuenta}"
        else:
            cuenta_sugerida = "Seleccione Cuenta..."

        lineas_asiento.append({
            "fecha": fecha_banco,
            "concepto": _normalizar_texto(descripcion)[:40].strip(),
            "cuenta": cuenta_sugerida,
            "tipo": tipo,
            "monto": monto,
        })
    return lineas_asiento


# ---------------------------------------------------------------------------
# Recategorización Monotributo — facturas electrónicas AFIP (PDF / ZIP)
# ---------------------------------------------------------------------------

_RE_FECHA_AFIP = r"(\d{2}/\d{2}/\d{4})"

# Códigos AFIP: NC restan, ND/FC suman, para recategorización monotributo.
_CODIGOS_AFIP_NC = {"003", "008", "013", "053", "203", "208", "213"}
_CODIGOS_AFIP_ND = {"002", "007", "012", "052", "202", "207", "212"}
_CODIGOS_AFIP_FC = {
    "001", "006", "011", "051", "081", "082", "083",
    "201", "206", "211",
}

COLUMNAS_MONOTRIBUTO = [
    "Archivo",
    "Tipo",
    "Código AFIP",
    "Fecha Emisión",
    "Período Desde",
    "Período Hasta",
    "Concepto",
    "Importe Total",
    "Comprobante",
    "CAE",
]


def iter_pdfs_desde_uploads(archivos) -> list[tuple[str, bytes]]:
    """Extrae PDFs sueltos o contenidos en ZIP (solo memoria, sin disco)."""
    salida: list[tuple[str, bytes]] = []
    vistos_hash: set[str] = set()
    for uploaded in archivos or []:
        nombre = str(getattr(uploaded, "name", "archivo.pdf"))
        data = uploaded.getvalue() if hasattr(uploaded, "getvalue") else bytes(uploaded)
        nombre_lower = nombre.lower()
        if nombre_lower.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for info in zf.infolist():
                    if info.is_dir() or not info.filename.lower().endswith(".pdf"):
                        continue
                    interno = Path(info.filename).name or info.filename
                    raw = zf.read(info.filename)
                    digest = hashlib.sha1(raw).hexdigest()
                    if digest in vistos_hash:
                        continue
                    vistos_hash.add(digest)
                    salida.append((f"{nombre}::{interno}", raw))
        elif nombre_lower.endswith(".pdf"):
            digest = hashlib.sha1(data).hexdigest()
            if digest in vistos_hash:
                continue
            vistos_hash.add(digest)
            salida.append((nombre, data))
    return salida


def extraer_texto_factura_afip(pdf_bytes: bytes) -> str:
    """Texto plano del comprobante AFIP vía pdfplumber."""
    partes: list[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            if texto.strip():
                partes.append(texto)
    return "\n".join(partes)


def formatear_comprobante_tango(punto_venta: str | int, nro_cmp: str | int) -> str:
    """Formato Tango: 00002-00000142 (5 dígitos PV + 8 dígitos número)."""
    return f"{int(str(punto_venta).strip()):05d}-{int(str(nro_cmp).strip()):08d}"


def _buscar_fecha_afip(texto: str, patron: str) -> str:
    m = re.search(patron, texto, flags=re.IGNORECASE | re.DOTALL)
    return m.group(1) if m else ""


def _normalizar_codigo_afip(valor: str | int | None) -> str:
    digitos = re.sub(r"\D", "", str(valor or ""))
    if not digitos:
        return ""
    return digitos.zfill(3)[-3:]


def detectar_tipo_comprobante_afip(texto: str) -> tuple[str, str, int]:
    """
    Detecta tipo AFIP del PDF.
    Retorna (tipo_legible, codigo_afip_3dig, signo) donde signo es -1 para NC.
    """
    bloque = str(texto or "")
    codigo = ""

    for patron in (
        r"C[oó]digo\s*(?:N[oº°\.]+)?\s*[:\s]*(\d{1,3})\b",
        r"\bCod(?:igo|\.)?\s*[:\s]*(\d{1,3})\b",
        r"Tipo\s*(?:de\s*)?Comprobante[:\s]*(\d{1,3})\b",
    ):
        m = re.search(patron, bloque, flags=re.IGNORECASE)
        if m:
            codigo = _normalizar_codigo_afip(m.group(1))
            break

    # En comprobantes AFIP suele aparecer "CODIGO" y el número en la línea siguiente.
    if not codigo:
        m = re.search(
            r"C[oó]digo\s*(?:\n|\r\n|\s)+(\d{1,3})\b",
            bloque,
            flags=re.IGNORECASE,
        )
        if m:
            codigo = _normalizar_codigo_afip(m.group(1))

    es_nc_txt = bool(
        re.search(r"NOTA\s+DE\s+CR[EÉ]DITO", bloque, flags=re.IGNORECASE)
    )
    es_nd_txt = bool(
        re.search(r"NOTA\s+DE\s+D[EÉ]BITO", bloque, flags=re.IGNORECASE)
    )
    es_fc_txt = bool(
        re.search(r"\bFACTURA\b", bloque, flags=re.IGNORECASE)
    ) and not es_nc_txt and not es_nd_txt

    if codigo in _CODIGOS_AFIP_NC or (es_nc_txt and codigo not in _CODIGOS_AFIP_ND):
        return "Nota de Crédito", codigo or ("013" if es_nc_txt else ""), -1
    if codigo in _CODIGOS_AFIP_ND or es_nd_txt:
        return "Nota de Débito", codigo or ("012" if es_nd_txt else ""), 1
    if codigo in _CODIGOS_AFIP_FC or es_fc_txt or not codigo:
        letra = ""
        m_let = re.search(
            r"\bFACTURA\s+([ABC])\b|\b([ABC])\s*(?:\n|\r\n)\s*C[oó]digo",
            bloque,
            flags=re.IGNORECASE,
        )
        if m_let:
            letra = (m_let.group(1) or m_let.group(2) or "").upper()
        tipo = f"Factura {letra}".strip() if letra else "Factura"
        return tipo, codigo, 1

    # Código desconocido: por defecto suma (no NC)
    return f"Comprobante {codigo}", codigo, 1


def _extraer_cae_afip(texto: str) -> str:
    """Extrae CAE/CAI de comprobantes AFIP (CAE N°, C.A.E. Nº, etc.)."""
    m = re.search(
        r"(?:C\s*\.?\s*A\s*\.?\s*E\s*\.?|CAI)\b"
        r"[\sNnº°o.:\-]*"
        r"([0-9]{10,14})\b",
        texto or "",
        flags=re.IGNORECASE,
    )
    return m.group(1) if m else ""


def _clave_dedupe_monotributo(fila: dict) -> str:
    """Clave estable para eliminar comprobantes duplicados."""
    cae = str(fila.get("CAE") or "").strip()
    if cae:
        return f"cae:{cae}"
    tipo = str(fila.get("Tipo") or "").strip().upper()
    codigo = str(fila.get("Código AFIP") or "").strip()
    cmpte = str(fila.get("Comprobante") or "").strip()
    periodo = str(fila.get("Período Desde") or "").strip()
    importe = round(float(fila.get("Importe Total") or 0), 2)
    if cmpte:
        return f"cmp:{codigo}|{tipo}|{cmpte}|{periodo}|{importe}"
    archivo = Path(str(fila.get("Archivo") or "")).name.lower()
    return f"file:{archivo}|{codigo}|{periodo}|{importe}"


def deduplicar_comprobantes_monotributo(
    filas: list[dict],
) -> tuple[list[dict], int]:
    """Conserva la primera aparición de cada comprobante. Retorna (únicos, descartados)."""
    vistos: set[str] = set()
    unicos: list[dict] = []
    descartados = 0
    for fila in filas:
        clave = _clave_dedupe_monotributo(fila)
        if clave in vistos:
            descartados += 1
            continue
        vistos.add(clave)
        unicos.append(fila)
    return unicos, descartados


def parsear_factura_afip_texto(texto: str, archivo: str = "") -> dict | None:
    """
    Extrae campos clave de un comprobante electrónico AFIP desde texto PDF.
    Notas de crédito quedan con Importe Total negativo.
    Retorna None si faltan datos mínimos (período desde + importe).
    """
    if not texto or not str(texto).strip():
        return None

    bloque = re.sub(r"[ \t]+", " ", texto)
    bloque = re.sub(r"\n{2,}", "\n", bloque)
    tipo, codigo_afip, signo = detectar_tipo_comprobante_afip(texto)

    fecha_emision = _buscar_fecha_afip(
        bloque,
        rf"Fecha\s+de\s+Emisi[oó]n[:\s]*{_RE_FECHA_AFIP}",
    )
    periodo_desde = _buscar_fecha_afip(
        bloque,
        rf"Per[ií]odo\s+Facturado\s+Desde[:\s]*{_RE_FECHA_AFIP}",
    )
    periodo_hasta = ""
    m_hasta = re.search(
        rf"Per[ií]odo\s+Facturado\s+Desde[:\s]*{_RE_FECHA_AFIP}\s*(?:Hasta|al)[:\s]*{_RE_FECHA_AFIP}",
        bloque,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if m_hasta:
        periodo_desde = m_hasta.group(1)
        periodo_hasta = m_hasta.group(2)
    else:
        periodo_hasta = _buscar_fecha_afip(
            bloque,
            rf"(?:Hasta|al)[:\s]*{_RE_FECHA_AFIP}",
        )

    concepto = ""
    m_conc = re.search(
        r"Concepto[:\s]*(\d+\s*-\s*(?:Productos|Servicios|Productos\s+y\s+Servicios))",
        bloque,
        flags=re.IGNORECASE,
    )
    if m_conc:
        concepto = re.sub(r"\s+", " ", m_conc.group(1)).strip()

    punto_venta = ""
    m_pv = re.search(r"Punto\s+de\s+Venta[:\s]*(\d{1,5})", bloque, flags=re.IGNORECASE)
    if m_pv:
        punto_venta = m_pv.group(1)

    nro_cmp = ""
    for patron_nro in (
        r"Comp(?:\.|\s)*Nro[:\s\.]*(\d+)",
        r"Nro\.?\s*(?:de\s+)?Comprobante[:\s]*(\d+)",
        r"N[uú]mero[:\s]*(\d{6,})",
    ):
        m_nro = re.search(patron_nro, bloque, flags=re.IGNORECASE)
        if m_nro:
            nro_cmp = m_nro.group(1)
            break

    importe_total = 0.0
    m_imp = re.search(
        r"Importe\s+Total[:\s]*\$?\s*([\d][\d.,]*)",
        bloque,
        flags=re.IGNORECASE,
    )
    if m_imp:
        importe_total = _limpiar_monto(m_imp.group(1))
    if importe_total <= 0:
        m_imp2 = re.search(
            r"(?:^|\n)\s*Total[:\s]*\$?\s*([\d][\d.,]*)",
            bloque,
            flags=re.IGNORECASE,
        )
        if m_imp2:
            importe_total = _limpiar_monto(m_imp2.group(1))

    comprobante = ""
    if punto_venta and nro_cmp:
        comprobante = formatear_comprobante_tango(punto_venta, nro_cmp)

    cae = _extraer_cae_afip(bloque)

    if not periodo_desde and fecha_emision:
        periodo_desde = fecha_emision
    if not periodo_hasta and periodo_desde:
        periodo_hasta = periodo_desde

    if not periodo_desde or importe_total <= 0:
        return None

    # NC siempre resta en el consolidado de facturación.
    importe_firmado = round(abs(importe_total) * signo, 2)

    return {
        "Archivo": archivo,
        "Tipo": tipo,
        "Código AFIP": codigo_afip,
        "Fecha Emisión": fecha_emision,
        "Período Desde": periodo_desde,
        "Período Hasta": periodo_hasta,
        "Concepto": concepto,
        "Importe Total": importe_firmado,
        "Comprobante": comprobante,
        "CAE": cae,
    }


def procesar_facturas_monotributo(archivos) -> tuple[pd.DataFrame, list[dict]]:
    """
    Procesa PDFs sueltos y/o ZIP con facturas AFIP.
    NC van en negativo; elimina duplicados (mismo CAE/comprobante).
    Retorna DataFrame ordenado por Período Desde y lista de errores por archivo.
    """
    filas: list[dict] = []
    errores: list[dict] = []

    for nombre, pdf_bytes in iter_pdfs_desde_uploads(archivos):
        try:
            texto = extraer_texto_factura_afip(pdf_bytes)
            parsed = parsear_factura_afip_texto(texto, nombre)
            if parsed:
                filas.append(parsed)
            else:
                errores.append({
                    "archivo": nombre,
                    "motivo": "No se detectaron período devengado e importe total en el PDF.",
                })
        except Exception as exc:
            errores.append({"archivo": nombre, "motivo": str(exc)})

    filas, n_dupes = deduplicar_comprobantes_monotributo(filas)
    if n_dupes:
        errores.append({
            "archivo": "(consolidado)",
            "motivo": f"Se descartaron {n_dupes} comprobante(s) duplicado(s).",
        })

    if not filas:
        return pd.DataFrame(columns=COLUMNAS_MONOTRIBUTO), errores

    df = pd.DataFrame(filas)
    for col in COLUMNAS_MONOTRIBUTO:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNAS_MONOTRIBUTO]
    df["_sort_desde"] = df["Período Desde"].map(
        lambda x: _parsear_fecha(str(x)) or date.min
    )
    df = df.sort_values("_sort_desde", kind="stable").drop(columns=["_sort_desde"])
    return df.reset_index(drop=True), errores


def exportar_monotributo_excel(df: pd.DataFrame) -> bytes:
    """Excel de trabajo monotributo en formato estándar del Estudio."""
    from excel_formato_estudio import exportar_informe_excel

    df_export = df.copy() if df is not None else pd.DataFrame()
    total = float(df_export["Importe Total"].sum()) if not df_export.empty and "Importe Total" in df_export.columns else 0.0
    facturas = (
        float(df_export.loc[df_export["Importe Total"] > 0, "Importe Total"].sum())
        if not df_export.empty and "Importe Total" in df_export.columns
        else 0.0
    )
    notas = (
        float(df_export.loc[df_export["Importe Total"] < 0, "Importe Total"].sum())
        if not df_export.empty and "Importe Total" in df_export.columns
        else 0.0
    )
    resumen = pd.DataFrame(
        [
            {"Concepto": "Facturas / ND (positivos)", "Importe": round(facturas, 2)},
            {"Concepto": "Notas de crédito (negativos)", "Importe": round(notas, 2)},
            {"Concepto": "Neto facturado", "Importe": round(total, 2)},
        ]
    )
    return exportar_informe_excel(
        titulo="Monotributo — Facturas devengadas",
        subtitulo="Análisis de períodos · Estudio Contable",
        kpis=[
            ("Neto facturado", round(total, 2), "money"),
            ("Cantidad de comprobantes", len(df_export), "int"),
        ],
        resumenes=[("Resumen FC / NC", resumen)],
        detalle=df_export,
        hoja_detalle="Facturas Devengadas",
        col_moneda=["Importe", "Importe Total"],
        col_fecha=["Fecha", "Fecha Emisión", "Fecha Contable"],
        total_col="Importe Total" if "Importe Total" in df_export.columns else None,
    )


# ---------------------------------------------------------------------------
# Liquidaciones de tarjeta (Prisma / Mercado Pago / First Data) → Excel
# ---------------------------------------------------------------------------

from procesadores.tarjeta_parser import (  # noqa: E402
    PLANTILLAS_TARJETAS,
    detectar_entidad_por_texto,
    extraer_con_plantilla,
    extraer_texto_liquidacion_pdf,
)

COLUMNAS_LIQUIDACION_TARJETA = [
    "Archivo",
    "Fecha",
    "Entidad",
    "Nro_Liquidacion",
    "Neto_Gravado",
    "IVA_21",
    "Percepcion_IVA",
    "Retencion_IVA",
    "Retencion_IIBB",
    "Percepcion_IIBB",
    "Total_Descontado",
]


def extraer_datos_liquidacion_pdf(file_buffer, nombre_archivo: str = "", entidad: str | None = None) -> dict:
    """
    Lee el PDF de liquidación de tarjeta, detecta (o usa) la entidad y extrae conceptos.
    `entidad` opcional: clave de PLANTILLAS_TARJETAS o 'Otra / No detectada'.
    """
    if not nombre_archivo and hasattr(file_buffer, "name"):
        nombre_archivo = str(getattr(file_buffer, "name", "") or "")
    texto = extraer_texto_liquidacion_pdf(file_buffer)
    entidad_usada = entidad or detectar_entidad_por_texto(texto)
    datos = extraer_con_plantilla(texto, entidad_usada)
    datos["Archivo"] = nombre_archivo or "liquidacion.pdf"
    datos["Entidad"] = (
        entidad_usada
        if entidad_usada != "Otra / No detectada"
        else datos.get("Entidad") or "Desconocida"
    )
    return datos


def procesar_liquidaciones_tarjeta_pdfs(
    archivos,
    entidades_por_archivo: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, list[dict]]:
    """
    Procesa uno o varios PDFs de liquidación de tarjeta → DataFrame unificado.
    `entidades_por_archivo`: mapa opcional {nombre_archivo: entidad confirmada}.
    """
    filas: list[dict] = []
    errores: list[dict] = []
    entidades_por_archivo = entidades_por_archivo or {}
    for uploaded in archivos or []:
        nombre, data = _leer_bytes_upload(uploaded)
        if not nombre.lower().endswith(".pdf"):
            errores.append({"archivo": nombre, "motivo": "Solo se admiten PDF."})
            continue
        try:
            entidad = entidades_por_archivo.get(nombre)
            fila = extraer_datos_liquidacion_pdf(data, nombre_archivo=nombre, entidad=entidad)
            if fila.get("Entidad") in {"Desconocida", "Otra / No detectada"} and float(
                fila.get("Total_Descontado") or 0
            ) <= 0:
                errores.append({
                    "archivo": nombre,
                    "motivo": "No se detectó entidad ni conceptos fiscales en el PDF.",
                    "entidad": fila.get("Entidad"),
                })
                continue
            filas.append(fila)
            if fila.get("Entidad") in {"Desconocida"}:
                errores.append({
                    "archivo": nombre,
                    "motivo": "Entidad no reconocida; se exportaron montos parciales si había.",
                    "entidad": "Desconocida",
                })
        except Exception as exc:
            errores.append({"archivo": nombre, "motivo": str(exc)})

    if not filas:
        return pd.DataFrame(columns=COLUMNAS_LIQUIDACION_TARJETA), errores
    df = pd.DataFrame(filas)
    for col in COLUMNAS_LIQUIDACION_TARJETA:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNAS_LIQUIDACION_TARJETA].reset_index(drop=True), errores


def exportar_liquidaciones_tarjeta_excel(df: pd.DataFrame) -> bytes:
    """
    Excel consolidado de liquidaciones (openpyxl):
    - Hoja Detalle con encabezado, montos y fila de TOTALES con fórmulas SUM
    - Hoja Totales_entidad
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cols_orden = [
        "Fecha", "Entidad", "Nro_Liquidacion", "Archivo",
        "Neto_Gravado", "IVA_21", "Percepcion_IVA", "Retencion_IVA",
        "Retencion_IIBB", "Percepcion_IIBB", "Total_Descontado",
    ]
    cols_monto = [
        "Neto_Gravado", "IVA_21", "Percepcion_IVA", "Retencion_IVA",
        "Retencion_IIBB", "Percepcion_IIBB", "Total_Descontado",
    ]
    work = df.copy() if df is not None else pd.DataFrame(columns=cols_orden)
    for c in cols_orden:
        if c not in work.columns:
            work[c] = None if c not in cols_monto else 0.0
    work = work[cols_orden]

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Liquidaciones"

    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_body = Font(name="Calibri", size=11, color="000000")
    font_total = Font(name="Calibri", bold=True, size=11, color="000000")
    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    fill_alt = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    # Encabezados
    for col_idx, nombre in enumerate(cols_orden, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border

    # Datos
    for row_idx, fila in enumerate(work.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(fila, start=1):
            col_name = cols_orden[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in cols_monto:
                try:
                    cell.value = float(valor or 0)
                except (TypeError, ValueError):
                    cell.value = 0.0
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            else:
                cell.value = "" if valor is None else str(valor)
                cell.alignment = align_left if col_name in {"Archivo", "Entidad"} else align_center
            cell.font = font_body
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = fill_alt

    n = len(work)
    if n > 0:
        total_row = n + 2
        ws.cell(row=total_row, column=1, value="TOTALES").font = font_total
        ws.cell(row=total_row, column=1).fill = fill_total
        ws.cell(row=total_row, column=1).border = border
        for col_idx in range(2, 5):
            c = ws.cell(row=total_row, column=col_idx, value="")
            c.fill = fill_total
            c.border = border
        for col_name in cols_monto:
            col_idx = cols_orden.index(col_name) + 1
            letra = get_column_letter(col_idx)
            cell = ws.cell(row=total_row, column=col_idx)
            cell.value = f"=SUM({letra}2:{letra}{n + 1})"
            cell.number_format = '#,##0.00'
            cell.font = font_total
            cell.fill = fill_total
            cell.border = border
            cell.alignment = align_right

    anchos = {
        "Fecha": 12, "Entidad": 28, "Nro_Liquidacion": 16, "Archivo": 36,
        "Neto_Gravado": 14, "IVA_21": 12, "Percepcion_IVA": 14, "Retencion_IVA": 14,
        "Retencion_IIBB": 14, "Percepcion_IIBB": 14, "Total_Descontado": 16,
    }
    for col_idx, nombre in enumerate(cols_orden, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(nombre, 14)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols_orden))}{max(1, n + 1)}"
    ws.freeze_panes = "A2"

    # Totales por entidad
    ws2 = wb.create_sheet("Totales_entidad")
    if n > 0 and "Entidad" in work.columns:
        tot = (
            work.groupby("Entidad", dropna=False)[cols_monto]
            .sum(numeric_only=True)
            .reset_index()
        )
    else:
        tot = pd.DataFrame(columns=["Entidad", *cols_monto])

    headers2 = ["Entidad", *cols_monto]
    for col_idx, nombre in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=nombre)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border

    for row_idx, fila in enumerate(tot.itertuples(index=False), start=2):
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if col_idx == 1:
                cell.value = str(valor)
                cell.alignment = align_left
            else:
                cell.value = float(valor or 0)
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            cell.font = font_body
            cell.border = border

    n2 = len(tot)
    if n2 > 0:
        total_row2 = n2 + 2
        ws2.cell(row=total_row2, column=1, value="TOTALES").font = font_total
        ws2.cell(row=total_row2, column=1).fill = fill_total
        ws2.cell(row=total_row2, column=1).border = border
        for col_idx in range(2, len(headers2) + 1):
            letra = get_column_letter(col_idx)
            cell = ws2.cell(row=total_row2, column=col_idx)
            cell.value = f"=SUM({letra}2:{letra}{n2 + 1})"
            cell.number_format = '#,##0.00'
            cell.font = font_total
            cell.fill = fill_total
            cell.border = border
            cell.alignment = align_right

    ws2.column_dimensions["A"].width = 28
    for col_idx in range(2, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 14
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Extractos Santander digitales (PDF) → Excel unificado + tabla dinámica
# ---------------------------------------------------------------------------

_RE_FECHA_EXT = re.compile(r"\b(\d{2}/\d{2}/\d{2,4})\b")
# Importes ARS. PRIORIDAD: coma decimal argentina (1.234.567,89) antes que
# el patrón OCR con punto decimal (1.234.56), si no `$ 2.790.747,23` se corta en `2.790.74`.
_RE_PESOS_EXT = re.compile(
    r"(?:(?:pesos|\$)\s*)("
    r"-?\d{1,3}(?:\.\d{3})+,\d{2}"       # 2.790.747,23
    r"|-?\d+,\d{2}"                        # 320,82
    r"|-?\d{1,3}(?:\.\d{3})+\.\d{2}(?!,)" # OCR 1.234.56 (no si sigue coma/dígito de miles AR)
    r"|-?\d{1,3}(?:\.\d{3})+(?!\d)"       # 1.078.978 sin decimales
    r"|-?\d+\.\d{2}(?![\d,])"             # 12.50 US / simple
    r"|-?\d+(?![.,\d])"                   # entero
    r")"
    r"|(?<![/\d.$])(-?\d{1,3}(?:\.\d{3})+,\d{2})(?!\d)"   # AR con $ implícito
    r"|(?<![/\d.$])(-?\d+,\d{2})(?!\d)"
    r"|(?<![/\d.$])(-?\d{1,3}(?:\.\d{3})+\.\d{2})(?![,\d])"  # OCR mangled
    r"|(?<![/\d.$])(-?\d+\.\d{2})(?![\d,])",
    re.IGNORECASE,
)
_RE_PERIODO_EXT = re.compile(
    r"Desde:\s*(\d{2}/\d{2}/\d{2,4}).*?Hasta:\s*(\d{2}/\d{2}/\d{2,4})",
    re.IGNORECASE | re.DOTALL,
)
_RE_CUIT_EXT = re.compile(r"CUIT:\s*([\d\-]+)", re.IGNORECASE)
_RE_CUENTA_EXT = re.compile(r"Cuenta Corriente\s*N[º°o.\s]*([\d\-/]+)", re.IGNORECASE)
_RE_CBU_EXT = re.compile(r"CBU:\s*(\d{22})", re.IGNORECASE)
_PALABRAS_DEBITO_EXT = (
    "debito", "débito", "compra", "transferencia realizada", "transferencia enviada",
    "transferencia a terceros", "transf a terceros", "transf. a terceros", "trf a terceros",
    "trf. a terceros", "transferencia inmediata a", "comision", "comisión",
    "impuesto", "iva ", "pago de ", "retencion", "retención",
    "cargo", "mantenimiento", "interes", "interés", "extraccion", "extracción",
    "retiro", "e-pago", "epago", "deb. automatico", "deb automatico", "debito automatico",
    "débito automatico", "pago haberes", "pago sueldos", "qr pct", "qrpct",
    "debin enviado", "echeq", "e-cheq", "banelco", "link ", "pago mis cuentas",
)
_PALABRAS_CREDITO_EXT = (
    "credito", "crédito", "recibid", "deposito", "depósito", "transferencia recibida",
    "transferencia de terceros", "transf de terceros", "transf. de terceros",
    "trf de terceros", "trf. de terceros", "transferencia desde terceros",
    "pago comercios", "pago a proveedores recibido", "acredita", "acreditacion",
    "acreditación", "cobro", "ingreso", "devolucion", "devolución", "reintegro",
    "interes ganado", "interés ganado", "pago con transferencia",
    "debin recibido", "debin aceptado", "venta", "cobranza",
    "rescate",  # rescate FCI / superfondo = ingreso a la cuenta
)


def _parse_fecha_extracto(txt: str) -> date | None:
    txt = str(txt or "").strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def _parse_monto_pesos_ar(txt: str) -> float:
    """Alias seguro: siempre pasa por ``_limpiar_monto`` (no destruir decimales)."""
    return float(_limpiar_monto(txt))


def _es_ruido_extracto_santander(linea: str) -> bool:
    n = linea.lower().strip()
    if not n:
        return True
    if n in {
        "fecha", "comprobante", "movimiento", "débito", "debito", "crédito", "credito",
        "saldo en cuenta", "movimientos en pesos", "resumen de cuenta",
        "* salvo error u omisión", "salvo error u omisión", "santander",
        "descripcion", "descripción", "origen", "saldo", "página", "pagina",
    }:
        return True
    ruido_prefijo = (
        "banco santander argentina", "movimientos en pesos", "saldo en cuenta",
        "cuenta corriente n", "período", "periodo", "emisión mensual", "emision mensual",
        "total en ", "resumen de cuenta", "responsable inscri", "unidad de información",
        "fondos comunes", "ningún accionista", "correlativo", "saldo total detalle impositivo",
        "detalle impositivo", "tipo de impuesto", "totales de retencion", "totales de retención",
        "total depositos del dia", "total depósitos del dia", "total depositos del día",
        "salvo error u omisi", "fecha comprobante",
        "resumen de caja de ahorro", "dispones de 30 dias", "el monto de iva discriminado",
    )
    if any(n.startswith(r) for r in ruido_prefijo):
        return True
    return bool(re.fullmatch(r"\d+\s*-\s*\d+", n))


def _montos_en_linea_extracto(ln: str) -> list[float]:
    """
    Extrae importes ARS de una línea **conservando el signo**.
    Regla Galicia / usuario: número negativo (resta) = débito; positivo = crédito.
    También: «pesos menos 26.377.263,87» (Santander, saldo en descubierto).
    """
    out: list[float] = []
    texto = ln or ""
    for m in _RE_PESOS_EXT.finditer(texto):
        raw = next((g for g in m.groups() if g), None)
        if not raw:
            continue
        start, end = m.start(), m.end()
        # Incluir guión inmediatamente anterior si el grupo no lo trajo
        if not raw.startswith("-") and start > 0 and texto[start - 1] == "-":
            raw = "-" + raw
            start -= 1
        # «pesos menos 1.234,56» / «$ menos 1.234,56» (saldo negativo en extracto Santander)
        if not raw.startswith("-"):
            ctx_izq = texto[max(0, start - 18):start].lower()
            if re.search(r"(?:pesos|\$)\s*menos\s*$", ctx_izq) or re.search(
                r"\bmenos\s*$", ctx_izq
            ):
                raw = "-" + raw
        ctx_izq = texto[max(0, start - 12):start].lower()
        ctx_der = texto[end:min(len(texto), end + 4)]
        if re.search(r"\bley\s*$", ctx_izq) or re.search(r"\bart\.?\s*$", ctx_izq):
            continue
        if "%" in ctx_der[:2]:
            continue
        val = _limpiar_monto(raw)  # conserva signo
        if val == 0:
            continue
        abs_val = abs(val)
        raw_digits = raw.lstrip("-")
        if "," not in raw_digits and not re.search(r"\.\d{2}$", raw_digits) and not re.fullmatch(
            r"\d{1,3}(?:\.\d{3})+", raw_digits
        ):
            if 1900 <= abs_val <= 2100:
                continue
            if re.fullmatch(r"\d{5,}", raw_digits) and "$" not in m.group(0).lower() and "pesos" not in m.group(0).lower():
                continue
        if abs_val < 1.0 and ("%" in texto[max(0, start - 2):end + 2] or "iva" in ctx_izq):
            continue
        out.append(round(val, 2))
    return out


def _elegir_monto_y_saldo_extracto(
    montos: list[float],
    saldo_prev: float | None,
) -> tuple[float, float | None]:
    """
    Elige (monto_movimiento_con_signo, saldo).
    El saldo es el último importe (conserva signo: la CC puede quedar en descubierto).
    El movimiento puede ser negativo (=débito).
    """
    if not montos:
        return 0.0, None
    if len(montos) == 1:
        return montos[0], None

    # No usar abs(): con saldo negativo (ej. -$ 26.377.263,87) abs() invertía
    # el delta y clasificaba mal retiro/rescate.
    saldo = float(montos[-1])
    candidatos = montos[:-1]

    if saldo_prev is not None:
        delta = round(saldo - float(saldo_prev), 2)
        delta_abs = abs(delta)
        # Preferir candidato cuyo valor absoluto cierra el delta; conservar signo del candidato
        for tol in (0.05, 0.51, 1.01):
            matches = [c for c in candidatos if abs(abs(c) - delta_abs) <= tol]
            if matches:
                matches.sort(key=lambda c: abs(abs(c) - delta_abs))
                elegido = matches[0]
                # Regla de oro: si el saldo BAJA → el movimiento es débito (negativo)
                if delta < 0:
                    return -abs(elegido), saldo
                if delta > 0:
                    return abs(elegido), saldo
                return elegido, saldo
        # Ningún candidato cierra: usar el delta con signo contable
        if delta_abs >= 0.01:
            return (-delta_abs if delta < 0 else delta_abs), saldo

    # Sin saldo previo: si hay signo en el penúltimo, respetarlo
    mov = candidatos[-1]
    if len(montos) >= 3:
        a, b = montos[-3], montos[-2]
        if a != 0 and b == 0:
            mov = a
        elif b != 0 and a == 0:
            mov = b
    return mov, saldo


def _dc_desde_monto_con_signo(monto_signed: float) -> tuple[float, float]:
    """Conveniencia: Importe− → Débito; Importe+ → Crédito. El signo vive en Importe."""
    m = round(float(monto_signed or 0), 2)
    if m < 0:
        return abs(m), 0.0
    if m > 0:
        return 0.0, abs(m)
    return 0.0, 0.0


def _importe_firmado_desde_fila(r: dict) -> float | None:
    """Obtiene el importe con signo original (− resta / + suma)."""
    if r.get("Importe") is not None and str(r.get("Importe")).strip() != "":
        try:
            return round(float(r["Importe"]), 2)
        except (TypeError, ValueError):
            pass
    deb = float(r.get("Debito") or 0)
    cred = float(r.get("Credito") or 0)
    if deb > 0 and cred <= 0:
        return -round(deb, 2)
    if cred > 0 and deb <= 0:
        return round(cred, 2)
    return None


def _aplicar_importe_y_dc(r: dict, importe_signed: float | None) -> dict:
    """Deja Importe con signo; Débito/Crédito solo como espejo opcional."""
    if importe_signed is None or abs(importe_signed) < 0.005:
        r["Importe"] = None
        r["Debito"] = None
        r["Credito"] = None
        return r
    imp = round(float(importe_signed), 2)
    r["Importe"] = imp
    deb, cred = _dc_desde_monto_con_signo(imp)
    r["Debito"] = round(deb, 2) if deb else None
    r["Credito"] = round(cred, 2) if cred else None
    return r


def _corregir_filas_extracto_por_saldos(movs: list[dict]) -> list[dict]:
    """
    Conserva el signo matemático del movimiento (− resta, + suma).
    No inventa débito/crédito en contra del signo: Importe es la fuente de verdad.
    Si falta el signo, usa el delta de saldo (negativo si el saldo bajó).
    """
    if not movs:
        return movs
    saldo_prev: float | None = None
    out: list[dict] = []
    for row in movs:
        r = dict(row)
        tipo = str(r.get("Tipo fila") or "Movimiento")
        saldo = r.get("Saldo")
        try:
            saldo_f = float(saldo) if saldo is not None and str(saldo).strip() != "" else None
        except (TypeError, ValueError):
            saldo_f = None

        if tipo == "Saldo inicial":
            r["Valido"] = True
            r["Importe"] = None
            r["Debito"] = None
            r["Credito"] = None
            if saldo_f is not None:
                saldo_prev = saldo_f
            out.append(r)
            continue

        importe = _importe_firmado_desde_fila(r)
        valido = True

        if saldo_prev is not None and saldo_f is not None:
            delta = round(saldo_f - saldo_prev, 2)
            if abs(delta) < 0.01:
                if importe is not None and abs(importe) > 0.01:
                    valido = False
            else:
                # Preferir signo original si cierra; si no hay importe, usar delta (con signo)
                if importe is None:
                    importe = delta
                elif abs(abs(importe) - abs(delta)) > 0.05:
                    # Magnitud no cierra: el delta dice qué restó/sumó (signo contable real)
                    importe = delta
                    valido = False
                else:
                    # Misma magnitud: forzar el signo del movimiento del saldo
                    # (resta → negativo) sin perder el valor
                    importe = -abs(importe) if delta < 0 else abs(importe)
                esperado = round(saldo_prev + float(importe), 2)
                valido = abs(esperado - saldo_f) <= 0.05

        r = _aplicar_importe_y_dc(r, importe)
        r["Valido"] = valido
        if saldo_f is not None:
            saldo_prev = saldo_f
        out.append(r)
    return out


_RE_CONCEPTO_MOV_EXT = re.compile(
    r"^(?:"
    r"impuesto|imp\b|retencion|retención|comision|comisión|iva\b|"
    r"pago\b|transferencia|transf\.?|trf\.?|cheque|compra\b|retiro\b|"
    r"debito\b|débito\b|credito\b|crédito\b|saldo inicial|mantenimiento|"
    r"interes|interés|imp al\b|terceros|debin|echeq|e-cheq|acredit|"
    r"deposito|depósito|extracci|cbu\b|cvu\b|banelco|link\b|qr\b|"
    r"deb\.?\s*aut|varios|cobro|ingreso|reintegro|devoluc"
    r")",
    re.IGNORECASE,
)

# Variantes frecuentes de transferencias a/de terceros (Galicia, BBVA, Macro, etc.)
_RE_TRANSF_TERCEROS = re.compile(
    r"(?i)\b(?:transf(?:erencia|\.)?|trf\.?)\s*"
    r"(?:inmediata\s+|electronica\s+|electrónica\s+|online\s+|cbu\s+)?"
    r"(?:a|de|desde|hacia)?\s*terceros?\b"
    r"|\btransferencia\s+(?:a|de)\s+terceros?\b"
    r"|\b(?:debito|débito|credito|crédito)\s+transf"
)


def _es_concepto_transferencia_terceros(texto: str) -> bool:
    t = str(texto or "")
    if not t.strip():
        return False
    if _RE_TRANSF_TERCEROS.search(t):
        return True
    low = _normalizar_texto(t)
    return ("terceros" in low and ("transf" in low or "trf" in low or "transfer" in low))


def _linea_solo_importes_extracto(ln: str) -> bool:
    """True si la línea es casi solo montos (pesos/$ + números), sin concepto."""
    resto = _RE_PESOS_EXT.sub("", ln or "")
    resto = re.sub(r"(?i)\bpesos\b|\$", "", resto)
    resto = re.sub(r"\s+", " ", resto).strip(" -|")
    return len(resto) < 3


def _parece_inicio_movimiento_extracto(ln: str, *, con_fecha_previa: bool) -> bool:
    """
    Detecta inicio de fila de movimiento (top-level).
    Debe ser INCLUSIVO: si solo aceptamos 'Impuesto…' se pierden ingresos/egresos.
    """
    if _es_ruido_extracto_santander(ln):
        return False
    low = ln.lower().strip()
    if low.startswith("resp:") or low.startswith("resp "):
        return False
    if "total depositos" in low or "total depósitos" in low:
        return False
    if low.startswith("desde:") or low.startswith("hasta:"):
        return False
    montos = _montos_en_linea_extracto(ln)
    m_fecha = _RE_FECHA_EXT.match(ln)
    if m_fecha:
        resto = ln[m_fecha.end():].strip()
        if not resto and not montos:
            return False
        return True
    # Fecha embebida a mitad de línea (OCR raro)
    if _RE_FECHA_EXT.search(ln) and (montos or re.search(r"[A-Za-zÁÉÍÓÚáéíóúñÑ]{3,}", ln)):
        return True
    if not con_fecha_previa:
        return False
    if re.fullmatch(r"\d{4,12}", ln.strip()):
        return True
    if re.match(r"^\d{4,12}\b", ln):
        return True
    # Concepto conocido (impuestos, transf, debin, etc.)
    if _RE_CONCEPTO_MOV_EXT.match(ln) or _es_concepto_transferencia_terceros(ln):
        return True
    # Cualquier texto con al menos un importe (ingresos/egresos genéricos)
    if montos and re.search(r"[A-Za-zÁÉÍÓÚáéíóúñÑ]{2,}", ln) and not _linea_solo_importes_extracto(ln):
        return True
    # Dos importes aunque el concepto sea un nombre de tercero / razón social
    if len(montos) >= 2 and not _linea_solo_importes_extracto(ln):
        return True
    # Razón social / descripción sin montos aún (montos en líneas siguientes)
    if (
        not _linea_solo_importes_extracto(ln)
        and re.search(r"[A-Za-zÁÉÍÓÚáéíóúñÑ]{4,}", re.sub(r"(?i)\bpesos\b|\$", "", ln))
        and not re.fullmatch(r"[\d\s./\-]+", ln)
        and "saldo" not in low
    ):
        return True
    return False


def _es_corte_nuevo_movimiento_extracto(l2: str, *, montos_actuales: list[float]) -> bool:
    """
    Criterio ESTRICTO para dejar de absorber líneas del movimiento actual.
    No cortar por nombres de beneficiario ni por líneas solo de importes.
    """
    if _es_ruido_extracto_santander(l2):
        return False
    if _RE_FECHA_EXT.match(l2):
        return True
    montos_l2 = _montos_en_linea_extracto(l2)
    if _linea_solo_importes_extracto(l2) and montos_l2:
        return False
    # Nuevo comprobante + importes ⇒ otro movimiento
    if re.match(r"^\d{4,12}\b", l2) and montos_l2 and montos_actuales:
        return True
    # Concepto fuerte + importes y el actual ya tiene importes
    if montos_actuales and montos_l2 and (
        _RE_CONCEPTO_MOV_EXT.match(l2) or _es_concepto_transferencia_terceros(l2)
    ):
        return True
    return False


def _clasificar_debito_credito_extracto(
    desc: str, monto: float, saldo_prev: float | None, saldo: float | None
) -> tuple[float, float]:
    """
    Asigna (débito, crédito). Prioridad inflexible:
    1) Si el saldo BAJA → SIEMPRE débito; si SUBE → crédito
    2) Si el monto viene con signo (negativo) → débito
    3) Palabras clave
    """
    monto_signed = float(monto or 0)
    monto_abs = abs(monto_signed)
    if monto_abs <= 0:
        return 0.0, 0.0
    if saldo_prev is not None and saldo is not None:
        delta = round(float(saldo) - float(saldo_prev), 2)
        if abs(abs(delta) - monto_abs) <= 0.05 or abs(delta) >= 0.01:
            if abs(abs(delta) - monto_abs) <= 1.01:
                if delta < 0:
                    return (abs(delta) if abs(abs(delta) - monto_abs) > 0.05 else monto_abs), 0.0
                if delta > 0:
                    return 0.0, (abs(delta) if abs(abs(delta) - monto_abs) > 0.05 else monto_abs)
    # Signo del importe (Galicia: -38.225,00 = débito)
    if monto_signed < 0:
        return monto_abs, 0.0
    dlow = (desc or "").lower()
    if _es_concepto_transferencia_terceros(desc):
        if re.search(r"(?i)\b(?:de|desde)\s+terceros\b", desc) and not re.search(
            r"(?i)\ba\s+terceros\b", desc
        ):
            return 0.0, monto_abs
        return monto_abs, 0.0
    if any(p in dlow for p in _PALABRAS_CREDITO_EXT):
        if "transferencia realizada" in dlow or "transferencia enviada" in dlow:
            return monto_abs, 0.0
        if re.search(r"(?i)\ba\s+terceros\b", dlow):
            return monto_abs, 0.0
        # "IMP. CRE." = impuesto al crédito → es DÉBITO en cuenta, no un crédito
        if re.search(r"(?i)\bimp\.?\s*cre\.?\b|\bimpuesto\s+.*credito", dlow):
            return monto_abs, 0.0
        return 0.0, monto_abs
    if any(p in dlow for p in _PALABRAS_DEBITO_EXT):
        if "pago comercios" in dlow or "pago a proveedores recibido" in dlow:
            return 0.0, monto_abs
        if "pago con transferencia" in dlow and "realizada" not in dlow:
            return 0.0, monto_abs
        return monto_abs, 0.0
    # Sin pista: si vino positivo asumir crédito solo si no parece egreso
    if monto_signed > 0 and any(
        k in dlow for k in ("acredit", "deposito", "depósito", "transferencia de", "recibid")
    ):
        return 0.0, monto_abs
    return monto_abs, 0.0


def unificar_concepto_extracto(descripcion: str) -> str:
    """Agrupa descripciones equivalentes (sin CUITs/refs variables) para tabla dinámica."""
    d = re.sub(r"\s+", " ", str(descripcion or "")).strip()
    if not d:
        return "Sin concepto"
    low = d.lower()
    if "saldo inicial" in low:
        return "Saldo Inicial"
    if low.startswith("transferencia realizada"):
        return "Transferencia realizada"
    if _es_concepto_transferencia_terceros(d):
        if re.search(r"(?i)\b(?:de|desde)\s+terceros\b", d) and not re.search(
            r"(?i)\ba\s+terceros\b", d
        ):
            return "Transferencia de terceros"
        return "Transferencia a terceros"
    if "pago con transferencia" in low:
        return "Pago con transferencia QRPCT"
    if "pago comercios prisma" in low:
        return "Pago comercios Prisma Visa"
    if "impuesto ley 25.413 debito" in low or "impuesto ley 25413 debito" in low:
        return "Impuesto ley 25.413 débito 0,6%"
    if "impuesto ley 25.413 credito" in low or "impuesto ley 25413 credito" in low:
        return "Impuesto ley 25.413 crédito 0,6%"
    if "com transf" in low or "comision transferencias" in low or "comisión transferencias" in low:
        return "Comisión transferencias"
    if "iva 21%" in low and ("transf" in low or "ley27743" in low or "ley 27743" in low):
        return "IVA 21% comisión transferencias"
    if "compra con tarjeta" in low:
        return "Compra con tarjeta de débito"
    if "pago a proveedores recibido" in low:
        return "Pago a proveedores recibido"
    if "pago haberes" in low:
        return "Pago haberes"
    # Limpieza genérica: sacar CUITs y referencias numéricas largas
    limpio = re.sub(r"\b\d{11}\b", "", d)
    limpio = re.sub(r"(?i)tarj\.?\s*nro\.?\s*\w+", "", limpio)
    limpio = re.sub(r"(?i)nro\.?\s*liq\.?\s*[\d/]+", "", limpio)
    limpio = re.sub(r"(?i)/\s*-\s*var\s*/?", "", limpio)
    limpio = re.sub(r"\b\d{5,}\b", "", limpio)
    limpio = re.sub(r"\s+", " ", limpio).strip(" -/|")
    return (limpio[:90] if limpio else d[:90])


def _mes_clave_desde_fecha(fecha_txt: str) -> str:
    f = _parse_fecha_extracto(fecha_txt)
    if not f:
        return "Sin mes"
    return f"{f.year:04d}-{f.month:02d}"


def _mes_etiqueta(mes_clave: str) -> str:
    meses = {
        "01": "Ene", "02": "Feb", "03": "Mar", "04": "Abr", "05": "May", "06": "Jun",
        "07": "Jul", "08": "Ago", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dic",
    }
    if not re.fullmatch(r"\d{4}-\d{2}", mes_clave or ""):
        return mes_clave or "Sin mes"
    anio, mes = mes_clave.split("-")
    return f"{meses.get(mes, mes)}-{anio[2:]}"


def _parsear_movimientos_santander_paginas(
    paginas: list[tuple[int, str]],
    archivo_origen: str = "",
) -> tuple[list[dict], dict]:
    meta = {"cliente": "", "cuit": "", "cuenta": "", "cbu": "", "periodos": [], "archivo": archivo_origen}
    movs: list[dict] = []
    periodo_actual = ""
    saldo_prev: float | None = None
    lineas: list[tuple[int, str, str]] = []

    for pag, texto in paginas:
        if not meta["cliente"]:
            for ln in texto.splitlines():
                if re.search(r"[A-ZÁÉÍÓÚÑ].*(S\.?R\.?L\.?|S\.?A\.?|S\.A\.S)", ln, re.I):
                    if "santander" not in ln.lower():
                        meta["cliente"] = ln.strip()
                        break
        m_cuit = _RE_CUIT_EXT.search(texto)
        if m_cuit and not meta["cuit"]:
            meta["cuit"] = m_cuit.group(1)
        m_cta = _RE_CUENTA_EXT.search(texto)
        if m_cta and not meta["cuenta"]:
            meta["cuenta"] = m_cta.group(1).strip()
        m_cbu = _RE_CBU_EXT.search(texto)
        if m_cbu and not meta["cbu"]:
            meta["cbu"] = m_cbu.group(1)
        m_per = _RE_PERIODO_EXT.search(texto)
        if m_per:
            periodo_actual = f"{m_per.group(1)} a {m_per.group(2)}"
            if periodo_actual not in meta["periodos"]:
                meta["periodos"].append(periodo_actual)
        for raw in texto.splitlines():
            ln = re.sub(r"[ \t]+", " ", raw).strip()
            if not ln:
                continue
            low = ln.lower()
            # Fin del detalle de movimientos (tarifario / impositivo / totales)
            if any(
                m in low
                for m in (
                    "detalle impositivo",
                    "cambio de comisiones",
                    "totales de retencion",
                    "totales de retención",
                    "tasas de acuerdos",
                    "concepto valor tipo",
                    "los precios no incluyen iva",
                )
            ):
                break
            if low.startswith("saldo total $") or low.startswith("saldo total$"):
                break
            lineas.append((pag, ln, periodo_actual))

    i = 0
    periodo_saldo: str | None = None
    fecha_anterior: date | None = None
    while i < len(lineas):
        pag, ln, periodo = lineas[i]
        if periodo != periodo_saldo:
            saldo_prev = None
            periodo_saldo = periodo
        # Fecha sola (sin concepto/montos): ancla el día y sigue — evita omitir
        # "TRANSFERENCIA A/DE TERCEROS" en la línea siguiente.
        m_fecha_sola = _RE_FECHA_EXT.match(ln)
        if m_fecha_sola:
            resto_fecha = ln[m_fecha_sola.end():].strip()
            if not resto_fecha and not _montos_en_linea_extracto(ln):
                f_ancla = _parse_fecha_extracto(m_fecha_sola.group(1))
                if f_ancla:
                    fecha_anterior = f_ancla
                i += 1
                continue
        if not _parece_inicio_movimiento_extracto(ln, con_fecha_previa=fecha_anterior is not None):
            i += 1
            continue

        m_fecha = _RE_FECHA_EXT.match(ln)
        fecha = _parse_fecha_extracto(m_fecha.group(1)) if m_fecha else fecha_anterior
        if not fecha:
            i += 1
            continue

        trabajo = ln
        if m_fecha:
            trabajo = ln[m_fecha.end():].strip()

        comprobante = ""
        m_comp = re.match(r"^(\d{4,12})\b", trabajo)
        if m_comp:
            comprobante = m_comp.group(1)
            trabajo = trabajo[m_comp.end():].strip()
        elif re.fullmatch(r"\d{4,12}", ln.strip()):
            comprobante = ln.strip()
            trabajo = ""

        montos = _montos_en_linea_extracto(ln)
        bloque: list[str] = []
        if trabajo:
            desc0 = _RE_PESOS_EXT.sub("", trabajo)
            desc0 = re.sub(r"\$\s*", "", desc0).strip(" -")
            if desc0:
                bloque.append(desc0)

        j = i + 1
        linea_completa = len(montos) >= 2
        es_saldo_ini = "saldo inicial" in (trabajo or ln).lower()
        es_terceros_mov = _es_concepto_transferencia_terceros(trabajo or ln)
        # Galicia CA: muchas líneas de CBU/CUIT/refs antes del importe → ventana amplia
        if es_saldo_ini and montos:
            max_extra = 0
        elif es_terceros_mov or not montos or re.fullmatch(r"\d{4,12}", ln.strip()):
            max_extra = 14
        else:
            max_extra = 4 if linea_completa else 12
        extras = 0
        while j < len(lineas) and extras < max_extra:
            _, l2, _ = lineas[j]
            # Nunca absorber una nueva fecha: es otro movimiento
            if _RE_FECHA_EXT.match(l2):
                break
            # Continuación de detalle (Resp:, CUIT, tarj.) sin montos de movimiento+saldo
            low2 = l2.lower().strip()
            es_detalle = (
                low2.startswith("resp:")
                or low2.startswith("resp ")
                or low2.startswith("tarj")
                or low2.startswith("cuit")
                or low2.startswith("aca ")
                or low2.startswith("deb. automatico")
                or low2.startswith("deb automatico")
                or "sobre $" in low2
                or (low2.startswith("pago de ") and len(_montos_en_linea_extracto(l2)) == 0)
                or (re.match(r"^\d{6,}/\d+", l2) is not None)
            )
            montos_l2 = _montos_en_linea_extracto(l2)
            # Corte estricto: no usar el detector inclusivo (si no, solo quedan impuestos)
            if (not es_detalle) and _es_corte_nuevo_movimiento_extracto(l2, montos_actuales=montos):
                # Excepción: aún sin montos, absorber concepto/beneficiario
                if not montos and (
                    _es_concepto_transferencia_terceros(l2)
                    or _RE_CONCEPTO_MOV_EXT.match(l2)
                    or re.search(r"[A-Za-zÁÉÍÓÚáéíóúñÑ]{3,}", l2)
                ):
                    pass
                else:
                    break
            if _es_ruido_extracto_santander(l2) and not montos_l2:
                j += 1
                extras += 1
                continue
            if linea_completa and montos_l2 and not es_detalle and not _linea_solo_importes_extracto(l2):
                # Otra fila con concepto+montos
                if _RE_CONCEPTO_MOV_EXT.match(l2) or re.match(r"^\d{4,12}\b", l2):
                    break
            if montos and montos_l2 and not es_detalle and re.match(r"^\d{4,12}\b", l2):
                break
            # Detalle (Resp:, % sobre $X, etc.): solo texto; no mezclar bases imponibles
            if not es_detalle:
                montos.extend(montos_l2)
            if re.fullmatch(r"\d{4,12}", l2.strip()) and not comprobante:
                comprobante = l2.strip()
            else:
                desc_part = _RE_PESOS_EXT.sub("", l2).strip(" -")
                desc_part = re.sub(r"\$\s*", "", desc_part).strip(" -")
                # No meter fechas sueltas ni importes sueltos en la descripción
                if (
                    desc_part
                    and not re.fullmatch(r"\d{4,12}", desc_part)
                    and not _RE_FECHA_EXT.match(desc_part)
                    and not _linea_solo_importes_extracto(l2)
                ):
                    bloque.append(desc_part)
            j += 1
            extras += 1
            if len(montos) >= 2 and not linea_completa:
                linea_completa = True
                max_extra = extras + 2

        # Peek: movimiento con 1 solo importe y la línea siguiente es el saldo
        if len(montos) == 1 and j < len(lineas):
            _, l_saldo, _ = lineas[j]
            m_saldo = _montos_en_linea_extracto(l_saldo)
            if (
                len(m_saldo) == 1
                and _linea_solo_importes_extracto(l_saldo)
                and not _RE_FECHA_EXT.match(l_saldo)
                and abs(m_saldo[0]) > abs(montos[0])
            ):
                montos.append(m_saldo[0])
                j += 1

        bloque_limpio = [re.sub(r"\s+", " ", b).strip(" -|") for b in bloque if b and str(b).strip()]
        if bloque_limpio:
            desc_raw, det_raw = _partir_concepto_y_detalle(
                bloque_limpio[0],
                "\n".join(bloque_limpio[1:]) if len(bloque_limpio) > 1 else None,
            )
        else:
            desc_raw, det_raw = "", ""
        descripcion = desc_raw
        detalle = det_raw
        descripcion_full = re.sub(
            r"\s+", " ", " ".join(x for x in (descripcion, detalle.replace("\n", " ")) if x)
        ).strip()
        if not descripcion_full and not montos:
            i = j if j > i else i + 1
            continue
        dlow = descripcion_full.lower()
        if "detalle impositivo" in dlow or "saldo total detalle" in dlow:
            i = j if j > i else i + 1
            continue

        fecha_txt = fecha.strftime("%d/%m/%Y")
        mes_clave = f"{fecha.year:04d}-{fecha.month:02d}"
        base = {
            "Fecha": fecha_txt,
            "Mes": mes_clave,
            "Mes etiqueta": _mes_etiqueta(mes_clave),
            "Periodo extracto": periodo,
            "Comprobante": comprobante,
            "Descripcion": descripcion or "Sin descripción",
            "Detalle": detalle or None,
            "Concepto unificado": "",
            "Debito": None,
            "Credito": None,
            "Saldo": None,
            "Pagina PDF": pag,
            "Archivo origen": archivo_origen,
            "Tipo fila": "Movimiento",
        }

        if "saldo inicial" in dlow or (not descripcion and montos and "saldo inicial" in ln.lower()):
            if "saldo inicial" in ln.lower() or "saldo inicial" in dlow:
                base["Saldo"] = montos[0] if montos else None
                base["Descripcion"] = "Saldo Inicial"
                base["Concepto unificado"] = "Saldo Inicial"
                base["Tipo fila"] = "Saldo inicial"
                movs.append(base)
                saldo_prev = base["Saldo"]
                fecha_anterior = fecha
                i = j if j > i else i + 1
                continue
        # Caso OCR: "Saldo Inicial" en línea previa y montos en la de fecha
        if not descripcion and len(montos) == 1 and i > 0:
            prev_ln = lineas[i - 1][1].lower()
            if "saldo inicial" in prev_ln:
                base["Saldo"] = montos[0]
                base["Descripcion"] = "Saldo Inicial"
                base["Concepto unificado"] = "Saldo Inicial"
                base["Tipo fila"] = "Saldo inicial"
                movs.append(base)
                saldo_prev = base["Saldo"]
                fecha_anterior = fecha
                i = j if j > i else i + 1
                continue

        if not montos:
            i = j if j > i else i + 1
            continue
        monto_mov, saldo = _elegir_monto_y_saldo_extracto(montos, saldo_prev)
        if abs(monto_mov) <= 0:
            i = j if j > i else i + 1
            continue
        # Importe con signo original (− resta / + suma). Sin reclasificar por texto.
        if saldo_prev is not None and saldo is not None:
            delta = round(float(saldo) - float(saldo_prev), 2)
            if abs(delta) >= 0.01 and abs(abs(delta) - abs(monto_mov)) <= 0.05:
                # Misma magnitud: el signo del movimiento del saldo es el contable
                monto_mov = -abs(monto_mov) if delta < 0 else abs(monto_mov)
            elif abs(delta) >= 0.01 and abs(monto_mov) > 0 and abs(abs(delta) - abs(monto_mov)) > 0.05:
                # Si el PDF traía signo (−), respetarlo; si no, usar delta
                if monto_mov >= 0 and delta < 0:
                    monto_mov = delta
                elif monto_mov <= 0 and delta > 0 and abs(monto_mov) != abs(delta):
                    # Conflicto raro: conservar signo del PDF
                    pass
        base = _aplicar_importe_y_dc(base, monto_mov)
        base["Saldo"] = round(saldo, 2) if saldo is not None else None
        base["Clasificacion"] = clasificar_movimiento_extracto(
            descripcion, detalle or "", monto_mov
        )
        base["Concepto unificado"] = base["Clasificacion"]
        base["Tipo Movimiento"] = (
            "Debito" if monto_mov < 0 else ("Credito" if monto_mov > 0 else "")
        )
        movs.append(base)
        fecha_anterior = fecha
        if saldo is not None:
            saldo_prev = saldo
        i = j if j > i else i + 1
    return _corregir_filas_extracto_por_saldos(movs), meta


# ---------------------------------------------------------------------------
# Extracto Banco Provincia (BIP) — tablas + bloques multi-línea sin "$"
# ---------------------------------------------------------------------------

_RE_MONTO_AR_LIBRE = re.compile(
    r"(?<![/\d.])("
    r"\d{1,3}(?:\.\d{3})+,\d{2}"          # 2.790.747,23
    r"|\d+,\d{2}"                           # 320,82
    r"|\d{1,3}(?:\.\d{3})+\.\d{2}(?![,\d])" # OCR 1.234.56
    r")(?!\d)"
)


def _montos_ar_libres(ln: str) -> list[float]:
    """Montos argentinos con o sin símbolo $ / 'pesos' (típicos del Provincia)."""
    out: list[float] = []
    for m in _RE_MONTO_AR_LIBRE.finditer(ln or ""):
        try:
            out.append(_parse_monto_pesos_ar(m.group(1)))
        except Exception:
            continue
    return out


def _es_ruido_extracto_provincia(linea: str) -> bool:
    n = _normalizar_texto(linea)
    if not n:
        return True
    if n in {
        "fecha", "descripcion", "descripcion del movimiento", "comprobante",
        "debito", "debitos", "credito", "creditos", "saldo", "debe", "haber",
        "movimientos", "detalle de movimientos",
    }:
        return True
    # Solo encabezados / pie cortos (no filas de movimiento largas)
    if len(n) <= 80 and any(
        r in n
        for r in (
            "banco de la provincia",
            "banca internet",
            "extracto de cuenta",
            "extracto electronico",
            "salvo error",
            "continua en",
            "tipo de cuenta",
        )
    ):
        return True
    if any(
        r in n
        for r in (
            "total debitos", "total creditos", "total debitos del periodo",
            "total creditos del periodo", "resumen del periodo",
        )
    ):
        return True
    if re.match(r"^(pagina|hoja)\s*\d+", n):
        return True
    return False


def _mapear_indices_tabla_provincia(header: list) -> dict[str, int]:
    idxs: dict[str, int] = {}
    for i, raw in enumerate(header or []):
        h = _normalizar_texto(str(raw or ""))
        if not h:
            continue
        if "fecha" in h and "fecha" not in idxs:
            idxs["fecha"] = i
        elif any(k in h for k in ("descrip", "concepto", "detalle", "movimiento", "leyenda")) and "desc" not in idxs:
            idxs["desc"] = i
        elif any(k in h for k in ("debito", "debe", "cargo")) and "debito" not in idxs:
            idxs["debito"] = i
        elif any(k in h for k in ("credito", "haber", "acredita")) and "credito" not in idxs:
            idxs["credito"] = i
        elif "saldo" in h and "saldo" not in idxs:
            idxs["saldo"] = i
        elif any(k in h for k in ("comp", "nro", "numero", "referencia")) and "comp" not in idxs:
            idxs["comp"] = i
    return idxs


def extraer_movimientos_provincia_tabla(
    data: bytes,
    archivo: str = "",
) -> list[MovimientoBanco]:
    """Tablas nativas del PDF digital BIP / Provincia."""
    movimientos: list[MovimientoBanco] = []
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                tablas = page.extract_tables() or []
                for tabla in tablas:
                    if not tabla or len(tabla) < 2:
                        continue
                    header = tabla[0]
                    idxs = _mapear_indices_tabla_provincia(header)
                    # A veces la 1.ª fila no es header: reintentar con fila 1
                    if "fecha" not in idxs and len(tabla) > 2:
                        idxs = _mapear_indices_tabla_provincia(tabla[1])
                        filas = tabla[2:]
                    else:
                        filas = tabla[1:]
                    if "fecha" not in idxs:
                        continue
                    i_fecha = idxs["fecha"]
                    i_desc = idxs.get("desc")
                    i_deb = idxs.get("debito")
                    i_cred = idxs.get("credito")
                    i_saldo = idxs.get("saldo")
                    i_comp = idxs.get("comp")
                    for fila in filas:
                        if not fila or i_fecha >= len(fila):
                            continue
                        fecha = _parsear_fecha(str(fila[i_fecha] or "").strip())
                        if not fecha or not _fecha_plausible_extracto(fecha):
                            continue
                        desc = ""
                        if i_desc is not None and i_desc < len(fila):
                            desc = str(fila[i_desc] or "").replace("\n", " ").strip()
                        if _es_ruido_extracto_provincia(desc) and "saldo" in _normalizar_texto(desc):
                            continue
                        dnorm = _normalizar_texto(desc)
                        if any(x in dnorm for x in ("saldo anterior", "saldo final", "saldo al inicio", "saldo al cierre")):
                            continue
                        debito = _limpiar_monto(fila[i_deb]) if i_deb is not None and i_deb < len(fila) else 0.0
                        credito = _limpiar_monto(fila[i_cred]) if i_cred is not None and i_cred < len(fila) else 0.0
                        saldo = _limpiar_monto(fila[i_saldo]) if i_saldo is not None and i_saldo < len(fila) else 0.0
                        # Una sola columna de importe: clasificar por signo o keywords
                        if debito <= 0 and credito <= 0:
                            # Buscar montos en toda la fila
                            montos_fila = []
                            for cel in fila:
                                montos_fila.extend(_montos_ar_libres(str(cel or "")))
                            if len(montos_fila) >= 2:
                                credito = montos_fila[-2]  # provisional; se ajusta abajo
                                saldo = montos_fila[-1]
                                debito = 0.0
                            elif len(montos_fila) == 1:
                                credito = montos_fila[0]
                        if debito <= 0 and credito <= 0:
                            continue
                        # Si solo hay un lado y hay saldo, no hace falta más
                        if debito > 0 and credito > 0:
                            # A veces duplican; preferir el no-cero distinto del saldo
                            pass
                        comp = ""
                        if i_comp is not None and i_comp < len(fila):
                            comp = str(fila[i_comp] or "").strip()
                        movimientos.append(
                            MovimientoBanco(
                                fecha=fecha,
                                comprobante=comp,
                                descripcion=desc or "Sin descripción",
                                debito=debito if debito > 0 else 0.0,
                                credito=credito if credito > 0 else 0.0,
                                saldo=saldo if saldo else None,
                                banco="provincia",
                                archivo_origen=archivo or "provincia.pdf",
                            )
                        )
    except Exception:
        return []
    return movimientos


def _parsear_movimientos_provincia_paginas(
    paginas: list[tuple[int, str]],
    archivo_origen: str = "",
) -> tuple[list[dict], dict]:
    """
    Parser de texto/OCR del Provincia:
    - montos sin '$' (1.234,56)
    - descripciones multi-línea
    - clasificación por variación de saldo cuando existe
    """
    meta = {"cliente": "", "cuit": "", "cuenta": "", "cbu": "", "periodos": [], "archivo": archivo_origen}
    movs: list[dict] = []
    lineas: list[tuple[int, str]] = []

    for pag, texto in paginas:
        meta_txt = _meta_basica_desde_texto(texto)
        for k in ("cliente", "cuit", "cuenta", "cbu"):
            if meta_txt.get(k) and not meta.get(k):
                meta[k] = meta_txt[k]
        # Cliente: línea con SA / SRL cerca del encabezado
        if not meta["cliente"]:
            for ln in (texto or "").splitlines()[:25]:
                if re.search(r"[A-ZÁÉÍÓÚÑ].*(S\.?R\.?L\.?|S\.?A\.?|S\.A\.S)", ln, re.I):
                    if "provincia" not in ln.lower() and "banco" not in ln.lower():
                        meta["cliente"] = ln.strip()
                        break
        for raw in (texto or "").splitlines():
            ln = re.sub(r"[ \t]+", " ", raw).strip()
            if not ln or _es_ruido_extracto_provincia(ln):
                continue
            low = ln.lower()
            if any(x in low for x in ("total deb", "total cred", "resumen del", "fin del extracto")):
                break
            lineas.append((pag, ln))

    saldo_prev: float | None = None
    fecha_anterior: date | None = None
    i = 0
    while i < len(lineas):
        pag, ln = lineas[i]
        m_fecha = _RE_FECHA_EXT.search(ln)
        if not m_fecha:
            i += 1
            continue
        fecha = _parse_fecha_extracto(m_fecha.group(1))
        if not fecha:
            if fecha_anterior is None:
                i += 1
                continue
            fecha = fecha_anterior

        bloque = [ln]
        j = i + 1
        while j < len(lineas) and j < i + 8:
            _, l2 = lineas[j]
            if _RE_FECHA_EXT.search(l2) and j > i:
                break
            if _es_ruido_extracto_provincia(l2) and not _montos_ar_libres(l2):
                j += 1
                continue
            bloque.append(l2)
            j += 1

        texto_bloque = " ".join(bloque)
        dnorm = _normalizar_texto(texto_bloque)
        if any(x in dnorm for x in ("saldo anterior", "saldo final", "saldo al inicio", "saldo al cierre", "saldo inicial")):
            montos_s = _montos_ar_libres(texto_bloque)
            if montos_s:
                saldo_prev = montos_s[-1]
            fecha_anterior = fecha
            i = j
            continue

        montos = _montos_ar_libres(texto_bloque)
        if not montos:
            i = j
            continue

        # Detección D/H explícita
        marca_dh = None
        m_dh = re.search(
            r"(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})\s*([DHdh])\b|\b([DHdh])\s*(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2})",
            texto_bloque,
        )
        if m_dh:
            marca_dh = (m_dh.group(2) or m_dh.group(3) or "").upper()

        if len(montos) >= 2:
            monto_mov, saldo = montos[-2], montos[-1]
        else:
            monto_mov, saldo = montos[0], None

        descripcion = _RE_FECHA_EXT.sub("", texto_bloque)
        descripcion = _RE_MONTO_AR_LIBRE.sub("", descripcion)
        descripcion = re.sub(r"\b[DHdh]\b", "", descripcion)
        descripcion = re.sub(r"\s+", " ", descripcion).strip(" -|/:")

        if marca_dh == "D":
            debito, credito = abs(monto_mov), 0.0
        elif marca_dh == "H":
            debito, credito = 0.0, abs(monto_mov)
        else:
            debito, credito = _clasificar_debito_credito_extracto(
                descripcion, monto_mov, saldo_prev, saldo
            )

        if debito <= 0 and credito <= 0:
            i = j
            continue

        mes_clave = f"{fecha.year:04d}-{fecha.month:02d}"
        movs.append({
            "Fecha": fecha.strftime("%d/%m/%Y"),
            "Mes": mes_clave,
            "Mes etiqueta": _mes_etiqueta(mes_clave),
            "Periodo extracto": mes_clave,
            "Comprobante": "",
            "Descripcion": descripcion or "Sin descripción",
            "Concepto unificado": unificar_concepto_extracto(descripcion),
            "Debito": debito if debito else None,
            "Credito": credito if credito else None,
            "Saldo": saldo,
            "Pagina PDF": pag,
            "Archivo origen": archivo_origen,
            "Tipo fila": "Movimiento",
            "Banco": "Banco Provincia",
        })
        if mes_clave not in meta["periodos"]:
            meta["periodos"].append(mes_clave)
        fecha_anterior = fecha
        if saldo is not None:
            saldo_prev = saldo
        i = j

    return movs, meta


def _leer_bytes_upload(uploaded) -> tuple[str, bytes]:
    nombre = str(getattr(uploaded, "name", "extracto.pdf"))
    data = uploaded.getvalue() if hasattr(uploaded, "getvalue") else bytes(uploaded)
    return nombre, data


def unir_pdfs_en_memoria(archivos) -> bytes:
    """Concatena varios PDFs en uno solo (orden de carga)."""
    pares: list[tuple[str, bytes]] = []
    for uploaded in archivos or []:
        nombre, data = _leer_bytes_upload(uploaded)
        if nombre.lower().endswith(".pdf"):
            pares.append((nombre, data))
    return unir_pdfs_desde_pares(pares)


def unir_pdfs_desde_pares(pares: list[tuple[str, bytes]]) -> bytes:
    """Concatena PDFs ya leídos como (nombre, bytes)."""
    doc_out = fitz.open()
    try:
        for _nombre, data in pares or []:
            with fitz.open(stream=data, filetype="pdf") as src:
                doc_out.insert_pdf(src)
        return doc_out.tobytes(deflate=True)
    finally:
        doc_out.close()


def _ocr_pagina_rapida(pagina_fitz, dpi: int = 160) -> list[str]:
    """OCR liviano para extractos escaneados (sin reintentos de rotación costosos)."""
    lector = _obtener_lector_ocr()
    pix = pagina_fitz.get_pixmap(dpi=dpi)
    img_pil = Image.open(io.BytesIO(pix.tobytes("png")))
    rot_metadata = getattr(pagina_fitz, "rotation", 0) or 0
    if rot_metadata != 0:
        img_pil = img_pil.rotate(-rot_metadata, expand=True)
    w, h = img_pil.size
    if rot_metadata == 0 and w > h:
        img_pil = img_pil.rotate(90, expand=True)
    # EasyOCR/torch comparte un modelo pesado entre sesiones Streamlit.
    # Serializar inferencias evita picos de RAM y carreras entre usuarios.
    with _lector_ocr_run_lock:
        resultados = lector.readtext(np.array(img_pil))
    filas: dict[int, list[tuple[float, str]]] = {}
    for bbox, texto, _ in resultados:
        y_centro = (bbox[0][1] + bbox[2][1]) / 2
        clave = int(y_centro / 18) * 18
        filas.setdefault(clave, []).append((bbox[0][0], texto))
    return [
        " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0])).strip()
        for y in sorted(filas)
        if any(t.strip() for _, t in filas[y])
    ]


COLUMNAS_EXTRACTO_UNIFICADO = [
    "Fecha", "Mes", "Mes etiqueta", "Periodo extracto", "Comprobante",
    "Descripcion", "Detalle",
    "Concepto unificado", "Clasificacion", "Nueva_Clasificacion", "Tipo Movimiento",
    "Importe", "Debito", "Credito", "Saldo",
    "Pagina PDF", "Archivo origen", "Tipo fila", "Banco", "Valido",
]

# Layout Excel oficial de la herramienta PDF extractos → Excel (todos los bancos).
COLUMNAS_EXCEL_EXTRACTO_BANCO = [
    "Fecha", "Descripcion", "Detalle", "Importe", "Saldo",
    "Clasificacion", "Nueva_Clasificacion", "Tipo Movimiento",
]
HOJAS_EXCEL_EXTRACTO_BANCO = ("Sheet1", "Resumen_Clasificacion")

# Estrategia de parseo por banco (investigación formatos AR 2026):
# - santander: columnas Débito/Crédito separadas → parser dedicado
# - galicia: bloques/tablas Office Banking → tablas pdfplumber + genérico
# - nacion/frances: marcadores D/H + montos ARS → genérico D/H
# - resto: genérico con montos ARS ± heurística de texto
PARSER_EXTRACTO_POR_BANCO: dict[str, str] = {
    "santander": "santander",
    "galicia": "galicia",
    "nacion": "generico_dh",
    "frances": "generico_dh",
    "macro": "generico_dh",
    "credicoop": "generico",
    "provincia": "provincia",
    "icbc": "generico",
    "hsbc": "generico",
    "supervielle": "generico",
    "ciudad": "generico",
    "comafi": "generico",
    "mercadopago": "generico",
    "patagonia": "generico",
    "brubank": "generico",
    "naranja": "generico",
    "desconocido": "generico",
}

# Etiquetas legibles de formato interno (para meta / UI / diagnóstico)
FORMATOS_EXTRACTO_LABEL: dict[str, str] = {
    "santander_dc_saldo": "Santander · Débito/Crédito/Saldo",
    "galicia_office_tabla": "Galicia Office Banking · tablas",
    "galicia_texto": "Galicia · texto/OCR",
    "provincia_bip_tabla": "Provincia BIP · tablas",
    "provincia_texto": "Provincia · texto/OCR",
    "columnas_dc_saldo": "Columnas Débito/Crédito/Saldo",
    "marcadores_dh": "Marcadores D/H (debe/haber)",
    "texto_generico": "Texto genérico / OCR",
}


def detectar_formato_extracto(
    texto: str,
    banco_slug: str = "",
    *,
    tiene_tablas_galicia: bool | None = None,
    tiene_tablas_provincia: bool | None = None,
) -> dict:
    """
    Identifica internamente el layout del extracto para elegir el parser correcto.
    Reduce errores al no forzar un único formato por banco cuando el PDF es distinto
    (digital vs escaneo, CC vs CA, homebanking viejo, etc.).
    """
    raw = texto or ""
    low = _normalizar_texto(raw)
    slug = (banco_slug or "").strip().lower() or "desconocido"

    tiene_dc = bool(
        re.search(r"\bdebito\b", low)
        and re.search(r"\bcredito\b", low)
        and re.search(r"\bsaldo\b", low)
    )
    tiene_dh = bool(
        re.search(r"\b\d{1,3}(?:\.\d{3})*,\d{2}\s*[dh]\b", raw, flags=re.I)
        or re.search(r"\b[dh]\s*\$?\s*\d{1,3}(?:\.\d{3})*,\d{2}\b", raw, flags=re.I)
        or (" debe " in f" {low} " and " haber " in f" {low} ")
    )
    tiene_terceros = bool(_RE_TRANSF_TERCEROS.search(raw) or "terceros" in low)
    office = "office banking" in low or "galicia" in low
    bip = "bip" in low or "banco de la provincia" in low or "bapro" in low

    # Prioridad por banco + pistas de layout
    if slug == "galicia":
        if tiene_tablas_galicia:
            fmt, parser = "galicia_office_tabla", "galicia"
        elif tiene_dc:
            fmt, parser = "columnas_dc_saldo", "santander"
        else:
            fmt, parser = "galicia_texto", "generico"
    elif slug == "santander":
        fmt, parser = "santander_dc_saldo", "santander"
    elif slug == "provincia":
        if tiene_tablas_provincia:
            fmt, parser = "provincia_bip_tabla", "provincia"
        elif tiene_dc:
            fmt, parser = "columnas_dc_saldo", "santander"
        else:
            fmt, parser = "provincia_texto", "provincia"
    elif tiene_dc and not tiene_dh:
        # BBVA/Macro/etc. a veces publican extracto con columnas D/C/Saldo
        fmt, parser = "columnas_dc_saldo", "santander"
    elif slug in {"nacion", "frances", "macro"} or tiene_dh:
        fmt, parser = "marcadores_dh", "generico_dh"
    else:
        # Fallback al mapa estático por banco
        parser = PARSER_EXTRACTO_POR_BANCO.get(slug, "generico")
        if parser == "santander":
            fmt = "santander_dc_saldo"
        elif parser == "galicia":
            fmt = "galicia_texto"
        elif parser == "provincia":
            fmt = "provincia_texto"
        elif parser == "generico_dh":
            fmt = "marcadores_dh"
        else:
            fmt = "texto_generico"

    return {
        "formato_id": fmt,
        "formato": FORMATOS_EXTRACTO_LABEL.get(fmt, fmt),
        "parser": parser,
        "banco_slug": slug,
        "pistas": {
            "columnas_debito_credito_saldo": tiene_dc,
            "marcadores_dh": tiene_dh,
            "transferencias_terceros": tiene_terceros,
            "office_banking": office,
            "bip": bip,
        },
    }


# Pistas fuertes en el nombre de archivo (ganan sobre el texto del cuerpo).
_HINTS_NOMBRE_BANCO: list[tuple[str, list[str]]] = [
    ("santander", [r"\bsantander\b"]),
    ("galicia", [r"\bgalicia\b", r"caja\s*de\s*ahorro"]),
    ("frances", [r"\bfrances\b", r"\bbbva\b"]),
    ("provincia", [r"\bprovincia\b", r"\bbapro\b"]),
    ("nacion", [r"\bnacion\b", r"\bbna\b"]),
    ("macro", [r"\bmacro\b"]),
    ("credicoop", [r"\bcredicoop\b"]),
    ("patagonia", [r"\bpatagonia\b"]),
    ("naranja", [r"\bnaranja\b", r"\bnaranjax\b"]),
    ("mercadopago", [r"\bmercado\s*pago\b", r"\bmercadopago\b"]),
]


def _puntaje_deteccion_banco(texto_norm: str, nombre_norm: str, keywords: list[str]) -> float:
    """
    Scoring ponderado: keywords más largas / específicas pesan más.
    El nombre de archivo pesa ×4 (suele traer el banco: '...Frances CC.pdf').
    """
    puntaje = 0.0
    for kw in keywords:
        kw_n = _normalizar_texto(kw)
        if not kw_n or len(kw_n) < 4:
            continue
        peso = 1.0 + min(3.0, len(kw_n) / 6.0)
        if kw_n in nombre_norm:
            puntaje += peso * 4.0
        elif kw_n in texto_norm:
            puntaje += peso
    return puntaje


def detectar_banco_desde_bytes(data: bytes, nombre_archivo: str = "") -> str:
    """
    Detecta banco desde nombre de archivo + encabezado del PDF.
    No usa el cuerpo de movimientos (evitar falsos positivos por transferencias a otros bancos).
    """
    nombre_norm = _normalizar_texto(nombre_archivo or "")
    # 1) Nombre de archivo decisivo (…Provincia CC.pdf, …Galicia…, etc.)
    for slug, pats in _HINTS_NOMBRE_BANCO:
        for pat in pats:
            if re.search(pat, nombre_norm, flags=re.IGNORECASE):
                return slug

    texto_pdf = ""
    try:
        doc = fitz.open(stream=data, filetype="pdf")
        try:
            # Primera página con texto útil (tapa/logo vacío no cuenta).
            for i in range(min(3, doc.page_count)):
                t = (doc[i].get_text("text") or "").strip()
                if len(t) >= 40:
                    texto_pdf = t
                    break
            if not texto_pdf and doc.page_count:
                texto_pdf = doc[0].get_text("text") or ""
        finally:
            doc.close()
    except Exception:
        texto_pdf = ""
    # Encabezado ~ primeras 1800 chars normalizados
    texto_norm = _normalizar_texto(texto_pdf)[:1800]

    # Marcadores literales de emisor
    if "banco de la provincia" in texto_norm or "bapro" in texto_norm:
        return "provincia"
    if "banco santander" in texto_norm or "santander rio" in texto_norm:
        return "santander"
    if (
        "banco galicia" in texto_norm
        or "office banking" in texto_norm
        or "caja de ahorro en pesos" in texto_norm
        or "resumen de caja de ahorro" in texto_norm
        or "prisma-comercios" in texto_norm
        or re.search(r"\bcbu\s*007", texto_norm)
        or texto_norm.startswith("0070")  # improbable
    ):
        return "galicia"
    if "bbva" in texto_norm or "banco frances" in texto_norm:
        return "frances"

    mejor_banco = ""
    mejor_puntaje = 0.0
    for clave, perfil in PERFILES_BANCO.items():
        if clave == "desconocido":
            continue
        kws = list(perfil.get("keywords") or [])
        p_total = _puntaje_deteccion_banco(texto_norm, nombre_norm, kws)
        if p_total > mejor_puntaje:
            mejor_puntaje = p_total
            mejor_banco = clave
    return mejor_banco if mejor_puntaje > 0 else "desconocido"


def _nombre_display_banco(slug: str) -> str:
    return str((PERFILES_BANCO.get(slug) or {}).get("nombre_display") or slug or "Banco Desconocido")


def _movimiento_banco_a_fila_extracto(mov: MovimientoBanco, periodo: str = "") -> dict:
    """Normaliza MovimientoBanco al schema unificado PDF→Excel (todos los bancos)."""
    fecha = mov.fecha
    mes_clave = f"{fecha.year:04d}-{fecha.month:02d}"
    desc_raw = mov.descripcion or "Sin descripción"
    desc, det = _partir_concepto_y_detalle(desc_raw, None)
    deb = float(mov.debito) if mov.debito else 0.0
    cred = float(mov.credito) if mov.credito else 0.0
    # Signo: egreso (−) / ingreso (+)
    if deb > 0 and cred <= 0:
        importe = -round(deb, 2)
    elif cred > 0:
        importe = round(cred, 2)
    else:
        importe = None
    clas = clasificar_movimiento_extracto(desc, det or "", importe)
    fila = {
        "Fecha": fecha.strftime("%d/%m/%Y"),
        "Mes": mes_clave,
        "Mes etiqueta": _mes_etiqueta(mes_clave),
        "Periodo extracto": periodo or mes_clave,
        "Comprobante": mov.comprobante or "",
        "Descripcion": desc,
        "Detalle": det or None,
        "Concepto unificado": clas,
        "Clasificacion": clas,
        "Nueva_Clasificacion": None,
        "Tipo Movimiento": (
            "Debito" if (importe is not None and importe < 0) else (
                "Credito" if (importe is not None and importe > 0) else ""
            )
        ),
        "Importe": importe,
        "Debito": round(deb, 2) if deb else None,
        "Credito": round(cred, 2) if cred else None,
        "Saldo": float(mov.saldo) if mov.saldo is not None else None,
        "Pagina PDF": int(mov.pagina or 0),
        "Archivo origen": mov.archivo_origen or "",
        "Tipo fila": "Movimiento",
        "Banco": _nombre_display_banco(mov.banco),
        "Valido": True,
    }
    return _aplicar_importe_y_dc(fila, importe)


def _anotar_fila_formato_banco(r: dict) -> dict:
    """Completa Descripcion/Detalle/Clasificacion/Tipo/Importe en cualquier fila de extracto."""
    out = dict(r)
    if str(out.get("Tipo fila") or "") == "Saldo inicial":
        out.setdefault("Detalle", None)
        out.setdefault("Clasificacion", "Saldo Inicial")
        out.setdefault("Nueva_Clasificacion", None)
        out.setdefault("Tipo Movimiento", "")
        out["Concepto unificado"] = "Saldo Inicial"
        return out

    imp = out.get("Importe")
    if imp is None or str(imp).strip() == "":
        deb = float(out.get("Debito") or 0)
        cred = float(out.get("Credito") or 0)
        if deb > 0 and cred <= 0:
            imp = -round(deb, 2)
        elif cred > 0:
            imp = round(cred, 2)
        else:
            imp = None
    try:
        imp_f = round(float(imp), 2) if imp is not None else None
    except (TypeError, ValueError):
        imp_f = None

    desc, det = _partir_concepto_y_detalle(
        str(out.get("Descripcion") or ""),
        str(out.get("Detalle") or "") or None,
    )
    clas = clasificar_movimiento_extracto(desc, det or "", imp_f)
    out["Descripcion"] = desc
    out["Detalle"] = det or None
    out["Clasificacion"] = clas
    out["Nueva_Clasificacion"] = out.get("Nueva_Clasificacion")
    out["Concepto unificado"] = clas
    out["Tipo Movimiento"] = (
        "Debito" if (imp_f is not None and imp_f < 0) else (
            "Credito" if (imp_f is not None and imp_f > 0) else ""
        )
    )
    return _aplicar_importe_y_dc(out, imp_f)


def _filas_desde_paginas_generico(
    paginas: list[tuple[int, str]],
    archivo: str,
    banco_slug: str,
) -> list[dict]:
    movs: list[MovimientoBanco] = []
    for num, texto in paginas:
        lineas = [ln.strip() for ln in (texto or "").splitlines() if ln.strip()]
        if lineas:
            movs.extend(_extraer_movimientos_desde_texto(lineas, num, banco_slug, archivo))
    movs = _deduplicar_y_corregir_saldos(movs)
    return [_movimiento_banco_a_fila_extracto(m) for m in movs]


def _filas_desde_galicia_bytes(data: bytes, archivo: str) -> list[dict]:
    """Intenta tablas nativas Galicia; si no hay, vacío (caller hace fallback genérico)."""
    try:
        movs = extraer_movimientos_galicia_tabla(io.BytesIO(data), archivo=archivo)
    except Exception:
        return []
    if not movs:
        return []
    movs = _deduplicar_y_corregir_saldos(movs)
    return [_movimiento_banco_a_fila_extracto(m) for m in movs]


def _meta_basica_desde_texto(texto: str) -> dict:
    """Heurística liviana de cliente/CUIT/cuenta/CBU en encabezado."""
    meta = {"cliente": "", "cuit": "", "cuenta": "", "cbu": ""}
    if not texto:
        return meta
    m_cuit = re.search(r"\b((?:20|23|24|27|30|33|34)-?\d{8}-?\d)\b", texto)
    if m_cuit:
        meta["cuit"] = m_cuit.group(1)
    m_cbu = re.search(r"\b(\d{22})\b", texto.replace(" ", ""))
    if m_cbu:
        meta["cbu"] = m_cbu.group(1)
    m_cta = re.search(
        r"(?i)(?:cuenta|c\.?\s*cte\.?|caja\s*de\s*ahorro|n[°ºo]\.?\s*cuenta)\s*[:\-]?\s*([\d\-/]{5,})",
        texto,
    )
    if not m_cta:
        # Banco Provincia imprime primero el número y después el tipo:
        # "578285/6 CAJA DE AHORROS EN PESOS".
        m_cta = re.search(
            r"(?i)\b([\d]{5,}/[\d])\s+(?:caja\s+de\s+ahorros?|cuenta\s+corriente)",
            texto,
        )
    if m_cta:
        meta["cuenta"] = m_cta.group(1).strip()
    return meta


def _filas_a_df_extracto(filas: list[dict]) -> pd.DataFrame:
    if not filas:
        return pd.DataFrame(columns=COLUMNAS_EXTRACTO_UNIFICADO)
    filas = [_anotar_fila_formato_banco(r) for r in _corregir_filas_extracto_por_saldos(list(filas))]
    df = pd.DataFrame(filas)
    for col in COLUMNAS_EXTRACTO_UNIFICADO:
        if col not in df.columns:
            df[col] = None
    df = df[COLUMNAS_EXTRACTO_UNIFICADO]
    df = df.drop_duplicates(
        subset=["Fecha", "Comprobante", "Descripcion", "Detalle", "Debito", "Credito", "Saldo", "Archivo origen"],
        keep="first",
    )
    df["_sort"] = df["Fecha"].map(lambda x: _parse_fecha_extracto(str(x)) or date.min)
    return df.sort_values(["_sort", "Pagina PDF"], kind="stable").drop(columns=["_sort"]).reset_index(drop=True)


def _procesar_un_pdf_extracto(nombre: str, data: bytes) -> tuple[list[dict], dict, dict | None]:
    """
    Procesa un PDF. Devuelve (filas, meta_archivo, error_dict|None).
    meta_archivo incluye banco_slug, banco (display), formato detectado, periodos, archivo.
    """
    banco_slug = detectar_banco_desde_bytes(data, nombre)
    display = _nombre_display_banco(banco_slug)
    meta_base = {
        "banco_slug": banco_slug,
        "banco": display,
        "formato_id": "",
        "formato": "",
        "parser": "",
        "archivos": [nombre],
        "periodos": [],
        "cliente": "",
        "cuit": "",
        "cuenta": "",
        "cbu": "",
    }
    cache_galicia: list[dict] | None = None

    def _parsear(paginas: list[tuple[int, str]], estrategia: str) -> tuple[list[dict], dict]:
        movs_loc: list[dict] = []
        meta_arch: dict = {}
        if estrategia == "santander":
            movs_loc, meta_arch = _parsear_movimientos_santander_paginas(paginas, nombre)
            for row in movs_loc:
                row["Banco"] = display
        elif estrategia == "galicia":
            movs_loc = list(cache_galicia) if cache_galicia else _filas_desde_galicia_bytes(data, nombre)
            if not movs_loc:
                # Fallback: layout tipo columnas D/C o texto genérico
                movs_alt, meta_alt = _parsear_movimientos_santander_paginas(paginas, nombre)
                if movs_alt:
                    movs_loc, meta_arch = movs_alt, meta_alt
                else:
                    movs_loc = _filas_desde_paginas_generico(paginas, nombre, "galicia")
                    texto_head = "\n".join(t for _, t in paginas[:2])
                    meta_arch = _meta_basica_desde_texto(texto_head)
            else:
                texto_head = "\n".join(t for _, t in paginas[:2])
                meta_arch = _meta_basica_desde_texto(texto_head)
            for row in movs_loc:
                row["Banco"] = display
        elif estrategia == "provincia":
            movs_tab = extraer_movimientos_provincia_tabla(data, nombre)
            if movs_tab:
                movs_loc = [_movimiento_banco_a_fila_extracto(m) for m in movs_tab]
                for row in movs_loc:
                    row["Banco"] = display
                texto_head = "\n".join(t for _, t in paginas[:2])
                meta_arch = _meta_basica_desde_texto(texto_head)
            else:
                movs_loc, meta_arch = _parsear_movimientos_provincia_paginas(paginas, nombre)
                for row in movs_loc:
                    row["Banco"] = display
                if not movs_loc:
                    movs_loc = _filas_desde_paginas_generico(paginas, nombre, "provincia")
                    texto_head = "\n".join(t for _, t in paginas[:2])
                    meta_arch = _meta_basica_desde_texto(texto_head)
                    for row in movs_loc:
                        row["Banco"] = display
        else:
            # generico / generico_dh
            movs_loc = _filas_desde_paginas_generico(paginas, nombre, banco_slug)
            if not movs_loc:
                # Último recurso: parser de columnas D/C/Saldo (captura transf. a/de terceros)
                movs_loc, meta_arch = _parsear_movimientos_santander_paginas(paginas, nombre)
            else:
                texto_head = "\n".join(t for _, t in paginas[:2])
                meta_arch = _meta_basica_desde_texto(texto_head)
            for row in movs_loc:
                row["Banco"] = display
        return movs_loc, meta_arch

    try:
        paginas = _paginas_texto_extracto_pdf(data, dpi_ocr=160)
        texto_all = "\n".join(t for _, t in paginas)
        chars = sum(len(t) for _, t in paginas)
        fechas_txt = len(re.findall(r"\b\d{2}/\d{2}/\d{2,4}\b", texto_all))
        # _paginas_texto_extracto_pdf ya puede haber aplicado OCR en la primera
        # pasada. Si el PDF no tenía texto nativo, detectar nuevamente el banco
        # sobre ese encabezado OCR antes de elegir formato/parser.
        if banco_slug == "desconocido":
            ocr_norm = _normalizar_texto(texto_all[:2500])
            banco_ocr = ""
            if "provincia" in ocr_norm:
                banco_ocr = "provincia"
            elif "santander" in ocr_norm:
                banco_ocr = "santander"
            elif "galicia" in ocr_norm or "office banking" in ocr_norm:
                banco_ocr = "galicia"
            elif "bbva" in ocr_norm or "banco frances" in ocr_norm:
                banco_ocr = "frances"
            elif "banco nacion" in ocr_norm or "banco de la nacion" in ocr_norm:
                banco_ocr = "nacion"
            elif "banco macro" in ocr_norm:
                banco_ocr = "macro"
            if banco_ocr:
                banco_slug = banco_ocr
                display = _nombre_display_banco(banco_slug)
                meta_base["banco_slug"] = banco_slug
                meta_base["banco"] = display

        # Sondeo liviano de tablas (solo para fingerprint de formato)
        tiene_tab_gal = False
        tiene_tab_prov = False
        if banco_slug == "galicia":
            try:
                cache_galicia = _filas_desde_galicia_bytes(data, nombre)
                tiene_tab_gal = bool(cache_galicia)
            except Exception:
                cache_galicia = None
                tiene_tab_gal = False
        if banco_slug == "provincia":
            try:
                tiene_tab_prov = bool(extraer_movimientos_provincia_tabla(data, nombre))
            except Exception:
                tiene_tab_prov = False

        info_fmt = detectar_formato_extracto(
            texto_all,
            banco_slug,
            tiene_tablas_galicia=tiene_tab_gal if banco_slug == "galicia" else None,
            tiene_tablas_provincia=tiene_tab_prov if banco_slug == "provincia" else None,
        )
        estrategia = str(info_fmt.get("parser") or PARSER_EXTRACTO_POR_BANCO.get(banco_slug, "generico"))
        meta_base["formato_id"] = info_fmt.get("formato_id") or ""
        meta_base["formato"] = info_fmt.get("formato") or ""
        meta_base["parser"] = estrategia

        if (chars < 40 or fechas_txt < 2) and estrategia != "galicia":
            paginas = _paginas_texto_extracto_pdf(data, dpi_ocr=170, forzar_ocr=True)
            chars = sum(len(t) for _, t in paginas)
            texto_all = "\n".join(t for _, t in paginas)
            # En PDFs 100 % escaneados no había texto para detectar el banco al
            # inicio. Repetir la detección sobre el encabezado producido por OCR.
            if banco_slug == "desconocido":
                ocr_norm = _normalizar_texto(texto_all[:2500])
                banco_ocr = ""
                if "banco provincia" in ocr_norm or "banco de la provincia" in ocr_norm:
                    banco_ocr = "provincia"
                elif "banco santander" in ocr_norm or "santander rio" in ocr_norm:
                    banco_ocr = "santander"
                elif "banco galicia" in ocr_norm or "office banking" in ocr_norm:
                    banco_ocr = "galicia"
                elif "bbva" in ocr_norm or "banco frances" in ocr_norm:
                    banco_ocr = "frances"
                elif "banco nacion" in ocr_norm or "banco de la nacion" in ocr_norm:
                    banco_ocr = "nacion"
                elif "banco macro" in ocr_norm:
                    banco_ocr = "macro"
                if banco_ocr:
                    banco_slug = banco_ocr
                    display = _nombre_display_banco(banco_slug)
                    meta_base["banco_slug"] = banco_slug
                    meta_base["banco"] = display
                    estrategia = PARSER_EXTRACTO_POR_BANCO.get(banco_slug, estrategia)
            info_fmt = detectar_formato_extracto(
                texto_all,
                banco_slug,
                tiene_tablas_galicia=tiene_tab_gal if banco_slug == "galicia" else None,
                tiene_tablas_provincia=tiene_tab_prov if banco_slug == "provincia" else None,
            )
            estrategia = str(info_fmt.get("parser") or estrategia)
            meta_base["formato_id"] = info_fmt.get("formato_id") or meta_base["formato_id"]
            meta_base["formato"] = info_fmt.get("formato") or meta_base["formato"]
            meta_base["parser"] = estrategia

        if chars < 40 and estrategia not in {"galicia", "provincia"}:
            return [], meta_base, {
                "archivo": nombre,
                "motivo": f"No se pudo leer texto del PDF ({display}).",
                "banco": display,
                "formato": meta_base.get("formato") or "",
            }

        movs, meta_arch = _parsear(paginas, estrategia)

        # Reintento OCR completo si el parser no encontró movimientos
        if not movs:
            paginas_ocr = _paginas_texto_extracto_pdf(data, dpi_ocr=180, forzar_ocr=True)
            if sum(len(t) for _, t in paginas_ocr) > chars:
                movs, meta_arch = _parsear(paginas_ocr, estrategia)

        # Si el formato elegido falló, probar parser alternativo (terceros suelen aparecer ahí)
        if not movs and estrategia not in {"santander"}:
            movs_alt, meta_alt = _parsear(paginas, "santander")
            if movs_alt:
                movs, meta_arch = movs_alt, meta_alt
                meta_base["parser"] = "santander"
                meta_base["formato_id"] = "columnas_dc_saldo"
                meta_base["formato"] = FORMATOS_EXTRACTO_LABEL["columnas_dc_saldo"]

        if not movs:
            return [], meta_base, {
                "archivo": nombre,
                "motivo": (
                    f"No se detectaron movimientos ({display} · "
                    f"{meta_base.get('formato') or estrategia}). "
                    "Si es escaneo, probá un PDF más nítido o el extracto digital del homebanking."
                ),
                "banco": display,
                "formato": meta_base.get("formato") or "",
            }

        meta_texto = _meta_basica_desde_texto(texto_all)
        for k in ("cliente", "cuit", "cuenta", "cbu"):
            if meta_arch.get(k):
                meta_base[k] = meta_arch[k]
            elif meta_texto.get(k):
                meta_base[k] = meta_texto[k]
        for p in meta_arch.get("periodos") or []:
            if p not in meta_base["periodos"]:
                meta_base["periodos"].append(p)
        for row in movs:
            p = str(row.get("Periodo extracto") or row.get("Mes") or "")
            if p and p not in meta_base["periodos"]:
                meta_base["periodos"].append(p)
        return movs, meta_base, None
    except Exception as exc:
        return [], meta_base, {
            "archivo": nombre,
            "motivo": str(exc),
            "banco": display,
            "formato": meta_base.get("formato") or "",
        }


def exportar_zip_extractos_por_banco(paquetes: list[dict], cuit: str = "") -> bytes:
    """Arma un ZIP con un Excel por banco (formato universal de la herramienta)."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H %M")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in paquetes or []:
            display = str(p.get("banco") or "Banco")
            slug = re.sub(r"[^A-Za-z0-9]+", "_", display).strip("_")[:28] or "banco"
            xlsx = p.get("xlsx")
            if xlsx:
                zf.writestr(f"{stamp}_{slug}.xlsx", xlsx)
            pdf_bytes = p.get("pdf_merged")
            archivos = p.get("meta", {}).get("archivos") or []
            if pdf_bytes and len(archivos) > 1:
                zf.writestr(f"{stamp}_{slug}_unificado.pdf", pdf_bytes)
    return buf.getvalue()


def procesar_extractos_bancarios_pdfs(archivos) -> tuple[pd.DataFrame, dict, list[dict]]:
    """
    Convertidor multi-banco: detecta el banco de cada PDF y aplica el parser adecuado.
    **Un paquete por banco** (no mezcla bancos distintos como 'meses' del mismo extracto).
    meta['por_banco'] tiene un ítem por banco con df/meta propios.
    El DataFrame retornado concatena todos (compatibilidad match/proveedores).
    """
    errores: list[dict] = []
    grupos: dict[str, dict] = {}
    # Conservar bytes por archivo para PDF merge intra-banco
    bytes_por_nombre: dict[str, bytes] = {}

    for uploaded in archivos or []:
        nombre, data = _leer_bytes_upload(uploaded)
        if not nombre.lower().endswith(".pdf"):
            errores.append({"archivo": nombre, "motivo": "Solo se admiten PDF."})
            continue
        bytes_por_nombre[nombre] = data
        filas, meta_arch, err = _procesar_un_pdf_extracto(nombre, data)
        if err:
            errores.append(err)
            continue
        slug = meta_arch["banco_slug"]
        g = grupos.get(slug)
        if g is None:
            g = {
                "banco_slug": slug,
                "banco": meta_arch["banco"],
                "filas": [],
                "meta": {
                    "cliente": "",
                    "cuit": "",
                    "cuenta": "",
                    "cbu": "",
                    "periodos": [],
                    "archivos": [],
                    "bancos": [meta_arch["banco"]],
                    "banco": meta_arch["banco"],
                    "banco_slug": slug,
                    "formato_id": meta_arch.get("formato_id") or "",
                    "formato": meta_arch.get("formato") or "",
                    "parser": meta_arch.get("parser") or "",
                    "formatos_por_archivo": [],
                },
            }
            grupos[slug] = g
        g["filas"].extend(filas)
        g["meta"]["archivos"].append(nombre)
        if meta_arch.get("formato"):
            g["meta"]["formatos_por_archivo"].append({
                "archivo": nombre,
                "formato": meta_arch.get("formato"),
                "formato_id": meta_arch.get("formato_id") or "",
                "parser": meta_arch.get("parser") or "",
            })
            # Conservar el último formato visto como resumen del banco
            g["meta"]["formato"] = meta_arch.get("formato") or g["meta"].get("formato")
            g["meta"]["formato_id"] = meta_arch.get("formato_id") or g["meta"].get("formato_id")
            g["meta"]["parser"] = meta_arch.get("parser") or g["meta"].get("parser")
        for k in ("cliente", "cuit", "cuenta", "cbu"):
            if meta_arch.get(k) and not g["meta"].get(k):
                g["meta"][k] = meta_arch[k]
        for p in meta_arch.get("periodos") or []:
            if p not in g["meta"]["periodos"]:
                g["meta"]["periodos"].append(p)

    por_banco: list[dict] = []
    frames: list[pd.DataFrame] = []
    for slug, g in grupos.items():
        df_b = _filas_a_df_extracto(g["filas"])
        meta_b = g["meta"]
        # PDF unificado solo dentro del mismo banco
        pdf_merged = None
        try:
            nombres = list(meta_b.get("archivos") or [])
            pares = [(n, bytes_por_nombre[n]) for n in nombres if n in bytes_por_nombre]
            if pares:
                pdf_merged = unir_pdfs_desde_pares(pares)
        except Exception:
            pdf_merged = None
        por_banco.append({
            "banco_slug": slug,
            "banco": g["banco"],
            "df": df_b,
            "meta": meta_b,
            "pdf_merged": pdf_merged,
        })
        if not df_b.empty:
            frames.append(df_b)

    bancos_display = [p["banco"] for p in por_banco]
    meta = {
        "cliente": next((p["meta"].get("cliente") for p in por_banco if p["meta"].get("cliente")), ""),
        "cuit": next((p["meta"].get("cuit") for p in por_banco if p["meta"].get("cuit")), ""),
        "cuenta": "",
        "cbu": "",
        "periodos": [],
        "archivos": [a for p in por_banco for a in (p["meta"].get("archivos") or [])],
        "bancos": bancos_display,
        "banco": " | ".join(bancos_display) if bancos_display else "Banco Desconocido",
        "por_banco": por_banco,
        "multi_banco": len(por_banco) > 1,
    }
    for p in por_banco:
        for per in p["meta"].get("periodos") or []:
            if per not in meta["periodos"]:
                meta["periodos"].append(per)

    if not frames:
        return pd.DataFrame(columns=COLUMNAS_EXTRACTO_UNIFICADO), meta, errores
    df = pd.concat(frames, ignore_index=True)
    df["_sort"] = df["Fecha"].map(lambda x: _parse_fecha_extracto(str(x)) or date.min)
    df = df.sort_values(["Banco", "_sort", "Pagina PDF"], kind="stable").drop(columns=["_sort"]).reset_index(drop=True)
    return df, meta, errores


def procesar_extractos_santander_pdfs(archivos) -> tuple[pd.DataFrame, dict, list[dict]]:
    """
    Compatibilidad: ahora delega al convertidor multi-banco.
    (Históricamente solo Santander; se mantiene el nombre por imports existentes.)
    """
    return procesar_extractos_bancarios_pdfs(archivos)


def _paginas_texto_extracto_pdf(
    data: bytes, dpi_ocr: int = 160, forzar_ocr: bool = False
) -> list[tuple[int, str]]:
    """
    Extrae texto por página: nativo si existe; si la página está vacía (escaneada), OCR.
    Con forzar_ocr=True reaplica OCR a todas las páginas (útiles para Provincia/BIP escaneados).
    """
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        paginas: list[tuple[int, str]] = []
        nativas_con_texto = 0
        for i in range(doc.page_count):
            texto = (doc[i].get_text("text") or "").strip()
            if texto:
                nativas_con_texto += 1
                paginas.append((i + 1, texto))
            else:
                paginas.append((i + 1, ""))
        necesita_ocr = forzar_ocr or nativas_con_texto < max(1, doc.page_count // 3)
        if necesita_ocr:
            for i, (num, texto) in enumerate(paginas):
                if forzar_ocr or not texto.strip():
                    lineas_ocr = _ocr_pagina_rapida(doc[i], dpi=dpi_ocr)
                    paginas[i] = (num, "\n".join(lineas_ocr))
        return paginas
    finally:
        doc.close()


# Prefijos de concepto Galicia / extractos AR (más largos primero) para partir Descripción/Detalle.
_PREFIJOS_CONCEPTO_EXTRACTO: tuple[str, ...] = tuple(
    sorted(
        (
            "TRANSFERENCIA DE TERCEROS",
            "TRANSFERENCIA A TERCEROS",
            "TRANSFERENCIA DE CUENTA PROPIA",
            "TRANSFERENCIA DE CUENTA",
            "TRANSFERENCIAS CASH PROVEEDORES",
            "TRANSFERENCIAS CASH",
            "TRANSFERENCIA REALIZADA",
            "CREDITO TRANSFERENCIA",
            "CREDITOS VARIOS",
            "TRF INMED PROVEED",
            "TRANSF INMED CP",
            "TRANSF. CTAS PROPIAS",
            "PAGO CON TRANSFERENCIA",
            "SERVICIO PAGO A PROVEEDORES",
            "SERVICIO ACREDITAMIENTO DE HABERES",
            "SERVICIO ACREDITAMIENTO DE",
            "ACREDITAMIENTO PRISMA-COMERCIOS VISA",
            "ACREDITAMIENTO PRISMA-COMERCIOS",
            "ING. BRUTOS S/ CRED",
            "IMP. CRE. LEY 25413 ALICUOTA GENERAL",
            "IMP. CRE. LEY 25413",
            "IMP. DEB. LEY 25413 GRAL.",
            "IMP. DEB. LEY 25413",
            "IMPUESTO LEY 25.413",
            "IMPUESTO LEY 25413",
            "IMP. ING. BRUTOS",
            "IMPUESTO DE SELLOS",
            "PAGO DE SERVICIOS",
            "PAGO TARJETA VISA D.A. AL VTO",
            "PAGO TARJETA VISA",
            "PAGO TARJETA MASTER",
            "PAGO TARJETA MASTERCARD",
            "DEBITO DEBIN",
            "TRANSF. AFIP",
            "RESCATE FIMA",
            "SUSCRIPCION FIMA",
            "COMISION SERVICIO DE CUENTA",
            "COMISION Y DERECHOS DE",
            "COM. DEPOSITO DE CHEQUE",
            "COMISION POR CUSTODIA DE",
            "PERCEP. IVA",
            "COMP. TITULOS / VAL.",
            "COMP. TITULOS",
            "INTERESES SOBRE SALDOS",
            "INTERES CAPITALIZADO",
            "DEB. AUTOM. DE SERV.",
            "DEP.EFVO.AUTOSERVICIO",
            "DEP. EFVO",
            "G.DE ECHEQ",
            "COMPRA DEBITO",
            "IVA",
        ),
        key=len,
        reverse=True,
    )
)

_RE_CORTE_DETALLE = re.compile(
    r"(?="
    r"\b\d{11}\b"  # CUIT
    r"|\b\d{16,22}\b"  # CBU / refs largas
    r"|\bVARIOS\b"
    r"|\bFACTURA\b"
    r"|\bNro\.?\s*Operacion\b"
    r"|\bOPERACION\b"
    r"|\bNRO\.\s*"
    r"|\bBANCO\b"
    r"|\bCITIBANK\b"
    r"|\bMERCADO\s+LIBRE\b"
    r"|\bCOELSA\b"
    r"|\bTICKET\s*:"
    r")",
    re.IGNORECASE,
)


def _es_basura_extracto_concepto(texto: str) -> bool:
    """Pie de página / leyendas / saldos USD que no son movimientos."""
    raw = str(texto or "").strip()
    if not raw:
        return False  # vacío ≠ basura (impuestos sin detalle son válidos)
    t = _normalizar_texto(raw)
    if t in {"nan", "none", "usd", "total", "saldos", "movimientos"}:
        return True
    if t.startswith("dispon") and "cuestionar" in t:
        return True
    if "garantia de hasta" in t or "fondo de garantia" in t:
        return True
    if "credito fiscal discriminado" in t:
        return True
    if t.startswith("sin movimientos"):
        return True
    if t.startswith("promedio "):
        return True
    if t.startswith("total retencion") or t.startswith("total mensual retencion"):
        return True
    if re.fullmatch(r"usd|ars|\$|u\$s", t):
        return True
    return False


def _formatear_detalle_extracto(texto: str) -> str:
    """
    Deja el Detalle en líneas separadas (como el Excel de referencia):
    nombre / CUIT / CBU / VARIOS / banco / operación.
    Así se puede leer y, si hace falta, partir en Excel.
    """
    raw = str(texto or "").replace("\r\n", "\n").replace("\r", "\n")
    partes: list[str] = []
    for bloque in raw.split("\n"):
        bloque = re.sub(r"[ \t]+", " ", bloque).strip(" -|/")
        if not bloque or _normalizar_texto(bloque) in {"nan", "none", "total", "usd", "ars"}:
            continue
        # Insertar cortes ante CUIT / CBU / VARIOS / banco / operación
        trozos = _RE_CORTE_DETALLE.split(bloque)
        buf = ""
        for trozo in trozos:
            trozo = (trozo or "").strip(" -|/")
            if not trozo:
                continue
            if _RE_CORTE_DETALLE.match(trozo) and buf:
                partes.append(buf.strip())
                buf = trozo
            elif _RE_CORTE_DETALLE.match(trozo) and not buf:
                buf = trozo
            else:
                buf = f"{buf} {trozo}".strip() if buf else trozo
        if buf:
            partes.append(buf.strip())
    # Limpiar duplicados consecutivos y basura corta
    out: list[str] = []
    for p in partes:
        p = re.sub(r"\s+", " ", p).strip(" -|/")
        if not p or _normalizar_texto(p) in {"nan", "none", "total", "usd"}:
            continue
        if out and _normalizar_texto(out[-1]) == _normalizar_texto(p):
            continue
        out.append(p)
    return "\n".join(out)


def _partir_concepto_y_detalle(descripcion: str, detalle: str | None = None) -> tuple[str, str]:
    """Separa concepto corto (Descripcion) del resto (Detalle), en líneas distintas."""
    det_in = _formatear_detalle_extracto(detalle or "")
    raw = str(descripcion or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw and not det_in:
        return "Sin descripción", ""

    # Si ya viene multilínea: 1ª = concepto, resto = detalle
    if "\n" in raw:
        partes = [p.strip() for p in raw.splitlines() if p.strip()]
        if len(partes) >= 2:
            head = re.sub(r"\s+", " ", partes[0]).strip()
            resto = _formatear_detalle_extracto("\n".join(partes[1:]))
            if det_in:
                resto = _formatear_detalle_extracto(f"{resto}\n{det_in}")
            # Solo prefijos sobre la 1ª línea (sin reentrar por newlines)
            return _partir_solo_prefijo(head, resto)

    return _partir_solo_prefijo(re.sub(r"\s+", " ", raw).strip(), det_in)


def _partir_solo_prefijo(raw_norm: str, det_in: str) -> tuple[str, str]:
    """Aplica prefijos conocidos y cortes por CUIT/nombre; no espera newlines en raw."""
    if not raw_norm:
        return "Sin descripción", det_in
    up = raw_norm.upper().strip()

    for pref in _PREFIJOS_CONCEPTO_EXTRACTO:
        if up.startswith(pref):
            resto_txt = raw_norm[len(pref):].strip(" -|/\t")
            m_cont = re.match(
                r"^(PROPIA|PROVEEDORES|ALICUOTA\s+GENERAL|GRAL\.?|D\.A\.\s+AL\s+VTO)\b[\s\-]*",
                resto_txt,
                flags=re.I,
            )
            concepto = pref
            if m_cont:
                concepto = re.sub(r"\s+", " ", f"{pref} {m_cont.group(1)}").strip()
                resto_txt = resto_txt[m_cont.end():].strip(" -|/")
            resto = _formatear_detalle_extracto(resto_txt)
            if det_in:
                resto = _formatear_detalle_extracto(f"{resto}\n{det_in}" if resto else det_in)
            return concepto.upper() if concepto == pref else concepto, resto

    m = re.search(
        r"\s+(\d{11}\b|\d{16,}\b|Nro\.?\s*Operacion|OPERACION\b|VARIOS\b|BANCO\b)",
        raw_norm,
        flags=re.I,
    )
    if m and m.start() > 8:
        head = raw_norm[: m.start()].strip(" -|/")
        tail = _formatear_detalle_extracto(raw_norm[m.start():])
        if det_in:
            tail = _formatear_detalle_extracto(f"{tail}\n{det_in}" if tail else det_in)
        return head or raw_norm, tail

    m2 = re.search(
        r"^(.{6,70}?)\s+([A-ZÁÉÍÓÚ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚa-záéíóúñ.]+){1,4}\b.*)$",
        raw_norm,
    )
    if m2 and not re.search(r"\d{4,}", m2.group(1)):
        # Evitar partir impuestos / leyendas
        head_l = _normalizar_texto(m2.group(1))
        if not any(k in head_l for k in ("imp", "iva", "ley", "comision", "interes")):
            head, tail = m2.group(1).strip(), _formatear_detalle_extracto(m2.group(2))
            if det_in:
                tail = _formatear_detalle_extracto(f"{tail}\n{det_in}" if tail else det_in)
            return head, tail

    return raw_norm, det_in


def clasificar_movimiento_extracto(
    descripcion: str,
    detalle: str = "",
    importe: float | None = None,
) -> str:
    """Clasificación contable alineada al Excel de referencia (Galicia)."""
    texto = f"{descripcion or ''} {detalle or ''}"
    low = _normalizar_texto(texto)
    imp = float(importe or 0)

    if "arba" in low:
        return "Pago ARBA"
    if "sircreb" in low:
        return "SIRCREB"
    if "tucuman" in low and ("ing" in low and "bruto" in low):
        return "Ingresos brutos Tucuman"
    if re.search(r"imp\.?\s*(deb|cre)\.?\s*ley\s*25\.?413", low) or "ley 25413" in low or "ley 25.413" in low:
        return "Impuestos a los débitos y créditos"
    if "impuesto de sellos" in low or "imp. sellos" in low:
        return "Impuesto a los sellos"
    if "percep" in low and "iva" in low:
        return "Percepción IVA"
    if re.search(r"(^|\s)iva(\s|$)", low) and "percep" not in low:
        return "IVA"
    if ("imp" in low and "ing" in low and "bruto" in low) or "iibb" in low:
        return "IIBB"
    if "ing" in low and "bruto" in low and "cred" in low:
        return "IIBB"
    if "rescate fima" in low or "rescate fci" in low:
        return "Rescate FIMA"
    if "suscripcion fima" in low or "suscripcion fci" in low or "suscripción fima" in low:
        return "Suscripción FCI"
    # Comisiones bancarias antes que cheques / títulos
    if (
        "comision" in low
        or "com. deposito" in low
        or "com deposito" in low
        or "derechos de" in low
        or ("gasto" in low and "banco" in low)
        or "servicio de cuenta" in low
        or "payway" in low
        or ("terminal" in low and "visa" in low)
    ):
        return "Gastos Bancarios"
    if "comp" in low and ("titulo" in low or "val." in low or "bono" in low) and "comision" not in low:
        return "Inversiones"
    if "interes" in low and (
        "saldo" in low or "capitaliz" in low or "acredit" in low
    ):
        return "Intereses"
    if "dep" in low and ("efvo" in low or "efectivo" in low):
        return "Depositos en efvo"
    if "echeq" in low or re.search(r"\bcheq", low):
        return "Cheques recibidos" if imp >= 0 else "Cheques emitidos"
    if "haberes" in low or "acreditamiento de haberes" in low:
        return "Pago de haberes"
    if "afip" in low or re.search(r"\bvep\b", low):
        return "Pagos AFIP"
    if "pago tarjeta" in low or "tarjeta visa" in low:
        return "Pagos tarjeta corporativa"
    if "compra debito" in low or "compra con tarjeta" in low:
        return "Compras"
    if "deb" in low and "autom" in low:
        # Débito automático: servicios vs compras (abono)
        if "abono" in low:
            return "Compras"
        return "Pago de Servicios"
    if "reintegro" in low or ("promo" in low and "galicia" in low):
        return "Transferencias recibidas" if imp >= 0 else "Compras"
    if "pago de servicio" in low:
        return "Pago de Servicios"
    if "prisma" in low or "acreditamiento prisma" in low:
        return "Acreditaciones comercios"
    if "pago a proveedores" in low and imp > 0:
        return "Pagos recibidos"
    if (
        "trf inmed" in low
        or "transf inmed" in low
        or "transf. ctas propias" in low
        or "transferencia a terceros" in low
        or (
            "transf" in low
            and "terceros" in low
            and " de terceros" not in f" {low} "
            and imp < 0
        )
    ):
        return "Transferencias emitidas"
    if "transferencia de terceros" in low or (
        "transf" in low and "terceros" in low and imp > 0
    ):
        return "Transferencias recibidas"
    if "credito transferencia" in low or ("transferencia" in low and imp > 0):
        return "Transferencias recibidas"
    if "transferencia" in low or "trf " in low or "transf " in low:
        return "Transferencias emitidas" if imp < 0 else "Transferencias recibidas"
    return "Sin clasificar"


def _fecha_extracto_corta(fecha_txt: object) -> str:
    """dd/mm/yy como en el Excel de referencia."""
    s = str(fecha_txt or "").strip()
    f = _parse_fecha_extracto(s)
    if f:
        return f.strftime("%d/%m/%y")
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d).strftime("%d/%m/%y")
        except ValueError:
            pass
    return s


def enriquecer_df_extracto_formato_banco(df: pd.DataFrame) -> pd.DataFrame:
    """
    Formato universal de la herramienta (todos los bancos):
    Fecha | Descripcion | Detalle | Importe | Saldo | Clasificacion | Nueva_Clasificacion | Tipo Movimiento
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=list(COLUMNAS_EXCEL_EXTRACTO_BANCO))
    out = df.copy()
    if "Tipo fila" in out.columns:
        out = out[out["Tipo fila"] != "Saldo inicial"].copy()

    if "Descripcion" in out.columns:
        desc_s = out["Descripcion"].astype(str)
        ruido = desc_s.map(_es_basura_extracto_concepto) | desc_s.str.contains(
            r"(?i)estan garantizados|garantizados por|fondo de garantia|detalle impositivo|cuestionar este resumen|credito fiscal discriminado",
            na=False,
        )
        if "Detalle" in out.columns:
            # Limpiar basura de celda, no borrar el movimiento
            out["Detalle"] = out["Detalle"].map(
                lambda v: None
                if v is None or _normalizar_texto(str(v)) in {"nan", "none", "null"}
                else v
            )
            det_s = out["Detalle"].astype(str)
            # Solo descartar si el DETALLE entero es leyenda legal (no un movimiento)
            ruido = ruido | det_s.str.contains(
                r"(?i)cuestionar este resumen|credito fiscal discriminado|garantia de hasta 0",
                na=False,
            )
        out = out.loc[~ruido].copy()

    anotadas = [_anotar_fila_formato_banco(row.to_dict()) for _, row in out.iterrows()]
    if not anotadas:
        return pd.DataFrame(columns=list(COLUMNAS_EXCEL_EXTRACTO_BANCO))
    out = pd.DataFrame(anotadas)
    out["Fecha"] = out["Fecha"].map(_fecha_extracto_corta)
    out["Nueva_Clasificacion"] = None
    for c in COLUMNAS_EXCEL_EXTRACTO_BANCO:
        if c not in out.columns:
            out[c] = None
    return out


def armar_resumen_clasificacion_extracto(df: pd.DataFrame) -> pd.DataFrame:
    """Hoja Resumen_Clasificacion: Debe / Haber / Importe_Neto por clasificación."""
    cols = ["Clasificacion", "Debe", "Haber", "Importe_Neto", "Moneda"]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)
    work = enriquecer_df_extracto_formato_banco(df)
    if work.empty:
        return pd.DataFrame(columns=cols)
    work["Importe"] = pd.to_numeric(work["Importe"], errors="coerce").fillna(0.0)
    rows = []
    for clas, g in work.groupby("Clasificacion", dropna=False):
        debe = float((-g.loc[g["Importe"] < 0, "Importe"]).sum())
        haber = float(g.loc[g["Importe"] > 0, "Importe"].sum())
        rows.append({
            "Clasificacion": clas or "Sin clasificar",
            "Debe": round(debe, 2),
            "Haber": round(haber, 2),
            "Importe_Neto": round(haber - debe, 2),
            "Moneda": None,
        })
    out = pd.DataFrame(rows, columns=cols).sort_values("Clasificacion").reset_index(drop=True)
    return out


def armar_tabla_dinamica_conceptos(df: pd.DataFrame) -> pd.DataFrame:
    """Compat: resumen por clasificación (antes concepto unificado × mes)."""
    return armar_resumen_clasificacion_extracto(df)


def exportar_extracto_santander_excel(df: pd.DataFrame, meta: dict | None = None) -> bytes:
    """
    Formato universal PDF extractos → Excel (Galicia, Santander, Provincia, etc.):
    - Sheet1: Fecha | Descripcion | Detalle | Importe | Saldo | Clasificacion | Nueva_Clasificacion | Tipo Movimiento
    - Resumen_Clasificacion: Clasificacion | Debe | Haber | Importe_Neto | Moneda
    Importe con signo: (−) resta / (+) suma.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    meta = meta or {}
    df_in = df.copy() if df is not None else pd.DataFrame()
    work = enriquecer_df_extracto_formato_banco(df_in)
    cols = list(COLUMNAS_EXCEL_EXTRACTO_BANCO)
    for c in cols:
        if c not in work.columns:
            work[c] = None
    df_out = work[cols].copy()
    df_out["Importe"] = pd.to_numeric(df_out["Importe"], errors="coerce")
    df_out["Saldo"] = pd.to_numeric(df_out["Saldo"], errors="coerce")

    resumen = armar_resumen_clasificacion_extracto(df_in)

    # Identidad visual Estudio (mismo header que informes Claude-style)
    from excel_formato_estudio import COLOR_PRIMARIO, COLOR_ZEBRA, HDR_FONT, MONEY_FMT_SIGNED

    header_font = HDR_FONT
    body_font = Font(name="Calibri", size=11, color="000000")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fill_h = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    zebra = PatternFill("solid", fgColor=COLOR_ZEBRA)

    wb = Workbook()
    ws = wb.active
    ws.title = HOJAS_EXCEL_EXTRACTO_BANCO[0]
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill_h
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=max(2, ws.max_row), max_col=8):
        for cell in row:
            cell.font = body_font
            cell.border = thin
            if cell.column in (4, 5) and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT_SIGNED
                cell.alignment = Alignment(horizontal="right")
            if cell.column in (2, 3):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.row % 2 == 0:
                cell.fill = zebra
    ws.row_dimensions[1].height = 18
    for r in range(2, ws.max_row + 1):
        det = ws.cell(r, 3).value
        if isinstance(det, str) and "\n" in det:
            ws.row_dimensions[r].height = min(15 * (1 + det.count("\n")), 75)
    anchos = {
        "A": 12, "B": 36, "C": 42, "D": 14, "E": 14, "F": 32, "G": 22, "H": 16,
    }
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:H{ws.max_row}"

    ws2 = wb.create_sheet(HOJAS_EXCEL_EXTRACTO_BANCO[1])
    headers_r = ["Clasificacion", "Debe", "Haber", "Importe_Neto", "Moneda"]
    ws2.append(headers_r)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = fill_h
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for _, row in resumen.iterrows():
        ws2.append([
            row.get("Clasificacion"),
            float(row.get("Debe") or 0),
            float(row.get("Haber") or 0),
            float(row.get("Importe_Neto") or 0),
            row.get("Moneda"),
        ])
    last_data = ws2.max_row
    if last_data >= 2:
        total_row = last_data + 1
        ws2.cell(total_row, 1, "TOTAL")
        ws2.cell(total_row, 2, f"=SUM(B2:B{last_data})")
        ws2.cell(total_row, 3, f"=SUM(C2:C{last_data})")
        ws2.cell(total_row, 4, f"=SUM(D2:D{last_data})")
        for col in range(1, 6):
            ws2.cell(total_row, col).font = Font(name="Calibri", bold=True, size=11)
            ws2.cell(total_row, col).border = thin
        for r in range(2, total_row + 1):
            for c in (2, 3, 4):
                cell = ws2.cell(r, c)
                if isinstance(cell.value, (int, float)) or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    cell.number_format = MONEY_FMT_SIGNED
                    cell.alignment = Alignment(horizontal="right")
                ws2.cell(r, 1).border = thin
                ws2.cell(r, 5).border = thin
                for c in (2, 3, 4):
                    ws2.cell(r, c).border = thin
    for i, w in enumerate([36, 14, 14, 14, 10], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    _ = meta
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_extracto_bancario_excel(df: pd.DataFrame, meta: dict | None = None) -> bytes:
    """Exportador oficial multi-banco (mismo layout universal para todos)."""
    return exportar_extracto_santander_excel(df, meta)


# ---------------------------------------------------------------------------
# Match débitos bancarios ↔ facturas de proveedores (ejercicio)
# ---------------------------------------------------------------------------

PROVEEDORES_DEFAULT_PATH = Path(r"T:\Estudio Contable\Proveedores.xlsx")
TOL_MATCH_MONTO_PESOS = 50.0
TOL_MATCH_MONTO_PCT = 0.005  # 0,5 %
TOL_MATCH_DIAS = 7  # factura vs pago (fin de semana / demora operativa)
TOL_MATCH_DIAS_SOLO_MONTO = 10  # sin nombre útil: monto + fecha
MIN_SCORE_MATCH_PROVEEDOR = 58.0
MIN_SCORE_MATCH_MONTO_FECHA = 70.0  # pase monto±fecha: monto casi exacto
MIN_SCORE_MATCH_RETARDO = 62.0  # pase retardo: nombre+monto sin fecha
MIN_FUZZY_NOMBRE_MATCH = 62
MIN_FUZZY_NOMBRE_RETARDO = 70  # retardo exige proveedor más claro
MIN_FUZZY_NOMBRE_SUMA = 70  # sumas parciales con nombre
MAX_NOMBRE_PARA_SOLO_MONTO = 55  # debajo = banco sin data útil de proveedor
TOL_MATCH_DIAS_SUMA_SIN_NOMBRE = 30  # sumas sin beneficiario: pagos parciales en el mes
TOL_MATCH_DIAS_MONTO_UNICO = 20  # 1:1 monto único remanente (sin nombre en banco)
MAX_POOL_SUMA = 50  # tope de candidatos por combinación (evita explosión)


def _es_descripcion_bancaria_sin_proveedor(desc_norm: str) -> bool:
    """True si el extracto no trae razón social (Macro N/D, TRANSF:id, etc.)."""
    d = (desc_norm or "").strip().lower()
    if not d:
        return True
    markers = (
        "macronline",
        "macr online",
        "n/d transf",
        "nd transf",
        "n d transf",
        "transf:",
        "db transf minorista",
        "e-set d/t",
        "eset d/t",
        "n/d db transf",
        "nd db transf",
    )
    # Macro / N/D: el pie de garantía u OCR no cuenta como beneficiario
    if any(m in d for m in markers):
        return True
    # Transferencia realizada / online sin nombre útil
    genericos_solo = (
        "transferencia realizada",
        "transf. online",
        "transf online",
    )
    if any(m in d for m in genericos_solo):
        genericos = {
            "transf", "transferencia", "realizada", "online", "banking", "macronline",
            "macr", "minorista", "dist", "tit", "nd", "n/d", "e-set", "eset", "d/t",
            "pago", "de", "a", "el", "la", "los", "las", "sa", "srl", "sas",
        }
        tokens = [t for t in d.replace(":", " ").split() if len(t) > 2 and t not in genericos]
        return len(tokens) < 2
    return False


def _dias_abs_fechas(a, b) -> int | None:
    if not a or not b:
        return None
    try:
        return abs((a - b).days)
    except Exception:
        return None


# Reexport SQL Tango (listado 21101 sin Excel manual)
try:
    from tango_sql import (  # noqa: E402
        TangoSQLError,
        cargar_facturas_proveedores_tango_sql,
        listar_empresas_tango,
        resolver_empresa_tango,
        tango_sql_disponible,
    )
except ImportError:  # pragma: no cover
    TangoSQLError = RuntimeError  # type: ignore[misc,assignment]

    def tango_sql_disponible() -> tuple[bool, str]:
        return False, "Módulo tango_sql no disponible"

    def listar_empresas_tango(*_a, **_k):
        return []

    def resolver_empresa_tango(*_a, **_k):
        raise TangoSQLError("Módulo tango_sql no disponible")

    def cargar_facturas_proveedores_tango_sql(*_a, **_k):
        raise TangoSQLError("Módulo tango_sql no disponible")


def _leer_excel_flexible(fuente) -> pd.DataFrame:
    """Lee Excel desde path, bytes, BytesIO o uploaded file."""
    if isinstance(fuente, (str, Path)):
        return pd.read_excel(fuente, sheet_name=None)
    data = fuente.getvalue() if hasattr(fuente, "getvalue") else fuente
    return pd.read_excel(io.BytesIO(data), sheet_name=None)


def cargar_facturas_proveedores_excel(fuente) -> pd.DataFrame:
    """
    Normaliza listado Tango 21101 (Proveedores.xlsx) u Excel similar.
    Facturas = Haber > 0 (FCC / NDC). Notas de crédito (Debe) se excluyen del pool a pagar.
    """
    hojas = _leer_excel_flexible(fuente)
    if isinstance(hojas, dict):
        # Preferir hoja con Razón social / Haber
        df = None
        for _nombre, frame in hojas.items():
            cols_n = {_normalizar_texto(str(c)) for c in frame.columns}
            if any("razon" in c or "social" in c for c in cols_n) and any(
                "haber" in c or "debe" in c for c in cols_n
            ):
                df = frame
                break
        if df is None:
            # segunda hoja típica del export Tango
            df = list(hojas.values())[1] if len(hojas) > 1 else list(hojas.values())[0]
    else:
        df = hojas

    mapeo: dict[str, str] = {}
    for col in df.columns:
        cn = _normalizar_texto(str(col))
        if "fecha" in cn and "fecha" not in mapeo.values():
            mapeo[col] = "fecha"
        elif ("razon" in cn or "social" in cn or cn == "proveedor") and "proveedor" not in mapeo.values():
            mapeo[col] = "proveedor"
        elif "cuit" in cn and "cuit" not in mapeo.values():
            mapeo[col] = "cuit"
        elif ("numero" in cn and "comprob" in cn) or cn in {"comprobante", "nro comprobante"}:
            mapeo[col] = "comprobante"
        elif "tipo comprob" in cn or cn == "tipo":
            mapeo[col] = "tipo"
        elif cn == "debe":
            mapeo[col] = "debe"
        elif cn == "haber":
            mapeo[col] = "haber"
        elif "cuenta" in cn and "cuenta" not in mapeo.values():
            mapeo[col] = "cuenta"
        elif "descrip" in cn and "descripcion" not in mapeo.values():
            mapeo[col] = "descripcion"
        elif ("cliente" in cn and "proveedor" in cn) or cn in {"codigo", "cod proveedor"}:
            mapeo[col] = "codigo_prov"

    work = df.rename(columns=mapeo).copy()
    for col in ("fecha", "proveedor", "importe", "comprobante", "tipo", "debe", "haber", "cuenta", "descripcion", "codigo_prov", "cuit"):
        if col not in work.columns:
            work[col] = None

    work["fecha"] = pd.to_datetime(work["fecha"], errors="coerce").dt.date
    work["debe"] = pd.to_numeric(work["debe"], errors="coerce").fillna(0.0)
    work["haber"] = pd.to_numeric(work["haber"], errors="coerce").fillna(0.0)
    # Facturas / ND: saldo a pagar = Haber
    work = work[work["haber"] > 0.009].copy()
    work["importe"] = work["haber"].round(2)
    work["proveedor"] = work["proveedor"].astype(str).fillna("").str.strip()
    work["proveedor_norm"] = work["proveedor"].map(_normalizar_texto)
    work["comprobante"] = work["comprobante"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    work["tipo"] = work["tipo"].apply(lambda x: "" if pd.isna(x) else str(x).strip())
    work = work[work["proveedor_norm"].str.len() > 1].copy()
    work = work.reset_index(drop=True)
    work["factura_id"] = work.index.astype(int)
    return work[
        [
            "factura_id", "fecha", "proveedor", "proveedor_norm", "importe",
            "comprobante", "tipo", "cuenta", "descripcion", "codigo_prov", "cuit",
            "debe", "haber",
        ]
    ]


def cargar_debitos_desde_extracto_df(df_extracto: pd.DataFrame) -> pd.DataFrame:
    """Filtra débitos del DF de extracto Santander (o compatible)."""
    if df_extracto is None or df_extracto.empty:
        return pd.DataFrame()
    work = df_extracto.copy()
    # Normalizar nombres de columnas frecuentes
    ren: dict[str, str] = {}
    for col in work.columns:
        cn = _normalizar_texto(str(col))
        if cn in {"debito", "debito $", "debitos"} and "Debito" not in ren.values():
            ren[col] = "Debito"
        elif cn in {"credito", "credito $", "creditos"} and "Credito" not in ren.values():
            ren[col] = "Credito"
        elif cn == "fecha" and "Fecha" not in ren.values():
            ren[col] = "Fecha"
        elif cn in {"descripcion", "movimiento", "concepto"} and "Descripcion" not in work.columns:
            if "Descripcion" not in ren.values():
                ren[col] = "Descripcion"
        elif "concepto unificado" in cn:
            ren[col] = "Concepto unificado"
        elif cn == "comprobante":
            ren[col] = "Comprobante"
        elif "tipo fila" in cn:
            ren[col] = "Tipo fila"
    work = work.rename(columns=ren)
    # Formato universal: débitos = Importe negativo si no hay columna Debito
    if "Debito" not in work.columns and "Importe" in work.columns:
        work["Importe"] = pd.to_numeric(work["Importe"], errors="coerce")
        work["Debito"] = work["Importe"].where(work["Importe"] < 0, 0).abs()
        work["Credito"] = work["Importe"].where(work["Importe"] > 0, 0)
    if "Debito" not in work.columns:
        return pd.DataFrame()
    if "Tipo Movimiento" in work.columns:
        # Asegurar coherencia con signo
        mask_deb = work["Tipo Movimiento"].astype(str).str.lower().eq("debito")
        if "Importe" in work.columns:
            work.loc[mask_deb & (pd.to_numeric(work["Importe"], errors="coerce") < 0), "Debito"] = (
                pd.to_numeric(work.loc[mask_deb, "Importe"], errors="coerce").abs()
            )
    if "Tipo fila" in work.columns:
        work = work[work["Tipo fila"].astype(str).str.lower().ne("saldo inicial")]
    work["Debito"] = pd.to_numeric(work["Debito"], errors="coerce").fillna(0.0)
    work = work[work["Debito"] > 0.009].copy()
    if work.empty:
        return work

    fechas = []
    for v in work["Fecha"]:
        if isinstance(v, date) and not isinstance(v, datetime):
            fechas.append(v)
        elif isinstance(v, datetime):
            fechas.append(v.date())
        elif pd.isna(v):
            fechas.append(None)
        else:
            fechas.append(_parse_fecha_extracto(str(v)) or _parsear_fecha(str(v)))
    work["fecha"] = fechas
    work["importe"] = work["Debito"].round(2)
    if "Descripcion" not in work.columns:
        work["Descripcion"] = work.get("Concepto unificado", "")
    work["descripcion"] = work["Descripcion"].astype(str).fillna("")
    # Galicia / otros: el beneficiario suele ir en Detalle (TRF INMED PROVEED + nombre)
    if "Detalle" in work.columns:
        det = work["Detalle"].fillna("").astype(str).replace("nan", "")
        work["descripcion"] = (
            work["descripcion"].str.strip()
            + " "
            + det.str.replace("\n", " ", regex=False).str.strip()
        ).str.strip()
    work["descripcion_norm"] = work["descripcion"].map(_normalizar_texto)
    work["comprobante"] = work["Comprobante"].astype(str) if "Comprobante" in work.columns else ""
    work = work.reset_index(drop=True)
    work["debito_id"] = work.index.astype(int)
    cols = ["debito_id", "fecha", "importe", "descripcion", "descripcion_norm", "comprobante"]
    for extra in ("Concepto unificado", "Archivo origen", "Pagina PDF", "Mes", "Detalle"):
        if extra in work.columns:
            cols.append(extra)
    return work[cols]


def es_debito_pago_o_transferencia(descripcion: str) -> bool:
    """
    True solo para pagos/transferencias que pueden calzar con facturas de proveedores.
    Excluye comisiones, impuestos, retenciones, haberes, extracciones, etc.
    """
    low = _normalizar_texto(descripcion or "")
    if not low or low in {"sin descripcion", "sin descripcion"}:
        return False

    exclusiones = (
        "impuesto",
        "iva 21",
        "iva21",
        "reg de transfisc",
        "transfisc",
        "comision",
        "comisi ",
        "retencion",
        "arba",
        "imp al debito",
        "iibb",
        "libb",
        "percepcion",
        "sircreb",
        "mantenimiento",
        "haberes",
        "sueldo",
        "jornales",
        "retiro de efectivo",
        "compra con tarjeta",
        "extracc",
        "ley 25.413",
        "ley 25413",
        "25.413",
        "25413",
        "debito 0,6",
        "credito 0,6",
        "debito extr efvo",
        "chequeras",
        "segun su uso",
        "saldo total",
        "saldo inicial",
    )
    if any(x in low for x in exclusiones):
        return False

    inclusiones = (
        "transferencia realizada",
        "transferencia inmediata",
        "debito transf",
        "transf. online",
        "transf online",
        "online banking",
        "pago a proveedor",
        "pago proveedores",
        "pago de servicios",
        "pago con transferencia",
        "trf inmed",
        "transf inmed",
        "trf. inmed",
        "trf inmed proveed",
        "inmed proveed",
    )
    if any(x in low for x in inclusiones):
        return True
    # Genéricos de pago/transf (ya pasaron el filtro de exclusión)
    if "transferencia" in low or "transf" in low or low.startswith("trf ") or " trf " in f" {low} ":
        return True
    if low.startswith("pago ") or " pago a " in low or " pago de " in low:
        return True
    return False


def filtrar_debitos_pagos_transferencias(
    df_debitos: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Separa débitos matcheables (pagos/transf) vs excluidos (imp/com/etc)."""
    if df_debitos is None or df_debitos.empty:
        vacio = pd.DataFrame()
        return vacio, vacio
    mask = df_debitos["descripcion"].astype(str).map(es_debito_pago_o_transferencia)
    pagos = df_debitos[mask].copy().reset_index(drop=True)
    excluidos = df_debitos[~mask].copy().reset_index(drop=True)
    if not pagos.empty:
        pagos["debito_id"] = pagos.index.astype(int)
    if not excluidos.empty:
        excluidos["debito_id"] = excluidos.index.astype(int)
    return pagos, excluidos


def _nombre_score_debito_proveedor(desc_norm: str, prov_norm: str) -> float:
    """Similitud 0–100 entre descripción bancaria y razón social del proveedor."""
    if not desc_norm or not prov_norm:
        return 0.0
    if prov_norm in desc_norm or desc_norm in prov_norm:
        return 95.0

    # Normalizar puntos (H.TRUJILLO → h trujillo) y genéricos bancarios
    def _toks(s: str) -> list[str]:
        s = re.sub(r"[^\w\s]", " ", s.lower())
        stop = {
            "trf", "inmed", "proveed", "varios", "banco", "galicia", "credicoop",
            "santander", "bbva", "macro", "rio", "coop", "sa", "srl", "sas", "soc",
            "de", "del", "la", "las", "los", "y", "e", "var", "fac", "factura",
        }
        return [p for p in s.split() if len(p) > 2 and p not in stop]

    d_toks = _toks(desc_norm)
    p_toks = _toks(prov_norm)
    if not p_toks:
        return float(fuzz.token_set_ratio(prov_norm, desc_norm))

    nombre_score = float(fuzz.token_set_ratio(prov_norm, desc_norm))

    def _hit(pt: str) -> bool:
        if pt in desc_norm.replace(".", " "):
            return True
        for dt in d_toks:
            if dt == pt:
                return True
            # Truncado Galicia: CONTADORE ↔ CONTADORES / TRUJILL ↔ TRUJILLO
            if len(pt) >= 6 and len(dt) >= 5 and (pt.startswith(dt) or dt.startswith(pt)):
                return True
            if len(pt) >= 8 and len(dt) >= 6 and pt[:6] == dt[:6]:
                return True
        return False

    hits = sum(1 for p in p_toks if _hit(p))
    nombre_score = max(nombre_score, 100.0 * hits / len(p_toks))

    # Token fuerte (≥6) compartido (ej. "trujillo") → piso alto
    fuertes = [p for p in p_toks if len(p) >= 6 and _hit(p)]
    if fuertes:
        nombre_score = max(nombre_score, 78.0 if len(fuertes) == 1 else 88.0)

    return min(100.0, float(nombre_score))


def _dias_retardo_pago(fecha_deb: date | None, fecha_fac: date | None) -> int | None:
    """Días entre factura y pago (positivo = pago posterior). Solo informativo."""
    if not fecha_deb or not fecha_fac:
        return None
    return (fecha_deb - fecha_fac).days


def _score_match_debito_factura(
    fecha_deb: date | None,
    desc_norm: str,
    monto_deb: float,
    fecha_fac: date | None,
    prov_norm: str,
    monto_fac: float,
    *,
    modo: str = "nombre",
    ignorar_fecha: bool | None = None,
) -> tuple[float, str]:
    """
    Retorna (score 0-100+, motivo).

    modos:
      - nombre: razón social en transf + monto; fecha ±3 bonus (si fecha lejos → 0, va a retardo)
      - monto_fecha: banco sin nombre útil; monto ≈ exacto y fechas ±3 días
      - retardo: nombre claro + monto, sin ventana de fecha

    Compat: ignorar_fecha=True equivale a modo=\"retardo\".
    """
    if ignorar_fecha is True:
        modo = "retardo"
    elif ignorar_fecha is False and modo == "nombre":
        pass  # default

    dif = abs(float(monto_deb) - float(monto_fac))
    tol = max(TOL_MATCH_MONTO_PESOS, abs(monto_fac) * TOL_MATCH_MONTO_PCT)
    dias = None
    if fecha_deb and fecha_fac:
        dias = abs((fecha_deb - fecha_fac).days)

    nombre_score = _nombre_score_debito_proveedor(desc_norm, prov_norm)

    if dif <= 0.05:
        s_monto = 100.0
    elif dif <= tol:
        s_monto = 90.0 - 25.0 * (dif / tol)
    else:
        s_monto = max(0.0, 70.0 - 40.0 * (dif / max(abs(monto_fac), 1.0)))

    # --- Retardo: nombre + monto, sin fecha ---
    if modo == "retardo":
        if dif > tol:
            return 0.0, "monto_lejos"
        if nombre_score < MIN_FUZZY_NOMBRE_RETARDO:
            return 0.0, "sin_nombre"
        score = 0.60 * s_monto + 0.40 * nombre_score
        if dif <= 0.05:
            score += 8.0
        if nombre_score >= 90:
            score += 5.0
        motivo = "exacto" if dif <= 0.05 else "tolerancia"
        motivo += "+nombre+retardo_nombre"
        return round(score, 2), motivo

    # --- Solo monto + fecha (banco sin data de proveedor) ---
    if modo == "monto_fecha":
        sin_prov = _es_descripcion_bancaria_sin_proveedor(desc_norm)
        # Si el extracto trae beneficiario, NUNCA calzar solo por monto
        # (evita Selva→Scarpello u otros empates de importe/fecha).
        if not sin_prov:
            return 0.0, "tiene_nombre"
        if nombre_score >= MAX_NOMBRE_PARA_SOLO_MONTO:
            return 0.0, "tiene_nombre"
        if dif > tol:
            return 0.0, "monto_lejos"
        if dias is None or dias > TOL_MATCH_DIAS_SOLO_MONTO:
            return 0.0, "fecha_lejos"
        score = 0.75 * s_monto + 0.25 * max(0.0, 100.0 - dias * 5)
        if dif <= 0.05:
            score += 10.0
        score += 3.0  # sin_prov
        score += max(0.0, 5.0 - float(dias) * 0.5)
        motivo = "exacto" if dif <= 0.05 else "tolerancia"
        motivo += "+monto+fecha"
        return round(score, 2), motivo

    # --- Monto único 1:1 con ventana ampliada (sin nombre en banco) ---
    if modo == "monto_unico":
        sin_prov = _es_descripcion_bancaria_sin_proveedor(desc_norm)
        if not sin_prov:
            return 0.0, "tiene_nombre"
        if nombre_score >= MAX_NOMBRE_PARA_SOLO_MONTO:
            return 0.0, "tiene_nombre"
        if dif > 0.05:  # solo exacto: la unicidad es el criterio
            return 0.0, "monto_lejos"
        if dias is None or dias > TOL_MATCH_DIAS_MONTO_UNICO:
            return 0.0, "fecha_lejos"
        score = 85.0 + max(0.0, 10.0 - float(dias))
        return round(score, 2), "exacto+monto+unico"

    # --- Nombre (prioridad): nombre + monto; fecha cercana (±3) ---
    if dif > tol:
        return 0.0, "monto_lejos"
    if nombre_score < MIN_FUZZY_NOMBRE_MATCH:
        return 0.0, "sin_nombre"
    # Fecha lejos → no calza acá; el pase retardo lo recupera
    if dias is not None and dias > TOL_MATCH_DIAS:
        return 0.0, "fecha_lejos"

    if dias is None:
        s_fecha = 50.0
    else:
        s_fecha = 100.0 - dias * 8

    score = 0.45 * s_monto + 0.40 * nombre_score + 0.15 * s_fecha
    if dif <= 0.05:
        score += 8.0
    if nombre_score >= 90 and dias is not None and dias <= TOL_MATCH_DIAS:
        score += 10.0

    motivo = "exacto" if dif <= 0.05 else "tolerancia"
    motivo += "+nombre"
    if dias is not None and dias <= TOL_MATCH_DIAS:
        motivo += "+fecha"
    return round(score, 2), motivo


def _recolectar_candidatos_monto_unico(
    debitos: pd.DataFrame,
    facturas: pd.DataFrame,
    *,
    excluir_d: set[int] | None = None,
    excluir_f: set[int] | None = None,
) -> list[tuple[float, int, int, str, float]]:
    """1 débito ↔ 1 factura con el mismo importe exacto y sin otros candidatos libres."""
    excluir_d = excluir_d or set()
    excluir_f = excluir_f or set()
    libres_d = debitos[~debitos["debito_id"].isin(excluir_d)] if len(debitos) else debitos
    libres_f = facturas[~facturas["factura_id"].isin(excluir_f)] if len(facturas) else facturas
    if libres_d is None or libres_d.empty or libres_f is None or libres_f.empty:
        return []

    from collections import defaultdict

    by_monto_d: dict[float, list[int]] = defaultdict(list)
    by_monto_f: dict[float, list[int]] = defaultdict(list)
    map_d = {int(r["debito_id"]): r for _, r in libres_d.iterrows()}
    map_f = {int(r["factura_id"]): r for _, r in libres_f.iterrows()}
    for did, r in map_d.items():
        by_monto_d[round(float(r.get("importe") or 0), 2)].append(did)
    for fid, r in map_f.items():
        by_monto_f[round(float(r.get("importe") or 0), 2)].append(fid)

    candidatos: list[tuple[float, int, int, str, float]] = []
    for mon, dids in by_monto_d.items():
        if len(dids) != 1:
            continue
        fids = by_monto_f.get(mon) or []
        if len(fids) != 1:
            continue
        did, fid = dids[0], fids[0]
        d, f = map_d[did], map_f[fid]
        score, motivo = _score_match_debito_factura(
            d.get("fecha"),
            str(d.get("descripcion_norm") or ""),
            float(d.get("importe") or 0),
            f.get("fecha"),
            str(f.get("proveedor_norm") or ""),
            float(f.get("importe") or 0),
            modo="monto_unico",
        )
        if score >= 70.0:
            candidatos.append((score, did, fid, motivo, 0.0))
    candidatos.sort(key=lambda x: (-x[0], x[1], x[2]))
    return candidatos


def _recolectar_candidatos_match(
    debitos: pd.DataFrame,
    facturas: pd.DataFrame,
    *,
    score_minimo: float,
    modo: str = "nombre",
    ignorar_fecha: bool | None = None,
    excluir_d: set[int] | None = None,
    excluir_f: set[int] | None = None,
) -> list[tuple[float, int, int, str, float]]:
    """Candidatos (score, debito_id, factura_id, motivo, dif_monto) sobre pools libres."""
    if ignorar_fecha is True:
        modo = "retardo"
    excluir_d = excluir_d or set()
    excluir_f = excluir_f or set()
    candidatos: list[tuple[float, int, int, str, float]] = []
    for _, d in debitos.iterrows():
        did = int(d["debito_id"])
        if did in excluir_d:
            continue
        for _, f in facturas.iterrows():
            fid = int(f["factura_id"])
            if fid in excluir_f:
                continue
            score, motivo = _score_match_debito_factura(
                d.get("fecha"),
                str(d.get("descripcion_norm") or ""),
                float(d.get("importe") or 0),
                f.get("fecha"),
                str(f.get("proveedor_norm") or ""),
                float(f.get("importe") or 0),
                modo=modo,
            )
            if score >= score_minimo:
                dif = abs(float(d["importe"]) - float(f["importe"]))
                candidatos.append((score, did, fid, motivo, dif))
    candidatos.sort(key=lambda x: (-x[0], x[4], x[1], x[2]))
    return candidatos


def _aplicar_candidatos_match(
    candidatos: list[tuple[float, int, int, str, float]],
    debitos: pd.DataFrame,
    facturas: pd.DataFrame,
    usados_d: set[int],
    usados_f: set[int],
    calces: list[dict],
) -> int:
    """Asigna 1:1 candidatos libres. Devuelve cuántos calces nuevos."""
    nuevos = 0
    for score, did, fid, motivo, dif in candidatos:
        if did in usados_d or fid in usados_f:
            continue
        usados_d.add(did)
        usados_f.add(fid)
        d = debitos.loc[debitos["debito_id"] == did].iloc[0]
        f = facturas.loc[facturas["factura_id"] == fid].iloc[0]
        calces.append({
            "Fecha banco": d.get("fecha"),
            "Descripcion banco": d.get("descripcion"),
            "Debito banco": float(d.get("importe") or 0),
            "Comprobante banco": d.get("comprobante") or "",
            "Fecha factura": f.get("fecha"),
            "Proveedor": f.get("proveedor"),
            "Importe factura": float(f.get("importe") or 0),
            "Comprobante factura": f.get("comprobante") or "",
            "Tipo factura": f.get("tipo") or "",
            "Dif monto": round(dif, 2),
            "Dias retardo": _dias_retardo_pago(d.get("fecha"), f.get("fecha")),
            "Score": score,
            "Criterio": motivo,
        })
        nuevos += 1
    return nuevos


def _tol_monto_match(monto: float) -> float:
    return max(TOL_MATCH_MONTO_PESOS, abs(float(monto)) * TOL_MATCH_MONTO_PCT)


def _filas_por_id(df: pd.DataFrame, id_col: str) -> dict[int, pd.Series]:
    if df is None or df.empty:
        return {}
    return {int(r[id_col]): r for _, r in df.iterrows()}


def _recolectar_candidatos_suma(
    debitos: pd.DataFrame,
    facturas: pd.DataFrame,
    usados_d: set[int],
    usados_f: set[int],
) -> list[tuple]:
    """
    Candidatos de suma parcial (solo pares):
      - 2 transferencias → 1 factura
      - 1 transferencia → 2 facturas
    Con nombre (fuzzy) o, si el banco no trae beneficiario, por monto + fechas cercanas.
    Cada ítem: (score, tipo, ids_d, ids_f, dif, detalle)
    """
    map_d = _filas_por_id(debitos, "debito_id")
    map_f = _filas_por_id(facturas, "factura_id")
    libres_d = [did for did in map_d if did not in usados_d]
    libres_f = [fid for fid in map_f if fid not in usados_f]
    if len(libres_d) < 1 or len(libres_f) < 1:
        return []

    candidatos: list[tuple] = []
    ventana_sin_nom = TOL_MATCH_DIAS_SUMA_SIN_NOMBRE

    # --- 2 transferencias → 1 factura ---
    for fid in libres_f:
        f = map_f[fid]
        target = float(f.get("importe") or 0)
        if target <= 0.009:
            continue
        tol = _tol_monto_match(target)
        prov_norm = str(f.get("proveedor_norm") or "")
        fecha_f = f.get("fecha")
        pool: list[tuple[float, int, bool]] = []
        for did in libres_d:
            d = map_d[did]
            monto_d = float(d.get("importe") or 0)
            if monto_d <= 0.009 or monto_d > target + tol:
                continue
            desc_n = str(d.get("descripcion_norm") or "")
            ns = _nombre_score_debito_proveedor(desc_n, prov_norm)
            sin_prov = _es_descripcion_bancaria_sin_proveedor(desc_n)
            if ns >= MIN_FUZZY_NOMBRE_SUMA:
                pool.append((ns, did, False))
            elif sin_prov:
                dias = _dias_abs_fechas(d.get("fecha"), fecha_f)
                if dias is not None and dias <= ventana_sin_nom:
                    pool.append((max(ns, 40.0), did, True))
        if len(pool) < 2:
            continue
        pool.sort(key=lambda x: (-x[0], x[1]))
        pool = pool[:MAX_POOL_SUMA]
        for i in range(len(pool)):
            for j in range(i + 1, len(pool)):
                ns_i, did_i, sin_i = pool[i]
                ns_j, did_j, sin_j = pool[j]
                s = float(map_d[did_i]["importe"]) + float(map_d[did_j]["importe"])
                dif = abs(s - target)
                if dif > tol:
                    continue
                ns = min(ns_i, ns_j)
                score = 70.0 + (10.0 if dif <= 0.05 else 0.0) + 0.2 * ns
                if sin_i and sin_j:
                    score -= 5.0  # un poco menos que con nombre
                detalle = (
                    f"{float(map_d[did_i]['importe']):.2f}+{float(map_d[did_j]['importe']):.2f}"
                    f"={s:.2f}≈{target:.2f}"
                )
                candidatos.append((score, "2transf_1fct", (did_i, did_j), (fid,), dif, detalle, ns))

    # --- 1 transferencia → 2 facturas ---
    por_prov: dict[str, list[int]] = {}
    for fid in libres_f:
        pn = str(map_f[fid].get("proveedor_norm") or "")
        if not pn:
            continue
        por_prov.setdefault(pn, []).append(fid)

    for did in libres_d:
        d = map_d[did]
        target = float(d.get("importe") or 0)
        if target <= 0.009:
            continue
        tol = _tol_monto_match(target)
        desc_norm = str(d.get("descripcion_norm") or "")
        sin_prov = _es_descripcion_bancaria_sin_proveedor(desc_norm)
        fecha_d = d.get("fecha")
        for prov_norm, fids in por_prov.items():
            if len(fids) < 2:
                continue
            ns = _nombre_score_debito_proveedor(desc_norm, prov_norm)
            if ns < MIN_FUZZY_NOMBRE_SUMA and not sin_prov:
                continue
            pool_f: list[int] = []
            for fid in fids:
                monto_f = float(map_f[fid].get("importe") or 0)
                if not (0.009 < monto_f <= target + tol):
                    continue
                if sin_prov and ns < MIN_FUZZY_NOMBRE_SUMA:
                    dias = _dias_abs_fechas(fecha_d, map_f[fid].get("fecha"))
                    if dias is None or dias > ventana_sin_nom:
                        continue
                pool_f.append(fid)
            if len(pool_f) < 2:
                continue
            pool_f.sort(key=lambda fid: abs(float(map_f[fid]["importe"]) - target / 2))
            pool_f = pool_f[:MAX_POOL_SUMA]
            for i in range(len(pool_f)):
                for j in range(i + 1, len(pool_f)):
                    fid_i, fid_j = pool_f[i], pool_f[j]
                    s = float(map_f[fid_i]["importe"]) + float(map_f[fid_j]["importe"])
                    dif = abs(s - target)
                    if dif > tol:
                        continue
                    score = 70.0 + (10.0 if dif <= 0.05 else 0.0) + 0.2 * max(ns, 40.0 if sin_prov else 0.0)
                    if sin_prov and ns < MIN_FUZZY_NOMBRE_SUMA:
                        score -= 5.0
                    detalle = (
                        f"{float(map_f[fid_i]['importe']):.2f}+{float(map_f[fid_j]['importe']):.2f}"
                        f"={s:.2f}≈{target:.2f}"
                    )
                    candidatos.append(
                        (score, "1transf_2fct", (did,), (fid_i, fid_j), dif, detalle, ns)
                    )

    candidatos.sort(key=lambda x: (-x[0], x[4], x[1], x[2], x[3]))
    return candidatos


def _aplicar_candidatos_suma(
    candidatos: list[tuple],
    debitos: pd.DataFrame,
    facturas: pd.DataFrame,
    usados_d: set[int],
    usados_f: set[int],
    calces: list[dict],
    *,
    grupo_desde: int = 1,
) -> tuple[int, float]:
    """
    Aplica sumas parciales 1:N / N:1. Una fila consolidada por grupo.
    Retorna (cantidad de grupos, importe débitos calzados).
    """
    map_d = _filas_por_id(debitos, "debito_id")
    map_f = _filas_por_id(facturas, "factura_id")
    n_grupos = 0
    importe_debitos = 0.0
    g = grupo_desde

    for score, tipo, ids_d, ids_f, dif, detalle, _ns in candidatos:
        if any(did in usados_d for did in ids_d):
            continue
        if any(fid in usados_f for fid in ids_f):
            continue
        for did in ids_d:
            usados_d.add(did)
        for fid in ids_f:
            usados_f.add(fid)

        rows_d = [map_d[did] for did in ids_d]
        rows_f = [map_f[fid] for fid in ids_f]
        suma_d = round(sum(float(r.get("importe") or 0) for r in rows_d), 2)
        suma_f = round(sum(float(r.get("importe") or 0) for r in rows_f), 2)

        fechas_d = [r.get("fecha") for r in rows_d if r.get("fecha")]
        fechas_f = [r.get("fecha") for r in rows_f if r.get("fecha")]
        fecha_banco = fechas_d[0] if len(fechas_d) == 1 else " | ".join(
            (f.strftime("%d/%m/%Y") if hasattr(f, "strftime") else str(f)) for f in fechas_d
        )
        fecha_fac = fechas_f[0] if len(fechas_f) == 1 else " | ".join(
            (f.strftime("%d/%m/%Y") if hasattr(f, "strftime") else str(f)) for f in fechas_f
        )
        dias = None
        if len(fechas_d) == 1 and len(fechas_f) == 1:
            dias = _dias_retardo_pago(fechas_d[0], fechas_f[0])
        elif fechas_d and fechas_f:
            # retardo desde la factura más vieja al pago más nuevo
            try:
                dias = (max(fechas_d) - min(fechas_f)).days
            except Exception:
                dias = None

        motivo = f"pase3_{tipo}"
        if dif <= 0.05:
            motivo = "exacto+nombre+" + motivo
        else:
            motivo = "tolerancia+nombre+" + motivo

        calces.append({
            "Fecha banco": fecha_banco,
            "Descripcion banco": " || ".join(str(r.get("descripcion") or "") for r in rows_d),
            "Debito banco": suma_d,
            "Comprobante banco": " + ".join(str(r.get("comprobante") or "") for r in rows_d if r.get("comprobante")),
            "Fecha factura": fecha_fac,
            "Proveedor": rows_f[0].get("proveedor"),
            "Importe factura": suma_f,
            "Comprobante factura": " + ".join(str(r.get("comprobante") or "") for r in rows_f),
            "Tipo factura": " + ".join(
                sorted({str(r.get("tipo") or "") for r in rows_f if r.get("tipo")})
            ),
            "Dif monto": round(dif, 2),
            "Dias retardo": dias,
            "Score": round(score, 2),
            "Criterio": motivo,
            "Grupo suma": f"S{g}",
            "Detalle suma": detalle,
        })
        n_grupos += 1
        importe_debitos += suma_d
        g += 1

    return n_grupos, round(importe_debitos, 2)


def matchear_debitos_con_facturas(
    df_debitos: pd.DataFrame,
    df_facturas: pd.DataFrame,
    *,
    score_minimo: float = MIN_SCORE_MATCH_PROVEEDOR,
    score_minimo_monto_fecha: float = MIN_SCORE_MATCH_MONTO_FECHA,
    score_minimo_retardo: float = MIN_SCORE_MATCH_RETARDO,
    priorizar_fecha_monto: bool = False,
) -> dict[str, pd.DataFrame]:
    """
    Empareja débitos bancarios (solo pagos/transferencias) con facturas.

    Por defecto:
    1) Nombre + monto (fecha ±3 días).
    2) Sin nombre útil: monto ≈ exacto + fecha ±3 días.
    3) Retardo: nombre claro + monto (sin ventana de fecha).
    4) Sumas: 2 transferencias ↔ 1 factura, o 1 ↔ 2 (mismo proveedor).

    Con ``priorizar_fecha_monto=True``: primero monto+fecha (±3) y monto único (±15),
    después nombre+monto, retardo y sumas.
    """
    debitos_all = df_debitos.copy() if df_debitos is not None else pd.DataFrame()
    facturas = df_facturas.copy() if df_facturas is not None else pd.DataFrame()
    if debitos_all.empty:
        debitos_all = pd.DataFrame(columns=["debito_id", "fecha", "importe", "descripcion", "descripcion_norm"])
    if facturas.empty:
        facturas = pd.DataFrame(columns=["factura_id", "fecha", "proveedor", "proveedor_norm", "importe", "comprobante"])

    debitos, excluidos = filtrar_debitos_pagos_transferencias(debitos_all)
    if debitos.empty:
        debitos = pd.DataFrame(columns=["debito_id", "fecha", "importe", "descripcion", "descripcion_norm", "comprobante"])

    usados_d: set[int] = set()
    usados_f: set[int] = set()
    calces: list[dict] = []

    def _pase_nombre() -> int:
        cand = _recolectar_candidatos_match(
            debitos, facturas, score_minimo=score_minimo, modo="nombre",
            excluir_d=usados_d, excluir_f=usados_f,
        )
        return _aplicar_candidatos_match(cand, debitos, facturas, usados_d, usados_f, calces)

    def _pase_monto_fecha() -> int:
        cand = _recolectar_candidatos_match(
            debitos,
            facturas,
            score_minimo=score_minimo_monto_fecha,
            modo="monto_fecha",
            excluir_d=usados_d,
            excluir_f=usados_f,
        )
        n = _aplicar_candidatos_match(cand, debitos, facturas, usados_d, usados_f, calces)
        cand_u = _recolectar_candidatos_monto_unico(
            debitos, facturas, excluir_d=usados_d, excluir_f=usados_f,
        )
        return n + _aplicar_candidatos_match(cand_u, debitos, facturas, usados_d, usados_f, calces)

    def _pase_retardo() -> int:
        cand = _recolectar_candidatos_match(
            debitos,
            facturas,
            score_minimo=score_minimo_retardo,
            modo="retardo",
            excluir_d=usados_d,
            excluir_f=usados_f,
        )
        return _aplicar_candidatos_match(cand, debitos, facturas, usados_d, usados_f, calces)

    if priorizar_fecha_monto:
        n_pase2 = _pase_monto_fecha()
        n_pase1 = _pase_nombre()
        n_pase3 = _pase_retardo()
    else:
        n_pase1 = _pase_nombre()
        n_pase2 = _pase_monto_fecha()
        n_pase3 = _pase_retardo()

    # Pase 4 — sumas parciales (2↔1)
    cand4 = _recolectar_candidatos_suma(debitos, facturas, usados_d, usados_f)
    n_pase4, imp_pase4 = _aplicar_candidatos_suma(
        cand4, debitos, facturas, usados_d, usados_f, calces, grupo_desde=1,
    )

    df_calzados = pd.DataFrame(calces)
    df_sin_fact = debitos[~debitos["debito_id"].isin(usados_d)].copy()
    if not df_sin_fact.empty:
        df_sin_fact = df_sin_fact.rename(columns={
            "fecha": "Fecha banco",
            "importe": "Debito banco",
            "descripcion": "Descripcion banco",
            "comprobante": "Comprobante banco",
        })
        keep = [c for c in (
            "Fecha banco", "Descripcion banco", "Debito banco", "Comprobante banco",
            "Concepto unificado", "Archivo origen",
        ) if c in df_sin_fact.columns]
        df_sin_fact = df_sin_fact[keep].reset_index(drop=True)

    df_impagas = facturas[~facturas["factura_id"].isin(usados_f)].copy()
    if not df_impagas.empty:
        df_impagas = df_impagas.rename(columns={
            "fecha": "Fecha factura",
            "proveedor": "Proveedor",
            "importe": "Importe factura",
            "comprobante": "Comprobante factura",
            "tipo": "Tipo factura",
        })
        keep_f = [c for c in (
            "Fecha factura", "Proveedor", "Importe factura", "Comprobante factura",
            "Tipo factura", "cuenta", "codigo_prov",
        ) if c in df_impagas.columns]
        df_impagas = df_impagas[keep_f].reset_index(drop=True)

    df_excluidos = pd.DataFrame()
    if excluidos is not None and not excluidos.empty:
        df_excluidos = excluidos.rename(columns={
            "fecha": "Fecha banco",
            "importe": "Debito banco",
            "descripcion": "Descripcion banco",
            "comprobante": "Comprobante banco",
        })
        keep_e = [c for c in (
            "Fecha banco", "Descripcion banco", "Debito banco", "Comprobante banco",
            "Concepto unificado", "Archivo origen",
        ) if c in df_excluidos.columns]
        df_excluidos = df_excluidos[keep_e].reset_index(drop=True)

    def _imp_criterio(substr: str) -> float:
        return round(
            float(
                sum(
                    float(c["Debito banco"])
                    for c in calces
                    if substr in str(c.get("Criterio") or "")
                )
            ),
            2,
        )

    imp_pase2 = round(_imp_criterio("monto+fecha") + _imp_criterio("monto+unico"), 2)
    imp_pase3 = _imp_criterio("retardo_nombre")
    imp_pase1 = round(
        float(
            sum(
                float(c["Debito banco"])
                for c in calces
                if "monto+fecha" not in str(c.get("Criterio") or "")
                and "monto+unico" not in str(c.get("Criterio") or "")
                and "retardo_nombre" not in str(c.get("Criterio") or "")
                and "pase3_" not in str(c.get("Criterio") or "")
            )
        ),
        2,
    )
    resumen = pd.DataFrame([
        {"Concepto": "Débitos totales extracto", "Cantidad": len(debitos_all), "Importe": round(float(debitos_all["importe"].sum()) if len(debitos_all) else 0, 2)},
        {"Concepto": "Excluidos (imp/com/haberes/etc)", "Cantidad": len(df_excluidos), "Importe": round(float(df_excluidos["Debito banco"].sum()) if len(df_excluidos) and "Debito banco" in df_excluidos.columns else 0, 2)},
        {"Concepto": "Pagos/transferencias a matchear", "Cantidad": len(debitos), "Importe": round(float(debitos["importe"].sum()) if len(debitos) else 0, 2)},
        {"Concepto": "Facturas proveedores", "Cantidad": len(facturas), "Importe": round(float(facturas["importe"].sum()) if len(facturas) else 0, 2)},
        {"Concepto": "Calzados pase 1 (nombre + monto)", "Cantidad": n_pase1, "Importe": imp_pase1},
        {"Concepto": f"Calzados pase 2 (monto ±{TOL_MATCH_DIAS_SOLO_MONTO}d / único ±{TOL_MATCH_DIAS_MONTO_UNICO}d)", "Cantidad": n_pase2, "Importe": imp_pase2},
        {"Concepto": "Calzados pase 3 (retardo, sin fecha)", "Cantidad": n_pase3, "Importe": imp_pase3},
        {"Concepto": "Calzados pase 4 (sumas 2↔1)", "Cantidad": n_pase4, "Importe": imp_pase4},
        {"Concepto": "Calzados total", "Cantidad": len(df_calzados), "Importe": round(float(df_calzados["Debito banco"].sum()) if len(df_calzados) else 0, 2)},
        {"Concepto": "Pagos sin factura", "Cantidad": len(df_sin_fact), "Importe": round(float(df_sin_fact["Debito banco"].sum()) if len(df_sin_fact) and "Debito banco" in df_sin_fact.columns else 0, 2)},
        {"Concepto": "Facturas impagas", "Cantidad": len(df_impagas), "Importe": round(float(df_impagas["Importe factura"].sum()) if len(df_impagas) and "Importe factura" in df_impagas.columns else 0, 2)},
    ])
    return {
        "calzados": df_calzados,
        "pagos_sin_factura": df_sin_fact if not df_sin_fact.empty else pd.DataFrame(),
        "facturas_impagas": df_impagas if not df_impagas.empty else pd.DataFrame(),
        "excluidos": df_excluidos if not df_excluidos.empty else pd.DataFrame(),
        "resumen": resumen,
    }


def filtrar_facturas_match_proveedores(df_facturas: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Quita del pool a matchear cargos del propio banco (FCC Galicia en cta cte, etc.).
    Retorna (facturas_a_matchear, facturas_excluidas).
    """
    if df_facturas is None or df_facturas.empty:
        vacio = pd.DataFrame()
        return vacio, vacio
    work = df_facturas.copy()
    prov = work["proveedor"].astype(str) if "proveedor" in work.columns else pd.Series([""] * len(work))
    mask = prov.str.contains(
        r"BANCO\s+DE\s+GALICIA|BANCO\s+GALICIA|BANCO\s+SANTANDER|BANCO\s+MACRO",
        case=False,
        na=False,
    )
    # Solo excluir si además parece movimiento de cuenta bancaria (no un proveedor real raro)
    if "cuenta" in work.columns:
        cta = work["cuenta"].astype(str).str.lower()
        mask = mask & cta.str.contains(r"banco|1110|cta\.?\s*cte|cuenta corriente", na=False)
    excluidas = work[mask].copy()
    ok = work[~mask].copy().reset_index(drop=True)
    if "factura_id" in ok.columns:
        ok["factura_id"] = ok.index.astype(int)
    return ok, excluidas


def ejecutar_match_debitos_proveedores(
    df_extracto: pd.DataFrame,
    fuente_proveedores,
    *,
    priorizar_fecha_monto: bool = False,
    excluir_cargos_banco: bool = True,
) -> tuple[dict[str, pd.DataFrame], dict]:
    """
    Pipeline único chat + web Streamlit.

    Incluye: Detalle en descripción, ventanas ±7/±10 días, fuzzy truncado (Trujillo),
    monto+fecha solo sin beneficiario, sumas 2↔1, exclusión de FCC del propio banco.
    """
    if df_extracto is None or getattr(df_extracto, "empty", True):
        raise ValueError("El extracto no tiene movimientos.")

    df_ext = enriquecer_df_extracto_formato_banco(df_extracto)
    debitos = cargar_debitos_desde_extracto_df(df_ext)
    if debitos is None or debitos.empty:
        raise ValueError("El extracto no tiene débitos para matchear.")

    facturas = cargar_facturas_proveedores_excel(fuente_proveedores)
    if facturas is None or facturas.empty:
        raise ValueError("No se leyeron facturas del Excel de proveedores.")

    excluidas_banco = pd.DataFrame()
    if excluir_cargos_banco:
        facturas, excluidas_banco = filtrar_facturas_match_proveedores(facturas)
        if facturas.empty:
            raise ValueError("Tras excluir cargos bancarios no quedan facturas para matchear.")

    resultado = matchear_debitos_con_facturas(
        debitos,
        facturas,
        priorizar_fecha_monto=priorizar_fecha_monto,
    )
    meta_extra = {
        "n_debitos": len(debitos),
        "n_facturas": len(facturas),
        "n_excluidas_banco": len(excluidas_banco),
        "tol_dias_nombre": TOL_MATCH_DIAS,
        "tol_dias_monto": TOL_MATCH_DIAS_SOLO_MONTO,
        "pipeline": "ejecutar_match_debitos_proveedores_v1",
    }
    if excluidas_banco is not None and not excluidas_banco.empty:
        # Informar en resumen
        res_df = resultado.get("resumen")
        if res_df is not None and not res_df.empty:
            extra = pd.DataFrame([{
                "Concepto": "Facturas excluidas (cargos banco propio)",
                "Cantidad": len(excluidas_banco),
                "Importe": round(float(excluidas_banco["importe"].sum()), 2)
                if "importe" in excluidas_banco.columns else 0.0,
            }])
            resultado["resumen"] = pd.concat([res_df, extra], ignore_index=True)
        resultado["facturas_excluidas_banco"] = excluidas_banco
    return resultado, meta_extra


def exportar_match_proveedores_excel(resultado: dict[str, pd.DataFrame], meta: dict | None = None) -> bytes:
    """
    Excel del match: sin bordes, centrado, moneda ARS, color de pestaña por hoja.
    Hojas: Resumen / Calzados / Pagos_sin_factura / Facturas_impagas / Excluidos_imp_com.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.page import PageMargins

    meta = meta or {}
    wb = Workbook()

    # Sin líneas
    sin_borde = Border()
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(name="Calibri", bold=True, size=12, color="000000")
    body_font = Font(name="Calibri", size=11, color="000000")
    # Moneda pesos argentinos (punto miles / coma decimal en Excel es-AR)
    fmt_ars = '"$"#.##0,00'
    fmt_entero = "0"

    # Color de pestaña por hoja (RGB hex sin #)
    colores_pestana = {
        "Resumen": "5B9BD5",           # azul
        "Calzados": "70AD47",          # verde
        "Pagos_sin_factura": "ED7D31", # naranja
        "Facturas_impagas": "C00000",  # rojo
        "Excluidos_imp_com": "7030A0", # violeta
    }
    # Fondo suave de encabezado (tono acorde a la hoja)
    fills_header = {
        "Resumen": PatternFill("solid", fgColor="D6EAF8"),
        "Calzados": PatternFill("solid", fgColor="D5F5E3"),
        "Pagos_sin_factura": PatternFill("solid", fgColor="FDEBD0"),
        "Facturas_impagas": PatternFill("solid", fgColor="FADBD8"),
        "Excluidos_imp_com": PatternFill("solid", fgColor="E8DAEF"),
    }

    def _es_monto(nombre_col: str) -> bool:
        n = _normalizar_texto(nombre_col)
        return any(k in n for k in ("importe", "debito", "credito", "dif monto", "monto"))

    def _aplicar_estilo_celda(cell, *, header: bool, es_money: bool, fill_h: PatternFill):
        cell.font = header_font if header else body_font
        cell.border = sin_borde
        cell.alignment = centro
        if header:
            cell.fill = fill_h
        elif es_money and isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            cell.number_format = fmt_ars

    def _write_sheet(
        title: str,
        frame: pd.DataFrame,
        money_cols: tuple[int, ...] | None = None,
    ):
        ws = wb.create_sheet(title) if title != "Resumen" else wb.active
        ws.title = title
        if title in colores_pestana:
            ws.sheet_properties.tabColor = colores_pestana[title]
        fill_h = fills_header.get(title, PatternFill("solid", fgColor="EAECEE"))
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
        ws.sheet_view.showGridLines = False

        info = [
            ("Cliente", meta.get("cliente") or ""),
            ("CUIT", meta.get("cuit") or ""),
            ("Extracto", meta.get("origen_extracto") or ""),
            ("Facturas", meta.get("origen_facturas") or ""),
            ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
            (
                "Nota",
                "Solo pagos/transferencias. Pase 1 con fecha; pase 2 retardo (monto+proveedor); "
                "pase 3 sumas 2↔1 (2 transf↔1 fct o 1 transf↔2 fcts).",
            ),
        ]

        if title == "Resumen":
            # Bloque meta
            for c_i, txt in enumerate(("Campo", "Valor"), start=1):
                cell = ws.cell(1, c_i, txt)
                _aplicar_estilo_celda(cell, header=True, es_money=False, fill_h=fill_h)
            for i, (k, v) in enumerate(info, start=2):
                c1 = ws.cell(i, 1, k)
                c2 = ws.cell(i, 2, v)
                _aplicar_estilo_celda(c1, header=False, es_money=False, fill_h=fill_h)
                _aplicar_estilo_celda(c2, header=False, es_money=False, fill_h=fill_h)

            start = len(info) + 3
            rows = list(dataframe_to_rows(frame if frame is not None else pd.DataFrame(), index=False, header=True))
            headers = [str(h) for h in (rows[0] if rows else [])]
            for r_i, row in enumerate(rows):
                for c_i, val in enumerate(row, start=1):
                    cell = ws.cell(start + r_i, c_i, val)
                    col_name = headers[c_i - 1] if c_i - 1 < len(headers) else ""
                    es_money = (r_i > 0) and (
                        (money_cols is not None and c_i in money_cols) or _es_monto(col_name)
                    )
                    if r_i > 0 and c_i == 2 and isinstance(val, (int, float)):
                        # Cantidad: entero centrado
                        cell.number_format = fmt_entero
                        _aplicar_estilo_celda(cell, header=False, es_money=False, fill_h=fill_h)
                    else:
                        _aplicar_estilo_celda(cell, header=(r_i == 0), es_money=es_money, fill_h=fill_h)
            for col in range(1, 4):
                ws.column_dimensions[get_column_letter(col)].width = 36 if col == 1 else 22
            ws.freeze_panes = "A2"
            return

        rows = list(dataframe_to_rows(frame if frame is not None else pd.DataFrame(), index=False, header=True))
        if not rows:
            cell = ws.cell(1, 1, "Sin datos")
            _aplicar_estilo_celda(cell, header=False, es_money=False, fill_h=fill_h)
            return

        headers = [str(h) for h in rows[0]]
        for r_i, row in enumerate(rows):
            for c_i, val in enumerate(row, start=1):
                cell = ws.cell(r_i + 1, c_i, val)
                col_name = headers[c_i - 1] if c_i - 1 < len(headers) else ""
                es_money = (r_i > 0) and (
                    (money_cols is not None and c_i in money_cols) or _es_monto(col_name)
                )
                _aplicar_estilo_celda(cell, header=(r_i == 0), es_money=es_money, fill_h=fill_h)

        for c_i, h in enumerate(headers, start=1):
            ancho = min(42, max(14, len(h) + 4))
            if _es_monto(h):
                ancho = max(ancho, 16)
            ws.column_dimensions[get_column_letter(c_i)].width = ancho
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

    _write_sheet(
        "Resumen",
        resultado.get("resumen") if resultado.get("resumen") is not None else pd.DataFrame(),
        money_cols=(3,),
    )
    _write_sheet(
        "Calzados",
        resultado.get("calzados") if resultado.get("calzados") is not None else pd.DataFrame(),
        money_cols=(3, 7, 10),
    )
    _write_sheet(
        "Pagos_sin_factura",
        resultado.get("pagos_sin_factura") if resultado.get("pagos_sin_factura") is not None else pd.DataFrame(),
        money_cols=(3,),
    )
    _write_sheet(
        "Facturas_impagas",
        resultado.get("facturas_impagas") if resultado.get("facturas_impagas") is not None else pd.DataFrame(),
        money_cols=(3,),
    )
    _write_sheet(
        "Excluidos_imp_com",
        resultado.get("excluidos") if resultado.get("excluidos") is not None else pd.DataFrame(),
        money_cols=(3,),
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def movimientos_banco_a_dataframe_conciliacion(movimientos: list) -> pd.DataFrame:
    """Adapta MovimientoBanco al layout Fecha/Descripción/Crédito/Débito/Saldo de Galicia."""
    filas: list[list] = []
    for mov in movimientos:
        credito = f"{mov.credito:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if mov.credito else ""
        debito = f"{mov.debito:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if mov.debito else ""
        saldo = ""
        if mov.saldo is not None:
            saldo = f"{mov.saldo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        fecha_txt = mov.fecha.strftime("%d/%m/%Y") if hasattr(mov.fecha, "strftime") else str(mov.fecha)
        filas.append([
            fecha_txt,
            mov.descripcion,
            mov.comprobante or "",
            credito,
            debito,
            saldo,
        ])
    return pd.DataFrame(
        filas,
        columns=["Fecha", "Descripción", "Origen", "Crédito", "Débito", "Saldo"],
    )


def extraer_movimientos_banco(
    ruta_pdf: str | Path,
    banco: Optional[str] = None,
    lineas_precargadas: Optional[list[str]] = None,
) -> list[MovimientoBanco]:
    """Extrae movimientos de un extracto PDF aplicando reglas del banco detectado."""
    ruta = Path(ruta_pdf)
    banco_detectado = banco or detectar_banco_pdf(ruta)
    movimientos: list[MovimientoBanco] = []

    if lineas_precargadas is None and banco_detectado == "galicia":
        movs_tabla = extraer_movimientos_galicia_tabla(ruta, archivo=ruta.name)
        if movs_tabla:
            return _deduplicar_y_corregir_saldos(movs_tabla)

    if lineas_precargadas is not None:
        movimientos.extend(
            _extraer_movimientos_desde_texto(lineas_precargadas, 1, banco_detectado, ruta.name)
        )
        return _deduplicar_y_corregir_saldos(movimientos)

    documento = fitz.open(ruta)
    try:
        with pdfplumber.open(ruta) as pdf:
            for idx, pagina in enumerate(pdf.pages):
                texto = (pagina.extract_text() or "").strip()
                if texto:
                    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
                elif idx < len(documento):
                    texto_fitz = (documento[idx].get_text() or "").strip()
                    if texto_fitz:
                        lineas = [l.strip() for l in texto_fitz.splitlines() if l.strip()]
                    else:
                        lineas = [t for _, t in _ocr_pagina(documento[idx])]
                else:
                    lineas = []
                movimientos.extend(
                    _extraer_movimientos_desde_texto(lineas, idx + 1, banco_detectado, ruta.name)
                )
    finally:
        documento.close()

    return _deduplicar_y_corregir_saldos(movimientos)


def _deduplicar_y_corregir_saldos(movimientos: list[MovimientoBanco]) -> list[MovimientoBanco]:
    """
    Quita duplicados exactos (mismo archivo/fecha/importe/descripción/comprobante/saldo)
    y corrige D/C según la cadena de saldos. No trunca la descripción a 40 chars
    (eso omitía movimientos distintos del mismo día y monto).
    """
    vistos: set[tuple] = set()
    unicos: list[MovimientoBanco] = []
    for mov in sorted(movimientos, key=lambda m: (m.archivo_origen, m.pagina, m.fecha)):
        clave = (
            mov.fecha,
            round(mov.importe_absoluto, 2),
            _normalizar_texto(mov.descripcion),
            (mov.comprobante or "").strip(),
            round(mov.saldo, 2) if mov.saldo is not None else None,
            mov.banco,
            mov.archivo_origen,
            mov.pagina,
        )
        if clave not in vistos:
            vistos.add(clave)
            unicos.append(mov)

    saldo_anterior: Optional[float] = None
    archivo_ant: Optional[str] = None
    for mov in unicos:
        if archivo_ant is not None and mov.archivo_origen != archivo_ant:
            saldo_anterior = None
        archivo_ant = mov.archivo_origen
        if mov.saldo is not None and saldo_anterior is not None:
            delta = round(mov.saldo - saldo_anterior, 2)
            importe = mov.importe_absoluto
            if abs(abs(delta) - importe) <= 1.0:
                if delta > 0:
                    mov.credito, mov.debito = importe, 0.0
                else:
                    mov.debito, mov.credito = importe, 0.0
        if mov.saldo is not None:
            saldo_anterior = mov.saldo

    return unicos


def extraer_movimientos_anuales(
    rutas_pdf: list[str | Path],
) -> tuple[list[MovimientoBanco], list[str], dict[tuple[int, int], BalanceMensual]]:
    """
    Procesa hasta 12 PDFs, consolida cronológicamente y encadena saldos mensuales.
    El saldo final de un mes pasa a ser saldo inicial del mes siguiente.
    """
    if len(rutas_pdf) > MAX_PDFS_ANUALES:
        rutas_pdf = rutas_pdf[:MAX_PDFS_ANUALES]

    # Ordenar PDFs por fecha del primer movimiento detectado (o nombre)
    extractos: list[tuple[Path, list[MovimientoBanco], list[str], BalanceMensual]] = []

    for ruta in rutas_pdf:
        ruta = Path(ruta)
        banco = detectar_banco_pdf(ruta)
        lineas = _extraer_lineas_pdf(ruta)
        movs = extraer_movimientos_banco(ruta, banco=banco, lineas_precargadas=lineas)
        balance = _construir_balance_mensual(movs, lineas, ruta.name)
        extractos.append((ruta, movs, lineas, balance))

    extractos.sort(key=lambda x: (x[3].anio, x[3].mes))

    todos: list[MovimientoBanco] = []
    bancos: list[str] = []
    saldos_pdf: dict[tuple[int, int], BalanceMensual] = {}

    for _ruta, movs, _lineas, balance in extractos:
        clave = (balance.anio, balance.mes)
        saldos_pdf[clave] = balance
        todos.extend(movs)
        banco = detectar_banco_pdf(_ruta)
        if banco not in bancos:
            bancos.append(banco)

    todos.sort(key=lambda m: (m.fecha, m.archivo_origen, m.pagina))

    # Recalcular balances mensuales con encadenamiento SI → SF → SI mes siguiente
    saldos_por_mes = _calcular_balances_mensuales_encadenados(todos, saldos_pdf)

    return todos, bancos, saldos_por_mes


def cargar_compras_tango(ruta: str | Path | None = None) -> pd.DataFrame:
    """Carga el padrón de compras/proveedores de Tango con CUITs."""
    ruta_final = Path(ruta) if ruta else COMPRAS_TANGO_PATH
    if not ruta_final.exists():
        return pd.DataFrame(columns=["nombre", "cuit", "nombre_norm"])

    df = pd.read_excel(ruta_final)
    mapeo: dict[str, str] = {}
    for col in df.columns:
        cn = _normalizar_texto(str(col)).replace(" ", "_")
        if "nombre" in cn or "legal" in cn or "razon" in cn:
            mapeo[col] = "nombre"
        elif "cuit" in cn or "c__u__i__t" in cn:
            mapeo[col] = "cuit"

    df = df.rename(columns=mapeo)
    if "nombre" not in df.columns:
        df["nombre"] = df.iloc[:, 0].astype(str)
    if "cuit" not in df.columns:
        return pd.DataFrame(columns=["nombre", "cuit", "nombre_norm"])

    df["cuit"] = df["cuit"].astype(str).str.replace(r"\D", "", regex=True)
    df["nombre_norm"] = df["nombre"].astype(str).map(_normalizar_texto)
    return df[df["cuit"].str.len() == 11].copy()


def _clasificar_movimiento(
    mov: MovimientoBanco,
    compras: pd.DataFrame,
) -> MovimientoBanco:
    """
    Clasifica un movimiento bancario usando reglas contables e IA heurística.
    Segrega: Proveedores, Préstamos, Inversiones, Planes fiscales o Anomalías.
    """
    desc = _normalizar_texto(mov.descripcion)
    categoria = "transferencia"
    es_anomalia = False

    # 1) Cruce por CUIT explícito en descripción o campo detectado
    cuit = mov.cuit_contraparte
    if not cuit:
        match = re.search(r"\b(20|23|24|27|30|33|34)\d{9}\b", desc.replace("-", ""))
        if match:
            cuit = match.group(0)
            mov.cuit_contraparte = cuit

    if cuit and not compras.empty:
        fila = compras[compras["cuit"] == cuit]
        if not fila.empty:
            mov.cuit_contraparte = cuit
            categoria = "proveedor"

    # 2) Cruce fuzzy con nombres de proveedores en compras_tango
    if categoria == "transferencia" and not compras.empty:
        opciones = compras["nombre_norm"].tolist()
        match = process.extractOne(desc, opciones, scorer=fuzz.partial_ratio)
        if match and match[1] >= 72:
            idx = opciones.index(match[0])
            mov.cuit_contraparte = str(compras.iloc[idx]["cuit"])
            categoria = "proveedor"

    # 3) Reglas heurísticas por palabras clave (motor IA contable)
    reglas = [
        (("plazo fijo", "pf ", "inversion", "colocacion", "u va"), "inversion"),
        (("prestamo", "cuota prest", "capital prest", "interes prest", "hipotec"), "prestamo"),
        (("afip", "arba", "mis facilidades", "plan de pago", "plan de pagos", "vep", "fiscal"), "plan_fiscal"),
        (("iva", "iibb", "ganancias", "retencion", "impuesto", "25.413", "25413"), "impuesto"),
        (("haberes", "sueldo", "jornales", "pago haberes"), "haberes"),
        (("comision", "comis", "mantenimiento cuenta", "cargo bancario"), "comision"),
        (("proveedor", "factura", "pago a"), "proveedor"),
    ]

    if categoria == "transferencia":
        for palabras, cat in reglas:
            if any(p in desc for p in palabras):
                categoria = cat
                break

    # 4) Detección de anomalías
    if categoria == "transferencia" and mov.importe_absoluto > 500000:
        es_anomalia = True
        categoria = "anomalia"
    if "sin descripcion" in desc or desc.strip() == "":
        es_anomalia = True
        categoria = "anomalia"

    mov.categoria_contable = categoria
    mov.etiqueta = ETIQUETAS.get(categoria, ETIQUETAS["anomalia"])
    mov.es_anomalia = es_anomalia
    return mov


def clasificar_movimientos(
    movimientos: list[MovimientoBanco],
    compras: Optional[pd.DataFrame] = None,
) -> list[MovimientoBanco]:
    """Aplica clasificación contable a todos los movimientos."""
    compras_df = compras if compras is not None else cargar_compras_tango()
    return [_clasificar_movimiento(m, compras_df) for m in movimientos]


def _plan_flag_si(val) -> bool:
    """Interpreta flags Si/No del plan Tango (incluye numpy.bool_ y strings)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if s in ("", "N", "NO", "FALSE", "0"):
        return False
    return s in ("S", "SI", "YES", "TRUE", "1")


def _detectar_imputables_por_estructura(codigos: pd.Series) -> pd.Series:
    """Detecta cuentas imputables por estructura del código.

    Una cuenta es imputable si:
    1. Su código tiene >= 2 separadores (niveles) — ej: "1.1.01.01" tiene 3 puntos
    2. No aparece como prefijo de ningún otro código (es hoja, no madre)
    """
    codigos_list = codigos.tolist()
    imputables = []
    for cod in codigos_list:
        cod_str = str(cod).strip()
        separadores = len(re.findall(r"[.\-]", cod_str))
        if separadores < 2:
            imputables.append(False)
            continue
        tiene_hijos = any(
            str(otro).strip().startswith(cod_str + ".") or str(otro).strip().startswith(cod_str + "-")
            for otro in codigos_list
            if str(otro).strip() != cod_str
        )
        imputables.append(not tiene_hijos)
    return pd.Series(imputables, index=codigos.index)


def _detectar_imputables_plan_tango(df: pd.DataFrame) -> pd.Series:
    """Imputable = Habilitado en Tango y sin cuentas hijas por prefijo numérico."""
    codigos = df["codigo"].astype(str).str.strip().tolist()
    col_hab = next((c for c in df.columns if "habilitado" in str(c).lower()), None)
    hab_series = df[col_hab].map(_plan_flag_si) if col_hab else None
    imputables: list[bool] = []
    for i, cod in enumerate(codigos):
        if hab_series is not None and not bool(hab_series.iloc[i]):
            imputables.append(False)
            continue
        tiene_hijos = any(
            otro.startswith(cod) and len(otro) > len(cod)
            for otro in codigos
            if otro != cod
        )
        imputables.append(not tiene_hijos)
    return pd.Series(imputables, index=df.index)


def _normalizar_plan_cuentas_df(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza columnas codigo/descripcion/imputable en un plan de cuentas."""
    df = df.copy()
    df = df.loc[:, ~df.columns.duplicated()]
    df.columns = df.columns.astype(str).str.strip().str.lower()

    mapeo_columnas: dict[str, str] = {}
    for col in df.columns:
        col_norm = _normalizar_texto(str(col))
        if col_norm == "codigo" or (col_norm.startswith("codigo") and "alternativo" not in col_norm):
            mapeo_columnas[col] = "codigo"
        elif "descrip" in col_norm:
            mapeo_columnas[col] = "descripcion"
    df = df.rename(columns=mapeo_columnas)
    df = df.loc[:, ~df.columns.duplicated()]

    if "codigo" not in df.columns or "descripcion" not in df.columns:
        raise ValueError("El plan de cuentas debe tener columnas de código y descripción.")

    if isinstance(df["codigo"], pd.DataFrame):
        df["codigo"] = df["codigo"].iloc[:, 0]
    if isinstance(df["descripcion"], pd.DataFrame):
        df["descripcion"] = df["descripcion"].iloc[:, 0]

    df["codigo"] = df["codigo"].fillna("").astype(str).str.strip()
    df["descripcion"] = df["descripcion"].fillna("").astype(str).str.strip()
    df["descripcion_norm"] = df["descripcion"].map(_normalizar_texto)
    col_imp = next((c for c in df.columns if "imputable" in str(c).lower()), None)
    col_hab = next((c for c in df.columns if "habilitado" in str(c).lower()), None)
    if col_imp:
        imp_series = df[col_imp]
        if isinstance(imp_series, pd.DataFrame):
            imp_series = imp_series.iloc[:, 0]
        df["imputable"] = imp_series.map(_plan_flag_si)
    elif col_hab:
        df["imputable"] = _detectar_imputables_plan_tango(df)
    else:
        df["imputable"] = _detectar_imputables_por_estructura(df["codigo"].astype(str))
    col_aux = next(
        (c for c in df.columns if "auxiliar" in str(c).lower() and "usa" in str(c).lower()),
        None,
    )
    if col_aux:
        aux_series = df[col_aux]
        if isinstance(aux_series, pd.DataFrame):
            aux_series = aux_series.iloc[:, 0]
        df["usa_auxiliares"] = aux_series.map(_plan_flag_si)
    return df


def _normalizar_nombre_solapa_balance(nombre: str) -> str:
    """Normaliza nombre de solapa para comparación tolerante (espacios y mayúsculas)."""
    return re.sub(r"\s+", " ", str(nombre).strip().lower())


def _aliases_solapa_por_banco(nombre_banco: str) -> tuple[str, ...]:
    """Nombres de solapa aceptados según el banco seleccionado en pantalla."""
    try:
        base = solapas_banco(nombre_banco)
    except ValueError:
        clave = _normalizar_nombre_solapa_balance(nombre_banco)
        return (clave,) if clave else ()
    extras: set[str] = set(base)
    clave_banco = _normalizar_nombre_solapa_balance(nombre_banco)
    if clave_banco:
        extras.add(clave_banco)
        if clave_banco.startswith("banco "):
            extras.add(clave_banco.replace("banco", "", 1).strip())
            extras.add(f"bco {clave_banco.replace('banco', '', 1).strip()}")
            extras.add(f"cta cte {clave_banco.replace('banco', '', 1).strip()}")
        else:
            extras.add(f"banco {clave_banco}")
            extras.add(f"bco {clave_banco}")
            extras.add(f"cta cte {clave_banco}")
    return tuple(extras)


def _tokens_busqueda_solapa_banco(nombre_banco: str) -> set[str]:
    """Tokens flexibles para match parcial case-insensitive de solapas bancarias."""
    tokens: set[str] = set()
    for alias in _aliases_solapa_por_banco(nombre_banco):
        norm = _normalizar_nombre_solapa_balance(alias)
        if not norm:
            continue
        tokens.add(norm)
        for part in re.split(r"[\s\-_/\.]+", norm):
            part = part.strip()
            if len(part) >= 3 and part not in ("banco", "bco", "cta", "cte"):
                tokens.add(part)
        if norm.startswith("banco "):
            resto = norm[6:].strip()
            if resto:
                tokens.add(resto)
    return {t for t in tokens if t and len(t) >= 3}


def _puntaje_solapa_banco(hoja: str, tokens: set[str]) -> int:
    """Puntaje de coincidencia parcial entre solapa Excel y tokens del banco."""
    h_norm = _normalizar_nombre_solapa_balance(hoja)
    if not h_norm:
        return 0
    score = 0
    for token in tokens:
        if token in h_norm:
            score += len(token) + 10
        elif len(token) >= 4 and any(token in part for part in re.split(r"[\s\-_/\.]+", h_norm)):
            score += len(token)
    return score


def resolver_solapa_por_banco(source, nombre_banco: str) -> str:
    """
    Resuelve la solapa del balance cuyo nombre coincide con el banco activo.
    Match exacto, alias del registry y coincidencia parcial tolerante
    (ej: «Banco Galicia» → «galicia», «GALICIA», «bco galicia», «cta cte galicia»).
    """
    aliases = _aliases_solapa_por_banco(nombre_banco)
    if not aliases or not aliases[0]:
        raise ValueError("Nombre de banco vacío para mapear la solapa del balance.")
    hojas = listar_solapas_excel(source)
    alias_norms = {_normalizar_nombre_solapa_balance(a) for a in aliases if a}
    tokens = _tokens_busqueda_solapa_banco(nombre_banco)

    for hoja in hojas:
        h_norm = _normalizar_nombre_solapa_balance(hoja)
        if h_norm in alias_norms:
            return hoja

    for hoja in hojas:
        h_norm = _normalizar_nombre_solapa_balance(hoja)
        for alias in alias_norms:
            if alias and len(alias) >= 3 and (alias in h_norm or h_norm in alias):
                return hoja

    mejor_hoja = ""
    mejor_puntaje = 0
    for hoja in hojas:
        puntaje = _puntaje_solapa_banco(hoja, tokens)
        if puntaje > mejor_puntaje:
            mejor_puntaje = puntaje
            mejor_hoja = hoja
    if mejor_hoja and mejor_puntaje >= 4:
        return mejor_hoja

    disponibles = ", ".join(hojas) if hojas else "(ninguna)"
    raise ValueError(
        f"No se encontró la solapa para '{nombre_banco}' en el balance. "
        f"Solapas disponibles: {disponibles}"
    )


def resolver_solapa_balance(source, nombre: str) -> str:
    """Resuelve solapa por impuesto o banco según el nombre activo en pantalla."""
    clave = _normalizar_nombre_solapa_balance(nombre)
    for reg_key in BANK_REGISTRY:
        if _normalizar_nombre_solapa_balance(reg_key) == clave:
            return resolver_solapa_por_banco(source, reg_key)
        ficha = BANK_REGISTRY[reg_key]
        slug = ficha.get("slug", "")
        if slug and _normalizar_nombre_solapa_balance(slug) == clave:
            return resolver_solapa_por_banco(source, reg_key)
    return resolver_solapa_por_impuesto(source, nombre)


def _aliases_solapa_por_impuesto(nombre_impuesto: str) -> tuple[str, ...]:
    """Nombres de solapa aceptados según el impuesto seleccionado en pantalla."""
    try:
        return solapas_impuesto(nombre_impuesto)
    except ValueError:
        clave = _normalizar_nombre_solapa_balance(nombre_impuesto)
        return (clave,) if clave else ()


def _puntaje_solapa_convenio_multilateral(hoja: str) -> int:
    """Puntaje de coincidencia parcial para solapas de Convenio Multilateral (case-insensitive)."""
    h = _normalizar_nombre_solapa_balance(hoja)
    h_compact = h.replace(" ", "")
    if h in ("iibb", "ingresos brutos") and "convenio" not in h and "cm" not in h:
        return 0
    if h == "cm":
        return 100
    puntaje = 0
    tokens = set(h.split())
    if "cm" in tokens or h.startswith("cm ") or h.endswith(" cm") or " cm " in f" {h} ":
        puntaje += 25
    if "conveniomultilateral" in h_compact or h == "convenio multilateral":
        puntaje += 40
    elif "convenio" in h:
        puntaje += 30
    if "ingresosbrutoscm" in h_compact or ("ingresos brutos" in h and "cm" in h):
        puntaje += 35
    return puntaje


def resolver_solapa_por_impuesto(source, nombre_impuesto: str) -> str:
    """
    Resuelve la solapa del balance cuyo nombre coincide con el impuesto activo.
    Tolerante a mayúsculas/minúsculas y espacios extra ('iva', 'Iva', 'IVA ').
    Para Ingresos Brutos acepta solapas «IIBB» o «Ingresos Brutos».
    Para Convenio Multilateral acepta coincidencia parcial con CM, Convenio, etc.
    """
    aliases = _aliases_solapa_por_impuesto(nombre_impuesto)
    if not aliases or not aliases[0]:
        raise ValueError("Nombre de impuesto vacío para mapear la solapa del balance.")
    hojas = listar_solapas_excel(source)
    alias_norms = {_normalizar_nombre_solapa_balance(a) for a in aliases}
    for hoja in hojas:
        if _normalizar_nombre_solapa_balance(hoja) in alias_norms:
            return hoja

    clave_imp = _normalizar_nombre_solapa_balance(nombre_impuesto)
    if clave_imp == "convenio multilateral":
        mejor_hoja = ""
        mejor_puntaje = 0
        for hoja in hojas:
            puntaje = _puntaje_solapa_convenio_multilateral(hoja)
            if puntaje > mejor_puntaje:
                mejor_puntaje = puntaje
                mejor_hoja = hoja
        if mejor_hoja and mejor_puntaje >= 25:
            return mejor_hoja

    disponibles = ", ".join(hojas) if hojas else "(ninguna)"
    raise ValueError(
        f"No se encontró la solapa para '{nombre_impuesto}' en el balance. "
        f"Solapas disponibles: {disponibles}"
    )


def listar_solapas_excel(source) -> list[str]:
    """Lista hojas de un workbook Excel (path, BytesIO o bytes)."""
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    pos = 0
    if hasattr(source, "tell"):
        pos = source.tell()
    try:
        with pd.ExcelFile(source, engine="openpyxl") as xl:
            return list(xl.sheet_names)
    finally:
        if hasattr(source, "seek"):
            source.seek(pos)


def leer_dataframe_balance_solapa(
    source,
    nombre_impuesto: str,
    *,
    es_csv: bool = False,
    header=None,
) -> pd.DataFrame:
    """Lee la solapa del balance asociada al impuesto (sheet_name dinámico)."""
    if es_csv:
        if hasattr(source, "seek"):
            source.seek(0)
        return pd.read_csv(source, header=header)
    hoja = resolver_solapa_balance(source, nombre_impuesto)
    if hasattr(source, "seek"):
        source.seek(0)
    return pd.read_excel(source, sheet_name=hoja, header=header, engine="openpyxl")


def descargar_excel_balance_url(url: str, *, timeout: int = 60) -> io.BytesIO:
    """Descarga bytes crudos de un Excel Cloud preservando todas las solapas."""
    import urllib.error
    import urllib.request

    try:
        with urllib.request.urlopen(url, timeout=timeout) as resp:
            data = resp.read()
    except urllib.error.URLError as exc:
        raise ConnectionError(
            "No se pudo descargar el archivo desde OneDrive/SharePoint. "
            "Verifica que el link tenga permisos de lectura pública o el modificador ?download=1."
        ) from exc
    if not data:
        raise ValueError("El Excel descargado desde la nube está vacío.")
    buf = io.BytesIO(data)
    buf.name = "balance_cloud.xlsx"
    try:
        listar_solapas_excel(buf)
    except Exception as exc:
        raise ValueError("El archivo descargado no es un Excel válido.") from exc
    buf.seek(0)
    return buf


def cargar_plan_cuentas(ruta: str | Path | None = None) -> pd.DataFrame:
    ruta_final = Path(ruta) if ruta else PLAN_CUENTAS_DEFAULT
    if ruta_final.suffix.lower() == ".csv":
        df = pd.read_csv(ruta_final, dtype=str)
    else:
        df = pd.read_excel(ruta_final, sheet_name="Cuentas contables")
    return _normalizar_plan_cuentas_df(df)


def cargar_movimientos_contables(archivo) -> list[MovimientoContable]:
    nombre = getattr(archivo, "name", str(archivo)).lower()
    df = pd.read_csv(archivo) if nombre.endswith(".csv") else pd.read_excel(archivo)
    df.columns = [_normalizar_texto(str(c)).replace(" ", "_") for c in df.columns]
    col_fecha = next((c for c in df.columns if "fecha" in c), df.columns[0])
    col_importe = next((c for c in df.columns if "importe" in c or "monto" in c), None)
    col_desc = next((c for c in df.columns if "descrip" in c or "concepto" in c or "leyenda" in c), None)
    col_cuenta = next((c for c in df.columns if "cuenta" in c or "codigo" in c), None)

    movimientos: list[MovimientoContable] = []
    for _, fila in df.iterrows():
        fecha_val = fila[col_fecha]
        fecha: Optional[date] = None
        if isinstance(fecha_val, (datetime, pd.Timestamp)):
            if pd.isna(fecha_val):
                continue
            fecha = fecha_val.date()
        elif isinstance(fecha_val, date):
            fecha = fecha_val
        elif pd.notna(fecha_val):
            fecha = _parsear_fecha(str(fecha_val).strip())
        if fecha is None:
            continue
        movimientos.append(
            MovimientoContable(
                fecha=fecha,
                descripcion=str(fila[col_desc]) if col_desc else "",
                importe=float(fila[col_importe]) if col_importe else 0.0,
                cuenta=str(fila[col_cuenta]) if col_cuenta else "",
                codigo_cuenta=str(fila[col_cuenta]) if col_cuenta and str(fila[col_cuenta]).isdigit() else "",
            )
        )
    return movimientos


def _diferencia_dias(f1: date, f2: date) -> int:
    return abs((f1 - f2).days)


def _coincide_importe(a: float, b: float) -> bool:
    return abs(abs(a) - abs(b)) <= TOLERANCIA_IMPORTE


def _categoria_planilla(descripcion: str, categoria: str = "", es_credito: bool = False) -> str:
    """Devuelve el nombre de fila (columna B) de la plantilla manual."""
    desc = _normalizar_texto(descripcion)
    for clave, fila in MAPEO_CATEGORIAS_PLANILLA.items():
        if clave in desc:
            return fila
  # Categoría contable solo si no es genérica (transferencia por defecto del parser)
    if categoria and categoria not in ("transferencia", "") and categoria in MAPEO_CATEGORIAS_PLANILLA:
        return MAPEO_CATEGORIAS_PLANILLA[categoria]
    if es_credito:
        return "Transferencia Recibida misma cuenta"
    return "Trasnferencia de 3eros"


def _acumular_resumen(resultado: ResultadoConciliacion, mov: MovimientoBanco, importe: float) -> None:
    cat = _categoria_planilla(mov.descripcion, mov.categoria_contable, mov.credito > 0)
    mes_clave = (mov.fecha.year, mov.fecha.month)
    resultado.resumen_por_categoria[cat] = resultado.resumen_por_categoria.get(cat, 0) + importe
    if mes_clave not in resultado.resumen_anual_por_mes:
        resultado.resumen_anual_por_mes[mes_clave] = {}
    resultado.resumen_anual_por_mes[mes_clave][cat] = (
        resultado.resumen_anual_por_mes[mes_clave].get(cat, 0) + importe
    )
    if mov.es_anomalia:
        resultado.anomalias.append(
            {
                "fecha": mov.fecha.strftime("%d/%m/%Y"),
                "descripcion": mov.descripcion,
                "importe": importe,
                "etiqueta": mov.etiqueta,
                "banco": mov.banco,
            }
        )


def conciliar_movimientos(
    movimientos_banco: list[MovimientoBanco],
    movimientos_contables: list[MovimientoContable],
    tolerancia_dias: int = TOLERANCIA_DIAS_CLEARING,
    compras: Optional[pd.DataFrame] = None,
) -> ResultadoConciliacion:
    """Cruza extracto bancario con contabilidad y clasifica movimientos."""
    movimientos_banco = clasificar_movimientos(movimientos_banco, compras)
    resultado = ResultadoConciliacion()
    resultado.total_movimientos = len(movimientos_banco)
    resultado.movimientos_todos = list(movimientos_banco)
    resultado.bancos_detectados = list({m.banco for m in movimientos_banco})

    contables_pendientes = list(movimientos_contables)
    banco_pendientes = list(movimientos_banco)

    if movimientos_banco:
        depositos, retiros = _calcular_totales_movimientos(movimientos_banco)
        resultado.total_depositos = depositos
        resultado.total_retiros = retiros

        saldos_mov = [m.saldo for m in movimientos_banco if m.saldo is not None]
        if saldos_mov:
            resultado.saldo_extracto = saldos_mov[-1]
            resultado.saldo_final = saldos_mov[-1]

        conteo_meses: dict[tuple[int, int], int] = {}
        for m in movimientos_banco:
            clave = (m.fecha.year, m.fecha.month)
            conteo_meses[clave] = conteo_meses.get(clave, 0) + 1
        if conteo_meses:
            anio, mes = max(conteo_meses, key=conteo_meses.get)
            resultado.mes_referencia = date(anio, mes, 1)

    for mov_banco in list(banco_pendientes):
        mejor_match: Optional[MovimientoContable] = None
        mejor_puntaje = 0.0

        for mov_cont in contables_pendientes:
            if not _coincide_importe(mov_banco.importe_neto, mov_cont.importe):
                continue
            if _diferencia_dias(mov_banco.fecha, mov_cont.fecha) > tolerancia_dias:
                continue
            puntaje_desc = fuzz.token_set_ratio(
                _normalizar_texto(mov_banco.descripcion),
                _normalizar_texto(mov_cont.descripcion),
            )
            puntaje_fecha = max(0, 100 - _diferencia_dias(mov_banco.fecha, mov_cont.fecha) * 15)
            puntaje_total = puntaje_desc * 0.7 + puntaje_fecha * 0.3
            if puntaje_total > mejor_puntaje:
                mejor_puntaje = puntaje_total
                mejor_match = mov_cont

        if mejor_match and mejor_puntaje >= 55:
            resultado.conciliados.append(
                {
                    "fecha_banco": mov_banco.fecha,
                    "fecha_contable": mejor_match.fecha,
                    "descripcion_banco": mov_banco.descripcion,
                    "descripcion_contable": mejor_match.descripcion,
                    "importe": mov_banco.importe_neto,
                    "cuenta": mejor_match.codigo_cuenta or mejor_match.cuenta,
                    "puntaje": round(mejor_puntaje, 1),
                    "dias_diferencia": _diferencia_dias(mov_banco.fecha, mejor_match.fecha),
                    "categoria": mov_banco.categoria_contable,
                    "etiqueta": mov_banco.etiqueta,
                    "cuit": mov_banco.cuit_contraparte,
                    "banco": mov_banco.banco,
                }
            )
            _acumular_resumen(resultado, mov_banco, abs(mov_banco.importe_neto))
            banco_pendientes.remove(mov_banco)
            contables_pendientes.remove(mejor_match)

    resultado.solo_banco = banco_pendientes
    resultado.solo_contabilidad = contables_pendientes

    for mov in resultado.solo_banco:
        _acumular_resumen(resultado, mov, mov.importe_absoluto)

    return resultado


def aplicar_saldos_al_resultado(
    resultado: ResultadoConciliacion,
    saldos_por_mes: dict[tuple[int, int], BalanceMensual],
) -> ResultadoConciliacion:
    """Inyecta balances de control y validación al resultado de conciliación."""
    resultado.saldos_por_mes = saldos_por_mes

    if saldos_por_mes:
        meses_ordenados = sorted(saldos_por_mes.keys())
        primero = meses_ordenados[0]
        ultimo = meses_ordenados[-1]
        bal_ini = saldos_por_mes[primero]
        bal_fin = saldos_por_mes[ultimo]

        resultado.saldo_inicial = bal_ini.saldo_inicial
        resultado.saldo_final = bal_fin.saldo_final or bal_fin.saldo_resumen
        resultado.saldo_extracto = bal_fin.saldo_resumen or bal_fin.saldo_final
        resultado.total_depositos = sum(b.total_ingresos for b in saldos_por_mes.values())
        resultado.total_retiros = sum(b.total_egresos for b in saldos_por_mes.values())
        resultado.balance_cierra = all(b.balance_cierra for b in saldos_por_mes.values())
        resultado.diferencia_balance = sum(abs(b.diferencia_balance) for b in saldos_por_mes.values() if not b.balance_cierra)
        resultado.mes_referencia = date(ultimo[0], ultimo[1], 1)

        # Alerta si algún mes no cierra
        for clave, b in saldos_por_mes.items():
            if not b.balance_cierra:
                resultado.anomalias.append(
                    {
                        "fecha": f"{b.mes:02d}/{b.anio}",
                        "descripcion": f"Balance no cierra — {b.archivo_origen}",
                        "importe": b.diferencia_balance,
                        "etiqueta": "⚠️ Diferencia de balance",
                        "banco": "",
                    }
                )

    return resultado


def imputar_cuenta(descripcion: str, plan_cuentas: pd.DataFrame, categoria: str = "") -> str:
    if categoria and categoria in MAPEO_CUENTAS_KEYWORDS:
        return MAPEO_CUENTAS_KEYWORDS[categoria]
    desc = _normalizar_texto(descripcion)
    for clave, codigo in MAPEO_CUENTAS_KEYWORDS.items():
        if clave in desc:
            return codigo
    if not plan_cuentas.empty:
        opciones = plan_cuentas["descripcion_norm"].tolist()
        codigos = plan_cuentas["codigo"].tolist()
        match = process.extractOne(desc, opciones, scorer=fuzz.partial_ratio)
        if match and match[1] >= 65:
            return codigos[opciones.index(match[0])]
    return PERFILES_BANCO["santander"]["cuenta_contable"]


def _columna_mes_en_planilla(hoja, mes: date) -> Optional[str]:
    for col in range(3, hoja.max_column + 1):
        valor = hoja.cell(5, col).value
        if isinstance(valor, datetime) and valor.year == mes.year and valor.month == mes.month:
            return get_column_letter(col)
        if isinstance(valor, date) and valor.year == mes.year and valor.month == mes.month:
            return get_column_letter(col)
    return None


def _agrupar_importes_planilla_por_mes(
    movimientos: list[MovimientoBanco],
) -> dict[tuple[int, int], dict[str, float]]:
    """
    Agrupa TODOS los movimientos del extracto por mes y fila de categoría (columna B).
    Usa el nombre exacto de categoría que figura en la plantilla manual.
    """
    por_mes: dict[tuple[int, int], dict[str, float]] = {}
    for mov in movimientos:
        if not _fecha_plausible_extracto(mov.fecha):
            continue
        cat = _categoria_planilla(mov.descripcion, mov.categoria_contable, mov.credito > 0)
        clave = (mov.fecha.year, mov.fecha.month)
        por_mes.setdefault(clave, {})
        por_mes[clave][cat] = por_mes[clave].get(cat, 0.0) + mov.importe_absoluto
    return por_mes


def _inyectar_categorias_mes(
    hoja,
    col_mes: str,
    categorias: dict[str, float],
    filas_categoria: dict[str, int],
) -> None:
    """Inyecta importes por categoría en las filas de la plantilla (solo celdas de valor)."""
    ref_estilo = f"{col_mes}8"
    for categoria, importe in categorias.items():
        if importe <= 0:
            continue
        fila_destino = filas_categoria.get(_normalizar_texto(categoria))
        if not fila_destino:
            # Coincidencia parcial por si el nombre difiere levemente
            cat_norm = _normalizar_texto(categoria)
            for clave, fila in filas_categoria.items():
                if cat_norm in clave or clave in cat_norm:
                    fila_destino = fila
                    break
        if not fila_destino:
            continue
        celda_addr = f"{col_mes}{fila_destino}"
        celda = hoja[celda_addr]
        if _es_formula(celda.value):
            continue
        _escribir_celda_valor(hoja, celda_addr, round(importe, 2), ref_estilo, forzar=True)


def _inyectar_balance_mes(hoja, col_mes: str, balance: BalanceMensual, es_primer_mes: bool) -> None:
    """
    Inyecta saldos de control en las celdas exactas de la plantilla manual.
    Respeta las fórmulas cruzadas: Ingresos = fila 6, Retiros = fila 41+54−49, SF = 57+58−59.
    """
    f = FILAS_PLANTILLA
    ref_estilo = f"{col_mes}{f['saldo_inicio']}"

    # Fila 57 — Saldo al inicio (valor del PDF o encadenado desde SF del mes anterior)
    if es_primer_mes and balance.saldo_inicial is not None:
        _escribir_celda_valor(
            hoja, f"{col_mes}{f['saldo_inicio']}", round(balance.saldo_inicial, 2), ref_estilo, forzar=True
        )
    elif not es_primer_mes:
        col_ant = _columna_anterior(col_mes)
        if col_ant:
            _escribir_celda_formula(
                hoja,
                f"{col_mes}{f['saldo_inicio']}",
                f"=+{col_ant}{f['saldo_final']}",
                ref_estilo,
            )

    # Filas 58-60 — Mismas fórmulas que la planilla manual (no totales directos del PDF)
    _escribir_celda_formula(hoja, f"{col_mes}{f['ingresos']}", f"=+{col_mes}6", ref_estilo)
    _escribir_celda_formula(
        hoja, f"{col_mes}{f['retiros']}", f"=+{col_mes}41+{col_mes}54-{col_mes}49", ref_estilo
    )
    _escribir_celda_formula(
        hoja,
        f"{col_mes}{f['saldo_final']}",
        f"=+{col_mes}{f['saldo_inicio']}+{col_mes}{f['ingresos']}-{col_mes}{f['retiros']}",
        ref_estilo,
    )

    # Fila 61 — Saldo según resumen del extracto bancario
    if balance.saldo_resumen is not None:
        _escribir_celda_valor(
            hoja, f"{col_mes}{f['saldo_resumen']}", round(balance.saldo_resumen, 2), ref_estilo, forzar=True
        )

    # Fila 62 — Diferencia c/ resumen (fórmula de plantilla)
    _escribir_celda_formula(
        hoja, f"{col_mes}{f['diferencia']}", f"=+{col_mes}{f['saldo_resumen']}-{col_mes}{f['saldo_final']}", ref_estilo
    )


def _inyectar_detalle_movimientos(wb, movimientos: list[MovimientoBanco], hoja_ref) -> None:
    """
    Volca el detalle de movimientos en hoja 'Detalle Extracto' clonando estilos de la plantilla.
    Columnas: Fecha | Concepto | Débito | Crédito | Saldo | Banco
    """
    nombre = HOJA_DETALLE_MOVIMIENTOS
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)

    encabezados = ["Fecha", "Concepto", "Débito", "Crédito", "Saldo", "Banco", "Archivo"]
    for col_idx, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        _copiar_estilo_celda(hoja_ref["B5"], celda)

    estilo_fila = hoja_ref["B8"]
    for i, mov in enumerate(sorted(movimientos, key=lambda m: (m.fecha, m.pagina)), start=FILA_INICIO_DETALLE):
        fila = [
            mov.fecha,
            mov.descripcion,
            mov.debito if mov.debito else None,
            mov.credito if mov.credito else None,
            mov.saldo,
            PERFILES_BANCO.get(mov.banco, {}).get("nombre_display", mov.banco),
            mov.archivo_origen,
        ]
        for col_idx, val in enumerate(fila, start=1):
            celda = ws.cell(row=i, column=col_idx, value=val)
            _copiar_estilo_celda(estilo_fila, celda)
            if col_idx in (3, 4, 5) and val is not None:
                celda.number_format = "#,##0.00"
        if i > 500:
            break

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14


def generar_planilla_conciliacion(
    resultado: ResultadoConciliacion,
    nombre_cliente: str,
    ruta_plantilla: str | Path | None = None,
    mes_cierre_balance: Optional[int] = 12,
) -> bytes:
    """
    Clona la plantilla original con openpyxl e inyecta saldos de control y movimientos.
    Preserva fórmulas, fuentes (Calibri), colores y bordes existentes.
    """
    origen = Path(ruta_plantilla) if ruta_plantilla else _resolver_ruta_plantilla()
    destino_temp = BASE_DIR / "_temp_planilla.xlsx"
    shutil.copy2(origen, destino_temp)

    wb = openpyxl.load_workbook(destino_temp)
    hoja_nombre = PERFILES_BANCO.get(
        resultado.bancos_detectados[0] if resultado.bancos_detectados else "santander", {}
    ).get("hoja_planilla", HOJA_BANCO_DEFAULT)

    if hoja_nombre not in wb.sheetnames:
        hoja_nombre = HOJA_BANCO_DEFAULT
    hoja = wb[hoja_nombre]

    # Reordenar meses en la fila 5 según mes_cierre_balance
    mes_inicio = (mes_cierre_balance % 12) + 1
    if resultado.mes_referencia:
        ref_year = resultado.mes_referencia.year
        ref_month = resultado.mes_referencia.month
        if ref_month < mes_inicio:
            anio_inicio = ref_year - 1
        else:
            anio_inicio = ref_year
    else:
        anio_inicio = date.today().year

    for i in range(12):
        col = 3 + i
        m = (mes_inicio - 1 + i) % 12 + 1
        y = anio_inicio + ((mes_inicio - 1 + i) // 12)
        hoja.cell(row=5, column=col, value=datetime(y, m, 1))

    # Solo modificar celdas de datos — encabezado cliente
    hoja["B1"] = nombre_cliente
    if resultado.mes_referencia:
        hoja["B2"] = datetime(resultado.mes_referencia.year, resultado.mes_referencia.month, 1)

    filas_categoria: dict[str, int] = {}
    for fila in range(1, hoja.max_row + 1):
        valor_b = hoja.cell(fila, 2).value
        if isinstance(valor_b, str) and valor_b.strip() and not valor_b.startswith("="):
            filas_categoria[_normalizar_texto(valor_b)] = fila

    # Categorías por mes: TODOS los movimientos del extracto → filas de la plantilla
    movs_fuente = resultado.movimientos_todos or resultado.solo_banco
    meses_categorias = _agrupar_importes_planilla_por_mes(movs_fuente)
    if not meses_categorias and resultado.resumen_anual_por_mes:
        meses_categorias = resultado.resumen_anual_por_mes

    for (anio, mes), categorias in sorted(meses_categorias.items()):
        col_mes = _columna_mes_en_planilla(hoja, date(anio, mes, 1))
        if not col_mes:
            continue
        _inyectar_categorias_mes(hoja, col_mes, categorias, filas_categoria)

    # Saldos de control por mes (filas 57-61)
    saldos_meses = resultado.saldos_por_mes or {}
    if not saldos_meses and resultado.mes_referencia:
        saldos_meses = {
            (resultado.mes_referencia.year, resultado.mes_referencia.month): BalanceMensual(
                anio=resultado.mes_referencia.year,
                mes=resultado.mes_referencia.month,
                saldo_inicial=resultado.saldo_inicial,
                saldo_final=resultado.saldo_final,
                saldo_resumen=resultado.saldo_extracto,
                total_ingresos=resultado.total_depositos,
                total_egresos=resultado.total_retiros,
                balance_cierra=resultado.balance_cierra,
                diferencia_balance=resultado.diferencia_balance,
            )
        }

    meses_ordenados = sorted(saldos_meses.keys())
    for idx, clave in enumerate(meses_ordenados):
        anio, mes = clave
        col_mes = _columna_mes_en_planilla(hoja, date(anio, mes, 1))
        if col_mes:
            _inyectar_balance_mes(hoja, col_mes, saldos_meses[clave], es_primer_mes=(idx == 0))

    # Detalle de movimientos con estilos clonados
    movs = resultado.movimientos_todos or resultado.solo_banco
    if movs:
        _inyectar_detalle_movimientos(wb, movs, hoja)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    destino_temp.unlink(missing_ok=True)
    return buffer.getvalue()


def generar_txt_tango(
    resultado: ResultadoConciliacion,
    plan_cuentas: pd.DataFrame,
    tipo_asiento: str = "CONC",
    concepto_base: str = "Conciliación bancaria",
) -> str:
    lineas = [
        "[ASIENTOS_CONTABLES]",
        "IDENTIFICADOR\tFECHA DEL ASIENTO\tCLASE DEL ASIENTO\tTIPO DE ASIENTO\t"
        "ESTADO DEL ASIENTO\tMONEDA DEL ASIENTO\tCONCEPTO\tOBSERVACIONES",
    ]
    cabeceras: list[str] = []
    renglones: list[str] = []
    identificador = 1
    cuenta_banco = PERFILES_BANCO.get(
        resultado.bancos_detectados[0] if resultado.bancos_detectados else "santander",
        PERFILES_BANCO["santander"],
    )["cuenta_contable"]

    for mov in resultado.solo_banco:
        fecha_str = mov.fecha.strftime("%d/%m/%Y")
        id_str = str(identificador)
        concepto = f"{concepto_base} - [{mov.etiqueta}] {mov.descripcion[:50]}"
        cabeceras.append(
            f"{id_str}\t{fecha_str}\tBASICO\t{tipo_asiento}\tINGRESADO\t"
            f"CORRIENTE\t{concepto}\tClasificación: {mov.categoria_contable}"
        )
        cuenta = imputar_cuenta(mov.descripcion, plan_cuentas, mov.categoria_contable)
        importe = mov.importe_absoluto
        if mov.debito > 0:
            renglones.append(f"{id_str}\t{cuenta}\t{importe:.2f}\t\t{fecha_str}\t{mov.descripcion}")
            renglones.append(f"{id_str}\t{cuenta_banco}\t\t{importe:.2f}\t{fecha_str}\t{mov.descripcion}")
        else:
            renglones.append(f"{id_str}\t{cuenta_banco}\t{importe:.2f}\t\t{fecha_str}\t{mov.descripcion}")
            renglones.append(f"{id_str}\t{cuenta}\t\t{importe:.2f}\t{fecha_str}\t{mov.descripcion}")
        identificador += 1

    lineas.extend(cabeceras)
    lineas.append("[RENGLONES]")
    lineas.append("IDENTIFICADOR\tCODIGO DE CUENTA\tIMPORTE DEBE\tIMPORTE HABER\tFECHA ORIGEN\tLEYENDA")
    lineas.extend(renglones)
    return "\r\n".join(lineas) + "\r\n"


def movimientos_a_dataframe(movimientos: list[MovimientoBanco]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Fecha": m.fecha.strftime("%d/%m/%Y"),
                "Banco": PERFILES_BANCO.get(m.banco, {}).get("nombre_display", m.banco),
                "Comprobante": m.comprobante,
                "Descripción": m.descripcion,
                "Débito": m.debito if m.debito else None,
                "Crédito": m.credito if m.credito else None,
                "Saldo": m.saldo if m.saldo else None,
                "CUIT": m.cuit_contraparte,
                "Categoría": m.etiqueta,
                "Anomalía": "⚠️ Sí" if m.es_anomalia else "",
                "Archivo": m.archivo_origen,
            }
            for m in movimientos
        ]
    )


def resultado_a_dataframe(resultado: ResultadoConciliacion) -> pd.DataFrame:
    filas = []
    for item in resultado.conciliados:
        filas.append(
            {
                "Estado": "Conciliado",
                "Fecha Banco": item["fecha_banco"].strftime("%d/%m/%Y"),
                "Fecha Contable": item["fecha_contable"].strftime("%d/%m/%Y"),
                "Descripción": item["descripcion_banco"],
                "Importe": item["importe"],
                "Categoría": item.get("etiqueta", ""),
                "CUIT": item.get("cuit", ""),
                "Banco": item.get("banco", ""),
                "Cuenta": item["cuenta"],
                "Puntaje": item["puntaje"],
            }
        )
    for mov in resultado.solo_banco:
        filas.append(
            {
                "Estado": "Solo Banco",
                "Fecha Banco": mov.fecha.strftime("%d/%m/%Y"),
                "Fecha Contable": "",
                "Descripción": mov.descripcion,
                "Importe": mov.importe_neto,
                "Categoría": mov.etiqueta,
                "CUIT": mov.cuit_contraparte,
                "Banco": mov.banco,
                "Cuenta": imputar_cuenta(mov.descripcion, pd.DataFrame(), mov.categoria_contable),
                "Puntaje": "",
            }
        )
    for mov in resultado.solo_contabilidad:
        filas.append(
            {
                "Estado": "Solo Contabilidad",
                "Fecha Banco": "",
                "Fecha Contable": mov.fecha.strftime("%d/%m/%Y"),
                "Descripción": mov.descripcion,
                "Importe": mov.importe,
                "Categoría": "",
                "CUIT": "",
                "Banco": "",
                "Cuenta": mov.codigo_cuenta or mov.cuenta,
                "Puntaje": "",
            }
        )
    return pd.DataFrame(filas)


# --- Auditoría de Préstamos y Mayor Contable ---

BANCOS_ARGENTINOS: list[str] = [
    "Banco Galicia",
    "Banco Nación",
    "Banco Provincia",
    "Banco Macro",
    "Banco Santander",
    "BBVA",
    "ICBC",
    "Banco Credicoop",
    "HSBC",
    "Banco Supervielle",
    "Banco Ciudad",
    "Banco Comafi",
    "Mercado Pago",
]

MAPEO_BANCO_DISPLAY_A_CLAVE: dict[str, str] = {
    "banco galicia": "galicia",
    "banco nacion": "nacion",
    "banco provincia": "provincia",
    "banco macro": "macro",
    "banco santander": "santander",
    "bbva": "frances",
    "banco frances": "frances",
    "icbc": "icbc",
    "banco credicoop": "credicoop",
    "hsbc": "hsbc",
    "banco supervielle": "supervielle",
    "banco ciudad": "ciudad",
    "banco comafi": "comafi",
    "mercado pago": "mercadopago",
    "mercadopago": "mercadopago",
}

# Mapa de keywords para detección canónica de banco en PDFs de préstamos
BANCOS_KEYWORDS_PRESTAMO: dict[str, list[str]] = {
    "Banco Nación": ["nacion", "bna", "banco de la nacion", "banco nacion"],
    "Banco Santander": ["santander", "santander rio"],
    "BBVA": ["bbva", "banco frances", "frances"],
    "Banco Provincia": ["provincia", "bapro", "banco provincia"],
    "Mercado Pago": ["mercado pago", "mercadopago", "mp financiero"],
    "Banco Galicia": ["galicia"],
    "ICBC": ["icbc", "industrial and commercial"],
    "Banco HSBC": ["hsbc"],
    "Banco Macro": ["macro"],
    "Banco Credicoop": ["credicoop"],
    "Banco Supervielle": ["supervielle"],
    "Banco Ciudad": ["ciudad"],
    "Banco Comafi": ["comafi"],
}

_PATTERNS_PRESTAMO: list[re.Pattern] = [
    re.compile(r"[Pp]r[eé]stamo\s*[Nn](?:ro\.?|[°\xbao])?\s*:?\s*(\w+)"),
    re.compile(r"[Cc]r[eé]dito\s+[Nn](?:ro\.?|[°\xbao])?\s*:?\s*(\w+)"),
    re.compile(r"[Oo]peraci[o\xf3]n\s*[Nn](?:ro\.?|[°\xbao])?\s*:?\s*(\w+)"),
    re.compile(r"[Cc]ontrato\s*[Nn](?:ro\.?|[°\xbao])?\s*:?\s*(\w+)"),
    re.compile(r"[Nn]ro\.?\s+[Pp]r[eé]stamo\s*:?\s*(\w+)"),
    re.compile(r"[Nn]ro\.?\s+[Cc]r[eé]dito\s*:?\s*(\w+)"),
    re.compile(r"[Nn][°\xbao]\s+[Oo]peraci[o\xf3]n\s*:?\s*(\w+)"),
    re.compile(r"[Nn][°\xbao]\s*:?\s*(\d{5,})"),
]

# Prefijos de cuenta Tango para clasificación préstamos
PREFIJOS_CUENTA_CAPITAL_PRESTAMO = ("221", "22101")
PREFIJOS_CUENTA_INTERES_PRESTAMO = ("425", "42502", "42501")
PALABRAS_NUEVO_CREDITO = ("desembolso", "acreditacion prestamo", "credito otorgado", "alta prestamo")


@dataclass
class CuotaPrestamo:
    """Cuota extraída de proyección / tabla de amortización."""

    numero_cuota: int
    fecha_vencimiento: date
    capital: float
    intereses: float
    total: float
    impuestos: float = 0.0
    saldo_restante: float = 0.0
    banco: str = ""
    archivo_origen: str = ""
    prestamo_id: str = ""


@dataclass
class CruceCuotaExtracto:
    """Resultado del cruce cuota ↔ débito bancario."""

    cuota: CuotaPrestamo
    movimiento: Optional[MovimientoBanco] = None
    coincidencia: bool = False
    diferencia_importe: float = 0.0
    diferencia_dias: int = 0
    observacion: str = ""


@dataclass
class ValidacionMayor:
    """Validación de imputación contable vs cuota esperada."""

    cuota: CuotaPrestamo
    capital_mayor: float = 0.0
    interes_mayor: float = 0.0
    capital_esperado: float = 0.0
    interes_esperado: float = 0.0
    error_imputacion: bool = False
    detalle: str = ""


@dataclass
class SaldoPasivoBanco:
    """Constitución de saldo pasivo de préstamos por banco."""

    banco: str
    saldo_inicial_pasivo: float
    capital_pagado: float
    nuevos_creditos: float
    saldo_calculado: float
    saldo_final_mayor: Optional[float] = None
    diferencia: float = 0.0
    cierra: bool = True
    alerta: str = ""


MODOS_AUDITORIA_PRESTAMOS = ("teorico", "parcial_extractos", "completo", "solo_extractos")

MSG_PENDIENTE_EXTRACTOS = "Pendiente (Requiere Extractos)"
MSG_PENDIENTE_MAYOR = "Pendiente (Requiere Mayor)"

DESCRIPCION_MODO_AUDITORIA = {
    "teorico": "Estructura Teórica — planilla base por saldos iniciales (sin extractos ni mayor)",
    "parcial_extractos": "Proyección + extractos — validación mayor pendiente",
    "completo": "Triple validación completa (proyección → extractos → mayor)",
    "solo_extractos": "Solo extractos bancarios (sin proyección de cuotas)",
}


@dataclass
class ResultadoAuditoriaPrestamos:
    """Resultado integral de la auditoría préstamos + mayor."""

    cuotas: list[CuotaPrestamo] = field(default_factory=list)
    cruces: list[CruceCuotaExtracto] = field(default_factory=list)
    validaciones_mayor: list[ValidacionMayor] = field(default_factory=list)
    saldos_por_banco: list[SaldoPasivoBanco] = field(default_factory=list)
    saldos_iniciales: list[dict] = field(default_factory=list)
    movimientos_banco: list[MovimientoBanco] = field(default_factory=list)
    alertas: list[dict] = field(default_factory=list)
    bancos_procesados: list[str] = field(default_factory=list)
    nombre_cliente: str = ""
    modo: str = "completo"


# Relleno rojo estándar Excel para desbalances / imputaciones erróneas
_FILL_ALERTA_ROJO = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_FILL_ALERTA_VERDE = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")


def _clave_banco_desde_display(nombre: str) -> str:
    """Convierte nombre visible del banco a clave interna."""
    norm = _normalizar_texto(nombre)
    if norm in MAPEO_BANCO_DISPLAY_A_CLAVE:
        return MAPEO_BANCO_DISPLAY_A_CLAVE[norm]
    for display, clave in MAPEO_BANCO_DISPLAY_A_CLAVE.items():
        if display in norm or norm in display:
            return clave
    for clave, perfil in PERFILES_BANCO.items():
        if _normalizar_texto(perfil["nombre_display"]) in norm:
            return clave
    return norm.replace(" ", "_")[:20] or "santander"


def _nombre_display_banco(clave: str) -> str:
    return PERFILES_BANCO.get(clave, {}).get("nombre_display", clave.replace("_", " ").title())


def _extraer_cuotas_desde_texto(lineas: list[str], banco: str = "", archivo: str = "") -> list[CuotaPrestamo]:
    """
    Interpreta líneas de PDF de proyección de préstamos buscando filas de amortización.
    Soporta fechas con / y con -, números de cuota al inicio o precedidos de espacios,
    y hasta 6 columnas numéricas (cuota, capital, intereses, IVA/gastos, total, saldo).
    También detecta el formato de recibo/liquidación de Banco Nación (pago bullet único).
    """
    cuotas: list[CuotaPrestamo] = []

    # --- Detección de recibo de Banco Nación (liquidación con pago único) ---
    texto_completo = " ".join(lineas).lower()
    es_recibo_nacion = (
        "liquidacion" in texto_completo
        and ("recibo" in texto_completo or "importe neto" in texto_completo or "db. transactor" in texto_completo)
        and "banco nacion" in texto_completo or "bna" in texto_completo
    ) or ("liquidacion" in texto_completo and "db. transactor" in texto_completo)

    if es_recibo_nacion:
        capital = 0.0
        fecha_vto = None
        for linea in lineas:
            lower = linea.lower()
            # Capital: "Capital   24.800.000,00 080 - PESOS" o "24.800.000,00 Importe Neto"
            if "capital" in lower or "importe neto" in lower:
                montos = re.findall(r"[\d\.][\d\.]*,\d{2}", linea)
                if montos:
                    capital = max(_limpiar_monto(m) for m in montos)
            # Fecha 1r.Vto: "Fecha 1r.Vto. 08-09-2025" o "Fecha 1r.Vto 08-09-2025"
            if "1r.vto" in lower or "primer vto" in lower or "fecha vto" in lower:
                m = re.search(r"(\d{2}[-/]\d{2}[-/]\d{4})", linea)
                if m:
                    fecha_vto = _parsear_fecha(m.group(1))
        if capital > 0 and fecha_vto:
            cuotas.append(CuotaPrestamo(
                numero_cuota=1,
                fecha_vencimiento=fecha_vto,
                capital=round(capital, 2),
                intereses=0.0,
                impuestos=0.0,
                total=round(capital, 2),
                banco=banco,
                archivo_origen=archivo,
            ))
            return cuotas
    # --- Fin detección recibo Nación ---

    # --- Detección específica: Banco Galicia (OCR de imagen limpia) ---
    # Formato: "1  A Vencer  2025-03-28  $ 1558534.38  $ 1282777.77  $ 239459.58  $ 25143.26  $ 3591.89  $ 7561.88"
    # Columnas: Cuota | Estado | Vencimiento | Monto total | Capital | Interés nominal | IVA interés | IVA percep | Otros gastos
    banco_norm = _normalizar_texto(banco)
    if "galicia" in banco_norm or any("galicia" in _normalizar_texto(l) for l in lineas[:10]):
        patron_galicia = re.compile(r"(\d{1,3})\s+\w[\w\s]*\s+(\d{4}[-/]\d{2}[-/]\d{2}|\d{2}[-/]\d{2}[-/]\d{4})")
        for linea in lineas:
            m = patron_galicia.search(linea)
            if not m:
                continue
            nums_raw = re.findall(r"\$?\s*[\d\.][\d\.]*,\d{2}|\$?\s*[\d,]*\.\d{2}|\d{5,}", linea)
            importes = [_limpiar_monto(n.replace("$", "").strip()) for n in nums_raw]
            importes = [v for v in importes if v >= 100.0]
            if len(importes) < 2:
                continue
            numero = int(m.group(1))
            if numero <= 0 or numero > 600:
                continue
            fecha = _parsear_fecha(m.group(2))
            if not fecha:
                continue
            # Galicia: Monto total (idx 0), Capital (idx 1), Int nominal (idx 2), IVA int (idx 3), IVA perc (idx 4), Otros (idx 5)
            total = importes[0] if len(importes) > 0 else 0.0
            capital = importes[1] if len(importes) > 1 else 0.0
            intereses = importes[2] if len(importes) > 2 else 0.0
            iva_int = importes[3] if len(importes) > 3 else 0.0
            iva_perc = importes[4] if len(importes) > 4 else 0.0
            otros = importes[5] if len(importes) > 5 else 0.0
            impuestos_galicia = round(iva_int + iva_perc + otros, 2)
            if capital <= 0 and total <= 0:
                continue
            cuotas.append(CuotaPrestamo(
                numero_cuota=numero,
                fecha_vencimiento=fecha,
                capital=round(capital, 2),
                intereses=round(intereses, 2),
                impuestos=round(impuestos_galicia, 2),
                total=round(total, 2),
                banco=banco,
                archivo_origen=archivo,
            ))
        if cuotas:
            return cuotas

    # --- Detección específica: Banco Santander "Desarrollo del Préstamo" ---
    # Formato: "18/04/2025  0,00  3.006.617,31  3.006.617,31  65.439.196,00"
    # Columnas: Fecha Vto | Capital Cuota | Int Compens. Per. | Total Cuota | Saldo Deuda Capital
    if "santander" in banco_norm or any("santander" in _normalizar_texto(l) for l in lineas[:10]):
        patron_sant_header = re.compile(r"fecha.*vto|capital.*cuota|desarrollo.*prestamo", re.IGNORECASE)
        en_tabla_sant = False
        num_cuota_sant = 0
        for linea in lineas:
            if patron_sant_header.search(linea):
                en_tabla_sant = True
                continue
            if not en_tabla_sant:
                continue
            m_fecha = re.search(r"\b(\d{2}[/\-]\d{2}[/\-]\d{4})\b", linea)
            if not m_fecha:
                continue
            fecha = _parsear_fecha(m_fecha.group(1))
            if not fecha:
                continue
            nums_raw = re.findall(r"[\d\.][\d\.]*,\d{2}", linea)
            importes = [_limpiar_monto(n) for n in nums_raw]
            importes = [v for v in importes if v >= 0]
            if len(importes) < 2:
                continue
            num_cuota_sant += 1
            # Columnas: Capital Cuota | Int Compens. | Total Cuota | Saldo
            capital_s = importes[0] if len(importes) > 0 else 0.0
            intereses_s = importes[1] if len(importes) > 1 else 0.0
            total_s = importes[2] if len(importes) > 2 else round(capital_s + intereses_s, 2)
            saldo_s = importes[3] if len(importes) > 3 else 0.0
            cuotas.append(CuotaPrestamo(
                numero_cuota=num_cuota_sant,
                fecha_vencimiento=fecha,
                capital=round(capital_s, 2),
                intereses=round(intereses_s, 2),
                impuestos=0.0,
                saldo_restante=round(saldo_s, 2),
                total=round(total_s, 2),
                banco=banco,
                archivo_origen=archivo,
            ))
        if cuotas:
            return cuotas

    # Acepta dd/mm/yy, dd/mm/yyyy, dd-mm-yy, dd-mm-yyyy
    patron_fecha = re.compile(r"\b(\d{2}[/\-]\d{2}[/\-]\d{2,4})\b")
    # Número de cuota: 1-3 dígitos al inicio de línea (con o sin espacios) o tras separador común
    patron_num_cuota = re.compile(r"(?:^|\|\s*)(\d{1,3})(?:\s|/|\||\.)")
    en_tabla = False

    for linea in lineas:
        lower = _normalizar_texto(linea)
        if any(k in lower for k in (
            "cuota", "amortiz", "vencim", "capital", "interes",
            "monto", "importe", "tabla de", "fecha de", "saldo",
            "proyeccion", "cronograma", "vto", "cuotas", "pagos",
            "plan de pago", "tabla amort",
        )):
            en_tabla = True
            continue
        if not en_tabla:
            continue

        match_fecha = patron_fecha.search(linea)
        if not match_fecha:
            continue

        match_cuota = patron_num_cuota.search(linea)
        if not match_cuota:
            continue

        numero = int(match_cuota.group(1))
        if numero <= 0 or numero > 600:
            continue

        fecha = _parsear_fecha(match_fecha.group(1))
        if not fecha:
            continue

        # Extraer todos los valores numéricos de la línea (excluir el número de cuota ya capturado)
        # Buscar secuencias con separadores de miles punto y decimal coma, o viceversa
        nums_raw = re.findall(r"\d[\d\.]*,\d{2}|\d[\d,]*\.\d{2}|\d{3,}", linea)
        importes = [_limpiar_monto(n) for n in nums_raw]
        importes = [v for v in importes if v >= 1.0]
        if len(importes) < 2:
            continue

        # Asignación por posición según cantidad de columnas encontradas:
        # ≥6: cuota_n, capital, intereses, iva_gastos, total, saldo
        # ≥5: capital, intereses, iva_gastos, total, saldo
        # ≥4: capital, intereses, iva_gastos, total
        # ≥3: capital, intereses, total
        # 2:  capital, total
        impuestos = 0.0
        saldo_restante = 0.0
        if len(importes) >= 6:
            capital, intereses, impuestos, total, saldo_restante = (
                importes[-6], importes[-5], importes[-4], importes[-2], importes[-1]
            )
        elif len(importes) >= 5:
            capital, intereses, impuestos, total, saldo_restante = (
                importes[-5], importes[-4], importes[-3], importes[-2], importes[-1]
            )
        elif len(importes) >= 4:
            capital, intereses, impuestos, total = (
                importes[-4], importes[-3], importes[-2], importes[-1]
            )
        elif len(importes) >= 3:
            capital, intereses, total = importes[-3], importes[-2], importes[-1]
        elif len(importes) == 2:
            capital, intereses, total = importes[0], 0.0, importes[1]
        else:
            continue

        if total <= 0:
            total = round(capital + intereses + impuestos, 2)

        cuotas.append(
            CuotaPrestamo(
                numero_cuota=numero,
                fecha_vencimiento=fecha,
                capital=round(capital, 2),
                intereses=round(intereses, 2),
                impuestos=round(impuestos, 2),
                saldo_restante=round(saldo_restante, 2),
                total=round(total, 2),
                banco=banco,
                archivo_origen=archivo,
            )
        )

    return cuotas


def _extraer_cuotas_desde_tabla_pdf(ruta: Path) -> list[CuotaPrestamo]:
    """Extrae cuotas desde tablas estructuradas con pdfplumber (detección de columnas ampliada).

    Intenta dos estrategias de extracción de tablas: settings por defecto y luego con
    tolerancias relajadas para PDFs con espaciado irregular.
    """
    cuotas: list[CuotaPrestamo] = []
    banco = detectar_banco_pdf(ruta)

    # Keywords ampliadas para soportar abreviaciones de bancos argentinos
    _HEADER_KW = {
        "capital", "cuota", "amortiz", "vencim", "interes",
        "monto", "saldo", "importe", "total", "fecha",
        "vto", "cap", "mto", "cuot", "int", "seg", "gast", "cuotas",
        "fec", "pago", "debito", "debit",
    }

    _TABLE_SETTINGS_LIST = [
        {},  # estrategia por defecto
        {"snap_tolerance": 5, "join_tolerance": 5, "edge_min_length": 3},
        {"snap_tolerance": 10, "join_tolerance": 10, "edge_min_length": 3,
         "text_tolerance": 5},
    ]

    def _procesar_tabla(tabla, banco_key, archivo_nombre):
        resultados = []
        if not tabla or len(tabla) < 2:
            return resultados
        header = [_normalizar_texto(str(c or "")) for c in tabla[0]]
        kw_count = sum(1 for h in header if any(k in h for k in _HEADER_KW))
        if kw_count < 1:
            return resultados

        idx_cuota = next(
            (i for i, h in enumerate(header)
             if h in ("n", "#", "cuota", "n°", "nro", "numero", "cuotas", "n.cuota", "nro.")
             or ("cuota" in h and "total" not in h and "monto" not in h)),
            0,
        )
        idx_fecha = next(
            (i for i, h in enumerate(header)
             if "fecha" in h or "venc" in h or "vto" in h or "fec" in h), 1
        )
        idx_cap = next(
            (i for i, h in enumerate(header)
             if "capital" in h or "amortiz" in h or ("cap" == h) or "cap." in h),
            2,
        )
        idx_int = next(
            (i for i, h in enumerate(header)
             if "interes" in h or ("int" == h) or "int." in h), 3
        )
        idx_imp = next(
            (i for i, h in enumerate(header)
             if "iva" in h or "impuesto" in h or "gasto" in h
             or "seguro" in h or "seg" in h or "cargo" in h),
            -1,
        )
        idx_saldo = next(
            (i for i, h in enumerate(header) if "saldo" in h), -1
        )
        idx_tot = next(
            (i for i, h in enumerate(header)
             if "total" in h
             or ("monto" in h and "capital" not in h)
             or ("importe" in h and "capital" not in h)
             or "mto" == h or "mto." in h),
            min(4, len(header) - 1),
        )

        for fila in tabla[1:]:
            if not fila or len(fila) < 2:
                continue
            try:
                num_txt = str(fila[idx_cuota] or "").strip()
                num_match = re.search(r"\d+", num_txt)
                if not num_match:
                    continue
                numero = int(num_match.group())
                if numero <= 0 or numero > 600:
                    continue
                fecha_txt = str(fila[idx_fecha] if idx_fecha < len(fila) else "")
                fecha = _parsear_fecha(fecha_txt)
                if not fecha:
                    continue
                capital = _limpiar_monto(fila[idx_cap] if idx_cap < len(fila) else None)
                intereses = _limpiar_monto(fila[idx_int] if idx_int < len(fila) else None)
                impuestos = _limpiar_monto(
                    fila[idx_imp] if idx_imp >= 0 and idx_imp < len(fila) else None
                )
                saldo_restante = _limpiar_monto(
                    fila[idx_saldo] if idx_saldo >= 0 and idx_saldo < len(fila) else None
                )
                total_raw = _limpiar_monto(fila[idx_tot] if idx_tot < len(fila) else None)
                total = total_raw if total_raw > 0 else round(capital + intereses + impuestos, 2)
                if capital <= 0 and intereses <= 0 and total <= 0:
                    continue
                resultados.append(CuotaPrestamo(
                    numero_cuota=numero,
                    fecha_vencimiento=fecha,
                    capital=round(capital, 2),
                    intereses=round(intereses, 2),
                    impuestos=round(impuestos, 2),
                    saldo_restante=round(saldo_restante, 2) if saldo_restante > 0 else 0.0,
                    total=round(total, 2),
                    banco=banco_key,
                    archivo_origen=archivo_nombre,
                ))
            except (ValueError, IndexError, TypeError):
                continue
        return resultados

    with pdfplumber.open(ruta) as pdf:
        for settings in _TABLE_SETTINGS_LIST:
            cuotas_intento: list[CuotaPrestamo] = []
            for pagina in pdf.pages:
                tablas = (pagina.extract_tables(settings) if settings else pagina.extract_tables()) or []
                for tabla in tablas:
                    cuotas_intento.extend(_procesar_tabla(tabla, banco, ruta.name))
            if cuotas_intento:
                cuotas = cuotas_intento
                break  # Se encontraron cuotas con esta estrategia; no continuar

    return cuotas


def _detectar_banco_nombre_canonico(texto: str) -> str:
    """Detecta banco en texto y devuelve nombre canónico de pantalla (e.g. 'Banco Santander')."""
    texto_norm = _normalizar_texto(texto)
    mejor = "Banco Desconocido"
    mejor_score = 0
    for nombre_canonico, keywords in BANCOS_KEYWORDS_PRESTAMO.items():
        score = sum(1 for kw in keywords if kw in texto_norm)
        if score > mejor_score:
            mejor_score = score
            mejor = nombre_canonico
    return mejor


def _detectar_numero_prestamo(texto: str) -> str:
    """Extrae número de préstamo/operación del texto usando patrones comunes."""
    for patron in _PATTERNS_PRESTAMO:
        m = patron.search(texto)
        if m:
            val = m.group(1).strip()
            if val.isdigit() and len(val) > 10:
                continue
            return val
    return ""


def _detectar_sistema_amortizacion(texto: str) -> str:
    """Detecta el sistema de amortización (Francés, Alemán, Americano)."""
    norm = _normalizar_texto(texto)
    if any(k in norm for k in ("frances", "french", "sistema f", "cuota fija")):
        return "Francés"
    if any(k in norm for k in ("aleman", "german", "sistema a", "cuota decreciente")):
        return "Alemán"
    if "americano" in norm:
        return "Americano"
    return ""


def _cuotas_a_dicts(cuotas: list[CuotaPrestamo]) -> list[dict]:
    """Convierte lista de CuotaPrestamo al formato de dicts de extraer_datos_pdf_prestamo."""
    result = []
    for c in sorted(cuotas, key=lambda x: (x.fecha_vencimiento, x.numero_cuota)):
        result.append({
            "cuota": c.numero_cuota,
            "vencimiento": c.fecha_vencimiento.strftime("%Y-%m-%d"),
            "capital": round(c.capital, 2),
            "intereses": round(c.intereses, 2),
            "iva_gastos": round(c.impuestos, 2),
            "monto_abonar": round(c.total, 2),
            "saldo_restante": 0.0,
        })
    return result


def _extraer_cuotas_desde_fitz_blocks(doc, banco: str = "", archivo: str = "") -> list[CuotaPrestamo]:
    """Extrae cuotas usando get_text('blocks') de fitz para mejor reconstrucción de columnas."""
    lineas: list[str] = []
    for page in doc:
        bloques = page.get_text("blocks")
        for bloque in sorted(bloques, key=lambda b: (round(b[1] / 12) * 12, b[0])):
            texto_bloque = (bloque[4] or "").strip()
            for linea in texto_bloque.splitlines():
                if linea.strip():
                    lineas.append(linea.strip())
    return _extraer_cuotas_desde_texto(lineas, banco=banco, archivo=archivo)


def extraer_datos_pdf_prestamo(pdf_path_o_bytes) -> dict:
    """
    Pipeline de 3 niveles para extraer datos estructurados de PDFs de préstamos bancarios.

    Acepta ruta (str/Path) o bytes del PDF.

    NIVEL 1 - pdfplumber: tablas nativas + texto plano.
    NIVEL 2 - fitz:       texto y bloques reconstruidos.
    NIVEL 3 - OCR:        easyocr sobre imagen de página (fitz pixmap).

    Returns:
        {
            "banco": str,            # nombre canónico detectado
            "prestamo_n": str,       # número de préstamo/operación
            "capital_original": float,
            "sistema": str,          # Francés / Alemán / Americano
            "cuotas": [
                {
                    "cuota": int,
                    "vencimiento": "YYYY-MM-DD",
                    "capital": float,
                    "intereses": float,
                    "iva_gastos": float,
                    "monto_abonar": float,
                    "saldo_restante": float,
                }, ...
            ],
            "metodo_extraccion": str,  # "pdfplumber" | "fitz" | "fitz_blocks" | "ocr"
        }
    """
    import tempfile as _tempfile

    _borrar_temp = False
    if isinstance(pdf_path_o_bytes, (bytes, bytearray)):
        _tmp = _tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        _tmp.write(pdf_path_o_bytes)
        _tmp.close()
        ruta = Path(_tmp.name)
        _borrar_temp = True
    else:
        ruta = Path(pdf_path_o_bytes)

    resultado: dict = {
        "banco": "Banco Desconocido",
        "prestamo_n": "",
        "capital_original": 0.0,
        "sistema": "",
        "cuotas": [],
        "metodo_extraccion": "ninguno",
    }

    try:
        # --- Detectar banco y número de préstamo (primeras 2 páginas) ---
        texto_enc = _extraer_texto_nativo_pdf(ruta, max_paginas=2)
        if not texto_enc.strip():
            try:
                _doc = fitz.open(ruta)
                partes_enc = [_doc[i].get_text("text") for i in range(min(2, len(_doc)))]
                _doc.close()
                texto_enc = "\n".join(partes_enc)
            except Exception:
                pass
        if not texto_enc.strip():
            try:
                lector = _obtener_lector_ocr()
                _doc = fitz.open(ruta)
                partes_enc = []
                for i in range(min(2, len(_doc))):
                    pix = _doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))
                    img = np.array(Image.open(io.BytesIO(pix.tobytes("png"))))
                    partes_enc.extend(t for _, t, _ in lector.readtext(img))
                _doc.close()
                texto_enc = " ".join(partes_enc)
            except Exception:
                pass

        resultado["banco"] = _detectar_banco_nombre_canonico(texto_enc)
        resultado["prestamo_n"] = _detectar_numero_prestamo(texto_enc)
        resultado["sistema"] = _detectar_sistema_amortizacion(texto_enc)

        banco_key = resultado["banco"]

        # --- NIVEL 1: pdfplumber tablas ---
        try:
            cuotas_t = _extraer_cuotas_desde_tabla_pdf(ruta)
            if cuotas_t:
                for c in cuotas_t:
                    c.banco = c.banco or banco_key
                resultado["cuotas"] = _cuotas_a_dicts(cuotas_t)
                resultado["metodo_extraccion"] = "pdfplumber"
        except Exception:
            pass

        # --- NIVEL 1b: pdfplumber texto ---
        if not resultado["cuotas"]:
            try:
                texto_full = _extraer_texto_nativo_pdf(ruta)
                if texto_full.strip():
                    lineas = [l.strip() for l in texto_full.splitlines() if l.strip()]
                    cuotas_tx = _extraer_cuotas_desde_texto(lineas, banco=banco_key, archivo=ruta.name)
                    if cuotas_tx:
                        resultado["cuotas"] = _cuotas_a_dicts(cuotas_tx)
                        resultado["metodo_extraccion"] = "pdfplumber"
            except Exception:
                pass

        # --- NIVEL 2: fitz texto plano ---
        if not resultado["cuotas"]:
            try:
                _doc = fitz.open(ruta)
                lineas_fitz = []
                for page in _doc:
                    lineas_fitz.extend(
                        l.strip() for l in page.get_text("text").splitlines() if l.strip()
                    )
                _doc.close()
                cuotas_fitz = _extraer_cuotas_desde_texto(
                    lineas_fitz, banco=banco_key, archivo=ruta.name
                )
                if cuotas_fitz:
                    resultado["cuotas"] = _cuotas_a_dicts(cuotas_fitz)
                    resultado["metodo_extraccion"] = "fitz"
            except Exception:
                pass

        # --- NIVEL 2b: fitz blocks (mejor reconstrucción de columnas) ---
        if not resultado["cuotas"]:
            try:
                _doc = fitz.open(ruta)
                cuotas_fb = _extraer_cuotas_desde_fitz_blocks(_doc, banco=banco_key, archivo=ruta.name)
                _doc.close()
                if cuotas_fb:
                    resultado["cuotas"] = _cuotas_a_dicts(cuotas_fb)
                    resultado["metodo_extraccion"] = "fitz_blocks"
            except Exception:
                pass

        # --- NIVEL 3: OCR visual ---
        if not resultado["cuotas"]:
            try:
                lector = _obtener_lector_ocr()
                _doc = fitz.open(ruta)
                lineas_ocr: list[str] = []
                for page in _doc:
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    img = np.array(Image.open(io.BytesIO(pix.tobytes("png"))))
                    resultados_ocr = lector.readtext(img)
                    filas: dict[int, list[tuple[float, str]]] = {}
                    for bbox, texto_ocr, _ in resultados_ocr:
                        y = int((bbox[0][1] + bbox[2][1]) / 2 / 20) * 20
                        filas.setdefault(y, []).append((bbox[0][0], texto_ocr))
                    for y in sorted(filas):
                        linea = " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0]))
                        lineas_ocr.append(linea)
                _doc.close()
                cuotas_ocr = _extraer_cuotas_desde_texto(
                    lineas_ocr, banco=banco_key, archivo=ruta.name
                )
                if cuotas_ocr:
                    resultado["cuotas"] = _cuotas_a_dicts(cuotas_ocr)
                    resultado["metodo_extraccion"] = "ocr"
            except Exception:
                pass

        # --- Capital original = suma de amortizaciones ---
        if resultado["cuotas"]:
            resultado["capital_original"] = round(
                sum(c.get("capital", 0) for c in resultado["cuotas"]), 2
            )

    finally:
        if _borrar_temp:
            try:
                ruta.unlink()
            except Exception:
                pass

    return resultado


def extraer_cuotas_prestamo(
    pdf_paths: list[str | Path],
    permitir_ocr: bool = True,
) -> list[CuotaPrestamo]:
    """
    Extrae tabla de amortización (Capital vs Intereses por cuota) desde PDFs de proyección.
    Combina extracción por tablas pdfplumber y parsing de texto/OCR.
    """
    todas: list[CuotaPrestamo] = []
    vistos: set[tuple] = set()

    for ruta_raw in pdf_paths:
        ruta = Path(ruta_raw)
        # Extraer texto encabezado una sola vez para detección de banco y N° préstamo
        texto_enc = _extraer_texto_encabezado(ruta, max_paginas=3)
        banco = detectar_banco_pdf(ruta)

        cuotas_tabla = _extraer_cuotas_desde_tabla_pdf(ruta)
        if len(cuotas_tabla) >= 1:
            cuotas = cuotas_tabla
        else:
            lineas = _extraer_lineas_pdf(ruta, permitir_ocr=permitir_ocr)
            cuotas_texto = _extraer_cuotas_desde_texto(lineas, banco=banco, archivo=ruta.name)
            cuotas = cuotas_texto

        # Usar texto del PDF para buscar N° de contrato antes de caer al nombre de archivo
        prestamo_id = _id_prestamo_desde_archivo(str(ruta), texto_pdf=texto_enc)
        for c in cuotas:
            c.prestamo_id = prestamo_id
            c.banco = c.banco or banco
            clave = (c.prestamo_id, c.numero_cuota, c.fecha_vencimiento, c.total)
            if clave not in vistos:
                vistos.add(clave)
                todas.append(c)

    todas.sort(key=lambda c: (c.fecha_vencimiento, c.numero_cuota))
    return todas


def _es_movimiento_prestamo(mov: MovimientoBanco) -> bool:
    desc = _normalizar_texto(mov.descripcion)
    return mov.debito > 0 and any(
        k in desc for k in ("prestamo", "cuota", "amortiz", "credito", "hipotec", "capital prest")
    )


def cruzar_extractos_prestamos(
    movimientos_banco: list[MovimientoBanco],
    cuotas_prestamo: list[CuotaPrestamo],
    tolerancia_dias: int = TOLERANCIA_DIAS_CLEARING,
) -> list[CruceCuotaExtracto]:
    """
    Cruce fuzzy: débitos bancarios vs cuotas de préstamo.
    Criterio: importe ±0.02, fecha ±3 días (configurable).
    """
    debits_prestamo = [m for m in movimientos_banco if _es_movimiento_prestamo(m) or m.debito > 0]
    debits_usados: set[int] = set()
    cruces: list[CruceCuotaExtracto] = []

    for cuota in cuotas_prestamo:
        mejor_mov: Optional[MovimientoBanco] = None
        mejor_score = -1.0
        mejor_dif_imp = 999999.0
        mejor_dif_dias = 999

        for idx, mov in enumerate(debits_prestamo):
            if idx in debits_usados:
                continue
            if not _coincide_importe(mov.debito, cuota.total):
                continue
            dif_dias = _diferencia_dias(mov.fecha, cuota.fecha_vencimiento)
            if dif_dias > tolerancia_dias:
                continue
            dif_imp = abs(mov.debito - cuota.total)
            score_desc = fuzz.partial_ratio(
                _normalizar_texto(mov.descripcion),
                "prestamo cuota amortizacion",
            )
            score = score_desc - dif_dias * 5 - dif_imp * 100
            if score > mejor_score:
                mejor_score = score
                mejor_mov = mov
                mejor_dif_imp = dif_imp
                mejor_dif_dias = dif_dias

        cruce = CruceCuotaExtracto(cuota=cuota)
        if mejor_mov:
            idx_usado = debits_prestamo.index(mejor_mov)
            debits_usados.add(idx_usado)
            cruce.movimiento = mejor_mov
            cruce.coincidencia = True
            cruce.diferencia_importe = round(mejor_dif_imp, 2)
            cruce.diferencia_dias = mejor_dif_dias
            cruce.observacion = "Cruce OK"
        else:
            cruce.observacion = "Sin débito bancario coincidente"
        cruces.append(cruce)

    return cruces


def cargar_mayor_tango(archivo) -> pd.DataFrame:
    """
    Carga mayor contable Tango (Excel/CSV) normalizando columnas estándar.
    Retorna DataFrame con: fecha, codigo_cuenta, descripcion, debe, haber, saldo.
    """
    nombre = getattr(archivo, "name", str(archivo)).lower()
    if isinstance(archivo, (str, Path)):
        df = pd.read_csv(archivo) if str(archivo).lower().endswith(".csv") else pd.read_excel(archivo)
    else:
        df = pd.read_csv(archivo) if nombre.endswith(".csv") else pd.read_excel(archivo)

    df.columns = [_normalizar_texto(str(c)).replace(" ", "_") for c in df.columns]
    col_fecha = next((c for c in df.columns if "fecha" in c), df.columns[0])
    col_cuenta = next((c for c in df.columns if "codigo" in c or "cuenta" in c), None)
    col_desc = next((c for c in df.columns if "descrip" in c or "concepto" in c or "leyenda" in c), None)
    col_debe = next((c for c in df.columns if c == "debe" or "debe_" in c or "importe_debe" in c), None)
    col_haber = next((c for c in df.columns if c == "haber" or "haber_" in c or "importe_haber" in c), None)
    col_saldo = next((c for c in df.columns if "saldo" in c), None)

    if col_debe is None and col_haber is None:
        col_importe = next((c for c in df.columns if "importe" in c or "monto" in c), None)
        if col_importe:
            df["debe"] = df[col_importe].apply(lambda x: float(x) if pd.notna(x) and float(x) > 0 else 0.0)
            df["haber"] = df[col_importe].apply(lambda x: abs(float(x)) if pd.notna(x) and float(x) < 0 else 0.0)
            col_debe, col_haber = "debe", "haber"

    filas = []
    for _, row in df.iterrows():
        fecha_val = row[col_fecha]
        fecha: Optional[date] = None
        if isinstance(fecha_val, (datetime, pd.Timestamp)):
            if pd.notna(fecha_val):
                fecha = fecha_val.date()
        elif isinstance(fecha_val, date):
            fecha = fecha_val
        elif pd.notna(fecha_val):
            fecha = _parsear_fecha(str(fecha_val).strip())
        if fecha is None:
            continue

        codigo = str(row[col_cuenta]).strip() if col_cuenta else ""
        codigo = re.sub(r"\D", "", codigo)[:8] if codigo else ""
        debe = float(row[col_debe]) if col_debe and pd.notna(row[col_debe]) else 0.0
        haber = float(row[col_haber]) if col_haber and pd.notna(row[col_haber]) else 0.0
        saldo = float(row[col_saldo]) if col_saldo and pd.notna(row[col_saldo]) else None
        desc = str(row[col_desc]) if col_desc else ""

        filas.append({
            "fecha": fecha,
            "codigo_cuenta": codigo,
            "descripcion": desc,
            "debe": round(debe, 2),
            "haber": round(haber, 2),
            "saldo": round(saldo, 2) if saldo is not None else None,
        })

    return pd.DataFrame(filas)


def _es_cuenta_capital_prestamo(codigo: str) -> bool:
    c = re.sub(r"\D", "", str(codigo))
    return any(c.startswith(p) for p in PREFIJOS_CUENTA_CAPITAL_PRESTAMO)


def _es_cuenta_interes_prestamo(codigo: str) -> bool:
    c = re.sub(r"\D", "", str(codigo))
    return any(c.startswith(p) for p in PREFIJOS_CUENTA_INTERES_PRESTAMO)


def validar_contra_mayor(
    mayor_df: pd.DataFrame,
    cuotas: list[CuotaPrestamo],
    movimientos_banco: list[MovimientoBanco],
    cruces: Optional[list[CruceCuotaExtracto]] = None,
) -> list[ValidacionMayor]:
    """
    Compara imputaciones del mayor Tango vs capital/intereses esperados por cuota.
    Marca errores de clasificación (capital imputado como interés o viceversa).
    """
    if mayor_df.empty:
        return [
            ValidacionMayor(
                cuota=c,
                capital_esperado=c.capital,
                interes_esperado=c.intereses,
                error_imputacion=True,
                detalle="Mayor contable vacío",
            )
            for c in cuotas
        ]

    cruces = cruces or cruzar_extractos_prestamos(movimientos_banco, cuotas)
    mapa_cruce = {id(c.cuota): c for c in cruces}
    validaciones: list[ValidacionMayor] = []

    for cuota in cuotas:
        cruce = mapa_cruce.get(id(cuota))
        fecha_ref = cuota.fecha_vencimiento
        if cruce and cruce.movimiento:
            fecha_ref = cruce.movimiento.fecha

        ventana_ini = fecha_ref - timedelta(days=TOLERANCIA_DIAS_CLEARING)
        ventana_fin = fecha_ref + timedelta(days=TOLERANCIA_DIAS_CLEARING)
        mask_fecha = (mayor_df["fecha"] >= ventana_ini) & (mayor_df["fecha"] <= ventana_fin)
        movs_mes = mayor_df[mask_fecha]

        capital_mayor = round(
            movs_mes[movs_mes["codigo_cuenta"].apply(_es_cuenta_capital_prestamo)]["haber"].sum(), 2
        )
        interes_mayor = round(
            movs_mes[movs_mes["codigo_cuenta"].apply(_es_cuenta_interes_prestamo)]["debe"].sum(), 2
        )

        error = False
        detalle_parts = []
        if capital_mayor > 0 and not _coincide_importe(capital_mayor, cuota.capital):
            error = True
            detalle_parts.append(f"Capital mayor ${capital_mayor:,.2f} ≠ esperado ${cuota.capital:,.2f}")
        if interes_mayor > 0 and not _coincide_importe(interes_mayor, cuota.intereses):
            error = True
            detalle_parts.append(f"Interés mayor ${interes_mayor:,.2f} ≠ esperado ${cuota.intereses:,.2f}")
        if capital_mayor == 0 and interes_mayor == 0:
            error = True
            detalle_parts.append("Sin imputación en mayor para el período")

        validaciones.append(
            ValidacionMayor(
                cuota=cuota,
                capital_mayor=capital_mayor,
                interes_mayor=interes_mayor,
                capital_esperado=cuota.capital,
                interes_esperado=cuota.intereses,
                error_imputacion=error,
                detalle="; ".join(detalle_parts) if detalle_parts else "Imputación correcta",
            )
        )

    return validaciones


def constituir_saldos_finales(
    saldos_iniciales: list[dict],
    cuotas: list[CuotaPrestamo],
    movimientos_banco: list[MovimientoBanco],
    mayor_df: pd.DataFrame,
) -> list[SaldoPasivoBanco]:
    """
    Por banco: Saldo Inicial Pasivo − Capital Pagado + Nuevos Créditos = Saldo Final Mayor.
    Genera alertas si no cierra.
    """
    resultados: list[SaldoPasivoBanco] = []

    for item in saldos_iniciales:
        banco_display = item.get("banco", "")
        saldo_ini = round(float(item.get("saldo_inicial", 0) or 0), 2)
        clave = _clave_banco_desde_display(banco_display)

        cuotas_banco = [
            c for c in cuotas
            if not c.banco or c.banco == clave or _normalizar_texto(banco_display) in _normalizar_texto(c.banco)
        ]
        movs_banco = [m for m in movimientos_banco if m.banco == clave or not m.banco]

        capital_pagado = round(sum(c.capital for c in cuotas_banco), 2)
        nuevos_creditos = round(
            sum(
                m.credito for m in movs_banco
                if m.credito > 0 and any(k in _normalizar_texto(m.descripcion) for k in PALABRAS_NUEVO_CREDITO)
            ),
            2,
        )

        saldo_calculado = round(saldo_ini - capital_pagado + nuevos_creditos, 2)

        saldo_final_mayor: Optional[float] = None
        if not mayor_df.empty:
            mask_cap = mayor_df["codigo_cuenta"].apply(_es_cuenta_capital_prestamo)
            saldos_mayor = mayor_df.loc[mask_cap, "saldo"].dropna()
            if not saldos_mayor.empty:
                saldo_final_mayor = round(float(saldos_mayor.iloc[-1]), 2)

        diferencia = round((saldo_final_mayor or saldo_calculado) - saldo_calculado, 2)
        cierra = saldo_final_mayor is None or abs(diferencia) <= TOLERANCIA_IMPORTE
        alerta = ""
        if saldo_final_mayor is not None and not cierra:
            alerta = (
                f"Desbalance: SI ${saldo_ini:,.2f} − Capital ${capital_pagado:,.2f} "
                f"+ Nuevos ${nuevos_creditos:,.2f} = ${saldo_calculado:,.2f} "
                f"≠ Mayor ${saldo_final_mayor:,.2f} (dif. ${diferencia:,.2f})"
            )

        resultados.append(
            SaldoPasivoBanco(
                banco=banco_display or _nombre_display_banco(clave),
                saldo_inicial_pasivo=saldo_ini,
                capital_pagado=capital_pagado,
                nuevos_creditos=nuevos_creditos,
                saldo_calculado=saldo_calculado,
                saldo_final_mayor=saldo_final_mayor,
                diferencia=diferencia,
                cierra=cierra,
                alerta=alerta,
            )
        )

    return resultados


def detectar_modo_auditoria(
    pdf_prestamos: list[str | Path],
    pdf_extractos: list[str | Path],
    ruta_mayor,
) -> str:
    """Determina el modo de procesamiento según archivos cargados."""
    tiene_ext = bool(pdf_extractos)
    tiene_mayor = ruta_mayor is not None
    tiene_proy = bool(pdf_prestamos)

    if not tiene_ext and not tiene_mayor:
        return "teorico"
    if tiene_proy and tiene_ext and tiene_mayor:
        return "completo"
    if tiene_proy and tiene_ext:
        return "parcial_extractos"
    if tiene_ext:
        return "solo_extractos"
    return "teorico"


def _cruces_pendientes_extractos(cuotas: list[CuotaPrestamo]) -> list[CruceCuotaExtracto]:
    return [
        CruceCuotaExtracto(cuota=c, observacion=MSG_PENDIENTE_EXTRACTOS)
        for c in cuotas
    ]


def _validaciones_pendientes_mayor(
    cuotas: list[CuotaPrestamo],
    mensaje: str = MSG_PENDIENTE_MAYOR,
) -> list[ValidacionMayor]:
    return [
        ValidacionMayor(
            cuota=c,
            capital_esperado=c.capital,
            interes_esperado=c.intereses,
            error_imputacion=False,
            detalle=mensaje,
        )
        for c in cuotas
    ]


def ejecutar_auditoria_teorica(
    pdf_prestamos: list[str | Path],
    saldos_iniciales: list[dict],
    nombre_cliente: str = "",
) -> ResultadoAuditoriaPrestamos:
    """Genera estructura teórica del préstamo sin extractos ni mayor."""
    return ejecutar_auditoria_prestamos(
        pdf_prestamos=pdf_prestamos,
        pdf_extractos=[],
        ruta_mayor=None,
        saldos_iniciales=saldos_iniciales,
        nombre_cliente=nombre_cliente,
        modo="teorico",
    )


def ejecutar_auditoria_prestamos(
    pdf_prestamos: list[str | Path],
    pdf_extractos: list[str | Path],
    ruta_mayor,
    saldos_iniciales: list[dict],
    nombre_cliente: str = "",
    modo: str | None = None,
    cuotas_precalculadas: list[CuotaPrestamo] | None = None,
    movimientos_precalculados: list[MovimientoBanco] | None = None,
    bancos_precalculados: list[str] | None = None,
) -> ResultadoAuditoriaPrestamos:
    """Orquesta auditoría de préstamos según archivos disponibles (modo adaptativo)."""
    modo_final = modo or detectar_modo_auditoria(pdf_prestamos, pdf_extractos, ruta_mayor)
    resultado = ResultadoAuditoriaPrestamos(nombre_cliente=nombre_cliente, modo=modo_final)
    resultado.saldos_iniciales = list(saldos_iniciales)

    if modo_final == "teorico":
        resultado.cuotas = []
        resultado.cruces = []
        resultado.validaciones_mayor = []
        resultado.saldos_por_banco = constituir_saldos_finales(
            saldos_iniciales, [], [], pd.DataFrame()
        )
        return resultado

    if cuotas_precalculadas is not None:
        cuotas = cuotas_precalculadas
    else:
        cuotas = extraer_cuotas_prestamo(pdf_prestamos) if pdf_prestamos else []
    resultado.cuotas = cuotas

    if pdf_extractos:
        if movimientos_precalculados is not None:
            movimientos = movimientos_precalculados
            bancos = bancos_precalculados or []
        else:
            movimientos, bancos, _ = extraer_movimientos_anuales(pdf_extractos)
    else:
        movimientos, bancos = [], []
    resultado.movimientos_banco = movimientos
    resultado.bancos_procesados = bancos

    cruces = cruzar_extractos_prestamos(movimientos, cuotas) if cuotas else []
    resultado.cruces = cruces

    if modo_final == "parcial_extractos":
        resultado.validaciones_mayor = _validaciones_pendientes_mayor(cuotas)
        mayor_df = pd.DataFrame()
    elif modo_final == "completo":
        mayor_df = cargar_mayor_tango(ruta_mayor)
        resultado.validaciones_mayor = validar_contra_mayor(mayor_df, cuotas, movimientos, cruces)
    else:
        mayor_df = cargar_mayor_tango(ruta_mayor) if ruta_mayor is not None else pd.DataFrame()
        if not mayor_df.empty and cuotas:
            resultado.validaciones_mayor = validar_contra_mayor(mayor_df, cuotas, movimientos, cruces)
        else:
            resultado.validaciones_mayor = _validaciones_pendientes_mayor(cuotas)

    resultado.saldos_por_banco = constituir_saldos_finales(
        saldos_iniciales, cuotas, movimientos, mayor_df
    )

    for cruce in cruces:
        if cruce.observacion in (MSG_PENDIENTE_EXTRACTOS, MSG_PENDIENTE_MAYOR):
            continue
        if not cruce.coincidencia:
            resultado.alertas.append({
                "tipo": "Cruce extracto",
                "cuota": cruce.cuota.numero_cuota,
                "fecha": cruce.cuota.fecha_vencimiento.strftime("%d/%m/%Y"),
                "importe": cruce.cuota.total,
                "detalle": cruce.observacion,
            })

    for val in resultado.validaciones_mayor:
        if val.detalle in (MSG_PENDIENTE_EXTRACTOS, MSG_PENDIENTE_MAYOR):
            continue
        if val.error_imputacion:
            resultado.alertas.append({
                "tipo": "Imputación mayor",
                "cuota": val.cuota.numero_cuota,
                "fecha": val.cuota.fecha_vencimiento.strftime("%d/%m/%Y"),
                "importe": val.cuota.total,
                "detalle": val.detalle,
            })

    for saldo in resultado.saldos_por_banco:
        if not saldo.cierra:
            resultado.alertas.append({
                "tipo": "Constitución saldo",
                "cuota": "",
                "fecha": "",
                "importe": saldo.diferencia,
                "detalle": saldo.alerta,
                "banco": saldo.banco,
            })

    return resultado


def auditoria_prestamos_a_dataframes(resultado: ResultadoAuditoriaPrestamos) -> dict[str, pd.DataFrame]:
    """Convierte resultado de auditoría a DataFrames para la UI Streamlit."""
    df_cuotas = pd.DataFrame([
        {
            "Cuota": c.numero_cuota,
            "Vencimiento": c.fecha_vencimiento.strftime("%d/%m/%Y"),
            "Capital": c.capital,
            "Intereses": c.intereses,
            "Total": c.total,
            "Banco": _nombre_display_banco(c.banco) if c.banco else "",
            "Archivo": c.archivo_origen,
        }
        for c in resultado.cuotas
    ])

    df_cruces = pd.DataFrame([
        {
            "Cuota": cr.cuota.numero_cuota,
            "Vencimiento": cr.cuota.fecha_vencimiento.strftime("%d/%m/%Y"),
            "Total Cuota": cr.cuota.total,
            "Cruce OK": "✅" if cr.coincidencia else "❌",
            "Fecha Extracto": cr.movimiento.fecha.strftime("%d/%m/%Y") if cr.movimiento else "",
            "Débito Extracto": cr.movimiento.debito if cr.movimiento else None,
            "Δ Importe": cr.diferencia_importe,
            "Δ Días": cr.diferencia_dias,
            "Observación": cr.observacion,
        }
        for cr in resultado.cruces
    ])

    df_mayor = pd.DataFrame([
        {
            "Cuota": v.cuota.numero_cuota,
            "Capital Esperado": v.capital_esperado,
            "Capital Mayor": v.capital_mayor,
            "Interés Esperado": v.interes_esperado,
            "Interés Mayor": v.interes_mayor,
            "Error": "⚠️ Sí" if v.error_imputacion else "✅",
            "Detalle": v.detalle,
        }
        for v in resultado.validaciones_mayor
    ])

    df_saldos = pd.DataFrame([
        {
            "Banco": s.banco,
            "Saldo Inicial Pasivo": s.saldo_inicial_pasivo,
            "Capital Pagado": s.capital_pagado,
            "Nuevos Créditos": s.nuevos_creditos,
            "Saldo Calculado": s.saldo_calculado,
            "Saldo Final Mayor": s.saldo_final_mayor,
            "Diferencia": s.diferencia,
            "Cierra": "✅" if s.cierra else "❌",
        }
        for s in resultado.saldos_por_banco
    ])

    df_alertas = pd.DataFrame(resultado.alertas) if resultado.alertas else pd.DataFrame(
        columns=["tipo", "cuota", "fecha", "importe", "detalle"]
    )

    return {
        "cuotas": df_cuotas,
        "cruces": df_cruces,
        "mayor": df_mayor,
        "saldos": df_saldos,
        "alertas": df_alertas,
    }


def _sanitizar_nombre_hoja_excel(nombre: str) -> str:
    """Sanitiza nombre para hoja Excel (máx. 31 caracteres, sin caracteres prohibidos)."""
    limpio = re.sub(r"[\\/*?:\[\]]", "", nombre).strip()
    return (limpio[:31] if limpio else "Prestamos")


_FONT_TITULO_PRESTAMOS = Font(bold=True, size=12)
_FONT_ENC_PRESTAMOS = Font(bold=True, size=10)
_FONT_FILA_PRESTAMOS = Font(size=10)


def _escribir_titulo_prestamo(ws, fila: int, col: int, valor) -> None:
    celda = ws.cell(row=fila, column=col, value=valor)
    celda.font = _FONT_TITULO_PRESTAMOS


def _escribir_encabezados_grilla_prestamos(ws, fila: int) -> None:
    for col, enc in enumerate(COLUMNAS_GRILLA_PRESTAMOS, start=1):
        celda = ws.cell(row=fila, column=col, value=enc)
        celda.font = _FONT_ENC_PRESTAMOS


def _escribir_fila_grilla_prestamo(
    ws,
    fila: int,
    fecha,
    capital: float | str,
    impuestos: float | str,
    intereses: float | str,
    total: float | str,
    resaltar_rojo: bool = False,
) -> None:
    valores = [fecha, capital, impuestos, intereses, total]
    for col, val in enumerate(valores, start=1):
        celda = ws.cell(row=fila, column=col, value=val)
        celda.font = _FONT_FILA_PRESTAMOS
        if isinstance(val, (int, float)):
            celda.number_format = "#,##0.00"
        if resaltar_rojo:
            celda.fill = _FILL_ALERTA_ROJO


def _ajustar_anchos_grilla_prestamos(ws) -> None:
    anchos = [14, 24, 14, 14, 18]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _nuevo_workbook_prestamos() -> openpyxl.Workbook:
    """Workbook independiente para papel de trabajo de préstamos (sin plantilla conciliación)."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _id_prestamo_desde_archivo(archivo: str, texto_pdf: str = "") -> str:
    """
    Identificador de préstamo. Primero busca N° de contrato dentro del texto del PDF;
    si no lo encuentra, usa el nombre original del archivo (preservado en el prefijo ec_).
    """
    if not archivo:
        return "Prestamo_1"

    if texto_pdf:
        n = _extraer_numero_prestamo_pdf(Path(archivo), texto_pdf=texto_pdf)
        if n and n != Path(archivo).stem:
            return n

    ruta = Path(archivo)
    stem = ruta.stem.strip()
    # Si viene de _guardar_upload con prefijo ec_{nombre}_{hash}
    if stem.startswith("ec_"):
        partes = stem[3:].rsplit("_", 1)
        nombre_limpio = partes[0].replace("_", " ").strip() if partes else stem[3:]
        if nombre_limpio:
            return nombre_limpio

    return stem or "Prestamo_1"


def _clave_cruce_cuota(cuota: CuotaPrestamo) -> tuple:
    """Clave única cuota dentro de un banco (soporta múltiples préstamos)."""
    return (cuota.prestamo_id or cuota.archivo_origen, cuota.numero_cuota)


def _mapa_cruces_por_cuota(cruces: list[CruceCuotaExtracto]) -> dict[tuple, CruceCuotaExtracto]:
    return {_clave_cruce_cuota(c.cuota): c for c in cruces}


def _mapa_validaciones_por_cuota(validaciones: list[ValidacionMayor]) -> dict[tuple, ValidacionMayor]:
    return {_clave_cruce_cuota(v.cuota): v for v in validaciones}


def _cuotas_pertenecen_banco(cuota: CuotaPrestamo, banco_display: str) -> bool:
    """Determina si una cuota corresponde al banco configurado en saldos iniciales."""
    clave_ui = _clave_banco_desde_display(banco_display)
    if cuota.banco == clave_ui:
        return True
    if _normalizar_texto(banco_display) in _normalizar_texto(cuota.banco):
        return True
    if cuota.banco and _normalizar_texto(_nombre_display_banco(cuota.banco)) == _normalizar_texto(banco_display):
        return True
    return not cuota.banco


def _agrupar_prestamos_por_banco(
    cuotas: list[CuotaPrestamo],
    saldos_iniciales: list[dict],
) -> dict[str, dict[str, list[CuotaPrestamo]]]:
    """
    Agrupa cuotas por banco (UI saldos iniciales) y por préstamo (PDF de proyección).
    Retorna {nombre_banco: {prestamo_id: [cuotas]}}.
    """
    bancos_ui = [item.get("banco", "") for item in saldos_iniciales if item.get("banco")]
    if not bancos_ui:
        bancos_ui = sorted({
            _nombre_display_banco(c.banco) if c.banco else "Préstamos"
            for c in cuotas
        })

    agrupado: dict[str, dict[str, list[CuotaPrestamo]]] = {b: {} for b in bancos_ui}

    for cuota in cuotas:
        banco_asignado = next(
            (b for b in bancos_ui if _cuotas_pertenecen_banco(cuota, b)),
            _nombre_display_banco(cuota.banco) if cuota.banco else bancos_ui[0],
        )
        pid = cuota.prestamo_id or _id_prestamo_desde_archivo(cuota.archivo_origen)
        agrupado.setdefault(banco_asignado, {}).setdefault(pid, []).append(cuota)

    for banco in bancos_ui:
        agrupado.setdefault(banco, {})

    return agrupado


def _capital_total_prestamo(cuotas: list[CuotaPrestamo]) -> float:
    return round(sum(c.capital for c in cuotas), 2)


def _escribir_celda_estilo(
    ws,
    fila: int,
    col: int,
    valor,
    estilo_ref,
    resaltar_rojo: bool = False,
    resaltar_verde: bool = False,
    formato_num: str | None = "#,##0.00",
) -> None:
    """Escribe celda clonando estilo de referencia; opcional resaltado condicional."""
    celda = ws.cell(row=fila, column=col, value=valor)
    _copiar_estilo_celda(estilo_ref, celda)
    if isinstance(valor, (int, float)) and formato_num:
        celda.number_format = formato_num
    if resaltar_rojo:
        celda.fill = _FILL_ALERTA_ROJO
    elif resaltar_verde:
        celda.fill = _FILL_ALERTA_VERDE


def _escribir_bloque_prestamo_hoja(
    ws,
    fila_inicio: int,
    titulo_prestamo: str,
    cuotas: list[CuotaPrestamo],
    cruces_map: dict[tuple, CruceCuotaExtracto],
    validaciones_map: dict[tuple, ValidacionMayor],
    modo: str = "completo",
) -> int:
    """Escribe bloque de cuotas con columnas estándar de proyección PDF."""
    fila = fila_inicio
    _escribir_titulo_prestamo(ws, fila, 1, f"Préstamo: {titulo_prestamo}")
    fila += 1
    _escribir_encabezados_grilla_prestamos(ws, fila)
    fila += 1

    cuotas_ord = sorted(cuotas, key=lambda c: (c.fecha_vencimiento, c.numero_cuota))
    if not cuotas_ord and modo == "teorico":
        _escribir_fila_grilla_prestamo(
            ws, fila, MSG_PENDIENTE_CUOTAS, "", "", "", "",
        )
        return fila + 2

    for cuota in cuotas_ord:
        clave = _clave_cruce_cuota(cuota)
        cruce = cruces_map.get(clave)
        validacion = validaciones_map.get(clave)
        error_imputacion = bool(
            validacion
            and validacion.error_imputacion
            and validacion.detalle not in (MSG_PENDIENTE_EXTRACTOS, MSG_PENDIENTE_MAYOR)
        )

        if modo == "teorico":
            fecha_val = MSG_PENDIENTE_EXTRACTOS
        elif cruce and cruce.movimiento:
            fecha_val = cruce.movimiento.fecha.strftime("%d/%m/%Y")
        elif cuota.fecha_vencimiento:
            fecha_val = cuota.fecha_vencimiento.strftime("%d/%m/%Y")
        else:
            fecha_val = ""

        capital = round(cuota.capital, 2)
        impuestos = round(cuota.impuestos, 2)
        intereses = round(cuota.intereses, 2)
        total = round(cuota.total or (capital + impuestos + intereses), 2)

        _escribir_fila_grilla_prestamo(
            ws, fila, fecha_val, capital, impuestos, intereses, total,
            resaltar_rojo=error_imputacion,
        )
        fila += 1

    return fila + 1


def _escribir_control_mayor_hoja(
    ws,
    fila_inicio: int,
    saldo_pasivo: SaldoPasivoBanco,
) -> int:
    """Control de balance pasivo vs mayor Tango al pie de la hoja del banco."""
    fila = fila_inicio + 1
    _escribir_titulo_prestamo(ws, fila, 1, "Control Mayor — Pasivo Préstamos")
    fila += 1

    filas_control = [
        ("Saldo Inicial Banco", saldo_pasivo.saldo_inicial_pasivo),
        ("Total Capital Amortizado", saldo_pasivo.capital_pagado),
        ("Saldo Calculado (SI − Capital + Nuevos)", saldo_pasivo.saldo_calculado),
        ("Saldo Final Mayor (Tango)", saldo_pasivo.saldo_final_mayor),
        ("Diferencia", saldo_pasivo.diferencia),
        ("Estado", "OK" if saldo_pasivo.cierra else "DESBALANCE"),
    ]
    alerta = not saldo_pasivo.cierra
    for etiqueta, valor in filas_control:
        ws.cell(row=fila, column=1, value=etiqueta).font = _FONT_FILA_PRESTAMOS
        celda_val = ws.cell(row=fila, column=2, value=valor)
        celda_val.font = _FONT_FILA_PRESTAMOS
        if isinstance(valor, (int, float)):
            celda_val.number_format = "#,##0.00"
        if alerta:
            celda_val.fill = _FILL_ALERTA_ROJO
        elif etiqueta == "Estado":
            celda_val.fill = _FILL_ALERTA_VERDE
        fila += 1

    if saldo_pasivo.alerta:
        celda = ws.cell(row=fila, column=1, value=saldo_pasivo.alerta)
        celda.font = _FONT_FILA_PRESTAMOS
        celda.fill = _FILL_ALERTA_ROJO
        fila += 1

    return fila


def _crear_hoja_auditoria_banco(
    wb,
    nombre_banco: str,
    prestamos: dict[str, list[CuotaPrestamo]],
    saldo_pasivo: Optional[SaldoPasivoBanco],
    cruces: list[CruceCuotaExtracto],
    validaciones: list[ValidacionMayor],
    modo: str = "completo",
) -> None:
    """Crea hoja de auditoría por banco con bloques apilados por préstamo."""
    nombre_hoja = _sanitizar_nombre_hoja_excel(nombre_banco)
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)

    _escribir_titulo_prestamo(ws, 1, 1, f"Préstamos Financieros — {nombre_banco}")

    cruces_map = _mapa_cruces_por_cuota(cruces)
    validaciones_map = _mapa_validaciones_por_cuota(validaciones)

    fila = 3
    if not prestamos:
        ws.cell(row=fila, column=1, value="Sin cuotas detectadas para este banco.").font = _FONT_FILA_PRESTAMOS
        fila += 2
    else:
        for idx, (pid, cuotas_prestamo) in enumerate(sorted(prestamos.items()), start=1):
            titulo = pid if len(prestamos) > 1 else "Préstamo único"
            fila = _escribir_bloque_prestamo_hoja(
                ws, fila, titulo, cuotas_prestamo,
                cruces_map, validaciones_map,
                modo=modo,
            )

    if saldo_pasivo and modo == "completo":
        _escribir_control_mayor_hoja(ws, fila, saldo_pasivo)
    elif saldo_pasivo and modo == "parcial_extractos":
        fila += 1
        ws.cell(row=fila, column=1, value=f"Control Mayor: {MSG_PENDIENTE_MAYOR}").font = _FONT_FILA_PRESTAMOS

    _ajustar_anchos_grilla_prestamos(ws)


def _crear_hoja_consolidado_auditoria(
    wb,
    resultado: ResultadoAuditoriaPrestamos,
) -> None:
    """Hoja Consolidado con resumen por banco."""
    nombre_consolidado = "Consolidado"
    if nombre_consolidado in wb.sheetnames:
        del wb[nombre_consolidado]
    ws = wb.create_sheet(nombre_consolidado, 0)

    _escribir_titulo_prestamo(ws, 1, 1, "Préstamos Financieros — Consolidado")
    ws.cell(row=2, column=1, value=f"Cliente: {resultado.nombre_cliente}").font = _FONT_FILA_PRESTAMOS
    modo_label = DESCRIPCION_MODO_AUDITORIA.get(resultado.modo, resultado.modo)
    ws.cell(row=3, column=1, value=f"Modo: {modo_label}").font = _FONT_FILA_PRESTAMOS

    enc_cons = [
        "Banco", "Saldo Inicial Pasivo", "Capital Pagado", "Nuevos Créditos",
        "Saldo Calculado", "Saldo Final Mayor", "Diferencia", "Estado",
    ]
    for col, enc in enumerate(enc_cons, start=1):
        ws.cell(row=5, column=col, value=enc).font = _FONT_ENC_PRESTAMOS

    for i, saldo in enumerate(resultado.saldos_por_banco, start=6):
        valores = [
            saldo.banco,
            saldo.saldo_inicial_pasivo,
            saldo.capital_pagado,
            saldo.nuevos_creditos,
            saldo.saldo_calculado,
            saldo.saldo_final_mayor,
            saldo.diferencia,
            "OK" if saldo.cierra else "ALERTA",
        ]
        for col, val in enumerate(valores, start=1):
            celda = ws.cell(row=i, column=col, value=val)
            celda.font = _FONT_FILA_PRESTAMOS
            if isinstance(val, (int, float)):
                celda.number_format = "#,##0.00"
            if not saldo.cierra:
                celda.fill = _FILL_ALERTA_ROJO
            elif col == 8:
                celda.fill = _FILL_ALERTA_VERDE

    base = len(resultado.saldos_por_banco) + 8
    ws.cell(row=base, column=1, value="Totales cuotas").font = _FONT_FILA_PRESTAMOS
    ws.cell(row=base, column=2, value=len(resultado.cuotas)).font = _FONT_FILA_PRESTAMOS
    ws.cell(row=base + 1, column=1, value="Cruces OK").font = _FONT_FILA_PRESTAMOS
    ws.cell(row=base + 1, column=2, value=sum(1 for c in resultado.cruces if c.coincidencia)).font = _FONT_FILA_PRESTAMOS
    ws.cell(row=base + 2, column=1, value="Alertas").font = _FONT_FILA_PRESTAMOS
    ws.cell(row=base + 2, column=2, value=len(resultado.alertas)).font = _FONT_FILA_PRESTAMOS


def _crear_hoja_alertas_auditoria(wb, resultado: ResultadoAuditoriaPrestamos) -> None:
    """Hoja Alertas de Auditoría."""
    nombre_alertas = "Alertas"
    if nombre_alertas in wb.sheetnames:
        del wb[nombre_alertas]
    ws = wb.create_sheet(nombre_alertas)

    _escribir_titulo_prestamo(ws, 1, 1, "Alertas — Préstamos vs Mayor")

    enc_alert = ["Tipo", "Cuota", "Fecha", "Importe / Diferencia", "Detalle", "Banco"]
    for col, enc in enumerate(enc_alert, start=1):
        ws.cell(row=3, column=col, value=enc).font = _FONT_ENC_PRESTAMOS

    for i, alerta in enumerate(resultado.alertas, start=4):
        fila_vals = [
            alerta.get("tipo", ""),
            alerta.get("cuota", ""),
            alerta.get("fecha", ""),
            alerta.get("importe", ""),
            alerta.get("detalle", ""),
            alerta.get("banco", ""),
        ]
        for col, val in enumerate(fila_vals, start=1):
            celda = ws.cell(row=i, column=col, value=val)
            celda.font = _FONT_FILA_PRESTAMOS
            celda.fill = _FILL_ALERTA_ROJO


MSG_PENDIENTE_CUOTAS = "Pendiente"


def _crear_hoja_teorica_banco(
    wb,
    nombre_banco: str,
    saldo_inicial: float,
) -> None:
    """Hoja teórica por banco: grilla vacía lista para completar o parsear PDF."""
    nombre_hoja = _sanitizar_nombre_hoja_excel(nombre_banco)
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)

    _escribir_titulo_prestamo(ws, 1, 1, f"Préstamos Financieros — {nombre_banco}")
    ws.cell(row=2, column=1, value=f"Saldo inicial pasivo: ${round(float(saldo_inicial or 0), 2):,.2f}").font = _FONT_FILA_PRESTAMOS

    fila = 4
    _escribir_titulo_prestamo(ws, fila, 1, "Préstamo: Pendiente de proyección PDF")
    fila += 1
    _escribir_encabezados_grilla_prestamos(ws, fila)
    fila += 1
    _escribir_fila_grilla_prestamo(ws, fila, MSG_PENDIENTE_CUOTAS, "", "", "", "")

    saldo_pasivo = SaldoPasivoBanco(
        banco=nombre_banco,
        saldo_inicial_pasivo=round(float(saldo_inicial or 0), 2),
        capital_pagado=0.0,
        nuevos_creditos=0.0,
        saldo_calculado=round(float(saldo_inicial or 0), 2),
        cierra=True,
    )
    _escribir_control_mayor_hoja(ws, fila + 2, saldo_pasivo)
    _ajustar_anchos_grilla_prestamos(ws)


def _crear_hoja_consolidado_teorico(
    wb,
    saldos_iniciales: list[dict],
    nombre_cliente: str,
) -> None:
    """Hoja Consolidado para modo teórico (solo saldos iniciales por banco)."""
    nombre_consolidado = "Consolidado"
    if nombre_consolidado in wb.sheetnames:
        del wb[nombre_consolidado]
    ws = wb.create_sheet(nombre_consolidado, 0)

    _escribir_titulo_prestamo(ws, 1, 1, "Préstamos Financieros — Consolidado")
    ws.cell(row=2, column=1, value=f"Cliente: {nombre_cliente}").font = _FONT_FILA_PRESTAMOS
    ws.cell(row=3, column=1, value=f"Modo: {DESCRIPCION_MODO_AUDITORIA['teorico']}").font = _FONT_FILA_PRESTAMOS

    enc_cons = [
        "Banco", "Saldo Inicial Pasivo", "Capital Pagado", "Nuevos Créditos",
        "Saldo Calculado", "Estado",
    ]
    for col, enc in enumerate(enc_cons, start=1):
        ws.cell(row=5, column=col, value=enc).font = _FONT_ENC_PRESTAMOS

    for i, item in enumerate(saldos_iniciales, start=6):
        banco = item.get("banco", "")
        saldo_ini = round(float(item.get("saldo_inicial", 0) or 0), 2)
        valores = [banco, saldo_ini, 0.0, 0.0, saldo_ini, "Pendiente extractos/mayor"]
        for col, val in enumerate(valores, start=1):
            celda = ws.cell(row=i, column=col, value=val)
            celda.font = _FONT_FILA_PRESTAMOS
            if isinstance(val, (int, float)):
                celda.number_format = "#,##0.00"
            if col == 6:
                celda.fill = _FILL_ALERTA_VERDE


def generar_planilla_teorica_bancos(
    saldos_iniciales_bancos: list[dict],
    cliente_nombre: str,
    ruta_plantilla: str | Path | None = None,
) -> bytes:
    """
    Genera planilla Excel teórica en el formato aprobado (demo_prestamos_auditoria.xlsx).
    Delega en generar_excel_prestamos_auditoria con cuotas vacías por banco.
    """
    _ = ruta_plantilla

    bancos = [item for item in saldos_iniciales_bancos if item.get("banco")]
    if not bancos:
        bancos = [{"banco": "Préstamos", "saldo_inicial": 0.0}]

    bancos_data = {
        item["banco"]: [{
            "prestamo_n":       1,
            "capital_original": float(item.get("saldo_inicial", 0) or 0),
            "sistema":          "Pendiente",
            "cuotas":           [],
        }]
        for item in bancos
    }
    saldos_iniciales = {
        item["banco"]: float(item.get("saldo_inicial", 0) or 0)
        for item in bancos
    }

    ruta = generar_excel_prestamos_auditoria(bancos_data, saldos_iniciales, modo="teorico")
    with open(ruta, "rb") as fh:
        return fh.read()


def generar_reporte_auditoria_prestamos(
    resultado: ResultadoAuditoriaPrestamos,
    ruta_plantilla: str | Path | None = None,
    modo: str | None = None,
) -> bytes:
    """
    Excel independiente multi-hoja para papel de trabajo de préstamos.
    Estructura: Consolidado | una hoja por banco | Alertas.
    """
    _ = ruta_plantilla  # compatibilidad; ya no usa plantilla de conciliación
    modo_reporte = modo or resultado.modo or "completo"
    if modo_reporte == "teorico" and not resultado.cuotas:
        saldos = resultado.saldos_iniciales or [
            {"banco": s.banco, "saldo_inicial": s.saldo_inicial_pasivo}
            for s in resultado.saldos_por_banco
        ]
        return generar_planilla_teorica_bancos(
            saldos,
            resultado.nombre_cliente,
        )

    wb = _nuevo_workbook_prestamos()

    saldos_iniciales = resultado.saldos_iniciales or [
        {"banco": s.banco, "saldo_inicial": s.saldo_inicial_pasivo}
        for s in resultado.saldos_por_banco
    ]
    agrupado = _agrupar_prestamos_por_banco(resultado.cuotas, saldos_iniciales)
    mapa_saldos = {s.banco: s for s in resultado.saldos_por_banco}

    _crear_hoja_consolidado_auditoria(wb, resultado)

    for banco_display in [item.get("banco", "") for item in saldos_iniciales if item.get("banco")]:
        prestamos = agrupado.get(banco_display, {})
        saldo_pasivo = mapa_saldos.get(banco_display)
        _crear_hoja_auditoria_banco(
            wb,
            banco_display,
            prestamos,
            saldo_pasivo,
            resultado.cruces,
            resultado.validaciones_mayor,
            modo=modo_reporte,
        )

    _crear_hoja_alertas_auditoria(wb, resultado)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel_prestamos_auditoria(
    bancos_data: dict,
    saldos_iniciales: dict,
    modo: str = "completo",
) -> str:
    """
    Genera Excel de auditoría de préstamos con arquitectura contable profesional.
    Formato idéntico al demo aprobado (demo_prestamos_auditoria.xlsx).

    Parámetros
    ----------
    bancos_data : dict
        {banco_nombre: [{"prestamo_n": int, "capital_original": float,
                         "sistema": str, "cuotas": [{"cuota": int,
                         "vencimiento": str, "capital": float,
                         "intereses": float, "iva_gastos": float,
                         "monto_abonar": float, "saldo_restante": float}]}]}
        También acepta claves "n" y "capital" como alias de "prestamo_n" y "capital_original".
    saldos_iniciales : dict
        {banco_nombre: float}
    modo : str
        "completo" | "teorico"

    Retorna
    -------
    str
        Ruta absoluta del archivo generado.
    """
    # ── Paleta de colores ──────────────────────────────────────────────────────
    _C_GRIS_HEADER = "FFB8B8B8"
    _C_AZUL_OSCURO = "FF1F4E79"
    _C_AZUL_MEDIO  = "FF2E75B6"
    _C_GRIS_CLARO  = "FFF2F2F2"
    _C_AMARILLO    = "FFFFFF00"
    _C_ROJO        = "FFFF0000"

    fill_gris_header = PatternFill(start_color=_C_GRIS_HEADER, end_color=_C_GRIS_HEADER, fill_type="solid")
    fill_azul_oscuro = PatternFill(start_color=_C_AZUL_OSCURO, end_color=_C_AZUL_OSCURO, fill_type="solid")
    fill_azul_medio  = PatternFill(start_color=_C_AZUL_MEDIO,  end_color=_C_AZUL_MEDIO,  fill_type="solid")
    fill_gris_claro  = PatternFill(start_color=_C_GRIS_CLARO,  end_color=_C_GRIS_CLARO,  fill_type="solid")
    fill_amarillo    = PatternFill(start_color=_C_AMARILLO,     end_color=_C_AMARILLO,    fill_type="solid")
    fill_rojo        = PatternFill(start_color=_C_ROJO,         end_color=_C_ROJO,        fill_type="solid")

    font_bold_11   = Font(name="Calibri", bold=True, size=11)
    font_bold_11_w = Font(name="Calibri", bold=True, size=11, color="FFFFFFFF")
    font_bold_10_w = Font(name="Calibri", bold=True, size=10, color="FFFFFFFF")
    font_bold_10   = Font(name="Calibri", bold=True, size=10)
    font_data_10   = Font(name="Calibri", size=10)

    FMT_MONEDA  = '$ #,##0.00'
    COLS_GRILLA = ["CUOTA", "VENCIMIENTO", "CAPITAL", "INTERESES",
                   "IVA/GASTOS", "MONTO A ABONAR", "SALDO RESTANTE"]
    ANCHOS      = [8, 14, 16, 16, 16, 18, 18]
    N_COLS      = len(COLS_GRILLA)  # 7

    alin_izq    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    alin_der    = Alignment(horizontal="right",  vertical="center")
    alin_centro = Alignment(horizontal="center", vertical="center")

    _thin        = Side(style="thin",   color="FFB8B8B8")
    _med         = Side(style="medium", color="FF2E75B6")
    border_thin  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    border_grilla = Border(left=_med,  right=_med,  top=_med,  bottom=_med)

    # ── Helpers internos ──────────────────────────────────────────────────────
    def _fila_merged(ws, fila: int, valor, fill, font, height: int = 22) -> None:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=N_COLS)
        c = ws.cell(row=fila, column=1, value=valor)
        c.fill      = fill
        c.font      = font
        c.alignment = alin_izq
        ws.row_dimensions[fila].height = height

    def _encabezado_grilla(ws, fila: int) -> None:
        for col, enc in enumerate(COLS_GRILLA, start=1):
            c = ws.cell(row=fila, column=col, value=enc)
            c.fill      = fill_azul_medio
            c.font      = font_bold_10_w
            c.alignment = alin_centro
            c.border    = border_grilla
        ws.row_dimensions[fila].height = 18

    def _fila_cuota(ws, fila: int, datos: dict, es_par: bool) -> None:
        relleno = fill_gris_claro if es_par else None
        valores = [
            datos.get("cuota",          "—"),
            datos.get("vencimiento",    "—"),
            datos.get("capital",        "—"),
            datos.get("intereses",      "—"),
            datos.get("iva_gastos",     "—"),
            datos.get("monto_abonar",   "—"),
            datos.get("saldo_restante", "—"),
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.font   = font_data_10
            c.border = border_thin
            if relleno:
                c.fill = relleno
            if col <= 2:
                c.alignment = alin_izq
            else:
                c.alignment = alin_der
                if isinstance(val, (int, float)):
                    c.number_format = FMT_MONEDA
        ws.row_dimensions[fila].height = 15

    def _ajustar_anchos(ws) -> None:
        for i, ancho in enumerate(ANCHOS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

    def _sum_q(cuotas: list, campo: str) -> float:
        return sum(
            c.get(campo, 0) for c in cuotas
            if isinstance(c.get(campo), (int, float))
        )

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for nombre_banco, prestamos_lista in bancos_data.items():
        nombre_hoja = _sanitizar_nombre_hoja_excel(nombre_banco)
        if nombre_hoja in wb.sheetnames:
            del wb[nombre_hoja]
        ws   = wb.create_sheet(nombre_hoja)
        fila = 1

        for idx_p, prestamo in enumerate(prestamos_lista):
            n           = prestamo.get("prestamo_n") or prestamo.get("n") or (idx_p + 1)
            capital_orig = float(prestamo.get("capital_original") or prestamo.get("capital") or 0)
            sistema     = prestamo.get("sistema", "—")
            cuotas      = prestamo.get("cuotas", [])

            # [A] Cabecera del préstamo ──────────────────────────────────────
            titulo = (
                f"PRÉSTAMO N° {n}  |  Capital Original: $ {capital_orig:,.0f}"
                f"  |  Sistema: {sistema}"
            )
            _fila_merged(ws, fila, titulo, fill_gris_header, font_bold_11, height=22)
            fila += 1

            # [B] Bloque resumen anual: título azul oscuro + labels + valores ──
            total_capital   = _sum_q(cuotas, "capital")
            total_intereses = _sum_q(cuotas, "intereses")
            total_iva       = _sum_q(cuotas, "iva_gastos")

            anio_label = ""
            for q in cuotas:
                v = str(q.get("vencimiento", ""))
                if len(v) >= 10 and v[6:10].isdigit():  # formato dd/mm/yyyy
                    anio_label = v[6:10]
                    break

            titulo_resumen = f"Resumen Año {anio_label}" if anio_label else "Resumen Anual"
            _fila_merged(ws, fila, titulo_resumen, fill_azul_oscuro, font_bold_11_w, height=18)
            fila += 1

            enc_res = ["Total Capital Amortizado", "Total Intereses Devengados", "Total IVA/Gastos"]
            for col, enc in enumerate(enc_res, start=1):
                c = ws.cell(row=fila, column=col, value=enc)
                c.font      = font_bold_10
                c.alignment = alin_izq
            ws.row_dimensions[fila].height = 16
            fila += 1

            for col, val in enumerate([total_capital, total_intereses, total_iva], start=1):
                c = ws.cell(row=fila, column=col, value=val)
                c.font          = font_data_10
                c.number_format = FMT_MONEDA
                c.alignment     = alin_der
            ws.row_dimensions[fila].height = 15
            fila += 1
            fila += 1  # línea en blanco entre resumen y grilla

            # [C] Grilla cronológica de cuotas ──────────────────────────────
            _encabezado_grilla(ws, fila)
            fila += 1

            if not cuotas:
                for col in range(1, N_COLS + 1):
                    c = ws.cell(row=fila, column=col, value="—")
                    c.font   = font_data_10
                    c.border = border_thin
                ws.row_dimensions[fila].height = 15
                fila += 1
            else:
                for idx_c, q in enumerate(cuotas):
                    _fila_cuota(ws, fila, q, idx_c % 2 == 1)
                    fila += 1

            # 4 filas en blanco entre préstamos
            fila += 4

        # [D] Conciliación contable final al pie de la hoja ─────────────────
        saldo_ini       = float(saldos_iniciales.get(nombre_banco, 0) or 0)
        total_cap_banco = sum(_sum_q(p.get("cuotas", []), "capital") for p in prestamos_lista)
        saldo_final     = saldo_ini - total_cap_banco

        fila += 2
        _fila_merged(ws, fila, "CONCILIACIÓN CONTABLE FINAL",
                     fill_azul_oscuro, font_bold_11_w, height=22)
        fila += 1

        filas_cierre = [
            ("Concepto",                                   "Importe",       None),
            ("Saldo Inicial del Banco",                    saldo_ini,       None),
            ("Total Capital Amortizado (todos préstamos)", total_cap_banco, None),
            ("Saldo Final Sugerido Mayor Contable",        saldo_final,
             fill_amarillo if saldo_final >= 0 else fill_rojo),
        ]

        for i, (concepto, importe, relleno_cierre) in enumerate(filas_cierre):
            c_concepto = ws.cell(row=fila, column=1, value=concepto)
            c_importe  = ws.cell(row=fila, column=2, value=importe)
            fnt = font_bold_10 if i == 0 else font_data_10
            c_concepto.font      = fnt
            c_importe.font       = fnt
            c_concepto.alignment = alin_izq
            c_importe.alignment  = alin_der
            if i > 0 and isinstance(importe, (int, float)):
                c_importe.number_format = FMT_MONEDA
            if relleno_cierre:
                c_importe.fill  = relleno_cierre
                c_concepto.fill = relleno_cierre
            ws.row_dimensions[fila].height = 16
            fila += 1

        _ajustar_anchos(ws)
        ws.sheet_view.zoomScale = 90

    # ── Hoja "Resumen Ejecutivo" (portada, siempre primera) ───────────────────
    ws_cover = wb.create_sheet("Resumen Ejecutivo", 0)
    ws_cover.column_dimensions["A"].width = 32
    ws_cover.column_dimensions["B"].width = 22
    ws_cover.column_dimensions["C"].width = 22

    total_saldos = sum(float(v or 0) for v in saldos_iniciales.values())

    cover_rows: list = [
        ("AUDITORÍA DE PRÉSTAMOS FINANCIEROS",),
        (f"Generado: {date.today().strftime('%d/%m/%Y')}  |  Modo: {modo}",),
        (),
        ("Fecha de generación", date.today().strftime("%d/%m/%Y"),),
        ("Modo de análisis", modo,),
        (),
        ("RESUMEN POR ENTIDAD", "Capital Total", "Saldo Inicial"),
    ]
    for banco_cv, prestamos_cv in bancos_data.items():
        cap_cv = sum(
            float(p.get("capital_original") or p.get("capital") or 0)
            for p in prestamos_cv
        )
        sal_cv = float(saldos_iniciales.get(banco_cv, 0) or 0)
        cover_rows.append((banco_cv, cap_cv, sal_cv))

    cover_rows += [
        (),
        ("DETALLE DE PRÉSTAMOS",),
        ("Banco", "Préstamo", "Sistema | Cuotas | Capital"),
    ]
    for banco_cv, prestamos_cv in bancos_data.items():
        for p in prestamos_cv:
            n_cv  = p.get("prestamo_n") or p.get("n") or "—"
            sis   = p.get("sistema", "—")
            n_q   = len(p.get("cuotas", []))
            cap_p = float(p.get("capital_original") or p.get("capital") or 0)
            cover_rows.append((
                banco_cv,
                f"Préstamo N° {n_cv}",
                f"{sis} | {n_q} cuotas | $ {cap_p:,.0f}",
            ))

    total_cap_gbl = sum(
        sum(float(p.get("capital_original") or p.get("capital") or 0) for p in pl)
        for pl in bancos_data.values()
    )
    total_int_gbl = sum(
        sum(_sum_q(p.get("cuotas", []), "intereses") for p in pl)
        for pl in bancos_data.values()
    )
    total_iva_gbl = sum(
        sum(_sum_q(p.get("cuotas", []), "iva_gastos") for p in pl)
        for pl in bancos_data.values()
    )

    cover_rows += [
        (),
        ("CONCILIACIÓN GLOBAL",),
        ("Total saldos iniciales",      total_saldos,),
        ("Total capital a amortizar",   total_cap_gbl,),
        ("Total intereses devengados",  total_int_gbl,),
        ("Total IVA/Gastos",            total_iva_gbl,),
    ]

    for i, row_data in enumerate(cover_rows, start=1):
        for j, val in enumerate(row_data, start=1):
            c = ws_cover.cell(row=i, column=j, value=val)
            c.font      = font_data_10
            c.alignment = alin_izq
            if j > 1 and isinstance(val, (int, float)):
                c.number_format = FMT_MONEDA
                c.alignment     = alin_der
        ws_cover.row_dimensions[i].height = 16

        if i == 1:
            ws_cover.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            ws_cover.cell(1, 1).fill = fill_azul_oscuro
            ws_cover.cell(1, 1).font = font_bold_11_w
            ws_cover.row_dimensions[1].height = 26
        elif i == 2:
            ws_cover.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
            ws_cover.cell(2, 1).fill = fill_azul_medio
            ws_cover.cell(2, 1).font = Font(name="Calibri", size=10, color="FFFFFFFF", italic=True)
        elif row_data and row_data[0] in (
            "RESUMEN POR ENTIDAD", "DETALLE DE PRÉSTAMOS",
            "CONCILIACIÓN GLOBAL",
        ):
            ws_cover.cell(i, 1).fill = fill_gris_header
            ws_cover.cell(i, 1).font = font_bold_10
        elif row_data and len(row_data) >= 2 and isinstance(row_data[1], float):
            for j in range(1, 4):
                ws_cover.cell(i, j).border = border_thin

    ws_cover.sheet_view.zoomScale = 90

    ruta_salida = BASE_DIR / "Reporte_Prestamos_Auditoria.xlsx"
    wb.save(str(ruta_salida))
    return str(ruta_salida)


# --- Devengamientos de fin de mes ---

BUSQUEDAS_CUENTA_DEVENGAMIENTO: dict[str, list[str]] = {
    "sueldos_jornales": ["sueldos y jornales"],
    "sueldos_pagar": ["sueldos y jornales a pagar", "sueldos a pagar"],
    "cargas_sociales": ["cargas sociales", "contribuciones patronales", "contribuciones y aportes"],
    "contribuciones_pagar": ["contribuciones seg social a pagar", "contribuciones patronales a pagar"],
    "aportes_pagar": ["aportes seg social a pagar", "retenciones aportes empleado"],
    "iva_debito": ["iva debito fiscal"],
    "iva_credito": ["iva credito fiscal"],
    "iva_pagar": ["iva a pagar"],
    "iva_saldo_favor": ["saldo tecnico iva a favor", "saldo libre disponibilidad iva"],
    "iibb_gasto": [
        "impuestos sobre los ingresos brutos",
        "impuesto ingresos brutos",
        "ingresos brutos gasto",
        "iibb gasto",
    ],
    "iibb_pagar": [
        "impuesto a los ingresos brutos a pagar",
        "iibb a pagar",
        "ingresos brutos a pagar",
        "impuesto ingresos brutos a pagar",
    ],
    "iibb_retenciones_banco": [
        "retenciones imp sobre los ing brutos",
        "retenciones ingresos brutos",
        "retenciones iibb",
        "retenciones imp ingresos brutos",
        "retenciones sobre ingresos brutos",
        "retenciones iibb banco",
        "retenciones ingresos brutos banco",
    ],
    "iibb_retenciones_agentes": [
        "retenc iibb agentes",
        "retenciones agentes ingresos brutos",
        "percepcion iibb agentes",
        "retenciones iibb tarjetas",
        "ret iibb agentes",
        "retenc agentes iibb",
    ],
    "iibb_percepciones": [
        "percepcion ingresos brutos pba",
        "percepcion iibb pba",
        "percepcion iibb",
        "percepcion ingresos brutos",
        "percepciones iibb",
        "percepciones iibb pba",
    ],
    "iva_retenciones": ["retenciones iva", "retenciones percepciones iva", "retenciones imp valor agregado", "ret iva tarjetas"],
    "iva_percepciones": ["percepcion iva", "percepcion al valor agregado", "percepcion iva 3337", "iva percepcion"],
    "iva_saldo_tecnico": ["saldo tecnico iva a favor", "saldo tecnico iva", "iva saldo tecnico favor"],
    "iva_saldo_libre": ["saldo libre disponibilidad iva", "iva saldo libre disponibilidad", "saldo libre iva"],
    "iva_credito_105": ["iva credito fiscal 10.5", "iva cf 10.5", "credito fiscal 10.5", "iva credito 10,5"],
    "iva_credito_210": ["iva credito fiscal 21", "iva cf 21", "credito fiscal 21", "iva credito 21%"],
    "iva_credito_270": ["iva credito fiscal 27", "iva cf 27", "credito fiscal 27", "iva credito 27%"],
    "iva_debito_105": ["iva debito fiscal 10.5", "iva df 10.5", "debito fiscal 10.5", "iva debito 10,5"],
    "iva_debito_210": ["iva debito fiscal 21", "iva df 21", "debito fiscal 21", "iva debito 21%"],
    "iva_ventas_105": ["iva ventas 10.5", "iva ventas 10,5", "iva ventas 10,50"],
    "iva_ventas_210": ["iva ventas 21", "iva ventas 21%", "iva ventas 21,00"],
    "iva_credito_general": ["iva credito fiscal"],
    "iva_debito_general": ["iva debito fiscal"],
    "tsh_gasto": ["gastos de seguridad", "tasa municipal", "tsh"],
    "tsh_pagar": ["tasa de seguridad e higiene a pagar"],
}

CODIGOS_CUENTA_DEVENGAMIENTO_FALLBACK: dict[str, str] = {
    "sueldos_jornales": "42201",
    "sueldos_pagar": "21309",
    "cargas_sociales": "42203",
    "contribuciones_pagar": "21302",
    "aportes_pagar": "21303",
    "iva_debito": "21401",
    "iva_credito": "11402",
    "iva_pagar": "21405",
    "iva_saldo_favor": "11411",
    "iibb_gasto": "42405",
    "iibb_pagar": "21404",
    "iibb_retenciones_banco": "11418",
    "iibb_retenciones_agentes": "11419",
    "iibb_percepciones": "11404",
    "iva_retenciones": "11410",
    "iva_percepciones": "11403",
    "iva_saldo_tecnico": "11411",
    "iva_saldo_libre": "11412",
    "iva_credito_105": "11401",
    "iva_credito_210": "11402",
    "iva_credito_270": "11409",
    "iva_debito_105": "21402",   # IVA Débito Fiscal 10,5%
    "iva_debito_210": "21401",   # IVA Débito Fiscal 21%
    "tsh_gasto": "42401",
    "tsh_pagar": "21407",
}


@dataclass
class RenglonAsiento:
    """Línea de un asiento contable de devengamiento."""

    codigo_cuenta: str
    descripcion_cuenta: str
    debe: float = 0.0
    haber: float = 0.0
    leyenda: str = ""


@dataclass
class AsientoDevengamiento:
    """Asiento contable balanceado de devengamiento."""

    identificador: int
    concepto: str
    fecha: date
    renglones: list[RenglonAsiento] = field(default_factory=list)

    @property
    def total_debe(self) -> float:
        return round(sum(r.debe for r in self.renglones), 2)

    @property
    def total_haber(self) -> float:
        return round(sum(r.haber for r in self.renglones), 2)

    @property
    def balanceado(self) -> bool:
        return abs(self.total_debe - self.total_haber) <= 0.01


@dataclass
class ResultadoDevengamiento:
    """Conjunto de asientos generados para un mes."""

    asientos: list[AsientoDevengamiento] = field(default_factory=list)
    cuentas_resolvidas: dict[str, tuple[str, str]] = field(default_factory=dict)
    advertencias: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Persistencia en disco — biblioteca de asientos y borrador de grillas
# ---------------------------------------------------------------------------

PERSISTENCIA_VERSION = 1
PERSISTENCIA_RED_DIR = Path(r"T:\Estudio Contable")
NOMBRE_BIBLIOTECA_JSON = "biblioteca_asientos.json"
NOMBRE_BORRADOR_JSON = "borrador_actual.json"
PERSISTENCIA_LOCAL_DIR = BASE_DIR / "data" / "persistencia"


def _slug_usuario_persistencia(usuario: str | None) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", str(usuario or "").strip().lower()).strip("_")
    return slug[:40]


def _nombre_archivo_persistencia(nombre_base: str, usuario: str | None = None) -> str:
    """biblioteca_asientos.json → biblioteca_asientos__recepcion.json"""
    slug = _slug_usuario_persistencia(usuario)
    if not slug:
        return nombre_base
    p = Path(nombre_base)
    return f"{p.stem}__{slug}{p.suffix}"


def _directorio_persistencia_escribible() -> Path:
    """Prioriza T:\\Estudio Contable en oficina Windows; en Cloud usa data/persistencia."""
    import os

    candidatos: list[Path] = []
    if os.name == "nt":
        candidatos.append(PERSISTENCIA_RED_DIR)
    candidatos.extend((PERSISTENCIA_LOCAL_DIR, BASE_DIR))
    for carpeta in candidatos:
        try:
            carpeta.mkdir(parents=True, exist_ok=True)
            probe = carpeta / ".persistencia_probe"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return carpeta
        except OSError:
            continue
    return BASE_DIR


def _ruta_persistencia_lectura(nombre: str) -> Path | None:
    import os

    carpetas = [PERSISTENCIA_LOCAL_DIR, BASE_DIR]
    if os.name == "nt":
        carpetas.insert(0, PERSISTENCIA_RED_DIR)
    for carpeta in carpetas:
        ruta = carpeta / nombre
        if ruta.is_file():
            return ruta
    return None


def _ruta_persistencia_escritura(nombre: str) -> Path:
    return _directorio_persistencia_escribible() / nombre


def _escribir_json_atomico(ruta: Path, payload: dict) -> Path:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    tmp = ruta.with_suffix(ruta.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2, default=str)
    tmp.replace(ruta)
    return ruta


def _leer_json_persistencia(ruta: Path) -> dict | None:
    try:
        with ruta.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else None
    except (OSError, json.JSONDecodeError, ValueError):
        return None


def _fecha_a_iso(val) -> str:
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, datetime):
        return val.date().isoformat()
    return str(val or "")


def _fecha_desde_iso(val: str) -> date:
    texto = str(val or "").strip()
    if not texto:
        return date.today()
    if "T" in texto:
        texto = texto.split("T", 1)[0]
    if " " in texto and len(texto) > 10:
        texto = texto.split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto[:10] if fmt == "%Y-%m-%d" else texto, fmt).date()
        except ValueError:
            continue
    try:
        ts = pd.to_datetime(texto, dayfirst=True, errors="coerce")
        if pd.notna(ts):
            return ts.date()
    except Exception:
        pass
    return date.today()


def asiento_devengamiento_a_dict(asiento: AsientoDevengamiento) -> dict:
    return {
        "identificador": asiento.identificador,
        "concepto": asiento.concepto,
        "fecha": _fecha_a_iso(asiento.fecha),
        "renglones": [
            {
                "codigo_cuenta": r.codigo_cuenta,
                "descripcion_cuenta": r.descripcion_cuenta,
                "debe": round(float(r.debe or 0), 2),
                "haber": round(float(r.haber or 0), 2),
                "leyenda": r.leyenda,
            }
            for r in asiento.renglones
        ],
        "tipo": getattr(asiento, "tipo", None),
        "periodo": getattr(asiento, "periodo", None),
        "fecha_tango_str": getattr(asiento, "fecha_tango_str", None),
        "advertencias": getattr(asiento, "advertencias", None),
        "roles_renglones": getattr(asiento, "_roles_renglones", None),
        "resumen_analitico": getattr(asiento, "_resumen_analitico", None),
    }


def asiento_devengamiento_desde_dict(data: dict) -> AsientoDevengamiento:
    renglones = [
        RenglonAsiento(
            codigo_cuenta=str(r.get("codigo_cuenta", "")),
            descripcion_cuenta=str(r.get("descripcion_cuenta", "")),
            debe=round(float(r.get("debe") or 0), 2),
            haber=round(float(r.get("haber") or 0), 2),
            leyenda=str(r.get("leyenda", "")),
        )
        for r in (data.get("renglones") or [])
    ]
    asiento = AsientoDevengamiento(
        identificador=data.get("identificador", 1),
        concepto=str(data.get("concepto", "")),
        fecha=_fecha_desde_iso(str(data.get("fecha", ""))),
        renglones=renglones,
    )
    for attr, key in (
        ("tipo", "tipo"),
        ("periodo", "periodo"),
        ("fecha_tango_str", "fecha_tango_str"),
        ("advertencias", "advertencias"),
    ):
        val = data.get(key)
        if val is not None:
            setattr(asiento, attr, val)
    roles = data.get("roles_renglones")
    if roles is not None:
        asiento._roles_renglones = roles  # type: ignore[attr-defined]
    resumen = data.get("resumen_analitico")
    if resumen is not None:
        asiento._resumen_analitico = resumen  # type: ignore[attr-defined]
    return asiento


def _entrada_biblioteca_a_dict(entrada: dict) -> dict:
    serializada = dict(entrada)
    asientos = entrada.get("asientos") or []
    serializada["asientos"] = [
        a if isinstance(a, dict) else asiento_devengamiento_a_dict(a)
        for a in asientos
    ]
    po = serializada.get("periodo_orden")
    if isinstance(po, tuple):
        serializada["periodo_orden"] = list(po)
    return serializada


def _entrada_biblioteca_desde_dict(data: dict) -> dict:
    entrada = dict(data)
    entrada["asientos"] = [
        asiento_devengamiento_desde_dict(a) if isinstance(a, dict) else a
        for a in (data.get("asientos") or [])
    ]
    po = entrada.get("periodo_orden")
    if isinstance(po, list) and len(po) == 2:
        entrada["periodo_orden"] = (int(po[0]), int(po[1]))
    return entrada


def guardar_biblioteca_persistida(
    *,
    biblioteca_asientos: list[dict],
    biblioteca_bancos: list[dict],
    periodos_procesados: list[dict],
    periodos_bancos_procesados: list[dict],
    usuario: str | None = None,
) -> Path:
    """Guarda la biblioteca completa en JSON físico (red T: o carpeta del proyecto)."""
    payload = {
        "version": PERSISTENCIA_VERSION,
        "actualizado": datetime.now().isoformat(timespec="seconds"),
        "usuario": usuario or "",
        "biblioteca_asientos": [_entrada_biblioteca_a_dict(e) for e in biblioteca_asientos],
        "biblioteca_bancos": [_entrada_biblioteca_a_dict(e) for e in biblioteca_bancos],
        "periodos_procesados": list(periodos_procesados),
        "periodos_bancos_procesados": list(periodos_bancos_procesados),
    }
    nombre = _nombre_archivo_persistencia(NOMBRE_BIBLIOTECA_JSON, usuario)
    return _escribir_json_atomico(_ruta_persistencia_escritura(nombre), payload)


def cargar_biblioteca_persistida(usuario: str | None = None) -> dict | None:
    """Lee biblioteca desde disco si existe (por usuario de oficina)."""
    nombre = _nombre_archivo_persistencia(NOMBRE_BIBLIOTECA_JSON, usuario)
    ruta = _ruta_persistencia_lectura(nombre)
    # Compat: si el usuario no tiene archivo propio, no heredar el global automáticamente
    # (evita pisar trabajo entre personas). Solo fallback sin usuario.
    if ruta is None and not _slug_usuario_persistencia(usuario):
        ruta = _ruta_persistencia_lectura(NOMBRE_BIBLIOTECA_JSON)
    if ruta is None:
        return None
    data = _leer_json_persistencia(ruta)
    if not data:
        return None
    return {
        "biblioteca_asientos": [
            _entrada_biblioteca_desde_dict(e)
            for e in (data.get("biblioteca_asientos") or [])
        ],
        "biblioteca_bancos": [
            _entrada_biblioteca_desde_dict(e)
            for e in (data.get("biblioteca_bancos") or [])
        ],
        "periodos_procesados": list(data.get("periodos_procesados") or []),
        "periodos_bancos_procesados": list(data.get("periodos_bancos_procesados") or []),
        "ruta": str(ruta),
    }


def guardar_borrador_grilla_persistido(borrador: dict, usuario: str | None = None) -> Path:
    """Auto-guardado de emergencia de la grilla en edición (por usuario)."""
    payload = {
        "version": PERSISTENCIA_VERSION,
        "actualizado": datetime.now().isoformat(timespec="seconds"),
        "usuario": usuario or "",
        "modulo": borrador.get("modulo"),
        "slug": borrador.get("slug"),
        "sociedad_id": borrador.get("sociedad_id"),
        "contexto": borrador.get("contexto"),
        "grilla_preview": borrador.get("grilla_preview") or [],
        "asientos": [
            a if isinstance(a, dict) else asiento_devengamiento_a_dict(a)
            for a in (borrador.get("asientos") or [])
        ],
        "resumen_analitico": borrador.get("resumen_analitico"),
        "columna_auditoria": borrador.get("columna_auditoria"),
        "auto_fp": borrador.get("auto_fp"),
    }
    nombre = _nombre_archivo_persistencia(NOMBRE_BORRADOR_JSON, usuario)
    return _escribir_json_atomico(_ruta_persistencia_escritura(nombre), payload)


def cargar_borrador_grilla_persistido(usuario: str | None = None) -> dict | None:
    nombre = _nombre_archivo_persistencia(NOMBRE_BORRADOR_JSON, usuario)
    ruta = _ruta_persistencia_lectura(nombre)
    if ruta is None and not _slug_usuario_persistencia(usuario):
        ruta = _ruta_persistencia_lectura(NOMBRE_BORRADOR_JSON)
    if ruta is None:
        return None
    data = _leer_json_persistencia(ruta)
    if not data:
        return None
    data["asientos"] = [
        asiento_devengamiento_desde_dict(a) if isinstance(a, dict) else a
        for a in (data.get("asientos") or [])
    ]
    data["ruta"] = str(ruta)
    return data


def limpiar_borrador_grilla_persistido(usuario: str | None = None) -> None:
    import os

    nombre = _nombre_archivo_persistencia(NOMBRE_BORRADOR_JSON, usuario)
    carpetas = [PERSISTENCIA_LOCAL_DIR, BASE_DIR]
    if os.name == "nt":
        carpetas.insert(0, PERSISTENCIA_RED_DIR)
    for carpeta in carpetas:
        ruta = carpeta / nombre
        try:
            ruta.unlink(missing_ok=True)
        except OSError:
            pass


def ruta_biblioteca_persistida_activa(usuario: str | None = None) -> str:
    """Ruta efectiva del JSON de biblioteca (para diagnóstico en UI)."""
    nombre = _nombre_archivo_persistencia(NOMBRE_BIBLIOTECA_JSON, usuario)
    lectura = _ruta_persistencia_lectura(nombre)
    if lectura is not None:
        return str(lectura)
    return str(_ruta_persistencia_escritura(nombre))


def _redondear_importe(valor: float) -> float:
    return round(float(valor or 0), 2)


def _agregar_renglon(
    renglones: list[RenglonAsiento],
    codigo: str,
    descripcion: str,
    debe: float = 0.0,
    haber: float = 0.0,
    leyenda: str = "",
) -> None:
    debe_r = _redondear_importe(debe)
    haber_r = _redondear_importe(haber)
    if debe_r <= 0 and haber_r <= 0:
        return
    renglones.append(
        RenglonAsiento(
            codigo_cuenta=codigo,
            descripcion_cuenta=descripcion,
            debe=debe_r,
            haber=haber_r,
            leyenda=leyenda,
        )
    )


def buscar_cuenta_devengamiento(
    clave_concepto: str,
    plan_cuentas: pd.DataFrame,
    resoluciones: dict[str, tuple[str, str]],
) -> tuple[str, str]:
    """Resuelve código y descripción de cuenta por concepto con búsqueda fuzzy."""
    if clave_concepto in resoluciones:
        return resoluciones[clave_concepto]

    terminos = BUSQUEDAS_CUENTA_DEVENGAMIENTO.get(clave_concepto, [clave_concepto])
    codigo_fallback = CODIGOS_CUENTA_DEVENGAMIENTO_FALLBACK.get(clave_concepto, "")
    mejor_codigo = codigo_fallback
    mejor_desc = clave_concepto
    mejor_score = 0

    if not plan_cuentas.empty:
        opciones = plan_cuentas["descripcion_norm"].tolist()
        codigos = plan_cuentas["codigo"].tolist()
        descripciones = plan_cuentas["descripcion"].tolist()
        for termino in terminos:
            term_norm = _normalizar_texto(termino)
            match = process.extractOne(term_norm, opciones, scorer=fuzz.partial_ratio)
            if match and match[1] > mejor_score:
                idx = opciones.index(match[0])
                mejor_score = match[1]
                mejor_codigo = codigos[idx]
                mejor_desc = descripciones[idx]

    if mejor_score < 65 and codigo_fallback:
        mejor_codigo = codigo_fallback
        if not plan_cuentas.empty:
            fila = plan_cuentas[plan_cuentas["codigo"] == codigo_fallback]
            if not fila.empty:
                mejor_desc = str(fila.iloc[0]["descripcion"])

    resoluciones[clave_concepto] = (mejor_codigo, mejor_desc)
    return mejor_codigo, mejor_desc


def _cuenta_tiene_alicuota_especifica(descripcion: str) -> bool:
    """True si la descripción parece una subcuenta por alícuota (10,5 / 21 / 27 %)."""
    import re as _re

    d = _normalizar_texto(descripcion)
    if _re.search(r"\d+[.,]?\d*\s*%", d):
        return True
    return bool(_re.search(r"\b(10[,.]?5|21|27)\b", d))


def buscar_cuenta_iva_semantica(
    clave_concepto: str,
    plan_cuentas: pd.DataFrame,
    resoluciones: dict[str, tuple[str, str]],
    *,
    solo_cuenta_generica: bool = False,
) -> tuple[str, str]:
    """
    Resuelve cuenta IVA únicamente por búsqueda semántica en el plan del cliente.
    No utiliza códigos fallback hardcodeados.
    """
    if clave_concepto in resoluciones:
        return resoluciones[clave_concepto]

    terminos = BUSQUEDAS_CUENTA_DEVENGAMIENTO.get(clave_concepto, [clave_concepto])
    plan = plan_cuentas
    if "imputable" in plan.columns:
        plan = plan[plan["imputable"] == True].copy()

    mejor_codigo = ""
    mejor_desc = clave_concepto
    mejor_score = 0

    if not plan.empty and "descripcion_norm" in plan.columns:
        opciones = plan["descripcion_norm"].tolist()
        codigos = plan["codigo"].tolist()
        descripciones = plan["descripcion"].tolist()
        for termino in terminos:
            term_norm = _normalizar_texto(termino)
            for idx, opcion in enumerate(opciones):
                if solo_cuenta_generica and _cuenta_tiene_alicuota_especifica(descripciones[idx]):
                    continue
                score = fuzz.partial_ratio(term_norm, opcion)
                if score > mejor_score:
                    mejor_score = score
                    mejor_codigo = str(codigos[idx])
                    mejor_desc = str(descripciones[idx])

    if mejor_score >= 65 and mejor_codigo:
        resoluciones[clave_concepto] = (mejor_codigo, mejor_desc)
        return mejor_codigo, mejor_desc

    resoluciones[clave_concepto] = ("", clave_concepto)
    return "", clave_concepto


def _validar_balance_asiento(asiento: AsientoDevengamiento) -> None:
    diff = _redondear_importe(asiento.total_debe - asiento.total_haber)
    if abs(diff) > 0.01 and asiento.renglones:
        ultimo = asiento.renglones[-1]
        if diff > 0:
            ultimo.haber = _redondear_importe(ultimo.haber + diff)
        else:
            ultimo.debe = _redondear_importe(ultimo.debe + abs(diff))


def _fecha_fin_mes(anio: int, mes: int) -> date:
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    return date(anio, mes, ultimo_dia)


def generar_asientos_devengamiento(
    cliente: dict,
    datos_mes: dict,
    plan_cuentas_df: pd.DataFrame,
) -> ResultadoDevengamiento:
    """
    Genera 4 asientos balanceados (Sueldos, IVA, IIBB, TSH) para el cierre mensual.
    Mapea conceptos al plan de cuentas del cliente mediante búsqueda fuzzy.
    """
    resultado = ResultadoDevengamiento()
    resoluciones: dict[str, tuple[str, str]] = {}
    anio = int(datos_mes.get("anio", date.today().year))
    mes = int(datos_mes.get("mes", date.today().month))
    fecha_asiento = _fecha_fin_mes(anio, mes)
    identificador = 1

    def cuenta(clave: str) -> tuple[str, str]:
        return buscar_cuenta_devengamiento(clave, plan_cuentas_df, resoluciones)

    sueldos = datos_mes.get("sueldos", {})
    bruto = _redondear_importe(sueldos.get("sueldo_bruto", 0))
    patronales = _redondear_importe(sueldos.get("contribuciones_patronales", 0))
    retenciones = _redondear_importe(sueldos.get("retenciones_aportes", 0))

    if bruto > 0 or patronales > 0 or retenciones > 0:
        renglones: list[RenglonAsiento] = []
        c_sueldos, d_sueldos = cuenta("sueldos_jornales")
        c_cargas, d_cargas = cuenta("cargas_sociales")
        c_pagar, d_pagar = cuenta("sueldos_pagar")
        c_contrib, d_contrib = cuenta("contribuciones_pagar")
        c_aportes, d_aportes = cuenta("aportes_pagar")

        neto_pagar = _redondear_importe(bruto - retenciones)

        if bruto > 0:
            _agregar_renglon(renglones, c_sueldos, d_sueldos, debe=bruto, leyenda="Devengamiento sueldos")
        if patronales > 0:
            _agregar_renglon(renglones, c_cargas, d_cargas, debe=patronales, leyenda="Contribuciones patronales")
        if neto_pagar > 0:
            _agregar_renglon(renglones, c_pagar, d_pagar, haber=neto_pagar, leyenda="Sueldos a pagar")
        if patronales > 0:
            _agregar_renglon(renglones, c_contrib, d_contrib, haber=patronales, leyenda="Contribuciones a pagar")
        if retenciones > 0:
            _agregar_renglon(renglones, c_aportes, d_aportes, haber=retenciones, leyenda="Retenciones y aportes empleado")

        asiento = AsientoDevengamiento(identificador, "Sueldos", fecha_asiento, renglones)
        _validar_balance_asiento(asiento)
        if asiento.renglones:
            resultado.asientos.append(asiento)
            identificador += 1

    iva = datos_mes.get("iva", {})
    debito_iva = _redondear_importe(iva.get("debito_fiscal", 0))
    credito_iva = _redondear_importe(iva.get("credito_fiscal", 0))
    saldo_ant = _redondear_importe(iva.get("saldo_favor_anterior", 0))

    if debito_iva > 0 or credito_iva > 0 or saldo_ant > 0:
        renglones_iva: list[RenglonAsiento] = []
        c_deb, d_deb = cuenta("iva_debito")
        c_cred, d_cred = cuenta("iva_credito")
        c_iva_pag, d_iva_pag = cuenta("iva_pagar")
        c_iva_sf, d_iva_sf = cuenta("iva_saldo_favor")

        posicion_bruta = _redondear_importe(debito_iva - credito_iva)

        if debito_iva > 0:
            _agregar_renglon(renglones_iva, c_deb, d_deb, debe=debito_iva, leyenda="IVA débito fiscal")
        if credito_iva > 0:
            _agregar_renglon(renglones_iva, c_cred, d_cred, haber=credito_iva, leyenda="IVA crédito fiscal")
        if posicion_bruta > 0:
            _agregar_renglon(
                renglones_iva, c_iva_pag, d_iva_pag, haber=posicion_bruta, leyenda="IVA a pagar"
            )
        elif posicion_bruta < 0:
            _agregar_renglon(
                renglones_iva,
                c_iva_sf,
                d_iva_sf,
                debe=abs(posicion_bruta),
                leyenda="Saldo técnico IVA a favor",
            )
        if saldo_ant > 0:
            _agregar_renglon(
                renglones_iva, c_iva_sf, d_iva_sf, debe=saldo_ant, leyenda="Aplicación saldo a favor anterior"
            )
            _agregar_renglon(
                renglones_iva, c_iva_pag, d_iva_pag, haber=saldo_ant, leyenda="Compensación saldo a favor anterior"
            )

        asiento_iva = AsientoDevengamiento(identificador, "IVA", fecha_asiento, renglones_iva)
        _validar_balance_asiento(asiento_iva)
        if asiento_iva.renglones:
            resultado.asientos.append(asiento_iva)
            identificador += 1

    iibb = datos_mes.get("iibb", {})
    imp_iibb = _redondear_importe(iibb.get("impuesto_determinado", 0))
    if imp_iibb > 0:
        renglones_iibb: list[RenglonAsiento] = []
        c_gasto, d_gasto = cuenta("iibb_gasto")
        c_pagar, d_pagar = cuenta("iibb_pagar")
        _agregar_renglon(renglones_iibb, c_gasto, d_gasto, debe=imp_iibb, leyenda="Devengamiento IIBB")
        _agregar_renglon(renglones_iibb, c_pagar, d_pagar, haber=imp_iibb, leyenda="IIBB a pagar")
        asiento_iibb = AsientoDevengamiento(identificador, "IIBB", fecha_asiento, renglones_iibb)
        resultado.asientos.append(asiento_iibb)
        identificador += 1

    tsh = datos_mes.get("tsh", {})
    tasa_tsh = _redondear_importe(tsh.get("tasa_determinada", 0))
    if tasa_tsh > 0:
        renglones_tsh: list[RenglonAsiento] = []
        c_gasto, d_gasto = cuenta("tsh_gasto")
        c_pagar, d_pagar = cuenta("tsh_pagar")
        _agregar_renglon(renglones_tsh, c_gasto, d_gasto, debe=tasa_tsh, leyenda="Devengamiento TSH")
        _agregar_renglon(renglones_tsh, c_pagar, d_pagar, haber=tasa_tsh, leyenda="TSH a pagar")
        asiento_tsh = AsientoDevengamiento(identificador, "TSH", fecha_asiento, renglones_tsh)
        resultado.asientos.append(asiento_tsh)

    resultado.cuentas_resolvidas = dict(resoluciones)
    for asiento in resultado.asientos:
        if not asiento.balanceado:
            resultado.advertencias.append(
                f"Asiento {asiento.concepto}: diferencia ${asiento.total_debe - asiento.total_haber:.2f}"
            )

    return resultado


class ExportacionTangoError(ValueError):
    """Error estructurado de validación pre-exportación a Tango Gestión."""

    def __init__(self, mensaje: str, errores: list[dict] | None = None):
        super().__init__(mensaje)
        self.errores = errores or []


def _normalizar_codigo_cuenta_export(codigo) -> str:
    s = str(codigo or "").strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def parsear_fecha_export_tango(val) -> date:
    """Normaliza fecha para export Tango (Excel nativo / TXT AAAAMMDD)."""
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    if val is None:
        return date.today()
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return date.today()
    if "T" in s:
        s = s.split("T", 1)[0]
    if " " in s and len(s) > 10:
        s = s.split(" ", 1)[0]
    if re.fullmatch(r"\d{8}", s):
        return datetime.strptime(s, "%Y%m%d").date()
    m = re.match(r"^(0?[1-9]|[12][0-9]|3[01])/(0?[1-9]|1[0-2])/(\d{4})$", s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(ts):
            return ts.date()
    except Exception:
        pass
    return date.today()


def formatear_fecha_tango_export_txt(val) -> str:
    """Fecha cruda AAAAMMDD para layouts TXT de intercambio Tango."""
    return parsear_fecha_export_tango(val).strftime("%Y%m%d")


def _plan_df_export_normalizado(plan_cuentas: pd.DataFrame | None) -> pd.DataFrame:
    if plan_cuentas is None or plan_cuentas.empty:
        return pd.DataFrame()
    try:
        return _normalizar_plan_cuentas_df(plan_cuentas.copy())
    except Exception:
        df = plan_cuentas.copy()
        if "codigo" not in df.columns:
            for col in df.columns:
                if "codigo" in str(col).lower():
                    df = df.rename(columns={col: "codigo"})
                    break
        if "codigo" in df.columns:
            df["codigo"] = df["codigo"].astype(str).str.strip()
        return df


def es_cuenta_no_imputable_tango(codigo: str, plan_df: pd.DataFrame) -> tuple[bool, str]:
    """
    True si la cuenta es Rubro/Madre o inválida para imputar en Tango.
    Retorna (es_invalida, motivo).
    """
    cod = _normalizar_codigo_cuenta_export(codigo)
    if not cod or cod == "99999":
        return True, "Cuenta sin asignar o código 99999"
    if plan_df is None or plan_df.empty:
        return False, ""
    if "codigo" not in plan_df.columns:
        return False, ""
    codigos = plan_df["codigo"].astype(str).str.strip()
    fila = plan_df[codigos == cod]
    if not fila.empty and "imputable" in plan_df.columns:
        imp = fila.iloc[0]["imputable"]
        if not _plan_flag_si(imp):
            desc = str(fila.iloc[0].get("descripcion", cod))
            return True, f"Cuenta Rubro/Madre no imputable en el plan ({desc})"
    for otro in codigos:
        if otro == cod:
            continue
        if otro.startswith(cod) and len(otro) > len(cod):
            return True, f"El código {cod} es prefijo de {otro} (cuenta madre / rubro)"
    return False, ""


def es_cuenta_auxiliar_sin_layout_export_tango(codigo: str, plan_df: pd.DataFrame) -> tuple[bool, str]:
    """
    True si la cuenta usa auxiliares contables en Tango y el layout Excel de importación
    no puede enviar apropiaciones (Tango rechaza: «no tiene asignado ningún tipo de auxiliar»).
    """
    cod = _normalizar_codigo_cuenta_export(codigo)
    if not cod or plan_df is None or plan_df.empty or "codigo" not in plan_df.columns:
        return False, ""
    if "usa_auxiliares" not in plan_df.columns:
        return False, ""
    codigos = plan_df["codigo"].astype(str).str.strip()
    fila = plan_df[codigos == cod]
    if fila.empty:
        return False, ""
    if not _plan_flag_si(fila.iloc[0]["usa_auxiliares"]):
        return False, ""
    desc = str(fila.iloc[0].get("descripcion", cod))
    return True, (
        f"La cuenta {cod} ({desc}) usa auxiliares contables en Tango. "
        "El importador Excel no incluye apropiaciones analíticas; Tango rechazará el asiento "
        "con «La cuenta no tiene asignado ningún tipo de auxiliar». "
        "En Tango: Plan de cuentas → asignar tipo de auxiliar y regla de apropiación al 100%, "
        "o elegí otra cuenta sin auxiliares en el selector de la grilla."
    )


def auditar_renglones_imputables_tango(
    asientos: list,
    plan_cuentas: pd.DataFrame | None,
) -> list[dict]:
    """Lista errores bloqueantes de cuentas no exportables a Tango (sin asignar / rubro)."""
    return auditar_exportacion_tango(asientos, plan_cuentas)["bloqueantes"]


def auditar_exportacion_tango(
    asientos: list,
    plan_cuentas: pd.DataFrame | None,
) -> dict[str, list[dict]]:
    """
    Auditoría pre-exportación.
    - bloqueantes: cuentas 99999, rubro/madre → impiden generar el archivo.
    - advertencias: cuentas con auxiliares contables → requieren configuración en Tango.
    """
    plan_df = _plan_df_export_normalizado(plan_cuentas)
    bloqueantes: list[dict] = []
    advertencias: list[dict] = []
    for asiento in asientos:
        id_as = getattr(asiento, "identificador", "?")
        concepto = str(getattr(asiento, "concepto", "") or "")
        periodo = str(getattr(asiento, "periodo", "") or "")
        for nro, ren in enumerate(getattr(asiento, "renglones", []) or [], start=1):
            cod = _normalizar_codigo_cuenta_export(ren.codigo_cuenta)
            base = {
                "identificador": id_as,
                "concepto": concepto,
                "periodo": periodo,
                "nro": nro,
                "codigo": cod,
                "descripcion": str(getattr(ren, "descripcion_cuenta", "") or ""),
            }
            usa_aux, mot_aux = es_cuenta_auxiliar_sin_layout_export_tango(cod, plan_df)
            if usa_aux:
                advertencias.append({**base, "motivo": mot_aux, "tipo": "auxiliar"})
                continue
            invalida, motivo = es_cuenta_no_imputable_tango(cod, plan_df)
            if invalida:
                bloqueantes.append({**base, "motivo": motivo, "tipo": "imputable"})
    return {"bloqueantes": bloqueantes, "advertencias": advertencias}


def resumir_informe_export_tango(informe: dict[str, list[dict]]) -> list[dict]:
    """Agrupa bloqueantes/advertencias por código de cuenta (evita repetir 158 líneas)."""
    resumen: list[dict] = []
    for tipo in ("bloqueantes", "advertencias"):
        vistos: dict[str, dict] = {}
        for err in informe.get(tipo) or []:
            cod = str(err.get("codigo", ""))
            clave = f"{tipo}:{cod}"
            if clave not in vistos:
                vistos[clave] = {
                    "tipo": tipo,
                    "codigo": cod,
                    "descripcion": err.get("descripcion", ""),
                    "motivo": err.get("motivo", ""),
                    "ocurrencias": 0,
                    "asientos": set(),
                }
            vistos[clave]["ocurrencias"] += 1
            vistos[clave]["asientos"].add(str(err.get("identificador", "?")))
        for item in vistos.values():
            item["asientos"] = sorted(item["asientos"], key=lambda x: (len(x), x))
            resumen.append(item)
    return resumen


def balancear_asiento_export_tango(
    asiento: AsientoDevengamiento,
    *,
    tolerancia_centavos: float = 5.0,
) -> AsientoDevengamiento:
    """Ajusta diferencias de redondeo ≤ tolerancia sobre el renglón de mayor importe."""
    diff = round(asiento.total_debe - asiento.total_haber, 2)
    if abs(diff) <= 0.01:
        return asiento
    if abs(diff) > tolerancia_centavos:
        raise ExportacionTangoError(
            f"Asiento {asiento.identificador} ({asiento.concepto}): "
            f"diferencia ${diff:,.2f} supera el ajuste automático de ${tolerancia_centavos:,.2f}."
        )
    if diff > 0:
        candidatos = [r for r in asiento.renglones if r.debe > 0]
        if not candidatos:
            raise ExportacionTangoError(
                f"Asiento {asiento.identificador}: Debe > Haber por ${diff:,.2f} sin renglón Debe ajustable."
            )
        objetivo = max(candidatos, key=lambda r: r.debe)
        objetivo.debe = round(objetivo.debe - diff, 2)
    else:
        candidatos = [r for r in asiento.renglones if r.haber > 0]
        if not candidatos:
            raise ExportacionTangoError(
                f"Asiento {asiento.identificador}: Haber > Debe por ${abs(diff):,.2f} sin renglón Haber ajustable."
            )
        objetivo = max(candidatos, key=lambda r: r.haber)
        objetivo.haber = round(objetivo.haber - abs(diff), 2)
    return asiento


def validar_asientos_lista_para_export(asientos: list) -> None:
    """Verifica partida doble al centavo para cada asiento de la lista."""
    if not asientos:
        raise ExportacionTangoError("No hay asientos para exportar.")
    desbalanceados = [
        f"#{getattr(a, 'identificador', '?')} {a.concepto} "
        f"(Debe ${a.total_debe:,.2f} ≠ Haber ${a.total_haber:,.2f})"
        for a in asientos
        if not a.balanceado
    ]
    if desbalanceados:
        raise ExportacionTangoError(
            "Los asientos no están balanceados al centavo: " + "; ".join(desbalanceados)
        )


def preparar_asientos_export_tango(
    asientos: list,
    plan_cuentas: pd.DataFrame | None = None,
    *,
    ajustar_centavos: bool = True,
    tolerancia_centavos: float = 5.0,
    validar_imputables: bool = True,
) -> list:
    """
    Pipeline pre-exportación: normaliza fechas/códigos, balancea centavos y audita imputables.
  La UI sigue mostrando DD/MM/YYYY; el Excel recibe objetos date nativos.
    """
    preparados: list = []
    for original in asientos:
        asiento = deepcopy(original)
        fecha_obj = parsear_fecha_export_tango(
            getattr(asiento, "fecha_tango_str", None) or asiento.fecha
        )
        asiento.fecha = fecha_obj
        asiento.fecha_tango_str = fecha_obj.strftime("%d/%m/%Y")  # type: ignore[attr-defined]
        for ren in asiento.renglones:
            ren.codigo_cuenta = _normalizar_codigo_cuenta_export(ren.codigo_cuenta)
            ren.debe = round(float(ren.debe or 0), 2)
            ren.haber = round(float(ren.haber or 0), 2)
        if ajustar_centavos:
            balancear_asiento_export_tango(asiento, tolerancia_centavos=tolerancia_centavos)
        preparados.append(asiento)
    validar_asientos_lista_para_export(preparados)
    if validar_imputables:
        errores = auditar_renglones_imputables_tango(preparados, plan_cuentas)
        if errores:
            raise ExportacionTangoError(
                f"Se detectaron {len(errores)} renglón/es con cuentas no imputables (Rubro/Madre). "
                "Corregilos en la grilla antes de exportar.",
                errores=errores,
            )
    return preparados


def validar_devengamientos_para_export(resultado_dev: ResultadoDevengamiento) -> None:
    """Verifica partida doble al centavo antes de exportar a Tango."""
    if not resultado_dev.asientos:
        raise ValueError("No hay asientos para exportar.")
    desbalanceados = [
        f"{a.concepto} (Debe ${a.total_debe:,.2f} ≠ Haber ${a.total_haber:,.2f})"
        for a in resultado_dev.asientos
        if not a.balanceado
    ]
    if desbalanceados:
        raise ValueError(
            "Los asientos no están balanceados al centavo: " + "; ".join(desbalanceados)
        )
    total_debe = round(sum(a.total_debe for a in resultado_dev.asientos), 2)
    total_haber = round(sum(a.total_haber for a in resultado_dev.asientos), 2)
    if abs(total_debe - total_haber) > 0.01:
        raise ValueError(
            f"Partida doble global desbalanceada: Debe ${total_debe:,.2f} ≠ Haber ${total_haber:,.2f}"
        )


def generar_txt_devengamientos(
    resultado_dev: ResultadoDevengamiento,
    tipo_asiento: str = "DEVENG",
    concepto_base: str = "Devengamiento fin de mes",
) -> str:
    """Exporta asientos de devengamiento en formato TXT Tango (tab-separated)."""
    validar_devengamientos_para_export(resultado_dev)
    lineas = [
        "[ASIENTOS_CONTABLES]",
        "IDENTIFICADOR\tFECHA DEL ASIENTO\tCLASE DEL ASIENTO\tTIPO DE ASIENTO\t"
        "ESTADO DEL ASIENTO\tMONEDA DEL ASIENTO\tCONCEPTO\tOBSERVACIONES",
    ]
    cabeceras: list[str] = []
    renglones_txt: list[str] = []

    for asiento in resultado_dev.asientos:
        fecha_str = formatear_fecha_tango_export_txt(asiento.fecha)
        id_str = str(asiento.identificador)
        concepto = f"{concepto_base} - {asiento.concepto}"
        cabeceras.append(
            f"{id_str}\t{fecha_str}\tBASICO\t{tipo_asiento}\tINGRESADO\t"
            f"PES\t{concepto}\tDevengamiento automático"
        )
        for ren in asiento.renglones:
            debe_str = f"{ren.debe:.2f}" if ren.debe > 0 else ""
            haber_str = f"{ren.haber:.2f}" if ren.haber > 0 else ""
            renglones_txt.append(
                f"{id_str}\t{ren.codigo_cuenta}\t{debe_str}\t{haber_str}\t{fecha_str}\t{ren.leyenda}"
            )

    lineas.extend(cabeceras)
    lineas.append("[RENGLONES]")
    lineas.append("IDENTIFICADOR\tCODIGO DE CUENTA\tIMPORTE DEBE\tIMPORTE HABER\tFECHA ORIGEN\tLEYENDA")
    lineas.extend(renglones_txt)
    return "\r\n".join(lineas) + "\r\n"


def devengamientos_a_dataframe(resultado_dev: ResultadoDevengamiento) -> pd.DataFrame:
    """Tabla plana de renglones para vista previa en Streamlit."""
    filas = []
    for asiento in resultado_dev.asientos:
        for ren in asiento.renglones:
            filas.append(
                {
                    "Asiento": asiento.identificador,
                    "Concepto": asiento.concepto,
                    "Fecha": asiento.fecha.strftime("%d/%m/%Y"),
                    "Código": ren.codigo_cuenta,
                    "Cuenta": ren.descripcion_cuenta,
                    "Debe": ren.debe if ren.debe > 0 else None,
                    "Haber": ren.haber if ren.haber > 0 else None,
                    "Leyenda": ren.leyenda,
                }
            )
    return pd.DataFrame(filas)


def generar_excel_asientos_devengamiento(
    resultado_dev: ResultadoDevengamiento,
    tipo_asiento: str = "DEVENG",
    ruta_plantilla: str | Path | None = None,
) -> bytes:
    """
    Exporta asientos de devengamiento clonando asientos contables.xlsx
    con columnas exactas para importación Tango.
    """
    validar_devengamientos_para_export(resultado_dev)
    estructura = leer_estructura_plantilla_asientos(ruta_plantilla)
    origen = Path(ruta_plantilla) if ruta_plantilla else _resolver_ruta_plantilla_asientos()
    destino_temp = BASE_DIR / "_temp_asientos_devengamiento.xlsx"
    shutil.copy2(origen, destino_temp)
    wb = openpyxl.load_workbook(destino_temp)

    ws_as = wb[HOJA_ASIENTOS]
    ws_ren = wb[HOJA_RENGLONES]
    enc_as = estructura[HOJA_ASIENTOS]
    enc_ren = estructura[HOJA_RENGLONES]

    mapa_as = {_normalizar_texto(h): idx + 1 for idx, h in enumerate(enc_as)}
    mapa_ren = {_normalizar_texto(h): idx + 1 for idx, h in enumerate(enc_ren)}

    def _col(mapa: dict[str, int], *claves: str, default: int | None = None) -> int | None:
        for clave in claves:
            norm = _normalizar_texto(clave)
            if norm in mapa:
                return mapa[norm]
        return default

    fila_as = ws_as.max_row + 1 if ws_as.max_row > 1 else 2
    fila_ren = ws_ren.max_row + 1 if ws_ren.max_row > 1 else 2

    for asiento in resultado_dev.asientos:
        valores_as = {
            _col(mapa_as, "identificador"): asiento.identificador,
            _col(mapa_as, "fecha de asiento", "fecha del asiento"): asiento.fecha,
            _col(mapa_as, "clase de asiento", "clase del asiento"): "BASICO",
            _col(mapa_as, "codigo de tipo de asiento", "tipo de asiento"): tipo_asiento,
            _col(mapa_as, "estado", "estado del asiento"): "INGRESADO",
            _col(mapa_as, "codigo de moneda", "moneda del asiento"): "CORRIENTE",
        }
        for col, val in valores_as.items():
            if col:
                celda = ws_as.cell(row=fila_as, column=col, value=val)
                if isinstance(val, date):
                    celda.number_format = "DD/MM/YYYY"
        fila_as += 1

        for nro, ren in enumerate(asiento.renglones, start=1):
            valores_ren = {
                _col(mapa_ren, "identificador"): asiento.identificador,
                _col(mapa_ren, "nro", "nro."): nro,
                _col(mapa_ren, "codigo de cuenta"): ren.codigo_cuenta,
                _col(mapa_ren, "debe", "importe debe"): ren.debe if ren.debe > 0 else None,
                _col(mapa_ren, "haber", "importe haber"): ren.haber if ren.haber > 0 else None,
                _col(mapa_ren, "fecha de origen", "fecha origen"): asiento.fecha,
            }
            for col, val in valores_ren.items():
                if col:
                    celda = ws_ren.cell(row=fila_ren, column=col, value=val)
                    if isinstance(val, date):
                        celda.number_format = "DD/MM/YYYY"
                    elif isinstance(val, (int, float)):
                        celda.number_format = "#,##0.00"
            fila_ren += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    try:
        destino_temp.unlink(missing_ok=True)
    except OSError:
        pass
    return out.getvalue()


def generar_excel_devengamientos(
    resultado_dev: ResultadoDevengamiento,
    datos_mes: dict,
    nombre_cliente: str,
    ruta_plantilla: str | Path | None = None,
) -> bytes:
    """Exporta devengamientos en plantilla Tango asientos contables.xlsx."""
    _ = datos_mes, nombre_cliente
    return generar_excel_asientos_devengamiento(resultado_dev, ruta_plantilla=ruta_plantilla)


MENSAJE_NOTIFICACION_DEFAULT = (
    "¡Conciliación Anual Terminada! El sistema procesó todo el año. "
    "Los Excels y asientos de Tango ya están listos para descargar."
)
TITULO_NOTIFICACION_DEFAULT = "Estudio Contable"


def _notificar_via_powershell(titulo: str, mensaje: str) -> bool:
    """Respaldo: notificación toast nativa de Windows 10/11 vía PowerShell."""
    if sys.platform != "win32":
        return False

    titulo_seguro = titulo.replace("'", "''")
    mensaje_seguro = mensaje.replace("'", "''")
    script = f"""
[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] | Out-Null
[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom.XmlDocument, ContentType = WindowsRuntime] | Out-Null
$template = @"
<toast>
  <visual>
    <binding template="ToastText02">
      <text id="1">{titulo_seguro}</text>
      <text id="2">{mensaje_seguro}</text>
    </binding>
  </visual>
</toast>
"@
$xml = New-Object Windows.Data.Xml.Dom.XmlDocument
$xml.LoadXml($template)
$toast = [Windows.UI.Notifications.ToastNotification]::new($xml)
[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("Estudio Contable").Show($toast)
"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            check=True,
            capture_output=True,
            timeout=15,
        )
        return True
    except (subprocess.SubprocessError, OSError):
        return False


def notificar_conciliacion_completada(
    titulo: str | None = None,
    mensaje: str | None = None,
) -> bool:
    """
    Dispara una notificación push flotante de escritorio en Windows.
    Usa plyer; si falla, intenta toast nativo de PowerShell.
    """
    titulo_final = titulo or TITULO_NOTIFICACION_DEFAULT
    mensaje_final = mensaje or MENSAJE_NOTIFICACION_DEFAULT

    try:
        from plyer import notification

        notification.notify(
            title=titulo_final,
            message=mensaje_final,
            app_name="Estudio Contable",
            timeout=10,
        )
        return True
    except Exception:
        return _notificar_via_powershell(titulo_final, mensaje_final)


# ---------------------------------------------------------------------------
# Módulo DDJJ IA: AsientoIABuilder + generar_excel_ddjj_tango
# ---------------------------------------------------------------------------

_TIPO_DISPLAY: dict[str, str] = {
    "IVA": "IVA",
    "IIBB_ARBA": "IIBB",
    "IIBB_CM03": "IIBB",
    "IIBB": "IIBB",
    "CM": "CM",
    "TSH": "TISH",
    "SUELDOS": "SUELDOS",
}


class AsientoIABuilder:
    """Genera asientos de devengamiento desde datos de DDJJ con bucle de auto-revisión."""

    MAX_INTENTOS = 5

    def __init__(self, plan_cuentas: pd.DataFrame):
        self.plan = plan_cuentas
        if "imputable" in plan_cuentas.columns:
            self.plan_imputable = plan_cuentas[plan_cuentas["imputable"] == True].copy()
        else:
            self.plan_imputable = plan_cuentas.copy()
        self._resoluciones: dict[str, str] = {}

    def procesar_ddjj(self, datos_ddjj: dict) -> AsientoDevengamiento:
        """Fases 2-4 del bucle de auto-revisión."""
        motivo_fallo: Optional[str] = None
        asiento: Optional[AsientoDevengamiento] = None
        for intento in range(self.MAX_INTENTOS):
            cuentas = self._seleccionar_cuentas(
                datos_ddjj["tipo"], datos_ddjj.get("montos", {}), motivo_fallo
            )
            asiento = self._armar_asiento(datos_ddjj, cuentas)
            ok, motivo_fallo = self._auditar(asiento)
            if ok:
                asiento.intentos = intento + 1  # type: ignore[attr-defined]
                # Las advertencias informativas (absorciones, redondeos) se preservan
                # en advertencias_info para mostrarlas como warnings no bloqueantes.
                advs_info = list(getattr(asiento, "advertencias", []))
                asiento.advertencias_info = advs_info  # type: ignore[attr-defined]
                asiento.advertencias_fallo = []  # type: ignore[attr-defined]
                asiento.advertencias = []  # type: ignore[attr-defined]
                return asiento
        # Agotó intentos: fallo real → advertencias_fallo bloquea el guardado
        if asiento is not None:
            msg_fallo = f"FALLO tras {self.MAX_INTENTOS} intentos: {motivo_fallo}"
            asiento.advertencias_fallo = [msg_fallo]  # type: ignore[attr-defined]
            asiento.advertencias_info = list(getattr(asiento, "advertencias", []))  # type: ignore[attr-defined]
            asiento.advertencias = [msg_fallo]  # type: ignore[attr-defined]
        return asiento  # type: ignore[return-value]

    def _seleccionar_cuentas(
        self, tipo: str, montos: dict, motivo_fallo: Optional[str]
    ) -> dict:
        """Fase 2: mapeo de cuentas por tipo de impuesto, solo imputables."""
        usar_fallback = motivo_fallo in ("cuenta_inexistente", "cuenta_no_imputable")

        def _buscar(clave: str) -> str:
            if clave in self._resoluciones and not usar_fallback:
                return self._resoluciones[clave]
            resultado = buscar_cuenta_devengamiento(clave, self.plan_imputable, {})
            # resultado es (codigo, descripcion)
            codigo_encontrado = resultado[0] if resultado else ""
            codigo_fallback = CODIGOS_CUENTA_DEVENGAMIENTO_FALLBACK.get(clave, "")
            if codigo_encontrado and codigo_encontrado != codigo_fallback:
                self._resoluciones[clave] = codigo_encontrado
                return codigo_encontrado
            self._resoluciones[clave] = codigo_fallback
            return codigo_fallback

        cuentas: dict = {}
        if tipo == "IVA":
            cuentas = {
                "debe_debito": _buscar("iva_debito"),
                "haber_credito": _buscar("iva_credito"),
                "haber_retenciones": _buscar("iva_retenciones"),
                "haber_percepciones": _buscar("iva_percepciones"),
                "haber_pagar": _buscar("iva_pagar"),
                "debe_saldo_favor": _buscar("iva_saldo_favor"),
                "debe_saldo_tecnico": _buscar("iva_saldo_tecnico"),
                "debe_saldo_libre": _buscar("iva_saldo_libre"),
            }
        elif tipo in ("IIBB_ARBA", "IIBB_CM03", "IIBB"):
            cuentas = {
                "debe_gasto": _buscar("iibb_gasto"),
                "haber_retenciones_banco": _buscar("iibb_retenciones_banco"),
                "haber_retenciones_agentes": _buscar("iibb_retenciones_agentes"),
                "haber_percepciones": _buscar("iibb_percepciones"),
                "haber_pagar": _buscar("iibb_pagar"),
            }
        elif tipo == "TSH":
            # Verificar si el plan tiene cuentas relacionadas con TSH
            _tsh_keywords = ["tsh", "tasa", "seguridad e higiene"]
            _tiene_tsh = False
            if not self.plan_imputable.empty and "descripcion" in self.plan_imputable.columns:
                _desc_lower = self.plan_imputable["descripcion"].str.lower().fillna("")
                _tiene_tsh = _desc_lower.str.contains(
                    "|".join(_tsh_keywords), case=False, na=False
                ).any()
            if not _tiene_tsh:
                return {"_omitir": True}
            cuentas = {
                "debe_gasto": _buscar("tsh_gasto"),
                "haber_pagar": _buscar("tsh_pagar"),
            }
        elif tipo == "SUELDOS":
            cuentas = {
                "debe_bruto": _buscar("sueldos_jornales"),
                "debe_patronales": _buscar("cargas_sociales"),
                "haber_sueldos_pagar": _buscar("sueldos_pagar"),
                "haber_contribuciones_pagar": _buscar("contribuciones_pagar"),
                "haber_aportes_pagar": _buscar("aportes_pagar"),
            }
        return cuentas

    def _armar_asiento(self, datos: dict, cuentas: dict) -> AsientoDevengamiento:
        """Fase 3: construye el borrador del asiento."""
        tipo = datos["tipo"]
        montos = datos.get("montos", {})
        periodo = datos.get("periodo", "") or ""
        try:
            mes, anio = int(periodo.split("/")[0]), int(periodo.split("/")[1])
        except Exception:
            hoy = datetime.today()
            mes, anio = hoy.month, hoy.year
        fecha = _fecha_fin_mes(anio, mes)
        leyenda_base = ""
        renglones: list[RenglonAsiento] = []

        def _desc(cod: Optional[str]) -> str:
            if not cod:
                return ""
            fila = self.plan[self.plan["codigo"].astype(str) == str(cod)]
            return str(fila["descripcion"].iloc[0]) if not fila.empty else str(cod)

        def _renglon(cod: Optional[str], debe: float, haber: float) -> None:
            if not cod:
                return
            _agregar_renglon(renglones, str(cod), _desc(cod), debe=debe, haber=haber, leyenda=leyenda_base)

        if cuentas.get("_omitir"):
            asiento_vacio = AsientoDevengamiento(
                identificador=f"DDJJ_{tipo}_{periodo.replace('/', '_')}",  # type: ignore[arg-type]
                concepto=f"Devengamiento {tipo}",
                fecha=fecha,
                renglones=[],
            )
            asiento_vacio.intentos = 1  # type: ignore[attr-defined]
            asiento_vacio.advertencias = [  # type: ignore[attr-defined]
                f"{tipo} no configurado en plan de cuentas de este cliente"
            ]
            asiento_vacio.tipo = tipo  # type: ignore[attr-defined]
            asiento_vacio.periodo = periodo  # type: ignore[attr-defined]
            return asiento_vacio

        if tipo == "IVA":
            deb  = float(montos.get("debito_fiscal", 0) or 0)
            cred = float(montos.get("credito_fiscal", 0) or 0)

            # PDF es la fuente de verdad para el total de retenciones+percepciones
            ret_pdf = float(montos.get("retenciones_percepciones", 0) or 0)
            # Excel aporta el desglose informativo (inyectado por app.py en el loop review)
            retenc  = float(montos.get("retenciones_iva", 0) or 0)
            percep  = float(montos.get("percepciones_iva", 0) or 0)
            # Si no hay desglose del Excel, usar el total del PDF en la cuenta de retenciones
            if retenc == 0 and percep == 0:
                retenc = ret_pdf

            # saldo_a_ingresar del PDF es el ancla; si no vino, calcular contra el total PDF
            neto_pdf = float(montos.get("saldo_a_ingresar", 0) or 0)
            if neto_pdf == 0 and deb > 0:
                neto_pdf = round(deb - cred - ret_pdf, 2)

            # Construir renglones con los valores de cada fuente
            if deb    > 0: _renglon(cuentas.get("debe_debito"),        deb,     0)
            if cred   > 0: _renglon(cuentas.get("haber_credito"),      0,       cred)
            if retenc > 0: _renglon(cuentas.get("haber_retenciones"),  0,       retenc)
            if percep > 0: _renglon(cuentas.get("haber_percepciones"), 0,       percep)
            if neto_pdf  > 0: _renglon(cuentas.get("haber_pagar"),      0,       neto_pdf)
            elif neto_pdf < 0: _renglon(cuentas.get("debe_saldo_favor"), abs(neto_pdf), 0)

            # ── Guard de signo: saldo a favor del contribuyente nunca va a pasivo ──────
            # Si neto_pdf < 0 (crédito) y algún renglon incorrecto apunta a IVA a Pagar
            # (por un error de parseo de signo en el extractor), se auto-corrige antes
            # del ajuste de balance.
            cod_pagar = str(cuentas.get("haber_pagar", "") or "")
            if neto_pdf < 0 and cod_pagar:
                renglones_corr: list[RenglonAsiento] = []
                for _r in renglones:
                    if str(_r.codigo_cuenta) == cod_pagar and _r.haber > 0:
                        # Mover al DEBE como saldo a favor del contribuyente
                        _renglon(cuentas.get("debe_saldo_favor"), _r.haber, 0)
                    else:
                        renglones_corr.append(_r)
                renglones[:] = renglones_corr

            # ── Ajuste dinámico de balance ──────────────────────────────────────────
            # Calcula el desequilibrio residual (normalmente causado por diferencias entre
            # el Excel de AFIP y el PDF) y lo absorbe en una cuenta patrimonial, de modo
            # que el auditor recibe siempre un asiento con Debe == Haber.
            t_debe  = round(sum(r.debe  for r in renglones), 2)
            t_haber = round(sum(r.haber for r in renglones), 2)
            ajuste  = round(t_debe - t_haber, 2)

            if abs(ajuste) > 0.001:
                if ajuste > 0:
                    # DEBE > HABER → el exceso es mayor obligación fiscal → IVA a Pagar (HABER)
                    _renglon(cuentas.get("haber_pagar"), 0, ajuste)
                else:
                    # HABER > DEBE → el exceso es crédito a favor → Saldo Técnico IVA (DEBE)
                    _renglon(cuentas.get("debe_saldo_tecnico"), abs(ajuste), 0)
        elif tipo in ("IIBB_ARBA", "IIBB_CM03", "IIBB"):
            total = float(
                montos.get("impuesto_determinado", montos.get("monto_total", 0)) or 0
            )
            ret_banco = float(montos.get("retenciones_banco", 0) or 0)
            ret_agentes = float(montos.get("retenciones_agentes", 0) or 0)
            percepciones = float(montos.get("percepciones", 0) or 0)
            saldo = float(montos.get("saldo_a_pagar", montos.get("saldo_dj", 0)) or 0)

            # Si saldo no vino del PDF, calcularlo para garantizar balance
            if saldo == 0:
                saldo = round(total - ret_banco - ret_agentes - percepciones, 2)

            # Verificar qué cuentas existen en el plan del cliente
            codigos_en_plan = set(self.plan["codigo"].astype(str).tolist())

            def _cuenta_en_plan(cod: Optional[str]) -> bool:
                return bool(cod) and str(cod) not in ("0", "None", "") and str(cod) in codigos_en_plan

            cod_ret_banco = cuentas.get("haber_retenciones_banco") or ""
            cod_ret_agentes = cuentas.get("haber_retenciones_agentes") or ""
            cod_percepciones = cuentas.get("haber_percepciones") or ""

            # Si una cuenta de retenciones no existe → absorber su monto en saldo a pagar
            _advertencias_iibb: list[str] = []
            if not _cuenta_en_plan(cod_ret_banco) and ret_banco > 0:
                saldo = round(saldo + ret_banco, 2)
                _advertencias_iibb.append(f"Cuenta retenciones banco '{cod_ret_banco}' no encontrada, neteada en pasivo")
                ret_banco = 0.0
            if not _cuenta_en_plan(cod_ret_agentes) and ret_agentes > 0:
                if _cuenta_en_plan(cod_ret_banco):
                    ret_banco = round(ret_banco + ret_agentes, 2)
                else:
                    saldo = round(saldo + ret_agentes, 2)
                _advertencias_iibb.append(f"Cuenta retenciones agentes '{cod_ret_agentes}' no encontrada, absorbida")
                ret_agentes = 0.0
            if not _cuenta_en_plan(cod_percepciones) and percepciones > 0:
                saldo = round(saldo + percepciones, 2)
                _advertencias_iibb.append(f"Cuenta percepciones '{cod_percepciones}' no encontrada, neteada en pasivo")
                percepciones = 0.0

            # Ajuste final de balance: recalcular saldo para garantizar DEBE == HABER
            haber_componentes = round(ret_banco + ret_agentes + percepciones + saldo, 2)
            if abs(total - haber_componentes) > 0.01:
                saldo = round(total - ret_banco - ret_agentes - percepciones, 2)

            _renglon(cuentas.get("debe_gasto"), total, 0)
            if ret_banco > 0:
                _renglon(cuentas.get("haber_retenciones_banco"), 0, ret_banco)
            if ret_agentes > 0:
                _renglon(cuentas.get("haber_retenciones_agentes"), 0, ret_agentes)
            if percepciones > 0:
                _renglon(cuentas.get("haber_percepciones"), 0, percepciones)
            if saldo > 0:
                _renglon(cuentas.get("haber_pagar"), 0, saldo)

            if _advertencias_iibb:
                # Guardar advertencias para agregar al asiento después de construirlo
                self._advertencias_iibb_pendientes = _advertencias_iibb
        elif tipo == "TSH":
            monto = float(montos.get("tasa_determinada", 0) or 0)
            _renglon(cuentas.get("debe_gasto"), monto, 0)
            _renglon(cuentas.get("haber_pagar"), 0, monto)
        elif tipo == "SUELDOS":
            bruto = float(montos.get("sueldo_bruto", 0) or 0)
            patron = float(montos.get("contribuciones_patronales", 0) or 0)
            retenc = float(montos.get("retenciones_aportes", 0) or 0)
            if bruto > 0:
                _renglon(cuentas.get("debe_bruto"), bruto, 0)
            if patron > 0:
                _renglon(cuentas.get("debe_patronales"), patron, 0)
            neto_sueldo = bruto - retenc
            if neto_sueldo > 0:
                _renglon(cuentas.get("haber_sueldos_pagar"), 0, neto_sueldo)
            if patron > 0:
                _renglon(cuentas.get("haber_contribuciones_pagar"), 0, patron)
            if retenc > 0:
                _renglon(cuentas.get("haber_aportes_pagar"), 0, retenc)

        identificador_str = f"DDJJ_{tipo}_{periodo.replace('/', '_')}"
        asiento = AsientoDevengamiento(
            identificador=identificador_str,  # type: ignore[arg-type]
            concepto=f"Devengamiento {tipo}",
            fecha=fecha,
            renglones=renglones,
        )
        asiento.intentos = 1  # type: ignore[attr-defined]
        # Incluir advertencias de IIBB si las hay
        advertencias_extra = getattr(self, "_advertencias_iibb_pendientes", [])
        asiento.advertencias = advertencias_extra  # type: ignore[attr-defined]
        if hasattr(self, "_advertencias_iibb_pendientes"):
            del self._advertencias_iibb_pendientes
        asiento.tipo = tipo  # type: ignore[attr-defined]
        asiento.periodo = periodo  # type: ignore[attr-defined]
        return asiento

    def _auditar(self, asiento: AsientoDevengamiento) -> tuple[bool, Optional[str]]:
        """Fase 4: auditoría — balance + existencia + imputabilidad."""
        codigos_plan = set(self.plan["codigo"].astype(str).tolist())
        if "imputable" in self.plan.columns:
            codigos_imputables = set(
                self.plan[self.plan["imputable"] == True]["codigo"].astype(str).tolist()
            )
        else:
            codigos_imputables = codigos_plan

        for r in asiento.renglones:
            cod = str(r.codigo_cuenta)
            if cod in ("0", "None", ""):
                continue
            if cod not in codigos_plan:
                return False, "cuenta_inexistente"
            if codigos_imputables and cod not in codigos_imputables:
                return False, "cuenta_no_imputable"

        total_debe = round(sum(r.debe for r in asiento.renglones), 2)
        total_haber = round(sum(r.haber for r in asiento.renglones), 2)
        diff = round(abs(total_debe - total_haber), 2)

        if diff > 0.10:
            return False, f"desbalance_{diff}"

        # Ajuste de centavo
        if 0 < diff <= 0.10 and asiento.renglones:
            ultimo = asiento.renglones[-1]
            if total_debe > total_haber:
                asiento.renglones[-1] = RenglonAsiento(
                    ultimo.codigo_cuenta,
                    ultimo.descripcion_cuenta,
                    ultimo.debe,
                    round(ultimo.haber + diff, 2),
                    ultimo.leyenda,
                )
            else:
                asiento.renglones[-1] = RenglonAsiento(
                    ultimo.codigo_cuenta,
                    ultimo.descripcion_cuenta,
                    round(ultimo.debe + diff, 2),
                    ultimo.haber,
                    ultimo.leyenda,
                )

        return True, None


def generar_excel_ddjj_tango(
    asientos: list,
    nombre_cliente: str,
    cuit: str,
    mes: int,
    anio: int,
    ruta_salida: "Path | str | None" = None,
) -> Path:
    """Genera Excel con 7 columnas para el importador masivo de Tango."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment as XlAlign

    COLS_TANGO = [
        "Fecha_Asiento",
        "Codigo_Cuenta",
        "Descripcion_Cuenta",
        "Debe",
        "Haber",
        "Leyenda",
        "Numero_Asiento",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Renglones"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(COLS_TANGO, 1):
        cell = ws.cell(1, col_idx, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = XlAlign(horizontal="center")

    widths = [14, 16, 35, 16, 16, 45, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fila = 2
    num_asiento = 1
    row_fill = PatternFill("solid", fgColor="EBF3FB")

    for asiento in asientos:
        fecha_str = (
            asiento.fecha.strftime("%d/%m/%Y")
            if hasattr(asiento.fecha, "strftime")
            else str(asiento.fecha)
        )
        for r in asiento.renglones:
            ws.cell(fila, 1, fecha_str)
            ws.cell(fila, 2, str(r.codigo_cuenta))
            ws.cell(fila, 3, r.descripcion_cuenta)
            debe_cell = ws.cell(fila, 4, r.debe if r.debe else 0)
            haber_cell = ws.cell(fila, 5, r.haber if r.haber else 0)
            debe_cell.number_format = '"$ "#,##0.00'
            haber_cell.number_format = '"$ "#,##0.00'
            ws.cell(fila, 6, r.leyenda)
            ws.cell(fila, 7, num_asiento)
            if fila % 2 == 0:
                for col in range(1, 8):
                    ws.cell(fila, col).fill = row_fill
            fila += 1
        num_asiento += 1

    if ruta_salida is None:
        ruta_salida = BASE_DIR / f"Devengamientos_Tango_{cuit}_{mes:02d}_{anio}.xlsx"
    wb.save(str(ruta_salida))
    return Path(ruta_salida)


_TIPOS_ASIENTO_TANGO_VALIDOS = frozenset({
    "ACT", "AJDEP", "AJU", "ALQUILER", "APER-PAT", "AXI", "AXIAF", "BAJA", "BAJAVTA",
    "CIE-PAT", "CN", "COB01", "COB01RES", "COM01", "CONVER", "DEP", "DEPEX", "FOND01",
    "FOND02", "FOND03", "FOND04", "FOND04RES", "FOND05", "FOND05RES", "FOND06", "FOND07",
    "INTERES", "IVA", "MEJ", "PAT01", "PAT02", "PROV01", "RDOACUM", "REFUND", "RES01",
    "RES02", "REV", "RXTAF", "SB", "SEGURO", "STOCK01", "SUELDOS", "SUELDOSRES", "TACC",
    "TACCRES", "TCIA", "TCIARES", "TEN", "VARIOS", "VARIOSRES", "VTA01", "VTA02", "VTA03",
    "VTARESU",
})

_MAPA_TIPO_ASIENTO_TANGO: dict[str, str] = {
    # IVA → VARIOS: en muchas empresas el tipo "IVA" no está habilitado
    # para el módulo Contable y Tango rechaza la importación.
    "IVA": "VARIOS",
    "IIBB": "VARIOS",
    "IIBB_ARBA": "VARIOS",
    "IIBB_CM03": "VARIOS",
    "SUELDOS": "SUELDOS",
    "TISH": "VARIOS",
    "TSH": "VARIOS",
    "CM": "VARIOS",
    "BANCO": "VARIOS",
}


def _codigo_tipo_asiento_tango(tipo_str: str) -> str:
    """Mapea el tipo interno del asiento al código de tipo de asiento del template Tango."""
    tipo = str(tipo_str or "").strip().upper()
    if not tipo:
        return "VARIOS"
    if tipo in _MAPA_TIPO_ASIENTO_TANGO:
        codigo = _MAPA_TIPO_ASIENTO_TANGO[tipo]
        if codigo in _TIPOS_ASIENTO_TANGO_VALIDOS:
            return codigo
    if tipo in _TIPOS_ASIENTO_TANGO_VALIDOS:
        return tipo
    # Códigos de banco del registry (GALICIA, MACRO, …) → VARIOS
    for ficha in BANK_REGISTRY.values():
        if str(ficha.get("codigo_tango") or "").upper() == tipo:
            return "VARIOS"
        if str(ficha.get("slug") or "").upper() == tipo:
            return "VARIOS"
    return "VARIOS"


def _es_tipo_banco_export(tipo_str: str) -> bool:
    tipo = str(tipo_str or "").strip().upper()
    if tipo == "BANCO":
        return True
    for ficha in BANK_REGISTRY.values():
        if str(ficha.get("codigo_tango") or "").upper() == tipo:
            return True
        if str(ficha.get("slug") or "").upper() == tipo:
            return True
    return False


def _nombre_banco_export(tipo_str: str) -> str:
    tipo = str(tipo_str or "").strip().upper()
    for nombre, ficha in BANK_REGISTRY.items():
        if str(ficha.get("codigo_tango") or "").upper() == tipo:
            return str(nombre)
        if str(ficha.get("slug") or "").upper() == tipo:
            return str(nombre)
    return "Banco"


_MONEDAS_EXTRANJERAS_TANGO = frozenset({"DOL", "USD", "U$S", "US$"})

# Estilos de columna del template oficial Tango (VACIO PARA LLENAR).
_TANGO_STYLES_ASIENTOS = [12, 14, 11, 11, 11, 11, 11, 11, None, None]
_TANGO_STYLES_COTIZ = [12, 11, 11, 13, None, None]
_TANGO_STYLES_RENGLONES = [12, 12, 11, 13, 13, 14, 11, 13, 13, None, None]


def _purgar_cotizaciones_moneda_extranjera(ws_c) -> None:
    """Elimina filas del template con moneda extranjera que Tango rechaza al importar."""
    for row_idx in range(ws_c.max_row, 1, -1):
        val = ws_c.cell(row=row_idx, column=2).value
        if val is None:
            continue
        moneda = str(val).strip().upper()
        if moneda in _MONEDAS_EXTRANJERAS_TANGO:
            ws_c.delete_rows(row_idx, 1)


def _assert_xlsx_sano(ruta: Path, *, exigir_shared_strings: bool = True) -> None:
    """Valida que el xlsx sea un ZIP OOXML legible (no puntero LFS / HTML / truncado)."""
    raw = Path(ruta).read_bytes()
    if len(raw) < 64 or not raw.startswith(b"PK"):
        raise ValueError(f"El archivo no es un Excel válido (firma PK ausente): {ruta}")
    with zipfile.ZipFile(ruta) as zf:
        bad = zf.testzip()
        if bad is not None:
            raise ValueError(f"ZIP corrupto en entrada '{bad}': {ruta}")
        names = set(zf.namelist())
        if "xl/workbook.xml" not in names or "[Content_Types].xml" not in names:
            raise ValueError(f"Excel incompleto (falta workbook/Content_Types): {ruta}")
        if exigir_shared_strings and "xl/sharedStrings.xml" not in names:
            raise ValueError(
                f"Plantilla Tango sin sharedStrings.xml (probable resave openpyxl): {ruta}"
            )


def _excel_serial_date(d: date) -> int:
    """Serial Excel (sistema 1900, epoch 1899-12-30)."""
    return (d - date(1899, 12, 30)).days


def _col_letter_xlsx(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _escape_xml_text(s: str) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _celda_xml_tango(col: int, row: int, value, style=None) -> str:
    if value is None or value == "":
        return ""
    ref = f"{_col_letter_xlsx(col)}{row}"
    s_attr = f' s="{style}"' if style is not None else ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return f'<c r="{ref}"{s_attr}><v>{_excel_serial_date(value)}</v></c>'
    if isinstance(value, bool):
        return f'<c r="{ref}"{s_attr} t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{ref}"{s_attr}><v>{value}</v></c>'
    return (
        f'<c r="{ref}"{s_attr} t="inlineStr">'
        f"<is><t>{_escape_xml_text(value)}</t></is></c>"
    )


def _fila_xml_tango(row_idx: int, values: list, styles: list | None, spans: str) -> str:
    cells: list[str] = []
    for i, v in enumerate(values, 1):
        st = styles[i - 1] if styles and i - 1 < len(styles) else None
        c = _celda_xml_tango(i, row_idx, v, st)
        if c:
            cells.append(c)
    joined = "".join(cells)
    return f'<row r="{row_idx}" spans="{spans}">{joined}</row>'


def _mapa_hojas_workbook_zip(zf: zipfile.ZipFile) -> dict[str, str]:
    from xml.etree import ElementTree as ET

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    nsr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    id_to_target = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    out: dict[str, str] = {}
    for sh in wb.findall(f"{ns}sheets/{ns}sheet"):
        target = id_to_target[sh.attrib[f"{nsr}id"]]
        if not target.startswith("xl/"):
            target = "xl/" + target.lstrip("/")
        out[sh.attrib["name"]] = target
    return out


def _append_rows_sheet_xml(
    xml_bytes: bytes, rows_xml: list[str], last_row: int, dim_end_col: str
) -> bytes:
    text = xml_bytes.decode("utf-8")
    text2, n = re.subn(
        r'(<dimension[^>]*ref=")A1:[A-Z]+\d+(")',
        rf"\g<1>A1:{dim_end_col}{last_row}\2",
        text,
        count=1,
    )
    if n:
        text = text2
    if "</sheetData>" not in text:
        raise ValueError("La hoja del template no tiene </sheetData>")
    text = text.replace("</sheetData>", "".join(rows_xml) + "</sheetData>", 1)
    return text.encode("utf-8")


def _rellenar_plantilla_tango_preservando_zip(
    template: Path,
    destino: Path,
    filas_asientos: list[list],
    filas_cotiz: list[list],
    filas_renglones: list[list],
) -> None:
    """
    Copia el xlsx oficial y agrega filas en sheet XML sin openpyxl.save().

    openpyxl reescribe el paquete OOXML (rompe sharedStrings / validaciones Tango)
    y en varias PCs el archivo deja de abrir. Este camino conserva las 29 entradas
    del template original.
    """
    _assert_xlsx_sano(template, exigir_shared_strings=True)
    shutil.copy2(template, destino)

    with zipfile.ZipFile(destino, "r") as zin:
        sm = _mapa_hojas_workbook_zip(zin)
        parts = {name: zin.read(name) for name in zin.namelist()}
        infos = {name: zin.getinfo(name) for name in zin.namelist()}

    for requerida in ("Asientos contables", "Cotizaciones", "Renglones"):
        if requerida not in sm:
            raise ValueError(f"El template no tiene la hoja '{requerida}'")

    a_xml = [
        _fila_xml_tango(i + 2, row, _TANGO_STYLES_ASIENTOS, "1:10")
        for i, row in enumerate(filas_asientos)
    ]
    c_xml = [
        _fila_xml_tango(i + 2, row, _TANGO_STYLES_COTIZ, "1:6")
        for i, row in enumerate(filas_cotiz)
    ]
    r_xml = [
        _fila_xml_tango(i + 2, row, _TANGO_STYLES_RENGLONES, "1:11")
        for i, row in enumerate(filas_renglones)
    ]

    parts[sm["Asientos contables"]] = _append_rows_sheet_xml(
        parts[sm["Asientos contables"]], a_xml, 1 + len(filas_asientos), "J"
    )
    parts[sm["Cotizaciones"]] = _append_rows_sheet_xml(
        parts[sm["Cotizaciones"]], c_xml, 1 + len(filas_cotiz), "F"
    )
    parts[sm["Renglones"]] = _append_rows_sheet_xml(
        parts[sm["Renglones"]], r_xml, 1 + len(filas_renglones), "K"
    )

    tmp = destino.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w") as zout:
        for name, data in parts.items():
            info = infos[name]
            zi = zipfile.ZipInfo(filename=name, date_time=info.date_time)
            zi.compress_type = info.compress_type
            zi.external_attr = info.external_attr
            zout.writestr(zi, data)
    tmp.replace(destino)
    _assert_xlsx_sano(destino, exigir_shared_strings=True)


def generar_excel_tango_nativo(
    asientos: list,
    nombre_cliente: str,
    cuit: str,
    mes: int,
    anio: int,
    id_base: int = 1000,
    nombre_archivo: str | None = None,
    plan_cuentas: pd.DataFrame | None = None,
    *,
    validar: bool = True,
) -> Path:
    """Clona el template vacío de Tango y llena Asientos contables + Renglones + Cotizaciones."""
    import tempfile

    _ = nombre_cliente  # reservado (nombre de archivo / trazas)

    if validar:
        asientos = preparar_asientos_export_tango(asientos, plan_cuentas)

    template = _resolver_ruta_plantilla_asientos_tango_vacio()
    if not template.exists():
        raise FileNotFoundError(
            f"No se encontró el template Tango en: {template}\n"
            "Copiá el archivo 'Asientos contables VACIO PARA LLENAR E IMPORTAR.xlsx' "
            "a la raíz del proyecto o a plantillas/."
        )
    _assert_xlsx_sano(template, exigir_shared_strings=True)

    if nombre_archivo:
        nombre_final = nombre_archivo
    else:
        nombre_final = f"Devengamientos_Tango_{cuit}_{mes:02d}_{anio}.xlsx"

    # Cloud: preferir tmp writable; local: también exportaciones/
    exportaciones_dir = BASE_DIR / "exportaciones"
    try:
        exportaciones_dir.mkdir(exist_ok=True)
        destino = exportaciones_dir / nombre_final
    except OSError:
        destino = Path(tempfile.gettempdir()) / nombre_final

    filas_asientos: list[list] = []
    filas_cotiz: list[list] = []
    filas_renglones: list[list] = []

    for i, asiento in enumerate(asientos):
        id_asiento = getattr(asiento, "identificador", None)
        if id_asiento is None:
            id_asiento = id_base + i
        fecha_obj = parsear_fecha_export_tango(
            getattr(asiento, "fecha_tango_str", None) or asiento.fecha
        )
        tipo_str = getattr(asiento, "tipo", "")
        periodo_str = getattr(asiento, "periodo", "")
        concepto_asiento = str(getattr(asiento, "concepto", "") or "").strip()
        if not concepto_asiento:
            if str(tipo_str).upper() == "CM":
                concepto_asiento = f"DEVENGAMIENTO CM {periodo_str}"
            elif _es_tipo_banco_export(tipo_str):
                concepto_asiento = f"{_nombre_banco_export(tipo_str)} {periodo_str}"
            else:
                concepto_asiento = (
                    f"Devengamiento {_TIPO_DISPLAY.get(tipo_str, tipo_str)} {periodo_str}"
                ).strip()

        codigo_tipo = _codigo_tipo_asiento_tango(tipo_str)
        # Concepto del asiento (cabecera). Leyenda de renglones: siempre vacía.
        filas_asientos.append([
            id_asiento, fecha_obj, "Básico", codigo_tipo, "Ingresado", "PES",
            concepto_asiento, None, "No", None,
        ])
        filas_cotiz.append([id_asiento, "PES", "COTIZACIÓN", 1.0, "No", None])

        for nro, renglon in enumerate(asiento.renglones, 1):
            debe = round(float(renglon.debe), 2) if renglon.debe else None
            haber = round(float(renglon.haber), 2) if renglon.haber else None
            filas_renglones.append([
                id_asiento, nro, renglon.codigo_cuenta,
                debe, haber, fecha_obj, None, None, None, "No", None,
            ])

    _rellenar_plantilla_tango_preservando_zip(
        template, destino, filas_asientos, filas_cotiz, filas_renglones
    )
    return destino


def segmentar_bloques_iva(texto_pdf: str) -> dict[str, str]:
    """
    Segmenta el texto del PDF interno IVA en tres bloques contables.
    Retorna texto por bloque: credito_fiscal, debito_fiscal, liquidacion.
    """
    lineas = texto_pdf.split("\n")
    bloques: dict[str, list[str]] = {
        "credito_fiscal": [],
        "debito_fiscal": [],
        "liquidacion": [],
    }
    seccion: str | None = None

    for linea in lineas:
        upper = linea.upper()
        if (
            "CREDITO FISCAL ACTIVIDADES" in upper
            or "CRÉDITO FISCAL ACTIVIDADES" in upper
            or "OPERACIONES DE COMPRAS" in upper
            or (
                ("CREDITO FISCAL COMPUTABLE" in upper or "CRÉDITO FISCAL COMPUTABLE" in upper)
                and ("ALICUOTA" in upper or "ALÍCUOTA" in upper)
            )
        ):
            seccion = "credito_fiscal"
            continue
        if "DEBITO FISCAL ACTIVIDADES" in upper or "DÉBITO FISCAL ACTIVIDADES" in upper:
            seccion = "debito_fiscal"
            continue
        if "liquidaci" in linea.lower():
            seccion = "liquidacion"
            continue

        if seccion:
            bloques[seccion].append(linea)

    return {k: "\n".join(v) for k, v in bloques.items()}


def extraer_datos_iva(texto_pdf: str) -> dict:
    """
    Extrae montos del PDF interno IVA por bloques (compras/ventas/liquidación).
    Retorna líneas por alícuota y totales agregados para compatibilidad.
    """
    import re

    MESES_ABR = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
    }

    def _limpiar_monto_arg(valor: str) -> float:
        if not valor:
            return 0.0
        t = str(valor).replace("$", "").replace("S", "").replace("\xa0", "")
        t = t.strip().replace(" ", "").replace(".", "").replace(",", ".")
        try:
            return round(float(t), 2)
        except ValueError:
            return 0.0

    def _monto_ultima_columna(linea: str) -> float:
        if "|" in linea:
            return _limpiar_monto_arg(linea.split("|")[-1])
        nums = re.findall(r"[\d.,]+", linea)
        return _limpiar_monto_arg(nums[-1]) if nums else 0.0

    def _parsear_alicuotas_bloque(
        texto_bloque: str,
        alicuotas: tuple[str, ...],
    ) -> list[dict]:
        lineas_out: list[dict] = []
        acumulado: dict[str, float] = {}
        for linea in texto_bloque.split("\n"):
            if re.search(r"total\s+(del\s+)?(cr[eé]dito|d[eé]bito)\s+fiscal", linea, re.IGNORECASE):
                continue
            match = re.search(
                rf"^({'|'.join(re.escape(a) for a in alicuotas)})\s*\|",
                linea.strip(),
            )
            if not match:
                match = re.match(
                    rf"^\s*({'|'.join(re.escape(a) for a in alicuotas)})\b",
                    linea.strip(),
                )
            if not match:
                continue
            alic = match.group(1).replace(".", ",")
            if alic in ("10,5", "10.5"):
                alic = "10,50"
            elif alic == "21":
                alic = "21,00"
            monto = _monto_ultima_columna(linea)
            if monto <= 0:
                continue
            acumulado[alic] = round(acumulado.get(alic, 0.0) + monto, 2)
        for alic, monto in sorted(acumulado.items()):
            lineas_out.append({"alicuota": alic, "monto": monto})
        return lineas_out

    def _parsear_nc_bloque(texto_bloque: str, tipo: str) -> float:
        total = 0.0
        lineas = texto_bloque.split("\n")
        for i, linea in enumerate(lineas):
            linea_l = linea.lower()
            siguiente = lineas[i + 1] if i + 1 < len(lineas) else ""
            monto = 0.0
            if tipo == "compras":
                if not (
                    "credito fiscal a restituir" in linea_l
                    or "crédito fiscal a restituir" in linea_l
                    or ("notas de credito" in linea_l and "compra" in linea_l)
                    or ("notas de crédito" in linea_l and "compra" in linea_l)
                    or ("nota de credito" in linea_l and "compra" in linea_l)
                ):
                    continue
            else:
                if not (
                    "debito fiscal a restituir" in linea_l
                    or "débito fiscal a restituir" in linea_l
                    or ("notas de credito" in linea_l and "venta" in linea_l)
                    or ("notas de crédito" in linea_l and "venta" in linea_l)
                    or ("nota de credito" in linea_l and "venta" in linea_l)
                    or "restitucion de debito" in linea_l
                    or "restitución de débito" in linea_l
                ):
                    continue
            monto = _monto_ultima_columna(linea)
            if monto <= 0 and siguiente:
                nums = re.findall(r"[\d.,]+", siguiente)
                monto = _limpiar_monto_arg(nums[-1]) if nums else 0.0
            if monto > 0:
                total = round(total + monto, 2)
        return total

    def _parsear_liquidacion(texto_bloque: str) -> dict:
        liq = {
            "percep": 0.0,
            "retenc": 0.0,
            "saldo_pagar": 0.0,
            "saldo_libre": 0.0,
            "nc_compras_cf": 0.0,
            "nc_ventas_df": 0.0,
        }
        lineas = texto_bloque.split("\n")
        for i, linea in enumerate(lineas):
            linea_l = linea.lower()
            if "total del credito fiscal" in linea_l or "total del crédito fiscal" in linea_l:
                continue
            if "total del debito fiscal" in linea_l or "total del débito fiscal" in linea_l:
                continue
            siguiente = lineas[i + 1] if i + 1 < len(lineas) else ""

            def _valor(ln: str, sig: str) -> float:
                v = _monto_ultima_columna(ln)
                if v > 0:
                    return v
                nums = re.findall(r"[\d.,]+", sig)
                return _limpiar_monto_arg(nums[-1]) if nums else 0.0

            if "percepciones impositivas sufridas" in linea_l and liq["percep"] == 0:
                liq["percep"] = _valor(linea, siguiente)
            elif "retenciones sufridas" in linea_l and liq["retenc"] == 0:
                liq["retenc"] = _valor(linea, siguiente)
            elif (
                "saldo de impuesto a favor de afip" in linea_l
                or "saldo del impuesto a favor de arca" in linea_l
                or "saldo del impuesto a favor de afip" in linea_l
            ) and liq["saldo_pagar"] == 0:
                liq["saldo_pagar"] = _valor(linea, siguiente)
            elif "saldo de libre disponibilidad" in linea_l and liq["saldo_libre"] == 0:
                liq["saldo_libre"] = _valor(linea, siguiente)
            elif (
                "credito fiscal a restituir" in linea_l or "crédito fiscal a restituir" in linea_l
            ):
                liq["nc_compras_cf"] = round(liq["nc_compras_cf"] + _valor(linea, siguiente), 2)
            elif (
                "debito fiscal a restituir" in linea_l
                or "débito fiscal a restituir" in linea_l
                or "restitucion de debito" in linea_l
                or "restitución de débito" in linea_l
            ):
                liq["nc_ventas_df"] = round(liq["nc_ventas_df"] + _valor(linea, siguiente), 2)
            elif "notas de credito" in linea_l or "notas de crédito" in linea_l:
                monto_nc = _valor(linea, siguiente)
                if monto_nc > 0:
                    if "compra" in linea_l:
                        liq["nc_compras_cf"] = round(liq["nc_compras_cf"] + monto_nc, 2)
                    elif "venta" in linea_l:
                        liq["nc_ventas_df"] = round(liq["nc_ventas_df"] + monto_nc, 2)
        return liq

    bloques = segmentar_bloques_iva(texto_pdf)
    lineas_compras = _parsear_alicuotas_bloque(
        bloques.get("credito_fiscal", ""), ("10,50", "21,00", "27,00")
    )
    lineas_ventas = _parsear_alicuotas_bloque(
        bloques.get("debito_fiscal", ""), ("10,50", "10,5", "21,00", "21")
    )
    nc_compras_bloque = _parsear_nc_bloque(bloques.get("credito_fiscal", ""), "compras")
    nc_ventas_bloque = _parsear_nc_bloque(bloques.get("debito_fiscal", ""), "ventas")
    liq = _parsear_liquidacion(bloques.get("liquidacion", ""))

    nc_compras_cf = round(nc_compras_bloque + liq["nc_compras_cf"], 2)
    nc_ventas_df = round(nc_ventas_bloque + liq["nc_ventas_df"], 2)

    cf_105 = cf_21 = cf_27 = df_105 = df_21 = 0.0
    for ln in lineas_compras:
        if ln["alicuota"] == "10,50":
            cf_105 = ln["monto"]
        elif ln["alicuota"] == "21,00":
            cf_21 = ln["monto"]
        elif ln["alicuota"] == "27,00":
            cf_27 = ln["monto"]
    for ln in lineas_ventas:
        if ln["alicuota"].startswith("10"):
            df_105 = round(df_105 + ln["monto"], 2)
        elif ln["alicuota"].startswith("21"):
            df_21 = round(df_21 + ln["monto"], 2)

    resultado: dict = {
        "periodo": None,
        "lineas_compras": lineas_compras,
        "lineas_ventas": lineas_ventas,
        "cf_105": cf_105,
        "cf_21": cf_21,
        "cf_27": cf_27,
        "df_105": df_105,
        "df_21": df_21,
        "nc_compras_cf": nc_compras_cf,
        "nc_ventas_df": nc_ventas_df,
        "percep": liq["percep"],
        "retenc": liq["retenc"],
        "saldo_pagar": liq["saldo_pagar"],
        "saldo_libre": liq["saldo_libre"],
        "bloques": bloques,
    }

    match_periodo = re.search(
        r"\b(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)-(\d{2})\b",
        texto_pdf,
        re.IGNORECASE,
    )
    if match_periodo:
        mes = MESES_ABR[match_periodo.group(1).lower()]
        yy = int(match_periodo.group(2))
        anio = 2000 + yy if yy < 50 else 1900 + yy
        resultado["periodo"] = (mes, anio)

    return resultado


def extraer_paginas_iva_pdf(
    ruta_pdf,
    _diagnosticos_out: "list | None" = None,
) -> list[dict]:
    """Lee PDF interno IVA y devuelve datos parseados por página/período."""
    import re as _re
    import pdfplumber

    MESES_ABR = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
    }
    MESES_NOMBRE = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12,
    }

    def _extraer_periodo(texto: str) -> "tuple[int, int] | None":
        cabecera = texto[:300].lower()
        for fuente in (cabecera, texto.lower()):
            m = _re.search(
                r"\b(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[.\-/\s](\d{2})\b", fuente
            )
            if m:
                yy = int(m.group(2))
                return MESES_ABR[m.group(1)], (2000 + yy if yy < 50 else 1900 + yy)
            m = _re.search(r"\b(0[1-9]|1[0-2])[/\-](20\d{2})\b", fuente)
            if m:
                return int(m.group(1)), int(m.group(2))
            m = _re.search(r"\b(20\d{2})[/\-](0[1-9]|1[0-2])\b", fuente)
            if m:
                return int(m.group(2)), int(m.group(1))
            for nombre, num in MESES_NOMBRE.items():
                m = _re.search(rf"\b{nombre}\b[^\d]{{0,30}}\b(20\d{{2}})\b", fuente)
                if m:
                    return num, int(m.group(1))
        return None

    paginas: list[dict] = []
    with pdfplumber.open(str(ruta_pdf)) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text() or ""
            if not texto.strip():
                continue

            datos = extraer_datos_iva(texto)
            periodo = datos.get("periodo") or _extraer_periodo(texto)
            if not periodo:
                continue

            mes, anio = periodo
            periodo_str = f"{mes:02d}/{anio}"
            paginas.append({
                "mes": mes,
                "anio": anio,
                "periodo_str": periodo_str,
                "datos": datos,
            })

            if _diagnosticos_out is not None:
                _diagnosticos_out.append({
                    "periodo": periodo_str,
                    "lineas_compras": len(datos.get("lineas_compras", [])),
                    "lineas_ventas": len(datos.get("lineas_ventas", [])),
                    "df_105": round(datos["df_105"], 2),
                    "df_21": round(datos["df_21"], 2),
                    "cf_105": round(datos["cf_105"], 2),
                    "cf_21": round(datos["cf_21"], 2),
                    "cf_27": round(datos["cf_27"], 2),
                    "nc_compras_cf": round(datos["nc_compras_cf"], 2),
                    "nc_ventas_df": round(datos["nc_ventas_df"], 2),
                    "percep": round(datos["percep"], 2),
                    "retenc": round(datos["retenc"], 2),
                    "saldo_pagar": round(datos["saldo_pagar"], 2),
                    "saldo_libre": round(datos["saldo_libre"], 2),
                    "metodo": "bloques_iva",
                })

    return paginas


def leer_pdf_iva_control(
    ruta_pdf,
    plan_cuentas: "pd.DataFrame",
    _diagnosticos_out: "list | None" = None,
    armador=None,
) -> "list[AsientoDevengamiento]":
    """Lee PDF interno IVA y genera asientos usando el armador provisto por la UI."""
    if armador is None:
        raise ValueError("leer_pdf_iva_control requiere un callable armador desde app.py")

    paginas = extraer_paginas_iva_pdf(ruta_pdf, _diagnosticos_out=_diagnosticos_out)
    asientos: list[AsientoDevengamiento] = []

    for idx, item in enumerate(paginas, start=1):
        asiento = armador(
            item["datos"],
            plan_cuentas,
            item["mes"],
            item["anio"],
            identificador=idx,
        )
        if asiento and asiento.renglones:
            asiento.periodo = item["periodo_str"]  # type: ignore[attr-defined]
            asientos.append(asiento)

    return asientos

