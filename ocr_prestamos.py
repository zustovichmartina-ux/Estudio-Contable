"""
OCR de PDFs escaneados de préstamos y generación del Excel final de auditoría.
Combina los datos ya extraídos en Auditoria_Prestamos_v2.xlsx con las cuotas de los PDFs escaneados.
"""
import re
import io
import sys
from pathlib import Path

import fitz
import easyocr
import numpy as np
from PIL import Image
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\recep\Desktop\Estudio Contable")
CARPETA = BASE / "extractos bancarios" / "Prestamos Financieros"

ARCHIVOS_ESCANEADOS = {
    "P1 20250311 Banco Santander__PréstamoN°_039100426399 CUOTAS.pdf": {
        "banco": "Banco Santander", "id": "039100426399"
    },
    "P5 20250312 Banco Nacion_32024192.pdf": {
        "banco": "Banco Nación", "id": "32024192"
    },
    "P21 20250225 Prestamo Banco Provincia (9866).pdf": {
        "banco": "Banco Provincia", "id": "9866"
    },
    "P22 20250228 Prestamo Banco Provincia (2461).pdf": {
        "banco": "Banco Provincia", "id": "2461"
    },
    "P23 20250421 Prestamo Banco Provincia (3405).pdf": {
        "banco": "Banco Provincia", "id": "3405"
    },
    "P24 20250423 Prestamo Banco Provincia (4418).pdf": {
        "banco": "Banco Provincia", "id": "4418"
    },
    "P25 20250423 Prestamo Banco Provincia (4370).pdf": {
        "banco": "Banco Provincia", "id": "4370"
    },
    "P26 y P27 20250617 Y 19 Prestamo Banco Provincia (8535-8407).pdf": {
        "banco": "Banco Provincia", "id": "8535-8407"
    },
    "P28 20250711 Prestamo Banco Provincia (2317).pdf": {
        "banco": "Banco Provincia", "id": "2317"
    },
    "P29 20250730 Prestamo Banco Provincia (7759).pdf": {
        "banco": "Banco Provincia", "id": "7759"
    },
    "P30 20250808 Prestamo Banco Provincia (5790).pdf": {
        "banco": "Banco Provincia", "id": "5790"
    },
}

# ── helpers ─────────────────────────────────────────────────────────────────

def limpiar_monto(s):
    if not s:
        return 0.0
    s = str(s).replace("$", "").strip()
    if re.search(r"\d\.\d{3},\d{2}", s):
        s = s.replace(".", "").replace(",", ".")
    elif re.search(r"\d,\d{3}\.\d{2}", s):
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    s = re.sub(r"[^\d.]", "", s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def parsear_fecha(s):
    for pat in [r"(\d{2}/\d{2}/\d{4})", r"(\d{2}-\d{2}-\d{4})", r"(\d{2}/\d{2}/\d{2})"]:
        m = re.search(pat, s)
        if m:
            return m.group(1)
    return None


def ocr_paginas(pdf_path, reader):
    """Devuelve lista de líneas reconstruidas por coordenada Y."""
    doc = fitz.open(str(pdf_path))
    lineas_total = []
    for idx in range(len(doc)):
        page = doc[idx]
        mat = fitz.Matrix(2.5, 2.5)
        pix = page.get_pixmap(matrix=mat)
        img = np.array(Image.open(io.BytesIO(pix.tobytes("png"))))
        resultados = reader.readtext(img)
        filas: dict[int, list] = {}
        for bbox, texto, _ in resultados:
            y = int((bbox[0][1] + bbox[2][1]) / 2 / 15) * 15
            filas.setdefault(y, []).append((bbox[0][0], texto))
        for y in sorted(filas):
            linea = " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0]))
            lineas_total.append(linea)
        print(f"    Página {idx+1}/{len(doc)} — {len(resultados)} elementos OCR")
    doc.close()
    return lineas_total


def extraer_cuotas(lineas):
    cuotas = []
    en_tabla = False
    pat_num = re.compile(r"(?:^|\s)(\d{1,3})(?:\s|/|\||\.)")

    for linea in lineas:
        lower = linea.lower()
        if any(k in lower for k in [
            "cuota", "capital", "interes", "vencim", "vto",
            "amort", "cronograma", "tabla", "fecha", "pagos",
        ]):
            en_tabla = True
            continue
        if not en_tabla:
            continue

        fecha = parsear_fecha(linea)
        if not fecha:
            continue

        m_num = pat_num.search(linea)
        if not m_num:
            continue

        numero = int(m_num.group(1))
        if numero <= 0 or numero > 600:
            continue

        nums_raw = re.findall(r"\d[\d\.]*,\d{2}|\d[\d,]*\.\d{2}|\d{4,}", linea)
        importes = [v for v in (limpiar_monto(n) for n in nums_raw) if v >= 1.0]
        if len(importes) < 2:
            continue

        if len(importes) >= 4:
            capital, intereses, impuestos, total = importes[-4], importes[-3], importes[-2], importes[-1]
        elif len(importes) >= 3:
            capital, intereses, impuestos, total = importes[-3], importes[-2], 0.0, importes[-1]
        else:
            capital, intereses, impuestos, total = importes[0], 0.0, 0.0, importes[1]

        if total <= 0:
            total = capital + intereses + impuestos

        cuotas.append({
            "cuota": numero, "vencimiento": fecha,
            "capital": round(capital, 2), "intereses": round(intereses, 2),
            "iva_gastos": round(impuestos, 2), "monto_abonar": round(total, 2),
            "saldo_restante": 0.0,
        })

    return cuotas


# ── formateo Excel ───────────────────────────────────────────────────────────

FILL_HEADER_PRESTAMO = PatternFill("solid", fgColor="B8B8B8")
FILL_RESUMEN_TITLE   = PatternFill("solid", fgColor="1F4E79")
FILL_GRILLA_HEAD     = PatternFill("solid", fgColor="2E75B6")
FILL_FILA_GRIS       = PatternFill("solid", fgColor="F2F2F2")
FILL_CIERRE_TITLE    = PatternFill("solid", fgColor="1F4E79")
FILL_SALDO_OK        = PatternFill("solid", fgColor="FFFF00")
FILL_SALDO_NEG       = PatternFill("solid", fgColor="FF0000")

FONT_BLANCO_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_NEGRO_BOLD  = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL      = Font(name="Calibri", size=10)
ALIN_IZQ = Alignment(horizontal="left", vertical="center")
ALIN_DER = Alignment(horizontal="right", vertical="center")
ALIN_CEN = Alignment(horizontal="center", vertical="center")

_side = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)

FMT_MONEDA = '$ #,##0.00'
COLS_ANCHO = [8, 14, 16, 16, 16, 18, 18]
COLS_NOMBRE = ["CUOTA", "VENCIMIENTO", "CAPITAL", "INTERESES", "IVA/GASTOS", "MONTO A ABONAR", "SALDO RESTANTE"]


def _estilo_celda(cell, fill=None, font=None, alin=None, fmt=None, border=True):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alin:
        cell.alignment = alin
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = BORDER


def agregar_prestamo_en_hoja(ws, prestamo_id, banco, cuotas, fila_inicio):
    """Dibuja un bloque completo de préstamo a partir de fila_inicio. Devuelve la próxima fila libre."""
    capital_total = sum(c["capital"] for c in cuotas)
    intereses_total = sum(c["intereses"] for c in cuotas)
    iva_total = sum(c["iva_gastos"] for c in cuotas)

    # A) Cabecera del préstamo
    ws.row_dimensions[fila_inicio].height = 22
    cell = ws.cell(fila_inicio, 1,
                   f"PRÉSTAMO N° {prestamo_id}  |  Capital: $ {capital_total:,.0f}  |  Banco: {banco}")
    ws.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=7)
    _estilo_celda(cell, FILL_HEADER_PRESTAMO, FONT_NEGRO_BOLD, ALIN_IZQ)
    f = fila_inicio + 1

    # B) Resumen anual
    ws.row_dimensions[f].height = 18
    titulo = ws.cell(f, 1, "RESUMEN AÑO 2025")
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=7)
    _estilo_celda(titulo, FILL_RESUMEN_TITLE, FONT_BLANCO_BOLD, ALIN_CEN)
    f += 1
    for col, (label, valor) in enumerate([
        ("Total Capital Amortizado", capital_total),
        ("Total Intereses Devengados", intereses_total),
        ("Total IVA/Gastos", iva_total),
    ], 1):
        ws.row_dimensions[f].height = 16
        cl = ws.cell(f, col * 2 - 1, label)
        _estilo_celda(cl, font=FONT_NEGRO_BOLD, alin=ALIN_IZQ)
        cv = ws.cell(f, col * 2, valor)
        _estilo_celda(cv, font=FONT_NORMAL, alin=ALIN_DER, fmt=FMT_MONEDA)
    f += 1

    # C) Encabezado grilla
    ws.row_dimensions[f].height = 18
    for col_idx, nombre in enumerate(COLS_NOMBRE, 1):
        cell = ws.cell(f, col_idx, nombre)
        _estilo_celda(cell, FILL_GRILLA_HEAD, FONT_BLANCO_BOLD, ALIN_CEN)
    f += 1

    # D) Filas de cuotas
    for i, c in enumerate(cuotas):
        ws.row_dimensions[f].height = 15
        fill_fila = None if i % 2 == 0 else FILL_FILA_GRIS
        valores = [
            c["cuota"], c["vencimiento"], c["capital"], c["intereses"],
            c["iva_gastos"], c["monto_abonar"], c["saldo_restante"],
        ]
        for col_idx, val in enumerate(valores, 1):
            cell = ws.cell(f, col_idx, val)
            is_num = col_idx >= 3
            _estilo_celda(cell, fill_fila, FONT_NORMAL,
                          ALIN_DER if is_num else ALIN_IZQ,
                          FMT_MONEDA if is_num else None)
        f += 1

    return f  # próxima fila libre (sin las 4 en blanco — el llamador las agrega)


def crear_o_abrir_wb(ruta):
    if ruta.exists():
        return openpyxl.load_workbook(str(ruta))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def siguiente_fila_libre(ws):
    """Busca la última fila con contenido y devuelve la siguiente."""
    max_f = ws.max_row
    for f in range(max_f, 0, -1):
        if any(ws.cell(f, c).value for c in range(1, 8)):
            return f + 5  # 4 filas en blanco
    return 1


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print("Iniciando EasyOCR (puede tardar la primera vez)...")
    reader = easyocr.Reader(["es", "en"], gpu=False, verbose=False)
    print("OCR listo.\n")

    # Cargar Excel base
    ruta_base = BASE / "Auditoria_Prestamos_v2.xlsx"
    ruta_salida = BASE / "Auditoria_Prestamos_Completa_Final.xlsx"
    wb = crear_o_abrir_wb(ruta_base)

    resultados = {}  # banco → lista de (prestamo_id, cuotas)

    for nombre_archivo, info in ARCHIVOS_ESCANEADOS.items():
        pdf_path = CARPETA / nombre_archivo
        banco = info["banco"]
        prestamo_id = info["id"]

        print(f"\n{'='*60}")
        print(f"Procesando: {nombre_archivo}")
        print(f"Banco: {banco} | Préstamo: {prestamo_id}")

        if not pdf_path.exists():
            # Intentar con nombre en unicode normalizado
            alternativas = list(CARPETA.glob(f"*{prestamo_id}*.pdf"))
            if alternativas:
                pdf_path = alternativas[0]
                print(f"  → Usando: {pdf_path.name}")
            else:
                print(f"  ✗ Archivo no encontrado, saltando.")
                continue

        print(f"  Ejecutando OCR...")
        lineas = ocr_paginas(pdf_path, reader)
        cuotas = extraer_cuotas(lineas)
        print(f"  → {len(cuotas)} cuotas extraídas")

        if not cuotas:
            # Guardar las líneas OCR para diagnóstico
            print(f"  Líneas OCR (primeras 20):")
            for l in lineas[:20]:
                print(f"    {l}")

        resultados.setdefault(banco, []).append((prestamo_id, cuotas))

    # Agregar al Excel
    print("\nAgregando datos al Excel...")
    for banco, prestamos in resultados.items():
        # Nombre de hoja: primeras 31 chars (límite Excel)
        nombre_hoja = banco[:31]
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
        else:
            ws = wb.create_sheet(nombre_hoja)
            # Anchos de columna
            for col_idx, ancho in enumerate(COLS_ANCHO, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        fila = siguiente_fila_libre(ws)
        for prestamo_id, cuotas in prestamos:
            if cuotas:
                fila = agregar_prestamo_en_hoja(ws, prestamo_id, banco, cuotas, fila)
            else:
                # Fila de advertencia
                ws.row_dimensions[fila].height = 20
                cell = ws.cell(fila, 1,
                    f"⚠ PRÉSTAMO {prestamo_id} — OCR no extrajo cuotas. Revisar PDF manualmente.")
                ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
                cell.fill = PatternFill("solid", fgColor="FFD700")
                cell.font = Font(name="Calibri", size=10, bold=True)
                fila += 2

            fila += 4  # 4 filas en blanco entre préstamos

    wb.save(str(ruta_salida))
    print(f"\n✓ Excel guardado: {ruta_salida}")
    print(f"  Tamaño: {ruta_salida.stat().st_size // 1024} KB")
    print(f"  Hojas: {wb.sheetnames}")

    # Resumen final
    print("\nResumen por banco:")
    for banco, prestamos in resultados.items():
        total_cuotas = sum(len(c) for _, c in prestamos)
        print(f"  {banco}: {len(prestamos)} préstamos, {total_cuotas} cuotas")


if __name__ == "__main__":
    main()
