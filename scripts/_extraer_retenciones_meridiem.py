# -*- coding: utf-8 -*-
"""Retenciones de liquidación Meridiem 04/2025–05/2026 → Excel formato estudio."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import guardar_informe_excel  # noqa: E402

SRC = Path(r"T:\CLIENTES\GRUPO MERIDIEM\SUELDOS Y JORNALES\Liquidacion Sueldos Grupo Meridiem.xlsx")
OUT = Path(r"C:\Users\recep\Desktop\Retenciones_Sueldos_Meridiem.xlsx")

DESDE = (2025, 4)
HASTA = (2026, 5)

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

# Conceptos de retención del empleado (restan al haber).
# Match flexible por encoding / tipografía.
CONCEPTOS_RET = [
    (r"^jubilaci", "Jubilación"),
    (r"^ley\s*19\.?\s*032", "Ley 19.032"),
    (r"^obra\s*social\b", "Obra Social"),
    (r"^adicional\s*obra\s*soc", "Adicional Obra Social"),
    (r"^faecys", "FAECyS"),
    (r"^sindicato", "Sindicato"),
    (r"^aporte\s*extraordinario\s*obra", "Aporte Extraordinario Obra Soc."),
    (r"^retenci[oó]n\s*ganancias?", "Retención Ganancias"),
    (r"^secza", "SECZA"),
    (r"^inacap", "INACAP"),
    (r"^osecac", "OSECAC"),
]


def parse_sheet_periodo(nombre: str) -> tuple[int, int] | None:
    m = re.fullmatch(r"(\d{2})-(\d{4})", str(nombre).strip())
    if not m:
        return None
    mes, anio = int(m.group(1)), int(m.group(2))
    if not (1 <= mes <= 12):
        return None
    return anio, mes


def en_rango(anio: int, mes: int) -> bool:
    return DESDE <= (anio, mes) <= HASTA


def limpia(n: str) -> str:
    return re.sub(r"\s+", " ", str(n or "").strip())


def normaliza_concepto(lab: str) -> str | None:
    t = limpia(lab)
    if not t:
        return None
    # No incluir totales
    if re.search(r"^total\s+", t, re.I):
        return None
    for pat, nombre in CONCEPTOS_RET:
        if re.search(pat, t, re.I):
            return nombre
    return None


def es_inicio_bloque_ret(lab: str) -> bool:
    return bool(re.search(r"total\s+no\s+remunerativos", limpia(lab), re.I))


def es_fin_bloque_ret(lab: str) -> bool:
    t = limpia(lab)
    return bool(
        re.search(r"total\s+descuentos", t, re.I)
        or re.search(r"sueldo\s*neto|^neto\b", t, re.I)
    )


def empleados_cols(header_row) -> list[tuple[int, str]]:
    cols: list[tuple[int, str]] = []
    for idx, val in enumerate(header_row):
        if idx < 4:
            continue
        txt = limpia(val)
        if not txt:
            continue
        if txt.upper().startswith("TOTAL"):
            break
        cols.append((idx, txt))
    return cols


def extraer_hoja(ws) -> list[dict]:
    rows = list(ws.iter_rows(min_row=1, max_row=120, max_col=30, values_only=True))
    if len(rows) < 4:
        return []
    cols = empleados_cols(rows[3])
    if not cols:
        return []

    out: list[dict] = []
    en_bloque = False
    for row in rows:
        lab = limpia(row[1] if len(row) > 1 else "")
        if not lab:
            continue
        if es_inicio_bloque_ret(lab):
            en_bloque = True
            continue
        if es_fin_bloque_ret(lab):
            en_bloque = False
            continue
        if not en_bloque:
            continue
        concepto = normaliza_concepto(lab)
        if not concepto:
            # Cualquier otra línea en el bloque de descuentos con monto
            if re.search(r"contribuci[oó]n|remuneraci[oó]n|importe\s+a\s+detraer", lab, re.I):
                continue
            concepto = lab

        for idx, nombre in cols:
            raw = row[idx] if idx < len(row) else None
            if raw is None or raw == "":
                continue
            try:
                importe = float(raw)
            except (TypeError, ValueError):
                continue
            if abs(importe) < 0.005:
                continue
            out.append(
                {
                    "empleado": nombre,
                    "concepto": concepto,
                    "importe": round(importe, 2),
                    "concepto_orig": lab,
                }
            )
    return out


def sk_mes(mes_label: str) -> tuple[int, int]:
    inv = {v: k for k, v in MESES.items()}
    m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", mes_label)
    if not m:
        return (9999, 13)
    return (int(m.group(2)), inv.get(m.group(1), 13))


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"No está el archivo: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    registros: list[dict] = []
    conceptos_vistos: set[str] = set()
    hojas = 0

    for sheet_name in wb.sheetnames:
        per = parse_sheet_periodo(sheet_name)
        if not per or not en_rango(*per):
            continue
        anio, mes = per
        etiqueta = f"{MESES[mes]} {anio}"
        periodo = f"{mes:02d}/{anio}"
        filas = extraer_hoja(wb[sheet_name])
        hojas += 1
        for e in filas:
            conceptos_vistos.add(e["concepto"])
            registros.append(
                {
                    "Periodo": periodo,
                    "Mes": etiqueta,
                    "Empleado": e["empleado"],
                    "Concepto": e["concepto"],
                    "Importe": e["importe"],
                    "_ord": (anio, mes),
                }
            )
    wb.close()

    if not registros:
        raise SystemExit("No se extrajeron retenciones en el rango pedido.")

    registros.sort(key=lambda r: (r["_ord"], r["Empleado"], r["Concepto"]))
    for r in registros:
        del r["_ord"]

    df = pd.DataFrame(registros)

    por_concepto = (
        df.groupby("Concepto", sort=False)
        .agg(Lineas=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )
    por_concepto["Total"] = por_concepto["Total"].round(2)

    por_mes = (
        df.groupby("Mes", sort=False)
        .agg(Lineas=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )
    por_mes["_o"] = por_mes["Mes"].map(sk_mes)
    por_mes = por_mes.sort_values("_o").drop(columns="_o")
    por_mes["Total"] = por_mes["Total"].round(2)

    por_mes_concepto = (
        df.groupby(["Mes", "Concepto"], sort=False)
        .agg(Total=("Importe", "sum"))
        .reset_index()
    )
    por_mes_concepto["_o"] = por_mes_concepto["Mes"].map(sk_mes)
    por_mes_concepto = por_mes_concepto.sort_values(["_o", "Concepto"]).drop(columns="_o")
    por_mes_concepto["Total"] = por_mes_concepto["Total"].round(2)

    periodos = sorted(df["Periodo"].unique(), key=lambda p: (int(p.split("/")[1]), int(p.split("/")[0])))
    rango = f"{periodos[0]} → {periodos[-1]}"

    guardar_informe_excel(
        OUT,
        titulo="Retenciones de sueldos — GRUPO MERIDIEM SRL",
        subtitulo="Descuentos del empleado por período (liquidación mensual)",
        periodo=rango,
        kpis=[
            ("Meses", df["Periodo"].nunique()),
            ("Líneas", len(df)),
            ("Conceptos", df["Concepto"].nunique()),
            ("Total retenciones", round(float(df["Importe"].sum()), 2)),
        ],
        resumenes=[
            ("Por tipo de retención", por_concepto),
            ("Por mes", por_mes),
            ("Por mes y concepto", por_mes_concepto),
        ],
        detalle=df[["Periodo", "Mes", "Empleado", "Concepto", "Importe"]],
        hoja_detalle="Detalle",
        col_moneda=["Importe", "Total"],
        total_col="Importe",
    )

    print("OUT", OUT)
    print("HOJAS", hojas)
    print("LINEAS", len(df))
    print("TOTAL", round(float(df["Importe"].sum()), 2))
    print("RANGO", rango)
    print("CONCEPTOS", sorted(conceptos_vistos))
    print(por_concepto.to_string(index=False))


if __name__ == "__main__":
    main()
