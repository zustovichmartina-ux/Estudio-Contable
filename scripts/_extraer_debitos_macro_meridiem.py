# -*- coding: utf-8 -*-
"""Débitos Banco Macro (Meridiem) → Excel claro por mes (mismo estilo que Galicia)."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    COLOR_PRIMARIO,
    MONEY_FMT,
    TITLE_FONT,
    ZEBRA,
    guardar_informe_excel,
)

PDF = Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged.pdf")
OUT = Path(r"c:\Users\recep\Desktop\Debitos_Macro_Meridiem_claro.xlsx")

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

# CUITs vistos en transferencias → nombre legible (si se conoce)
CUIT_NOMBRES = {
    "20440418288": "CUIT 20-44041828-8",
    "27425390703": "CUIT 27-42539070-3",
    "27432414448": "CUIT 27-43241444-8",
    "30714058386": "Grupo Meridiem SRL",
    "30663205621": "Caja de Seguros S.A.",
    "30693504186": "BHN Seguros Generales",
}

RE_MONEY = re.compile(r"-?\d{1,3}(?:\.\d{3})*(?:,\d{2})")
RE_MOV = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+(.+)$"
)
RE_SALDO_INI = re.compile(
    r"SALDO ULTIMO EXTRACTO AL .+?\s+(" + RE_MONEY.pattern + r")\s*$",
    re.I,
)


def parse_ar_money(s: str) -> float:
    s = (s or "").strip().replace(" ", "")
    neg = s.startswith("-")
    s = s.lstrip("-")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    v = float(s)
    return -v if neg else v


def fmt_ar(n: float) -> str:
    v = abs(float(n))
    entero = int(round(v))
    if abs(v - entero) < 0.005:
        return f"{entero:,}".replace(",", ".")
    s = f"{v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def concepto_legible(desc: str) -> str:
    d = re.sub(r"\s+", " ", (desc or "").strip())
    # TRANSF:...-CUIT
    m = re.search(r"TRANSF:[^-]+-(\d{11})", d, re.I)
    if m:
        cuit = m.group(1)
        return CUIT_NOMBRES.get(cuit, f"Transferencia CUIT {cuit}")
    # TEF DATANET PR NOMBRE CUIT
    m = re.match(r"TEF DATANET PR\s+(.+?)\s+(\d{11})\b", d, re.I)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip().title()
    if re.search(r"Comision Trf|COMISION TRANSFERENCIAS", d, re.I):
        return "Comisión transferencia"
    if re.search(r"DBCR 25413 S/DB|IDCB GRAL", d, re.I):
        return "Imp. débito Ley 25413"
    if re.search(r"DBCR 25413 S/CR", d, re.I):
        return "Imp. crédito Ley 25413"
    if re.search(r"RET\.?\s*ING\.?\s*BRUTOS\s*SIRCREB|SIRCREB", d, re.I):
        return "Ret. IIBB SIRCREB"
    if re.search(r"RET\s*IIBB\s*TUCUMAN|IIBB TUCUMAN", d, re.I):
        return "Ret. IIBB Tucumán"
    if re.search(r"IMP\.?\s*AFIP", d, re.I):
        return "Impuestos AFIP"
    if re.search(r"Transf\.?\s*MacrOnline|Transf\.?\s*MacrOL|DB TRANSF MINORISTA|DB TR\.\.AUT", d, re.I):
        return "Transferencia MacroOnline"
    if re.search(r"TRF MO CCDO|CR TRANSF AUT SDO MISMO TIT|DB TR\.\.AUT\.SDO", d, re.I):
        return "Transferencia mismo titular"
    if re.search(r"RETIRO CAJ|COM RETIRO EFECTIVO|EXTRAC EFVO", d, re.I):
        return "Retiro / extracción efectivo"
    if re.search(r"DEBITO FISCAL IVA|RETENCION IVA PERCEPCION", d, re.I):
        return "IVA / percepción"
    if re.search(r"PERCEPCION.*BRUTOS|PERCEPCION I I B B", d, re.I):
        return "Percepción IIBB"
    if re.search(r"WNPOWER", d, re.I):
        return "WNPOWER.COM"
    if re.search(r"INACAP", d, re.I):
        return "INACAP cuota"
    if re.search(r"PAGO SERV", d, re.I):
        return "Pago de servicios"
    if re.search(r"TARJETA DE CREDITO VISA|DB TARJETA", d, re.I):
        return "Tarjeta de crédito Visa"
    if re.search(r"Liq\.Susc|Sol\.Resc", d, re.I):
        return "Fondos / rescate (no debería ser débito)"
    # recortar refs numéricas largas al final
    d = re.sub(r"\s+\d{5,}\s*$", "", d)
    d = re.sub(r"\s+SUC\.:\s*\d+", "", d, flags=re.I)
    d = re.sub(r"\s+0\s*$", "", d)
    return d[:60].strip() or "Débito"


def extraer_debitos(pdf_path: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    debitos: list[dict] = []
    saldo_prev: float | None = None
    en_cc_bancaria = False

    for ln in lineas:
        if "CUENTA CORRIENTE BANCARIA NRO" in ln.upper():
            en_cc_bancaria = True
            saldo_prev = None
            continue
        if ln.startswith("CUENTA CORRIENTE ESPECIAL"):
            en_cc_bancaria = False
            saldo_prev = None
            continue
        if not en_cc_bancaria:
            continue

        m_ini = RE_SALDO_INI.search(ln)
        if m_ini:
            saldo_prev = parse_ar_money(m_ini.group(1))
            continue

        m = RE_MOV.match(ln)
        if not m:
            continue
        fecha_s, resto = m.groups()
        montos = RE_MONEY.findall(resto)
        if len(montos) < 2:
            continue
        # último = saldo; penúltimo = importe del movimiento (columna débito o crédito)
        saldo = parse_ar_money(montos[-1])
        importe = abs(parse_ar_money(montos[-2]))
        if importe < 0.005:
            saldo_prev = saldo
            continue

        # texto sin los montos finales
        desc = resto
        for mon in montos[-2:]:
            # quitar última aparición
            idx = desc.rfind(mon)
            if idx >= 0:
                desc = (desc[:idx] + desc[idx + len(mon) :]).strip()
        desc = re.sub(r"\s+", " ", desc).strip()
        # quitar referencia numérica suelta al final si quedó
        desc = re.sub(r"\s+\d{5,8}\s*$", "", desc).strip()

        es_debito: bool | None = None
        if saldo_prev is not None:
            delta = round(saldo - saldo_prev, 2)
            if abs(abs(delta) - importe) <= 0.05:
                es_debito = delta < 0
            elif delta < -0.01:
                es_debito = True
            elif delta > 0.01:
                es_debito = False

        if es_debito is None:
            # heurística por descripción
            if re.search(r"^N/D|^IMP\.|^TRANSF:|^WNPOWER|^INACAP|^DEBITO|^RETIRO|^DB ", desc, re.I):
                es_debito = True
            elif re.search(r"^N/C|^TEF DATANET|Sol\.Resc|Liq\.Susc", desc, re.I):
                es_debito = False
            else:
                # sin certeza: si N/D o transf out típica
                es_debito = bool(re.search(r"N/D|TRANSF:|Comision|COMISION|IMP\. AFIP", desc, re.I))

        saldo_prev = saldo
        if not es_debito:
            continue

        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            continue

        debitos.append(
            {
                "fecha": fecha,
                "mes": fecha.month,
                "anio": fecha.year,
                "importe": round(importe, 2),
                "concepto": concepto_legible(desc),
                "descripcion_raw": desc,
            }
        )

    return debitos


def armar_filas(debitos: list[dict]) -> list[dict]:
    grupos: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for d in debitos:
        grupos[(d["anio"], d["mes"])].append(d)
    filas: list[dict] = []
    for anio, mes in sorted(grupos.keys()):
        etiqueta = f"{MESES[mes]} {anio}"
        for d in sorted(grupos[(anio, mes)], key=lambda x: (x["fecha"], -x["importe"])):
            filas.append(
                {
                    "Mes": etiqueta,
                    "Fecha": d["fecha"].strftime("%d/%m/%Y"),
                    "Importe": d["importe"],
                    "Detalle": d["concepto"],
                    "Línea": f"{fmt_ar(d['importe'])}  {d['concepto']}",
                }
            )
    return filas


def main() -> None:
    if not PDF.exists():
        raise SystemExit(f"No está el PDF: {PDF}")
    debitos = extraer_debitos(PDF)
    if not debitos:
        raise SystemExit("No se extrajeron débitos Macro.")
    filas = armar_filas(debitos)

    import pandas as pd

    df = pd.DataFrame(filas)
    resumen = (
        df.groupby("Mes", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )

    guardar_informe_excel(
        OUT,
        titulo="Débitos Banco Macro — GRUPO MERIDIEM SRL",
        subtitulo="Solo débitos del extracto (filtrados y ordenados)",
        periodo=f"{filas[0]['Mes']} → {filas[-1]['Mes']}" if filas else "",
        kpis=[
            ("Débitos", len(filas)),
            ("Total débitos", round(sum(d["importe"] for d in debitos), 2)),
            ("Meses", df["Mes"].nunique()),
        ],
        resumenes=[("Totales por mes", resumen)],
        detalle=df[["Mes", "Fecha", "Importe", "Detalle"]],
        hoja_detalle="Lista",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    wb = load_workbook(OUT)
    if "Por mes" in wb.sheetnames:
        del wb["Por mes"]
    ws = wb.create_sheet("Por mes", 0)
    ws["A1"] = "Débitos Macro — vista clara"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "GRUPO MERIDIEM SRL · Solo egresos (débitos) · CC Bancaria"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")

    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in filas:
        grupos[r["Mes"]].append(r)

    inv = {v: k for k, v in MESES.items()}

    def sk(lab: str):
        m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", lab)
        if not m:
            return (9999, 13)
        return (int(m.group(2)), inv.get(m.group(1), 13))

    fila = 4
    for mes_label in sorted(grupos.keys(), key=sk):
        items = grupos[mes_label]
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        cell = ws.cell(fila, 1, mes_label)
        cell.fill = fill_mes
        cell.font = font_mes
        ws.cell(fila, 2).fill = fill_mes
        fila += 1
        for i, r in enumerate(items):
            ws.cell(fila, 1, r["Importe"]).number_format = MONEY_FMT
            ws.cell(fila, 1).font = BODY_FONT
            ws.cell(fila, 2, r["Detalle"]).font = BODY_FONT
            if i % 2 == 1:
                ws.cell(fila, 1).fill = ZEBRA
                ws.cell(fila, 2).fill = ZEBRA
            fila += 1
        tot = round(sum(x["Importe"] for x in items), 2)
        ws.cell(fila, 1, tot).number_format = MONEY_FMT
        ws.cell(fila, 1).font = font_tot
        ws.cell(fila, 2, f"TOTAL {mes_label}").font = font_tot
        fila += 2

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    wb.save(OUT)

    print("OUT", OUT)
    print("DEBITOS", len(debitos))
    print("TOTAL", round(sum(d["importe"] for d in debitos), 2))
    print(df.groupby("Mes")["Importe"].agg(["count", "sum"]).to_string())


if __name__ == "__main__":
    main()
