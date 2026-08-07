# -*- coding: utf-8 -*-
"""Débitos Galicia desde extracto PDF → Excel claro por mes."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BOLD_FONT,
    BODY_FONT,
    COLOR_PRIMARIO,
    HDR_FILL,
    HDR_FONT,
    MONEY_FMT,
    SECTION_FONT,
    TITLE_FONT,
    ZEBRA,
    guardar_informe_excel,
)
from openpyxl.styles import Alignment, Font, PatternFill

PDF = Path(r"c:\Users\recep\Downloads\ilovepdf_merged (24).pdf")
OUT = Path(r"c:\Users\recep\Desktop") / "Debitos_Galicia_Meridiem_claro.xlsx"

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

RE_MOV = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s+"
    r"(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s*$"
)
RE_MONEY = re.compile(r"^-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2}$")
RE_CUIT = re.compile(r"^\d{11}$")
RE_SKIP_CONT = re.compile(
    r"^(VARIOS|BANCO |INDUSTRIAL |CAJA DE |PROVEEDORES|ACRED\.|Sucursal:|terminal:|"
    r"ENTRE BCOS|FIMA |P[aá]gina |Resumen de |Total \$|Consolidado|PERIODO |"
    r"TOTAL |Los dep|Dispon|El cr[eé]dito|Tasa |Datos de |Tipo de |N[uú]mero |"
    r"Cantidad |IVA:|CUIT |GRUPO MERIDIEM|Movimientos|Fecha Descri|CBU |"
    r"\d{10,}|DT\.|REG\.|INDUSTRIAL AND)",
    re.I,
)
RE_MES_LABEL = re.compile(
    r"^(Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|Septiembre|Octubre|Noviembre|Diciembre)\s+20\d{2}$",
    re.I,
)


def parse_ar_money(s: str) -> float:
    s = (s or "").strip().replace(" ", "")
    neg = s.startswith("-") or s.startswith("(")
    s = s.replace("-", "").replace("(", "").replace(")", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return -abs(float(s)) if neg else float(s)


def fmt_ar(n: float) -> str:
    """16.000,00 → estilo AR; enteros como 16.000."""
    v = abs(float(n))
    entero = int(round(v))
    if abs(v - entero) < 0.005:
        return f"{entero:,}".replace(",", ".")
    s = f"{v:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def limpiar_nombre(raw: str) -> str:
    t = re.sub(r"\s+", " ", (raw or "").strip())
    t = t.replace(",", " ").replace("  ", " ").strip()
    # Title-ish but keep acronyms
    partes = []
    for w in t.split():
        if w.isupper() and len(w) <= 4:
            partes.append(w)
        else:
            partes.append(w.capitalize() if w.islower() or w.isupper() else w)
    return " ".join(partes)


def es_nombre_beneficiario(line: str) -> bool:
    t = line.strip()
    if not t or RE_MONEY.match(t) or RE_CUIT.match(t):
        return False
    if RE_SKIP_CONT.match(t):
        return False
    if RE_MES_LABEL.match(t):
        return False
    if re.match(r"^\d{2}/\d{2}/\d{2}", t):
        return False
    # nombres / razones sociales
    if re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]", t) and len(t) >= 3:
        return True
    return False


def concepto_legible(desc: str, conts: list[str]) -> str:
    """Prioriza beneficiario (ej. MARIA EUGENIA CARDEN); si no, la descripción del débito."""
    for c in conts:
        if es_nombre_beneficiario(c):
            return limpiar_nombre(c)
    d = re.sub(r"\s+", " ", desc).strip()
    # quitar ruido típico
    d = re.sub(r"\s*-\s*$", "", d)
    return limpiar_nombre(d) if d else "Débito"


def extraer_debitos(pdf_path: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            for ln in t.splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    debitos: list[dict] = []
    i = 0
    while i < len(lineas):
        ln = lineas[i]
        m = RE_MOV.match(ln)
        if not m:
            i += 1
            continue
        fecha_s, desc, mon1, mon2 = m.groups()
        # En este extracto: monto del movimiento + saldo (débito = negativo en mon1)
        monto = parse_ar_money(mon1)
        # créditos positivos → saltar
        if monto >= -0.009:
            i += 1
            continue
        # saltar totales
        if desc.lower().startswith("total"):
            i += 1
            continue

        conts: list[str] = []
        j = i + 1
        while j < len(lineas):
            nxt = lineas[j]
            if RE_MOV.match(nxt) or nxt.startswith("Total $") or nxt.startswith("Resumen de"):
                break
            if nxt.startswith("Página") or "Página" in nxt and "/" in nxt:
                break
            conts.append(nxt)
            j += 1

        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            i = j
            continue

        # Si hay label "Abril 2025" como continuación de comisión del mes anterior,
        # usamos ese mes para agrupar (más fiel al resumen Galicia).
        mes_grupo = fecha.month
        anio_grupo = fecha.year
        for c in conts:
            mm = RE_MES_LABEL.match(c.strip())
            if mm:
                nombre = mm.group(1).lower()
                mapa = {v.lower(): k for k, v in MESES.items()}
                # también sin tilde
                for k, v in list(mapa.items()):
                    mapa[k.replace("á", "a").replace("é", "e").replace("ó", "o")] = v
                key = nombre.replace("á", "a").replace("é", "e").replace("ó", "o")
                if key in mapa or nombre in mapa:
                    mes_grupo = mapa.get(nombre) or mapa[key]
                    anio_m = re.search(r"(20\d{2})", c)
                    if anio_m:
                        anio_grupo = int(anio_m.group(1))
                break

        debitos.append(
            {
                "fecha": fecha,
                "mes": mes_grupo,
                "anio": anio_grupo,
                "importe": abs(monto),
                "concepto": concepto_legible(desc, conts),
                "descripcion_raw": desc,
            }
        )
        i = j

    return debitos


def armar_filas_claras(debitos: list[dict]) -> list[dict]:
    """Una fila por débito, ordenadas por mes/fecha, con etiqueta de mes."""
    grupos: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for d in debitos:
        grupos[(d["anio"], d["mes"])].append(d)

    filas: list[dict] = []
    for anio, mes in sorted(grupos.keys()):
        items = sorted(grupos[(anio, mes)], key=lambda x: (x["fecha"], -x["importe"]))
        etiqueta = f"{MESES[mes]} {anio}"
        for d in items:
            filas.append(
                {
                    "Mes": etiqueta,
                    "Fecha": d["fecha"].strftime("%d/%m/%Y"),
                    "Importe": round(d["importe"], 2),
                    "Detalle": d["concepto"],
                    "Línea": f"{fmt_ar(d['importe'])}  {d['concepto']}",
                }
            )
    return filas


def guardar_excel_por_bloques(ruta: Path, filas: list[dict], debitos: list[dict]) -> Path:
    """Hoja Lista (formato estudio) + hoja clara tipo ABRIL / líneas."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(filas)
    resumen = (
        df.groupby("Mes", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )

    guardar_informe_excel(
        ruta,
        titulo="Débitos Banco Galicia — GRUPO MERIDIEM SRL",
        subtitulo="Solo débitos del extracto (filtrados y ordenados)",
        periodo=f"{filas[0]['Mes']} → {filas[-1]['Mes']}" if filas else "",
        kpis=[
            ("Débitos", len(filas)),
            ("Total débitos", round(sum(d["importe"] for d in debitos), 2)),
            ("Meses", df["Mes"].nunique() if not df.empty else 0),
        ],
        resumenes=[("Totales por mes", resumen)],
        detalle=df[["Mes", "Fecha", "Importe", "Detalle"]],
        hoja_detalle="Lista",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    # Segunda hoja: bloques mes + líneas "16.000  NOMBRE"
    from openpyxl import load_workbook

    wb = load_workbook(ruta)
    if "Por mes" in wb.sheetnames:
        del wb["Por mes"]
    ws = wb.create_sheet("Por mes", 0)
    ws["A1"] = "Débitos Galicia — vista clara"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "GRUPO MERIDIEM SRL · Solo egresos (débitos)"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")

    fila = 4
    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in filas:
        grupos[r["Mes"]].append(r)

    # orden cronológico
    def sort_key(mes_label: str):
        m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", mes_label)
        if not m:
            return (9999, 13)
        nombre, anio = m.group(1), int(m.group(2))
        inv = {v: k for k, v in MESES.items()}
        return (anio, inv.get(nombre, 13))

    for mes_label in sorted(grupos.keys(), key=sort_key):
        items = grupos[mes_label]
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
        cell = ws.cell(fila, 1, mes_label)
        cell.fill = fill_mes
        cell.font = font_mes
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(fila, 2).fill = fill_mes
        fila += 1
        for i, r in enumerate(items):
            ws.cell(fila, 1, r["Importe"]).number_format = MONEY_FMT
            ws.cell(fila, 1).font = BODY_FONT
            ws.cell(fila, 2, r["Detalle"]).font = BODY_FONT
            if i % 2 == 1:
                ws.cell(fila, 1).fill = ZEBRA
                ws.cell(fila, 2).fill = ZEBRA
            fila += 1
        tot = round(sum(x["Importe"] for x in items), 2)
        ws.cell(fila, 1, tot).number_format = MONEY_FMT
        ws.cell(fila, 1).font = font_tot
        ws.cell(fila, 2, f"TOTAL {mes_label}").font = font_tot
        fila += 2

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    wb.save(ruta)
    return ruta


def main() -> None:
    if not PDF.exists():
        raise SystemExit(f"No está el PDF: {PDF}")
    debitos = extraer_debitos(PDF)
    if not debitos:
        raise SystemExit("No se extrajeron débitos.")
    filas = armar_filas_claras(debitos)
    out = guardar_excel_por_bloques(OUT, filas, debitos)
    print("OUT", out)
    print("DEBITOS", len(debitos))
    print("TOTAL", round(sum(d["importe"] for d in debitos), 2))
    # muestra mayo-like
    for r in filas:
        if "CARDEN" in r["Detalle"].upper() or "EUGENIA" in r["Detalle"].upper():
            print("EX", r["Mes"], r["Línea"])


if __name__ == "__main__":
    main()
