# -*- coding: utf-8 -*-
"""
Calce Ventas + Retenciones sufridas vs Créditos bancarios — GRUPO MERIDIEM.

Layout horizontal (1 fila por factura):
  Fecha venta | Cliente | Total venta | Ret IVA | Ret Gan | Ret IIBB | Ret SUSS |
  Ret OSSEG | Ingresos bancarios | Nº FCT | Fecha cobro | Banco | Observación

Ecuación:
  Factura total ≈ Banco + Ret IVA + Ret Gan + Ret IIBB + Ret SUSS + Ret OSSEG

Regla OSSEG (Caja de Seguros / SUSS):
  - Donde haya Ret SUSS, Ret OSSEG = mismo importe (no figura en AFIP).
  - Sin SUSS → OSSEG = 0.
  - Ret IIBB nueva = Ret IIBB anterior − OSSEG (clip a 0 si negativo).

IIBB por factura (antes del ajuste OSSEG):
  - Si hay cobro bancario: residual = Total − Banco − Ret IVA − Ret Gan − Ret SUSS
    (cierra la ecuación pre-OSSEG; suele ~8% s/neto en Caja/Experta).
  - Si no hay cobro: prorrateo de la fila «Retenciones» (clientes) del papel CM
    por total de ventas del mes.
  - Control: suma IIBB vs totales mensuales CM (documentado en Resumen/KPI).
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from itertools import combinations
from pathlib import Path

import pandas as pd
import pdfplumber
import xlrd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    SECTION_FONT,
    SUB_FONT,
    _escribir_encabezado_hoja,
    _escribir_tabla,
    guardar_informe_excel,
)

VENTAS = Path(r"c:\Users\recep\Desktop\VENTAS GRUPO.xlsx")
RET_SS = Path(r"c:\Users\recep\Desktop\RETENCIONES DE SEGURIDAD SOCIAL GRUPO.xls")
RET_IVA = Path(r"c:\Users\recep\Desktop\RETENCIONES DE IVA GRUPO.xls")
RET_GAN = Path(r"c:\Users\recep\Desktop\RETENCIONES DE GANANCIAS GRUPO.xls")
CM_PATH = Path(r"T:\CLIENTES\GRUPO MERIDIEM\CONVENIO MULTILATERAL\LIQUIDACION CONVENIO MULTILATERAL.xlsx")
CM_PATH_UNC = Path(
    r"\\TANGOSRV\Compartido\CLIENTES\GRUPO MERIDIEM\CONVENIO MULTILATERAL\LIQUIDACION CONVENIO MULTILATERAL.xlsx"
)
PDF_GALICIA = Path(r"c:\Users\recep\Downloads\ilovepdf_merged (24).pdf")
PDF_MACRO = Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged (1).pdf")
PDF_MACRO_FB = Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged.pdf")
OUT = Path(r"c:\Users\recep\Desktop\Calce_Ventas_Retenciones_Bancos_Meridiem.xlsx")

TOL = 5.0
TOL_COMBO = 2000.0
TOL_PAGO_UNICO_ABS = 50.0
TOL_PAGO_UNICO_PCT = 0.001  # 0,1%

CUIT_CAJA = "30663205621"
CUIT_PROVINCIA = "30688254090"
CUIT_EXPERTA = "30687156168"
CUIT_BHN = "30693504186"
CUIT_MERIDIEM = "30714058386"

CUIT_NOMBRES = {
    CUIT_CAJA: "CAJA DE SEGUROS SA",
    CUIT_PROVINCIA: "PROVINCIA ART SA",
    CUIT_EXPERTA: "EXPERTA ART S.A",
    CUIT_BHN: "BHN SEGUROS GENERALES SA",
    CUIT_MERIDIEM: "GRUPO MERIDIEM SRL",
}

RE_MOV_GAL = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s+"
    r"(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s*$"
)
RE_MONEY = re.compile(r"-?\d{1,3}(?:\.\d{3})*(?:,\d{2})")
RE_MOV_MACRO = re.compile(r"^(\d{2}/\d{2}/\d{2})\s+(.+)$")
RE_SALDO_INI = re.compile(
    r"SALDO ULTIMO EXTRACTO AL .+?\s+(" + RE_MONEY.pattern + r")\s*$",
    re.I,
)
RE_CUIT = re.compile(r"\b(\d{11})\b")
RE_SKIP_CONT = re.compile(
    r"^(VARIOS|BANCO |INDUSTRIAL |CAJA DE |PROVEEDORES|ACRED\.|Sucursal:|terminal:|"
    r"ENTRE BCOS|FIMA |P[aá]gina |Resumen de |Total \$|Consolidado|PERIODO |"
    r"TOTAL |Los dep|Dispon|El cr[eé]dito|Tasa |Datos de |Tipo de |N[uú]mero |"
    r"Cantidad |IVA:|CUIT |GRUPO MERIDIEM|Movimientos|Fecha Descri|CBU |"
    r"\d{10,}|DT\.|REG\.|INDUSTRIAL AND)",
    re.I,
)

COLS_DETALLE = [
    "Fecha de la venta",
    "Cliente",
    "Total de la venta",
    "Retenciones de IVA",
    "Retenciones de Ganancias",
    "Retenciones de Ingresos Brutos",
    "Retenciones SUSS",
    "Retenciones OSSEG",
    "Ingresos bancarios",
    "Nº factura",
    "Fecha cobro",
    "Banco",
    "Observación / Match",
    "Estado",
    "Diferencia",
]

MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def parse_ar_money(s: str) -> float:
    s = (s or "").strip().replace(" ", "")
    neg = s.startswith("-") or s.startswith("(")
    s = s.replace("-", "").replace("(", "").replace(")", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return -abs(float(s)) if neg else float(s)


def parse_fecha(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).date()
    except Exception:
        return None


def norm_cuit(v) -> str:
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def norm_cli(nombre: str) -> str:
    t = re.sub(r"\s+", " ", (nombre or "").upper().strip())
    t = t.replace(".", "").replace(",", "")
    for suf in (" SA", " S A", " SRL", " S A S"):
        if t.endswith(suf):
            t = t[: -len(suf)].strip()
    return t


def cliente_a_cuit(cli: str) -> str | None:
    n = norm_cli(cli)
    if "CAJA DE SEGUROS" in n:
        return CUIT_CAJA
    if "PROVINCIA" in n and ("ART" in n or "RIESGOS" in n):
        return CUIT_PROVINCIA
    if "EXPERTA" in n:
        return CUIT_EXPERTA
    if "BHN" in n:
        return CUIT_BHN
    if "VARESE" in n:
        return None
    return None


def fmt_ddmm(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def texto_cobro(nro: int | None, fecha: date | None, banco: str) -> str:
    if not nro or not fecha or not banco:
        return ""
    return f"FCT {nro} la recibe el {fmt_ddmm(fecha)} en Banco {banco}"


# ── Lectura ventas / retenciones AFIP / CM ───────────────────────────────────


def cargar_ventas() -> list[dict]:
    wb = load_workbook(VENTAS, data_only=True)
    ws = wb["Ventas"]
    out = []
    for r in range(3, ws.max_row + 1):
        fecha = parse_fecha(ws.cell(r, 1).value)
        if not fecha:
            continue
        nro = int(float(ws.cell(r, 4).value or 0))
        cli = str(ws.cell(r, 5).value or "").strip()
        neto = float(ws.cell(r, 7).value or 0)
        exento = float(ws.cell(r, 9).value or 0)
        iva = float(ws.cell(r, 10).value or 0)
        total = float(ws.cell(r, 11).value or 0)
        if neto <= 0 and exento > 0:
            neto = exento
        out.append(
            {
                "id": f"F{nro}",
                "fecha": fecha,
                "pv": int(float(ws.cell(r, 3).value or 0)),
                "nro": nro,
                "cliente": cli,
                "cuit": cliente_a_cuit(cli),
                "neto": round(neto, 2),
                "iva": round(iva, 2),
                "total": round(total, 2),
                "tipo": str(ws.cell(r, 2).value or ""),
                "ret_iibb": 0.0,
                "ret_iibb_cm": 0.0,
                "iibb_metodo": "",
            }
        )
    return out


def _leer_xls_ret(path: Path, tipo: str) -> list[dict]:
    book = xlrd.open_workbook(str(path))
    sh = book.sheet_by_index(0)
    hdr = [str(sh.cell_value(0, c)).strip().lower() for c in range(sh.ncols)]

    def col(*names):
        for n in names:
            for i, h in enumerate(hdr):
                if n in h:
                    return i
        return None

    i_cuit = col("cuit agente")
    i_den = col("denominaci", "razón", "razon")
    i_fecha = col("fecha ret")
    i_imp = col("importe ret")
    i_cert = col("número certificado", "numero certificado")
    i_comp = col("número comprobante", "numero comprobante")
    i_fcomp = col("fecha comprobante")

    rows = []
    for r in range(1, sh.nrows):
        cuit = norm_cuit(sh.cell_value(r, i_cuit) if i_cuit is not None else "")
        fecha = parse_fecha(sh.cell_value(r, i_fecha) if i_fecha is not None else None)
        imp = float(sh.cell_value(r, i_imp) if i_imp is not None else 0)
        if not fecha or imp == 0:
            continue
        comp = ""
        if i_comp is not None:
            raw = sh.cell_value(r, i_comp)
            if isinstance(raw, float):
                comp = str(int(raw)) if raw == int(raw) else str(raw)
            else:
                comp = str(raw or "").strip()
        comp_n = re.sub(r"\D", "", comp).lstrip("0") or comp
        rows.append(
            {
                "tipo": tipo,
                "cuit": cuit,
                "agente": str(sh.cell_value(r, i_den) if i_den is not None else "").strip(),
                "fecha": fecha,
                "importe": round(imp, 2),
                "certificado": str(sh.cell_value(r, i_cert) if i_cert is not None else "").strip(),
                "comprobante": comp_n,
                "fecha_comp": parse_fecha(sh.cell_value(r, i_fcomp) if i_fcomp is not None else None),
                "usado": False,
            }
        )
    return rows


def cargar_retenciones() -> list[dict]:
    rows = []
    rows += _leer_xls_ret(RET_IVA, "IVA")
    rows += _leer_xls_ret(RET_GAN, "GAN")
    rows += _leer_xls_ret(RET_SS, "SS")
    return rows


def agrupar_retenciones(rets: list[dict]) -> list[dict]:
    by_comp: dict[tuple, dict] = {}
    ss_sin_comp: list[dict] = []

    for r in rets:
        if r["tipo"] == "SS":
            ss_sin_comp.append(r)
            continue
        key = (r["fecha"], r["cuit"], r["comprobante"] or "")
        g = by_comp.setdefault(
            key,
            {
                "fecha": r["fecha"],
                "cuit": r["cuit"],
                "agente": r["agente"],
                "comprobante": r["comprobante"],
                "iva": 0.0,
                "gan": 0.0,
                "ss": 0.0,
                "ids": [],
            },
        )
        if r["tipo"] == "IVA":
            g["iva"] += r["importe"]
        elif r["tipo"] == "GAN":
            g["gan"] += r["importe"]
        g["ids"].append(r)

    grupos = list(by_comp.values())

    for s in ss_sin_comp:
        cands = [g for g in grupos if g["fecha"] == s["fecha"] and g["cuit"] == s["cuit"] and g["ss"] == 0]
        if len(cands) == 1:
            cands[0]["ss"] += s["importe"]
            cands[0]["ids"].append(s)
        elif len(cands) > 1:
            best = max(cands, key=lambda g: g["iva"] + g["gan"])
            best["ss"] += s["importe"]
            best["ids"].append(s)
        else:
            grupos.append(
                {
                    "fecha": s["fecha"],
                    "cuit": s["cuit"],
                    "agente": s["agente"],
                    "comprobante": "",
                    "iva": 0.0,
                    "gan": 0.0,
                    "ss": s["importe"],
                    "ids": [s],
                }
            )

    for g in grupos:
        g["suma_ret"] = round(g["iva"] + g["gan"] + g["ss"], 2)
        g["usado"] = False
    return grupos


def _fnum(x) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def cargar_iibb_cm(meses: set[tuple[int, int]]) -> dict[tuple[int, int], float]:
    """Totales mensuales fila «Retenciones» (clientes) del papel CM."""
    path = CM_PATH if CM_PATH.exists() else CM_PATH_UNC
    wb = load_workbook(path, data_only=True, read_only=True)
    out: dict[tuple[int, int], float] = {}
    for y, m in sorted(meses):
        sheet = f"{m:02d}-{y}"
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=14, values_only=True):
            if row[2] == "Retenciones":
                tot = round(sum(_fnum(x) for x in row[3:11]), 2)
                out[(y, m)] = tot
                break
    wb.close()
    return out


def asignar_iibb_prorrateo(ventas: list[dict], iibb_mes: dict[tuple[int, int], float]) -> str:
    """Prorratea IIBB CM del mes por total de factura (fallback si no hay cobro)."""
    by_mes: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for v in ventas:
        by_mes[(v["fecha"].year, v["fecha"].month)].append(v)

    metodos: list[str] = []
    for key, vs in by_mes.items():
        tot_iibb = iibb_mes.get(key, 0.0)
        sum_tot = round(sum(v["total"] for v in vs), 2)
        periodo = f"{key[1]:02d}-{key[0]}"
        if tot_iibb <= 0 or sum_tot <= 0:
            for v in vs:
                v["ret_iibb_cm"] = 0.0
                v["ret_iibb"] = 0.0
                v["iibb_metodo"] = f"Sin IIBB CM {periodo}"
            continue
        assigned = 0.0
        for i, v in enumerate(vs):
            if i < len(vs) - 1:
                w = v["total"] / sum_tot
                val = round(tot_iibb * w, 2)
                assigned += val
            else:
                val = round(tot_iibb - assigned, 2)
            v["ret_iibb_cm"] = val
            v["ret_iibb"] = val  # default; se reemplaza por residual si hay cobro
            if len(vs) == 1:
                v["iibb_metodo"] = f"IIBB CM {periodo} (100% mes ${tot_iibb:,.2f})"
            else:
                v["iibb_metodo"] = (
                    f"IIBB CM {periodo} prorrateo s/total mes "
                    f"(${tot_iibb:,.2f} / {len(vs)} FCT)"
                )
        metodos.append(f"{periodo}: ${tot_iibb:,.2f} a {len(vs)} FCT")
    return "; ".join(metodos) if metodos else "Sin datos CM"


# ── Créditos bancarios ───────────────────────────────────────────────────────


def _es_no_cobranza(desc: str, conts: list[str] | None = None) -> bool:
    t = " ".join([desc] + (conts or [])).upper()
    if re.search(r"RESCATE FIMA|FIMA PREMIUM|Sol\.Resc|Liq\.Susc|SOL\.RESC", t, re.I):
        return True
    if re.search(r"MISMA\s*TITULARIDAD|TRANSFER\. CASH MISMA|CR TRANSF AUT SDO MISMO TIT", t, re.I):
        return True
    if CUIT_MERIDIEM in t.replace("-", ""):
        return True
    if re.search(r"BENEFICIO PYME|CASHBACK|INTERES|INTERÉS|AJUSTE", t, re.I):
        return True
    return False


def _clasificar_cobranza(desc: str, conts: list[str] | None = None) -> tuple[str, str | None]:
    t = " ".join([desc] + (conts or []))
    if _es_no_cobranza(desc, conts):
        if re.search(r"RESCATE|FIMA|Sol\.Resc|Liq\.Susc", t, re.I):
            return "Fondos/rescate", None
        if re.search(r"MISMA|TITULARIDAD|MERIDIEM", t, re.I):
            return "Traspaso mismo titular", CUIT_MERIDIEM
        return "Otro no cobranza", None
    cuits = RE_CUIT.findall(t.replace("-", ""))
    for c in RE_CUIT.findall(t):
        if c not in cuits:
            cuits.append(c)
    cuit = None
    for c in cuits:
        if c != CUIT_MERIDIEM:
            cuit = c
            break
    tu = t.upper()
    if "CAJA DE SEGUROS" in tu or cuit == CUIT_CAJA:
        return "Cobranza cliente", CUIT_CAJA
    if "EXPERTA" in tu or cuit == CUIT_EXPERTA:
        return "Cobranza cliente", CUIT_EXPERTA
    if "PROVINCIA" in tu or cuit == CUIT_PROVINCIA:
        return "Cobranza cliente", CUIT_PROVINCIA
    if "BHN" in tu or cuit == CUIT_BHN:
        return "Cobranza cliente", CUIT_BHN
    if re.search(r"TEF DATANET|TRANSFERENCIAS CASH|SNP PAGO|SERVICIO PAGO|N/C|ACRED", tu):
        return "Cobranza cliente", cuit
    return "Cobranza/otro ingreso", cuit


def extraer_creditos_galicia(pdf: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf) as pdf_obj:
        for page in pdf_obj.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    creds: list[dict] = []
    i = 0
    while i < len(lineas):
        m = RE_MOV_GAL.match(lineas[i])
        if not m:
            i += 1
            continue
        fecha_s, desc, mon1, _mon2 = m.groups()
        monto = parse_ar_money(mon1)
        if monto <= 0.009 or desc.lower().startswith("total"):
            i += 1
            continue
        conts: list[str] = []
        j = i + 1
        while j < len(lineas):
            nxt = lineas[j]
            if RE_MOV_GAL.match(nxt) or nxt.startswith("Total $") or nxt.startswith("Resumen de"):
                break
            if nxt.startswith("Página") or ("Página" in nxt and "/" in nxt):
                break
            conts.append(nxt)
            j += 1
        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            i = j
            continue
        cat, cuit = _clasificar_cobranza(desc, conts)
        concepto = desc
        for c in conts:
            if re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]", c) and not RE_SKIP_CONT.match(c):
                if not re.match(r"^\d{11}$", c.strip()):
                    concepto = f"{desc} / {c.strip()}"
                    break
        creds.append(
            {
                "banco": "Galicia",
                "fecha": fecha,
                "importe": round(monto, 2),
                "concepto": concepto[:120],
                "descripcion_raw": desc,
                "categoria": cat,
                "cuit": cuit,
                "es_cobranza": cat == "Cobranza cliente",
                "usado": False,
            }
        )
        i = j
    return _dedupe_creds(creds)


def extraer_creditos_macro(pdf: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf) as pdf_obj:
        for page in pdf_obj.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    creds: list[dict] = []
    saldo_prev: float | None = None
    en_cc = False

    for ln in lineas:
        if "CUENTA CORRIENTE BANCARIA NRO" in ln.upper():
            en_cc = True
            saldo_prev = None
            continue
        if ln.startswith("CUENTA CORRIENTE ESPECIAL"):
            en_cc = False
            saldo_prev = None
            continue
        if not en_cc:
            continue

        m_ini = RE_SALDO_INI.search(ln)
        if m_ini:
            saldo_prev = parse_ar_money(m_ini.group(1))
            continue

        m = RE_MOV_MACRO.match(ln)
        if not m:
            continue
        fecha_s, resto = m.groups()
        montos = RE_MONEY.findall(resto)
        if len(montos) < 2:
            continue
        saldo = parse_ar_money(montos[-1])
        importe = abs(parse_ar_money(montos[-2]))
        if importe < 0.005:
            saldo_prev = saldo
            continue

        desc = resto
        for mon in montos[-2:]:
            idx = desc.rfind(mon)
            if idx >= 0:
                desc = (desc[:idx] + desc[idx + len(mon) :]).strip()
        desc = re.sub(r"\s+", " ", desc).strip()

        es_debito: bool | None = None
        if saldo_prev is not None:
            delta = round(saldo - saldo_prev, 2)
            if abs(abs(delta) - importe) <= 0.05:
                es_debito = delta < 0
            elif delta < -0.01:
                es_debito = True
            elif delta > 0.01:
                es_debito = False
        if es_debito is None:
            if re.search(r"^N/D|^IMP\.|^TRANSF:|^WNPOWER|^INACAP|^DEBITO|^RETIRO|^DB ", desc, re.I):
                es_debito = True
            elif re.search(r"^N/C|^TEF DATANET|Sol\.Resc|Liq\.Susc", desc, re.I):
                es_debito = False
            else:
                es_debito = bool(re.search(r"N/D|TRANSF:|Comision|COMISION|IMP\. AFIP", desc, re.I))

        saldo_prev = saldo
        if es_debito:
            continue

        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            continue

        cat, cuit = _clasificar_cobranza(desc)
        creds.append(
            {
                "banco": "Macro",
                "fecha": fecha,
                "importe": round(importe, 2),
                "concepto": desc[:120],
                "descripcion_raw": desc,
                "categoria": cat,
                "cuit": cuit,
                "es_cobranza": cat == "Cobranza cliente",
                "usado": False,
            }
        )
    return _dedupe_creds(creds)


def _dedupe_creds(creds: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    out = []
    for c in creds:
        key = (c["banco"], c["fecha"], c["importe"], c["descripcion_raw"][:60])
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return out


# ── Matching ─────────────────────────────────────────────────────────────────


def score_ret_vs_factura(g: dict, v: dict) -> float | None:
    if v["cuit"] and g["cuit"] and v["cuit"] != g["cuit"]:
        return None
    if g["fecha"] < v["fecha"]:
        delta_d = (v["fecha"] - g["fecha"]).days
        if delta_d > 2:
            return None
    dias = abs((g["fecha"] - v["fecha"]).days)
    if dias > 45:
        return None

    score = dias * 10.0
    if v["neto"] > 0 and g["ss"] > 0:
        exp_ss = round(v["neto"] * 0.01, 2)
        if abs(g["ss"] - exp_ss) <= 1.0:
            score -= 1000
        elif abs(g["ss"] - exp_ss) / max(exp_ss, 1) < 0.02:
            score -= 500
        else:
            score += abs(g["ss"] - exp_ss) / max(v["neto"], 1) * 5000

    if v["iva"] > 0 and g["iva"] > 0:
        exp_iva = round(v["iva"] * 0.80, 2)
        if abs(g["iva"] - exp_iva) <= 1.0:
            score -= 1000
        elif abs(g["iva"] - exp_iva) / max(exp_iva, 1) < 0.02:
            score -= 500
        else:
            score += abs(g["iva"] - exp_iva) / max(v["iva"], 1) * 5000

    if v["neto"] > 0 and g["gan"] > 0:
        exp_gan = round(v["neto"] * 0.02, 2)
        if abs(g["gan"] - exp_gan) <= 2.0:
            score -= 800
        elif abs(g["gan"] - exp_gan) / max(exp_gan, 1) < 0.05:
            score -= 400
        else:
            if g["cuit"] == CUIT_PROVINCIA:
                score += 50
            else:
                score += abs(g["gan"] - exp_gan) / max(v["neto"], 1) * 3000

    if g["ss"] == 0 and g["iva"] == 0 and g["gan"] > 0:
        if g["gan"] > v["total"] * 0.15:
            return None
        score += 200

    return score


def _iibb_proxy_8(v: dict) -> float:
    """Heurística solo para localizar el crédito (~8% s/neto en Caja/Experta)."""
    base = v["neto"] if v.get("neto", 0) > 0 else round(v["total"] / 1.21, 2)
    return round(base * 0.08, 2)


def expected_bank(v: dict, g: dict | None, iibb: float | None = None) -> float:
    ret_afip = g["suma_ret"] if g else 0.0
    if iibb is None:
        iibb = float(v.get("ret_iibb_cm") or 0.0)
    return round(v["total"] - ret_afip - float(iibb), 2)


def iibb_residual(total: float, banco: float, ret_iva: float, ret_gan: float, ret_ss: float) -> float:
    """IIBB pre-OSSEG que cierra: Total − Banco − AFIP (IVA+Gan+SUSS)."""
    return round(float(total) - float(banco) - float(ret_iva) - float(ret_gan) - float(ret_ss), 2)


def aplicar_osseg(ret_ss: float, ret_iibb_prev: float) -> tuple[float, float, str]:
    """
    Regla OSSEG: mismo monto que SUSS; se resta de IIBB.
    Retorna (osseg, iibb_nueva, nota_clip).
    """
    osseg = round(float(ret_ss or 0), 2)
    prev = round(float(ret_iibb_prev or 0), 2)
    nuevo = round(prev - osseg, 2)
    note = ""
    if nuevo < 0:
        note = (
            f"IIBB clip a 0 (era ${prev:,.2f} − OSSEG ${osseg:,.2f} = ${nuevo:,.2f})"
        )
        nuevo = 0.0
    return osseg, nuevo, note


def _recalcular_diff(r: dict) -> None:
    """Ecuación: Total = Banco + IVA + Gan + IIBB + SUSS + OSSEG."""
    if r.get("Ingresos bancarios") is None or r.get("Total de la venta") is None:
        r["Diferencia"] = None
        return
    suma = round(
        float(r.get("Retenciones de IVA") or 0)
        + float(r.get("Retenciones de Ganancias") or 0)
        + float(r.get("Retenciones de Ingresos Brutos") or 0)
        + float(r.get("Retenciones SUSS") or 0)
        + float(r.get("Retenciones OSSEG") or 0)
        + float(r.get("Ingresos bancarios") or 0),
        2,
    )
    r["Diferencia"] = round(float(r["Total de la venta"]) - suma, 2)


def _aplicar_osseg_a_fila(r: dict, ret_iibb_prev: float | None = None) -> None:
    """Aplica OSSEG=SUSS y ajusta IIBB en una fila de detalle ya armada."""
    ss = float(r.get("Retenciones SUSS") or 0)
    prev = (
        float(ret_iibb_prev)
        if ret_iibb_prev is not None
        else float(r.get("Retenciones de Ingresos Brutos") or 0)
    )
    # Si la fila ya tiene OSSEG y IIBB post-ajuste, no re-restar:
    # el caller pasa ret_iibb_prev (pre-OSSEG) cuando recalcula residual.
    osseg, iibb, note = aplicar_osseg(ss, prev)
    r["Retenciones OSSEG"] = osseg
    r["Retenciones de Ingresos Brutos"] = iibb
    if note:
        obs = str(r.get("Observación / Match") or "")
        if note not in obs:
            r["Observación / Match"] = f"{obs} | {note}".strip(" |")
    _recalcular_diff(r)


def _row_detalle(
    *,
    estado: str,
    v: dict,
    g: dict | None,
    c: dict | None,
    ret_iva: float,
    ret_gan: float,
    ret_ss: float,
    ret_iibb: float,
    credito: float | None,
    obs_extra: str = "",
    iibb_origen: str = "",
) -> dict:
    """ret_iibb es el importe PRE-OSSEG (residual cobro o prorrateo CM)."""
    osseg, iibb_adj, clip_note = aplicar_osseg(ret_ss, ret_iibb)
    banco_val = credito if credito is not None else 0.0
    suma = round(ret_iva + ret_gan + iibb_adj + ret_ss + osseg + banco_val, 2)
    diff = round(v["total"] - suma, 2) if credito is not None else None
    obs_parts = []
    cobro = texto_cobro(v["nro"], c["fecha"] if c else None, c["banco"] if c else "")
    if cobro:
        obs_parts.append(cobro)
    origen = iibb_origen or v.get("iibb_metodo") or ""
    if origen:
        obs_parts.append(origen)
    if osseg > 0:
        obs_parts.append(f"OSSEG=SUSS ${osseg:,.2f}; IIBB post-OSSEG")
    if clip_note:
        obs_parts.append(clip_note)
    if obs_extra:
        obs_parts.append(obs_extra)
    if c and c.get("concepto"):
        obs_parts.append(c["concepto"][:80])
    return {
        "Fecha de la venta": v["fecha"],
        "Cliente": v["cliente"],
        "Total de la venta": v["total"],
        "Retenciones de IVA": ret_iva,
        "Retenciones de Ganancias": ret_gan,
        "Retenciones de Ingresos Brutos": iibb_adj,
        "Retenciones SUSS": ret_ss,
        "Retenciones OSSEG": osseg,
        "Ingresos bancarios": credito,
        "Nº factura": v["nro"],
        "Fecha cobro": c["fecha"] if c else None,
        "Banco": c["banco"] if c else "",
        "Observación / Match": " | ".join(p for p in obs_parts if p),
        "Estado": estado,
        "Diferencia": diff,
        "_id": v["id"],
    }


def _find_bank(
    cobranzas: list[dict], exp: float, cuit: str | None, ref: date, tol: float = TOL
) -> dict | None:
    best = None
    bestdiff = 1e18
    for c in cobranzas:
        if c["usado"]:
            continue
        if cuit and c["cuit"] and cuit != c["cuit"]:
            continue
        if abs((c["fecha"] - ref).days) > 45:
            continue
        diff = abs(c["importe"] - exp)
        if diff < bestdiff:
            bestdiff = diff
            best = c
    if best and bestdiff <= tol:
        return best
    return None


def _find_bank_multi(
    cobranzas: list[dict],
    targets: list[float],
    cuit: str | None,
    ref: date,
    tol: float = TOL,
) -> tuple[dict | None, float | None]:
    best = None
    best_exp = None
    bestdiff = 1e18
    for c in cobranzas:
        if c["usado"]:
            continue
        if cuit and c["cuit"] and cuit != c["cuit"]:
            continue
        if abs((c["fecha"] - ref).days) > 45:
            continue
        for exp in targets:
            diff = abs(c["importe"] - exp)
            if diff < bestdiff:
                bestdiff = diff
                best = c
                best_exp = exp
    if best and bestdiff <= tol:
        return best, best_exp
    return None, None


def _tol_pago_unico(monto: float) -> float:
    return max(TOL_PAGO_UNICO_ABS, abs(float(monto)) * TOL_PAGO_UNICO_PCT)


def _ym(d: date) -> tuple[int, int]:
    return (d.year, d.month)


def _next_ym(ym: tuple[int, int]) -> tuple[int, int]:
    y, m = ym
    return (y, m + 1) if m < 12 else (y + 1, 1)


def _creditos_cliente_ventana(
    cobranzas: list[dict],
    cuit: str | None,
    ym: tuple[int, int],
    *,
    prefer_same_month: bool = True,
) -> list[dict]:
    """
    Créditos del mismo CUIT.
    Preferir mes de cobro = mes venta; si no hay, mes siguiente
    (salvo que el mes siguiente tenga ventas propias — eso lo decide el caller).
    """
    same = []
    nxt = []
    ym2 = _next_ym(ym)
    for c in cobranzas:
        if c["usado"]:
            continue
        if cuit and c.get("cuit") and c["cuit"] != cuit:
            continue
        if not cuit and c.get("cuit"):
            continue
        cy = _ym(c["fecha"])
        if cy == ym:
            same.append(c)
        elif cy == ym2:
            nxt.append(c)
    if prefer_same_month and same:
        return same
    return same + nxt


def _rets_afip_para_lote(
    lote: list[dict],
    factura_ret: dict[str, dict],
    grupos: list[dict],
) -> tuple[float, float, float, list[dict]]:
    """Suma IVA/GAN/SS ya asignadas + grupos AFIP libres del mismo CUIT aplicables al lote."""
    iva = gan = ss = 0.0
    usados_g: list[dict] = []
    seen_ids: set[int] = set()
    for v in lote:
        g = factura_ret.get(v["id"])
        if g and id(g) not in seen_ids:
            seen_ids.add(id(g))
            iva += g["iva"]
            gan += g["gan"]
            ss += g["ss"]
            usados_g.append(g)
    cuit = lote[0].get("cuit") if lote else None
    if not cuit:
        return round(iva, 2), round(gan, 2), round(ss, 2), usados_g
    fmin = min(v["fecha"] for v in lote)
    fmax = max(v["fecha"] for v in lote)
    for g in grupos:
        if g["usado"] or g.get("cuit") != cuit or id(g) in seen_ids:
            continue
        if g["fecha"] < fmin and (fmin - g["fecha"]).days > 5:
            continue
        if g["fecha"] > fmax and (g["fecha"] - fmax).days > 60:
            continue
        iva += g["iva"]
        gan += g["gan"]
        ss += g["ss"]
        usados_g.append(g)
        seen_ids.add(id(g))
    return round(iva, 2), round(gan, 2), round(ss, 2), usados_g


def _asignar_pago_unico(
    lote: list[dict],
    c: dict,
    ret_iva: float,
    ret_gan: float,
    ret_ss: float,
    obs: str,
    estado: str = "Calza OK (pago único mes)",
) -> list[dict]:
    """Prorratea 1 crédito + ret AFIP entre N facturas; IIBB residual cierra la ecuación."""
    sum_tot = round(sum(v["total"] for v in lote), 2)
    rows: list[dict] = []
    for i, v in enumerate(lote):
        if i < len(lote) - 1:
            w = v["total"] / sum_tot if sum_tot else 0
            r_iva = round(ret_iva * w, 2)
            r_gan = round(ret_gan * w, 2)
            r_ss = round(ret_ss * w, 2)
            credito = round(c["importe"] * w, 2)
        else:
            r_iva = round(ret_iva - sum(x["Retenciones de IVA"] for x in rows), 2)
            r_gan = round(ret_gan - sum(x["Retenciones de Ganancias"] for x in rows), 2)
            r_ss = round(ret_ss - sum(x["Retenciones SUSS"] for x in rows), 2)
            credito = round(
                c["importe"] - sum(x["Ingresos bancarios"] or 0 for x in rows), 2
            )
        iibb = iibb_residual(v["total"], credito, r_iva, r_gan, r_ss)
        rows.append(
            _row_detalle(
                estado=estado,
                v=v,
                g=None,
                c=c,
                ret_iva=r_iva,
                ret_gan=r_gan,
                ret_ss=r_ss,
                ret_iibb=iibb,
                credito=credito,
                obs_extra=obs,
                iibb_origen=(
                    f"IIBB residual cobro (cierra eq.); CM mes prorrata "
                    f"${v.get('ret_iibb_cm') or 0:,.2f}"
                ),
            )
        )
    return rows


def _lote_cierra_con_credito(
    lote: list[dict],
    importe: float,
    ret_iva: float,
    ret_gan: float,
    ret_ss: float,
) -> tuple[bool, float]:
    """True si suma(totales) ≈ crédito + AFIP + IIBB razonable. Devuelve (ok, score)."""
    sum_tot = round(sum(v["total"] for v in lote), 2)
    sum_afip = round(ret_iva + ret_gan + ret_ss, 2)
    sum_iibb_cm = round(sum(float(v.get("ret_iibb_cm") or 0) for v in lote), 2)
    sum_iibb_8 = round(sum(_iibb_proxy_8(v) for v in lote), 2)
    tol = _tol_pago_unico(sum_tot)
    best = 1e18
    for iibb in (sum_iibb_cm, sum_iibb_8, 0.0):
        d = abs(importe + sum_afip + iibb - sum_tot)
        if d < best:
            best = d
    if best <= tol:
        return True, best
    iibb_imp = round(sum_tot - importe - sum_afip, 2)
    # IIBB típico Caja/Experta ~8%; tope 12% del total o CM*1.2
    max_iibb = max(sum_iibb_cm * 1.20, sum_iibb_8 * 1.20, sum_tot * 0.12, tol)
    if -tol <= iibb_imp <= max_iibb:
        pen = min(abs(iibb_imp - sum_iibb_cm), abs(iibb_imp - sum_iibb_8), abs(iibb_imp))
        return True, pen * 0.01
    return False, 1e18


def _buscar_subset_pago(
    lote: list[dict],
    c: dict,
    factura_ret: dict[str, dict],
    grupos: list[dict],
) -> list[dict] | None:
    """Busca subconjunto de 2..N FCT que cierre con el crédito c.
    Prefiere el que mejor acerque crédito ≈ total−AFIP−IIBB y mayor cobertura."""
    if len(lote) < 2:
        return None
    best_combo = None
    best_key = None  # (score, -sum_tot, -n)  minimize score, maximize coverage
    pool = sorted(lote, key=lambda v: -v["total"])[:10]
    for k in range(len(pool), 1, -1):
        if k >= 8:
            continue
        if k == len(lote) and len(lote) <= 10:
            iter_combos = [tuple(lote)]
        else:
            iter_combos = combinations(pool, k)
        for combo in iter_combos:
            combo_l = list(combo)
            r_iva, r_gan, r_ss, _gs = _rets_afip_para_lote(combo_l, factura_ret, grupos)
            ok, score = _lote_cierra_con_credito(combo_l, c["importe"], r_iva, r_gan, r_ss)
            if not ok:
                continue
            sum_tot = sum(v["total"] for v in combo_l)
            # Penalizar si el banco cubre <75% del total (retenciones no deberían ser >25%)
            cobertura = c["importe"] / sum_tot if sum_tot else 0
            if cobertura < 0.75:
                score += (0.75 - cobertura) * sum_tot * 0.01
            key = (score, -sum_tot, -len(combo_l))
            if best_key is None or key < best_key:
                best_key = key
                best_combo = combo_l
    return best_combo


def _match_pago_unico_mes(
    ventas: list[dict],
    cobranzas: list[dict],
    factura_ret: dict[str, dict],
    grupos: list[dict],
    used_facturas: set[str],
) -> list[dict]:
    """
    N≥2 FCT mismo cliente+mes:
    - Si hay 1 solo crédito (mes venta, o mes sig. si no hay en el mes): pago único de todas.
    - Si hay varios créditos: cada crédito vs subconjunto de FCT que cierre.
    """
    by_key: dict[tuple, list[dict]] = defaultdict(list)
    meses_con_venta: dict[str, set[tuple[int, int]]] = defaultdict(set)
    for v in ventas:
        cuit = v.get("cuit")
        if not cuit:
            continue
        ym = _ym(v["fecha"])
        meses_con_venta[cuit].add(ym)
        if v["id"] not in used_facturas:
            by_key[(cuit, ym)].append(v)

    ok_rows: list[dict] = []
    for (cuit, ym), lote in sorted(by_key.items(), key=lambda x: (x[0][1], x[0][0])):
        lote = [v for v in lote if v["id"] not in used_facturas]
        if len(lote) < 2:
            continue
        lote = sorted(lote, key=lambda v: (v["fecha"], v["nro"]))

        same = _creditos_cliente_ventana(cobranzas, cuit, ym, prefer_same_month=True)
        # Mes siguiente solo si ese mes no tiene ventas del mismo cliente
        ym2 = _next_ym(ym)
        if same:
            creds = same
        elif ym2 not in meses_con_venta.get(cuit, set()):
            creds = _creditos_cliente_ventana(
                cobranzas, cuit, ym, prefer_same_month=False
            )
        else:
            # hay ventas en mes siguiente: permitir créditos next-month
            # solo si aún quedan libres (no bloquear), pero priorizar subsets
            creds = _creditos_cliente_ventana(
                cobranzas, cuit, ym, prefer_same_month=False
            )

        if not creds:
            continue

        mm_aaaa = f"{ym[1]:02d}/{ym[0]}"

        if len(creds) == 1:
            c = creds[0]
            ret_iva, ret_gan, ret_ss, gs_afip = _rets_afip_para_lote(
                lote, factura_ret, grupos
            )
            ok, _sc = _lote_cierra_con_credito(
                lote, c["importe"], ret_iva, ret_gan, ret_ss
            )
            combo = lote if ok else _buscar_subset_pago(lote, c, factura_ret, grupos)
            if not combo:
                continue
            ret_iva, ret_gan, ret_ss, gs_afip = _rets_afip_para_lote(
                combo, factura_ret, grupos
            )
            c["usado"] = True
            for g in gs_afip:
                g["usado"] = True
                for r in g.get("ids") or []:
                    r["usado"] = True
            for v in combo:
                if v["id"] not in factura_ret and gs_afip:
                    factura_ret[v["id"]] = gs_afip[0]
            obs = f"Pago único mes {mm_aaaa} — {len(combo)} facturas"
            rows = _asignar_pago_unico(combo, c, ret_iva, ret_gan, ret_ss, obs)
            for r, v in zip(rows, combo):
                used_facturas.add(v["id"])
                ok_rows.append(r)
            continue

        # Varios créditos: asignar subsets greedy (crédito más grande primero)
        restantes = list(lote)
        for c in sorted(creds, key=lambda x: -x["importe"]):
            if c["usado"] or len(restantes) < 2:
                break
            combo = _buscar_subset_pago(restantes, c, factura_ret, grupos)
            if not combo:
                continue
            ret_iva, ret_gan, ret_ss, gs_afip = _rets_afip_para_lote(
                combo, factura_ret, grupos
            )
            c["usado"] = True
            for g in gs_afip:
                g["usado"] = True
                for r in g.get("ids") or []:
                    r["usado"] = True
            obs = f"Pago único mes {mm_aaaa} — {len(combo)} facturas"
            rows = _asignar_pago_unico(combo, c, ret_iva, ret_gan, ret_ss, obs)
            ids_combo = {v["id"] for v in combo}
            for r, v in zip(rows, combo):
                used_facturas.add(v["id"])
                ok_rows.append(r)
            restantes = [v for v in restantes if v["id"] not in ids_combo]

    return ok_rows


def _buscar_combo_facturas(g: dict, libres: list[dict]) -> list[dict] | None:
    if g["ss"] <= 0 or g["iva"] <= 0:
        return None
    cands = [
        v
        for v in libres
        if v.get("cuit") == g["cuit"]
        and v["fecha"] <= g["fecha"]
        and (g["fecha"] - v["fecha"]).days <= 45
    ]
    if len(cands) < 2:
        return None
    target_neto = g["ss"] / 0.01
    target_iva = g["iva"] / 0.80
    cands.sort(key=lambda v: (abs((g["fecha"] - v["fecha"]).days), -v["total"]))

    best = None
    best_err = 1e18
    for k in (2, 3, 4):
        if len(cands) < k:
            continue
        pool = cands[:18] if len(cands) > 18 else cands
        for combo in combinations(pool, k):
            sn = sum(v["neto"] for v in combo)
            si = sum(v["iva"] for v in combo)
            err = abs(sn - target_neto) + abs(si - target_iva)
            if err < best_err:
                best_err = err
                best = list(combo)
            if err <= 2.0:
                return list(combo)
    if best and best_err <= 5.0:
        return best
    return None


def calzar(ventas: list[dict], grupos: list[dict], creds: list[dict]) -> tuple[list[dict], list[dict]]:
    cobranzas = [c for c in creds if c["es_cobranza"]]
    ok: list[dict] = []
    pendientes: list[dict] = []
    used_facturas: set[str] = set()
    sin_match: list[dict] = []

    pairs: list[tuple[float, dict, dict]] = []
    for g in grupos:
        for v in ventas:
            sc = score_ret_vs_factura(g, v)
            if sc is None:
                continue
            pairs.append((sc, g, v))
    pairs.sort(key=lambda x: x[0])

    factura_ret: dict[str, dict] = {}
    for sc, g, v in pairs:
        if g["usado"] or v["id"] in factura_ret:
            continue
        if g["ss"] > 0 and g["iva"] > 0 and v["neto"] > 0:
            if abs(g["ss"] - round(v["neto"] * 0.01, 2)) > 1.0:
                continue
            if abs(g["iva"] - round(v["iva"] * 0.80, 2)) > 1.0:
                continue
        elif g["iva"] > 0 and v["iva"] > 0:
            if abs(g["iva"] - round(v["iva"] * 0.80, 2)) > 1.0 and sc > -500:
                continue
        elif g["ss"] == 0 and g["iva"] == 0 and g["gan"] > 0:
            if g["cuit"] == CUIT_PROVINCIA:
                continue
            if sc > 500:
                continue
        else:
            continue
        g["usado"] = True
        for r in g["ids"]:
            r["usado"] = True
        factura_ret[v["id"]] = g

    # ── Primero: combos AFIP (N FCT ↔ 1 retención) + banco ───────────────────
    for g in list(grupos):
        if g["usado"]:
            continue
        libres = [
            v for v in ventas if v["id"] not in used_facturas and v["id"] not in factura_ret
        ]
        combo = _buscar_combo_facturas(g, libres)
        if not combo:
            continue
        if any(v["id"] in used_facturas for v in combo):
            continue
        sum_total = round(sum(v["total"] for v in combo), 2)
        sum_iibb_cm = round(sum(v.get("ret_iibb_cm") or 0 for v in combo), 2)
        sum_iibb_8 = round(sum(_iibb_proxy_8(v) for v in combo), 2)
        targets = [
            round(sum_total - g["suma_ret"] - sum_iibb_8, 2),
            round(sum_total - g["suma_ret"] - sum_iibb_cm, 2),
            round(sum_total - g["suma_ret"], 2),
        ]
        c, _ = _find_bank_multi(
            cobranzas, targets, g["cuit"], g["fecha"], tol=max(TOL, TOL_COMBO)
        )
        nros = "+".join(str(v["nro"]) for v in sorted(combo, key=lambda x: x["nro"]))
        g["usado"] = True
        for r in g["ids"]:
            r["usado"] = True
        if c:
            c["usado"] = True
        yms = {_ym(v["fecha"]) for v in combo}
        if len(yms) == 1 and c:
            ym0 = next(iter(yms))
            obs_extra = f"Pago único mes {ym0[1]:02d}/{ym0[0]} — {len(combo)} facturas"
            estado_c = "Calza OK (pago único mes)"
        else:
            obs_extra = f"TEF cubre FCT {nros}"
            estado_c = "Calza OK (pago combo)"
        for v in combo:
            w = v["total"] / sum_total if sum_total else 0
            ret_iva = round(g["iva"] * w, 2)
            ret_gan = round(g["gan"] * w, 2)
            ret_ss = round(g["ss"] * w, 2)
            credito = round(c["importe"] * w, 2) if c else None
            if credito is not None:
                iibb = iibb_residual(v["total"], credito, ret_iva, ret_gan, ret_ss)
                origen = (
                    f"IIBB residual cobro (cierra eq.); CM mes prorrata "
                    f"${v.get('ret_iibb_cm') or 0:,.2f}"
                )
            else:
                iibb = float(v.get("ret_iibb_cm") or 0)
                origen = v.get("iibb_metodo") or ""
            ok.append(
                _row_detalle(
                    estado=estado_c,
                    v=v,
                    g=g,
                    c=c,
                    ret_iva=ret_iva,
                    ret_gan=ret_gan,
                    ret_ss=ret_ss,
                    ret_iibb=iibb,
                    credito=credito,
                    obs_extra=obs_extra,
                    iibb_origen=origen,
                )
            )
            used_facturas.add(v["id"])
            factura_ret[v["id"]] = g
        if c and combo:
            block = ok[-len(combo) :]
            for key, total_g in (
                ("Retenciones de IVA", g["iva"]),
                ("Retenciones de Ganancias", g["gan"]),
                ("Retenciones SUSS", g["ss"]),
            ):
                s = round(sum(r[key] for r in block), 2)
                block[-1][key] = round(block[-1][key] + (total_g - s), 2)
            scb = round(sum(r["Ingresos bancarios"] or 0 for r in block), 2)
            block[-1]["Ingresos bancarios"] = round(
                (block[-1]["Ingresos bancarios"] or 0) + (c["importe"] - scb), 2
            )
            for r in block:
                ib = r["Ingresos bancarios"] or 0
                iibb_prev = iibb_residual(
                    r["Total de la venta"],
                    ib,
                    r["Retenciones de IVA"],
                    r["Retenciones de Ganancias"],
                    r["Retenciones SUSS"],
                )
                _aplicar_osseg_a_fila(r, ret_iibb_prev=iibb_prev)

    # ── Luego: pago único cliente+mes (N FCT → 1 crédito) ───────────────────
    ok.extend(
        _match_pago_unico_mes(ventas, cobranzas, factura_ret, grupos, used_facturas)
    )

    def try_match_bank(v: dict, g: dict | None) -> dict | None:
        ref = g["fecha"] if g else v["fecha"]
        targets = [
            expected_bank(v, g, _iibb_proxy_8(v)),
            expected_bank(v, g, float(v.get("ret_iibb_cm") or 0)),
            expected_bank(v, g, 0.0),
            v["total"],
        ]
        seen: set[float] = set()
        uniq: list[float] = []
        for t in targets:
            if t not in seen:
                seen.add(t)
                uniq.append(t)
        best, exp = _find_bank_multi(cobranzas, uniq, v.get("cuit"), ref, tol=TOL)
        if best:
            best["usado"] = True
            return {"credito": best, "esperado": exp}
        # Ecuación residual: total ≈ crédito + AFIP + IIBB razonable
        ret_iva = g["iva"] if g else 0.0
        ret_gan = g["gan"] if g else 0.0
        ret_ss = g["ss"] if g else 0.0
        sum_afip = ret_iva + ret_gan + ret_ss
        tol = _tol_pago_unico(v["total"])
        best_c = None
        best_sc = 1e18
        for c in cobranzas:
            if c["usado"]:
                continue
            if v.get("cuit") and c.get("cuit") and v["cuit"] != c["cuit"]:
                continue
            if abs((c["fecha"] - ref).days) > 45:
                continue
            ok, sc = _lote_cierra_con_credito([v], c["importe"], ret_iva, ret_gan, ret_ss)
            if ok and sc < best_sc:
                best_sc = sc
                best_c = c
            # también diff directo vs targets con tol relativa
            for t in uniq:
                d = abs(c["importe"] - t)
                if d <= tol and d < best_sc:
                    best_sc = d
                    best_c = c
        if best_c:
            best_c["usado"] = True
            return {"credito": best_c, "esperado": best_c["importe"]}
        return None

    for v in sorted(ventas, key=lambda x: (x["fecha"], x["nro"])):
        if v["id"] in used_facturas:
            continue
        g = factura_ret.get(v["id"])
        m = try_match_bank(v, g) if g else None

        if m and g:
            c = m["credito"]
            iibb = iibb_residual(v["total"], c["importe"], g["iva"], g["gan"], g["ss"])
            ok.append(
                _row_detalle(
                    estado="Calza OK",
                    v=v,
                    g=g,
                    c=c,
                    ret_iva=g["iva"],
                    ret_gan=g["gan"],
                    ret_ss=g["ss"],
                    ret_iibb=iibb,
                    credito=c["importe"],
                    iibb_origen=(
                        f"IIBB residual cobro (cierra eq.); CM mes prorrata "
                        f"${v.get('ret_iibb_cm') or 0:,.2f}"
                    ),
                )
            )
            used_facturas.add(v["id"])
        elif g and not m:
            # No marcar usado aún: puede entrar en pago único residual / lote abajo
            sin_match.append(v)
        else:
            matched = False
            if v["cuit"] == CUIT_EXPERTA:
                for g2 in grupos:
                    if g2["usado"] or g2["cuit"] != CUIT_EXPERTA:
                        continue
                    targets = [
                        expected_bank(v, g2, _iibb_proxy_8(v)),
                        expected_bank(v, g2, float(v.get("ret_iibb_cm") or 0)),
                        expected_bank(v, g2, 0.0),
                    ]
                    c_e, _ = _find_bank_multi(
                        cobranzas,
                        targets,
                        CUIT_EXPERTA,
                        g2["fecha"],
                        tol=_tol_pago_unico(v["total"]),
                    )
                    if c_e:
                        c_e["usado"] = True
                        g2["usado"] = True
                        for r in g2["ids"]:
                            r["usado"] = True
                        iibb = iibb_residual(
                            v["total"], c_e["importe"], g2["iva"], g2["gan"], g2["ss"]
                        )
                        ok.append(
                            _row_detalle(
                                estado="Calza OK",
                                v=v,
                                g=g2,
                                c=c_e,
                                ret_iva=g2["iva"],
                                ret_gan=g2["gan"],
                                ret_ss=g2["ss"],
                                ret_iibb=iibb,
                                credito=c_e["importe"],
                                iibb_origen=(
                                    f"IIBB residual cobro (cierra eq.); CM mes prorrata "
                                    f"${v.get('ret_iibb_cm') or 0:,.2f}"
                                ),
                            )
                        )
                        used_facturas.add(v["id"])
                        matched = True
                        break
            if matched:
                continue
            m2 = try_match_bank(v, None)
            if m2:
                c = m2["credito"]
                iibb = iibb_residual(v["total"], c["importe"], 0.0, 0.0, 0.0)
                ok.append(
                    _row_detalle(
                        estado="Calza OK (sin ret AFIP)",
                        v=v,
                        g=None,
                        c=c,
                        ret_iva=0.0,
                        ret_gan=0.0,
                        ret_ss=0.0,
                        ret_iibb=iibb if abs(iibb) > TOL else 0.0,
                        credito=c["importe"],
                        iibb_origen=(
                            "IIBB residual cobro"
                            if abs(iibb) > TOL
                            else "Sin IIBB (crédito ≈ total factura)"
                        ),
                    )
                )
                used_facturas.add(v["id"])
                continue
            best = None
            bestdiff = 1e18
            for c in cobranzas:
                if c["usado"]:
                    continue
                if v["cuit"] and c["cuit"] and v["cuit"] != c["cuit"]:
                    continue
                for exp_try in (
                    v["total"],
                    expected_bank(v, None, _iibb_proxy_8(v)),
                    expected_bank(v, None, float(v.get("ret_iibb_cm") or 0)),
                ):
                    d = abs(c["importe"] - exp_try)
                    if d < bestdiff:
                        bestdiff = d
                        best = (c, exp_try)
            if best and bestdiff <= _tol_pago_unico(v["total"]):
                c, _exp_try = best
                c["usado"] = True
                iibb = iibb_residual(v["total"], c["importe"], 0.0, 0.0, 0.0)
                ok.append(
                    _row_detalle(
                        estado="Calza OK (aprox)",
                        v=v,
                        g=None,
                        c=c,
                        ret_iva=0.0,
                        ret_gan=0.0,
                        ret_ss=0.0,
                        ret_iibb=iibb if abs(iibb) > TOL else 0.0,
                        credito=c["importe"],
                        iibb_origen="IIBB residual cobro",
                    )
                )
                used_facturas.add(v["id"])
            else:
                sin_match.append(v)

    # ── Reintento pago único sobre lo que quedó (p.ej. con ret AFIP pero sin banco 1:1)
    again = _match_pago_unico_mes(ventas, cobranzas, factura_ret, grupos, used_facturas)
    ok.extend(again)
    sin_match = [v for v in sin_match if v["id"] not in used_facturas]

    # ── Fallback lote Provincia/BHN (misma ecuación; etiqueta pago único si cierra)
    for cuit_lote, _nombre in (
        (CUIT_PROVINCIA, "PROVINCIA ART"),
        (CUIT_BHN, "BHN"),
        (CUIT_EXPERTA, "EXPERTA"),
        (CUIT_CAJA, "CAJA"),
    ):
        # Agrupar libres por mes
        by_mes: dict[tuple[int, int], list[dict]] = defaultdict(list)
        for v in ventas:
            if v.get("cuit") != cuit_lote or v["id"] in used_facturas:
                continue
            by_mes[_ym(v["fecha"])].append(v)
        for ym, lote in by_mes.items():
            if len(lote) < 2:
                continue
            lote = sorted(lote, key=lambda x: (x["fecha"], x["nro"]))
            creds_c = _creditos_cliente_ventana(cobranzas, cuit_lote, ym)
            if not creds_c:
                continue
            ret_iva, ret_gan, ret_ss, gs_afip = _rets_afip_para_lote(
                lote, factura_ret, grupos
            )
            sum_tot = round(sum(v["total"] for v in lote), 2)
            sum_iibb_cm = round(sum(float(v.get("ret_iibb_cm") or 0) for v in lote), 2)
            sum_iibb_8 = round(sum(_iibb_proxy_8(v) for v in lote), 2)
            sum_afip = round(ret_iva + ret_gan + ret_ss, 2)
            tol = _tol_pago_unico(sum_tot)
            best_c = None
            best_d = 1e18
            for c in creds_c:
                for t in (
                    round(sum_tot - sum_afip - sum_iibb_cm, 2),
                    round(sum_tot - sum_afip - sum_iibb_8, 2),
                    round(sum_tot - sum_afip, 2),
                    sum_tot,
                ):
                    d = abs(c["importe"] - t)
                    if d < best_d:
                        best_d = d
                        best_c = c
                iibb_imp = round(sum_tot - c["importe"] - sum_afip, 2)
                max_iibb = max(sum_iibb_cm * 1.15, sum_tot * 0.20, tol)
                if -tol <= iibb_imp <= max_iibb:
                    pen = abs(iibb_imp - sum_iibb_cm)
                    score = pen * 0.01
                    if score < best_d:
                        best_d = score
                        best_c = c
            if not best_c:
                continue
            ok_close = best_d <= tol
            ok_gap = False
            if len(creds_c) == 1 and best_c["importe"] >= sum_tot * 0.80 and best_c["importe"] <= sum_tot * 1.02:
                gap_iibb = round(sum_tot - best_c["importe"] - sum_afip, 2)
                ok_gap = -tol <= gap_iibb <= max(sum_iibb_cm * 1.15, sum_tot * 0.20)
            if not ok_close and not ok_gap:
                continue
            best_c["usado"] = True
            for g in gs_afip:
                g["usado"] = True
                for r in g.get("ids") or []:
                    r["usado"] = True
            mm_aaaa = f"{ym[1]:02d}/{ym[0]}"
            obs = f"Pago único mes {mm_aaaa} — {len(lote)} facturas"
            rows = _asignar_pago_unico(
                lote, best_c, ret_iva, ret_gan, ret_ss, obs
            )
            for r, v in zip(rows, lote):
                used_facturas.add(v["id"])
                ok.append(r)
        sin_match = [x for x in sin_match if x["id"] not in used_facturas]

    # Facturas con ret AFIP pero sin banco (quedaron en sin_match)
    for v in list(sin_match):
        if v["id"] in used_facturas:
            continue
        g = factura_ret.get(v["id"])
        if g:
            pendientes.append(
                _row_detalle(
                    estado="Falta banco",
                    v=v,
                    g=g,
                    c=None,
                    ret_iva=g["iva"],
                    ret_gan=g["gan"],
                    ret_ss=g["ss"],
                    ret_iibb=float(v.get("ret_iibb_cm") or 0),
                    credito=None,
                    obs_extra=(
                        f"Ret AFIP matchean; crédito esperado ~ "
                        f"${expected_bank(v, g, _iibb_proxy_8(v)):,.2f}"
                    ),
                    iibb_origen=v.get("iibb_metodo") or "",
                )
            )
            used_facturas.add(v["id"])
            sin_match = [x for x in sin_match if x["id"] != v["id"]]

    for v in sin_match:
        if v["id"] in used_facturas:
            continue
        pendientes.append(
            _row_detalle(
                estado="Falta retención / banco",
                v=v,
                g=None,
                c=None,
                ret_iva=0.0,
                ret_gan=0.0,
                ret_ss=0.0,
                ret_iibb=float(v.get("ret_iibb_cm") or 0),
                credito=None,
                obs_extra="Sin match de retenciones AFIP ni crédito bancario",
                iibb_origen=v.get("iibb_metodo") or "",
            )
        )

    for g in grupos:
        if g["usado"]:
            continue
        osseg_g, iibb_g, note_g = aplicar_osseg(g["ss"], 0.0)
        obs_g = f"Retención AFIP sin factura (comp {g.get('comprobante') or '-'})"
        if note_g:
            obs_g = f"{obs_g} | {note_g}"
        pendientes.append(
            {
                "Fecha de la venta": None,
                "Cliente": g["agente"],
                "Total de la venta": None,
                "Retenciones de IVA": g["iva"],
                "Retenciones de Ganancias": g["gan"],
                "Retenciones de Ingresos Brutos": iibb_g,
                "Retenciones SUSS": g["ss"],
                "Retenciones OSSEG": osseg_g,
                "Ingresos bancarios": None,
                "Nº factura": None,
                "Fecha cobro": None,
                "Banco": "",
                "Observación / Match": obs_g,
                "Estado": "Sin factura",
                "Diferencia": None,
            }
        )

    for c in cobranzas:
        if c["usado"]:
            continue
        pendientes.append(
            {
                "Fecha de la venta": None,
                "Cliente": CUIT_NOMBRES.get(c["cuit"] or "", c["concepto"][:40]),
                "Total de la venta": None,
                "Retenciones de IVA": None,
                "Retenciones de Ganancias": None,
                "Retenciones de Ingresos Brutos": None,
                "Retenciones SUSS": None,
                "Retenciones OSSEG": None,
                "Ingresos bancarios": c["importe"],
                "Nº factura": None,
                "Fecha cobro": c["fecha"],
                "Banco": c["banco"],
                "Observación / Match": f"{fmt_ddmm(c['fecha'])} | {c['concepto'][:80]}",
                "Estado": "Crédito sin factura",
                "Diferencia": None,
            }
        )

    return ok, pendientes


def _limpiar_df(rows: list[dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=COLS_DETALLE)
    df = pd.DataFrame(rows)
    for col in COLS_DETALLE:
        if col not in df.columns:
            df[col] = None
    return df[COLS_DETALLE]


# ── Excel salida ─────────────────────────────────────────────────────────────


def main() -> None:
    print("Cargando ventas...")
    ventas = cargar_ventas()
    print(f"  {len(ventas)} facturas")

    meses = {(v["fecha"].year, v["fecha"].month) for v in ventas}
    print("Cargando IIBB del Convenio Multilateral...")
    iibb_mes = cargar_iibb_cm(meses)
    metodo_iibb = asignar_iibb_prorrateo(ventas, iibb_mes)
    tot_iibb_cm = round(sum(iibb_mes.values()), 2)
    print(f"  IIBB CM meses={len(iibb_mes)} total CM=${tot_iibb_cm:,.2f}")
    print(f"  Método: {metodo_iibb}")

    print("Cargando retenciones AFIP...")
    rets = cargar_retenciones()
    grupos = agrupar_retenciones(rets)
    print(f"  {len(rets)} retenciones -> {len(grupos)} clusters")

    print("Extrayendo créditos Galicia...")
    creds_g = extraer_creditos_galicia(PDF_GALICIA)
    print(f"  {len(creds_g)} créditos ({sum(1 for c in creds_g if c['es_cobranza'])} cobranzas)")

    pdf_m = PDF_MACRO if PDF_MACRO.exists() else PDF_MACRO_FB
    print(f"Extrayendo créditos Macro ({pdf_m.name})...")
    creds_m = extraer_creditos_macro(pdf_m)
    print(f"  {len(creds_m)} créditos ({sum(1 for c in creds_m if c['es_cobranza'])} cobranzas)")

    creds = creds_g + creds_m
    print("Calzando...")
    ok, pend = calzar(ventas, grupos, creds)

    tot_ventas = round(sum(v["total"] for v in ventas), 2)
    tot_ret_iva = round(sum(r["importe"] for r in rets if r["tipo"] == "IVA"), 2)
    tot_ret_gan = round(sum(r["importe"] for r in rets if r["tipo"] == "GAN"), 2)
    tot_ret_ss = round(sum(r["importe"] for r in rets if r["tipo"] == "SS"), 2)
    cobranzas = [c for c in creds if c["es_cobranza"]]
    tot_cob = round(sum(c["importe"] for c in cobranzas), 2)
    tot_ok_banco = round(sum(r["Ingresos bancarios"] or 0 for r in ok), 2)
    n_ok = sum(1 for r in ok if str(r["Estado"]).startswith("Calza OK"))
    n_parcial = sum(1 for r in ok if "parcial" in str(r["Estado"]).lower())
    n_exact_diff = sum(
        1
        for r in ok
        if str(r["Estado"]).startswith("Calza OK")
        and r.get("Diferencia") is not None
        and abs(r["Diferencia"]) <= TOL
    )
    n_pend_fac = sum(1 for p in pend if p["Estado"] in {"Falta banco", "Falta retención / banco"})
    pct_ok = round(100.0 * n_ok / len(ventas), 1) if ventas else 0.0

    # Detalle completo = calces + pendientes de factura (para 1 fila por factura)
    fac_rows = ok + [
        p for p in pend if p["Estado"] in {"Falta banco", "Falta retención / banco"}
    ]
    fac_rows.sort(
        key=lambda r: (
            r["Fecha de la venta"] or date(2099, 1, 1),
            r["Nº factura"] or 0,
        )
    )

    tot_iibb_asig = round(
        sum(float(r.get("Retenciones de Ingresos Brutos") or 0) for r in fac_rows),
        2,
    )
    tot_osseg = round(
        sum(float(r.get("Retenciones OSSEG") or 0) for r in fac_rows),
        2,
    )
    tot_ss_detalle = round(
        sum(float(r.get("Retenciones SUSS") or 0) for r in fac_rows),
        2,
    )
    tot_iva_detalle = round(
        sum(float(r.get("Retenciones de IVA") or 0) for r in fac_rows),
        2,
    )
    tot_gan_detalle = round(
        sum(float(r.get("Retenciones de Ganancias") or 0) for r in fac_rows),
        2,
    )
    n_clip_iibb = sum(
        1
        for r in fac_rows
        if "IIBB clip a 0" in str(r.get("Observación / Match") or "")
    )
    gap_iibb = round(tot_iibb_asig - tot_iibb_cm, 2)
    tot_iibb_pre_osseg = round(tot_iibb_asig + tot_osseg, 2)

    df_detalle = _limpiar_df(fac_rows)
    df_pend = _limpiar_df(pend)
    df_creds = pd.DataFrame(
        [
            {
                "Banco": c["banco"],
                "Fecha": c["fecha"],
                "Importe": c["importe"],
                "Categoría": c["categoria"],
                "CUIT": c["cuit"] or "",
                "Concepto": c["concepto"],
                "Usado en calce": "Sí" if c["usado"] else "No",
            }
            for c in sorted(creds, key=lambda x: (x["fecha"], x["banco"]))
        ]
    )
    df_rets = pd.DataFrame(
        [
            {
                "Tipo": r["tipo"],
                "Fecha": r["fecha"],
                "CUIT agente": r["cuit"],
                "Agente": r["agente"],
                "Importe": r["importe"],
                "Comprobante": r["comprobante"],
                "Usado en calce": "Sí" if r["usado"] else "No",
            }
            for r in sorted(rets, key=lambda x: (x["fecha"], x["tipo"]))
        ]
    )
    df_iibb = pd.DataFrame(
        [
            {
                "Período": f"{m:02d}-{y}",
                "Retenciones IIBB CM (clientes)": iibb_mes.get((y, m), 0.0),
                "Facturas del mes": sum(
                    1 for v in ventas if v["fecha"].year == y and v["fecha"].month == m
                ),
                "Total ventas mes": round(
                    sum(v["total"] for v in ventas if v["fecha"].year == y and v["fecha"].month == m),
                    2,
                ),
            }
            for y, m in sorted(meses)
        ]
    )

    por_mes_rows = []
    for y, m in sorted(meses):
        rows_m = [
            r
            for r in fac_rows
            if r.get("Fecha de la venta")
            and r["Fecha de la venta"].year == y
            and r["Fecha de la venta"].month == m
        ]
        n_ok_m = sum(1 for r in rows_m if str(r.get("Estado") or "").startswith("Calza OK"))
        por_mes_rows.append(
            {
                "Mes": f"{MESES_ES.get(m, str(m))} {y}",
                "Período": f"{m:02d}-{y}",
                "Nº facturas": len(rows_m),
                "Calce OK": n_ok_m,
                "Total ventas": round(
                    sum(float(r.get("Total de la venta") or 0) for r in rows_m), 2
                ),
                "Ret IVA": round(
                    sum(float(r.get("Retenciones de IVA") or 0) for r in rows_m), 2
                ),
                "Ret Ganancias": round(
                    sum(float(r.get("Retenciones de Ganancias") or 0) for r in rows_m), 2
                ),
                "Ret IIBB (post-OSSEG)": round(
                    sum(float(r.get("Retenciones de Ingresos Brutos") or 0) for r in rows_m),
                    2,
                ),
                "Ret SUSS": round(
                    sum(float(r.get("Retenciones SUSS") or 0) for r in rows_m), 2
                ),
                "Ret OSSEG": round(
                    sum(float(r.get("Retenciones OSSEG") or 0) for r in rows_m), 2
                ),
                "Ingresos bancarios": round(
                    sum(float(r.get("Ingresos bancarios") or 0) for r in rows_m), 2
                ),
                "IIBB CM control": iibb_mes.get((y, m), 0.0),
            }
        )
    df_por_mes = pd.DataFrame(por_mes_rows)

    resumen_estados = (
        pd.DataFrame(
            [
                {
                    "Estado": r["Estado"],
                    "Cantidad": 1,
                    "Total facturas": r.get("Total de la venta") or 0,
                    "Ingresos bancarios": r.get("Ingresos bancarios") or 0,
                }
                for r in ok + pend
            ]
        )
        .groupby("Estado", as_index=False)
        .agg(
            Cantidad=("Cantidad", "sum"),
            Total_facturas=("Total facturas", "sum"),
            Ingresos=("Ingresos bancarios", "sum"),
        )
        .rename(columns={"Total_facturas": "Total facturas", "Ingresos": "Ingresos bancarios"})
    )

    criterio = pd.DataFrame(
        [
            {
                "Ítem": "Ecuación",
                "Detalle": (
                    "Factura = Banco + Ret IVA + Ret Gan + Ret IIBB + Ret SUSS + Ret OSSEG"
                ),
            },
            {
                "Ítem": "OSSEG (Caja de Seguros)",
                "Detalle": (
                    "Ret OSSEG = Ret SUSS (mismo monto; no figura en AFIP). "
                    "Ret IIBB nueva = Ret IIBB anterior − OSSEG (clip a 0 si negativo)."
                ),
            },
            {
                "Ítem": "IIBB",
                "Detalle": (
                    "Con cobro: residual pre-OSSEG = Total−Banco−Ret AFIP (cierra eq. pre-OSSEG; "
                    "suele ~8% neto en Caja/Experta). Luego se resta OSSEG. "
                    "Sin cobro: prorrateo fila «Retenciones» (clientes) del CM s/total ventas del mes, "
                    "luego − OSSEG. Control vs CM en KPI. Ret Bancos CM = SIRCREB (aparte)."
                ),
            },
            {
                "Ítem": "Prorrateo IIBB CM (fallback)",
                "Detalle": metodo_iibb,
            },
            {
                "Ítem": "Match retención→factura",
                "Detalle": "CUIT + fecha; SS≈1% neto; IVA ret≈80% IVA; GAN≈2% neto",
            },
            {
                "Ítem": "Match factura→banco",
                "Detalle": (
                    f"Crédito ≈ Total − Ret AFIP − IIBB (heurística ~8% neto o CM); "
                    f"tol. ${TOL:.0f}; combos ±${TOL_COMBO:.0f}"
                ),
            },
            {
                "Ítem": "Pago único (cliente+mes)",
                "Detalle": (
                    "Si N≥2 FCT del mismo cliente en el mismo mes y hay 1 crédito: "
                    "suma(totales) ≈ crédito + ret IVA/Gan/IIBB/SUSS (tol $50 o 0,1%). "
                    "Obs: «Pago único mes MM/AAAA — N facturas». También match 1:1."
                ),
            },
            {
                "Ítem": "Fuentes",
                "Detalle": "Ventas + XLS AFIP (IVA/Gan/SUSS) + CM IIBB + Galicia + Macro",
            },
        ]
    )

    fechas = [v["fecha"] for v in ventas]
    periodo = f"{min(fechas).strftime('%d/%m/%Y')} → {max(fechas).strftime('%d/%m/%Y')}"

    col_moneda = [
        "Total de la venta",
        "Retenciones de IVA",
        "Retenciones de Ganancias",
        "Retenciones de Ingresos Brutos",
        "Retenciones SUSS",
        "Retenciones OSSEG",
        "Ingresos bancarios",
        "Diferencia",
        "Total facturas",
        "Importe",
        "Retenciones IIBB CM (clientes)",
        "Total ventas mes",
        "Crédito esperado",
        "Total ventas",
        "Ret IVA",
        "Ret Ganancias",
        "Ret IIBB (post-OSSEG)",
        "Ret SUSS",
        "Ret OSSEG",
        "IIBB CM control",
    ]
    col_fecha = ["Fecha de la venta", "Fecha cobro", "Fecha"]

    guardar_informe_excel(
        OUT,
        titulo="Calce Ventas + Retenciones vs Bancos — GRUPO MERIDIEM",
        subtitulo=(
            "Factura = Banco + Ret IVA + Ret Gan + Ret IIBB + Ret SUSS + Ret OSSEG. "
            "OSSEG = SUSS; IIBB post = IIBB pre − OSSEG. "
            f"OSSEG total ${tot_osseg:,.2f} · IIBB post ${tot_iibb_asig:,.2f} "
            f"(pre≈${tot_iibb_pre_osseg:,.2f}) · CM ${tot_iibb_cm:,.2f} (gap ${gap_iibb:,.2f})"
        ),
        periodo=periodo,
        kpis=[
            ("Facturas", len(ventas), "int"),
            ("Calce OK", n_ok, "int"),
            ("% calce OK", pct_ok, "text"),
            ("Calce OK diff≤$5", n_exact_diff, "int"),
            ("Calce parcial (lotes)", n_parcial, "int"),
            ("Pendientes (falta banco/ret)", n_pend_fac, "int"),
            ("Total ventas", tot_ventas, "money"),
            ("Ret. IVA (detalle)", tot_iva_detalle, "money"),
            ("Ret. Ganancias (detalle)", tot_gan_detalle, "money"),
            ("Ret. SUSS (detalle)", tot_ss_detalle, "money"),
            ("Ret. OSSEG (=SUSS)", tot_osseg, "money"),
            ("IIBB post-OSSEG", tot_iibb_asig, "money"),
            ("IIBB pre-OSSEG (≈post+OSSEG)", tot_iibb_pre_osseg, "money"),
            ("IIBB CM Retenciones (control)", tot_iibb_cm, "money"),
            ("Gap IIBB post vs CM", gap_iibb, "money"),
            ("Filas IIBB clip a 0", n_clip_iibb, "int"),
            ("Ret. IVA AFIP (fuente)", tot_ret_iva, "money"),
            ("Ret. Ganancias AFIP (fuente)", tot_ret_gan, "money"),
            ("Ret. SUSS AFIP (fuente)", tot_ret_ss, "money"),
            ("Cobranzas bancarias", tot_cob, "money"),
            ("Ingresos bancarios vinculados", tot_ok_banco, "money"),
            ("Tolerancia $", TOL, "money"),
        ],
        resumenes=[
            ("Por mes", df_por_mes),
            ("Por estado", resumen_estados),
            ("IIBB CM por mes", df_iibb),
            ("Criterio de match", criterio),
        ],
        detalle=df_detalle if not df_detalle.empty else pd.DataFrame({"Info": ["Sin filas"]}),
        hoja_detalle="Detalle",
        col_moneda=col_moneda,
        col_fecha=col_fecha,
        total_col="Total de la venta",
    )

    wb = load_workbook(OUT)

    ws_mes = wb.create_sheet("Por mes", 1)
    fila = _escribir_encabezado_hoja(
        ws_mes,
        "Resumen mensual — retenciones e ingresos",
        "Ret IIBB ya restado OSSEG · OSSEG = SUSS (Caja de Seguros)",
    )
    _escribir_tabla(
        ws_mes,
        df_por_mes,
        fila,
        col_moneda=[
            "Total ventas",
            "Ret IVA",
            "Ret Ganancias",
            "Ret IIBB (post-OSSEG)",
            "Ret SUSS",
            "Ret OSSEG",
            "Ingresos bancarios",
            "IIBB CM control",
        ],
        total_col="Total ventas",
    )

    ws_p = wb.create_sheet("Pendientes", 2)
    fila = _escribir_encabezado_hoja(
        ws_p,
        "Pendientes / no calza",
        "Falta banco · Falta retención · Sin factura · Crédito sin factura",
    )
    if not df_pend.empty:
        _escribir_tabla(
            ws_p,
            df_pend,
            fila,
            col_moneda=col_moneda,
            col_fecha=col_fecha,
        )
    else:
        ws_p.cell(fila, 1, "Sin pendientes").font = BODY_FONT

    ws_c = wb.create_sheet("Créditos bancos", 3)
    fila = _escribir_encabezado_hoja(
        ws_c,
        "Créditos Galicia + Macro",
        "Cobranzas de clientes vs fondos/traspasos",
    )
    _escribir_tabla(
        ws_c,
        df_creds,
        fila,
        col_moneda=["Importe"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    ws_r = wb.create_sheet("Retenciones AFIP", 4)
    fila = _escribir_encabezado_hoja(
        ws_r,
        "Retenciones sufridas AFIP (IVA + Ganancias + SUSS)",
        "Fuente: XLS AFIP del grupo · OSSEG no figura en AFIP (se deriva de SUSS)",
    )
    _escribir_tabla(
        ws_r,
        df_rets,
        fila,
        col_moneda=["Importe"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    ws_i = wb.create_sheet("IIBB CM", 5)
    fila = _escribir_encabezado_hoja(
        ws_i,
        "Retenciones IIBB — Convenio Multilateral",
        "Fila «Retenciones» (clientes) del papel CM · prorrateo s/total ventas del mes",
    )
    _escribir_tabla(
        ws_i,
        df_iibb,
        fila,
        col_moneda=["Retenciones IIBB CM (clientes)", "Total ventas mes"],
        total_col="Retenciones IIBB CM (clientes)",
    )

    ws0 = wb["Resumen"]
    last = ws0.max_row + 2
    ws0.cell(last, 1, "Nota IIBB / OSSEG").font = SECTION_FONT
    ws0.cell(
        last + 1,
        1,
        "OSSEG = SUSS (mismo monto; no en AFIP). IIBB post = IIBB pre − OSSEG (clip≥0). "
        "Ecuación: Total = Banco + IVA + Gan + IIBB + SUSS + OSSEG. "
        f"OSSEG ${tot_osseg:,.2f} restado de IIBB (pre≈${tot_iibb_pre_osseg:,.2f} → "
        f"post ${tot_iibb_asig:,.2f}; clips a 0: {n_clip_iibb}). "
        f"Control CM ${tot_iibb_cm:,.2f} vs IIBB post ${tot_iibb_asig:,.2f} (gap ${gap_iibb:,.2f}). "
        f"Calce OK: {n_ok}/{len(ventas)} · Parciales: {n_parcial} · Pendientes: {n_pend_fac}. "
        "Resumen mensual en hoja «Por mes» (y bloque en Resumen). "
        "Ret Bancos del CM (SIRCREB) no entra en esta columna.",
    ).font = SUB_FONT

    wb.save(OUT)

    print("OUT", OUT)
    print(f"OK {n_ok}/{len(ventas)} ({pct_ok}%) · parcial {n_parcial} · pend {n_pend_fac}")
    print(
        f"OSSEG ${tot_osseg:,.2f} (=SUSS detalle ${tot_ss_detalle:,.2f}) | "
        f"IIBB pre~${tot_iibb_pre_osseg:,.2f} -> post ${tot_iibb_asig:,.2f} | "
        f"CM ${tot_iibb_cm:,.2f} | gap ${gap_iibb:,.2f} | clips {n_clip_iibb}"
    )
    print(f"Ventas ${tot_ventas:,.2f} | Cobranzas ${tot_cob:,.2f}")
    samples = [r for r in ok if r.get("Fecha cobro") and str(r["Estado"]).startswith("Calza OK")]
    print("Ejemplos cobro:")
    for r in samples[:5]:
        print(
            " ",
            texto_cobro(r["Nº factura"], r["Fecha cobro"], r["Banco"]),
            f"| tot={r['Total de la venta']:,.2f} banco={r['Ingresos bancarios']:,.2f} "
            f"IIBB={r['Retenciones de Ingresos Brutos']:,.2f} "
            f"OSSEG={r.get('Retenciones OSSEG') or 0:,.2f} diff={r['Diferencia']}",
        )
    if not df_pend.empty:
        print("Pendientes por estado:")
        for est, grp in df_pend.groupby("Estado"):
            print(f"  {est}: {len(grp)}")
    print("Por mes:")
    for row in por_mes_rows:
        print(
            f"  {row['Mes']}: OSSEG=${row['Ret OSSEG']:,.2f} "
            f"IIBB=${row['Ret IIBB (post-OSSEG)']:,.2f} "
            f"OK={row['Calce OK']}/{row['Nº facturas']}"
        )

if __name__ == "__main__":
    main()
