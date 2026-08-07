# -*- coding: utf-8 -*-
"""Retenciones/percepciones IIBB bancarias Galicia + Macro (Meridiem)."""
from __future__ import annotations

import calendar
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    COLOR_PRIMARIO,
    MONEY_FMT,
    ZEBRA,
    _escribir_encabezado_hoja,
    _escribir_tabla,
    guardar_informe_excel,
)
from _extraer_retenciones_galicia_macro_meridiem import (  # noqa: E402
    PDF_GALICIA,
    PDF_MACRO_CANDIDATES,
    _sort_mes,
    extraer_retenciones_macro,
    filas_banco,
    resolver_pdf_macro,
)
from _extraer_retenciones_galicia_meridiem import (  # noqa: E402
    MESES,
    parse_ar_money,
    extraer_retenciones as extraer_retenciones_galicia,
)

OUT = Path(r"C:\Users\recep\Desktop\Retenciones_IIBB_Bancarias_Meridiem.xlsx")
PDF_TUCUMAN_BANCARIAS = Path(r"c:\Users\recep\Desktop\retenciones tucuman bancarias.pdf")

TIPOS_IIBB = {
    "Ret. IIBB SIRCREB",
    "Ret. IIBB Tucumán",
    "Ret. IIBB s/ créditos",
    "Ret. IIBB",
    "Retención IIBB",
    "Percepción IIBB",
}

RE_PDF_TUC = re.compile(
    r"^(\d{6})\s+(.+?)\s+(\d{2}-\d{8}-\d)\s+(\d+)\s+([\d.]+,\d{2})\s*$"
)


def es_iibb(tipo: str, desc: str = "") -> bool:
    t = (tipo or "").strip()
    if t in TIPOS_IIBB:
        return True
    blob = f"{t} {desc}".upper()
    if re.search(r"25413|PERCEP\.?\s*IVA|RET\.?\s*IVA|LEY\s*25", blob):
        return False
    if re.search(r"SIRCREB|IIBB|ING\.?\s*BRUTOS|INGRESOS\s+BRUTOS", blob):
        # excluir AFIP genérico / IVA si se coló
        if re.search(r"\bIVA\b", blob) and not re.search(r"IIBB|SIRCREB|BRUTOS", t.upper()):
            return False
        return True
    return False


def filtrar_iibb(rets: list[dict]) -> list[dict]:
    return [
        r
        for r in rets
        if es_iibb(r.get("tipo", ""), r.get("descripcion_raw", ""))
    ]


def map_banco_pdf(nombre: str) -> str:
    u = (nombre or "").upper()
    if "GALICIA" in u:
        return "Galicia"
    if "BANSUD" in u or "MACRO" in u:
        return "Macro"
    if "INDUSTRIAL AND COMMERCIAL" in u or "ICBC" in u:
        return "ICBC"
    return (nombre or "Banco").strip()[:40]


def parse_pdf_tucuman_bancarias(pdf_path: Path) -> list[dict]:
    """Listado AFIP/Rentas: período YYYYMM + banco + CBU + importe (totales mensuales)."""
    rows: list[dict] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    for line in text.splitlines():
        m = RE_PDF_TUC.match(line.strip())
        if not m:
            continue
        periodo, banco_raw, cuit_agente, cbu, importe_s = m.groups()
        anio, mes_n = int(periodo[:4]), int(periodo[4:])
        ultimo = calendar.monthrange(anio, mes_n)[1]
        fecha = date(anio, mes_n, ultimo)
        mes_label = f"{MESES[mes_n]} {anio}"
        banco = map_banco_pdf(banco_raw)
        rows.append(
            {
                "periodo": periodo,
                "fecha": fecha,
                "mes": mes_label,
                "banco": banco,
                "banco_raw": banco_raw.strip(),
                "cuit_agente": cuit_agente,
                "cbu": cbu,
                "importe": round(parse_ar_money(importe_s), 2),
                "tipo": "Ret. IIBB Tucumán",
            }
        )
    return rows


def _fecha_key(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _fmt_fecha(v) -> str:
    d = _fecha_key(v)
    return d.strftime("%d/%m/%Y") if d else str(v or "")


def cargar_filas_excel(ruta: Path) -> list[dict]:
    wb = load_workbook(ruta, data_only=True)
    ws = wb["Lista"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    filas: list[dict] = []
    for r in range(2, ws.max_row + 1):
        vals = {headers[c - 1]: ws.cell(r, c).value for c in range(1, len(headers) + 1)}
        if not vals.get("Tipo") or vals.get("Tipo") == "TOTAL":
            continue
        fecha = vals.get("Fecha")
        filas.append(
            {
                "Fecha": _fmt_fecha(fecha),
                "Banco": vals.get("Banco"),
                "Tipo": vals.get("Tipo"),
                "Importe": round(float(vals.get("Importe") or 0), 2),
                "Descripción": vals.get("Descripción") or "",
                "Mes": vals.get("Mes") or "",
            }
        )
    return filas


def incorporar_pdf_tucuman(
    filas: list[dict], pdf_rows: list[dict]
) -> tuple[list[dict], dict]:
    """Mergea totales PDF Tucumán: no duplica si ya hay mismo banco+mes (suma) o fecha+importe."""
    out = [dict(f) for f in filas]
    sumas_tuc: dict[tuple[str, str], float] = defaultdict(float)
    keys_fecha_imp: set[tuple[str, str, float]] = set()
    for f in out:
        if f.get("Tipo") != "Ret. IIBB Tucumán":
            continue
        sumas_tuc[(f["Banco"], f["Mes"])] += float(f["Importe"])
        keys_fecha_imp.add((f["Banco"], f["Fecha"], round(float(f["Importe"]), 2)))

    matched = 0
    nuevas = 0
    importe_nuevo = 0.0
    for p in pdf_rows:
        banco, mes, imp = p["banco"], p["mes"], round(p["importe"], 2)
        fecha_s = p["fecha"].strftime("%d/%m/%Y")
        ya_mes = abs(round(sumas_tuc.get((banco, mes), 0.0), 2) - imp) < 0.02
        ya_exacto = (banco, fecha_s, imp) in keys_fecha_imp
        if ya_mes or ya_exacto:
            matched += 1
            marca = f" · confirmado PDF Tucumán {p['periodo']}"
            for f in out:
                if (
                    f.get("Tipo") == "Ret. IIBB Tucumán"
                    and f.get("Banco") == banco
                    and f.get("Mes") == mes
                    and marca not in (f.get("Descripción") or "")
                ):
                    f["Descripción"] = (f.get("Descripción") or "").rstrip() + marca
            continue
        desc = (
            f"Ret. IIBB Tucumán bancaria (PDF) · período {p['periodo']} · "
            f"{p['banco_raw']} · CBU {p['cbu']}"
        )
        out.append(
            {
                "Fecha": fecha_s,
                "Banco": banco,
                "Tipo": "Ret. IIBB Tucumán",
                "Importe": imp,
                "Descripción": desc,
                "Mes": mes,
            }
        )
        sumas_tuc[(banco, mes)] += imp
        keys_fecha_imp.add((banco, fecha_s, imp))
        nuevas += 1
        importe_nuevo += imp

    stats = {
        "pdf_filas": len(pdf_rows),
        "pdf_total": round(sum(p["importe"] for p in pdf_rows), 2),
        "matched": matched,
        "nuevas": nuevas,
        "importe_nuevo": round(importe_nuevo, 2),
    }
    return out, stats


def _hoja_por_mes(wb, filas: list[dict]) -> None:
    if "Por mes" in wb.sheetnames:
        del wb["Por mes"]
    ws = wb.create_sheet("Por mes", 1)
    fila = _escribir_encabezado_hoja(
        ws,
        "Retenciones IIBB bancarias — por mes",
        "GRUPO MERIDIEM SRL · Galicia + Macro + PDF Tucumán",
        f"{filas[0]['Mes']} → {filas[-1]['Mes']}" if filas else "",
    )
    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in filas:
        grupos[r["Mes"]].append(r)

    for mes_label in sorted(grupos.keys(), key=_sort_mes):
        items = grupos[mes_label]
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
        cell = ws.cell(fila, 1, mes_label)
        cell.fill = fill_mes
        cell.font = font_mes
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in (2, 3, 4):
            ws.cell(fila, col).fill = fill_mes
        fila += 1
        for idx, r in enumerate(items):
            ws.cell(fila, 1, r["Fecha"]).font = BODY_FONT
            ws.cell(fila, 2, r["Banco"]).font = BODY_FONT
            ws.cell(fila, 3, r["Tipo"]).font = BODY_FONT
            ws.cell(fila, 4, r["Importe"]).number_format = MONEY_FMT
            ws.cell(fila, 4).font = BODY_FONT
            if idx % 2 == 1:
                for col in (1, 2, 3, 4):
                    ws.cell(fila, col).fill = ZEBRA
            fila += 1
        tot = round(sum(x["Importe"] for x in items), 2)
        ws.cell(fila, 3, f"TOTAL {mes_label}").font = font_tot
        ws.cell(fila, 4, tot).number_format = MONEY_FMT
        ws.cell(fila, 4).font = font_tot
        fila += 2

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16


def guardar(ruta: Path, filas: list[dict], stats_pdf: dict | None = None) -> Path:
    df = pd.DataFrame(filas)
    tot = round(float(df["Importe"].sum()), 2) if len(df) else 0.0

    por_banco = (
        df.groupby("Banco", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )
    por_tipo = (
        df.groupby("Tipo", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )
    por_banco_tipo = (
        df.groupby(["Banco", "Tipo"], sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values(["Banco", "Total"], ascending=[True, False])
    )
    por_mes = (
        df.groupby("Mes", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )
    por_mes["_sk"] = por_mes["Mes"].map(_sort_mes)
    por_mes = por_mes.sort_values("_sk").drop(columns=["_sk"])

    por_banco_mes = (
        df.groupby(["Banco", "Mes"], sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )
    por_banco_mes["_sk"] = por_banco_mes["Mes"].map(_sort_mes)
    por_banco_mes = por_banco_mes.sort_values(["Banco", "_sk"]).drop(columns=["_sk"])

    tot_tuc = round(
        float(df.loc[df["Tipo"] == "Ret. IIBB Tucumán", "Importe"].sum()) if len(df) else 0.0,
        2,
    )
    banco_totales = {
        str(r["Banco"]): round(float(r["Total"]), 2) for _, r in por_banco.iterrows()
    }

    meses = sorted({r["Mes"] for r in filas}, key=_sort_mes) if filas else []
    periodo = f"{meses[0]} → {meses[-1]}" if meses else ""

    fechas = [r["Fecha"] for r in filas]
    rango_fechas = f"{min(fechas)} → {max(fechas)}" if fechas else ""

    kpis = [
        ("Movimientos IIBB", len(filas)),
        ("Total IIBB", tot),
        ("Total Tucumán", tot_tuc),
    ]
    for b in ("Galicia", "Macro", "ICBC"):
        if b in banco_totales:
            kpis.append((b, banco_totales[b]))

    sub = (
        "SIRCREB / IIBB Tucumán / Perc. IIBB (Galicia + Macro"
        + (" + PDF Tucumán bancarias" if stats_pdf else "")
        + "). Sin Ley 25413 ni IVA."
    )
    if stats_pdf:
        sub += (
            f" PDF: {stats_pdf['nuevas']} filas nuevas"
            f" / {stats_pdf['matched']} ya en extractos (dedupe)."
        )

    guardar_informe_excel(
        ruta,
        titulo="Retenciones IIBB bancarias — GRUPO MERIDIEM SRL",
        subtitulo=sub,
        periodo=periodo or rango_fechas,
        kpis=kpis,
        resumenes=[
            ("Totales por banco", por_banco),
            ("Por tipo", por_tipo),
            ("Por banco y tipo", por_banco_tipo),
            ("Por mes", por_mes),
            ("Por banco y mes", por_banco_mes),
        ],
        detalle=df[["Fecha", "Banco", "Tipo", "Importe", "Descripción", "Mes"]],
        hoja_detalle="Lista",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    wb = load_workbook(ruta)
    _hoja_por_mes(wb, filas)
    orden = ["Resumen", "Por mes", "Lista"]
    for i, name in enumerate(orden):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    wb.save(ruta)
    return ruta


def main() -> None:
    # Si ya hay Excel y llega el PDF Tucumán, partimos del Excel (incorporación).
    # Si no, regeneramos desde extractos Galicia/Macro.
    filas: list[dict] = []
    if OUT.exists() and PDF_TUCUMAN_BANCARIAS.exists():
        filas = cargar_filas_excel(OUT)
        print("FUENTE Excel existente", OUT)
    elif PDF_GALICIA.exists():
        try:
            pdf_m = resolver_pdf_macro()
            rets_g = filtrar_iibb(extraer_retenciones_galicia(PDF_GALICIA))
            rets_m = filtrar_iibb(extraer_retenciones_macro(pdf_m)) if pdf_m else []
            filas = filas_banco(rets_g, "Galicia") + filas_banco(rets_m, "Macro")
            print("FUENTE extractos Galicia+Macro")
        except Exception as e:
            print("WARN extractos:", e)
            filas = []
    if not filas and OUT.exists():
        filas = cargar_filas_excel(OUT)
        print("FUENTE Excel existente (fallback)", OUT)

    if not filas:
        raise SystemExit("No hay filas base (ni extractos ni Excel).")

    stats_pdf: dict | None = None
    if PDF_TUCUMAN_BANCARIAS.exists():
        pdf_rows = parse_pdf_tucuman_bancarias(PDF_TUCUMAN_BANCARIAS)
        filas, stats_pdf = incorporar_pdf_tucuman(filas, pdf_rows)
        print("PDF_TUC", PDF_TUCUMAN_BANCARIAS)
        print(
            "PDF stats",
            stats_pdf["pdf_filas"],
            "filas PDF,",
            stats_pdf["nuevas"],
            "nuevas,",
            stats_pdf["matched"],
            "dedupe,",
            "imp_nuevo",
            stats_pdf["importe_nuevo"],
            "pdf_total",
            stats_pdf["pdf_total"],
        )
    else:
        print("WARN sin PDF Tucumán:", PDF_TUCUMAN_BANCARIAS)

    def _sk(f: dict) -> tuple:
        return (
            _sort_mes(f.get("Mes") or ""),
            f.get("Fecha") or "",
            f.get("Banco") or "",
            f.get("Tipo") or "",
        )

    filas = sorted(filas, key=_sk)
    out = guardar(OUT, filas, stats_pdf)

    df = pd.DataFrame(filas)
    tot = round(float(df["Importe"].sum()), 2)
    tot_tuc = round(float(df.loc[df["Tipo"] == "Ret. IIBB Tucumán", "Importe"].sum()), 2)
    print("OUT", out)
    print("CANTIDAD", len(filas))
    print("TOTAL", tot)
    print("TOTAL_TUCUMAN", tot_tuc)
    for b, g in df.groupby("Banco"):
        print("BANCO", b, len(g), round(float(g["Importe"].sum()), 2))
    for t, g in df.groupby("Tipo"):
        print("TIPO", t, len(g), round(float(g["Importe"].sum()), 2))


if __name__ == "__main__":
    main()
