# -*- coding: utf-8 -*-
"""Débitos de RETENCIONES Galicia Meridiem → Excel formato estudio."""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    COLOR_PRIMARIO,
    MONEY_FMT,
    TITLE_FONT,
    ZEBRA,
    guardar_informe_excel,
)

PDF = Path(r"c:\Users\recep\Downloads\ilovepdf_merged (24).pdf")
OUT = Path(r"C:\Users\recep\Desktop\Debitos_Retenciones_Galicia_Meridiem.xlsx")

MESES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

RE_MOV = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s+"
    r"(-?\d{1,3}(?:\.\d{3})*(?:,\d{2})|-?\d+,\d{2})\s*$"
)
RE_MES_LABEL = re.compile(
    r"^(Enero|Febrero|Marzo|Abril|Mayo|Junio|Julio|Agosto|Septiembre|Octubre|Noviembre|Diciembre)\s+20\d{2}$",
    re.I,
)


def parse_ar_money(s: str) -> float:
    s = (s or "").strip().replace(" ", "")
    neg = s.startswith("-") or s.startswith("(")
    s = s.replace("-", "").replace("(", "").replace(")", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return -abs(float(s)) if neg else float(s)


def clasificar_retencion(desc: str, conts: list[str]) -> str | None:
    """Tipo de retención, o None si no aplica.

    En Galicia Meridiem las retenciones aparecen como:
      - ING. BRUTOS S/ CRED + DT.301/03-TUCUMAN  → Ret. IIBB Tucumán
      - ING. BRUTOS S/ CRED + REG.RECAU.SIRCREB → Ret. IIBB SIRCREB
      - PERCEP. IVA → Percepción IVA
    Se excluyen Ley 25413, IVA genérico s/ débitos, comisiones, etc.
    """
    u = (desc or "").strip().upper()
    blob = " ".join([desc or ""] + list(conts)).upper()

    # Excluir impuestos bancarios genéricos (aunque el consolidado diga "retención")
    if re.search(r"LEY\s*25\.?413|25413", u):
        return None
    if re.search(r"IMPUESTO\s+DE\s+SELLOS|INTERESES\s+SOBRE", u):
        return None

    if re.search(r"ING\.?\s*BRUTOS", u) or "SIRCREB" in blob:
        if "SIRCREB" in blob:
            return "Ret. IIBB SIRCREB"
        if "TUCUMAN" in blob or "301/03" in blob or re.search(r"DT\.?\s*301", blob):
            return "Ret. IIBB Tucumán"
        if re.search(r"ING\.?\s*BRUTOS", u):
            return "Ret. IIBB s/ créditos"
        return None

    if re.search(r"PERCEP\.?\s*IVA|PERCEPCION\s+IVA", u):
        return "Percepción IVA"

    if re.search(r"RET\.?\s*IVA|RETENCION\s+IVA", u):
        return "Ret. IVA"

    if re.search(r"RET\.?\s*GANAN|RETENCION\s+GANAN", u):
        return "Ret. Ganancias"

    if re.search(r"\bRETENCION\b|\bRETENCIONES\b|\bRET\.\b", u):
        if re.search(r"25413|25\.413", blob):
            return None
        return "Retención (otra)"

    # "IVA" solo = impuesto s/ débitos/comisiones, no retención
    return None


def mes_desde_conts(fecha, conts: list[str]) -> tuple[int, int]:
    mes_grupo, anio_grupo = fecha.month, fecha.year
    for c in conts:
        mm = RE_MES_LABEL.match(c.strip())
        if not mm:
            continue
        nombre = mm.group(1).lower()
        mapa = {v.lower(): k for k, v in MESES.items()}
        for k, v in list(mapa.items()):
            mapa[k.replace("á", "a").replace("é", "e").replace("ó", "o")] = v
        key = nombre.replace("á", "a").replace("é", "e").replace("ó", "o")
        if key in mapa or nombre in mapa:
            mes_grupo = mapa.get(nombre) or mapa[key]
            anio_m = re.search(r"(20\d{2})", c)
            if anio_m:
                anio_grupo = int(anio_m.group(1))
        break
    return mes_grupo, anio_grupo


def extraer_retenciones(pdf_path: Path) -> list[dict]:
    lineas: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for ln in (page.extract_text() or "").splitlines():
                s = ln.strip()
                if s:
                    lineas.append(s)

    retenciones: list[dict] = []
    i = 0
    while i < len(lineas):
        m = RE_MOV.match(lineas[i])
        if not m:
            i += 1
            continue
        fecha_s, desc, mon1, _mon2 = m.groups()
        monto = parse_ar_money(mon1)
        if monto >= -0.009 or desc.lower().startswith("total"):
            i += 1
            continue

        conts: list[str] = []
        j = i + 1
        while j < len(lineas):
            nxt = lineas[j]
            if RE_MOV.match(nxt) or nxt.startswith("Total $") or nxt.startswith("Resumen de"):
                break
            if nxt.startswith("Página") or ("Página" in nxt and "/" in nxt):
                break
            conts.append(nxt)
            j += 1

        try:
            fecha = datetime.strptime(fecha_s, "%d/%m/%y").date()
        except ValueError:
            i = j
            continue

        tipo = clasificar_retencion(desc, conts)
        if tipo is None:
            i = j
            continue

        mes_grupo, anio_grupo = mes_desde_conts(fecha, conts)
        retenciones.append(
            {
                "fecha": fecha,
                "mes": mes_grupo,
                "anio": anio_grupo,
                "importe": abs(monto),
                "tipo": tipo,
                "descripcion_raw": desc,
                "continuacion": " | ".join(conts[:3]) if conts else "",
            }
        )
        i = j

    return retenciones


def armar_filas(retenciones: list[dict]) -> list[dict]:
    grupos: dict[tuple[int, int], list[dict]] = defaultdict(list)
    for d in retenciones:
        grupos[(d["anio"], d["mes"])].append(d)

    filas: list[dict] = []
    for anio, mes in sorted(grupos.keys()):
        items = sorted(grupos[(anio, mes)], key=lambda x: (x["fecha"], x["tipo"], -x["importe"]))
        etiqueta = f"{MESES[mes]} {anio}"
        for d in items:
            filas.append(
                {
                    "Mes": etiqueta,
                    "Fecha": d["fecha"].strftime("%d/%m/%Y"),
                    "Tipo": d["tipo"],
                    "Importe": round(d["importe"], 2),
                    "Descripción": d["descripcion_raw"],
                    "Detalle": d["continuacion"] or d["tipo"],
                }
            )
    return filas


def guardar_excel(ruta: Path, filas: list[dict], retenciones: list[dict]) -> Path:
    import pandas as pd

    df = pd.DataFrame(filas)
    por_mes = (
        df.groupby("Mes", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )
    por_tipo = (
        df.groupby("Tipo", sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )

    guardar_informe_excel(
        ruta,
        titulo="Retenciones Banco Galicia — GRUPO MERIDIEM SRL",
        subtitulo="Solo débitos de retenciones / percepciones del extracto (sin Ley 25413)",
        periodo=f"{filas[0]['Mes']} → {filas[-1]['Mes']}" if filas else "",
        kpis=[
            ("Retenciones", len(filas)),
            ("Total", round(sum(d["importe"] for d in retenciones), 2)),
            ("Tipos", int(df["Tipo"].nunique()) if not df.empty else 0),
            ("Meses", int(df["Mes"].nunique()) if not df.empty else 0),
        ],
        resumenes=[
            ("Por tipo de retención", por_tipo),
            ("Totales por mes", por_mes),
        ],
        detalle=df[["Mes", "Fecha", "Tipo", "Importe", "Descripción", "Detalle"]],
        hoja_detalle="Lista",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    wb = load_workbook(ruta)
    if "Por mes" in wb.sheetnames:
        del wb["Por mes"]
    ws = wb.create_sheet("Por mes", 0)
    ws["A1"] = "Retenciones Galicia — vista clara"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "GRUPO MERIDIEM SRL · Solo débitos de retenciones"
    ws["A2"].font = Font(name="Calibri", size=11, color="666666")

    fila = 4
    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in filas:
        grupos[r["Mes"]].append(r)

    def sort_key(mes_label: str):
        m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", mes_label)
        if not m:
            return (9999, 13)
        nombre, anio = m.group(1), int(m.group(2))
        inv = {v: k for k, v in MESES.items()}
        return (anio, inv.get(nombre, 13))

    for mes_label in sorted(grupos.keys(), key=sort_key):
        items = grupos[mes_label]
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        cell = ws.cell(fila, 1, mes_label)
        cell.fill = fill_mes
        cell.font = font_mes
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in (2, 3):
            ws.cell(fila, col).fill = fill_mes
        fila += 1
        for idx, r in enumerate(items):
            ws.cell(fila, 1, r["Importe"]).number_format = MONEY_FMT
            ws.cell(fila, 1).font = BODY_FONT
            ws.cell(fila, 2, r["Tipo"]).font = BODY_FONT
            ws.cell(fila, 3, r["Fecha"]).font = BODY_FONT
            if idx % 2 == 1:
                for col in (1, 2, 3):
                    ws.cell(fila, col).fill = ZEBRA
            fila += 1
        tot = round(sum(x["Importe"] for x in items), 2)
        ws.cell(fila, 1, tot).number_format = MONEY_FMT
        ws.cell(fila, 1).font = font_tot
        ws.cell(fila, 2, f"TOTAL {mes_label}").font = font_tot
        fila += 2

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    wb.save(ruta)
    return ruta


def main() -> None:
    if not PDF.exists():
        raise SystemExit(f"No está el PDF: {PDF}")
    retenciones = extraer_retenciones(PDF)
    if not retenciones:
        raise SystemExit("No se extrajeron retenciones.")
    filas = armar_filas(retenciones)
    out = guardar_excel(OUT, filas, retenciones)
    total = round(sum(d["importe"] for d in retenciones), 2)
    print("OUT", out)
    print("CANTIDAD", len(retenciones))
    print("TOTAL", total)
    print("PERIODO", filas[0]["Mes"], "->", filas[-1]["Mes"])
    por_tipo: dict[str, float] = defaultdict(float)
    cnt: Counter[str] = Counter()
    for d in retenciones:
        por_tipo[d["tipo"]] += d["importe"]
        cnt[d["tipo"]] += 1
    for t, n in cnt.most_common():
        print(f"TIPO\t{n}\t{round(por_tipo[t], 2)}\t{t}")


if __name__ == "__main__":
    main()
