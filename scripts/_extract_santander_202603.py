import sys
from pathlib import Path
_SCRIPTS_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _SCRIPTS_DIR.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

"""Extrae extracto Santander 2026-03 (PDF sin texto) a Excel vía OCR."""
from __future__ import annotations

import io
import json
from datetime import datetime
from pathlib import Path

import fitz
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image

from procesador import (
    _extraer_movimientos_desde_texto,
    _extraer_saldos_desde_lineas,
    _obtener_lector_ocr,
)

PDF = Path(__file__).resolve().parent / "_tmp_santander_2026-03.pdf"
CACHE = Path(__file__).resolve().parent / "_tmp_santander_2026-03_ocr.json"
OUT = Path(r"C:\Users\recep\Desktop") / "Extracto_Santander_Oftalmologia_RELE_2026-03.xlsx"
OUT_NET = Path(
    r"\\TANGOSRV\Compartido\CLIENTES\OFTALMOLOGIA RELE MAR DEL PLATA SRL"
    r"\Balances\2026\Banco Santander\Extracto_Santander_2026-03_OCR.xlsx"
)
DPI = 180


def ocr_pagina_rapida(pagina_fitz) -> list[str]:
    lector = _obtener_lector_ocr()
    pix = pagina_fitz.get_pixmap(dpi=DPI)
    img_pil = Image.open(io.BytesIO(pix.tobytes("png")))
    rot_metadata = getattr(pagina_fitz, "rotation", 0) or 0
    if rot_metadata != 0:
        img_pil = img_pil.rotate(-rot_metadata, expand=True)
    w, h = img_pil.size
    if rot_metadata == 0 and w > h:
        img_pil = img_pil.rotate(90, expand=True)

    resultados = lector.readtext(np.array(img_pil))
    filas: dict[int, list[tuple[float, str]]] = {}
    for bbox, texto, _ in resultados:
        y_centro = (bbox[0][1] + bbox[2][1]) / 2
        clave = int(y_centro / 18) * 18
        filas.setdefault(clave, []).append((bbox[0][0], texto))
    return [
        " ".join(t for _, t in sorted(filas[y], key=lambda x: x[0])).strip()
        for y in sorted(filas)
        if any(t.strip() for _, t in filas[y])
    ]


def main() -> None:
    print("Inicio", datetime.now().isoformat(timespec="seconds"), flush=True)
    if not PDF.exists():
        raise SystemExit(f"No existe PDF: {PDF}")

    if CACHE.exists():
        print(f"Usando cache OCR: {CACHE}", flush=True)
        lineas_por_pagina = json.loads(CACHE.read_text(encoding="utf-8"))
    else:
        doc = fitz.open(PDF)
        print(f"Paginas: {doc.page_count}", flush=True)
        lineas_por_pagina = []
        for i in range(doc.page_count):
            print(f"OCR pagina {i + 1}/{doc.page_count}...", flush=True)
            lineas = [l for l in ocr_pagina_rapida(doc[i]) if l]
            lineas_por_pagina.append(lineas)
            print(f"  -> {len(lineas)} lineas", flush=True)
            CACHE.write_text(
                json.dumps(lineas_por_pagina, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        doc.close()

    todas = [l for page in lineas_por_pagina for l in page]
    movs = []
    for pidx, lineas in enumerate(lineas_por_pagina, start=1):
        movs.extend(
            _extraer_movimientos_desde_texto(lineas, pidx, "santander", PDF.name)
        )

    # Dedup simple
    vistos: set[tuple] = set()
    unicos = []
    for m in sorted(movs, key=lambda x: (x.pagina, x.fecha or datetime.min.date(), x.descripcion)):
        clave = (
            m.fecha,
            round(m.debito or 0, 2),
            round(m.credito or 0, 2),
            (m.descripcion or "")[:50],
            m.pagina,
        )
        if clave in vistos:
            continue
        vistos.add(clave)
        unicos.append(m)
    movs = unicos

    si, sf = _extraer_saldos_desde_lineas(todas)
    print(f"Movimientos: {len(movs)} | SI={si} SF={sf}", flush=True)

    df_mov = pd.DataFrame(
        [
            {
                "Fecha": m.fecha.strftime("%d/%m/%Y") if m.fecha else "",
                "Comprobante": m.comprobante or "",
                "Descripcion": m.descripcion or "",
                "Debito": m.debito if m.debito else None,
                "Credito": m.credito if m.credito else None,
                "Saldo": m.saldo if m.saldo is not None else None,
                "CUIT contraparte": m.cuit_contraparte or "",
                "Pagina PDF": m.pagina,
                "Banco": "Santander",
                "Archivo": PDF.name,
            }
            for m in movs
        ]
    )
    df_ocr = pd.DataFrame(
        [
            {"Pagina": pidx, "Orden": n, "Texto OCR": linea}
            for pidx, lineas in enumerate(lineas_por_pagina, start=1)
            for n, linea in enumerate(lineas, start=1)
        ]
    )
    total_deb = float(df_mov["Debito"].fillna(0).sum()) if len(df_mov) else 0.0
    total_cred = float(df_mov["Credito"].fillna(0).sum()) if len(df_mov) else 0.0
    resumen = pd.DataFrame(
        [
            {"Campo": "Cliente", "Valor": "OFTALMOLOGIA RELE MAR DEL PLATA SRL"},
            {"Campo": "Banco", "Valor": "Banco Santander"},
            {"Campo": "Periodo archivo", "Valor": "2026-03"},
            {"Campo": "PDF origen (red)", "Valor": r"\\TANGOSRV\Compartido\CLIENTES\OFTALMOLOGIA RELE MAR DEL PLATA SRL\Balances\2026\Banco Santander\2026-03.pdf"},
            {"Campo": "Paginas PDF", "Valor": len(lineas_por_pagina)},
            {"Campo": "Lineas OCR", "Valor": len(df_ocr)},
            {"Campo": "Movimientos parseados", "Valor": len(df_mov)},
            {"Campo": "Saldo inicial detectado", "Valor": si if si is not None else ""},
            {"Campo": "Saldo final detectado", "Valor": sf if sf is not None else ""},
            {"Campo": "Total Debitos", "Valor": round(total_deb, 2)},
            {"Campo": "Total Creditos", "Valor": round(total_cred, 2)},
            {"Campo": "Generado", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M")},
            {
                "Campo": "Nota",
                "Valor": "PDF sin texto nativo: OCR easyocr. Revisar importes/fechas ante posibles errores de lectura.",
            },
        ]
    )

    header_font = Font(name="Calibri", bold=True, size=12, color="000000")
    body_font = Font(name="Calibri", size=11, color="000000")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fill_h = PatternFill("solid", fgColor="D9E2F3")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    for r in dataframe_to_rows(resumen, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill_h
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.font = body_font
            cell.border = thin
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    ws2 = wb.create_sheet("Movimientos")
    for r in dataframe_to_rows(df_mov, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = fill_h
        cell.border = thin
    for row in ws2.iter_rows(min_row=2, max_row=max(2, ws2.max_row), max_col=ws2.max_column):
        for cell in row:
            cell.font = body_font
            cell.border = thin
            if cell.column in (4, 5, 6) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
    for col, w in {
        "A": 12,
        "B": 14,
        "C": 55,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 12,
        "I": 12,
        "J": 18,
    }.items():
        ws2.column_dimensions[col].width = w
    if ws2.max_row > 1:
        ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("OCR_completo")
    for r in dataframe_to_rows(df_ocr, index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = fill_h
        cell.border = thin
    for row in ws3.iter_rows(min_row=2, max_row=max(2, ws3.max_row), max_col=3):
        for cell in row:
            cell.font = body_font
            cell.border = thin
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 100
    ws3.freeze_panes = "A2"

    wb.save(OUT)
    print("Guardado local:", OUT, flush=True)
    try:
        wb.save(OUT_NET)
        print("Guardado red:", OUT_NET, flush=True)
    except Exception as exc:
        print("No se pudo guardar en red:", exc, flush=True)
    print("OK", flush=True)


if __name__ == "__main__":
    main()
