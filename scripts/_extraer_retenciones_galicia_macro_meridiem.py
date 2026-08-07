# -*- coding: utf-8 -*-
"""Débitos de RETENCIONES Galicia + Macro (Meridiem) → Excel único por banco."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from excel_formato_estudio import (  # noqa: E402
    BODY_FONT,
    COLOR_PRIMARIO,
    MONEY_FMT,
    TITLE_FONT,
    ZEBRA,
    _escribir_encabezado_hoja,
    _escribir_tabla,
    guardar_informe_excel,
)
from _extraer_retenciones_galicia_meridiem import (  # noqa: E402
    MESES,
    armar_filas as armar_filas_galicia,
    extraer_retenciones as extraer_retenciones_galicia,
)
from _extraer_debitos_macro_meridiem import (  # noqa: E402
    extraer_debitos as extraer_debitos_macro,
)

PDF_GALICIA = Path(r"c:\Users\recep\Downloads\ilovepdf_merged (24).pdf")
PDF_MACRO_CANDIDATES = [
    Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged (1).pdf"),
    Path(r"c:\Users\recep\Downloads\Resumen (1) (1)_merged.pdf"),
]
OUT = Path(r"C:\Users\recep\Desktop\Debitos_Retenciones_Galicia_Macro_Meridiem.xlsx")


def clasificar_retencion_macro(desc: str) -> str | None:
    """Tipo de retención Macro, o None si no aplica (excluye Ley 25413 / AFIP genérico)."""
    u = (desc or "").strip().upper()
    if re.search(r"25413|IDCB\s+GRAL|DBCR\s+25413", u):
        return None
    if re.search(r"^IMP\.?\s*AFIP|IMPUESTOS AFIP", u) and not re.search(
        r"RET|PERCEP|SIRCREB|IIBB", u
    ):
        return None
    if re.search(r"SIRCREB|RET\.?\s*ING\.?\s*BRUTOS", u):
        return "Ret. IIBB SIRCREB"
    if re.search(r"RET\s*IIBB\s*TUCUMAN|IIBB\s*TUCUMAN", u):
        return "Ret. IIBB Tucumán"
    if re.search(r"PERCEPCION.*BRUTOS|PERCEPCION\s+I\s*I\s*B\s*B|PERCEP.*IIBB", u):
        return "Percepción IIBB"
    if re.search(r"RETENCION\s+IVA\s+PERCEPCION|PERCEP\.?\s*IVA|PERCEPCION\s+IVA", u):
        return "Percepción IVA"
    if re.search(r"DEBITO\s+FISCAL\s+IVA", u):
        return "Débito fiscal IVA"
    if re.search(r"RET\.?\s*IVA|RETENCION\s+IVA", u):
        return "Ret. IVA"
    if re.search(r"\bRETENCION\b|\bRET\.\b|\bRET\s", u):
        return "Retención (otra)"
    return None


def extraer_retenciones_macro(pdf_path: Path) -> list[dict]:
    retenciones: list[dict] = []
    for d in extraer_debitos_macro(pdf_path):
        tipo = clasificar_retencion_macro(d["descripcion_raw"])
        if tipo is None:
            continue
        retenciones.append(
            {
                "fecha": d["fecha"],
                "mes": d["mes"],
                "anio": d["anio"],
                "importe": round(d["importe"], 2),
                "tipo": tipo,
                "descripcion_raw": d["descripcion_raw"],
                "continuacion": "",
            }
        )
    return retenciones


def filas_banco(retenciones: list[dict], banco: str) -> list[dict]:
    base = armar_filas_galicia(retenciones)
    out: list[dict] = []
    for r in base:
        out.append(
            {
                "Banco": banco,
                "Mes": r["Mes"],
                "Fecha": r["Fecha"],
                "Tipo": r["Tipo"],
                "Importe": r["Importe"],
                "Descripción": r["Descripción"],
                "Detalle": r["Detalle"],
            }
        )
    return out


def _sort_mes(mes_label: str) -> tuple[int, int]:
    m = re.match(r"([A-ZÁÉÍÓÚ]+)\s+(20\d{2})", mes_label)
    if not m:
        return (9999, 13)
    inv = {v: k for k, v in MESES.items()}
    return (int(m.group(2)), inv.get(m.group(1), 13))


def _hoja_banco(wb, nombre: str, titulo: str, filas: list[dict]) -> None:
    import pandas as pd

    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)
    fila = _escribir_encabezado_hoja(
        ws,
        titulo,
        "GRUPO MERIDIEM SRL · Solo débitos de retenciones / percepciones",
        f"{filas[0]['Mes']} → {filas[-1]['Mes']}" if filas else "",
    )
    df = pd.DataFrame(filas)[["Mes", "Fecha", "Tipo", "Importe", "Descripción"]]
    _escribir_tabla(
        ws,
        df,
        fila,
        col_moneda=["Importe"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    # Vista por mes al costado
    fila_pm = fila + len(df) + 4
    ws.cell(fila_pm, 1, "Por mes").font = Font(
        name="Calibri", bold=True, size=12, color=COLOR_PRIMARIO
    )
    fila_pm += 1
    fill_mes = PatternFill("solid", fgColor=COLOR_PRIMARIO)
    font_mes = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_tot = Font(name="Calibri", bold=True, size=11, color=COLOR_PRIMARIO)

    grupos: dict[str, list[dict]] = defaultdict(list)
    for r in filas:
        grupos[r["Mes"]].append(r)

    for mes_label in sorted(grupos.keys(), key=_sort_mes):
        items = grupos[mes_label]
        ws.merge_cells(start_row=fila_pm, start_column=1, end_row=fila_pm, end_column=3)
        cell = ws.cell(fila_pm, 1, mes_label)
        cell.fill = fill_mes
        cell.font = font_mes
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in (2, 3):
            ws.cell(fila_pm, col).fill = fill_mes
        fila_pm += 1
        for idx, r in enumerate(items):
            ws.cell(fila_pm, 1, r["Importe"]).number_format = MONEY_FMT
            ws.cell(fila_pm, 1).font = BODY_FONT
            ws.cell(fila_pm, 2, r["Tipo"]).font = BODY_FONT
            ws.cell(fila_pm, 3, r["Fecha"]).font = BODY_FONT
            if idx % 2 == 1:
                for col in (1, 2, 3):
                    ws.cell(fila_pm, col).fill = ZEBRA
            fila_pm += 1
        tot = round(sum(x["Importe"] for x in items), 2)
        ws.cell(fila_pm, 1, tot).number_format = MONEY_FMT
        ws.cell(fila_pm, 1).font = font_tot
        ws.cell(fila_pm, 2, f"TOTAL {mes_label}").font = font_tot
        fila_pm += 2


def guardar_excel(
    ruta: Path,
    filas_g: list[dict],
    filas_m: list[dict],
    rets_g: list[dict],
    rets_m: list[dict],
) -> Path:
    import pandas as pd

    filas = filas_g + filas_m
    df = pd.DataFrame(filas)
    tot_g = round(sum(d["importe"] for d in rets_g), 2)
    tot_m = round(sum(d["importe"] for d in rets_m), 2)
    tot = round(tot_g + tot_m, 2)

    por_banco = pd.DataFrame(
        [
            {"Banco": "Galicia", "Cantidad": len(filas_g), "Total": tot_g},
            {"Banco": "Macro", "Cantidad": len(filas_m), "Total": tot_m},
        ]
    )
    por_tipo = (
        df.groupby(["Banco", "Tipo"], sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
        .sort_values(["Banco", "Total"], ascending=[True, False])
    )
    por_mes = (
        df.groupby(["Banco", "Mes"], sort=False)
        .agg(Cantidad=("Importe", "count"), Total=("Importe", "sum"))
        .reset_index()
    )
    # ordenar meses cronológicamente dentro de cada banco
    por_mes["_sk"] = por_mes["Mes"].map(_sort_mes)
    por_mes = por_mes.sort_values(["Banco", "_sk"]).drop(columns=["_sk"])

    periodo = ""
    if filas:
        meses = sorted({r["Mes"] for r in filas}, key=_sort_mes)
        periodo = f"{meses[0]} → {meses[-1]}"

    guardar_informe_excel(
        ruta,
        titulo="Retenciones Galicia + Macro — GRUPO MERIDIEM SRL",
        subtitulo="Solo débitos de retenciones / percepciones (sin Ley 25413 genérico)",
        periodo=periodo,
        kpis=[
            ("Retenciones", len(filas)),
            ("Total general", tot),
            ("Galicia", tot_g),
            ("Macro", tot_m),
        ],
        resumenes=[
            ("Totales por banco", por_banco),
            ("Por banco y tipo", por_tipo),
            ("Por banco y mes", por_mes),
        ],
        detalle=df[["Banco", "Mes", "Fecha", "Tipo", "Importe", "Descripción"]],
        hoja_detalle="Lista",
        col_moneda=["Importe", "Total"],
        col_fecha=["Fecha"],
        total_col="Importe",
    )

    wb = load_workbook(ruta)
    _hoja_banco(wb, "Galicia", "Retenciones Banco Galicia", filas_g)
    _hoja_banco(wb, "Macro", "Retenciones Banco Macro", filas_m)

    # Orden: Resumen, Galicia, Macro, Lista
    orden = ["Resumen", "Galicia", "Macro", "Lista"]
    for i, name in enumerate(orden):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
    wb.save(ruta)
    return ruta


def resolver_pdf_macro() -> Path:
    for p in PDF_MACRO_CANDIDATES:
        if p.exists():
            return p
    raise SystemExit("No está el PDF Macro en Downloads.")


def main() -> None:
    if not PDF_GALICIA.exists():
        raise SystemExit(f"No está el PDF Galicia: {PDF_GALICIA}")
    pdf_m = resolver_pdf_macro()

    rets_g = extraer_retenciones_galicia(PDF_GALICIA)
    rets_m = extraer_retenciones_macro(pdf_m)
    if not rets_g and not rets_m:
        raise SystemExit("No se extrajeron retenciones.")

    filas_g = filas_banco(rets_g, "Galicia")
    filas_m = filas_banco(rets_m, "Macro")
    out = guardar_excel(OUT, filas_g, filas_m, rets_g, rets_m)

    tot_g = round(sum(d["importe"] for d in rets_g), 2)
    tot_m = round(sum(d["importe"] for d in rets_m), 2)
    meses = sorted({r["Mes"] for r in filas_g + filas_m}, key=_sort_mes)
    print("OUT", out)
    print("PDF_MACRO", pdf_m)
    print("GALICIA", len(rets_g), tot_g)
    print("MACRO", len(rets_m), tot_m)
    print("TOTAL", len(rets_g) + len(rets_m), round(tot_g + tot_m, 2))
    print("PERIODO", f"{meses[0]} -> {meses[-1]}" if meses else "")


if __name__ == "__main__":
    main()
